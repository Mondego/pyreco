__FILENAME__ = admin
from django.contrib import admin
from simple_import.models import ImportLog, ColumnMatch

class ColumnMatchInline(admin.TabularInline):
    model = ColumnMatch
    extra = 0

class ImportLogAdmin(admin.ModelAdmin):
    list_display = ('name', 'user', 'date',)
    #inlines = [ColumnMatchInline]
admin.site.register(ImportLog, ImportLogAdmin)

########NEW FILE########
__FILENAME__ = compat
# This file has been borrowed from TastyPie (https://github.com/toastdriven/django-tastypie) under a BSD license:
#Copyright (c) 2010, Daniel Lindsley
#All rights reserved.

#Redistribution and use in source and binary forms, with or without
#modification, are permitted provided that the following conditions are met:
#    * Redistributions of source code must retain the above copyright
#      notice, this list of conditions and the following disclaimer.
#    * Redistributions in binary form must reproduce the above copyright
#      notice, this list of conditions and the following disclaimer in the
#      documentation and/or other materials provided with the distribution.
#    * Neither the name of the tastypie nor the
#      names of its contributors may be used to endorse or promote products
#      derived from this software without specific prior written permission.

#THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS" AND
#ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED
#WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE
#DISCLAIMED. IN NO EVENT SHALL tastypie BE LIABLE FOR ANY
#DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES
#(INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES;
#LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND
#ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY, OR TORT
#(INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF THIS
#SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.

from django.conf import settings
from django.core.exceptions import ImproperlyConfigured
import django

__all__ = ['User', 'AUTH_USER_MODEL']

AUTH_USER_MODEL = getattr(settings, 'AUTH_USER_MODEL', 'auth.User')

# Django 1.5+ compatibility
if django.VERSION >= (1, 5):
    try:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        username_field = User.USERNAME_FIELD
    except ImproperlyConfigured:
        # The the users model might not be read yet.
        # This can happen is when setting up the create_api_key signal, in your
        # custom user module.
        User = None
else:
    from django.contrib.auth.models import User
    username_field = 'username'

########NEW FILE########
__FILENAME__ = forms
from django import forms
from django.contrib.contenttypes.models import ContentType

from simple_import.models import ImportLog, ColumnMatch, RelationalMatch


class ImportForm(forms.ModelForm):
    class Meta:
        model = ImportLog
        fields = ('name', 'import_file', 'import_type')
    model = forms.ModelChoiceField(ContentType.objects.all())
    
    
class MatchForm(forms.ModelForm):
    class Meta:
        model = ColumnMatch
        exclude = ['header_position']

class MatchRelationForm(forms.ModelForm):
    class Meta:
        model = RelationalMatch
        widgets = {
            'related_field_name': forms.Select(choices=(('', '---------'),))
        }

########NEW FILE########
__FILENAME__ = 0001_initial
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding model 'ImportSetting'
        db.create_table(u'simple_import_importsetting', (
            (u'id', self.gf('django.db.models.fields.AutoField')(primary_key=True)),
            ('user', self.gf('django.db.models.fields.related.ForeignKey')(to=orm['auth.User'])),
            ('content_type', self.gf('django.db.models.fields.related.ForeignKey')(to=orm['contenttypes.ContentType'])),
        ))
        db.send_create_signal(u'simple_import', ['ImportSetting'])

        # Adding model 'ColumnMatch'
        db.create_table(u'simple_import_columnmatch', (
            (u'id', self.gf('django.db.models.fields.AutoField')(primary_key=True)),
            ('column_name', self.gf('django.db.models.fields.CharField')(max_length=200)),
            ('field_name', self.gf('django.db.models.fields.CharField')(max_length=255, blank=True)),
            ('import_setting', self.gf('django.db.models.fields.related.ForeignKey')(to=orm['simple_import.ImportSetting'])),
            ('default_value', self.gf('django.db.models.fields.CharField')(max_length=2000, blank=True)),
        ))
        db.send_create_signal(u'simple_import', ['ColumnMatch'])

        # Adding unique constraint on 'ColumnMatch', fields ['column_name', 'import_setting']
        db.create_unique(u'simple_import_columnmatch', ['column_name', 'import_setting_id'])

        # Adding model 'ImportLog'
        db.create_table(u'simple_import_importlog', (
            (u'id', self.gf('django.db.models.fields.AutoField')(primary_key=True)),
            ('name', self.gf('django.db.models.fields.CharField')(max_length=255)),
            ('user', self.gf('django.db.models.fields.related.ForeignKey')(related_name='simple_import_log', to=orm['auth.User'])),
            ('date', self.gf('django.db.models.fields.DateTimeField')(auto_now_add=True, blank=True)),
            ('import_file', self.gf('django.db.models.fields.files.FileField')(max_length=100)),
            ('error_file', self.gf('django.db.models.fields.files.FileField')(max_length=100, blank=True)),
            ('import_setting', self.gf('django.db.models.fields.related.ForeignKey')(to=orm['simple_import.ImportSetting'])),
            ('import_type', self.gf('django.db.models.fields.CharField')(max_length=1)),
        ))
        db.send_create_signal(u'simple_import', ['ImportLog'])


    def backwards(self, orm):
        # Removing unique constraint on 'ColumnMatch', fields ['column_name', 'import_setting']
        db.delete_unique(u'simple_import_columnmatch', ['column_name', 'import_setting_id'])

        # Deleting model 'ImportSetting'
        db.delete_table(u'simple_import_importsetting')

        # Deleting model 'ColumnMatch'
        db.delete_table(u'simple_import_columnmatch')

        # Deleting model 'ImportLog'
        db.delete_table(u'simple_import_importlog')


    models = {
        u'auth.group': {
            'Meta': {'object_name': 'Group'},
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': u"orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        u'auth.permission': {
            'Meta': {'ordering': "(u'content_type__app_label', u'content_type__model', u'codename')", 'unique_together': "((u'content_type', u'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['contenttypes.ContentType']"}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        u'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': u"orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': u"orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        u'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        u'simple_import.columnmatch': {
            'Meta': {'unique_together': "(('column_name', 'import_setting'),)", 'object_name': 'ColumnMatch'},
            'column_name': ('django.db.models.fields.CharField', [], {'max_length': '200'}),
            'default_value': ('django.db.models.fields.CharField', [], {'max_length': '2000', 'blank': 'True'}),
            'field_name': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_setting': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportSetting']"})
        },
        u'simple_import.importlog': {
            'Meta': {'object_name': 'ImportLog'},
            'date': ('django.db.models.fields.DateTimeField', [], {'auto_now_add': 'True', 'blank': 'True'}),
            'error_file': ('django.db.models.fields.files.FileField', [], {'max_length': '100', 'blank': 'True'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_file': ('django.db.models.fields.files.FileField', [], {'max_length': '100'}),
            'import_setting': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportSetting']"}),
            'import_type': ('django.db.models.fields.CharField', [], {'max_length': '1'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'simple_import_log'", 'to': u"orm['auth.User']"})
        },
        u'simple_import.importsetting': {
            'Meta': {'object_name': 'ImportSetting'},
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['contenttypes.ContentType']"}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['auth.User']"})
        }
    }

    complete_apps = ['simple_import']
########NEW FILE########
__FILENAME__ = 0002_auto__add_importedobject__add_field_importlog_update_key
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding model 'ImportedObject'
        db.create_table(u'simple_import_importedobject', (
            (u'id', self.gf('django.db.models.fields.AutoField')(primary_key=True)),
            ('import_log', self.gf('django.db.models.fields.related.ForeignKey')(to=orm['simple_import.ImportLog'])),
            ('object_id', self.gf('django.db.models.fields.IntegerField')()),
            ('content_type', self.gf('django.db.models.fields.related.ForeignKey')(to=orm['contenttypes.ContentType'])),
        ))
        db.send_create_signal(u'simple_import', ['ImportedObject'])

        # Adding field 'ImportLog.update_key'
        db.add_column(u'simple_import_importlog', 'update_key',
                      self.gf('django.db.models.fields.CharField')(default='', max_length=200, blank=True),
                      keep_default=False)


    def backwards(self, orm):
        # Deleting model 'ImportedObject'
        db.delete_table(u'simple_import_importedobject')

        # Deleting field 'ImportLog.update_key'
        db.delete_column(u'simple_import_importlog', 'update_key')


    models = {
        u'auth.group': {
            'Meta': {'object_name': 'Group'},
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': u"orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        u'auth.permission': {
            'Meta': {'ordering': "(u'content_type__app_label', u'content_type__model', u'codename')", 'unique_together': "((u'content_type', u'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['contenttypes.ContentType']"}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        u'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': u"orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': u"orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        u'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        u'simple_import.columnmatch': {
            'Meta': {'unique_together': "(('column_name', 'import_setting'),)", 'object_name': 'ColumnMatch'},
            'column_name': ('django.db.models.fields.CharField', [], {'max_length': '200'}),
            'default_value': ('django.db.models.fields.CharField', [], {'max_length': '2000', 'blank': 'True'}),
            'field_name': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_setting': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportSetting']"})
        },
        u'simple_import.importedobject': {
            'Meta': {'object_name': 'ImportedObject'},
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['contenttypes.ContentType']"}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_log': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportLog']"}),
            'object_id': ('django.db.models.fields.IntegerField', [], {})
        },
        u'simple_import.importlog': {
            'Meta': {'object_name': 'ImportLog'},
            'date': ('django.db.models.fields.DateTimeField', [], {'auto_now_add': 'True', 'blank': 'True'}),
            'error_file': ('django.db.models.fields.files.FileField', [], {'max_length': '100', 'blank': 'True'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_file': ('django.db.models.fields.files.FileField', [], {'max_length': '100'}),
            'import_setting': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportSetting']"}),
            'import_type': ('django.db.models.fields.CharField', [], {'max_length': '1'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'update_key': ('django.db.models.fields.CharField', [], {'max_length': '200', 'blank': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'simple_import_log'", 'to': u"orm['auth.User']"})
        },
        u'simple_import.importsetting': {
            'Meta': {'object_name': 'ImportSetting'},
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['contenttypes.ContentType']"}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['auth.User']"})
        }
    }

    complete_apps = ['simple_import']
########NEW FILE########
__FILENAME__ = 0003_auto__add_relationalmatch
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding model 'RelationalMatch'
        db.create_table(u'simple_import_relationalmatch', (
            (u'id', self.gf('django.db.models.fields.AutoField')(primary_key=True)),
            ('import_log', self.gf('django.db.models.fields.related.ForeignKey')(to=orm['simple_import.ImportLog'])),
            ('field_name', self.gf('django.db.models.fields.CharField')(max_length=255)),
            ('related_field_name', self.gf('django.db.models.fields.CharField')(max_length=255)),
        ))
        db.send_create_signal(u'simple_import', ['RelationalMatch'])


    def backwards(self, orm):
        # Deleting model 'RelationalMatch'
        db.delete_table(u'simple_import_relationalmatch')


    models = {
        u'auth.group': {
            'Meta': {'object_name': 'Group'},
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': u"orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        u'auth.permission': {
            'Meta': {'ordering': "(u'content_type__app_label', u'content_type__model', u'codename')", 'unique_together': "((u'content_type', u'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['contenttypes.ContentType']"}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        u'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': u"orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': u"orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        u'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        u'simple_import.columnmatch': {
            'Meta': {'unique_together': "(('column_name', 'import_setting'),)", 'object_name': 'ColumnMatch'},
            'column_name': ('django.db.models.fields.CharField', [], {'max_length': '200'}),
            'default_value': ('django.db.models.fields.CharField', [], {'max_length': '2000', 'blank': 'True'}),
            'field_name': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_setting': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportSetting']"})
        },
        u'simple_import.importedobject': {
            'Meta': {'object_name': 'ImportedObject'},
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['contenttypes.ContentType']"}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_log': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportLog']"}),
            'object_id': ('django.db.models.fields.IntegerField', [], {})
        },
        u'simple_import.importlog': {
            'Meta': {'object_name': 'ImportLog'},
            'date': ('django.db.models.fields.DateTimeField', [], {'auto_now_add': 'True', 'blank': 'True'}),
            'error_file': ('django.db.models.fields.files.FileField', [], {'max_length': '100', 'blank': 'True'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_file': ('django.db.models.fields.files.FileField', [], {'max_length': '100'}),
            'import_setting': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportSetting']"}),
            'import_type': ('django.db.models.fields.CharField', [], {'max_length': '1'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'update_key': ('django.db.models.fields.CharField', [], {'max_length': '200', 'blank': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'simple_import_log'", 'to': u"orm['auth.User']"})
        },
        u'simple_import.importsetting': {
            'Meta': {'object_name': 'ImportSetting'},
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['contenttypes.ContentType']"}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['auth.User']"})
        },
        u'simple_import.relationalmatch': {
            'Meta': {'object_name': 'RelationalMatch'},
            'field_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_log': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportLog']"}),
            'related_field_name': ('django.db.models.fields.CharField', [], {'max_length': '255'})
        }
    }

    complete_apps = ['simple_import']
########NEW FILE########
__FILENAME__ = 0004_auto__add_field_columnmatch_null_on_empty
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding field 'ColumnMatch.null_on_empty'
        db.add_column(u'simple_import_columnmatch', 'null_on_empty',
                      self.gf('django.db.models.fields.BooleanField')(default=False),
                      keep_default=False)


    def backwards(self, orm):
        # Deleting field 'ColumnMatch.null_on_empty'
        db.delete_column(u'simple_import_columnmatch', 'null_on_empty')


    models = {
        u'auth.group': {
            'Meta': {'object_name': 'Group'},
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': u"orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        u'auth.permission': {
            'Meta': {'ordering': "(u'content_type__app_label', u'content_type__model', u'codename')", 'unique_together': "((u'content_type', u'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['contenttypes.ContentType']"}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        u'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': u"orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': u"orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        u'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        u'simple_import.columnmatch': {
            'Meta': {'unique_together': "(('column_name', 'import_setting'),)", 'object_name': 'ColumnMatch'},
            'column_name': ('django.db.models.fields.CharField', [], {'max_length': '200'}),
            'default_value': ('django.db.models.fields.CharField', [], {'max_length': '2000', 'blank': 'True'}),
            'field_name': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_setting': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportSetting']"}),
            'null_on_empty': ('django.db.models.fields.BooleanField', [], {'default': 'False'})
        },
        u'simple_import.importedobject': {
            'Meta': {'object_name': 'ImportedObject'},
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['contenttypes.ContentType']"}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_log': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportLog']"}),
            'object_id': ('django.db.models.fields.IntegerField', [], {})
        },
        u'simple_import.importlog': {
            'Meta': {'object_name': 'ImportLog'},
            'date': ('django.db.models.fields.DateTimeField', [], {'auto_now_add': 'True', 'blank': 'True'}),
            'error_file': ('django.db.models.fields.files.FileField', [], {'max_length': '100', 'blank': 'True'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_file': ('django.db.models.fields.files.FileField', [], {'max_length': '100'}),
            'import_setting': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportSetting']"}),
            'import_type': ('django.db.models.fields.CharField', [], {'max_length': '1'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'update_key': ('django.db.models.fields.CharField', [], {'max_length': '200', 'blank': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'simple_import_log'", 'to': u"orm['auth.User']"})
        },
        u'simple_import.importsetting': {
            'Meta': {'object_name': 'ImportSetting'},
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['contenttypes.ContentType']"}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['auth.User']"})
        },
        u'simple_import.relationalmatch': {
            'Meta': {'object_name': 'RelationalMatch'},
            'field_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_log': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportLog']"}),
            'related_field_name': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'})
        }
    }

    complete_apps = ['simple_import']
########NEW FILE########
__FILENAME__ = 0005_auto__add_field_columnmatch_header_position
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding field 'ColumnMatch.header_position'
        db.add_column(u'simple_import_columnmatch', 'header_position',
                      self.gf('django.db.models.fields.IntegerField')(default=1),
                      keep_default=False)


    def backwards(self, orm):
        # Deleting field 'ColumnMatch.header_position'
        db.delete_column(u'simple_import_columnmatch', 'header_position')


    models = {
        u'auth.group': {
            'Meta': {'object_name': 'Group'},
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': u"orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        u'auth.permission': {
            'Meta': {'ordering': "(u'content_type__app_label', u'content_type__model', u'codename')", 'unique_together': "((u'content_type', u'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['contenttypes.ContentType']"}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        u'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': u"orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': u"orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        u'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        u'simple_import.columnmatch': {
            'Meta': {'unique_together': "(('column_name', 'import_setting'),)", 'object_name': 'ColumnMatch'},
            'column_name': ('django.db.models.fields.CharField', [], {'max_length': '200'}),
            'default_value': ('django.db.models.fields.CharField', [], {'max_length': '2000', 'blank': 'True'}),
            'field_name': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'header_position': ('django.db.models.fields.IntegerField', [], {}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_setting': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportSetting']"}),
            'null_on_empty': ('django.db.models.fields.BooleanField', [], {'default': 'False'})
        },
        u'simple_import.importedobject': {
            'Meta': {'object_name': 'ImportedObject'},
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['contenttypes.ContentType']"}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_log': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportLog']"}),
            'object_id': ('django.db.models.fields.IntegerField', [], {})
        },
        u'simple_import.importlog': {
            'Meta': {'object_name': 'ImportLog'},
            'date': ('django.db.models.fields.DateTimeField', [], {'auto_now_add': 'True', 'blank': 'True'}),
            'error_file': ('django.db.models.fields.files.FileField', [], {'max_length': '100', 'blank': 'True'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_file': ('django.db.models.fields.files.FileField', [], {'max_length': '100'}),
            'import_setting': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportSetting']"}),
            'import_type': ('django.db.models.fields.CharField', [], {'max_length': '1'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'update_key': ('django.db.models.fields.CharField', [], {'max_length': '200', 'blank': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'simple_import_log'", 'to': u"orm['auth.User']"})
        },
        u'simple_import.importsetting': {
            'Meta': {'object_name': 'ImportSetting'},
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['contenttypes.ContentType']"}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['auth.User']"})
        },
        u'simple_import.relationalmatch': {
            'Meta': {'object_name': 'RelationalMatch'},
            'field_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_log': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportLog']"}),
            'related_field_name': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'})
        }
    }

    complete_apps = ['simple_import']
########NEW FILE########
__FILENAME__ = 0006_auto__add_unique_importsetting_user_content_type
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding unique constraint on 'ImportSetting', fields ['user', 'content_type']
        db.create_unique(u'simple_import_importsetting', ['user_id', 'content_type_id'])


    def backwards(self, orm):
        # Removing unique constraint on 'ImportSetting', fields ['user', 'content_type']
        db.delete_unique(u'simple_import_importsetting', ['user_id', 'content_type_id'])


    models = {
        u'auth.group': {
            'Meta': {'object_name': 'Group'},
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': u"orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        u'auth.permission': {
            'Meta': {'ordering': "(u'content_type__app_label', u'content_type__model', u'codename')", 'unique_together': "((u'content_type', u'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['contenttypes.ContentType']"}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        u'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': u"orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': u"orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        u'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        u'simple_import.columnmatch': {
            'Meta': {'unique_together': "(('column_name', 'import_setting'),)", 'object_name': 'ColumnMatch'},
            'column_name': ('django.db.models.fields.CharField', [], {'max_length': '200'}),
            'default_value': ('django.db.models.fields.CharField', [], {'max_length': '2000', 'blank': 'True'}),
            'field_name': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'header_position': ('django.db.models.fields.IntegerField', [], {}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_setting': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportSetting']"}),
            'null_on_empty': ('django.db.models.fields.BooleanField', [], {'default': 'False'})
        },
        u'simple_import.importedobject': {
            'Meta': {'object_name': 'ImportedObject'},
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['contenttypes.ContentType']"}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_log': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportLog']"}),
            'object_id': ('django.db.models.fields.IntegerField', [], {})
        },
        u'simple_import.importlog': {
            'Meta': {'object_name': 'ImportLog'},
            'date': ('django.db.models.fields.DateTimeField', [], {'auto_now_add': 'True', 'blank': 'True'}),
            'error_file': ('django.db.models.fields.files.FileField', [], {'max_length': '100', 'blank': 'True'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_file': ('django.db.models.fields.files.FileField', [], {'max_length': '100'}),
            'import_setting': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportSetting']"}),
            'import_type': ('django.db.models.fields.CharField', [], {'max_length': '1'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'update_key': ('django.db.models.fields.CharField', [], {'max_length': '200', 'blank': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'simple_import_log'", 'to': u"orm['auth.User']"})
        },
        u'simple_import.importsetting': {
            'Meta': {'unique_together': "(('user', 'content_type'),)", 'object_name': 'ImportSetting'},
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['contenttypes.ContentType']"}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['auth.User']"})
        },
        u'simple_import.relationalmatch': {
            'Meta': {'object_name': 'RelationalMatch'},
            'field_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            u'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'import_log': ('django.db.models.fields.related.ForeignKey', [], {'to': u"orm['simple_import.ImportLog']"}),
            'related_field_name': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'})
        }
    }

    complete_apps = ['simple_import']
########NEW FILE########
__FILENAME__ = models
from django.contrib.contenttypes.models import ContentType
from django.contrib.contenttypes import generic
from django.db import models
from django.conf import settings
from django.db import transaction
from django.utils.encoding import smart_text
import datetime

from simple_import.compat import AUTH_USER_MODEL

import sys
if sys.version_info >= (3,0):
    unicode = str

class ImportSetting(models.Model):
    """ Save some settings per user per content type """
    user = models.ForeignKey(AUTH_USER_MODEL)
    content_type = models.ForeignKey(ContentType)
    
    class Meta():
        unique_together = ('user', 'content_type',)


class ColumnMatch(models.Model):
    """ Match column names from the user uploaded file to the database """
    column_name = models.CharField(max_length=200)
    field_name = models.CharField(max_length=255, blank=True)
    import_setting = models.ForeignKey(ImportSetting)
    default_value = models.CharField(max_length=2000, blank=True)
    null_on_empty = models.BooleanField(default=False, help_text="If cell is blank, clear out the field setting it to blank.")
    header_position = models.IntegerField(help_text="Annoying way to order the columns to match the header rows")
    
    class Meta:
        unique_together = ('column_name', 'import_setting')
    
    def __unicode__(self):
        return unicode('{0} {1}'.format(self.column_name, self.field_name))

    def guess_field(self):
        """ Guess the match based on field names
        First look for an exact field name match
        then search defined alternative names
        then normalize the field name and check for match
        """
        model = self.import_setting.content_type.model_class()
        field_names = model._meta.get_all_field_names()
        if self.column_name in field_names:
            self.field_name = self.column_name
            return
        #TODO user defined alt names
        normalized_field_name = self.column_name.lower().replace(' ', '_')
        if normalized_field_name in field_names:
            self.field_name = self.column_name
        # Try verbose name
        for field_name in field_names:
            field = model._meta.get_field_by_name(field_name)[0]
            if hasattr(field, 'verbose_name'):
                if field.verbose_name.lower().replace(' ', '_') == normalized_field_name:
                    self.field_name = field_name


class ImportLog(models.Model):
    """ A log of all import attempts """
    name = models.CharField(max_length=255)
    user = models.ForeignKey(AUTH_USER_MODEL, editable=False, related_name="simple_import_log")
    date = models.DateTimeField(auto_now_add=True, verbose_name="Date Created")
    import_file = models.FileField(upload_to="import_file")
    error_file = models.FileField(upload_to="error_file", blank=True)
    import_setting = models.ForeignKey(ImportSetting, editable=False)
    import_type_choices = (
        ("N", "Create New Records"),
        ("U", "Create and Update Records"),
        ("O", "Only Update Records"),
    )
    import_type = models.CharField(max_length=1, choices=import_type_choices)
    update_key = models.CharField(max_length=200, blank=True)
    
    def __unicode__(self):
        return unicode(self.name)
    
    def clean(self):
        from django.core.exceptions import ValidationError
        filename = str(self.import_file).lower()
        if not filename[-3:] in ('xls', 'ods', 'csv', 'lsx'):
            raise ValidationError('Invalid file type. Must be xls, xlsx, ods, or csv.')
    
    @transaction.commit_on_success
    def undo(self):
        if self.import_type != "N":
            raise Exception("Cannot undo this type of import!")
        for obj in self.importedobject_set.all():
            if obj.content_object:
                obj.content_object.delete()
            obj.delete()

    @staticmethod
    def is_empty(value):
        """ Check `value` for emptiness by first comparing with None and then
        by coercing to string, trimming, and testing for zero length """
        return value is None or not len(smart_text(value).strip())
    
    def get_matches(self):
        """ Get each matching header row to database match
        Returns a ColumnMatch queryset"""
        header_row = self.get_import_file_as_list(only_header=True)
        match_ids = []
        
        for i, cell in enumerate(header_row):
            if self.is_empty(cell): # Sometimes we get blank headers, ignore them.
                continue
            
            try:
                match = ColumnMatch.objects.get(
                    import_setting = self.import_setting,
                    column_name = cell,
                )
            except ColumnMatch.DoesNotExist:
                match = ColumnMatch(
                    import_setting = self.import_setting,
                    column_name = cell,
                )
                match.guess_field()
            
            match.header_position = i
            match.save()
            
            match_ids += [match.id]
        
        return ColumnMatch.objects.filter(pk__in=match_ids).order_by('header_position')

    def get_import_file_as_list(self, only_header=False):
        file_ext = str(self.import_file).lower()[-3:]
        data = []
        
        self.import_file.seek(0)
        
        if file_ext == "xls":
            import xlrd
            import os
            
            wb = xlrd.open_workbook(file_contents=self.import_file.read())
            sh1 = wb.sheet_by_index(0)
            for rownum in range(sh1.nrows):
                row_values = []
                for cell in sh1.row(rownum):
                    # xlrd is too dumb to just check for dates. So we have to ourselves
                    if cell.ctype == 3: # 3 is date - http://www.lexicon.net/sjmachin/xlrd.html#xlrd.Cell-class
                        row_values += [datetime.datetime(*xlrd.xldate_as_tuple(cell.value, wb.datemode))]
                    else:
                        row_values += [cell.value]
                data += [row_values]
                if only_header:
                    break
        elif file_ext == "csv":
            import csv
            reader = csv.reader(self.import_file)
            for row in reader:
                data += [row]
                if only_header:
                    break
        elif file_ext == "lsx":
            from openpyxl.reader.excel import load_workbook
            # load_workbook actually accepts a file-like object for the filename param
            wb = load_workbook(filename=self.import_file, use_iterators = True)
            sheet = wb.get_active_sheet()
            for row in sheet.iter_rows():
                data_row = []
                for cell in row:
                    data_row += [cell.internal_value]
                data += [data_row]
                if only_header:
                    break
        elif file_ext == "ods":
            from .odsreader import ODSReader
            doc = ODSReader(self.import_file)
            table = doc.SHEETS.items()[0]

            # Remove blank columns that ods files seems to have
            blank_columns = []
            for i, header_cell in enumerate(table[1][0]):
                if self.is_empty(header_cell):
                    blank_columns += [i]
            # just an overly complicated way to remove these
            # indexes from a list
            for offset, blank_column in enumerate(blank_columns):
                for row in table[1]:
                    del row[blank_column - offset]

            if only_header:
                data += [table[1][0]]
            else:
                data += table[1]
        # Remove blank columns. We use the header row as a unique index. Can't handle blanks.
        columns_to_del = []
        for i, header_cell in enumerate(data[0]):
            if self.is_empty(header_cell):
                columns_to_del += [i]
        num_deleted = 0
        for column_to_del in columns_to_del:
            for row in data:
                del row[column_to_del - num_deleted]
            num_deleted += 1
        if only_header:
            return data[0]
        return data


class RelationalMatch(models.Model):
    """Store which unique field is being use to match.
    This can be used only to set a FK or one M2M relation
    on the import root model. It does not add them.
    With Multple rows set to the same field, you could set more
    than one per row.
    EX Lets say a student has an ID and username and both
    are marked as unique in Django orm. The user could reference
    that student by either."""
    import_log = models.ForeignKey(ImportLog)
    field_name = models.CharField(max_length=255) # Ex student_number_set
    related_field_name = models.CharField(max_length=255, blank=True) # Ex username


class ImportedObject(models.Model):
    import_log = models.ForeignKey(ImportLog)
    object_id = models.IntegerField()
    content_type = models.ForeignKey(ContentType)
    content_object = generic.GenericForeignKey('content_type', 'object_id')


########NEW FILE########
__FILENAME__ = odsreader
# Copyright 2011 Marco Conti

# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at

# 	http://www.apache.org/licenses/LICENSE-2.0

# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

# Thanks to grt for the fixes

import odf.opendocument
from odf.table import *
from odf.text import P

class ODSReader:

	# loads the file
	def __init__(self, file):
		self.doc = odf.opendocument.load(file)
		self.SHEETS = {}
		for sheet in self.doc.spreadsheet.getElementsByType(Table):
			self.readSheet(sheet)
	

	# reads a sheet in the sheet dictionary, storing each sheet as an array (rows) of arrays (columns)
	def readSheet(self, sheet):
		name = sheet.getAttribute("name")
		rows = sheet.getElementsByType(TableRow)
		arrRows = []
		
		# for each row
		for row in rows:
			row_comment = ""
			arrCells = []
			cells = row.getElementsByType(TableCell)
			
			# for each cell
			for cell in cells:
				# repeated value?
				repeat = cell.getAttribute("numbercolumnsrepeated")
				if(not repeat):
					repeat = 1
					
				ps = cell.getElementsByType(P)
				textContent = ""
								
				# for each text node
				for p in ps:
					for n in p.childNodes:
						if (n.nodeType == 3):
							textContent = textContent + unicode(n.data)
					
				if(textContent or textContent == ""):
					if(textContent == "" or textContent[0] != "#"): # ignore comments cells
						for rr in range(int(repeat)): # repeated?
							arrCells.append(textContent)
						
					else:
						row_comment = row_comment + textContent + " ";
						
			# if row contained something
			if(len(arrCells)):
				arrRows.append(arrCells)
				
			#else:
			#	print "Empty or commented row (", row_comment, ")"
		
		self.SHEETS[name] = arrRows
		
	# returns a sheet as an array (rows) of arrays (columns)
	def getSheet(self, name):
		return self.SHEETS[name]
########NEW FILE########
__FILENAME__ = tests
import os

from django.contrib.contenttypes.models import ContentType
from django.core.urlresolvers import reverse
from django.test import TestCase
from simple_import.compat import User
from simple_import.models import *
from django.core.files import File

class SimpleTest(TestCase):
    def setUp(self):
        user = User.objects.create_user('temporary', 'temporary@gmail.com', 'temporary')
        user.is_staff = True
        user.is_superuser = True
        user.save()
        self.client.login(username='temporary', password='temporary')
        self.absolute_path = os.path.join(os.path.dirname(__file__), 'static', 'test_import.xls')
        self.import_setting = ImportSetting.objects.create(
            user=user,
            content_type=ContentType.objects.get_for_model(ImportLog)
        )
        with open(self.absolute_path) as fp:
            self.import_log = ImportLog.objects.create(
                name=u'test',
                user=user,
                import_file = File(fp),
                import_setting = self.import_setting,
                import_type = u'N',
            )   
        
    def test_import(self):
        """ Make sure we can upload the file and match columns """
        import_log_ct_id = ContentType.objects.get_for_model(ImportLog).id
        
        self.assertEqual(ImportLog.objects.count(), 1)
        
        with open(self.absolute_path) as fp:
            response = self.client.post(reverse('simple_import-start_import'), {
                'name': 'This is a test',
                'import_file': fp,
                'import_type': "N",
                'model': import_log_ct_id}, follow=True)
        
        self.assertEqual(ImportLog.objects.count(), 2)
        
        self.assertRedirects(response, reverse('simple_import-match_columns', kwargs={'import_log_id': ImportLog.objects.all()[1].id}))
        self.assertContains(response, '<h1>Match Columns</h1>')
        # Check matching
        self.assertContains(response, '<option value="name" selected="selected">')
        self.assertContains(response, '<option value="user" selected="selected">')
        self.assertContains(response, '<option value="import_file" selected="selected">')
        self.assertContains(response, '<option value="import_setting">import setting (Required) (Related)</option>')
        # Check Sample Data
        self.assertContains(response, '/tmp/foo.xls')
    
    def test_match_columns(self):
        """ Test matching columns view  """
        self.assertEqual(ColumnMatch.objects.count(), 0)
        
        response = self.client.post(
            reverse('simple_import-match_columns', kwargs={'import_log_id': self.import_log.id}), {
            'columnmatch_set-TOTAL_FORMS':6,
            'columnmatch_set-INITIAL_FORMS':6,
            'columnmatch_set-MAX_NUM_FORMS':1000,
            'columnmatch_set-0-id':1,
            'columnmatch_set-0-column_name':'name',
            'columnmatch_set-0-import_setting':self.import_setting.id,
            'columnmatch_set-0-field_name':'name',
            'columnmatch_set-1-id':2,
            'columnmatch_set-1-column_name':'UseR',
            'columnmatch_set-1-import_setting':self.import_setting.id,
            'columnmatch_set-1-field_name':'user',
            'columnmatch_set-2-id':3,
            'columnmatch_set-2-column_name':'nothing',
            'columnmatch_set-2-import_setting':self.import_setting.id,
            'columnmatch_set-2-field_name':'',
            'columnmatch_set-3-id':4,
            'columnmatch_set-3-column_name':'import file',
            'columnmatch_set-3-import_setting':self.import_setting.id,
            'columnmatch_set-3-field_name':'import_file',
            'columnmatch_set-4-id':5,
            'columnmatch_set-4-column_name':'import_setting',
            'columnmatch_set-4-import_setting':self.import_setting.id,
            'columnmatch_set-4-field_name':'import_setting',
            'columnmatch_set-5-id':6,
            'columnmatch_set-5-column_name':'importtype',
            'columnmatch_set-5-import_setting':self.import_setting.id,
            'columnmatch_set-5-field_name':'import_type',
        }, follow=True)
        
        self.assertRedirects(response, reverse('simple_import-match_relations', kwargs={'import_log_id': self.import_log.id}))
        self.assertContains(response, '<h1>Match Relations and Prepare to Run Import</h1>')
        self.assertEqual(ColumnMatch.objects.count(), 6)
    
    def test_match_relations(self):
        """ Test matching relations view  """
        self.assertEqual(RelationalMatch.objects.count(), 0)
        
        response = self.client.post(
            reverse('simple_import-match_relations', kwargs={'import_log_id': self.import_log.id}), {
            'relationalmatch_set-TOTAL_FORMS':2,
            'relationalmatch_set-INITIAL_FORMS':2,
            'relationalmatch_set-MAX_NUM_FORMS':1000,
            'relationalmatch_set-0-id':1,
            'relationalmatch_set-0-import_log':self.import_log.id,
            'relationalmatch_set-0-field_name':'user',
            'relationalmatch_set-0-related_field_name':'id',
            'relationalmatch_set-1-id':2,
            'relationalmatch_set-1-import_log':self.import_log.id,
            'relationalmatch_set-1-field_name':'import_setting',
            'relationalmatch_set-1-related_field_name':'id',
        }, follow=True)

        self.assertRedirects(response, reverse('simple_import-do_import', kwargs={'import_log_id': self.import_log.id}))
        self.assertContains(response, '<h1>Import Results</h1>')
        
        self.assertEqual(RelationalMatch.objects.count(), 2)


########NEW FILE########
__FILENAME__ = urls
from django.conf.urls import *
from simple_import import views

urlpatterns = patterns('',
    url('^start_import/$', views.start_import, name='simple_import-start_import'),
    url('^match_columns/(?P<import_log_id>\d+)/$', views.match_columns, name='simple_import-match_columns'),
    url('^match_relations/(?P<import_log_id>\d+)/$', views.match_relations, name='simple_import-match_relations'),
    url('^do_import/(?P<import_log_id>\d+)/$', views.do_import, name='simple_import-do_import'),
)

########NEW FILE########
__FILENAME__ = views
from django import forms
from django.contrib.contenttypes.models import ContentType
from django.contrib.admin.models import LogEntry, ADDITION, CHANGE
from django.contrib.admin.views.decorators import staff_member_required
from django.conf import settings
from django.core.exceptions import SuspiciousOperation
from django.core.urlresolvers import reverse
from django.db.models import Q, ForeignKey
from django.db import transaction
from django.db import IntegrityError
from django.core.exceptions import ObjectDoesNotExist
from django.forms.models import inlineformset_factory
from django.http import HttpResponseRedirect
from django.shortcuts import render_to_response, get_object_or_404
from django.template import RequestContext
import sys
from django.db.models.fields import AutoField
from django.utils.encoding import smart_text

from simple_import.compat import User
from simple_import.models import ImportLog, ImportSetting, ColumnMatch, ImportedObject, RelationalMatch
from simple_import.forms import ImportForm, MatchForm, MatchRelationForm

if sys.version_info >= (3,0):
    unicode = str

def validate_match_columns(import_log, model_class, header_row):
    """ Perform some basic pre import validation to make sure it's
    even possible the import can work
    Returns list of errors
    """
    errors = []
    column_matches = import_log.import_setting.columnmatch_set.all()
    field_names = model_class._meta.get_all_field_names()
    for field_name in field_names:
        field_object, model, direct, m2m = model_class._meta.get_field_by_name(field_name)
        # Skip if update only and skip ptr which suggests it's a django inherited field
        # Also some hard coded ones for Django Auth
        if import_log.import_type != "O" and field_name[-3:] != "ptr" and \
            not field_name in ['password', 'date_joined', 'last_login']: 
            if (direct and model and not field_object.blank) or (not getattr(field_object, "blank", True)):
                field_matches = column_matches.filter(field_name=field_name)
                match_in_header = False
                if field_matches:
                    for field_match in field_matches:
                        if field_match.column_name.lower() in header_row:
                            match_in_header = True
                    if not match_in_header:
                        errors += [u"{0} is required but is not in your spreadsheet. ".format(field_object.verbose_name.title())]
                else:
                    errors += [u"{0} is required but has no match.".format(field_object.verbose_name.title())]
    return errors


def get_custom_fields_from_model(model_class):
    """ django-custom-fields support
    """
    if 'custom_field' in settings.INSTALLED_APPS:
        from custom_field.models import CustomField
        try:
            content_type = ContentType.objects.get(model=model_class._meta.module_name,app_label=model_class._meta.app_label)
        except ContentType.DoesNotExist:
            content_type = None
        custom_fields = CustomField.objects.filter(content_type=content_type)
        return custom_fields


@staff_member_required
def match_columns(request, import_log_id):
    """ View to match import spreadsheet columns with database fields
    """
    import_log = get_object_or_404(ImportLog, id=import_log_id)
    
    if not request.user.is_superuser and import_log.user != request.user:
        raise SuspiciousOperation("Non superuser attempting to view other users import")
    
    # need to generate matches if they don't exist already
    existing_matches = import_log.get_matches()
    
    MatchFormSet = inlineformset_factory(ImportSetting, ColumnMatch, form=MatchForm, extra=0)
    
    import_data = import_log.get_import_file_as_list()
    header_row = [x.lower() for x in import_data[0]] # make all lower 
    sample_row = import_data[1]
    errors = []
    
    model_class = import_log.import_setting.content_type.model_class()
    field_names = model_class._meta.get_all_field_names()
    for field_name in field_names:
        field_object, model, direct, m2m = model_class._meta.get_field_by_name(field_name)
        # We can't add a new AutoField and specify it's value
        if import_log.import_type == "N" and isinstance(field_object, AutoField):
            field_names.remove(field_name)
        
    if request.method == 'POST':
        formset = MatchFormSet(request.POST, instance=import_log.import_setting)
        if formset.is_valid():
            formset.save()
            if import_log.import_type in ["U", "O"]:
                update_key = request.POST.get('update_key', '')
                
                if update_key:
                    field_name = import_log.import_setting.columnmatch_set.get(column_name=update_key).field_name
                    if field_name:
                        field_object, model, direct, m2m = model_class._meta.get_field_by_name(field_name)
                        
                        if direct and field_object.unique:
                            import_log.update_key = update_key
                            import_log.save()
                        else:
                            errors += ['Update key must be unique. Please select a unique field.']
                    else:
                        errors += ['Update key must matched with a column.']
                else:
                    errors += ['Please select an update key. This key is used to linked records for updating.']
            errors += validate_match_columns(
                import_log,
                model_class,
                header_row)
            
            all_field_names = []
            for clean_data in formset.cleaned_data:
                if clean_data['field_name']:
                    if clean_data['field_name'] in all_field_names:
                        errors += ["{0} is duplicated.".format(clean_data['field_name'])]
                    all_field_names += [clean_data['field_name']]
            if not errors:
                return HttpResponseRedirect(reverse(
                    match_relations,
                    kwargs={'import_log_id': import_log.id}))
    else:
        formset = MatchFormSet(instance=import_log.import_setting, queryset=existing_matches)
    
    field_choices = (('', 'Do Not Use'),)
    for field_name in field_names:
        field_object, model, direct, m2m = model_class._meta.get_field_by_name(field_name)
        add = True
        
        if direct:
            field_verbose = field_object.verbose_name
        else:
            field_verbose = field_name
        
        if direct and  not field_object.blank:
            field_verbose += " (Required)"
        if direct and field_object.unique:
            field_verbose += " (Unique)"
        if m2m or isinstance(field_object, ForeignKey):
            field_verbose += " (Related)"
        elif not direct:
            add = False
        
        if add:
            field_choices += ((field_name, field_verbose),)
    
    # Include django-custom-field support
    custom_fields = get_custom_fields_from_model(model_class)
    if custom_fields:
        for custom_field in custom_fields:
            field_choices += (("simple_import_custom__{0}".format(custom_field),
                           "{0} (Custom)".format(custom_field)),)
    # Include defined methods
    # Model must have a simple_import_methods defined
    if hasattr(model_class, 'simple_import_methods'):
        for import_method in model_class.simple_import_methods:
            field_choices += (("simple_import_method__{0}".format(import_method),
                               "{0} (Method)".format(import_method)),)
    # User model should allow set password
    if issubclass(model_class, User):
        field_choices += (("simple_import_method__{0}".format('set_password'),
                               "Set Password (Method)"),) 
    
    for i, form in enumerate(formset):
        form.fields['field_name'].widget = forms.Select(choices=(field_choices))
        form.sample = sample_row[i]
    
    return render_to_response(
        'simple_import/match_columns.html',
        {'import_log': import_log, 'formset': formset, 'errors': errors},
        RequestContext(request, {}),)


def get_direct_fields_from_model(model_class):
    direct_fields = []
    all_fields_names = model_class._meta.get_all_field_names()
    for field_name in all_fields_names:
        field = model_class._meta.get_field_by_name(field_name)
        # Direct, not m2m, not FK
        if field[2] and not field[3] and field[0].__class__.__name__ != "ForeignKey":
            direct_fields += [field[0]]
    return direct_fields


@staff_member_required
def match_relations(request, import_log_id):
    import_log = get_object_or_404(ImportLog, id=import_log_id)
    model_class = import_log.import_setting.content_type.model_class()
    matches = import_log.get_matches()
    field_names = []
    choice_set = []
    
    for match in matches.exclude(field_name=""):
        field_name = match.field_name
        
        if not field_name.startswith('simple_import_custom__') and \
                not field_name.startswith('simple_import_method__'):
            field, model, direct, m2m = model_class._meta.get_field_by_name(field_name)
            
            if m2m or isinstance(field, ForeignKey): 
                RelationalMatch.objects.get_or_create(
                    import_log=import_log,
                    field_name=field_name)
                
                field_names.append(field_name)
                choices = ()
                for field in get_direct_fields_from_model(field.related.parent_model()):
                    if field.unique:
                        choices += ((field.name, unicode(field.verbose_name)),)
                choice_set += [choices]
    
    existing_matches = import_log.relationalmatch_set.filter(field_name__in=field_names)
    
    MatchRelationFormSet = inlineformset_factory(
        ImportLog,
        RelationalMatch,
        form=MatchRelationForm, extra=0)
    
    if request.method == 'POST':
        formset = MatchRelationFormSet(request.POST, instance=import_log)
        
        if formset.is_valid():
            formset.save()
            
            url = reverse('simple_import-do_import',
                kwargs={'import_log_id': import_log.id})
            
            if 'commit' in request.POST:
                url += "?commit=True"
            
            return HttpResponseRedirect(url)
    else:
        formset = MatchRelationFormSet(instance=import_log)
    
    for i, form in enumerate(formset.forms):
        choices = choice_set[i]
        form.fields['related_field_name'].widget = forms.Select(choices=choices)
    
    return render_to_response(
        'simple_import/match_relations.html',
        {'formset': formset,
         'existing_matches': existing_matches},
        RequestContext(request, {}),)

def set_field_from_cell(import_log, new_object, header_row_field_name, cell):
    """ Set a field from a import cell. Use referenced fields the field
    is m2m or a foreign key.
    """
    if (not header_row_field_name.startswith('simple_import_custom__') and
            not header_row_field_name.startswith('simple_import_method__')):
        field, model, direct, m2m =  new_object._meta.get_field_by_name(header_row_field_name)
        if m2m:
            new_object.simple_import_m2ms[header_row_field_name] = cell
        elif isinstance(field, ForeignKey):
            related_field_name = RelationalMatch.objects.get(import_log=import_log, field_name=field.name).related_field_name
            related_model = field.related.parent_model
            related_object = related_model.objects.get(**{related_field_name:cell})
            setattr(new_object, header_row_field_name, related_object)
        elif field.choices and getattr(settings, 'SIMPLE_IMPORT_LAZY_CHOICES', True):
            # Prefer database values over choices lookup
            database_values, verbose_values = zip(*field.choices)
            if cell in database_values:
                setattr(new_object, header_row_field_name, cell)
            elif cell in verbose_values:
                for choice in field.choices:
                    if smart_text(cell) == smart_text(choice[1]):
                        setattr(new_object, header_row_field_name, choice[0])
        else:
            setattr(new_object, header_row_field_name, cell)
    
    
def set_method_from_cell(import_log, new_object, header_row_field_name, cell):
    """ Run a method from a import cell.
    """
    if (not header_row_field_name.startswith('simple_import_custom__') and
            not header_row_field_name.startswith('simple_import_method__')):
        pass
    elif header_row_field_name.startswith('simple_import_custom__'):
        new_object.set_custom_value(header_row_field_name[22:], cell)
    elif header_row_field_name.startswith('simple_import_method__'):
        getattr(new_object, header_row_field_name[22:])(cell)
       

@staff_member_required
def do_import(request, import_log_id):
    """ Import the data!
    """
    import_log = get_object_or_404(ImportLog, id=import_log_id)
    if import_log.import_type == "N" and 'undo' in request.GET and request.GET['undo'] == "True":
        import_log.undo()
        return HttpResponseRedirect(reverse(
                    do_import,
                    kwargs={'import_log_id': import_log.id}) + '?success_undo=True')
    
    if 'success_undo' in request.GET and request.GET['success_undo'] == "True":
        success_undo = True
    else:
        success_undo = False
    
    model_class = import_log.import_setting.content_type.model_class()
    import_data = import_log.get_import_file_as_list()
    header_row = import_data.pop(0)
    header_row_field_names = []
    header_row_default = []
    header_row_null_on_empty = []
    error_data = [header_row + ['Error Type', 'Error Details']]
    create_count = 0
    update_count = 0
    fail_count = 0
    if 'commit' in request.GET and request.GET['commit'] == "True":
        commit = True
    else:
        commit = False
    
    key_column_name = None
    if import_log.update_key and import_log.import_type in ["U", "O"]:
        key_match = import_log.import_setting.columnmatch_set.get(column_name=import_log.update_key)
        key_column_name = key_match.column_name
        key_field_name = key_match.field_name
    for i, cell in enumerate(header_row):
        match = import_log.import_setting.columnmatch_set.get(column_name=cell)
        header_row_field_names += [match.field_name]
        header_row_default += [match.default_value]
        header_row_null_on_empty += [match.null_on_empty]
        if key_column_name == cell.lower():
            key_index = i
    
    with transaction.commit_manually():
        for row in import_data:
            try:
                is_created = True
                if import_log.import_type == "N":
                    new_object = model_class()
                elif import_log.import_type == "O":
                    filters = {key_field_name: row[key_index]}
                    new_object = model_class.objects.get(**filters)
                    is_created = False
                elif import_log.import_type == "U":
                    filters = {key_field_name: row[key_index]}
                    try:
                        new_object = model_class.objects.get(**filters)
                        is_created = False
                    except model_class.DoesNotExist:
                        new_object = model_class()
                new_object.simple_import_m2ms = {} # Need to deal with these after saving
                for i, cell in enumerate(row):
                    if header_row_field_names[i]: # skip blank
                        if not import_log.is_empty(cell) or header_row_null_on_empty[i]:
                            set_field_from_cell(import_log, new_object, header_row_field_names[i], cell)
                        elif header_row_default[i]:
                            set_field_from_cell(import_log, new_object, header_row_field_names[i], header_row_default[i])
                new_object.save()

                for i, cell in enumerate(row):
                    if header_row_field_names[i]: # skip blank
                        if not import_log.is_empty(cell) or header_row_null_on_empty[i]:
                            set_method_from_cell(import_log, new_object, header_row_field_names[i], cell)
                        elif header_row_default[i]:
                            set_method_from_cell(import_log, new_object, header_row_field_names[i], header_row_default[i])
                # set_custom_value() calls save() on its own, but the same cannot be assumed
                # for other methods, e.g. set_password()
                new_object.save()

                for key in new_object.simple_import_m2ms.keys():
                    value = new_object.simple_import_m2ms[key]
                    m2m = getattr(new_object, key)
                    m2m_model = type(m2m.model())
                    related_field_name = RelationalMatch.objects.get(import_log=import_log, field_name=key).related_field_name
                    m2m_object = m2m_model.objects.get(**{related_field_name:value})
                    m2m.add(m2m_object)
                
                if is_created:
                    LogEntry.objects.log_action(
                        user_id         = request.user.pk, 
                        content_type_id = ContentType.objects.get_for_model(new_object).pk,
                        object_id       = new_object.pk,
                        object_repr     = smart_text(new_object), 
                        action_flag     = ADDITION
                    )
                    create_count += 1
                else:
                    LogEntry.objects.log_action(
                        user_id         = request.user.pk, 
                        content_type_id = ContentType.objects.get_for_model(new_object).pk,
                        object_id       = new_object.pk,
                        object_repr     = smart_text(new_object), 
                        action_flag     = CHANGE
                    )
                    update_count += 1
                ImportedObject.objects.create(
                    import_log = import_log,
                    object_id = new_object.pk,
                    content_type = import_log.import_setting.content_type)
            except IntegrityError:
                exc = sys.exc_info()
                error_data += [row + ["Integrity Error", smart_text(exc[1][1])]]
                fail_count += 1
            except ObjectDoesNotExist:
                exc = sys.exc_info()
                error_data += [row + ["No Record Found to Update", smart_text(exc[1])]]
                fail_count += 1
            except ValueError:
                exc = sys.exc_info()
                if unicode(exc[1]).startswith('invalid literal for int() with base 10'):
                    error_data += [row + ["Incompatible Data - A number was expected, but a character was used", smart_text(exc[1])]] 
                else:
                    error_data += [row + ["Value Error", smart_text(exc[1])]]
                fail_count += 1
            except:
                exc = sys.exc_info()
                error_data += [row + ["Unknown Error", smart_text(exc[1])]]
                fail_count += 1
        if commit:
            transaction.commit()
        else:
            transaction.rollback()
    
            
    if fail_count:
        from io import StringIO
        from django.core.files.base import ContentFile
        from openpyxl.workbook import Workbook
        from openpyxl.writer.excel import save_virtual_workbook
        
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.title = "Errors"
        filename = 'Errors.xlsx'
        for row in error_data:
            ws.append(row)
        buf = StringIO()
        # Not Python 3 compatible 
        #buf.write(str(save_virtual_workbook(wb)))
        import_log.error_file.save(filename, ContentFile(save_virtual_workbook(wb)))
        import_log.save()
    
    return render_to_response(
        'simple_import/do_import.html',
        {
            'error_data': error_data,
            'create_count': create_count,
            'update_count': update_count,
            'fail_count': fail_count,
            'import_log': import_log,
            'commit': commit,
            'success_undo': success_undo,},
        RequestContext(request, {}),)


@staff_member_required
def start_import(request):
    """ View to create a new import record
    """
    if request.method == 'POST':
        form = ImportForm(request.POST, request.FILES)
        if form.is_valid():
            import_log = form.save(commit=False)
            import_log.user = request.user
            import_log.import_setting, created = ImportSetting.objects.get_or_create(
                user=request.user,
                content_type=form.cleaned_data['model'],
            )
            import_log.save()
            return HttpResponseRedirect(reverse(match_columns, kwargs={'import_log_id': import_log.id}))
    else:
        form = ImportForm()
    if not request.user.is_superuser:
        form.fields["model"].queryset = ContentType.objects.filter(
            Q(permission__group__user=request.user, permission__codename__startswith="change_") |
            Q(permission__user=request.user, permission__codename__startswith="change_")).distinct()
    
    return render_to_response('simple_import/import.html', {'form':form,}, RequestContext(request, {}),)


########NEW FILE########
__FILENAME__ = test_settings
import os
PROJECT_DIR = os.path.dirname(__file__)

STATIC_URL = PROJECT_DIR + '/static/'

DATABASES = {
    'default': {
        'ENGINE': 'django.db.backends.sqlite3',
        'NAME': 'testdb',
    }
}

INSTALLED_APPS = (
    'simple_import',
    'django.contrib.auth',
    'django.contrib.contenttypes',
    'django.contrib.sessions',
    'django.contrib.sites',
    'django.contrib.messages',
    'django.contrib.staticfiles',
    'django.contrib.admin',
    'south',
)

ROOT_URLCONF = "test_urls"

########NEW FILE########
__FILENAME__ = test_urls
from django.conf.urls import patterns, include, url

# Uncomment the next two lines to enable the admin:
from django.contrib import admin
admin.autodiscover()

urlpatterns = patterns('',
    # Examples:
    # url(r'^$', 'test_project.views.home', name='home'),
    # url(r'^test_project/', include('test_project.foo.urls')),

    # Uncomment the admin/doc line below to enable admin documentation:
    # url(r'^admin/doc/', include('django.contrib.admindocs.urls')),

    # Uncomment the next line to enable the admin:
    url(r'^admin/', include(admin.site.urls)),
    url(r'^simple_import/', include('simple_import.urls'))
)

########NEW FILE########
