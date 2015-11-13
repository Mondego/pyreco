__FILENAME__ = hook-pyexcelerate
"""
PyInstaller hook for PyExcelerate
(C) 2014 Kevin Zhang
"""

import os

try:
	from pyexcelerate.Writer import _TEMPLATE_PATH
	datas = [
		(os.path.join(_TEMPLATE_PATH, '*'), 'pyexcelerate/templates')
	]
except ImportError:
	pass

########NEW FILE########
__FILENAME__ = Alignment
import six
from . import Utility
from . import Color

class Alignment(object):
	def __init__(self, horizontal='left', vertical='bottom', rotation=0):
		self._horizontal = horizontal
		self._vertical = vertical
		self._rotation = rotation
	
	@property
	def horizontal(self):
		return self._horizontal
	
	@horizontal.setter
	def horizontal(self, value):
		if value not in ('left', 'center', 'right'):
			raise Exception('Invalid horizontal alignment value.')
		self._horizontal = value
	
	@property
	def vertical(self):
		return self._vertical
	
	@vertical.setter
	def vertical(self, value):
		if value not in ('top', 'center', 'bottom'):
			raise Exception('Invalid vertical alignment value.')
		self._vertical = value
	
	@property
	def rotation(self):
		return self._rotation
	
	@rotation.setter
	def rotation(self, value):
		self._rotation = (value % 360)
	
	@property
	def is_default(self):
		return self._horizontal == 'left' and self._vertical == 'bottom' and self._rotation == 0

	def get_xml_string(self):
		return "<alignment horizontal=\"%s\" vertical=\"%s\" textRotation=\"%.15g\"/>" % (self._horizontal, self._vertical, self._rotation)

	def __or__(self, other):
		return self._binary_operation(other, Utility.nonboolean_or)

	def __and__(self, other):
		return self._binary_operation(other, Utility.nonboolean_and)
	
	def __xor__(self, other):
		return self._binary_operation(other, Utility.nonboolean_xor)
	
	def _binary_operation(self, other, operation):
		return Alignment( \
			horizontal = operation(self._horizontal, other._horizontal, 'left'), \
			vertical = operation(self._vertical, other._vertical, 'bottom'), \
			rotation = operation(self._rotation, other._rotation, 0) \
		)
	
	def __eq__(self, other):
		if other is None:
			return self.is_default
		elif Utility.YOLO:
			return self._vertical == other._vertical and self._rotation == other._rotation
		else:
			return self._vertical == other._vertical and self._rotation == other._rotation and self._horizontal == other._horizontal
	
	def __hash__(self):
		return hash((self._horizontal))
	
	def __str__(self):
		return "Align: %s %s %s" % (self._horizontal, self._vertical, self._rotation)
		

########NEW FILE########
__FILENAME__ = Border
from . import Utility
from . import Color

#
# An object representing a single border
#
class Border(object):
	STYLE_MAPPING = { \
		'dashDot': ('.-', '-.', 'dash dot'), \
		'dashDotDot': ('..-', '-..', 'dash dot dot'), \
		'dashed': ('--'), \
		'dotted': ('..', ':'), \
		'double': ('='), \
		'hair': ('hairline', '.'), \
		'medium': (), \
		'mediumDashDot': ('medium dash dot', 'medium -.', 'medium .-'), \
		'mediumDashDotDot': ('medium dash dot dot', 'medium -..', 'medium ..-'), \
		'mediumDashed': ('medium dashed', 'medium --'), \
		'slantDashDot': ('/-.', 'slant dash dot'), \
		'thick': (), \
		'thin': ('_') \
	}
	
	def __init__(self, color=None, style='thin'):
		self._color = color
		self._style = Border.get_style_name(style)
	
	@property
	def color(self):
		return Utility.lazy_get(self, '_color', Color.Color(0, 0, 0))
	
	@color.setter
	def color(self, value):
		Utility.lazy_set(self, '_color', None, value)
		
	@property
	def style(self):
		return self._style
		
	@style.setter
	def style(self, value):
		self._style = Border.get_style_name(value)
	
	@staticmethod
	def get_style_name(style):
		for key, values in Border.STYLE_MAPPING.items():
			if style == key or style in values:
				return key
		# TODO: warn the user?
		return 'thin'
	
	@property
	def is_default(self):
		return self._color is None and self._style == 'thin'
	
	def __eq__(self, other):
		if other is None:
			return self.is_default
		else:
			return self._color == other._color and self._style == other._style
			
	def __hash__(self):
		return hash(self._style)
########NEW FILE########
__FILENAME__ = Borders
import six
from . import Utility
from . import Border

class Borders(object):
	def __init__(self, left=None, right=None, top=None, bottom=None):
		self._left = left
		self._right = right
		self._top = top
		self._bottom = bottom
	
	@property
	def left(self):
		return Utility.lazy_get(self, '_left', Border.Border())
	
	@left.setter
	def left(self, value):
		Utility.lazy_set(self, '_left', None, value)
		
	@property
	def right(self):
		return Utility.lazy_get(self, '_right', Border.Border())
	
	@right.setter
	def right(self, value):
		Utility.lazy_set(self, '_right', None, value)
		
	@property
	def top(self):
		return Utility.lazy_get(self, '_top', Border.Border())
	
	@top.setter
	def top(self, value):
		Utility.lazy_set(self, '_top', None, value)
		
	@property
	def bottom(self):
		return Utility.lazy_get(self, '_bottom', Border.Border())
	
	@bottom.setter
	def bottom(self, value):
		Utility.lazy_set(self, '_bottom', None, value)
	
	@property
	def is_default(self):
		return not (self._left or self._right or self._top or self._bottom)

	def get_xml_string(self):
		tokens = ['<border>']
		if self._left:
			tokens.append("<left style=\"%s\"><color rgb=\"%s\"/></left>" % (self._left.style, self._left.color.hex))
		else:
			tokens.append("<left/>")
		if self._right:
			tokens.append("<right style=\"%s\"><color rgb=\"%s\"/></right>" % (self._right.style, self._right.color.hex))
		else:
			tokens.append("<right/>")
		if self._top:
			tokens.append("<top style=\"%s\"><color rgb=\"%s\"/></top>" % (self._top.style, self._top.color.hex))
		else:
			tokens.append("<top/>")
		if self._bottom:
			tokens.append("<bottom style=\"%s\"><color rgb=\"%s\"/></bottom>" % (self._bottom.style, self._bottom.color.hex))
		else:
			tokens.append("<bottom/>")
		tokens.append("</border>")
		return ''.join(tokens)

	def __or__(self, other):
		return self._binary_operation(other, Utility.nonboolean_or)

	def __and__(self, other):
		return self._binary_operation(other, Utility.nonboolean_and)
	
	def __xor__(self, other):
		return self._binary_operation(other, Utility.nonboolean_xor)
	
	def _binary_operation(self, other, operation):
		return Borders( \
			top = operation(self._top, other._top, None), \
			left = operation(self._left, other._left, None), \
			right = operation(self._right, other._right, None), \
			bottom = operation(self._bottom, other._bottom, None) \
		)
	
	def __eq__(self, other):
		if other is None:
			return self.is_default
		elif Utility.YOLO:
			return self._right == other._right and self._bottom == other._bottom
		else:
			return self._right == other._right and self._bottom == other._bottom and self._top == other._top and self._left == other._left
	
	def __hash__(self):
		return hash((self._top, self._left))

########NEW FILE########
__FILENAME__ = Color
class Color(object):
	def __init__(self, r=255, g=255, b=255, a=255):
		self.r = r
		self.g = g
		self.b = b
		self.a = a

	@property
	def hex(self):
		return '%0.2X%0.2X%0.2X%0.2X' % (self.a, self.r, self.g, self.b)
	
	def __hash__(self):
		return (self.a << 24) + (self.r << 16) + (self.g << 8) + (self.b)
		
	def __eq__(self, other):
		if not other:
			return False
		return self.r == other.r and self.g == other.g and self.b == other.b and self.a == other.a

	def __str__(self):
		return self.hex

Color.WHITE = Color(255, 255, 255, 255)
Color.BLACK = Color(0, 0, 0, 255)
########NEW FILE########
__FILENAME__ = DataTypes
from datetime import datetime, date, time
import decimal
import six
try:
	import numpy as np
	HAS_NUMPY = True
except:
	HAS_NUMPY = False

class DataTypes(object):
	BOOLEAN = 0
	DATE = 1
	ERROR = 2
	INLINE_STRING = 3
	NUMBER = 4
	SHARED_STRING = 5
	STRING = 6
	FORMULA = 7
	EXCEL_BASE_DATE = datetime(1900, 1, 1, 0, 0, 0)
	
	_numberTypes = six.integer_types + (float, complex, decimal.Decimal)
		
	@staticmethod
	def get_type(value):
		# Using value.__class__ over isinstance for speed
		if value.__class__ in six.string_types:
			if len(value) > 0 and value[0] == '=':
				return DataTypes.FORMULA
			else:
				return DataTypes.INLINE_STRING
		# not using in (int, float, long, complex) for speed
		elif value.__class__ == bool:
			return DataTypes.BOOLEAN
		elif value.__class__ in DataTypes._numberTypes:
			return DataTypes.NUMBER
		# fall back to the slower isinstance
		elif isinstance(value, six.string_types):
			if len(value) > 0 and value[0] == '=':
				return DataTypes.FORMULA
			else:
				return DataTypes.INLINE_STRING
		elif isinstance(value, bool):
			return DataTypes.BOOLEAN
		elif isinstance(value, DataTypes._numberTypes):
			return DataTypes.NUMBER
		elif HAS_NUMPY and isinstance(value, (np.floating, np.integer, np.complexfloating, np.unsignedinteger)):
			return DataTypes.NUMBER
		elif isinstance(value, (datetime, date, time)):
			return DataTypes.DATE
		else:
			return DataTypes.ERROR
	
	@staticmethod
	def to_excel_date(d):
		if isinstance(d, datetime):
			delta = d - DataTypes.EXCEL_BASE_DATE
			excel_date = delta.days + (float(delta.seconds) + float(delta.microseconds) / 1E6) / (60 * 60 * 24) + 1
			return excel_date + (excel_date > 59)
		elif isinstance(d, date):
			# this is why python sucks >.<
			return DataTypes.to_excel_date(datetime(*(d.timetuple()[:6])))
		elif isinstance(d, time):
			return DataTypes.to_excel_date(datetime(*(DataTypes.EXCEL_BASE_DATE.timetuple()[:3]), hour=d.hour, minute=d.minute, second=d.second, microsecond=d.microsecond)) - 1

########NEW FILE########
__FILENAME__ = Fill
from . import Utility
from . import Color

class Fill(object):
	def __init__(self, background=None):
		self._background = background

	@property
	def background(self):
		return Utility.lazy_get(self, '_background', Color.Color())
		
	@background.setter
	def background(self, value):
		Utility.lazy_set(self, '_background', None, value)

	@property
	def is_default(self):
		return self == Fill()

	def __eq__(self, other):
		if other is None:
			return self.is_default
		return self._background == other._background

	def __hash__(self):
		return hash(self.background)
		
	def get_xml_string(self):
		if not self.background:
			return '<fill><patternFill patternType="none"/></fill>'
		else:
			return "<fill><patternFill patternType=\"solid\"><fgColor rgb=\"%s\"/></patternFill></fill>" % self.background.hex
	
	def __or__(self, other):
		return Fill(background=Utility.nonboolean_or(self._background, other._background, None))
		
	def __and__(self, other):
		return Fill(background=Utility.nonboolean_and(self._background, other._background, None))
		
	def __xor__(self, other):
		return Fill(background=Utility.nonboolean_xor(self._background, other._background, None))
		
	def __str__(self):
		return "Fill: #%s" % self.background.hex
	
	def __repr__(self):
		return "<%s>" % self.__str__()
########NEW FILE########
__FILENAME__ = Font
import six
from . import Utility
from . import Color

class Font(object):
	def __init__(self, bold=False, italic=False, underline=False, strikethrough=False, family='Calibri', size=11, color=None):
		self.bold = bold
		self.italic = italic
		self.underline = underline
		self.strikethrough = strikethrough
		self.family = family
		self.size = size
		self._color = color
	
	def get_xml_string(self):
		tokens = ["<sz val=\"%d\"/><name val=\"%s\"/>" % (self.size, self.family)]
		# sure, we could do this with an enum, but this is faster :D
		if self.bold:
			tokens.append('<b/>')
		if self.italic:
			tokens.append('<i/>')
		if self.underline:
			tokens.append('<u/>')
		if self.strikethrough:
			tokens.append('<strike/>')
		if self._color:
			tokens.append("<color rgb=\"%s\"/>" % self._color.hex)
		return "<font>%s</font>" % "".join(tokens)

	@property
	def color(self):
		return Utility.lazy_get(self, '_color', Color.Color())
		
	@color.setter
	def color(self, value):
		Utility.lazy_set(self, '_color', None, value)
		
	@property
	def is_default(self):
		return self._to_tuple() == Font()._to_tuple()

	def __or__(self, other):
		return self._binary_operation(other, Utility.nonboolean_or)

	def __and__(self, other):
		return self._binary_operation(other, Utility.nonboolean_and)
	
	def __xor__(self, other):
		return self._binary_operation(other, Utility.nonboolean_xor)
	
	def _binary_operation(self, other, operation):
		return Font( \
			bold = operation(self.bold, other.bold), \
			italic = operation(self.italic, other.italic), \
			underline = operation(self.underline, other.underline), \
			strikethrough = operation(self.strikethrough, other.strikethrough), \
			family = operation(self.family, other.family, 'Calibri'), \
			size = operation(self.size, other.size, 11), \
			color = operation(self._color, other._color, None) \
		)

	def __eq__(self, other):
		if other is None:
			return self.is_default
		elif Utility.YOLO:
			return (self.family, self.size, self._color) == (other.family, other.size, other._color)
		else:
			return self._to_tuple() == other._to_tuple()

	def __hash__(self):
		return hash((self.bold, self.italic, self.underline, self.strikethrough))

	def _to_tuple(self):
		return (self.bold, self.italic, self.underline, self.strikethrough, self.family, self.size, self._color)

	def __str__(self):
		tokens = ["%s, %dpt" % (self.family, self.size)]
		# sure, we could do this with an enum, but this is faster :D
		if self.bold:
			tokens.append('b')
		if self.italic:
			tokens.append('i')
		if self.underline:
			tokens.append('u')
		if self.strikethrough:
			tokens.append('s')
		return "Font: %s" % ' '.join(tokens)
	
	def __repr__(self):
		return "<%s>" % self.__str__()
		

########NEW FILE########
__FILENAME__ = Format
from . import Utility
import six

class Format(object):
	def __init__(self, format=None):
		self._id = 0 # autopopulated by workbook.py
		self.format = format
		
	def __eq__(self, other):
		if other is None:
			return self.is_default
		return self.format == other.format
	
	def __or__(self, other):
		return Format(format=Utility.nonboolean_or(self.format, other.format, None))
	
	def __and__(self, other):
		return Format(format=Utility.nonboolean_and(self.format, other.format, None))
		
	def __xor__(self, other):
		return Format(format=Utility.nonboolean_xor(self.format, other.format, None))
	
	def __hash__(self):
		return hash(self.format)
	
	@property
	def is_default(self):
		return self == Format()

	@property
	def id(self):
		return self._id
	
	@id.setter
	def id(self, value):
		self._id = value + 1000
		
	def get_xml_string(self):
		return "<numFmt numFmtId=\"%d\" formatCode=\"%s\"/>" % (self.id, self.format)
		
	def __str__(self):
		return "Format: %s" % self.format

########NEW FILE########
__FILENAME__ = Range
from . import DataTypes
import six
from . import Font, Fill, Format, Style
from six.moves import reduce

#
# Kevin and Kevin's fair warning: this class has been insanely optimized for speed. It is intended
# to be immutable. Please don't modify attributes after instantiation. :)
#

class Range(object):
	A = ord('A')
	Z = ord('Z')
	def __init__(self, start, end, worksheet, validate=True):
		self._start = (Range.string_to_coordinate(start) if validate and isinstance(start, six.string_types) else start)
		self._end = (Range.string_to_coordinate(end) if validate and isinstance(end, six.string_types) else end)
		if (not (1 <= self._start[0] <= 65536) and self._start[0] != float('inf')) \
			or (not (1 <= self._end[0] <= 65536) and self._end[0] != float('inf')):
			raise Exception("Row index out of bounds")
		if (not (1 <= self._start[1] <= 256) and self._start[1] != float('inf')) \
			or (not (1 <= self._end[1] <= 256) and self._end[1] != float('inf')):
			raise Exception("Column index out of bounds")
		self.worksheet = worksheet
		self.is_cell = (self._start == self._end)
		
		self.is_row = (self._end[1] == float('inf') and self._start[0] == self._end[0] and self._start[1] == 1)
		self.is_column = (self._end[0] == float('inf') and self._start[1] == self._end[1] and self._start[0] == 1)
 		
		self.x = (self._start[0] if self.is_row or self.is_cell else None)
		self.y = (self._start[1] if self.is_column or self.is_cell else None)
		self.height = (self._end[0] - self._start[0] + 1)
		self.width = (self._end[1] - self._start[1] + 1)
		if self.is_cell:
			worksheet._columns = max(worksheet._columns, self.y)
	
	@property
	def coordinate(self):
		if self.is_cell:
			return self._start
		else:
			raise Exception("Non-singleton range selected")
	
	@property
	def style(self):
		if self.is_row:
			return self.__get_attr(self.worksheet.get_cell_style, Range.AttributeInterceptor(self.worksheet.get_row_style(self.x), ''))
		return self.__get_attr(self.worksheet.get_cell_style, Range.AttributeInterceptor(self, 'style'))
		
	@style.setter
	def style(self, data):
		self.__set_attr(self.worksheet.set_cell_style, data)

	@property
	def value(self):
		return self.__get_attr(self.worksheet.get_cell_value)
		
	@value.setter
	def value(self, data):
		self.__set_attr(self.worksheet.set_cell_value, data)
	
	# this class permits doing things like range().style.font.bold = True
	class AttributeInterceptor(object):
		def __init__(self, parent, attribute = ''):
			self.__dict__['_parent'] = parent
			self.__dict__['_attribute'] = attribute
		def __getattr__(self, name):
			if self._attribute == '':
				return Range.AttributeInterceptor(self._parent, name)
			return Range.AttributeInterceptor(self._parent, "%s.%s" % (self._attribute, name))
		def __setattr__(self, name, value):
			if isinstance(self._parent, Style.Style):
				setattr(reduce(getattr, self._attribute.split('.'), self._parent), name, value)
			else:
				for cell in self._parent:
					setattr(reduce(getattr, self._attribute.split('.'), cell), name, value)

	# note that these are not the python __getattr__/__setattr__
	def __get_attr(self, method, default=None):
		if self.is_cell:
			for merge in self.worksheet.merges:
				if self in merge:
					return method(merge._start[0], merge._start[1])
			return method(self.x, self.y)
		elif default:
			return default
		else:
			raise Exception('Non-singleton range selected')
	
	def __set_attr(self, method, data):
		if self.is_cell:
			for merge in self.worksheet.merges:
				if self in merge:
					method(merge._start[0], merge._start[1], data)
					return
			method(self.x, self.y, data)
		elif self.is_row and isinstance(data, Style.Style):
			# Applying a row style
			self.worksheet.set_row_style(self.x, data)
		elif DataTypes.DataTypes.get_type(data) != DataTypes.DataTypes.ERROR:
			# Attempt to apply in batch
			for cell in self:
				cell.__set_attr(method, data)
		else:
			if len(data) <= self.height:
				for row in data:
					if len(row) > self.width:
						raise Exception("Row too large for range, row has %s columns, but range only has %s" % (len(row), self.width))
				for x, row in enumerate(data):
					for y, value in enumerate(row):
						method(x + self._start[0], y + self._start[1], value)
			else:
				raise Exception("Too many rows for range, data has %s rows, but range only has %s" % (len(data), self.height))

	def intersection(self, range):
		"""
		Calculates the intersection with another range object
		"""
		if self.worksheet != range.worksheet:
			# Different worksheet
			return None
		start = (max(self._start[0], range._start[0]), max(self._start[1], range._start[1]))
		end = (min(self._end[0], range._end[0]), min(self._end[1], range._end[1]))
		if end[0] < start[0] or end[1] < start[1]:
			return None
		return Range(start, end, self.worksheet, validate=False)
	
	__and__ = intersection
	
	def intersects(self, range):
		return self.intersection(range) is not None
	
	def merge(self):
		self.worksheet.add_merge(self)

	def __iter__(self):
		if self.is_row or self.is_column:
			raise Exception('Can\'t iterate over an infinite row/column')
		for x in range(self._start[0], self._end[0] + 1):
			for y in range(self._start[1], self._end[1] + 1):
				yield Range((x, y), (x, y), self.worksheet, validate=False)

	def __contains__(self, item):
		return self.intersection(item) == item

	def __hash__(self):
		def hash(val):
			return val[0] << 8 + val[1]
		return hash(self._start) << 24 + hash(self._end)

	def __str__(self):
		return Range.coordinate_to_string(self._start) + ":" + Range.coordinate_to_string(self._end)

	def __len__(self):
		if self._start[0] == self._end[0]:
			return self.width
		else:
			return self.height

	def __eq__(self, other):
		if other is None:
			return False
		return self._start == other._start and self._end == other._end
	
	def __ne__(self, other):
		return not (self == other)
	
	def __getitem__(self, key):
		if self.is_row:
			# return the key'th column
			if isinstance(key, six.string_types):
				key = Range.string_to_coordinate(key)
			return Range((self.x, key), (self.x, key), self.worksheet, validate=False)
		elif self.is_column:
			#return the key'th row
			return Range((key, self.y), (key, self.y), self.worksheet, validate=False)			
		else:
			raise Exception("Selection not valid")
	
	def __setitem__(self, key, value):
		if self.is_row:
			self.worksheet.set_cell_value(self.x, key, value)
		else:
			raise Exception("Couldn't set that")
	
	@staticmethod
	def string_to_coordinate(s):
		# Convert a base-26 name to integer
		y = 0
		l = len(s)
		for index, c in enumerate(s):
			if ord(c) < Range.A or ord(c) > Range.Z:
				s = s[index:]
				break
			y *= 26
			y += ord(c) - Range.A + 1
		if len(s) == l:
			return y
		else:
			return (int(s), y)

	_cts_cache = {}
	@staticmethod
	def coordinate_to_string(coord):
		if coord[1] == float('inf'):
			return 'IV%s' % str(coord[0])
		
		# convert an integer to base-26 name
		y = coord[1] - 1
		if y not in Range._cts_cache:
			s = ""	
			while y >= 0:
				s = chr((y % 26) + Range.A) + s
				y = int(y / 26) - 1
			Range._cts_cache[y] = s
		return Range._cts_cache[y] + str(coord[0])

########NEW FILE########
__FILENAME__ = SharedStrings
# NB: Not actually used atm

class SharedStrings(object):
	def __init__(self, workbook):
		self._parent = workbook
		self._map = {}
		self._index = 1
		
	@property
	def workbook(self):
		return self._parent
		
	def get_key(self, s):
		# get the key for s
		if s not in self._map:
			self._map[s] = self._index
			self._index += 1
		return self._map[s]
		
########NEW FILE########
__FILENAME__ = Style
from . import Font, Fill, Format, Alignment, Borders
from . import Utility
import six

class Style(object):
	def __init__(self, font=None, fill=None, format=None, alignment=None, borders=None, size=None):
		self._font = font
		self._fill = fill
		self._format = format
		self._alignment = alignment
		self._borders = borders
		self._size = size

	@property
	def size(self):
		return self._size		
	
	@property
	def is_default(self):
		return not (self._font or self._fill or self._format or self._alignment or self._borders or self._size is not None)

	@property
	def borders(self):
		return Utility.lazy_get(self, '_borders', Borders.Borders())
		
	@borders.setter
	def borders(self, value):
		Utility.lazy_set(self, '_borders', None, value)
		
	@property
	def alignment(self):
		return Utility.lazy_get(self, '_alignment', Alignment.Alignment())
		
	@alignment.setter
	def alignment(self, value):
		Utility.lazy_set(self, '_alignment', None, value)
		
	@property
	def format(self):
		# don't use default because default should be const
		return Utility.lazy_get(self, '_format', Format.Format())
	
	@format.setter
	def format(self, value):
		Utility.lazy_set(self, '_format', None, value)
	
	@property
	def font(self):
		return Utility.lazy_get(self, '_font', Font.Font())
	
	@font.setter
	def font(self, value):
		Utility.lazy_set(self, '_font', None, value)
	
	@property
	def fill(self):
		return Utility.lazy_get(self, '_fill', Fill.Fill())
	
	@fill.setter
	def fill(self, value):
		Utility.lazy_set(self, '_fill', None, value)
	
	def get_xml_string(self):
		# Precondition: Workbook._align_styles has been run.
		# Be careful when using this function as id's may be inaccurate if precondition not met.
		tag = []
		if not self._format is None:
			tag.append("numFmtId=\"%d\"" % self._format.id)
		if not self._font is None:
			tag.append("applyFont=\"1\" fontId=\"%d\"" % (self._font.id))
		if not self._fill is None:
			tag.append("applyFill=\"1\" fillId=\"%d\"" % (self._fill.id + 1))
		if not self._borders is None:
			tag.append("applyBorder=\"1\" borderId=\"%d\"" % (self._borders.id))
		if self._alignment is None:
			return "<xf xfId=\"0\" %s/>" % (" ".join(tag))
		else:
			return "<xf xfId=\"0\"  %s applyAlignment=\"1\">%s</xf>" % (" ".join(tag), self._alignment.get_xml_string())
		
	def __hash__(self):
		return hash((self._font, self._fill))
	
	def __eq__(self, other):
		if other is None:
			return self.is_default
		elif Utility.YOLO:
			return self._format == other._format and self._alignment == other._alignment and self._borders == other._borders
		else:
			return self._to_tuple() == other._to_tuple()
	
	def __or__(self, other):
		return self._binary_operation(other, Utility.nonboolean_or)
			
	def __and__(self, other):
		return self._binary_operation(other, Utility.nonboolean_and)

	def __xor__(self, other):
		return self._binary_operation(other, Utility.nonboolean_xor)

	def _binary_operation(self, other, operation):
		return Style( \
			font=operation(self.font, other.font), \
			fill=operation(self.fill, other.fill), \
			format=operation(self.format, other.format), \
			alignment=operation(self.alignment, other.alignment), \
			borders=operation(self.borders, other.borders) \
		)
	
	def _to_tuple(self):
		return (self._font, self._fill, self._format, self._alignment, self._borders, self._size)
	
			
	def __str__(self):
		return "%s %s %s %s" % (self.font, self.fill, self.format, self.alignment)
		
	def __repr__(self):
		return "<%s>" % self.__str__()

########NEW FILE########
__FILENAME__ = benchmark
from ..Workbook import Workbook
from ..Color import Color
from ..Style import Style
from ..Font import Font
from ..Fill import Fill
import time
from .utils import get_output_path
from random import randint

ROWS = 1000
COLUMNS = 100
BOLD = 1
ITALIC = 2
UNDERLINE = 4
RED_BG = 8
testData = [[1] * COLUMNS] * ROWS
formatData = [[1] * COLUMNS] * ROWS
def run_pyexcelerate_value_fastest():
	wb = Workbook()
	stime = time.clock()
	ws = wb.new_sheet('Test 1', data=testData)
	wb.save(get_output_path('test_pyexcelerate_value_fastest.xlsx'))
	elapsed = time.clock() - stime
	print("pyexcelerate value fastest, %s, %s, %s" % (ROWS, COLUMNS, elapsed))
	return elapsed

def run_pyexcelerate_value_faster():
	wb = Workbook()
	stime = time.clock()
	ws = wb.new_sheet('Test 1')
	for row in range(ROWS):
		for col in range(COLUMNS):
			ws.set_cell_value(row + 1, col + 1, 1)
	wb.save(get_output_path('test_pyexcelerate_value_faster.xlsx'))
	elapsed = time.clock() - stime
	print("pyexcelerate value faster, %s, %s, %s" % (ROWS, COLUMNS, elapsed))
	return elapsed
	

def run_pyexcelerate_value_fast():
	wb = Workbook()
	stime = time.clock()
	ws = wb.new_sheet('Test 1')
	for row in range(ROWS):
		for col in range(COLUMNS):
			ws[row + 1][col + 1].value = 1
	wb.save(get_output_path('test_pyexcelerate_value_fast.xlsx'))
	elapsed = time.clock() - stime
	print("pyexcelerate value fast, %s, %s, %s" % (ROWS, COLUMNS, elapsed))
	return elapsed
	
def run_openpyxl():
	try:
		import openpyxl
	except ImportError:
		raise Exception('openpyxl not installled')
	stime = time.clock()
	wb = openpyxl.workbook.Workbook(optimized_write=True) 
	ws = wb.create_sheet()
	ws.title = 'Test 1'
	for row in testData:
		ws.append(row)
	wb.save(get_output_path('test_openpyxl.xlsx'))
	elapsed = time.clock() - stime
	print("openpyxl, %s, %s, %s" % (ROWS, COLUMNS, elapsed))
	return elapsed

def run_xlsxwriter_value():
	try:
		import xlsxwriter.workbook
	except ImportError:
		raise Exception('XlsxWriter not installled')
	stime = time.clock()
	wb = xlsxwriter.workbook.Workbook(get_output_path('test_xlsxwriter.xlsx'), {'constant_memory': True})
	ws = wb.add_worksheet()
	for row in range(ROWS):
		for col in range(COLUMNS):
			ws.write_number(row, col, 1)
	wb.close()
	elapsed = time.clock() - stime
	print("xlsxwriter value, %s, %s, %s" % (ROWS, COLUMNS, elapsed))
	return elapsed
	
def generate_format_data():
	for row in range(ROWS):
		for col in range(COLUMNS):
			formatData[row][col] = randint(1, (1 << 4) - 1)
	
def run_pyexcelerate_style_fastest():
	wb = Workbook()
	stime = time.clock()
	ws = wb.new_sheet('Test 1')
	bold = Style(font=Font(bold=True))
	italic = Style(font=Font(italic=True))
	underline = Style(font=Font(underline=True))
	red = Style(fill=Fill(background=Color(255,0,0,0)))
	for row in range(ROWS):
		for col in range(COLUMNS):
			ws.set_cell_value(row + 1, col + 1, 1)
			style = Style()
			if formatData[row][col] & BOLD:
				style.font.bold = True
			if formatData[row][col] & ITALIC:
				style.font.italic = True
			if formatData[row][col] & UNDERLINE:
				style.font.underline = True
			if formatData[row][col] & RED_BG:
				style.fill.background = Color(255, 0, 0)
			ws.set_cell_style(row + 1, col + 1, style)
	wb.save(get_output_path('test_pyexcelerate_style_fastest.xlsx'))
	elapsed = time.clock() - stime
	print("pyexcelerate style fastest, %s, %s, %s" % (ROWS, COLUMNS, elapsed))
	return elapsed
	
def run_pyexcelerate_style_faster():
	wb = Workbook()
	stime = time.clock()
	ws = wb.new_sheet('Test 1')
	for row in range(ROWS):
		for col in range(COLUMNS):
			ws.set_cell_value(row + 1, col + 1, 1)
			if formatData[row][col] & BOLD:
				ws.get_cell_style(row + 1, col + 1).font.bold = True
			if formatData[row][col] & ITALIC:
				ws.get_cell_style(row + 1, col + 1).font.italic = True
			if formatData[row][col] & UNDERLINE:
				ws.get_cell_style(row + 1, col + 1).font.underline = True
			if formatData[row][col] & RED_BG:
				ws.get_cell_style(row + 1, col + 1).fill.background = Color(255, 0, 0)
	wb.save(get_output_path('test_pyexcelerate_style_faster.xlsx'))
	elapsed = time.clock() - stime
	print("pyexcelerate style faster, %s, %s, %s" % (ROWS, COLUMNS, elapsed))
	return elapsed
	
def run_pyexcelerate_style_fast():
	wb = Workbook()
	stime = time.clock()
	ws = wb.new_sheet('Test 1')
	for row in range(ROWS):
		for col in range(COLUMNS):
			ws[row + 1][col + 1].value = 1
			if formatData[row][col] & BOLD:
				ws[row + 1][col + 1].style.font.bold = True
			if formatData[row][col] & ITALIC:
				ws[row + 1][col + 1].style.font.italic = True
			if formatData[row][col] & UNDERLINE:
				ws[row + 1][col + 1].style.font.underline = True
			if formatData[row][col] & RED_BG:
				ws[row + 1][col + 1].style.fill.background = Color(255, 0, 0, 0)
	wb.save(get_output_path('test_pyexcelerate_style_fast.xlsx'))
	elapsed = time.clock() - stime
	print("pyexcelerate style fast, %s, %s, %s" % (ROWS, COLUMNS, elapsed))
	return elapsed
	
def run_pyexcelerate_style_cheating():
	wb = Workbook()
	stime = time.clock()
	ws = wb.new_sheet('Test 1')

	cell_formats = []

	for i in range(16):
		cell_format = Style()
		if i & BOLD:
			cell_format.font.bold = True
		if i & ITALIC:
			cell_format.font.italic = True
		if i & UNDERLINE:
			cell_format.font.underline = True
		if i & RED_BG:
			cell_format.fill.background = Color(255, 0, 0)
		cell_formats.append(cell_format)

	for row in range(ROWS):
		for col in range(COLUMNS):
			ws.set_cell_value(row + 1, col + 1, 1)
			ws.set_cell_style(row + 1, col + 1, cell_formats[formatData[row][col]])
	wb.save(get_output_path('test_pyexcelerate_style_fastest.xlsx'))
	elapsed = time.clock() - stime
	print("pyexcelerate style cheating, %s, %s, %s" % (ROWS, COLUMNS, elapsed))
	return elapsed

def run_xlsxwriter_style_cheating():
	try:
		import xlsxwriter.workbook
	except ImportError:
		raise Exception('XlsxWriter not installled')
	stime = time.clock()
	wb = xlsxwriter.workbook.Workbook(get_output_path('test_xlsxwriter_style.xlsx'), {'constant_memory': True})
	ws = wb.add_worksheet()
	
	cell_formats = []

	for i in range(16):
		cell_format = wb.add_format()
		if i & BOLD:
			cell_format.set_bold()
		if i & ITALIC:
			cell_format.set_italic()
		if i & UNDERLINE:
			cell_format.set_underline()
		if i & RED_BG:
			cell_format.set_bg_color('red')
		cell_formats.append(cell_format)

	for row in range(ROWS):
		for col in range(COLUMNS):
			ws.write_number(row, col, 1, cell_formats[formatData[row][col]]) 
	wb.close()
	elapsed = time.clock() - stime
	print("xlsxwriter style cheating, %s, %s, %s" % (ROWS, COLUMNS, elapsed))
	return elapsed
	
def run_xlsxwriter_style():
	try:
		import xlsxwriter.workbook
	except ImportError:
		raise Exception('XlsxWriter not installled')
	stime = time.clock()
	wb = xlsxwriter.workbook.Workbook(get_output_path('test_xlsxwriter_style.xlsx'), {'constant_memory': True})
	ws = wb.add_worksheet()
	
	for row in range(ROWS):
		for col in range(COLUMNS):
			format = wb.add_format()
			if formatData[row][col] & BOLD:
				format.set_bold()
			if formatData[row][col] & ITALIC:
				format.set_italic()
			if formatData[row][col] & UNDERLINE:
				format.set_underline()
			if formatData[row][col] & RED_BG:
				format.set_bg_color('red')
			ws.write_number(row, col, 1, format) 
	wb.close()
	elapsed = time.clock() - stime
	print("xlsxwriter style, %s, %s, %s" % (ROWS, COLUMNS, elapsed))
	return elapsed

def run_openpyxl_optimization():
	try:
		import openpyxl
	except ImportError:
		raise Exception('openpyxl not installled')
	stime = time.clock()
	wb = openpyxl.workbook.Workbook(optimized_write=True) 
	ws = wb.create_sheet()
	ws.title = 'Test 1'
	for col_idx in range(COLUMNS):
		col = openpyxl.cell.get_column_letter(col_idx + 1)
		for row in range(ROWS):
			ws.cell('%s%s'%(col, row + 1)).value = 1
			if formatData[row][col_idx] & BOLD:
				ws.cell('%s%s'%(col, row + 1)).style.font.bold = True
			if formatData[row][col_idx] & ITALIC:
				ws.cell('%s%s'%(col, row + 1)).style.font.italic = True
			if formatData[row][col_idx] & UNDERLINE:
				ws.cell('%s%s'%(col, row + 1)).style.font.underline = True
			if formatData[row][col_idx] & RED_BG:
				ws.cell('%s%s'%(col, row + 1)).style.fill.fill_type = openpyxl.style.Fill.FILL_SOLID
				ws.cell('%s%s'%(col, row + 1)).style.fill.start_color = openpyxl.style.Color(openpyxl.style.Color.RED)
				ws.cell('%s%s'%(col, row + 1)).style.fill.end_color = openpyxl.style.Color(openpyxl.style.Color.RED)
			ws.cell('%s%s'%(col, row + 1)).value = 1
	wb.save(get_output_path('test_openpyxl_opt.xlsx'))
	elapsed = time.clock() - stime
	print("openpyxl, %s, %s, %s" % (ROWS, COLUMNS, elapsed))
	return elapsed


	

def test_all():
	run_pyexcelerate_value_fastest()
	run_pyexcelerate_value_faster()
	run_pyexcelerate_value_fast()
	run_xlsxwriter_value()
	run_openpyxl()
	generate_format_data()
	run_pyexcelerate_style_cheating()
	run_pyexcelerate_style_fastest()
	run_pyexcelerate_style_faster()
	run_pyexcelerate_style_fast()
	run_xlsxwriter_style_cheating()
	run_xlsxwriter_style()
	run_openpyxl_optimization()

########NEW FILE########
__FILENAME__ = python_benchmarks
import time
from nose.tools import *

def test_benchmark():
	TRIALS = range(1000000)
	
	integer = 1
	float = 3.0
	long = 293203948032948023984023948023957245
	
	# attempt isinstance
	
	stime = time.clock()
	for i in TRIALS:
		answer = isinstance(integer, (int, float, long, complex))
		ok_(answer)
	print("isinstance, %s" % (time.clock() - stime))

	# attempt __class__
	
	stime = time.clock()
	for i in TRIALS:
		answer = (integer.__class__ in set((int, float, long, complex)))
		ok_(answer)
	print("__class__, set, %s" % (time.clock() - stime))
	
	stime = time.clock()
	for i in TRIALS:
		answer = (integer.__class__ == int or integer.__class__ == float or integer.__class__ == long or integer.__class__ == complex)
		ok_(answer)
	print("__class__, or, %s" % (time.clock() - stime))
	

########NEW FILE########
__FILENAME__ = test_DataTypes
from ..DataTypes import DataTypes
from nose.tools import eq_
from datetime import datetime, date, time
from ..Workbook import Workbook
from .utils import get_output_path
from decimal import Decimal
import numpy

def test__get_type():
	eq_(DataTypes.get_type(15), DataTypes.NUMBER)
	eq_(DataTypes.get_type(15.0), DataTypes.NUMBER)
	eq_(DataTypes.get_type(Decimal('15.0')), DataTypes.NUMBER)
	eq_(DataTypes.get_type("test"), DataTypes.INLINE_STRING)
	eq_(DataTypes.get_type(datetime.now()), DataTypes.DATE)
	eq_(DataTypes.get_type(True), DataTypes.BOOLEAN)
	
def test_numpy():
	testData = numpy.ones((5, 5), dtype = int)
	wb = Workbook()
	ws = wb.new_sheet("Test 1", data=testData)
	eq_(ws[1][1].value, 1)
	eq_(DataTypes.get_type(ws[1][1].value), DataTypes.NUMBER)
	wb.save(get_output_path("numpy-test.xlsx"))

def test_ampersand_escaping():
	testData = [["http://example.com/?one=1&two=2"]]
	wb = Workbook()
	ws = wb.new_sheet("Test 1", data=testData)
	data = list(ws.get_xml_data())
	assert "http://example.com/?one=1&amp;two=2" in data[0][1][0]

def test_to_excel_date():
	eq_(DataTypes.to_excel_date(datetime(1900, 1, 1, 0, 0, 0)), 1.0)
	eq_(DataTypes.to_excel_date(datetime(1900, 1, 1, 12, 0, 0)), 1.5)
	eq_(DataTypes.to_excel_date(datetime(1900, 1, 1, 12, 0, 0)), 1.5)
	eq_(DataTypes.to_excel_date(datetime(2013, 5, 10, 6, 0, 0)), 41404.25)
	eq_(DataTypes.to_excel_date(date(1900, 1, 1)), 1.0)
	eq_(DataTypes.to_excel_date(date(2013, 5, 10)), 41404.0)
	eq_(DataTypes.to_excel_date(time(6, 0, 0)), 0.25)
	# check excel's improper handling of leap year
	eq_(DataTypes.to_excel_date(datetime(1900, 2, 28, 0, 0, 0)), 59.0)
	eq_(DataTypes.to_excel_date(datetime(1900, 3, 1, 0, 0, 0)), 61.0)

########NEW FILE########
__FILENAME__ = test_Range
from ..Workbook import Workbook
from ..Range import Range
from nose.tools import eq_

def test__string_to_coordinate():
    stc = Range.string_to_coordinate
    eq_(stc("A1"), (1, 1))
    eq_(stc("A2"), (2, 1))
    eq_(stc("A3"), (3, 1))
    eq_(stc("Z1"), (1, 26))
    eq_(stc("B10"), (10, 2))
    eq_(stc("B24"), (24, 2))
    eq_(stc("B39"), (39, 2))
    eq_(stc("AA1"), (1, 27))
    eq_(stc("AB1"), (1, 28))

def test__coordinate_to_string():
    cts = Range.coordinate_to_string
    eq_(cts((1, 1)), "A1")
    eq_(cts((2, 1)), "A2")
    eq_(cts((3, 1)), "A3")
    eq_(cts((1, 26)), "Z1")
    eq_(cts((1, 2)), "B1")
    eq_(cts((10, 2)), "B10")
    eq_(cts((24, 2)), "B24")
    eq_(cts((39, 2)), "B39")
    eq_(cts((1, 27)), "AA1")
    eq_(cts((1, 28)), "AB1")

def test_merge():
     wb = Workbook()
     ws = wb.new_sheet("Test")
     r1 =  Range("A1", "A5", ws)
     r1.merge()
     r2 =  Range("B1", "B5", ws)
     r2.merge()
     eq_(len(ws.merges), 2)

def test_horizontal_intersection():
    wb = Workbook()
    ws = wb.new_sheet("Test")
    r1 =  Range("A1", "A3", ws)
    r2 =  Range("A2", "A4", ws)
    eq_(r1.intersects(r2), True)
    eq_(r1.intersection(r2), Range("A2", "A3", ws))

def test_vertical_intersection():
    wb = Workbook()
    ws = wb.new_sheet("Test")
    r1 =  Range("A1", "C1", ws)
    r2 =  Range("B1", "D1", ws)
    eq_(r1.intersects(r2), True)
    eq_(r1.intersection(r2), Range("B1", "C1", ws))

def test_rectangular_intersection():
    wb = Workbook()
    ws = wb.new_sheet("Test")
    r1 =  Range("A1", "C3", ws)
    r2 =  Range("B2", "D4", ws)
    eq_(r1.intersects(r2), True)
    eq_(r1.intersection(r2), Range("B2", "C3", ws))

def test_no_intersection():
    wb = Workbook()
    ws = wb.new_sheet("Test")
    r1 =  Range("A1", "B2", ws)
    r2 =  Range("C3", "D4", ws)
    eq_(r1.intersects(r2), False)
    eq_(r1.intersection(r2), None)

def test_range_equal_to_none():
    wb = Workbook()
    ws = wb.new_sheet("Test")
    r1 =  Range("A1", "C3", ws)
    r2 =  Range("B2", "D4", ws)
    eq_(r1.intersection(r2) == None, False)

def test_range_equal_to_itself():
    wb = Workbook()
    ws = wb.new_sheet("Test")
    r1 =  Range("A1", "C3", ws)
    eq_(r1 == r1, True)

"""
def test_get_xml_data():
    wb = Workbook()
    ws = wb.new_sheet("Test")
    ws[1][1].value = 1
    eq_(ws[1][1].value, 1) 
    ws[1][3].value = 3
    eq_(ws[1][3].value, 3)
    gxd = ws[1].get_xml_data()
    eq_(gxd.next(), ('A1', 1, 4))
    eq_(gxd.next(), ('C1', 3, 4))
"""

########NEW FILE########
__FILENAME__ = test_speed
from .benchmark import run_pyexcelerate_value_fastest, run_xlsxwriter_value, run_openpyxl
from nose.tools import ok_
import nose

def test_vs_xlsxwriter():
	raise nose.SkipTest('Skipping speed test')
	ours = run_pyexcelerate_value_fastest()
	theirs = run_xlsxwriter_value()
	ok_(ours / theirs <= 0.67, msg='PyExcelerate is too slow! Better Excelerate it some more!')

########NEW FILE########
__FILENAME__ = test_Style
from ..Workbook import Workbook
from ..Color import Color
from ..Font import Font
from ..Fill import Fill
from ..Style import Style
from ..Alignment import Alignment
import time
import numpy
from datetime import datetime
from nose.tools import eq_
from .utils import get_output_path

def test_style():
	wb = Workbook()
	ws = wb.new_sheet("test")
	ws[1][1].value = 1
	ws[1][2].value = 1
	ws[1][3].value = 1
	ws[1][1].style.font.bold = True
	ws[1][2].style.font.italic = True
	ws[1][3].style.font.underline = True
	ws[1][1].style.font.strikethrough = True
	ws[1][1].style.font.color = Color(255, 0, 255)
	ws[1][1].style.fill.background = Color(0, 255, 0)
	ws[1][2].style.fill.background = Color(255, 255, 0)
	ws[2][1].value = "asdf"
	ws.range("A2", "B2").merge()
	eq_(ws[1][2].value, ws[1][1].value)
	ws[2][2].value = "qwer"
	eq_(ws[1][2].value, ws[1][1].value)
	ws[2][1].style.fill.background = Color(0, 255, 0)
	ws[1][1].style.alignment.vertical = 'top'
	ws[1][1].style.alignment.horizontal = 'right'
	ws[1][1].style.alignment.rotation = 90
	ws[3][3].style.borders.top.color = Color(255, 0, 0)
	ws[3][3].style.borders.left.color = Color(0, 255, 0)
	ws[3][4].style.borders.right.style = '-.'
	wb.save(get_output_path("style-test.xlsx"))

def test_style_compression():
	wb = Workbook()
	ws = wb.new_sheet("test")
	ws.range("A1","C3").value = 1
	ws.range("A1","C1").style.font.bold = True
	ws.range("A2","C3").style.font.italic = True
	ws.range("A3","C3").style.fill.background = Color(255, 0, 0)
	ws.range("C1","C3").style.font.strikethrough = True
	wb.save(get_output_path("style-compression-test.xlsx"))
	
def test_style_reference():
	wb = Workbook()
	ws = wb.new_sheet("test")
	ws[1][1].value = 1
	font = Font(bold=True, italic=True, underline=True, strikethrough=True)
	ws[1][1].style.font = font
	wb.save(get_output_path("style-reference-test.xlsx"))

def test_style_row():
	wb = Workbook()
	ws = wb.new_sheet("test")
	ws[1].style.fill.background = Color(255, 0, 0)
	ws[1][3].style.fill.background = Color(0, 255, 0)
	wb.save(get_output_path("style-row-test.xlsx"))

def test_style_row_col():
	wb = Workbook()
	ws = wb.new_sheet("test")
	ws.range("A1", "D4").value = 'sdfgs5b56seb6se56bse5jsdfljg'
	eq_(Style(), ws.get_row_style(1))
	eq_(Style(), ws.get_col_style(1))
	ws.set_row_style(1, Style(size=-1))
	ws.set_row_style(2, Style(size=0))
	ws.set_row_style(3, Style(size=100, fill=Fill(background=Color(0, 255, 0, 0))))
	ws.set_col_style(1, Style(size=-1))
	ws.set_col_style(2, Style(size=0))
	ws.set_col_style(3, Style(size=100, fill=Fill(background=Color(255, 0, 0, 0))))
	wb.save(get_output_path("style-auto-row-col-test.xlsx"))

def test_and_or_xor():
	bolditalic = Font(bold=True, italic=True)
	italicunderline = Font(italic=True, underline=True)
	eq_(Font(italic=True), bolditalic & italicunderline)
	eq_(Font(bold=True, italic=True, underline=True), bolditalic | italicunderline)
	eq_(Font(bold=True, underline=True), bolditalic ^ italicunderline)
	
	fontstyle = Style(font=Font(bold=True))
	fillstyle = Style(fill=Fill(background=Color(255, 0, 0, 0)))
	eq_(Style(), fontstyle & fillstyle)
	eq_(Style(font=Font(bold=True), fill=Fill(background=Color(255, 0, 0, 0))), fontstyle | fillstyle)
	eq_(Style(font=Font(bold=True), fill=Fill(background=Color(255, 0, 0, 0))), fontstyle ^ fillstyle)
	
	leftstyle = Style(alignment=Alignment('right', 'top'))
	bottomstyle = Style(alignment=Alignment(vertical='top', rotation=15))
	eq_(Style(alignment=Alignment('right', 'top', 15)), leftstyle | bottomstyle)
	eq_(Style(alignment=Alignment(vertical='top')), leftstyle & bottomstyle)
	eq_(Style(alignment=Alignment('right', rotation=15)), leftstyle ^ bottomstyle)
	
def test_str_():
	font = Font(bold=True, italic=True, underline=True, strikethrough=True)
	eq_(font.__repr__(), "<Font: Calibri, 11pt b i u s>")
########NEW FILE########
__FILENAME__ = test_Workbook
from ..Workbook import Workbook
from ..Color import Color
import time
import numpy
import nose
import os
from datetime import datetime
from nose.tools import eq_
from .utils import get_output_path

def test_get_xml_data():
    wb = Workbook()
    ws = wb.new_sheet("Test")
    ws[1][1].value = 1
    eq_(ws[1][1].value, 1)
    ws[1][3].value = 3
    eq_(ws[1][3].value, 3)
    
def test_save():
    ROWS = 65
    COLUMNS = 100
    wb = Workbook()
    testData = [[1] * COLUMNS] * ROWS
    stime = time.clock()
    ws = wb.new_sheet("Test 1", data=testData)
    wb.save(get_output_path("test.xlsx"))
    #print("%s, %s, %s" % (ROWS, COLUMNS, time.clock() - stime))

def test_formulas():
	wb = Workbook()
	ws = wb.new_sheet("test")
	ws[1]['A'].value = 1
	ws[1][2].value = 2
	ws[1][3].value = '=SUM(A1,B1)'
	ws[1][4].value = datetime.now()
	ws[1][5].value = datetime(1900,1,1,1,0,0)
	ws[1][6].value = True
	wb.save(get_output_path("formula-test.xlsx"))
	
def test_merge():
	wb = Workbook()
	ws = wb.new_sheet("test")
	ws[1][1].value = "asdf"
	ws.range("A1", "B1").merge()
	eq_(ws[1][2].value, ws[1][1].value)
	ws[1][2].value = "qwer"
	eq_(ws[1][2].value, ws[1][1].value)
	wb.save(get_output_path("merge-test.xlsx"))
	
def test_cell():
	wb = Workbook()
	ws = wb.new_sheet("test")
	ws.cell("C3").value = "test"
	eq_(ws[3][3].value, "test")

def test_range():
	wb = Workbook()
	ws = wb.new_sheet("test")
	ws.range("B2", "D3").value = [[1, 2, 3], [4, 5, 6]]
	eq_(ws[2][2].value, 1)
	eq_(ws[2][3].value, 2)
	eq_(ws[2][4].value, 3)
	eq_(ws[3][2].value, 4)
	eq_(ws[3][3].value, 5)
	eq_(ws[3][4].value, 6)
	
def test_numpy_range():
	wb = Workbook()
	ws = wb.new_sheet("test")
	ws.range("A1", "GN13").value = numpy.zeros((13,196))
	wb.save(get_output_path("numpy-range-test.xlsx"))

def test_none():
     testData = [[1,2,None]]
     wb = Workbook()
     ws = wb.new_sheet("Test 1", data=testData)
     ws[1][1].style.font.bold = True
     wb.save(get_output_path("none-test.xlsx"))
     
def test_number_precision():
	try:
		import xlrd
	except ImportError:
		raise nose.SkipTest('xlrd not installed')

	filename = get_output_path('precision.xlsx')
	sheetname = 'Sheet1'

	nums = [
		1,
		1.2,
		1.23,
		1.234,
		1.2345,
		1.23456,
		1.234567,
		1.2345678,
		1.23456789,
		1.234567890,
		1.2345678901,
		1.23456789012,
		1.234567890123,
		1.2345678901234,
		1.23456789012345,
	]

	write_workbook = Workbook()
	write_worksheet = write_workbook.new_sheet(sheetname)

	for index, value in enumerate(nums):
		write_worksheet[index + 1][1].value = value

	write_workbook.save(filename)

	read_workbook = xlrd.open_workbook(filename)
	read_worksheet = read_workbook.sheet_by_name(sheetname)

	for row_num in range(len(nums)):
		expected = nums[row_num]
		got = read_worksheet.cell(row_num, 0).value

	if os.path.exists(filename):
		os.remove(filename)

#def test_column_select():
#	wb = Workbook()
#	ws = wb.new_sheet("Test")
#	print(ws[1:3])
#	ws[1:3][1].style.fill.background = Color(255, 0, 0)

########NEW FILE########
__FILENAME__ = utils
import os
import shutil

def get_output_path(fn):
	out_dir = os.path.join(os.path.dirname(__file__), 'output')
	if not os.path.exists(out_dir):
		os.mkdir(out_dir)
	return os.path.join(out_dir, fn)

########NEW FILE########
__FILENAME__ = Utility
def nonboolean_or(left, right, default=False):
	if default == False:
		return left | right
	if left == default:
		return right
	if right == default or left == right:
		return left
	return left | right # this scenario doesn't actually make sense, but it might be implemented

def nonboolean_and(left, right, default=False):
	if default == False:
		return left & right
	if left == right:
		return left
	return default
	
def nonboolean_xor(left, right, default=False):
	if default == False:
		return left ^ right
	if left == default:
		return right
	if right == default:
		return left
	return default
	
def lazy_get(self, attribute, default):
	value = getattr(self, attribute)
	if not value:
		setattr(self, attribute, default)
		return default
	else:
		return value
		
def lazy_set(self, attribute, default, value):
	if value == default:
		setattr(self, attribute, default)
	else:
		setattr(self, attribute, value)

YOLO = False # are we aligning?
########NEW FILE########
__FILENAME__ = vfs
from zipfile import ZipFile, ZIP_DEFLATED

class VirtualDirectory(object):
    def __init__(self, *args, **kwargs):
        self.directories = {}
        self.files = {}

    def __add_directory(self, path):
        subdirs = self.path.split("/")
        if subdirs[0] not in self.directories:
            d = Directory()
            self.directories[subdirs[0]] = d
        if len(subdirs) > 1:
            return self.directories[subdirs[0]].add_directory(subdirs[1:])
        else:
            return self.directories[subdirs[0]]

    def get_or_create_directory(self, path):
        subdirs = self.path.split("/")
        if subdirs[0] not in self.directories:
            d = self.__add_directory(subdirs[0])
        if len(subdirs) > 1:
            return d.get_or_create_directory(subdirs[1:])
        return d

    def add_file(self, path, data):
        ps = path.split("/") 
        if len(ps) == 1:
            self.files[ps[0]] = data
        else:
            d = self.get_or_create_directory(ps[:-1])
            d.files[ps[-1]] = data

    def walk(self):
        dirs = [self] + self.directories.keys()
        while dirs:
            d = dirs.pop()
            yield (d.directories, d.files)
            if d.directories:
                dirs.extend(d.directories)
            

class Filesystem(VirtualDirectory):
    def writeZip(self, path):
        zf = ZipFile(path, 'w', 'ZIP_DEFLATED') 
        for f, d in self.files.items(): 
            zf.writestr(f, d)

########NEW FILE########
__FILENAME__ = Workbook
from . import Worksheet
from .Writer import Writer
from . import Utility
import time

class Workbook(object):
	# map for attribute sets => style attribute id's
	STYLE_ATTRIBUTE_MAP = {'fonts':'_font', 'fills':'_fill', 'num_fmts':'_format', 'borders':'_borders'}
	STYLE_ID_ATTRIBUTE = 'id'
	alignment = None
	def __init__(self, encoding='utf-8'):
		self._worksheets = []
		self._styles = []
		self._items = {} #dictionary containing lists of fonts, fills, etc.
		self._encoding = encoding
		self._writer = Writer(self)

	def add_sheet(self, worksheet):
		self._worksheets.append(worksheet)
		
	def new_sheet(self, sheet_name, data=None):
		worksheet = Worksheet.Worksheet(sheet_name, self, data)
		self._worksheets.append(worksheet)
		return worksheet

	def add_style(self, style):
		# keep them all, even if they're deleted. compress later.
		self._styles.append(style)
	
	@property
	def has_styles(self):
		return len(self._styles) > 0

	@property
	def styles(self):
		self._align_styles()
		return self._styles

	def get_xml_data(self):
		if Workbook.alignment != self:
			self._align_styles() # because it will be used by the worksheets later
		for index, ws in enumerate(self._worksheets, start=1):
			yield (index, ws)

	def _align_styles(self):
		if Workbook.alignment != self:
			Utility.YOLO = True
			Workbook.alignment = self
			items = dict([(x, {}) for x in Workbook.STYLE_ATTRIBUTE_MAP.keys()])
			styles = {}
			for index, style in enumerate(self._styles):
				# compress style
				if not style.is_default:
					if style not in styles:
						styles[style] = len(styles) + 1
						setattr(style, Workbook.STYLE_ID_ATTRIBUTE, styles[style])
						# compress individual attributes
						for attr, attr_id in Workbook.STYLE_ATTRIBUTE_MAP.items():
							obj = getattr(style, attr_id)
							if obj and not obj.is_default: # we only care about it if it's not default
								if obj not in items[attr]:
									items[attr][obj] = len(items[attr]) + 1 # insert it
								obj.id = items[attr][obj] # apply
					else:
						setattr(style, Workbook.STYLE_ID_ATTRIBUTE, styles[style])
			for k, v in items.items():
				# ensure it's sorted properly
				items[k] = [tup[0] for tup in sorted(v.items(), key=lambda x: x[1])]
			self._items = items
			self._styles = [tup[0] for tup in sorted(styles.items(), key=lambda x: x[1])]
			Utility.YOLO = False
			
	def __getattr__(self, name):
		if Workbook.alignment != self:
			self._align_styles()
		return self._items[name]

	def __len__(self):
		return len(self._worksheets)

	def _save(self, file_handle):
		self._align_styles()
		self._writer.save(file_handle)

	def save(self, filename):
		self._save(open(filename, 'wb'))

########NEW FILE########
__FILENAME__ = Worksheet
from . import Range
from . import Style
from . import Format
from .DataTypes import DataTypes
import six
from datetime import datetime
from xml.sax.saxutils import escape

class Worksheet(object):
	def __init__(self, name, workbook, data=None):
		self._columns = 0 # cache this for speed
		self._name = name
		self._cells = {}
		self._cell_cache = {}
		self._styles = {}
		self._row_styles = {}
		self._col_styles = {}
		self._parent = workbook
		self._merges = [] # list of Range objects
		self._attributes = {}
		if data != None:
			for x, row in enumerate(data, 1):
				for y, cell in enumerate(row, 1):
					if x not in self._cells:
						self._cells[x] = {}
					self._cells[x][y] = cell
					self._columns = max(self._columns, y)

	def __getitem__(self, key):
		if isinstance(key, slice):
			if key.step is not None and key.step > 1:
				raise Exception("PyExcelerate doesn't support slicing with steps")
			else:
				return Range.Range((key.start or 1, 1), (key.stop or float('inf'), float('inf')), self)
		else:
			if key not in self._cells:
				self._cells[key] = {}
			return Range.Range((key, 1), (key, float('inf')), self) # return a row range

	@property
	def stylesheet(self):
		return self._stylesheet

	@property
	def col_styles(self):
		return self._col_styles.items()

	@property
	def name(self):
		return self._name
	
	@property
	def merges(self):
		return self._merges
	
	@property
	def num_rows(self):
		if len(self._cells) > 0:
			return max(self._cells.keys())
		else:
			return 1
	
	@property
	def num_columns(self):
		return max(1, self._columns)
	
	def cell(self, name):
		# convenience method
		return self.range(name, name)
	
	def range(self, start, end):
		# convenience method
		return Range.Range(start, end, self)
		
	def add_merge(self, range):
		for merge in self._merges:
			if range.intersects(merge):
				raise Exception("Invalid merge, intersects existing")
		self._merges.append(range)
	
	def get_cell_value(self, x, y):
		if x not in self._cells:
			self._cells[x] = {}
		if y not in self._cells[x]:
			return None
		type = DataTypes.get_type(self._cells[x][y])
		if type == DataTypes.FORMULA:
			# remove the equals sign
			return self._cells[x][y][:1]
		elif type == DataTypes.INLINE_STRING and self._cells[x][y][2:] == '\'=':
			return self._cells[x][y][:1]
		else:
			return self._cells[x][y]
	
	def set_cell_value(self, x, y, value):
		if x not in self._cells:
			self._cells[x] = {}
		if DataTypes.get_type(value) == DataTypes.DATE:
			self.get_cell_style(x, y).format = Format.Format('yyyy-mm-dd')
		self._cells[x][y] = value
	
	def get_cell_style(self, x, y):
		if x not in self._styles:
			self._styles[x] = {}
		if y not in self._styles[x]:
			self.set_cell_style(x, y, Style.Style())
		return self._styles[x][y]
	
	def set_cell_style(self, x, y, value):
		if x not in self._styles:
			self._styles[x] = {}
		self._styles[x][y] = value
		self._parent.add_style(value)
		if not self.get_cell_value(x, y):
			self.set_cell_value(x, y, '')
	
	def get_row_style(self, row):
		if row not in self._row_styles:
			self.set_row_style(row, Style.Style())
		return self._row_styles[row]
		
	def set_row_style(self, row, value):
		self._row_styles[row] = value
		self.workbook.add_style(value)
		
	def get_col_style(self, col):
		if col not in self._col_styles:
			self.set_col_style(col, Style.Style())
		return self._col_styles[col]
		
	def set_col_style(self, col, value):
		self._col_styles[col] = value
		self.workbook.add_style(value)
	
	@property
	def workbook(self):
			return self._parent

	def __get_cell_data(self, cell, x, y, style):
		if cell is None:
			return "" # no cell data
		if cell not in self._cell_cache or cell.__class__ == bool:
			# boolean values are treated oddly in dictionaries, manually override
			type = DataTypes.get_type(cell)
			
			if type == DataTypes.NUMBER:
				self._cell_cache[cell] = '"><v>%.15g</v></c>' % (cell)
			elif type == DataTypes.INLINE_STRING:
				self._cell_cache[cell] = '" t="inlineStr"><is><t>%s</t></is></c>' % escape(cell)
			elif type == DataTypes.DATE:
				self._cell_cache[cell] = '"><v>%s</v></c>' % (DataTypes.to_excel_date(cell))
			elif type == DataTypes.FORMULA:
				self._cell_cache[cell] = '"><f>%s</f></c>' % (cell)
			elif type == DataTypes.BOOLEAN:
				self._cell_cache[cell] = '" t="b"><v>%d</v></c>' % (cell)
		
		if style:
			return "<c r=\"%s\" s=\"%d%s" % (Range.Range.coordinate_to_string((x, y)), style.id, self._cell_cache[cell])
		else:
			return "<c r=\"%s%s" % (Range.Range.coordinate_to_string((x, y)), self._cell_cache[cell])
	
	def get_col_xml_string(self, col):
		if col in self._col_styles and not self._col_styles[col].is_default:
			style = self._col_styles[col]
			if style.size == -1:
				size = 0
				for x, row in self._cells.items():
					for y, cell in row.items():
						size = max((len(str(cell)) * 7 + 5) / 7, size)
			else:
				size = style.size if style.size else 15
				
			return "<col min=\"%d\" max=\"%d\" hidden=\"%d\" bestFit=\"%d\" customWidth=\"%d\" width=\"%f\" style=\"%d\">" % (
				col, col,
				1 if style.size == 0 else 0, # hidden
				1 if style.size == -1 else 0, # best fit
				1 if style.size is not None else 0, # customWidth
				size,
				style.id)
		else:
			return "<col min=\"%d\" max=\"%d\">" % (col, col)
	
	def get_row_xml_string(self, row):
		if row in self._row_styles and not self._row_styles[row].is_default:
			style = self._row_styles[row]
			if style.size == -1:
				size = 0
				for x, r in self._cells.items():
					for y, cell in r.items():
						try:
							font_size = self._styles[x][y].font.size
						except:
							font_size = 11
						size = max(font_size * (cell.count('\n') + 1) + 4, size)
			else:
				size = style.size if style.size else 15
			return "<row r=\"%d\" s=\"%d\" customFormat=\"1\" hidden=\"%d\" customHeight=\"%d\" ht=\"%f\">" % (
				row,
				style.id,
				1 if style.size == 0 else 0, # hidden
				1 if style.size is not None else 0, # customHeight
				size
			)
		else:
			return "<row r=\"%d\">" % row
		
	def get_xml_data(self):
		# Precondition: styles are aligned. if not, then :v
		for x, row in six.iteritems(self._cells):
			row_data = []
			for y, cell in six.iteritems(self._cells[x]):
				if x not in self._styles or y not in self._styles[x]:
					style = None
				else:
					style = self._styles[x][y]
				row_data.append(self.__get_cell_data(cell, x, y, style))
			yield x, row_data

########NEW FILE########
__FILENAME__ = Writer
import os
import sys
import tempfile
from zipfile import ZipFile, ZIP_DEFLATED
from datetime import datetime
import time
from jinja2 import Environment, FileSystemLoader
from . import Color

if getattr(sys, 'frozen', False):
	_basedir = os.path.join(sys._MEIPASS, 'pyexcelerate')
else:
	_basedir = os.path.dirname(__file__)
_TEMPLATE_PATH = os.path.join(_basedir, 'templates')

class Writer(object):
	env = Environment(loader=FileSystemLoader(_TEMPLATE_PATH), auto_reload=False)
	_docProps_app_template = env.get_template("docProps/app.xml")
	_docProps_core_template = env.get_template("docProps/core.xml")
	_content_types_template = env.get_template("[Content_Types].xml")
	_rels_template = env.get_template("_rels/.rels")
	_styles_template = env.get_template("xl/styles.xml") 
	_workbook_template = env.get_template("xl/workbook.xml")
	_workbook_rels_template = env.get_template("xl/_rels/workbook.xml.rels")
	_worksheet_template = env.get_template("xl/worksheets/sheet.xml")

	def __init__(self, workbook):
		self.workbook = workbook

	def _render_template_wb(self, template, extra_context=None):
		context = {'workbook': self.workbook}
		if extra_context:
			context.update(extra_context)
		return template.render(context).encode('utf-8')

	def _get_utc_now(self):
		now = datetime.utcnow()
		return now.strftime("%Y-%m-%dT%H:%M:00Z")


	def save(self, f):
		zf = ZipFile(f, 'w', ZIP_DEFLATED)
		zf.writestr("docProps/app.xml", self._render_template_wb(self._docProps_app_template))
		zf.writestr("docProps/core.xml", self._render_template_wb(self._docProps_core_template, {'date': self._get_utc_now()}))
		zf.writestr("[Content_Types].xml", self._render_template_wb(self._content_types_template))
		zf.writestr("_rels/.rels", self._rels_template.render().encode('utf-8'))
		if self.workbook.has_styles:
			zf.writestr("xl/styles.xml", self._render_template_wb(self._styles_template))
		zf.writestr("xl/workbook.xml", self._render_template_wb(self._workbook_template))
		zf.writestr("xl/_rels/workbook.xml.rels", self._render_template_wb(self._workbook_rels_template))
		for index, sheet in self.workbook.get_xml_data():
			tfd, tfn = tempfile.mkstemp()
			tf = os.fdopen(tfd, 'wb')
			sheetStream = self._worksheet_template.generate({'worksheet': sheet})
			for s in sheetStream:
				tf.write(s.encode('utf-8'))
			tf.close()
			zf.write(tfn, "xl/worksheets/sheet%s.xml" % (index))
			os.remove(tfn)
		zf.close()

########NEW FILE########
__FILENAME__ = _version
__version__ = '0.6.1'

########NEW FILE########
