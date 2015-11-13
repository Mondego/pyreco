__FILENAME__ = objects

# -*- coding: utf-8 -*-
from ...objects import *
from ...errors import *
from ...common import get_logger
from ...metadata import Field, FieldList
from ...stores import DataStore
from ...datautil import collapse_record

__all__ = (
        "MongoDBStore",
        "MongoDBCollection",
    )

try:
    import pymongo
except:
    from ...common import MissingPackage
    pymongo = MissingPackage("pymongo", "MongoDB streams", "http://www.mongodb.org/downloads/")

"""List of default shared stores."""
_default_stores = {}

def default_store(database, host, port):
    """Gets a default store for connectable or URL. If store does not exist
    one is created and added to shared default store pool."""

    key = (database, host, port)
    try:
        store = _default_stores[key]
    except KeyError:
        store = MongoDBStore(database, host=host, port=port)
        _default_stores[key] = store

    return store

class MongoDBStore(DataStore):
    __identifier__ = "mongo"

    def __init__(self, database, host='localhost', port=27017, client=None):
        """Creates a MongoDB data object store."""

        if client and (host or port):
            raise ArgumentError("Either client or host/port should be "
                                "specified, not both.")

        if client:
            self.client = client
        else:
            self.client = pymongo.MongoClient(host=host, port=port)

        self.database = self.client[database]

        self.host = host
        self.port = port

    def objects(self, names=None):
        raise NotImplementedError

    def create(self, name, fields, replace=False):
        obj = MongoDBCollection(name, fields=fields, store=self)

        if replace:
            # The only way how to check for collection existence is to look if
            # there are any objects in it
            if len(obj) > 0:
                raise ObjectExistsError("Collection %s already exists" % name)
            else:
                obj.truncate()

        return obj

    def get_object(self, name):
        obj = MongoDBCollection(name, store=self)


class MongoDBCollection(DataObject):
    """docstring for ClassName
    """
    __identifier__ = "mongo"

    def __init__(self, collection, fields, truncate=False,
                 expand=False,
                 database=None, host='localhsot', port=27017,
                 store=None):
        """Creates a MongoDB data object.

        Attributes
        * `collection`: mongo collection name
        * `database`: database name
        * `host`: mongo database server host, default is ``localhost``
        * `port`: mongo port, default is ``27017``
        * `expand`: expand dictionary values and treat children as top-level
          keys with dot '.' separated key path to the child.
        * `store`: MongoDBStore owning the object

        Specify either store or database, not both.
        """
        super(MongoDBCollection, self).__init__()

        if store and database:
            raise ArgumentError("Both store and database spectified")

        if store:
            self.store = store
        else:
            self.store = _default_store(database, host, port)

        if isinstance(collection, str):
            self.collection = self.store.database[collection]
        else:
            self.collection = collection

        if expand is None:
            self.expand = store.expand
        else:
            self.expand = expand

        if not fields:
            raise NotImplementedError("MongoDB field detection is not yet "
                                      "implemented, please specify them "
                                      "manually")
        self.fields = fields

        if truncate:
            self.truncate()

    def clone(self, fields=None, expand=None):
        fields = fields or self.fields
        if expand is None:
            expand = self.expand

        return MongoDBCollection(collection=self.collection, store=self.store,
                             fields=fields, expand=expand)

    def representations(self):
        return ["mongo", "records", "rows"]

    def is_consumable(self):
        return False

    def truncate(self):
        self.collection.remove()

    def __len__(self):
        return self.collection.count()

    def rows(self):
        fields = self.fields.names()
        iterator = self.collection.find(fields=fields)
        return MongoDBRowIterator(iterator, fields, self.expand)

    def records(self):
        fields = self.field.names()
        iterator = self.collection.find(fields=fields)
        return MongoDBRecordIterator(iterator, self.expand)

    def append(self, obj):
        if type(obj) == dict:
            record = obj
        else:
            record = dict(zip(self.fields.names(), obj))

        if self.expand:
            record = expand_record(record)

        self.collection.insert(record)

class MongoDBRowIterator(object):
    """Wrapper for pymongo.cursor.Cursor to be able to return rows() as tuples and records() as
    dictionaries"""
    def __init__(self, cursor, field_names, expand):
        self.cursor = cursor
        self.field_names = field_names
        self.expand = expand

    def __iter__(self):
        return self

    def __next__(self):
        record = next(self.cursor)

        if not record:
            raise StopIteration

        if self.expand:
            row = []
            for field in self.field_names:
                value = record
                for key in field.split('.'):
                    if key in value:
                        value = value[key]
                    else:
                        break
                array.append(value)
        else:
            row = [record[field] for field in self.field_names]

        return row

class MongoDBRecordIterator(object):
    """Wrapper for pymongo.cursor.Cursor to be able to return rows() as tuples and records() as
    dictionaries"""
    def __init__(self, cursor, expand=False):
        self.cursor = cursor
        self.expand = expand

    def __iter__(self):
        return self

    def __next__(self):
        record = next(cursor)

        if not record:
            raise StopIteration

        if not self.expand:
            return record
        else:
            return collapse_record(record)

########NEW FILE########
__FILENAME__ = ops
# -*- coding: utf-8 -*-
from .objects import MongoDBCollection

from ...metadata import *
from ...errors import *
from ...prototypes import *
from ...objects import *

def prepare_mongo_key(key):
    key = prepare_key(key)
    return {name:1 for name in key}


#############################################################################
# Metadata Operations

@field_filter.register("mongo")
def _(ctx, obj, keep=None, drop=None, rename=None, filter=None):

    if rename:
        raise NotImplementedError("Renaming of MongoDB fields is not "
                                    "implemented")

    if filter:
        if keep or drop or rename:
            raise OperationError("Either filter or keep, drop, rename should "
                                 "be used")
        else:
            field_filter = filter
    else:
        field_filter = FieldFilter(keep=keep, drop=drop, rename=rename)

    new_fields = field_filter.filter(obj.fields)

    # No need to actually do anything just pretend that we have new fields.

    return obj.clone(fields=new_fields)

#############################################################################
# Row Operations


@distinct.register("mongo")
def _(ctx, obj, key=None):

    if not key:
        key = obj.fields.names()

    key = prepare_mongo_key(key)

    new_fields = obj.fields.fields(key)
    cursor = obj.collection.group(key, {}, {}, "function(obj, prev){}")
    return IterableRecordsDataSource(cursor, new_fields)


########NEW FILE########
__FILENAME__ = objects
# -*- coding: utf-8 -*-
from ...objects import *
from ...errors import *
from ...common import get_logger
from ...metadata import Field, FieldList
from ...stores import DataStore

__all__ = (
        "SQLDataStore",
        "SQLTable",
        "SQLStatement",
        "reflect_fields"
    )

try:
    import sqlalchemy
    import sqlalchemy.sql as sql
    # (sql type, storage type, analytical type)
    _sql_to_bubbles_types = (
        (sqlalchemy.types.UnicodeText, "text", "typeless"),
        (sqlalchemy.types.Text, "text", "typeless"),
        (sqlalchemy.types.Unicode, "string", "nominal"),
        (sqlalchemy.types.String, "string", "nominal"),
        (sqlalchemy.types.Integer, "integer", "discrete"),
        (sqlalchemy.types.Numeric, "float", "measure"),
        (sqlalchemy.types.DateTime, "datetime", "typeless"),
        (sqlalchemy.types.Date, "date", "typeless"),
        (sqlalchemy.types.Time, "time", "typeless"),
        (sqlalchemy.types.Interval, "unknown", "typeless"),
        (sqlalchemy.types.Boolean, "boolean", "flag"),
        (sqlalchemy.types.Binary, "binary", "typeless")
    )

    concrete_sql_type_map = {
        "string": sqlalchemy.types.Unicode,
        "text": sqlalchemy.types.UnicodeText,
        "date": sqlalchemy.types.Date,
        "time": sqlalchemy.types.Time,
        "datetime": sqlalchemy.types.DateTime,
        "binary": sqlalchemy.types.Binary,
        "integer": sqlalchemy.types.Integer,
        "number": sqlalchemy.types.Numeric,
        "boolean": sqlalchemy.types.SmallInteger
    }

    from sqlalchemy.sql.expression import Executable, ClauseElement
    from sqlalchemy.ext.compiler import compiles

    class CreateTableAsSelect(Executable, ClauseElement):
        def __init__(self, table, select):
            self.table = table
            self.select = select

    @compiles(CreateTableAsSelect)
    def visit_create_table_as_select(element, compiler, **kw):
        return "CREATE TABLE %s AS (%s)" % (
            element.table,
            compiler.process(element.select)
        )

    class InsertFromSelect(Executable, ClauseElement):
        _execution_options = \
            Executable._execution_options.union({'autocommit': True})

        def __init__(self, table, select):
            self.table = table
            self.select = select

    @compiles(InsertFromSelect)
    def visit_insert_from_select(element, compiler, **kw):
        # TODO: there is different syntax for postgresql: uses (query)
        preparer = compiler.dialect.preparer(compiler.dialect)
        colnames = [preparer.format_column(c) for c in element.select.columns]
        collist = ", ".join(colnames)
        return "INSERT INTO %s (%s) %s" % (
            compiler.process(element.table, asfrom=True),
            collist,
            compiler.process(element.select)
        )

except ImportError:
    from ...common import MissingPackage
    sqlalchemy = MissingPackage("sqlalchemy", "SQL streams", "http://www.sqlalchemy.org/",
                                comment = "Recommended version is > 0.7")
    _sql_to_bubbles_types = ()
    concrete_sql_type_map = {}

# Length of string data types such as varchar in dialects that require length
# to be specified. This value is used when no `Field.size` is specified.
DEFAULT_STRING_LENGTH = 126


def concrete_storage_type(field, type_map={}, dialect=None):
    """Derives a concrete storage type for the field based on field conversion
       dictionary"""

    concrete_type = field.concrete_storage_type

    if not isinstance(concrete_type, sqlalchemy.types.TypeEngine):
        if type_map:
            concrete_type = type_map.get(field.storage_type)

        if not concrete_type:
            concrete_type = concrete_sql_type_map.get(field.storage_type)

        # Account for type specific options, like "length"
        if (concrete_type == sqlalchemy.types.Unicode):
            # TODO: add all dialects that require length for strings
            if not field.size and dialect in ("mysql", ):
                length = DEFAULT_STRING_LENGTH
            else:
                length = field.size

            concrete_type = sqlalchemy.types.Unicode(length=length)

        if not concrete_type:
            raise ValueError("unable to find concrete storage type for field '%s' "
                                "of type '%s' (concrete: %s)" \
                            % (field.name, field.storage_type,
                                field.concrete_storage_type))

    return concrete_type

"""List of default shared stores."""
_default_stores = {}

def reflect_fields(selectable):
    """Get fields from a table. Field types are normalized to the bubbles
    data types. Analytical type is set according to a default conversion
    dictionary."""

    fields = []

    for column in selectable.columns:
        field = Field(name=column.name)
        field.concrete_storage_type = column.type

        for conv in _sql_to_bubbles_types:
            if issubclass(column.type.__class__, conv[0]):
                field.storage_type = conv[1]
                field.analytical_type = conv[2]
                break

        if field.storage_type is not None:
            field.storaget_tpye = "unknown"

        if field.analytical_type is not None:
            field.analytical_type = "unknown"

        fields.append(field)

    return FieldList(*fields)

def default_store(url=None, connectable=None, schema=None):
    """Gets a default store for connectable or URL. If store does not exist
    one is created and added to shared default store pool."""

    if url and connectable:
        raise ArgumentError("Only one of URL or connectable should be " \
                            "specified, not both")

    if url:
        try:
            store = _default_stores[url]
        except KeyError:
            store = SQLDataStore(url=url, schema=schema)
            _default_stores[url] = store
            _default_stores[store.connectable] = store
    else:
        try:
            store = _default_stores[connectable]
        except KeyError:
            store = SQLDataStore(connectable=connectable)
            _default_stores[store.connectable] = store

    return store

def _postgres_copy_from(self, connection, table, stream, is_csv=True,
                         null_string=None, delimiter=None):
    """Loads data from file-like object `stream` into a `table`."""
    connection = engine.raw_connection()

    cursor = connection.cursor()

    preparer = sqlalchemy.sql.compiler.IdentifierPreparer("postgres")
    table_name = format_table(table, use_schema=True)

    options = []

    if is_csv:
        options.append("CSV")
    if null_string:
        options.append("NULL '%s'" % null_string)
    if delimiter:
        options.append("DELIMITER '%s'" % delimiter)

    if options:
        options_string = "WITH "+" ".join(options)
    else:
        options_string = ""

    sql = "COPY %s FROM STDIN %s" % (table_name, options_string)

    cursor.copy_expert(sql, stream)
    cursor.close()
    connection.commit()

class SQLDataStore(DataStore):
    """Holds context of SQL store operations."""

    __extension_name__ = "sql"

    _bubbles_info = {
        "options": [
            {"name":"url", "description": "Database URL"},
            {"name":"schema", "description":"Database schema"}
        ],
        "requirements": ["sqlalchemy"]
    }

    def __init__(self, url=None, connectable=None, schema=None,
            concrete_type_map=None, sqlalchemy_options=None):
        """Opens a SQL data store.

        * `url` – connection URL (see SQLAlchemy documentation for more
          information)
        * `connectable` – a connection or an engine
        * `schema` – default database schema
        * `concrete_Type_map` – a dictionary where keys are generic storage
          types and values are concrete storage types
        * `sqlalchemy_options` – options passed to `create_engine()`

        Either `url` or `connectable` should be specified, but not both.
        """

        if not url and not connectable:
            raise AttributeError("Either url or connectable should be provided" \
                                 " for SQL data source")

        super(SQLDataStore, self).__init__()

        if connectable:
            self.connectable = connectable
            self.should_close = False
        else:
            sqlalchemy_options = sqlalchemy_options or {}
            self.connectable = sqlalchemy.create_engine(url,
                    **sqlalchemy_options)
            self.should_close = True

        self.concrete_type_map = concrete_type_map or concrete_sql_type_map

        self.metadata = sqlalchemy.MetaData(bind=self.connectable)
        self.schema = schema
        self.logger = get_logger()

    def clone(self, schema=None, concrete_type_map=None):
        store = SQLDataStore(connectable=self.connectable,
                             schema=schema or self.schema,
                             concrete_type_map=concrete_type_map or
                                                     self.concrete_type_map
                             )
        return store

    def objects(self, names=None):
        """Return list of tables and views.

        * `names`: only objects with given names are returned
        """

        self.metadata.reflect(schema=self.schema, views=True, only=names)
        tables = self.metadata.sorted_tables

        objects = []
        for table in tables:
            obj = SQLTable(table=table, schema=self.schema, store=self)
            objects.append(obj)

        return objects

    def get_object(self, name):
        """Returns a `SQLTable` object for a table with name `name`."""

        obj = SQLTable(table=name, schema=self.schema, store=self)
        return obj

    def exists(self, name):
        try:
            table = self.table(name)
            return table.exists()
        except NoSuchObjectError:
            return False

    def statement(self, statement, fields=None):
        """Returns a statement object belonging to this store"""
        if not fields:
            fields = reflect_fields(statement)

        return SQLStatement(statement=statement,
                            store=self,
                            fields=fields,
                            schema=self.schema)

    def create(self, name, fields, replace=False, from_obj=None, schema=None,
               id_column=None):
        """Creates a table and returns `SQLTable`. See `create_table()`
        for more information"""
        table = self.create_table(name, fields=fields, replace=replace,
                                  from_obj=from_obj, schema=schema,
                                  id_column=id_column)
        return SQLTable(store=self, table=table, fields=fields,
                                schema=schema)

    def create_table(self, name, fields, replace=False, from_obj=None, schema=None,
               id_column=None):
        """Creates a new table.

        * `fields`: field list for new columns
        * `replace`: if table exists, it will be dropped, otherwise an
          exception is raised
        * `from_obj`: object with SQL selectable compatible representation
          (table or statement)
        * `schema`: schema where new table is created. When ``None`` then
          store's default schema is used.
        """

        schema = schema or self.schema

        table = self.table(name, schema, autoload=False)
        if table.exists():
            if replace:
                self.delete(name, schema)
                # Create new table object
                table = self.table(name, schema, autoload=False)
            else:
                schema_str = " (in schema '%s')" % schema if schema else ""
                raise ObjectExistsError("Table %s%s already exists" % (table, schema_str))

        if from_obj:
            if id_column:
                raise ArgumentError("id_column should not be specified when "
                                    "creating table from another object")

            return self._create_table_from(table, from_obj)
        elif id_column:
            sequence_name = "seq_%s_%s" % (name, id_column)
            sequence = sqlalchemy.schema.Sequence(sequence_name,
                                                  optional=True)
            col = sqlalchemy.schema.Column(id_column,
                                           sqlalchemy.types.Integer,
                                           sequence, primary_key=True)
            table.append_column(col)

        for field in fields:
            concrete_type = concrete_storage_type(field, self.concrete_type_map)
            if field.name == id_column:
                sequence_name = "seq_%s_%s" % (name, id_column)
                sequence = sqlalchemy.schema.Sequence(sequence_name,
                                                      optional=True)
                col = sqlalchemy.schema.Column(id_column,
                                           concrete_type,
                                           sequence, primary_key=True)
            else:
                col = sqlalchemy.schema.Column(field.name, concrete_type)

            table.append_column(col)

        table.create()

        return table

    def _create_table_from(table, from_obj):
        """Creates a table using ``CREATE TABLE ... AS SELECT ...``. The
        `from_obj` should have SQL selectable compatible representation."""

        source = from_obj.selectable()
        statement = CreateTableAsSelect(table, source)
        self.connectable.execute(statement)
        return self.table(name=table, autoload=True)

    def delete(self, name, schema):
        """Drops table"""
        schema = schema or self.schema
        table = self.table(name, schema, autoload=False)
        if not table.exists():
            raise Exception("Trying to delete table '%s' that does not exist" \
                                                                    % name)
        table.drop(checkfirst=False)
        self.metadata.drop_all(tables=[table])
        self.metadata.remove(table)

    def table(self, table, schema=None, autoload=True):
        """Returns a table with `name`. If schema is not provided, then
        store's default schema is used."""
        if table is None:
            raise Exception("Table name should not be None")
        if isinstance(table, sqlalchemy.schema.Table):
            return table

        schema = schema or self.schema

        try:
            return sqlalchemy.Table(table, self.metadata,
                                    autoload=autoload, schema=schema,
                                    autoload_with=self.connectable)
        except sqlalchemy.exc.NoSuchTableError:
            if schema:
                slabel = " in schema '%s'" % schema
            else:
                slabel = ""

            raise NoSuchObjectError("Unable to find table '%s'%s" % \
                                    (table, slabel))

    def execute(self, statement, *args, **kwargs):
        """Executes `statement` in store's connectable"""
        # TODO: Place logging here
        self.logger.debug("EXECUTE SQL: %s" % str(statement))
        return self.connectable.execute(statement, *args, **kwargs)

class SQLDataObject(DataObject):
    _bubbles_info = { "abstract": True }

    def __init__(self, store=None, schema=None):
        """Initializes new `SQLDataObject`. `store` might be a `SQLDataStore`
        object, a URL string or SQLAlchemy connectable object. If it is
        not a concrete store, then default store for that URL/connectable is
        created and/or reused if already exists. `schema` is a database schema
        for this object. It might be different from `store`'s schema."""

        if isinstance(store, SQLDataStore):
            self.store = store
        elif isinstance(store, str):
            self.store = default_store(url=store, schema=schema)
        else:
            self.store = default_store(connectable=store, schema=schema)

        self.schema = schema

    def can_compose(self, obj):
        """Returns `True` if `obj` can be composed with the receiver – that
        is, whether the target object is also a SQL object within same
        engine"""

        if not isinstance(obj, SQLDataObject):
            return False
        if not hasattr(obj, "store"):
            return False
        if self.store.connectable == obj.store.connectable:
            return True
        else:
            return False

    def records(self):
        # SQLAlchemy result is dict-like object where values can be accessed
        # by field names as well, so we just return the same iterator
        return self.rows()

    def __iter__(self):
        return self.rows()

    def is_consumable(self):
        return False

    def clone_statement(self, statement=None, fields=None):
        """Clone statement representation from the receiver. If `statement` or
        `fields` are not specified, then they are copied from the receiver.

        Use this method in operations to create derived statements from the
        receiver.
        """
        if statement is None:
            statement = self.sql_statement()

        fields = fields or self.fields.clone()
        obj = SQLStatement(statement, self.store, fields=fields,
                                    schema=self.schema)
        return obj

class SQLStatement(SQLDataObject):
    """Object representing a SQL statement (from SQLAlchemy)."""

    _bubbles_info = {
        "attributes": [
            {"name":"statement", "description": "SQL statement"},
            {"name":"store", "description":"SQL data store"},
            {"name":"schema", "description":"default schema"},
            {"name":"fields", "description":"statement fields (columns)"}
        ],
        "requirements": ["sqlalchemy"]
    }

    def __init__(self, statement, store, fields=None, schema=None):
        """Creates a relational database data object.

        Attributes:

        * `statement`: SQL statement to be used as a data source
        * `store`: SQL data store the object belongs to. Might be a
          `SQLDataStore` instance, URL string or a connectable object.
        * `schema` - database schema, if different than schema of `store`
        * `fields` - list of fields that override automatic field reflection
          from the statement

        If `store` is not provided, then default store is used for given
        connectable or URL. If no store exists, one is created.
        """

        super(SQLStatement, self).__init__(store=store, schema=schema)

        self.statement = statement
        try:
            self.name = statement.name
        except AttributeError:
            schema_label = "%s_" % schema if schema else ""
            self.name = "anonymous_statement_%s%d" % (schema_label, id(self))

        self.statement = statement

        if fields:
            self.fields = fields
        else:
            self.fields = reflect_fields(statement)

    def as_target(self):
        raise DataObjectError("SQL statement (%s) can not be used "
                                "as target object" % self.name)
    def __len__(self):
        """Returns number of rows selected by the statement."""
        cnt = sqlalchemy.sql.func.count(1)
        statement = sql.expression.select([cnt], from_obj=self.statement)

        return self.store.connectable.scalar(statement)

    def rows(self):
        return iter(self.store.execute(self.statement))

    def selectable(self):
        return self.statement

    def sql_statement(self):
        return self.statement

    def representations(self):
        """Return list of possible object representations"""
        return ["sql", "rows", "records"]

    def columns(self, fields=None):
        """Returns Column objects for `fields`. If no `fields` are specified,
        then all columns are returned"""

        if fields is None:
            return self.statement.columns

        cols = [self.statement.c[str(field)] for field in fields]

        return cols

    def column(self, field):
        """Returns a column for field"""
        return self.statement.c[str(field)]

class SQLTable(SQLDataObject):
    """Object representing a SQL database table or view (from SQLAlchemy)."""

    _bubbles_info = {
        "attributes": [
            {"name":"table", "description": "table name"},
            {"name":"store", "description":"SQL data store"},
            {"name":"schema", "description":"default schema"},
            {"name":"fields", "description":"statement fields (columns)"},

            {"name":"buffer_size", "description":"size of insert buffer"},
            {
                "name":"create",
                "description":"flag whether table is created",
                "type":"boolean"
            },
            {
                "name":"truncate",
                "description":"flag whether table is truncated",
                "type":"boolean"
            },
            {
                "name":"replace",
                "description":"flag whether table is replaced when created",
                "type":"boolean"
            },
        ],
        "requirements": ["sqlalchemy"]
    }

    def __init__(self, table, store, fields=None, schema=None,
                 create=False, replace=False, truncate=False,
                 id_key_name=None, buffer_size=1024):
        """Creates a relational database data object.

        Attributes:

        * `table`: table name
        * `store`: SQL data store the object belongs to
        * `fields`: fieldlist for a new table
        * `truncate`: if `True` table will be truncated upon initiaization
        * `create`: if `True` then table will be created during object
          initialization. Default is `False` and exception is raised when
          table does not exist.
        * `replace`:  if `True` and `create` is requested, then existing table
          will be dropped and created as new.
        * `id_key_name`: name of the auto-increment key during table creation.
            If specified, then key column is created, otherwise no key column
            is created.
        * `buffer_size`: size of buffer for table INSERTs - how many records
          are collected before they are inserted using multi-insert statement.
          Default is 1000.

        """

        super().__init__(store=store, schema=schema)

        self.fields = None

        if create:
            if fields is None:
                raise ArgumentError("No fields specified for new table")

            self.fields = fields
            self.table = self.store.create_table(table, self.fields,
                                                 replace=replace,
                                                 schema=schema,
                                                 id_column=id_key_name)
            # FIXME: reflect concrete storage types
        else:
            if isinstance(table, str):
                self.table = self.store.table(table, schema=schema, autoload=True)
            else:
                self.table = table

            if not self.fields:
                self.fields = reflect_fields(self.table)

        self.name = self.table.name

        if truncate:
            self.table.delete().execute()

        self._field_names = self.fields.names()

        # Bulk INSERT buffer (if backends supports bulk inserts)
        self.buffer_size = buffer_size
        self._insert_buffer = []
        self.insert_statement = self.table.insert()
        # SQL Statement representation

    def rows(self):
        return iter(self.table.select().execute())

    def representations(self):
        """Return list of possible object representations"""
        return ["sql_table", "sql", "records", "rows"]

    def selectable(self):
        return self.table.select()

    def sql_statement(self):
        return self.table

    def sql_table(self):
        return self.table

    def __len__(self):
        """Returns number of rows in a table"""
        statement = self.table.count()
        result = self.store.connectable.scalar(statement)
        return result

    def truncate(self):
        if self.table is None:
            raise RepresentationError("Can not truncate: "
                                      "SQL object is a statement not a table")
        self.store.execute(self.table.delete())

    def append(self, row):
        # FIXME: remove this once SQLAlchemy will allow list based multi-insert
        row = dict(zip(self._field_names, row))

        self._insert_buffer.append(row)
        if len(self._insert_buffer) >= self.buffer_size:
            self.flush()

    def flush(self):
        if self._insert_buffer:
            self.store.execute(self.insert_statement, self._insert_buffer)
            self._insert_buffer = []

    def append_from(self, obj):
        """Appends data from object `obj` which might be a `DataObject`
        instance or an iterable. If `obj` is a `DataObject`, then it should
        have one of following representations, listed in order of preferrence:

        * sql_table
        * sql_statement
        * rows

        If the `obj` is just an iterable, then it is treated as `rows`
        representation of data object.

        `flush()` is called:

        * before SQL insert from a table or a statement
        * after insert of all rows of `rows` representation
        """

        # TODO: depreciate this in favor of the insert() operation
        reprs = obj.representations()

        if self.can_compose(obj):
            self.store.logger.debug("append_from: composing into %s" %
                                                                self.name)
            # Flush all data that were added through append() to preserve
            # insertion order (just in case)
            self.flush()

            # Preare INSERT INTO ... SELECT ... statement
            source = obj.selectable()
            statement = InsertFromSelect(self.table, source)
            self.store.execute(statement)

        elif "rows" in reprs:
            self.store.logger.debug("append_from: appending rows into %s" %
                                                                self.name)
            for row in obj.rows():
                self.append(row)

            # Clean-up after bulk insertion
            self.flush()

        else:
            raise RepresentationError(
                            "Incopatible representations '%s'", (reprs, ) )

    def probe(self, probeset):
        return probe_table(self, probeset)

    def columns(self, fields=None):
        """Returns Column objects for `fields`. If no `fields` are specified,
        then all columns are returned"""

        if fields is None:
            return self.table.columns

        cols = [self.table.c[str(field)] for field in fields]

        return cols

    def column(self, field):
        """Returns a column for field"""
        return self.table.c[str(field)]


########NEW FILE########
__FILENAME__ = ops
import functools
from ...operation import RetryOperation
from ...prototypes import *
from ...metadata import Field, FieldList, FieldFilter
from ...metadata import prepare_aggregation_list, prepare_order_list
from ...objects import IterableDataSource
from ...errors import *
from .utils import prepare_key, zip_condition, join_on_clause

try:
    import sqlalchemy
    from sqlalchemy import sql
except ImportError:
    from ...common import MissingPackage
    sqlalchemy = MissingPackage("sqlalchemy", "SQL streams", "http://www.sqlalchemy.org/",
                                comment = "Recommended version is > 0.7")
    sql = sqlalchemy

def _unary(func):
    @functools.wraps(func)
    def decorator(ctx, obj, *args, **kwargs):
        result = func(ctx, obj.sql_statement(), *args, **kwargs)
        return obj.clone_statement(statement=result)

    return decorator

#############################################################################
# Metadata Operations

@field_filter.register("sql")
def _(ctx, obj, keep=None, drop=None, rename=None):
    """Returns a statement with fields according to the field filter"""
    # TODO: preserve order of "keep" -> see FieldFilter

    field_filter = FieldFilter(keep=keep, drop=drop, rename=rename)

    statement = obj.sql_statement()
    statement = statement.alias("__ff")

    columns = []

    for field in obj.fields:
        name = str(field)
        column = statement.c[name]
        if name in field_filter.rename:
            column = column.label(field_filter.rename[name])
        columns.append(column)

    row_filter = field_filter.row_filter(obj.fields)
    selection = row_filter(columns)

    statement = sql.expression.select(selection, from_obj=statement)
    fields = field_filter.filter(obj.fields)

    result = obj.clone_statement(statement=statement, fields=fields)
    return result


#############################################################################
# Row Operations


@filter_by_value.register("sql")
def _(ctx, src, key, value, discard=False):
    """Returns difference between left and right statements"""

    if isinstance(key, (list, tuple)) and \
                    not isinstance(value, (list, tuple)):
        raise ArgumentError("If key is compound then value should be as well")

    cols = src.columns()
    statement = src.sql_statement()
    filter_cols = src.columns(prepare_key(key))
    if len(filter_cols) == 1:
        value = (value, )
    condition = zip_condition(filter_cols, value)
    statement = sql.expression.select(cols, from_obj=statement,
                                        whereclause=condition)

    statement = statement.alias("__value_filter")
    return src.clone_statement(statement=statement)


@filter_by_range.register("sql")
def _(ctx, src, field, low, high, discard=False):
    """Filter by range: field should be between low and high."""

    statement = src.sql_statement()
    key_column = statement.c[str(key)]

    if low is not None and high is None:
        cond = key_column >= low
    elif high is not None and low is None:
        cond = key_column < high
    else:
        cond = sql.expression.between(key_column, low, high)

    if discard:
        cond = sql.expression.not_(cond)

    statement = sql.expression.select(statement.columns,
                                      from_obj=statement,
                                      whereclause=cond)

    # TODO: remove this in newer SQLAlchemy version
    statement = statement.alias("__range_filter")
    return src.clone_statement(statement=statement)


@filter_by_set.register("sql")
def _(ctx, obj, key, value_set, discard=False):
    raise RetryOperation(["rows"], reason="Not implemented")


@filter_not_empty.register("sql")
def _(ctx, obj, field):
    statement = obj.sql_statement()

    column = statement.c[str(field)]
    condition = column != None
    selection = obj.columns()

    statement = sql.expression.select(selection, from_obj=statement,
                                            whereclause=condition)

    return obj.clone_statement(statement=statement)

@filter_by_predicate.register("sql")
def _(ctx, obj, fields, predicate, discard=False):
    raise RetryOperation(["rows"], reason="Not implemented")


@distinct.register("sql")
def _(ctx, obj, keys=None):
    """Returns a statement that selects distinct values for `keys`"""

    statement = obj.sql_statement()
    if keys:
        keys = prepare_key(keys)
    else:
        keys = obj.fields.names()

    cols = [statement.c[str(key)] for key in keys]
    statement = sql.expression.select(cols, from_obj=statement, group_by=cols)

    fields = obj.fields.fields(keys)

    return obj.clone_statement(statement=statement, fields=fields)


@first_unique.register("sql")
def _(ctx, statement, keys=None):
    """Returns a statement that selects whole rows with distinct values
    for `keys`"""
    # TODO: use prepare_key
    raise NotImplementedError

@sample.register("sql")
@_unary
def _(ctx, statement, value, mode="first"):
    """Returns a sample. `statement` is expected to be ordered."""

    if mode == "first":
        return statement.select(limit=value)
    else:
        raise RetryOperation(["rows"], reason="Unhandled mode '%s'" % mode)


@sort.register("sql")
@_unary
def _(ctx, statement, orderby):
    """Returns a ordered SQL statement. `orders` should be a list of
    two-element tuples `(field, order)`"""

    # Each attribute mentioned in the order should be present in the selection
    # or as some column from joined table. Here we get the list of already
    # selected columns and derived aggregates

    orderby = prepare_order_list(orderby)

    columns = []
    for field, order in orderby:
        column = statement.c[str(field)]
        order = order.lower()
        if order.startswith("asc"):
            column = column.asc()
        elif order.startswith("desc"):
            column = column.desc()
        else:
            raise ValueError("Unknown order %s for column %s") % (order, column)

        columns.append(column)

    statement = sql.expression.select(statement.columns,
                                   from_obj=statement,
                                   order_by=columns)
    return statement

aggregation_functions = {
    "sum": sql.functions.sum,
    "min": sql.functions.min,
    "max": sql.functions.max,
    "count": sql.functions.count
}


@aggregate.register("sql")
def _(ctx, obj, key, measures=None, include_count=True,
              count_field="record_count"):

    """Aggregate `measures` by `key`"""

    keys = prepare_key(key)

    if measures:
        measures = prepare_aggregation_list(measures)
    else:
        measures = []

    out_fields = FieldList()
    out_fields += obj.fields.fields(keys)

    statement = obj.sql_statement()
    group = [statement.c[str(key)] for key in keys]

    selection = [statement.c[str(key)] for key in keys]

    for measure, agg_name in measures:
        func = aggregation_functions[agg_name]
        label = "%s_%s" % (str(measure), agg_name)
        aggregation = func(obj.column(measure)).label(label)
        selection.append(aggregation)

        # TODO: make this a metadata function
        field = obj.fields.field(measure)
        field = field.clone(name=label, analytical_type="measure")
        out_fields.append(field)

    if include_count:
        out_fields.append(Field(count_field,
                            storage_type="integer",
                            analytical_type="measure"))

        count =  sql.functions.count(1).label(count_field)
        selection.append(count)

    statement = sql.expression.select(selection,
                                      from_obj=statement,
                                      group_by=group)

    return obj.clone_statement(statement=statement, fields=out_fields)


#############################################################################
# Field Operations

@append_constant_fields.register("sql")
def _(ctx, obj, fields, values):
    statement = obj.sql_statement()

    new_fields = obj.fields + FieldList(*fields)

    selection = statement.c
    selection += values

    statement = sql.expression.select(selection, from_obj=statement)
    result = obj.clone_statement(statement=statement, fields=new_fields)

    return result

@dates_to_dimension.register("sql")
def _(ctx, obj, fields=None, unknown_date=0):
    """Update all date fields to be date IDs. `unknown_date` is a key to date
    dimension table for unspecified date (NULL in the source).
    `fields` is list of date fields. If not specified, then all date fields
    are considered."""

    statement = obj.sql_statement()
    statement = statement.alias("__to_dim")

    if fields:
        date_fields = obj.fields(fields)
    else:
        date_fields = obj.fields.fields(storage_type="date")

    selection = []
    fields = []
    for field in obj.fields:
        col = statement.c[str(field)]
        if field in date_fields:
            # Safe way
            year = sql.func.extract("year", col)
            month = sql.func.extract("month", col)
            day = sql.func.extract("day", col)
            key = year * 10000 + month * 100 + day
            col = sql.expression.case([ (col != None, key) ],
                    else_=unknown_date)

            # print("CAST TO: %s" % str(col))

            col = col.label(field.name)
            fields.append(field.clone(storage_type="integer",
                                      concrete_storage_type=None))
        else:
            fields.append(field.clone())

        selection.append(col)

    fields = FieldList(*fields)
    statement = sql.expression.select(selection,
                                      from_obj=statement)

    # TODO: mark date fields to be integers
    return obj.clone_statement(statement=statement, fields=fields)

@split_date.register("sql")
def _(ctx, obj, fields, parts=["year", "month", "day"]):
    """Extract `parts` from date objects replacing the original date field
    with parts field."""

    statement = obj.sql_statement()
    date_fields = prepare_key(fields)

    # Validate date fields
    for f in obj.fields.fields(date_fields):
        if f.storage_type != "date":
            raise FieldError("Field '%s' is not a date field" % f)

    # Prepare output fields
    fields = FieldList()
    proto = Field(name="p", storage_type="integer", analytical_type="ordinal")

    selection = []

    for field in obj.fields:
        name = str(field)
        col = statement.c[name]
        if name in date_fields:
            for part in parts:
                label = "%s_%s" % (str(field), part)
                fields.append(proto.clone(name=label))
                col_part = sql.expression.extract(part, col)
                col_part = col_part.label(label)
                selection.append(col_part)
        else:
            fields.append(field.clone())
            selection.append(col)

    statement = sql.expression.select(selection,
                                      from_obj=statement)

    return obj.clone_statement(statement=statement, fields=fields)

# @operation("sql")
# def string_to_date(ctx, obj, fields, fmt="%Y-%m-%dT%H:%M:%S.Z"):
# 
#     date_fields = prepare_key(fields)
# 
#     selection = []
#     # Prepare output fields
#     fields = FieldList()
#     for field in obj.fields:
#         col = statement.c[str(field)]
#         if str(field) in date_fields:
#             fields.append(field.clone(storage_type="date",
#                                       concrete_storage_type=None))
# 
#         else:
#             fields.append(field.clone())
#         selection.append(col)
# 
#     return IterableDataSource(iterator(indexes), fields)


#############################################################################
# Compositions

@append.register("sql[]")
def _(ctx, objects):
    """Returns a statement with sequentialy concatenated results of the
    `statements`. Statements are chained using ``UNION``."""
    first = objects[0]

    if not all(first.can_compose(o) for o in objects[1:]):
        raise RetryOperation(["rows", "rows[]"],
                             reason="Can not compose")

    statements = [o.sql_statement() for o in objects]
    statement = sqlalchemy.sql.expression.union(*statements)

    return first.clone_statement(statement=statement)


@join_details.register("sql", "sql")
def _(ctx, master, detail, master_key, detail_key):
    """Creates a master-detail join using simple or composite keys. The
    columns used as a key in the `detail` object are not included in the
    result.
    """
    # TODO: add left inner, left outer

    if not master.can_compose(detail):
        raise RetryOperation(["rows", "rows"], reason="Can not compose")

    master_key = prepare_key(master_key)
    detail_key = prepare_key(detail_key)

    master_stat = master.sql_statement().alias("__m")
    detail_stat = detail.sql_statement().alias("__d")

    # Prepare the ON left=right ... clause
    onclause = join_on_clause(master_stat, detail_stat, master_key, detail_key)

    # Prepare output fields and columns selection - the selection skips detail
    # columns that are used as key, because they are already present in the
    # master table.

    out_fields = master.fields.clone()
    selection = list(master_stat.columns)
    for field in detail.fields:
        if str(field) not in detail_key:
            out_fields.append(field)
            selection.append(detail_stat.c[str(field)])

    joined = sql.expression.join(master_stat,
                                 detail_stat,
                                 onclause=onclause)

    # Alias the output fields to match the field names
    aliased = []
    for col, field in zip(selection, out_fields):
        aliased.append(col.label(field.name))

    select = sql.expression.select(aliased,
                                from_obj=joined,
                                use_labels=True)

    return master.clone_statement(statement=select, fields=out_fields)

# TODO: depreciated
@join_details.register("sql", "sql[]", name="join_details")
def _(ctx, master, details, joins):
    """Creates left inner master-detail join (star schema) where `master` is an
    iterator if the "bigger" table `details` are details. `joins` is a list of
    tuples `(master, detail)` where the master is index of master key and
    detail is index of detail key to be matched.

    If `inner` is `True` then inner join is performed. That means that only
    rows from master that have corresponding details are returned.

    .. warning::

        all detail iterators are consumed and result is held in memory. Do not
        use for large datasets.
    """
    # TODO: update documentation

    if not details:
        raise ArgumentError("No details provided, nothing to join")

    if not joins:
        raise ArgumentError("No joins specified")

    if len(details) != len(joins):
        raise ArgumentError("For every detail there should be a join "
                            "(%d:%d)." % (len(details), len(joins)))

    if not all(master.can_compose(detail) for detail in details):
        raise RetryOperation(["rows", "rows[]"], reason="Can not compose")

    out_fields = master.fields.clone()

    master_stmt = master.sql_statement().alias("master")
    selection = list(master_stmt.columns)

    joined = master_stmt
    i = 0
    for detail, join in zip(details, joins):
        alias = "detail%s" % i
        det_stmt = detail.sql_statement().alias(alias)
        master_key = join["master"]
        detail_key = join["detail"]

        onclause = master_stmt.c[master_key] == det_stmt.c[detail_key]
        # Skip detail key in the output

        for field, col in zip(detail.fields, det_stmt.columns):
            if str(field) != str(detail_key):
                selection.append(col)
                out_fields.append(field.clone())

        joined = sql.expression.join(joined,
                                     det_stmt,
                                     onclause=onclause)

    aliased = []
    for col, field in zip(selection, out_fields):
        aliased.append(col.label(field.name))

    select = sql.expression.select(aliased,
                                from_obj=joined,
                                use_labels=True)

    return master.clone_statement(statement=select, fields=out_fields)


@added_keys.register("sql", "sql")
def _(ctx, src, target, src_key, target_key=None):
    """Returns difference between left and right statements"""

    # FIXME: add composition checking

    if not src.can_compose(target):
        raise RetryOperation(("rows", "sql"), reason="Can not compose")

    target_key = target_key or src_key

    src_cols = src.columns(prepare_key(src_key))
    target_cols = target.columns(prepare_key(target_key))

    src_stat = src.sql_statement()
    target_stat = target.sql_statement()

    src_selection = sql.expression.select(src_cols, from_obj=src_stat)
    target_selection = sql.expression.select(target_cols, from_obj=target_stat)

    diff = src_selection.except_(target_selection)

    return src.clone_statement(statement=diff)


@added_rows.register("sql", "sql")
def _(ctx, src, target, src_key, target_key=None):
    diff = ctx.added_keys(src, target, src_key, target_key)

    diff_stmt = diff.sql_statement()
    diff_stmt = diff_stmt.alias("__added_keys")
    src_stmt = src.sql_statement()

    key = prepare_key(src_key)

    src_cols = src.columns(key)
    diff_cols = [diff_stmt.c[f] for f in key]

    cond = zip_condition(src_cols, diff_cols)

    join = sql.expression.join(src_stmt, diff_stmt, onclause=cond)
    join = sql.expression.select(src.columns(), from_obj=join)

    return src.clone_statement(statement=join)


@added_rows.register("rows", "sql", name="added_rows")
def _(ctx, src, target, src_key, target_key=None):

    src_key = prepare_key(src_key)

    if target_key:
        target_key = prepare_key(target_key)
    else:
        target_key = src_key

    statement = target.sql_statement()
    target_cols = target.columns(target_key)

    field_filter = FieldFilter(keep=src_key).row_filter(src.fields)

    def iterator():
        for row in src.rows():
            row_key = field_filter(row)

            cond = zip_condition(target_cols, row_key)

            select = sql.expression.select([sql.func.count(1)],
                                           from_obj=statement,
                                           whereclause=cond)

            result = target.store.execute(select)
            result = list(result)
            if len(result) >= 1 and result[0][0] == 0:
                yield row

    return IterableDataSource(iterator(), fields=src.fields)


# TODO: dimension loading: new values
# TODO: create decorator @target that will check first argument whether it is
# a target or not

@changed_rows.register("sql", "sql")
def _(ctx, dim, source, dim_key, source_key, fields, version_field):
    """Return an object representing changed dimension rows.

    Arguments:
        * `dim` – dimension table (target)
        * `source` – source statement
        * `dim_key` – dimension table key
        * `source_key` – source table key
        * `fields` – fields to be compared for changes
        * `version_field` – field that is optionally checked to be empty (NULL)

    """
    src_columns = source.columns(fields)
    dim_columns = dim.columns(fields)
    src_stmt = source.sql_statement()
    dim_stmt = dim.sql_statement()

    # 1. find changed records
    # TODO: this might be a separate operation

    join_cond = zip_condition(dim.columns(prepare_key(dim_key)),
                              source.columns(prepare_key(source_key)))
    join = sql.expression.join(src_stmt, dim_stmt, onclause=join_cond)

    change_cond = [ d != s for d, s in zip(dim_columns, src_columns) ]
    change_cond = sql.expression.or_(*change_cond)

    if version_field:
        version_column = dim.column(version_field)
        change_cond = sql.expression.and_(change_cond, version_column == None)

    join = sql.expression.select(source.columns(),
                                 from_obj=join,
                                 whereclause=change_cond)

    return source.clone_statement(statement=join)

#############################################################################
# Loading

# TODO: continue here
@load_versioned_dimension.register("sql_table", "sql")
def _(ctx, dim, source, dim_key, fields,
                             version_fields=None, source_key=None):
    """Type 2 dimension loading."""
    # Now I need to stay in the same kernel!

    new = added_rows(dim, source)


#############################################################################
# Auditing


@count_duplicates.register("sql")
def _(ctx, obj, keys=None, threshold=1,
                       record_count_label="record_count"):
    """Returns duplicate rows based on `keys`. `threshold` is lowest number of
    duplicates that has to be present to be returned. By default `threshold`
    is 1. If no keys are specified, then all columns are considered."""

    if not threshold or threshold < 1:
        raise ValueError("Threshold should be at least 1 "
                         "meaning 'at least one duplcate'.")

    statement = obj.sql_statement()

    count_field = Field(record_count_label, "integer")

    if keys:
        keys = prepare_key(keys)
        group = [statement.c[str(field)] for field in keys]
        fields = list(keys)
        fields.append(count_field)
        out_fields = FieldList(*fields)
    else:
        group = list(statement.columns)
        out_fields = obj.fields.clone() + FieldList(count_field)

    counter = sqlalchemy.func.count("*").label(record_count_label)
    selection = group + [counter]
    condition = counter > threshold

    statement = sql.expression.select(selection,
                                   from_obj=statement,
                                   group_by=group,
                                   having=condition)

    result = obj.clone_statement(statement=statement, fields=out_fields)
    return result

@duplicate_stats.register("sql_statement")
def _(ctx, obj, fields=None, threshold=1):
    """Return duplicate statistics of a `statement`"""
    count_label = "__record_count"
    dups = duplicates(obj, threshold, count_label)
    statement = dups.statement
    statement = statement.alias("duplicates")

    counter = sqlalchemy.func.count("*").label("record_count")
    group = statement.c[count_label]
    result_stat = sqlalchemy.sql.expression.select([counter, group],
                                              from_obj=statement,
                                              group_by=[group])

    fields = dups.fields.clone()
    fields.add(count_label)

    result = obj.clone_statement(statement=result_stat, fields=fields)
    return result

@nonempty_count.register("sql")
def _(ctx, obj, fields=None):
    """Return count of empty fields for the object obj"""

    # FIXME: add fields
    # FIXME: continue here

    statement = obj.sql_statement()
    statement = statement.alias("empty")

    if not fields:
        fields = obj.fields
    fields = prepare_key(fields)

    cols = [statement.c[f] for f in fields]
    selection = [sqlalchemy.func.count(col) for col in cols]
    statement = sqlalchemy.sql.expression.select(selection,
                                                  from_obj=statement)

    out_fields = obj.fields.fields(fields)
    return obj.clone_statement(statement=statement, fields=out_fields)

    # field, key, key, key, empty_count

@distinct_count.register("sql")
def _(ctx, obj, fields=None):
    """Return count of empty fields for the object obj"""

    # FIXME: add fields
    # FIXME: continue here

    statement = obj.sql_statement()
    statement = statement.alias("distinct_count")

    if not fields:
        fields = obj.fields
    fields = prepare_key(fields)

    cols = [statement.c[f] for f in fields]
    selection = []
    for col in cols:
        c = sqlalchemy.func.count(sqlalchemy.func.distinct(col))
        selection.append(c)

    statement = sqlalchemy.sql.expression.select(selection,
                                                  from_obj=statement)

    out_fields = obj.fields.fields(fields)
    return obj.clone_statement(statement=statement, fields=out_fields)


#############################################################################
# Assertions


@assert_unique.register("sql")
def _(ctx, obj, key=None):
    """Checks whether the receiver has unique values for `key`. If `key` is
    not specified, then all fields from `obj` are considered."""

    statement = obj.sql_statement().alias("__u")

    if key:
        key = prepare_key(key)
        group = [statement.c[field] for field in key]
    else:
        group = list(statement.columns)

    counter = sqlalchemy.func.count("*").label("duplicate_count")
    selection = [counter]
    condition = counter > 1

    statement = sql.expression.select(selection,
                                       from_obj=statement,
                                       group_by=group,
                                       having=condition,
                                       limit=1)

    result = list(obj.store.execute(statement))

    if len(result) != 0:
        raise ProbeAssertionError

    return obj

@assert_contains.register("sql")
def _(ctx, obj, field, value):
    statement = obj.sql_statement().alias("__u")
    column = statement.c[str(field)]

    condition = column == value
    selection = [1]

    statement = sql.expression.select(selection,
                                       from_obj=statement,
                                       group_by=group,
                                       having=condition,
                                       limit=1)

    result = list(obj.store.execute(statement))
    if len(result) != 1:
        raise ProbeAssertionError

    return obj

@assert_missing.register("sql")
def _(ctx, obj, field, value):
    # statement = obj.sql_statement().alias("__u")
    statement = obj.sql_statement()
    column = statement.c[str(field)]

    condition = column == value
    selection = [1]

    statement = sql.expression.select(selection,
                                       from_obj=statement,
                                       whereclause=condition,
                                       limit=1)

    result = list(obj.store.execute(statement))
    if len(result) != 0:
        raise ProbeAssertionError

    return obj

#############################################################################
# Conversions


@as_records.register("sql")
def _(ctx, obj):
    """Return object with records representation."""
    # SQL Alchemy result can be used as both - records or rows, so we just
    # return the object:

    return obj


#############################################################################
# Loading

@insert.register("sql", "sql")
def _(ctx, source, target):
    if not target.can_compose(src):
        raise RetryOperation(["rows"])

    # Flush all data that were added through append() to preserve
    # insertion order (just in case)
    target.flush()

    # Preare INSERT INTO ... SELECT ... statement
    statement = InsertFromSelect(target.table, source.selectable())

    target.store.execute(statement)

    return target

@insert.register("rows", "sql")
def _(ctx, source, target):

    if len(source.fields) > len(target.fields):
         raise OperationError("Number of source fields %s is greater than "
                              "number of target fields %s" % (len(source.fields),
                                                             len(target.fields)))

    missing = set(source.fields.names()) - set(target.fields.names())
    if missing:
        raise OperationError("Source contains fields that are not in the "
                "target: %s" % (missing, ))

    indexes = []
    for name in target.fields.names():
        if name in source.fields:
            indexes.append(source.fields.index(name))
        else:
            indexes.append(None)

    target.flush()
    for row in source.rows():
        row = [row[i] if i is not None else None for i in indexes]
        target.append(row)

    target.flush()

    return target


########NEW FILE########
__FILENAME__ = utils
from ...errors import *
from ...metadata import Field, FieldList

__all__ = (
            "prepare_key",
            "zip_condition",
            "join_on_clause"
        )

def prepare_key(key):
    """Returns a list of columns for `key`. Key might be a list of fields or
    just one field represented by `Field` object or a string. Examples:

    >>> key = prepare_key(obj, "id")
    >>> key = prepare_key(obj, ["code", "name"])

    If `statement` is not `None`, then columns for the key are returned.
    """

    if isinstance(key, (str, Field)):
        key = (key, )

    return tuple(str(f) for f in key)


def join_on_clause(left, right, left_key, right_key):
    """Returns a conditional statement for joining `left` and `right`
    statements by keys."""

    left_cols = [left.c[str(key)] for key in left_key]
    right_cols = [right.c[str(key)] for key in right_key]

    return zip_condition(left_cols, right_cols)


def zip_condition(left, right):
    """Creates an equality condition by zipping `left` and `right` list of
    elements"""

    cond = [l == r for l, r in zip(left, right)]

    assert len(cond) >= 1
    if len(cond) > 1:
        cond = sql.expression.and_(*cond)
    else:
        cond = cond[0]

    return cond


########NEW FILE########
__FILENAME__ = objects
#!/usr/bin/env python
# -*- coding: utf-8 -*-

import csv
import io
import os.path
from collections import defaultdict, namedtuple
import itertools
from ...objects import *
from ...metadata import *
from ...errors import *
from ...resource import Resource
from ...stores import DataStore
import json
from datetime import datetime
from time import strptime
from base64 import b64decode
import json

__all__ = (
        "CSVStore",
        "CSVSource",
        "CSVTarget",
        )


CSVData = namedtuple("CSVData", ["handle", "dialect", "encoding", "fields"])

# TODO: add type converters
# TODO: handle empty strings as NULLs

class CSVStore(DataStore):
    def __init__(self, path, extension=".csv", role=None, **kwargs):
        super(CSVStore, self).__init__()
        self.path = path
        self.extension = extension
        if role:
            self.role = role.lower
        else:
            self.role = "source"

        self.kwargs = kwargs

    def get_object(self, name):
        """Returns a CSVSource object with filename constructed from store's
        path and extension"""
        fields_file = "%s_fields.json" % name
        fields_path = os.path.join(self.path, fields_file)

        name = name + self.extension
        path = os.path.join(self.path, name)

        args = dict(self.kwargs)
        if os.path.exists(fields_path):
            with open(fields_path) as f:
                metadata = json.load(f)
            args["fields"] = FieldList(*metadata)

        if self.role in ["s", "src", "source"]:
            return CSVSource(path, **args)
        elif self.role in ["t", "target"]:
            return CSVTarget(path, **args)
        else:
            raise ArgumentError("Unknown CSV object role '%s'" % role)

    def create(self, name, fields, replace=False):
        """Create a file"""
        name = name + self.extension
        path = os.path.join(self.path, name)

        target = CSVTarget(path, fields=fields, truncate=True)
        return target

class CSVSource(DataObject):
    """Comma separated values text file as a data source."""

    _bubbles_info = {
        "attributes": [
            {
                "name":"resource",
                "description": "file name, URL or a file handle with CVS data"
            },
            {
                "name":"fields",
                "description": "fields in the file. Should be set if read_header "
                               "is false"
            },
            {
                "name":"fields",
                "description": "flag determining whether first line contains "
                                "header or not. ``True`` by default."
            },
            {
                "name":"encoding",
                "description":"file encoding"
            },
            {
                "name":"read_header",
                "description":"flag whether file header is read or not"
            },
            {
                "name":"skip_rows",
                "description":"number of rows to be skipped"
            },
            {
                "name": "empty_as_null",
                "description": "Treat emtpy strings as NULL values"
            },
            {
                "name": "type_converters",
                "description": "dictionary of data type converters"
            }
        ]
    }

    def __init__(self, resource, read_header=True, dialect=None,
            delimiter=None, encoding=None, skip_rows=None,
            empty_as_null=True, fields=None, type_converters=None, **options):
        """Creates a CSV data source stream.

        * `resource`: file name, URL or a file handle with CVS data
        * `read_header`: flag determining whether first line contains header
          or not. ``True`` by default.
        * `encoding`: source character encoding, by default no conversion is
          performed.
        * `fields`: optional `FieldList` object. If not specified then
          `read_header` should be used.
        * `skip_rows`: number of rows to be skipped. Default: ``None``
        * `empty_as_null`: treat empty strings as ``Null`` values
        * `type_converters`: dictionary of converters (functions). It has
          to cover all knowd types.

        Note: avoid auto-detection when you are reading from remote URL
        stream.

        Rules for fields:

        * if `fields` are specified, then they are used, header is ignored
          depending on `read_header` flag
        * if `detect_types` is not requested, then each field is of type
          `string` (this is the default)
        """

        self.file = None

        if not any((fields, read_header)):
            raise ArgumentError("At least one of fields or read_header"
                                " should be specified")

        self.read_header = read_header
        self.encoding = encoding
        self.empty_as_null = empty_as_null

        self.dialect = dialect
        self.delimiter = delimiter

        self.skip_rows = skip_rows or 0
        self.fields = fields
        # TODO: use default type converters
        self.type_converters = type_converters or {}

        self.resource = Resource(resource, encoding=self.encoding)
        self.handle = self.resource.open()

        options = dict(options) if options else {}
        if self.dialect:
            if isinstance(self.dialect, str):
                options["dialect"] = csv.get_dialect(self.dialect)
            else:
                options["dialect"] = self.dialect
        if self.delimiter:
            options["delimiter"] = self.delimiter

        # CSV reader options
        self.options = options

        # self.reader = csv.reader(handle, **self.reader_args)
        self.reader = csv.reader(self.handle, **options)


        if self.skip_rows:
            for i in range(0, self.skip_rows):
                next(self.reader)

        # Initialize field list
        if self.read_header:
            field_names = next(self.reader)

            # Fields set explicitly take priority over what is read from the
            # header. (Issue #17 might be somehow related)
            if not self.fields:
                fields = [ (name, "string", "default") for name in field_names]
                self.fields = FieldList(*fields)

        if not self.fields:
            raise RuntimeError("Fields are not initialized. "
                               "Either read fields from CSV header or "
                               "set them manually")

        self.set_fields(self.fields)


    def set_fields(self, fields):
        self.converters = [self.type_converters.get(f.storage_type) for f in fields]

        if not any(self.converters):
            self.converters = None

    def release(self):
        if self.resource:
            self.resource.close()

    def representations(self):
        return ["csv", "rows", "records"]

    def rows(self):
        missing_values = [f.missing_value for f in self.fields]

        for row in self.reader:
            result = []

            for i, value in enumerate(row):
                if self.empty_as_null and not value:
                    result.append(None)
                    continue

                if missing_values[i] and value == missing_values[i]:
                    result.append(None)
                    continue

                func = self.converters[i] if self.converters else None

                if func:
                    result.append(func(value))
                else:
                    result.append(value)
            yield result

    def csv_stream(self):
        return self.handle

    def records(self):
        fields = self.fields.names()
        for row in self.reader:
            yield dict(zip(fields, row))

    def is_consumable(self):
        return True

    def retained(self):
        """Returns retained copy of the consumable"""
        # Default implementation is naive: consumes whole CSV into Python
        # memory
        # TODO: decide whether source is seek-able or not

        return RowListDataObject(list(self.rows()), self.fields)


class CSVTarget(DataObject):
    """Comma separated values text file as a data target."""

    _bubbles_info = {
        "attributes": [
            {
                "name":"resource",
                "description": "Filename or URL"
            },
            {
                "name": "write_headers",
                "description": "Flag whether first row will contain field names"
            },
            {
                "name": "truncate",
                "description": "If `True` (default) then target file is truncated"
            },
            {
                "name": "encoding",
                "description": "file character encoding"
            },
            {
                "name": "fields",
                "description": "data fields"
            }
        ]
    }

    def __init__(self, resource, write_headers=True, truncate=True,
                 encoding="utf-8", dialect=None,fields=None, **kwds):
        """Creates a CSV data target

        :Attributes:
            * resource: target object - might be a filename or file-like
              object
            * write_headers: write field names as headers into output file
            * truncate: remove data from file before writing, default: True

        """
        self.write_headers = write_headers
        self.truncate = truncate
        self.encoding = encoding
        self.dialect = dialect
        self.fields = fields
        self.kwds = kwds

        self.close_file = False
        self.handle = None

        mode = "w" if self.truncate else "a"

        self.handle = open(resource, mode=mode, encoding=encoding)

        self.writer = csv.writer(self.handle, dialect=self.dialect, **self.kwds)

        if self.write_headers:
            if not self.fields:
                raise BubblesError("No fields provided")
            self.writer.writerow(self.fields.names())

        self.field_names = self.fields.names()

    def finalize(self):
        if self.handle:
            self.handle.close()

    def append(self, row):
        self.writer.writerow(row)

    def append_from(self, obj):
        for row in obj:
            self.append(row)



########NEW FILE########
__FILENAME__ = xls
# -*- coding: utf-8 -*-
from ..objects import *
from ..errors import *
from ..common import get_logger
from ..metadata import Field, FieldList
from ..stores import DataStore
from ..resource import Resource
import datetime

try:
    import xlrd
    _cell_types = {
        xlrd.XL_CELL_EMPTY:   'unknown',
        xlrd.XL_CELL_TEXT:    'string',
        xlrd.XL_CELL_NUMBER:  'float',
        xlrd.XL_CELL_DATE:    'date',
        xlrd.XL_CELL_BOOLEAN: 'boolean',
        xlrd.XL_CELL_ERROR:   'unknown',
        xlrd.XL_CELL_BLANK:   'unknown',
    }

except ImportError:
    from ..common import MissingPackage
    xlrd = MissingPackage("xlrd", "Data objects from MS Excel spreadsheets")

def _load_workbook(resource, encoding):
        resource = Resource(resource, binary=True)

        with resource.open() as f:
            data = f.read()

        workbook = xlrd.open_workbook(file_contents=data,
                                      encoding_override=encoding)

        return workbook

class XLSStore(DataStore):
    def __init__(self, resource, encoding=None):
        super(XLSStore, self).__init__()

        self.resource = resource
        self.encoding = encoding
        self.book = None

    def get_object(self, name, skip_rows=0, has_header=True):
        if not self.book:
            self.book = _load_workbook(self.resource, self.encoding)

        return XLSObject(workbook=self.book, sheet=name, skip_rows=skip_rows,
                            has_header=has_header)

    def object_names(self):
        if not self.book:
            self.book = _load_workbook(self.resource, self.encoding)

        return self.book.sheet_names()

    def create(self, name):
        raise BubblesError("XLS store is read-only")


class XLSObject(DataObject):
    __identifier__ = "xls"
    def __init__(self, resource=None, fields=None, sheet=0, encoding=None,
                 skip_rows=0, has_header=True, workbook=None):
        """Creates a XLS spreadsheet data source stream.

        Attributes:

        * resource: file name, URL or file-like object
        * sheet: sheet index number (as int) or sheet name (as str)
        * read_header: flag determining whether first line contains header or
          not. ``True`` by default.
        """
        if workbook:
            if resource:
                raise ArgumentError("Can not specify both resource and "
                                    "workbook")

            self.workbook = workbook
        else:
            self.workbook = _load_workbook(resource, encoding)

        if isinstance(sheet, int):
            self.sheet = self.workbook.sheet_by_index(sheet)
        else:
            self.sheet = self.workbook.sheet_by_name(sheet)

        if skip_rows >= self.sheet.nrows:
            raise ArgumentError("First row number is larger than number of rows")

        if has_header:
            self.first_row = skip_rows + 1
        else:
            self.first_row  = skip_rows

        if fields:
            self.fields = fields
        else:
            # Read fields
            if has_header:
                names = self.sheet.row_values(skip_rows)

            self.fields = FieldList()

            row = self.sheet.row(self.first_row)
            for name, cell in zip(names, row):
                storage_type = _cell_types.get(cell.ctype, "unknown")
                field = Field(name, storage_type=storage_type)
                self.fields.append(field)

    def representations(self):
        return ["rows", "records"]

    def __len__(self):
        return self.sheet.nrows - self.first_row

    def rows(self):
        if not self.fields:
            raise RuntimeError("Fields are not initialized")
        return XLSRowIterator(self.workbook,
                              self.sheet,
                              self.first_row,
                              self.fields)

    def records(self):
        fields = self.fields.names()
        for row in self.rows():
            yield dict(zip(fields, row))

    def is_consumable(self):
        return False

class XLSRowIterator(object):
    """
    Iterator that reads XLS spreadsheet
    """
    def __init__(self, workbook, sheet, first_row=0, fields=None):
        self.workbook = workbook
        self.sheet = sheet
        self.row_count = sheet.nrows
        self.current_row = first_row
        if fields:
            self.field_count = len(fields)
        else:
            self.field_count = None

    def __iter__(self):
        return self

    def __next__(self):
        if self.current_row >= self.row_count:
            raise StopIteration

        row = self.sheet.row(self.current_row)
        if self.field_count is not None:
            row = row[:self.field_count]
        row = tuple(self._cell_value(cell) for cell in row)
        self.current_row += 1
        return row

    def _cell_value(self, cell):
        if cell.ctype == xlrd.XL_CELL_NUMBER:
            return float(cell.value)
        elif cell.ctype == xlrd.XL_CELL_DATE:
            # TODO: distinguish date and datetime
            args = xlrd.xldate_as_tuple(cell.value, self.workbook.datemode)
            return datetime.date(args[0], args[1], args[2])
        elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return bool(cell.value)
        else:
            return cell.value


########NEW FILE########
__FILENAME__ = xlsx
from itertools import dropwhile

from bubbles.errors import ArgumentError, BubblesError
from bubbles.metadata import FieldList, Field
from bubbles.objects import DataObject
from bubbles.stores import DataStore

try:
    import openpyxl
except ImportError:
    from ..common import MissingPackage
    openpyxl = MissingPackage("openpyxl", "Data objects from MS Excel spreadsheets")


CELL_TYPES = {openpyxl.cell.Cell.TYPE_STRING: 'string',
              openpyxl.cell.Cell.TYPE_FORMULA: 'unknown',
              openpyxl.cell.Cell.TYPE_NUMERIC: 'float',
              openpyxl.cell.Cell.TYPE_BOOL: 'boolean',
              openpyxl.cell.Cell.TYPE_NULL: 'unknown',
              openpyxl.cell.Cell.TYPE_INLINE: 'string',
              openpyxl.cell.Cell.TYPE_ERROR: 'unknown',
              openpyxl.cell.Cell.TYPE_FORMULA_CACHE_STRING: 'string'}


def _load_workbook(resource):
    return openpyxl.load_workbook(resource, use_iterators=True,
                                  keep_vba=False, data_only=True)


class XLSXStore(DataStore):
    def __init__(self, resource, encoding=None):
        super(XLSXStore, self).__init__()

        self.resource = resource
        self.encoding = encoding
        self._book = None

    @property
    def book(self):
        if self._book is None:
            self._book = _load_workbook(self.resource)
        return self._book

    def get_object(self, name, skip_rows=0, has_header=True):
        return XLSXObject(self.book, sheet=name, encoding=self.encoding,
                          skip_rows=skip_rows, has_header=has_header)

    def object_names(self):
        return self.book.get_sheet_names()

    def create(self, name):
        raise BubblesError("XLSX store is read-only")


class XLSXObject(DataObject):
    __identifier__ = "xlsx"

    def __init__(self, resource=None, fields=None, sheet=0,
                 encoding=None, skip_rows=0, has_header=True):
        """Creates a XLSX spreadsheet data source stream.

        Attributes:

        * resource: file name, URL or file-like object
        * sheet: sheet index number (as int) or sheet name
        * has_header: flag determining whether first line contains header or
          not. ``True`` by default.
        """
        if isinstance(resource, openpyxl.Workbook):
            self.workbook = resource
        else:
            self.workbook = _load_workbook(resource)

        if isinstance(sheet, int):
            self.sheet = self.workbook.worksheets[sheet]
        elif isinstance(sheet, str):
            self.sheet = self.workbook[sheet]
        else:
            raise ArgumentError('sheet has to be a string or an integer')

        if has_header:
            self.first_row = skip_rows + 1
        else:
            self.first_row = skip_rows

        if fields:
            self.fields = fields
        else:
            rows = enumerate(self.sheet.rows)
            first_row = next(dropwhile(lambda x: x[0] < self.first_row,
                                       rows))[1]
            if has_header:
                header_row = next(self.sheet.rows)
                names = [str(c.value) for c in header_row]
            else:
                names = ['col%d' % i for i in range(len(first_row))]

            self.fields = FieldList()
            for name, cell in zip(names, first_row):
                if cell.is_date:
                    storage_type = 'date'
                else:
                    storage_type = CELL_TYPES.get(cell.data_type, 'unknown')
                field = Field(name, storage_type=storage_type)
                self.fields.append(field)

    def _nrows(self):
        return sum([1 for _ in self.sheet.rows])

    def representations(self):
        return ["rows", "records"]

    def __len__(self):
        return self._nrows() - self.first_row

    def rows(self):
        if not self.fields:
            raise RuntimeError("Fields are not initialized")

        field_count = len(self.fields)
        rows = enumerate(self.sheet.rows)
        rows = dropwhile(lambda x: x[0] < self.first_row, rows)
        for _, row in rows:
            yield tuple([c.value for c in row[:field_count]])

    def records(self):
        fields = self.fields.names()
        for row in self.rows():
            yield dict(zip(fields, row))

    def is_consumable(self):
        return False

########NEW FILE########
__FILENAME__ = common
# -*- coding: utf-8 -*-
# Common bubbles functions and classes - related to data processing or process
# management
#
# For language utility functions see module util

import re
import sys
import logging

from collections import OrderedDict

__all__ = [
    "logger_name",
    "get_logger",
    "create_logger",

    "MissingPackage",

    "decamelize",
    "to_identifier",

    "IgnoringDictionary"
]

logger_name = "bubbles"
logger = None

def get_logger():
    """Get bubbles default logger"""
    global logger

    if logger:
        return logger
    else:
        return create_logger()

def create_logger():
    """Create a default logger"""
    global logger
    logger = logging.getLogger(logger_name)

    formatter = logging.Formatter(fmt='%(asctime)s %(levelname)s %(message)s')

    handler = logging.StreamHandler()
    handler.setFormatter(formatter)
    logger.addHandler(handler)

    return logger


class IgnoringDictionary(OrderedDict):
    """Simple dictionary extension that will ignore any keys of which values
    are empty (None/False)"""
    def __setitem__(self, key, value):
        if value is not None:
            super(IgnoringDictionary, self).__setitem__(key, value)

    def set(self, key, value):
        """Sets `value` for `key` even if value is null."""
        super(IgnoringDictionary, self).__setitem__(key, value)

    def __repr__(self):
        items = []
        for key, value in self.items():
            item = '%s: %s' % (repr(key), repr(value))
            items.append(item)

        return "{%s}" % ", ".join(items)


class MissingPackageError(Exception):
    """Exception raised when encountered a missing package."""
    pass

class MissingPackage(object):
    """Bogus class to handle missing optional packages - packages that are not
    necessarily required for Cubes, but are needed for certain features."""

    def __init__(self, package, feature = None, source = None, comment = None):
        self.package = package
        self.feature = feature
        self.source = source
        self.comment = comment

    def __call__(self, *args, **kwargs):
        self._fail()

    def __getattr__(self, name):
        self._fail()

    def _fail(self):
        if self.feature:
            use = " to be able to use: %s" % self.feature
        else:
            use = ""

        if self.source:
            source = " from %s" % self.source
        else:
            source = ""

        if self.comment:
            comment = ". %s" % self.comment
        else:
            comment = ""

        raise MissingPackageError("Optional package '%s' is not installed. "
                                  "Please install the package%s%s%s" %
                                      (self.package, source, use, comment))

def decamelize(name):
    s1 = re.sub('(.)([A-Z][a-z]+)', r'\1 \2', name)
    return re.sub('([a-z0-9])([A-Z])', r'\1 \2', s1)

def to_identifier(name):
    return re.sub(r' ', r'_', name).lower()




########NEW FILE########
__FILENAME__ = datapackage
# -*- Encoding: utf8 -*-

from .stores import DataStore
from .resource import is_local, read_json
from .objects import data_object
from .errors import *
from .metadata import Field, FieldList, DEFAULT_ANALYTICAL_TYPES
import os
import json
from collections import OrderedDict

from urllib.parse import urljoin, urlparse


OBJECT_TYPES = {
    "csv": "csv_source"
}

# Conversion from JSON Table Schema
# Source: http://dataprotocols.org/json-table-schema/

def schema_to_fields(fields):
    """Convert simple data set field list into bubbles field list."""
    flist = []
    for i, md in enumerate(fields):
        if not "name" in md and not "id" in md:
            raise MetadataError("Field #%d has no name" % i)

        name = md.get("name")
        if not name:
            name = md["id"]

        storage_type = md.get("type", "unknown")
        if storage_type == "any":
            storage_type = "unknown"

        atype = DEFAULT_ANALYTICAL_TYPES.get(storage_type, "typeless")

        field = Field(name,
                      storage_type=storage_type,
                      analytical_type=atype,
                      label=md.get("title"),
                      description=md.get("description"),
                      info=md.get("info"),
                      size=md.get("size"),
                      missing_value=md.get("missing_value"))
        flist.append(field)

    return FieldList(*flist)

class DataPackageResource(object):
    def __init__(self, package, resource):
        self.package = package

        path = resource.get("path")
        if path:
            self.url = urljoin(package.url, path)
        elif "url" in resource:
            self.url = resource["url"]
        elif "data" in resource:
            raise NotImplementedError("Embedded datapackage resource data "
                                      "are not supported")
        else:
            raise MetadataError("No path or url specified in a package "
                                "resource.")

        self.name = resource.get("name")
        self.title = resource.get("title")
        self.url = urljoin(package.url, path)

        schema = resource.get("schema")
        if schema:
            fields = schema.get("fields")
            self.fields = schema_to_fields(fields)
        else:
            self.fields = None

        self.type = resource.get("type", os.path.splitext(self.url)[1][1:])

        if self.type not in OBJECT_TYPES:
            raise TypeError("Data object type '%s' is not supported in "
                            "datapackage." % self.type)

    def dataobject(self):
        return data_object(OBJECT_TYPES[self.type], self.url,
                           fields=self.fields)

class DataPackage(object):
    def __init__(self, url):
        # TODO: currently only local paths are supported
        if is_local(url) and not url.endswith("/"):
            url = url + "/"

        self.url = url

        infopath = urljoin(url, "datapackage.json")
        metadata = read_json(infopath)
        with open(infopath) as f:
            try:
                metadata = json.load(f)
            except Exception as e:
                raise Exception("Unable to read %s: %s"
                                % (infopath, str(e)))

        self.name = metadata.get("name")
        self._resources = OrderedDict()
        for i, res in enumerate(metadata["resources"]):
            resource = DataPackageResource(self, res)
            if not resource.name:
                resource.name = "resource%d" % i

            if resource.name in self._resources:
                raise Exception("Duplicate resource '%s' in data package '%s'"
                                % (resource.name, self.name))
            self._resources[resource.name] = resource

    def __getitem__(self, item):
        return self._resources[item]

    def resource(self, name):
        return self._resources[name]

    @property
    def resources(self):
        return list(self._resources.values())

    @property
    def resource_count(self):
        return len(self._resources)

class DataPackageCollectionStore(DataStore):
    __extension_name__ = "datapackages"

    def __init__(self, url):
        """Creates a store that contains collection of data packages. The
        datasets are referred as `store_name.dataset_name`."""

        if not is_local(url):
            raise NotImplementedError("Remote package collections are "
                                      "not supported yet")

        self.packages = []

        paths = []
        for path in os.listdir(url):
            path = os.path.join(url, path)
            if not os.path.isdir(path):
                continue

            if not os.path.exists(os.path.join(path, "datapackage.json")):
                continue

            package = DataPackage(path)
            self.packages.append(package)

        self.datasets = OrderedDict()
        self._index_datasets()

    def _index_datasets(self):
        """Collect dataset names."""

        self.resources = OrderedDict()
        for package in self.packages:
            if package.resource_count == 1:
                name = package.name
                if name in self.resources:
                    raise MetadataError("Two single-resource packages with the"
                                        " same name '%s'" % name)
                self.resources[name] = package.resources[0]
            else:
                for resource in package.resources:
                    name = "%s.%s" % (package.name, resource.name)
                    if name in self.resources:
                        raise MetadataError("Duplicate datapackage resource %s in "
                                            "%s." % (resource.name, package.name) )
                    self.resources[name] = resource

    def object_names(self):
        return self.resources.keys()

    def get_object(self, name, **args):
        try:
            return self.resources[name].dataobject()
        except KeyError:
            raise NoSuchObjectError(name)


########NEW FILE########
__FILENAME__ = datautil
# -*- Encoding: utf8 -*-
"""Various utility functions"""

import datetime

__all__ = (
        "expand_record",
        "collapse_record",
        "guess_type",
        "to_bool"
        )


def guess_type(string, date_format="%Y-%m-%dT%H:%M:%S.Z"):
    """Guess one of basic types that the `string` might contain. Returns a
    string with basic type name. If `date_format` is ``None`` then string is
    not tested for date type. Default is ISO date format."""

    if string is None:
        return None

    try:
        int(string)
        return "integer"
    except ValueError:
        pass

    try:
        float(string)
        return "float"
    except ValueError:
        pass

    if date_format:
        try:
            datetime.datetime.strptime(string, date_format)
            return "date"
        except ValueError:
            pass

    return "string"


def expand_record(record, separator = '.'):
    """Expand record represented as dict object by treating keys as key paths separated by
    `separator`, which is by default ``.``. For example: ``{ "product.code": 10 }`` will become
    ``{ "product" = { "code": 10 } }``

    See :func:`bubbles.collapse_record` for reverse operation.
    """
    result = {}
    for key, value in list(record.items()):
        current = result
        path = key.split(separator)
        for part in path[:-1]:
            if part not in current:
                current[part] = {}
            current = current[part]
        current[path[-1]] = value
    return result


def collapse_record(record, separator = '.', root = None):
    """See :func:`bubbles.expand_record` for reverse operation.
    """

    result = {}
    for key, value in list(record.items()):
        if root:
            collapsed_key = root + separator + key
        else:
            collapsed_key = key

        if type(value) == dict:
            collapsed = collapse_record(value, separator, collapsed_key)
            result.update(collapsed)
        else:
            result[collapsed_key] = value
    return

def to_bool(value):
    """Return boolean value. Convert string to True when "true", "yes" or "on"
    """
    if isinstance(value, str):
        value = value.lower()
        return value in ["1", "true", "yes", "on"] and value != "0"
    else:
        return bool(value)



########NEW FILE########
__FILENAME__ = dev
# -*- Encoding: utf8 -*-

"""Development utilities - functions used for development and
documentation of bubbles functionalities and objcects."""

def experimental(fn):
    """Mark a method as experimental. Interface or implementation of an
    experimental method might very likely change. Use with caution. This
    decorator just appends a doc string."""

    warning = \
    """
    .. warning::

       This method is experimental. Interface or implementation might
       change. Use with caution and note that you might have to modify
       your code later.
    """

    if fn.__doc__ is not None:
        fn.__doc__ += warning
    fn._bubbles_experimental = True

    return fn

def is_experimental(fn):
    """Returns `True` if function `fn` is experimental."""

    if hasattr(fn, "_bubbles_experimental"):
        return fn._bubbles_experimental
    else:
        return False

def required(fn):
    """Mark method as required to be implemented by scubclasses"""
    fn._bubbles_required = True
    return fn

def recommended(fn):
    """Mark method as recommended to be implemented by subclasses"""
    fn._bubbles_recommended = True
    return fn


########NEW FILE########
__FILENAME__ = errors
# -*- coding: utf-8 -*-
"""Bubbles Exceptions"""

class BubblesError(Exception):
	"""Basic error class"""
	pass
class UserError(BubblesError):
    """Superclass for all errors caused by the bubbles users. Error
    messages from this error might be safely passed to the front-end. Do not
    include any information that you would not like to be public"""
    error_type = "unknown_user_error"

class InternalError(BubblesError):
    """Superclass for all errors that happened internally: configuration
    issues, connection problems, model inconsistencies..."""

class ConfigurationError(InternalError):
    """Raised when there is a problem with workspace configuration assumed."""

class MetadataError(BubblesError):
	"""Error raised on metadata inconsistency"""
	pass

class NoSuchFieldError(MetadataError):
	"""Error raised on metadata inconsistency"""
	pass

class FieldOriginError(MetadataError):
    """Error with field origin, such as circular reference."""
    pass

class ArgumentError(BubblesError):
    """Raised when wrong argument is passed to a function"""
    pass

class ProbeAssertionError(BubblesError):
    """Raised when probe assertion fails"""
    def __init__(self, reason=None):
        self.reason = reason

class ConsumedError(BubblesError):
    """Raised when trying to read from already consumed object"""
    pass
#
# DataObject and DataStore errors
#

class DataObjectError(BubblesError):
    """Generic error in a data object."""

class NoSuchObjectError(DataObjectError):
    """Raised when object does not exist."""
    pass

class ObjectExistsError(DataObjectError):
    """Raised when attempting to create and object that already exists"""
    pass

class IsNotTargetError(DataObjectError):
    """Raised when trying to use data object as target - appending data,
    truncating or any other target-only operation."""
    pass

class IsNotSourceError(DataObjectError):
    """Raised when trying to use data object as source, for example reading
    data"""
    pass

class RepresentationError(DataObjectError):
    """Raised when requested unknown, invalid or not available
    representation"""
    pass

#
# Operations errors
#

class OperationError(BubblesError):
    """Raised when operation is not found or is mismatched"""
    pass

class RetryOperation(Exception):
    """Raised within operation to signal to the kernel that it should try
    another operation with different signature."""
    def __init__(self, signature=None, reason=None):
        self.signature = signature
        self.reason = reason

class RetryError(BubblesError):
    """Raised when operation was retried too many times"""
    pass

class GraphError(BubblesError):
    pass

class FieldError(BubblesError):
    """Raised when wrong field types are passed to an operation."""
    pass

########NEW FILE########
__FILENAME__ = context
# -*- coding: utf-8 -*-
import itertools
from collections import defaultdict
from ..errors import *
from ..dev import is_experimental
from ..operation import Operation, Signature, get_representations
from ..common import get_logger
from ..threadlocal import LocalProxy

__all__ = (
            "OperationContext",
            "default_context",
            "LoggingContextObserver",
            "CollectingContextObserver",
        )

"""List of modules with operations to be loaded for the default contex."""

# FIXME: make this more intelligent and lazy
_default_op_modules = (
            "bubbles.backends.sql.ops",
            "bubbles.backends.mongo.ops",
            "bubbles.ops.rows",
            "bubbles.ops.generic",
        )


class _OperationReference(object):
    def __init__(self, context, name):
        """Creates a reference to an operation within a context"""
        self.context = context
        self.name = name

    def __call__(self, *args, **kwargs):
        return self.context.call(self.name, *args, **kwargs)


class OperationContext(object):
    # TODO: add parent_context
    def __init__(self, retry_count=10):
        """Creates an operation context. Default `retry_count` is 10.

        Use:

        .. code-block:: python

            c = Context()
            duplicate_records = c.op.duplicates(table)

        Context uses multiple dispatch based on operation arguments. The
        `k.duplicates()` operation might be different, based on available
        representations of `table` object.
        """

        super().__init__()
        self.operations = {}

        self.op = _OperationGetter(self)

        self.logger = get_logger()
        self.observer = LoggingContextObserver(self.logger)
        self.retry_count = retry_count

        self.retry_allow = []
        self.retry_deny = []

    def operation(self, name):
        """Get operation by `name`. If operatin does not exist, then
        `operation_not_found()` is called and the lookup is retried."""

        try:
            return self.operations[name]
        except KeyError:
            return self.operation_not_found(name)

    def add_operations_from(self, obj):
        """Import operations from `obj`. All attributes of `obj` are
        inspected. If they are instances of Operation, then the operation is
        imported."""

        for name in dir(obj):
            op = getattr(obj, name)
            if isinstance(op, Operation):
                self.add_operation(op)

    def add_operation(self, op):
        """Registers a decorated operation.  operation is considered to be a
        context's method and would receive the context as first argument."""

        self.operations[op.name] = op

    def remove_operation(self, name):
        """Removes all operations with `name` and `signature`. If no
        `signature` is specified, then all operations with given name are
        removed, regardles of the signature."""

        del self.operations[name]

    def operation_not_found(self, name):
        """Subclasses might override this method to load necessary modules and
        register operations using `register_operation()`. Default
        implementation raises an exception."""
        raise OperationError("Operation '%s' not found" % name)

    def get_operation(self, name, signature):
        """Returns an operation with given name and signature. Requires exact
        signature match."""
        # Get all signatures registered for the operation
        operations = self.operation_list(name)

        for op in operations:
            if op.signature == signature:
                return op
        raise OperationError("No operation '%s' with signature '%s'" %
                                                        (name, signature))
    def debug_print_catalogue(self):
        """Pretty prints the operation catalogue – all operations, their
        signatures and function references. Used for debugging purposes."""

        print("== OPERATIONS ==")
        keys = list(self.operations.keys())
        keys.sort()
        for name in keys:
            ops = self.operations[name]
            print("* %s:" % name)

            for op in ops:
                print("    - %s:%s" % \
                            (op.signature.signature, op.function) )

    def can_retry(self, opname):
        """Returns `True` if an operation can be retried. By default, all
        operations can be retried. `Context.retry_allow` contains list of
        allowed operations, `Context.retry_deny` contains list of denied
        operations. When the allow list is set, then only operations from the
        list are allowed. When the deny list is set, the operations in the
        deny list are not allowed."""

        if self.retry_deny and opname in self.retry_deny:
            return False

        if self.retry_allow and opname not in self.retry_allow:
            return False

        return True

    def call(self, op_name, *args, **kwargs):
        """Dispatch and call operation with `name`. Arguments are passed to the
        operation, If the operation raises `RetryOperation` then another
        function with signature from the exception is tried. If no signature
        is provided in the exception, then next matching signature is used."""

        op = self.operation(op_name)
        operands = args[:op.opcount]

        reps = get_representations(*operands)
        resolution_order = op.resolution_order(reps)
        first_signature = resolution_order[0]

        self.logger.debug("op %s(%s)" % (op_name, reps))

        if self.observer:
            self.observer.will_call_operation(self, op, first_signature)

        result = None

        # We try to perform requested operation. If the operation raises
        # RetryOperation exception, then we use signature from the exception
        # for another operation. Operation can be retried retry_count number
        # of times.
        # Observer is notified about each retry.

        visited = set()

        while resolution_order:
            sig = resolution_order.pop(0)
            visited.add(sig)

            try:
                function = op.function(sig)
            except KeyError:
                raise OperationError("No signature (%s) in operation %s"
                                     % (sig, op_name))

            try:
                if op.experimental:
                    self.logger.warn("operation %s is experimental" % \
                                        op_name)
                result = function(self, *args, **kwargs)

            except RetryOperation as e:
                if not self.can_retry(op_name):
                    raise RetryError("Retry of operation '%s' is not allowed")

                retry = e.signature

                if retry:
                    retry = Signature(*retry)

                    if retry in visited:
                        raise RetryError("Asked to retry operation %s with "
                                         "signature %s, but it was already "
                                         "visited."
                                         % (op_name, retry))

                    resolution_order.insert(0, retry)

                elif retry:
                    retry = resolution_order[0]

                if self.observer:
                    self.observer.will_retry_operation(self, op, retry,
                                                       first_signature,
                                                       str(e))
            else:
                # Let the observer know which operation was called at last and
                # completed sucessfully
                if self.observer:
                    self.observer.did_call_operation(self, op, sig,
                                                     first_signature)
                return result

        raise RetryError("No remaining signature to rerty when calling "
                         "operation %s with %s"
                         % (op_name, first_signature))


class LoggingContextObserver(object):
    def __init__(self, logger=None):
        self.logger = logger

    def will_call_operation(self, ctx, op, signature):
        logger = self.logger or ctx.logger
        logger.info("calling %s(%s)" % (op, signature))

    def did_call_operation(self, ctx, op, signature, first):
        logger = self.logger or ctx.logger

        if first == signature:
            logger.debug("called %s(%s)" % (op, signature))
        else:
            logger.debug("called %s(%s) as %s" % \
                                (op, signature, first))

    def will_retry_operation(self, ctx, op, signature, first, reason):
        logger = self.logger or ctx.logger
        logger.info("retry %s(%s) as %s, reason: %s" %
                                        (op, first, signature, reason))

class CollectingContextObserver(object):
    def __init__(self):
        self.history = []
        self.current = None
        self.tried = []

    def will_call_operation(self, ctx, op):
        self.current = op
        self.tried = []

    def will_retry_operation(self, ctx, op, reason):
        self.tried.append( (self.current, reason) )
        self.current = op

    def did_call_operation(self, ctx, op, retries):
        line = (op, retries, self.tried)
        self.history.append(line)


class BoundOperation(object):
    def __init__(self, context, opname):
        self.context = context
        self.opname = opname

    def __call__(self, *args, **kwargs):
        return self.context.call(self.opname, *args, **kwargs)


class _OperationGetter(object):
    def __init__(self, context):
        self.context = context

    def __getattr__(self, name):
        return BoundOperation(self.context, name)

    def __getitem__(self, name):
        return BoundOperation(self.context, name)


def create_default_context():
    """Creates a ExecutionContext with default operations."""
    context = OperationContext()

    for modname in _default_op_modules:
        mod = _load_module(modname)
        context.add_operations_from(mod)

    return context

def _load_module(modulepath):
    mod = __import__(modulepath)
    path = []
    for token in modulepath.split(".")[1:]:
       path.append(token)
       try:
           mod = getattr(mod, token)
       except AttributeError:
           raise BubblesError("Unable to get %s" % (path, ))
    return mod


default_context = LocalProxy("default_context",
                             factory=create_default_context)


########NEW FILE########
__FILENAME__ = engine
# -*- coding: utf-8 -*-
from collections import namedtuple, Counter
from ..errors import *

__all__ = (
    "ExecutionEngine",

    # TODO: Not quite public yet, but we export it nonetheless
    "ExecutionStep",
    "ExecutionPlan"
)


class ExecutionStep(object):
    def __init__(self, node, outlets=None, result=None):
        self.node = node
        self.outlets = outlets or []
        self.result = result

    def evaluate(self, engine, context, operands):
        """Evaluates the wrapped node within `context` and with `operands`.
        Stores the evaluation result."""

        self.result = self.node.evaluate(engine, context, operands)
        return self.result

    def __str__(self):
        return "evaluate %s" % str(self.node)


# Execution Engine
# ================
#
# Main graph execution class. Current execution method is simple:
# 1. topologically sort nodes
# 2. prepare execution node for each node and connects them by outlets
# 3. execute in the topological order and set result to the execution node
#    – result is used as input for other nodes

# TODO: allow use of lists of objects, such as rows[] or sql[]. Currently
# there is no way how to specify this kind of connections in the graph.

ExecutionPlan = namedtuple("ExecutionPlan", ["steps", "consumption"])

# Execution Engine
#
# This is the core graph execution class. In the future it will contain more
# helper methods for node execution.
#
# Sublcasses should implement the prepare_execution_plan method that returns a
# tuple ExecutionPlan(steps, consumption) where steps is a list of
# ExecutionSteps and consumption is a dictionary mapping a node to number of
# times the node output is consumed within the graph.

class ExecutionEngine(object):

    def __init__(self, context, stores=None):
        """Creates an instance of execution engine within an execution
        `context`.

        `stores` is a mapping of store names and opened data stores. Stores
        are used when resoving data sources by reference.

        Execution engine is also used in :class:`Pipeline` objects to run the
        pipelines.
        """

        self.stores = stores or {}
        self.context = context
        self.logger = context.logger

    def execution_plan(self, graph):
        """Returns a list of topologically sorted `ExecutionSteps`, ready to
        be used for execution.

        If node is an operation node, it will contain references to nodes
        holding results that will be passed as operands to the operation.

        """

        # TODO: this method will be customizable in subclasses in the future

        # Operation -> Node -> Execution Step
        # Node is an operation with parameters set (configured operation)
        # Execution Node is Node in execution context with bound outlets

        sorted_nodes = graph.sorted_nodes()

        node_steps = {}
        steps = []

        # Count consumption of node's output and add consumption hints to the
        # execution plan. ExecutionEngine should handle (or refuse to handle)
        # multiple consumptions of consumable objects
        consumption = Counter()

        for node in sorted_nodes:
            sources = graph.sources(node)
            outlets = node.outlets(self.context)

            outlet_nodes = []
            for i, outlet in enumerate(outlets):
                if i == 0:
                    outlet_node = sources.get(outlet) or sources.get("default")
                else:
                    outlet_node = sources.get(outlet)

                if not outlet_node:
                    raise BubblesError("Outlet '%s' is not connected" %
                            outlet)

                # Get execution node wrapper for the outlet node
                outlet_nodes.append(node_steps[outlet_node])

                # Count the consumption (see note before the outer loop)
                consumption[outlet_node] += 1

            step = ExecutionStep(node, outlets=outlet_nodes)

            node_steps[node] = step
            steps.append(step)

        plan = ExecutionPlan(steps, consumption)

        return plan

    def run(self, graph):
        """Runs the `graph` nodes. First an execution plan is prepared, then
        the nodes are executed according to the plan. See
        :meth:`ExecutionEngine.prepare_execution_plan` for more information.
        """

        # TODO: write documentation about consumable objects

        plan = self.execution_plan(graph)

        # Set of already consumed nodes
        consumed = set()

        for i, step in enumerate(plan.steps):
            self.logger.debug("step %s: %s" % (i, str(step)))

            operands = []

            for outlet in step.outlets:

                # Check how many times the outlet node that is about to be
                # used is going to be consumed. If it is consumable and will
                # be consumed more than once, then a retained version of the
                # object is created. Retention policy is defined by the
                # backend. In most of the cases it is just python list wrapper
                # over consumed iterator of rows, which might be quite costly.

                consume_times = plan.consumption[outlet.node]
                if outlet.result.is_consumable() and consume_times > 1:
                    if outlet.node not in consumed:
                        self.logger.debug("retaining consumable %s. it will "
                                          "be consumed %s times" % \
                                                 (outlet.node, consume_times))
                        outlet.result = outlet.result.retained()

                consumed.add(outlet.node)
                operands.append(outlet.result)

            step.evaluate(self, self.context, operands)

########NEW FILE########
__FILENAME__ = graph
from collections import OrderedDict, namedtuple, Counter
from ..objects import data_object
from ..common import get_logger
from ..errors import *

__all__ = (
    "Graph",
    "Node",
    "Connection",

    # Not quite public
    "Node",
    "StoreObjectNode",
    "ObjectNode",
    "CreateObjectNode",
    "ObjectFactoryNode"
)

class NodeBase(object):
    def outlets(self, context):
        """Default node has no outlets."""
        return []

class Node(NodeBase):
    def __init__(self, opname, *args, **kwargs):
        """Creates a `Node` with operation `op` and operation `options`"""

        self.opname = opname
        self.args = args
        self.kwargs = kwargs

    def is_source(self):
        return False

    def evaluate(self, engine, context, operands=None):
        """Evaluates the operation with name `opname` within `context`"""
        # fixme: identify operands in *args
        args = list(operands) + list(self.args)
        result = context.call(self.opname, *args, **self.kwargs)
        return result

    def __str__(self):
        return "operation %s" % self.opname

    def outlets(self, context):
        prototype = context.operation(self.opname)
        return prototype.operands


class ObjectFactoryNode(NodeBase):
    def __init__(self, factory, *args, **kwargs):

        self.factory = factory
        self.args = args
        self.kwargs = kwargs

    def is_source(self):
        return True

    def evaluate(self, engine, context, operands=None):
        return data_object(self.factory, *self.args, **self.kwargs)

    def __str__(self):
        return "factory source %s" % self.factory

class StoreObjectNode(NodeBase):
    def __init__(self, store, objname, **parameters):
        self.store = store
        self.objname = objname
        self.parameters = parameters

    def is_source(self):
        return True

    def evaluate(self, engine, context, operands=None):
        """Looks up the object `objname` in `store` from `engine`."""

        try:
            store = engine.stores[self.store]
        except KeyError:
            raise ArgumentError("Unknown store '%s'" % self.store)

        return store.get_object(self.objname, **self.parameters)


    def __str__(self):
        return "source %s in %s" % (self.objname, self.store)

class ObjectNode(NodeBase):
    def __init__(self, obj):
        self.obj = obj

    def is_source(self):
        return True

    def evaluate(self, engine, context, operands=None):
        """Returns the contained object."""
        return self.obj

    def __str__(self):
        return "object %s" % (self.obj, )


class CreateObjectNode(NodeBase):
    def __init__(self, store, name, *args, **kwargs):
        self.store = store
        self.name = name
        self.args = args
        self.kwargs = kwargs

    def is_source(self):
        return False

    def evaluate(self, engine, context, operands=None):
        if len(operands) != 1:
            raise ArgumentError("Number of operands for 'create object' should be 1")

        source = operands[0]

        try:
            store = engine.stores[self.store]
        except KeyError:
            raise ArgumentError("Unknown store %s" % self.store)

        target = store.create(self.name, source.fields,
                              *self.args, **self.kwargs)
        target.append_from(source)

        return target

    def outlets(self, context):
        """`Create` node has one outlet for an object that will be used to
        fill the created object's content."""
        return ["default"]

    def __str__(self):
        return("create %s in %s" % (self.name, self.store))


Connection = namedtuple("Connection", ["source", "target", "outlet"])


class Graph(object):
    """Data processing graph.

    .. note:
            Modifications are not thread safe – as intended.
    """

    def __init__(self, nodes=None, connections=None):
        """Creates a node graph with connections.

        :Parameters:
            * `nodes` - dictionary with keys as node names and values as nodes
            * `connections` - list of two-item tuples. Each tuple contains source and target node
              or source and target node name.
        """

        super(Graph, self).__init__()
        self.nodes = OrderedDict()
        self.connections = set()

        self.logger = get_logger()

        self._name_sequence = 1

        if nodes:
            try:
                for name, node in nodes.items():
                    self.add(node, name)
            except:
                raise ValueError("Nodes should be a dictionary, is %s" % type(nodes))

        if connections:
            for connection in connections:
                self.connect(*connectio)

    def _generate_node_name(self):
        """Generates unique name for a node"""
        while 1:
            name = "node" + str(self._name_sequence)
            if name not in self.nodes.keys():
                break
            self._name_sequence += 1

        return name

    def add(self, node, name=None):
        """Add a `node` into the stream. Does not allow to add named node if
        node with given name already exists. Generate node name if not
        provided. Node name is generated as ``node`` + sequence number.
        Uniqueness is tested."""

        name = name or self._generate_node_name()

        if name in self.nodes:
            raise KeyError("Node with name %s already exists" % name)

        self.nodes[name] = node

        return name

    def node_name(self, node):
        """Returns name of `node`."""
        # There should not be more
        if not node:
            raise ValueError("No node provided")

        names = [key for key,value in self.nodes.items() if value==node]

        if len(names) == 1:
            return names[0]
        elif len(names) > 1:
            raise Exception("There are more references to the same node")
        else: # if len(names) == 0
            raise Exception("Can not find node '%s'" % node)

    def rename_node(self, node, name):
        """Sets a name for `node`. Raises an exception if the `node` is not
        part of the stream, if `name` is empty or there is already node with
        the same name. """

        if not name:
            raise ValueError("No node name provided for rename")
        if name in self.nodes():
            raise ValueError("Node with name '%s' already exists" % name)

        old_name = self.node_name(node)

        del self.nodes[old_name]
        self.nodes[name] = node

    def node(self, node):
        """Coalesce node reference: `reference` should be either a node name
        or a node. Returns the node object."""

        if isinstance(node, str):
            return self.nodes[node]
        elif node in self.nodes.values():
            return node
        else:
            raise ValueError("Unable to find node '%s'" % node)

    def remove(self, node):
        """Remove a `node` from the stream. Also all connections will be
        removed."""

        # Allow node name, get the real node object
        if isinstance(node, basestring):
            name = node
            node = self.nodes[name]
        else:
            name = self.node_name(node)

        del self.nodes[name]

        remove = [c for c in self.connections if c[0] == node or c[1] == node]

        for connection in remove:
            self.connections.remove(connection)

    def connect(self, source, target, outlet="default"):
        """Connects source node and target node. Nodes can be provided as
        objects or names."""
        # Get real nodes if names are provided
        source = self.node(source)
        target = self.node(target)

        sources = self.sources(target)
        if outlet in sources:
            raise GraphError("Target has already connection for outlet '%s'" % \
                                outlet)
        connection = Connection(source, target, outlet)
        self.connections.add(connection)

    def remove_connection(self, source, target):
        """Remove connection between source and target nodes, if exists."""

        connection = (self.coalesce_node(source), self.coalesce_node(target))
        self.connections.discard(connection)

    def sorted_nodes(self):
        """
        Returns topologically sorted nodes.

        Algorithm::

            L = Empty list that will contain the sorted elements
            S = Set of all nodes with no incoming edges
            while S is non-empty do
                remove a node n from S
                insert n into L
                for each node m with an edge e from n to m do
                    remove edge e from the graph
                    if m has no other incoming edges then
                        insert m into S
            if graph has edges then
                raise exception: graph has at least one cycle
            else
                return proposed topologically sorted order: L
        """
        def is_source(node, connections):
            for connection in connections:
                if node == connection.target:
                    return False
            return True

        def source_connections(node, connections):
            conns = set()
            for connection in connections:
                if node == connection.source:
                    conns.add(connection)
            return conns

        nodes = set(self.nodes.values())
        connections = self.connections.copy()
        sorted_nodes = []

        # Find source nodes:
        source_nodes = set([n for n in nodes if is_source(n, connections)])

        # while S is non-empty do
        while source_nodes:
            # remove a node n from S
            node = source_nodes.pop()
            # insert n into L
            sorted_nodes.append(node)

            # for each node m with an edge e from n to m do
            s_connections = source_connections(node, connections)
            for connection in s_connections:
                #     remove edge e from the graph
                m = connection.target
                connections.remove(connection)
                #     if m has no other incoming edges then
                #         insert m into S
                if is_source(m, connections):
                    source_nodes.add(m)

        # if graph has edges then
        #     output error message (graph has at least one cycle)
        # else
        #     output message (proposed topologically sorted order: L)

        if connections:
            raise Exception("Steram has at least one cycle (%d connections left of %d)" % (len(connections), len(self.connections)))

        return sorted_nodes

    def targets(self, node):
        """Return nodes that `node` passes data into."""
        node = self.node(node)
        nodes =[conn.target for conn in self.connections if conn.target == node]
        return nodes

    def sources(self, node):
        """Return a dictionary where keys are outlet names and values are
        nodes."""
        node = self.node(node)

        nodes = {}
        for conn in self.connections:
            if conn.target == node:
                nodes[conn.outlet] = conn.source

        return nodes



########NEW FILE########
__FILENAME__ = pipeline
# -*- coding: utf-8 -*-

from .context import default_context
from .engine import ExecutionEngine
from ..stores import open_store
from .graph import *
from ..errors import *
from ..dev import experimental

__all__ = [
            "Pipeline"
        ]

class Pipeline(object):
    def __init__(self, stores=None, context=None, graph=None, name=None):
        """Creates a new pipeline with `context` and sets current object to
        `obj`. If no context is provided, default one is used.

        Pipeline inherits operations from the `context` and uses context's
        dispatcher to call the operations. Operations are provided as
        pipeline's methods:

        .. code-block:: python

            p = Pipeline(stores={"default":source_store})
            p.source("default", "data")
            # Call an operation within context
            p.distinct("city")

            p.create("default", "cities")
            p.run()

        `name` is an optional user's pipeline identifier that is used for
        debugging purposes.

        .. note::

            You can set the `engine_class` variable to your own custom
            execution engine class with custom execution policy.

        """
        self.context = context or default_context

        self.stores = {}

        # List of owned and therefore opened stores
        self._owned_stores = []

        stores = stores or {}
        for name, store in stores.items():
            if isinstance(store, dict):
                store = dict(store)
                type_ = store.pop("type")
                store = open_store(type_, **store)
                self._owned_stores.append(store)

            self.stores[name] = store

        self.graph = graph or Graph()
        self.name = name

        # Set default execution engine
        self.engine_class = ExecutionEngine

        self.node = None

        self._test_if_needed = None
        self._test_if_satisfied = None

    def source(self, store, objname, **params):
        """Appends a source node to an empty pipeline. The source node will
        reference an object `objname` in store `store`. The actual object will
        be fetched during execution."""

        if self.node is not None:
            raise BubblesError("Can not set pipeline source: there is already "
                                "a node. Use new pipeline.")

        self.node = StoreObjectNode(store, objname, **params)
        self.graph.add(self.node)

        return self

    def source_object(self, obj, **params):
        """If `obj` is a data object, then it is set as source. If `obj` is a
        string, then it is considered as a factory and object is obtained
        using :fun:`data_object` with `params`"""
        if self.node is not None:
            raise BubblesError("Can not set pipeline source: there is already "
                                "a node. Use new pipeline.")

        if isinstance(obj, str):
            node = ObjectFactoryNode(obj, **params)
        else:
            if params:
                raise ArgumentError("params should not be specified if "
                                    "object is provided.")
            node = ObjectNode(obj)

        self.graph.add(node)
        self.node = node

        return self

    def insert_into(self, store, objname, **params):
        """Appends a node that inserts into `objname` object in `store`.  The
        actual object will be fetched during execution."""

        if self.node is None:
            raise BubblesError("Cannot insert from a empty or disconnected "
                                "pipeline.")

        target = StoreObjectNode(store, objname, **params)
        self._append_insert_into(target)

        return self

    def insert_into_object(self, obj, **params):
        """If `obj` is a data object, then it is set as target for insert. If
        `obj` is a string, then it is considered as a factory and object is
        obtained using :fun:`data_object` with `params`"""

        if self.node is None:
            raise BubblesError("Cannot insert from a empty or disconnected "
                                "pipeline.")

        if isinstance(obj, str):
            node = ObjectFactoryNode(obj, **params)
        else:
            if params:
                raise ArgumentError("params should not be specified if "
                                    "object is provided.")
            node = ObjectNode(obj)

        self._append_insert_into(node)

        return self

    def _append_insert_into(self, target, **params):
        """Appends an `insert into` node."""

        insert = Node("insert")
        self.graph.add(insert)
        self.graph.add(target)
        self.graph.connect(target, insert, "target")
        self.graph.connect(self.node, insert, "source")

        self.node = insert

    def create(self, store, name, *args, **kwargs):
        """Create new object `name` in store `name`. """

        node = CreateObjectNode(store, name, *args, **kwargs)
        self._append_node(node)

        return self

    def __getattr__(self, name):
        """Adds operation node into the stream"""

        return _PipelineOperation(self, name)

    def fork(self, empty=None):
        """Forks the current pipeline. If `empty` is ``True`` then the forked
        fork has no node set and might be used as source."""
        fork = Pipeline(self.stores, self.context, self.graph)

        if not empty:
            fork.node = self.node

        return fork

    def _append_node(self, node, outlet="default"):
        """Appends a node to the pipeline stream. The new node becomes actual
        node."""

        node_id = self.graph.add(node)
        if self.node:
            self.graph.connect(self.node, node, outlet)
        self.node = node

        return self

    def run(self, context=None):
        """Runs the pipeline in Pipeline's context. If `context` is provided
        it overrides the default context.

        There are two prerequisities for the pipeline to be run:

        * *test if needed* – pipeline is run only when needed, only when the
          test is satisfied. If the test is not satisfied
          (`ProbeAssertionError` is raised), the pipeline is gracefully
          skipped and considered successful without running.

        * *test if satisfied* - pipeline is run only when certain requirements
          are satisfied. If the requirements are not met, then an exception is
          raised and pipeline is not run.

        The difference between *"if needed"* and *"if satisfied"* is that the
        first one test whether the pipeline did already run or we already have
        the data. The second one *"if satisfied"* tests whether the pipeline
        will be able to run successfuly.

        """

        engine = self._get_engine(context)

        run = True
        if self._test_if_needed:
            try:
                engine.run(self._test_if_needed.graph)
            except ProbeAssertionError:
                name = self.name or "(unnamed)"
                self.context.logger.info("Skipping pipeline '%s', "
                                         "no need to run according to the test." %
                                         name)
                run = False

        if run and self._test_if_satisfied:
            try:
                engine.run(self._test_if_satisfied.graph)
            except ProbeAssertionError as e:
                name = self.name or "(unnamed)"
                reason = e.reason or "(unknown reason)"
                self.context.logger.error("Requirements for pipeline '%s' "
                            "are not satisfied. Reason: %s" % (name, reason))
                raise

        if run:
            result = engine.run(self.graph)
        else:
            result = None


        # TODO: run self._test_successful
        # TODO: run self.rollback if not sucessful

        return result

    def execution_plan(self, context=None):
        """Returns an execution plan of the pipeline as provided by the
        execution engine. For more information see
        :meth:`ExecutionEngine.execution_plan`.  """

        engine = self._get_engine(context)
        return engine.execution_plan(self.graph)

    def _get_engine(self, context=None):
        """Return a fresh engine instance that uses either target's context or
        explicitly specified other `context`."""
        context = context or self.context
        engine = self.engine_class(context=context, stores=self.stores)
        return engine

    def test_if_needed(self):
        """Create a branch that will be run before the pipeline is run. If the
        test passes, then pipeline will not be run. Use this, for example, to
        determine whether the data is already processed."""
        self._test_if_needed = Pipeline(self.stores, self.context)
        return self._test_if_needed

    def test_if_satisfied(self):
        """Create a branch that will be run before the pipeline is run. If the
        test fails, then pipeline will not be run and an error will be raised.
        Use this to test dependencies of the pipeline and avoid running an
        expensive process."""

        self._test_if_satisfied = Pipeline(self.stores, self.context)
        return self._test_if_satisfied

class _PipelineOperation(object):
    def __init__(self, pipeline, opname):
        """Creates a temporary object that will append an operation to the
        pipeline"""
        self.pipeline = pipeline
        self.opname = opname

    def __call__(self, *args, **kwargs):
        """Appends an operation (previously stated) with called arguments as
        operarion's parameters"""
        # FIXME: works only with unary operations, otherwise undefined
        # TODO: make this work with binary+ operations
        #
        #
        # Dev note:
        # – `operands` might be either a node - received throught
        # pipeline.node or it might be a pipeline forked from the receiver

        operation = self.pipeline.context.operation(self.opname)

        if operation.opcount == 1:
            # Unary operation
            node = Node(self.opname, *args, **kwargs)
            self.pipeline._append_node(node)

        else:
            # n-ary operation. Take operands from arguments. There is one less
            # - the one that is just being created by this function as default
            # operand within the processing pipeline.

            operands = args[0:operation.opcount-1]
            args = args[operation.opcount-1:]

            # Get name of first (pipeline default) outlet and rest of the
            # operand outlets
            firstoutlet, *restoutlets = operation.operands

            # Pipeline node - default
            node = Node(self.opname, *args, **kwargs)
            self.pipeline._append_node(node, firstoutlet)

            # All operands should be pipeline objects with a node
            if not all(isinstance(o, Pipeline) for o in operands):
                raise BubblesError("All operands should come from a Pipeline")

            # Operands are expected to be pipelines forked from this one
            for outlet, operand in zip(restoutlets, operands):
                # Take current node of the "operand pipeline"
                src_node = operand.node

                # ... and connect it to the outlet of the currently created
                # node
                self.pipeline.graph.connect(src_node, node, outlet)

        return self.pipeline

def create_pipeline(description):
    """Create a pipeline from a description. Description should be a list of
    operation descriptions. The operation is described as a tuple where first
    item is the operation name followed by operands. Last argument is a
    dictionary with options.

        ["operation", "operand", "operand", {foo}]
    """

    raise NotImplementedError


########NEW FILE########
__FILENAME__ = expression
from expressions import Compiler, ExpressionError, default_dialect

class bubbles_dialect(default_dialect):
    operators = {
        # "^": (1000, RIGHT, BINARY),
        "~": (1000, None, UNARY),
        "¬": (1000, None, UNARY),
        "*": (900, LEFT, BINARY),
        "/": (900, LEFT, BINARY),
        "%": (900, LEFT, BINARY),

        "+":  (500, LEFT, BINARY),
        "-":  (500, LEFT, UNARY | BINARY),

        "<<": (350, LEFT, BINARY),
        ">>": (350, LEFT, BINARY),

        "&":  (300, LEFT, BINARY),
        "|":  (300, LEFT, BINARY),

        "<":  (200, LEFT, BINARY),
        "≤": (200, LEFT, BINARY),
        "<=": (200, LEFT, BINARY),
        ">":  (200, LEFT, BINARY),
        "≥": (200, LEFT, BINARY),
        ">=": (200, LEFT, BINARY),
        "≠": (200, LEFT, BINARY),
        "!=": (200, LEFT, BINARY),
        "=":  (200, LEFT, BINARY),

        "not": (120, None, UNARY),
        "and": (110, LEFT, BINARY),
        "or":  (100, LEFT, BINARY),
    }
    case_sensitive = False

    # TODO: future function list (unused now)
    functions = (
            "min", "max",
            )

# IS blank
# IS null

# Translate bubbles operator to Python operator
_python_operators = {
    "=": "=="
}


class PythonExpressionCompiler(Compiler):
    def compile_literal(self, context, literal):
        # TODO: support string literals (quote them properely)
        if isinstance(literal, str):
            raise NotImplementedError("String literals are not yet supported")
        else:
            return str(literal)

    def compile_variable(self, context, variable):
        if variable in context:
            return variable
        else:
            raise ExpressionError("Unknown variable %s" % variable)

    def compile_operator(self, context, operator, op1, op2):
        operator = _python_operators.get(operator, operator)
        return "(%s%s%s)" % (op1, operator, op2)

    def compile_unary(self, context, operator, operand):
        operator = _python_operators.get(operator, operator)
        return "(%s%s)" % (operand, operator)

    def compile_function(self, context, function, args):
        # TODO: support functions
        raise NotImplementedError("Functions are not yet implemented")


def PythonLambdaCompiler(PythonExpressionCompiler):
    def finalize(self, context):
        # TODO: sanitize field names in context
        field_list = ", ".join(context)
        lambda_str = "lambda %s: %s" % (context, expression)
        return lambda_str


def lambda_from_predicate(predicate, key):
    compiler = PythonLambdaCompiler()
    expression = compiler.compile(predicate, context=key)
    return expression

########NEW FILE########
__FILENAME__ = extensions
# -*- coding=utf -*-
from .common import decamelize, to_identifier
from .errors import *
from collections import defaultdict


__all__ = [
    "Extensible",
    "extensions"
]

# Known extension types.
# Keys:
#     base: extension base class name
#     suffix: extension class suffix to be removed for default name (same as
#         base class nameif not specified)
#     modules: a dictionary of extension names and module name to be loaded
#         laily

_default_modules = {
    "store": {
        "sql":"bubbles.backends.sql.objects",
        "mongo":"bubbles.backends.sql.mongo",
        "csv":"bubbles.backends.text.objects",
        "datapackage":"bubbles.datapackage",
        "datapackages":"bubbles.datapackage",
    },
    "object": {
        "csv_source":"bubbles.backends.text.objects",
        "csv_target":"bubbles.backends.text.objects",
        "xls":"bubbles.backends.xls"
    },
}

_base_modules = {
    "store": "bubbles.stores",
    "object": "bubbles.objects",
}

class Extensible(object):
    """For now just an extension superclass to find it's subclasses."""

    """Extension type, such as `store` or `browser`. Default is derived
    from the extension root class name."""
    __extension_type__ = None

    """Class name suffix to be stripped to get extension's base name. Default
    is the root class name"""
    __extension_suffix__ = None

    """Extension name, such as `sql`. Default is derived from the extension
    class name."""
    __extension_name__ = None
    __extension_aliases__ = []

    """List of extension options.  The options is a list of dictionaries with
    keys:

    * `name` – option name
    * `type` – option data type (default is ``string``)
    * `description` – description (optional)
    * `label` – human readable label (optional)
    * `values` – valid values for the option.
    """
    __options__ = None

class ExtensionsFactory(object):
    def __init__(self, root):
        """Creates an extension factory for extension root class `root`."""
        self.root = root
        name = root.__name__

        # Get extension collection name, such as 'stores', 'browsers', ...
        self.name = root.__extension_type__
        if not self.name:
            self.name = to_identifier(decamelize(name))

        self.suffix = root.__extension_suffix__
        # Accept "" empty string as no suffix. Treat `None` as default.
        if self.suffix is None:
            self.suffix = name

        self.options = {}
        self.option_types = {}

        self.extensions = {}

        for option in root.__options__ or []:
            name = option["name"]
            self.options[name] = option
            self.option_types[name] = option.get("type", "string")

    def __call__(self, _extension_name, *args, **kwargs):
        return self.create(_extension_name, *args, **kwargs)

    def create(self, _extension_name, *args, **kwargs):
        """Creates an extension. First argument should be extension's name."""
        extension = self.get(_extension_name)

        option_types = dict(self.option_types)
        for option in extension.__options__ or []:
            name = option["name"]
            option_types[name] = option.get("type", "string")

        kwargs = coalesce_options(dict(kwargs), option_types)

        return extension(*args, **kwargs)


    def get(self, name):
        if name in self.extensions:
            return self.extensions[name]

        # Load module...
        modules = _default_modules.get(self.name)
        if modules and name in modules:
            # TODO don't load module twice (once for manager once here)
            _load_module(modules[name])

        self.discover()

        try:
            return self.extensions[name]
        except KeyError:
            raise InternalError("Unknown extension '%s' of type %s"
                                     % (name, self.name))

    def discover(self):
        extensions = collect_subclasses(self.root, self.suffix)
        self.extensions.update(extensions)

        aliases = {}
        for name, ext in extensions.items():
            if ext.__extension_aliases__:
                for alias in ext.__extension_aliases__:
                    aliases[alias] = ext

        self.extensions.update(aliases)


class ExtensionsManager(object):
    def __init__(self):
        self.managers = {}
        self._is_initialized = False

    def __lazy_init__(self):
        for root in Extensible.__subclasses__():
            manager = ExtensionsFactory(root)
            self.managers[manager.name] = manager
        self._is_initialized = True

    def __getattr__(self, type_):
        if not self._is_initialized:
            self.__lazy_init__()

        if type_ in self.managers:
            return self.managers[type_]

        # Retry with loading the required base module

        _load_module(_base_modules[type_])
        self.__lazy_init__()

        try:
            return self.managers[type_]
        except KeyError:
            raise InternalError("Unknown extension type '%s'" % type_)

"""Extensions provider. Use::

    browser = extensions.browser("sql", ...)

"""
extensions = ExtensionsManager()


def collect_subclasses(parent, suffix=None):
    """Collect all subclasses of `parent` and return a dictionary where keys
    are object names. Obect name is decamelized class names transformed to
    identifiers and with `suffix` removed. If a class has class attribute
    `__identifier__` then the attribute is used as name."""

    subclasses = {}
    for c in subclass_iterator(parent):
        name = None
        if hasattr(c, "__extension_name__"):
            name = getattr(c, "__extension_name__")
        elif hasattr(c, "__identifier__"):
            # TODO: depreciated
            name = getattr(c, "__identifier__")

        if not name:
            name = c.__name__
            if suffix and name.endswith(suffix):
                name = name[:-len(suffix)]
            name = to_identifier(decamelize(name))

        subclasses[name] = c

    return subclasses


def subclass_iterator(cls, _seen=None):
    """
    Generator over all subclasses of a given class, in depth first order.

    Source: http://code.activestate.com/recipes/576949-find-all-subclasses-of-a-given-class/
    """

    if not isinstance(cls, type):
        raise TypeError('_subclass_iterator must be called with '
                        'new-style classes, not %.100r' % cls)

    _seen = _seen or set()

    try:
        subs = cls.__subclasses__()
    except TypeError: # fails only when cls is type
        subs = cls.__subclasses__(cls)
    for sub in subs:
        if sub not in _seen:
            _seen.add(sub)
            yield sub
            for sub in subclass_iterator(sub, _seen):
                yield sub

def _load_module(modulepath):
    mod = __import__(modulepath)
    path = []
    for token in modulepath.split(".")[1:]:
       path.append(token)
       mod = getattr(mod, token)
    return mod

def coalesce_options(options, types):
    """Coalesce `options` dictionary according to types dictionary. Keys in
    `types` refer to keys in `options`, values of `types` are value types:
    string, list, float, integer or bool."""

    out = {}

    for key, value in options.items():
        if key in types:
            out[key] = coalesce_option_value(value, types[key], key)
        else:
            out[key] = value

    return out

def coalesce_option_value(value, value_type, label=None):
    """Convert string into an object value of `value_type`. The type might be:
        `string` (no conversion), `integer`, `float`, `list` – comma separated
        list of strings.
    """
    value_type = value_type.lower()

    try:
        if value_type in ('string', 'str'):
            return_value = str(value)
        elif value_type == 'list':
            if isinstance(value, basestring):
                return_value = value.split(",")
            else:
                return_value = list(value)
        elif value_type == "float":
            return_value = float(value)
        elif value_type in ["integer", "int"]:
            return_value = int(value)
        elif value_type in ["bool", "boolean"]:
            if not value:
                return_value = False
            elif isinstance(value, basestring):
                return_value = value.lower() in ["1", "true", "yes", "on"]
            else:
                return_value = bool(value)
        else:
            raise ArgumentError("Unknown option value type %s" % value_type)

    except ValueError:
        if label:
            label = "parameter %s " % label
        else:
            label = ""

        raise ArgumentError("Unable to convert %svalue '%s' into type %s" %
                                                (label, astring, value_type))
    return return_value

########NEW FILE########
__FILENAME__ = metadata
# -*- coding: utf-8 -*-
import copy
import itertools
import functools
import re
import inspect
import warnings
from .common import get_logger, IgnoringDictionary
from .errors import *

# from collections import OrderedDict

__all__ = [
    "to_field",
    "Field",
    "FieldList",
    "FieldFilter",
    "storage_types",
    "analytical_types",
    "distill_aggregate_measures",
    "prepare_key",
    "prepare_aggregation_list",
    "prepare_order_list",
    "DEFAULT_ANALYTICAL_TYPES"
]

"""Abstracted field storage types"""
storage_types = (
        "unknown",  # Unspecified storage type, processing behavior is undefined
        "string",   # names, labels, up to hundreds of hundreds of chars
        "text",     # bigger text storage
        "integer",  # integer numeric types
        "number",    # floating point types

        "boolean",  # two-state value

        "datetime", # Full date and time with time zone
        "time",     # time without a date
        "date",     # 

        "array",    # ordered collection type
        "object",   # JSON-like object
        "binary",   # any representation of binary data (byte string or base64
                    # encoded string

        "geopoint", # a tuple: (longitude, latitude)
    )



"""Analytical types used by analytical nodes"""
analytical_types = ("default",  # unspecified or based on storage type
                    "typeless", # not relevant
                    "flag",     # two-element set
                    "discrete", # mostly integer with allowed arithmentic
                    "measure",  # mostly floating point number
                    "nominal",  # unordered set
                    "ordinal"   # ordered set
                    )

"""Mapping between storage types and their respective default analytical
types"""
# NOTE: For the time being, this is private
DEFAULT_ANALYTICAL_TYPES = {
                "unknown": "typeless",
                "string": "typeless",
                "text": "typeless",
                "integer": "discrete",
                "number": "measure",
                "date": "typeless",
                "array": "typeless",
                "document": "typeless"
            }


FIELD_ATTRIBUTES = (
    "name",
    "label",
    "storage_type",
    "analytical_type",
    "concrete_storage_type",
    "size",
    "missing_value",
    "info",
    "origin",
    "description"
)


def to_field(obj):
    """Converts `obj` to a field object. `obj` can be ``str``, ``tuple``
    (``list``), ``dict`` object or :class:`Field` object. If it is `Field`
    instance, then same object is passed.

    If field is not a `Field` instance, then construction of new field is as follows:

    ``str``:
        `field name` is set

    ``tuple``:
        (`field_name`, `storaget_type`, `analytical_type`), the `field_name` is
        obligatory, rest is optional

    ``dict``
        contains key-value pairs for initializing a :class:`Field` object

    Attributes of a field that are not specified in the `obj` are filled as:
    `storage_type` is set to ``unknown``, `analytical_type` is set to
    ``typeless``
    """

    if isinstance(obj, Field):
        field = obj
    else:
        # Set defaults first
        d = { }

        if isinstance(obj, str):
            d["name"] = obj
        elif isinstance(obj, (list, tuple)):
            d["name"] = obj[0]

            try:
                d["storage_type"] = obj[1]
                try:
                    d["analytical_type"] = obj[2]
                except IndexError:
                    pass
            except IndexError:
                pass

        elif isinstance(obj, dict):
            for attr in FIELD_ATTRIBUTES:
                if attr in obj:
                    d[attr] = obj[attr]

        else:
            raise ArgumentError("Unable to create field from %s" % (type(obj), ))

        if "analytical_type" not in d:
            storage_type = d.get("storage_type")
            if storage_type:
                deftype = DEFAULT_ANALYTICAL_TYPES.get(storage_type)

        field = Field(**d)

    return field


class Field(object):
    """`Field` is a metadata that describes part of data object's structure.
    It might refer to a table column, document key or any other descriptor of
    common data. Fields are identified by name.

    Attributes:

    * `name` - field name.
    * `label` - optional human readable field label that might be used in
      end-user applications.
    * `storage_type` - Normalized data storage type to one of the Bubble's
      basic storage types.
    * `concrete_storage_type` (optional, recommended) - Data store/database
      dependent storage type - this is the real data type descriptor as used
      in a backend system, such as database, where the field comes from or
      where the field is going to be created. If it is `None`, then backend
      tries to use a default type for the field's `storage_type`. Value might
      be any object understandable by the backend that will be "touching" the
      object's data.
    * `analytical_type` - data type from user's perspective. Describes the
      intention how the field is used. Might also be used for restricting some
      functionality based on the type.
    * `size` – optional field size - interpretation of the value is
      related to the `storage_type` and/or `concrete_storaget_type`. For
      example it might be length of text fields or list of array
      dimensions for array type.
    * `missing_values` (optional) - Array of values that represent missing
      values in the dataset for given field
    * `info` – user specific field information, might contain formatting
      information for example

    """
    # TODO: make this public once ownership mechanism is redesigned
    # * `origin` – field or field list from which this field was derived
    # * `owner` – object that created this field

    attribute_defaults = {
                "storage_type":"string",
                "analytical_type": None
            }

    def __init__(self, name=None, storage_type=None, analytical_type=None,
                 concrete_storage_type=None, size=None, missing_value=None,
                 label=None, info=None, origin=None, description=None):

        self.name = name
        self.storage_type = storage_type
        self.analytical_type = analytical_type
        self.concrete_storage_type = concrete_storage_type
        self.size = size
        self.missing_value = missing_value
        self.label = label
        self.info = info
        self.origin = origin
        self.description = description

    def clone(self, **attributes):
        """Clone a field and set attributes"""
        d = self.to_dict()
        d.update(attributes)
        return Field(**d)

    def to_dict(self):
        """Return dictionary representation of the field."""
        d = IgnoringDictionary()

        # Note: the ignoring dictionary skips empty values. It also preserves
        # the orider (for nicer JSON output)
        for attr in FIELD_ATTRIBUTES:
            d[attr] = getattr(self, attr)

        return d

    def __deepcopy__(self, memo):
        field = Field(self.name,
                      self.storage_type,
                      self.analytical_type,
                      self.concrete_storage_type,
                      self.size,
                      self.missing_value,
                      self.label,
                      copy.deepcopy(self.info, memo),
                      self.origin,
                      self.description)
        return field

    def __str__(self):
        """Return field name as field string representation."""
        return self.name

    def __repr__(self):
        return "%s(%s)" % (self.__class__.__name__, self.to_dict())

    def __eq__(self, other):
        if self is other:
            return True

        if not isinstance(other, Field):
            return False

        for attr in FIELD_ATTRIBUTES:
            if getattr(self, attr) != getattr(other, attr):
                return False
        return True

    def __ne__(self,other):
        return not self.__eq__(other)


class FieldList(object):
    """List of fields"""
    def __init__(self, *fields):
        """
        Main description of data object structure. Created from a list of
        :class:`Field` objects from a list of strings, dictionaries or tuples


        How fields are consutrcuted:

        * string: `field name` is set
        * tuple: (`field_name`, `storaget_type`, `analytical_type`), the `field_name` is
          obligatory, rest is optional
        * dict: contains key-value pairs for initializing a :class:`Field` object

        Example: `FieldList('name', 'address')`.

        For strings and in if not explicitly specified in a tuple or a dict case, then following rules
        apply:

        * `storage_type` is set to ``unknown``
        * `analytical_type` is set to ``typeless``
        """
        super(FieldList, self).__init__()

        # FIXME: use OrderedDict (Python 2.7+)
        self._fields = []
        self._field_dict = {}
        self._field_names = []

        if fields:
            # Convert input to Field instances
            # This is convenience, so one can pass list of strsings, for example

            for field in fields:
                self.append(field)

    def append(self, field):
        """Appends a field to the list. This method requires `field` to be
        instance of `Field`"""

        # FIXME: depreciated: FieldList should be immutable
        field = to_field(field)
        self._fields.append(field)
        self._field_dict[field.name] = field
        self._field_names.append(field.name)

    def names(self, indexes=None):
        """Return names of fields in the list.

        :Parameters:
            * `indexes` - list of indexes for which field names should be collected. If set to
              ``None`` then all field names are collected - this is default behaviour.
        """

        if indexes:
            names = [self._field_names[i] for i in indexes]
            return names
        else:
            return self._field_names


    def indexes(self, fields):
        """Return a tuple with indexes of fields from ``fields`` in a data row. Fields
        should be a list of ``Field`` objects or strings.

        This method is useful when it is more desirable to process data as rows (arrays), not as
        dictionaries, for example for performance purposes.
        """

        indexes = [self.index(field) for field in fields]

        return tuple(indexes)

    def index_map(self):
        """Returns a map of field name to field index"""
        return dict( (f, i) for f, i in enumerate(self._field_names))

    def mask(self, fields=None):
        """Return a list representing field selector - which fields are
        selected from a row."""

        sel_names = [str(field) for field in fields]

        mask = [str(name) in sel_names for name in self.names()]
        return mask

    def index(self, field):
        """Return index of a field"""

        try:
            index = self._field_names.index(str(field))
        except ValueError:
            raise NoSuchFieldError("Field list has no field with name '%s'" % str(field))

        return index

    def fields(self, names=None, storage_type=None, analytical_type=None):
        """Return a tuple with fields. `names` specifies which fields are returned. When names is
        ``None`` all fields are returned. `storage_type` or `analytical_type`
        is specified, only fields of that type are returned.
        """

        if not names:
            fields = self._fields
        else:
            fields = [self._field_dict[str(name)] for name in names]

        if storage_type:
            fields = [f for f in fields if f.storage_type == storage_type]

        if analytical_type:
            fields = [f for f in fields if f.analytical_type == analytical_type]

        return FieldList(*fields)

    def aggregated_fields(self, aggregation_list, include_count=True,
                          count_field="record_count"):
        """Returns a `FieldList` containing fields after aggregations of
        measures in the `aggregation_list`. The list should be a list of
        tuples `(field, aggregtion)` and the `field` of the tuple should be in
        the receiving `FieldList`. You can prepare the aggregation list using
        the :func:`prepare_aggregation_list` function.

        Resulting fields are cloned from the original fields and will have
        analytical type set to ``measure``.

        Example:

        >>> agg_list = aggregation_list(['amount', ('discount', 'avg')])
        >>> agg_fields = fields.aggregated_fields(agg_list)

        Will return fields with names: `('amount_sum', 'discount_avg',
        'record_count')`
        """

        agg_fields = FieldList()

        for measure in measures:
            if isinstance(measure, (str, Field)):
                field = str(measure)
                index = fields.index(field)
                aggregate = "sum"
            elif isinstance(measure, (list, tuple)):
                field = measure[0]
                index = fields.index(field)
                aggregate = measure[1]

            field = fields.field(measure)
            field = field.clone(name="%s_%s" % (str(measure), aggregate),
                                analytical_type="measure")
            agg_fields.append(field)

        return agg_fields

    def field(self, ref):
        """Return a field with name `ref` if `ref` is a string, or if it is an
        integer, returns a field at that index."""

        if isinstance(ref, int):
            return self._fields[ref]
        else:
            try:
                return self._field_dict[ref]
            except KeyError:
                raise NoSuchFieldError("Field list has no field with name "
                                        "'%s'" % (ref,) )

    def __len__(self):
        return len(self._fields)

    def __getitem__(self, reference):
        return self.field(reference)

    def __setitem__(self, index, new_field):
        field = self._fields[index]
        del self._field_dict[field.name]
        self._fields[index] = new_field
        self._field_names[index] = new_field.name
        self._field_dict[new_field.name] = new_field

    def __delitem__(self, index):
        field = self._fields[index]
        del self._field_dict[field.name]
        del self._fields[index]
        del self._field_names[index]

    def __iter__(self):
        return self._fields.__iter__()

    def __contains__(self, field):
        if isinstance(field, str):
            return field in self._field_names

        return field in self._fields

    def __iadd__(self, array):
        for field in array:
            self.append(field)

        return self

    def __add__(self, array):
        fields = self.copy()
        fields += array
        return fields

    def __str__(self):
        return "[" + ", ".join(self.names()) + "]"

    def __repr__(self):
        frepr = [repr(field) for field in self._fields]
        return "%s([%s])" % (self.__class__.__name__, ",".join(frepr))

    def __eq__(self, other):
        if not isinstance(other, FieldList):
            return False
        return other._fields == self._fields

    def copy(self):
        """Return a shallow copy of the list.
        """
        return FieldList(*self._fields)

    def clone(self, fields=None, origin=None, owner=None):
        """Creates a copy of the list and copy of the fields. Automatically
        sets the origin of fields to be the original field or `origin` if
        specified.
        """
        fields = self.fields(fields)

        cloned_fields = FieldList()
        for field in fields:
            new_field = copy.copy(field)
            new_field.origin = origin or field
            cloned_fields.append(new_field)

        return cloned_fields


class FieldFilter(object):
    """Filters fields in a stream"""
    # TODO: preserve order of "keep"
    def __init__(self, keep=None, drop=None, rename=None):
        """Creates a field map. `rename` is a dictionary where keys are input
        field names and values are output field names. `drop` is list of
        field names that will be dropped from the stream. If `keep` is used,
        then all fields are dropped except those specified in `keep` list."""
        if drop and keep:
            raise MetadataError("You can nott specify both 'keep' and 'drop' "
                                "options in FieldFilter.")

        super(FieldFilter, self).__init__()

        self.rename = rename or {}

        if drop:
            self.drop = [str(f) for f in drop]
        else:
            self.drop = []

        if keep:
            self.keep = [str(f) for f in keep]
        else:
            self.keep = []

    def filter(self, fields):
        """Map `fields` according to the FieldFilter: rename or drop fields as
        specified. Returns a new FieldList object.

        .. note::

            For each renamed field a new copy is created. Not renamed fields
            are the same as in `fields`. To use filtered fields in a node
            you have to clone the field list.
        """
        output_fields = FieldList()

        # Check whether we have all fields requested
        fields = FieldList(*fields)

        names = fields.names()

        for field in self.keep:
            if field not in names:
                raise NoSuchFieldError(field)

        for field in self.drop:
            if field not in names:
                raise NoSuchFieldError(field)

        for field in fields:
            if field.name in self.rename:
                # Create a copy and rename field if it is mapped
                new_field = copy.copy(field)
                new_field.name = self.rename[field.name]
            else:
                new_field = field

            if (self.drop and field.name not in self.drop) or \
                (self.keep and field.name in self.keep) or \
                not (self.keep or self.drop):
                output_fields.append(new_field)

        return output_fields

    def row_filter(self, fields):
        """Returns an object that will convert rows with structure specified in
        `fields`. You can use the object to filter fields from a row (list,
        array) according to this map.
        """
        return RowFieldFilter(self.field_mask(fields))

    def field_mask(self, fields):
        """Returns a list where ``True`` value is set for field that is selected
        and ``False`` for field that has to be ignored. Selectors of fields can
        be used by `itertools.compress()`. This is the preferred way of field
        filtering.
        """

        selectors = []

        for field in fields:
            flag = (self.drop and field.name not in self.drop) \
                    or (self.keep and field.name in self.keep) \
                    or not (self.keep or self.drop)
            selectors.append(flag)

        return selectors


class RowFieldFilter(object):
    """Class for filtering fields in array"""

    def __init__(self, mask=None):
        """Create an instance of RowFieldFilter. `mask` is a list of indexes
        that are passed to output."""
        super(RowFieldFilter, self).__init__()
        self.mask = mask or []

    def __call__(self, row):
        return self.filter(row)

    def filter(self, row):
        """Filter a `row` according to ``indexes``."""
        return tuple(value for value,mask in zip(row, self.mask) if mask)

    def __repr__(self):
        return "%s(%s)" % (self.__class__.__name__, self.mask)


def distill_aggregate_measures(measures, default_aggregates=None):
    """Normalizes list of measures. Element of the source can be:
    * a string - default aggregations will be used or `sum`` if defaults are
      not provided
    * a field - default aggregations will be used or `sum` if defaults are not
      provided
    * tuple with first element as measure specification and second element
      either aggregation string or list of aggregations

    Returns list of tuples: (`measure`, `aggregate`)
    """
    default_aggregates = default_aggregates or []

    distilled_measures = []
    for measure_aggs in measures:
        # Extract measure and aggregate. If no aggregate is specified,
        # "sum" is assumed.

        if isinstance(measure_aggs, list) or isinstance(measure_aggs, tuple):
            measure = measure_aggs[0]
            aggregates = measure_aggs[1]
            if not (isinstance(aggregates, list) or
                        isinstance(aggregates,tuple)):
                aggregates = [aggregates]
        else:
            measure = measure_aggs
            aggregates = default_aggregates or ["sum"]

        for aggregate in aggregates:
            distilled_measures.append( (measure, aggregate) )

    return distilled_measures


def prepare_key(key):
    """Returns a list of columns for `key`. Key might be a list of fields or
    just one field represented by `Field` object or a string. Examples:

    >>> key = prepare_key(obj, "id")
    >>> key = prepare_key(obj, ["code", "name"])

    If `statement` is not `None`, then columns for the key are returned.
    """

    if isinstance(key, (str, Field)):
        key = (key, )

    return tuple(str(f) for f in key)


##
# Various metadata helpers
#

def prepare_aggregation_list(measures):
    """Coalesces list of measures for aggregation.  `measures` should be a
    list of tuples `(field, aggregtion)`or just fields.  If just a field is
    specified, then `sum` aggregation is assumed. Returns correct list of
    tuples."""

    return prepare_tuple_list(measures, "sum")

def prepare_order_list(fields):
    """Coalesces list of fields for ordering. Accepts: a string, list of
    strings, list of tuples `(field, order)`. Default order is ``asc``."""

    return prepare_tuple_list(fields, "asc")

def prepare_tuple_list(fields, default_value):
    """Coalesces list of fields to list of tuples. Accepts: a string, list of
    strings, list of tuples `(field, value)`. """

    if not fields:
        return []

    result = []

    if not isinstance(fields, (list, tuple, FieldList)):
        fields = [fields]

    for obj in fields:
        if isinstance(obj, (str, Field)):
            field = str(obj)
            value = default_value
        elif isinstance(obj, (list, tuple)):
            field, value = obj
        result.append( (field, value) )

    return result

########NEW FILE########
__FILENAME__ = objects
# -*- Encoding: utf8 -*-

"""Data Objects"""

# FIXME: add this
# from .ops.iterator import as_records
from .errors import *
from .extensions import Extensible, extensions
from .metadata import *
from .dev import required, experimental

__all__ = [
        "DataObject",
        "IterableDataSource",
        "RowListDataObject",
        "IterableRecordsDataSource",

        "shared_representations",
        "data_object",
        ]

def data_object(type_, *args, **kwargs):
    """Returns a data object of specified `type_`. Arguments are passed to
    respective data object factory.

    Available data object type in the base framework:

    * `sql_table`
    * `sql`
    * `csv_source`
    * `csv_target`
    * `iterable`
    * `iterable_records`
    * `row_list`
    """

    return extensions.object(type_, *args, **kwargs)


def iterator_object(iterator, fields):
    """Returns a data object wrapping an `iterator` with rows of `fields`."""
    return IterableDataSource(iterator, fields)


class DataObject(Extensible):
    __extension_type__ = "object"
    __extension_suffix__ = "Object"

    def representations(self):
        """Returns list of representation names of this data object. Default
        implementation raises an exception, as subclasses are required to
        implement this method.

        Representations do not have to be known up front – they might depend
        on various run-time conditions. Refer to particular object
        documentation.
        """
        raise NotImplementedError("Subclasses are required to specify "
                                  "representations of a data object.")

    def is_compatible(self, obj, required=None, ignored=None):
        """Returns `True` when the receiver and the `object` share
        representations. `required` contains list of representations that at
        least one of them is required. If not present, then returns `False`.
        `ignored` is a list of representations that are not relevant for the
        comparison."""
        required = set(required or [])
        ignored = set(ignored or [])
        ours = set(self.representations() or [])
        theirs = set(obj.representations or [])

        reprs = (ours - ignored) & (theirs - ignored)
        if required:
            reprs = reprs & required

        return len(reprs) > 0

    @required
    def can_compose(self, obj):
        """Returns `True` when any of the representations can be naturally
        (without a proxy) composed by any of the representations of `obj` to
        form new representation.  Example of composable objects are SQL
        statement object from the same engine. Subclasses should implement
        this method. Default implementation returns `False`, which means that
        the only suggested composition is to use iterators.

        The method should be transient. That is, if A can be naturally
        composed with B and B with C, then A can be naturally composed with C
        as well. This property is for simplification of finding whether a list
        of objects can be composed together or not.
        """

        return False

    def truncate(self):
        """Removes all records from the target table"""
        raise NotImplementedError

    def flush(self):
        """Flushes oustanding data. Default implementation does nothing.
        Subclasses might implement this method if necessary."""
        pass

    def is_consumable(self):
        """Returns `True` if the object is consumed when used, returns `False`
        when object can be used indefinitely number of times."""
        raise NotImplementedError("Data objects are required to implement "
                                  "is_consumable() method")

    def retained(self, count=1):
        """Returns object's replacement which can be consumed `count` times.
        Implementation of object retention depends on the backend.

        For example default iterable data object consumes the rows into a list
        and provides data object which wraps the list and deletes the list
        after `count` number of uses.

        .. note::

            If the object's retention policy is not appropriate for your task
            (memory or time hungry), it is recommended to cache the object
            content before consumption.

        .. note::

            Not all objects or all representations might be retained. Refer
            to particular object implementation for more information.

        Default implementation returns the receiver if it is not consumable
        and raises an exception if it is consumable. Consumable objects should
        implement this method.
        """
        if self.is_consumable():
            return self
        else:
            raise NotImplementedError

            # TODO: should we use this? Isn't it too dangerous to use this
            # very naive and hungry implementation? It is definitely
            # convenient.
            #
            # return RowListDataSource(self.rows(), self.fields)

    def __iter__(self):
        return self.rows()

    def records(self):
        """Returns an iterator of records - dictionary-like objects that can
        be acessed by field names. Default implementation returns
        dictionaries, however other objects mith return their own structures.
        For example the SQL object returns the same iterator as for rows, as
        it can serve as key-value structure as well."""

        names = [str(field) for field in fields]

        for row in self.rows():
            yield dict(zip(names, row))

    def append(self, row):
        """Appends `row` to the object. Subclasses should implement this."""
        raise NotImplementedError

    def append_from(self, obj):
        """Appends data from object `obj` which might be a `DataObject`
        instance or an iterable. Default implementation uses iterable and
        calls `append()` for each element of the iterable.

        This method executes the append instead of returning a composed
        object. For object composition see `append` operation.
        """
        for row in iter(obj):
            self.append(row)

    def append_from_iterable(self, iterator):
        """Appends data from iterator which has same fields as the receiver.
        This method actualy executes the append instead of returning a
        composed object. For object composition see `append` operation"""
        obj = IterableDataSource(iterator, self.fields)
        self.append_from(obj)

    def as_source(self):
        """Returns version of the object that can be used as source. Subclasses
        might return an object that will raise exception on attempt to use
        target-only methods such as appending.

        Default implementation returns the receiver."""
        return self

    def as_target(self):
        """Returns version of the object that can be used as target. Subclasses
        might return an object that will raise exception on attempt to use
        source-only methods such as iterating rows.

        Default implementation returns the receiver."""
        return self

    def finalize(self):
        """Subclasses should implement this method if they need to release
        resources (memory, database connection, open files, ...) acquired
        during the object's lifetime. Default implementation does nothing."""
        pass

class ObjectPool(set):

    def release(self):
        """Release all retained objects. The release order should not be
        anticipated."""

        while self:
            obj = objects.pop()
            obj.release()


def shared_representations(objects):
    """Returns representations that are shared by all `objects`"""
    objects = list(objects.values())
    reps = set(objects[0].representations())
    for obj in objects[1:]:
        reps &= set(obj.representations())

    return reps

class IterableDataSource(DataObject):
    """Wrapped Python iterator that serves as data source. The iterator should
    yield "rows" – list of values according to `fields` """

    __identifier__ = "iterable"

    _bubbles_info = {
        "attributes": [
            {"name":"iterable", "description": "Python iterable object"},
            {"name":"fields", "description":"fields of the iterable"}
        ]
    }


    def __init__(self, iterable, fields):
        """Create a data object that wraps an iterable."""
        self.fields = fields
        self.iterable = iterable

    def representations(self):
        """Returns the only representation of iterable object, which is
        `rows`"""
        return ["rows", "records"]

    def rows(self):
        return iter(self.iterable)

    def records(self):
        return as_records(self.iterable, self.fields)

    def is_consumable(self):
        return True

    def retained(self, retain_count=1):
        """Returns retained replacement of the receiver. Default
        implementation consumes the iterator into a list and returns a data
        object wrapping the list, which might leave big memory footprint on
        larger datasets. In this case it is recommended to explicitly cache
        the consumable object using other means.
        """

        return RowListDataObject(list(self.iterable), self.fields)

    def filter(self, keep=None, drop=None, rename=None):
        """Returns another iterable data source with filtered fields"""

        ffilter = FieldFilter(keep=keep, drop=drop, rename=rename)
        fields = ffilter.filter(self.fields)

        if keep or drop:
            iterator = ops.iterator.field_filter(self.iterable, self.fields,
                                                 ffilter)
        else:
            # No need to filter if we are just renaming, reuse the iterator
            iterator = self.iterable

        return IterableDataSource(iterator, fields)


class IterableRecordsDataSource(IterableDataSource):
    """Wrapped Python iterator that serves as data source. The iterator should
    yield "records" – dictionaries with keys as specified in `fields` """

    __identifier__ = "iterable_records"

    _bubbles_info = {
        "attributes": [
            {"name":"iterable", "description": "Python iterable object"},
            {"name":"fields", "description":"fields of the iterable"}
        ]
    }

    def rows(self):
        names = [str(field) for field in self.fields]
        for record in self.iterable:
            yield [record[f] for f in names]

    def records(self):
        return iter(self.iterable)

    def is_consumable(self):
        return True


class RowListDataObject(DataObject):
    """Wrapped Python list that serves as data source or data target. The list
    content are "rows" – lists of values corresponding to `fields`.

    If list is not provided, one will be created.
    """

    __identifier__ = "list"

    _bubbles_info = {
        "attributes": [
            {"name":"data", "description": "List object."},
            {"name":"fields", "description":"fields of the iterable"}
        ]
    }

    def __init__(self, data=None, fields=None):
        """Create a data object that wraps an iterable. The object is dumb,
        does not perform any field checking, accepts anything passed to it.
        `data` should be appendable list-like object, if not provided, empty
        list is created."""

        self.fields = fields
        if data is None:
            self.data = []
        else:
            self.data = data

    def representations(self):
        """Returns the only representation of iterable object, which is
        `rows`"""
        return ["rows", "records"]

    def rows(self):
        return iter(self.data)

    def records(self):
        return as_records(self.rows(), self.fields)

    def is_consumable(self):
        return False

    def append(self, row):
        self.data.append(row)

    def truncate(self):
        self.data = []


class RowToRecordConverter(object):
    def __init__(self, fields, ignore_empty=False):
        """Creates a converter from rows (list) to a record (dictionary). If
        `ignore_empty` is `True` then keys for empty field values are not
        created."""

        self.field_names = (str(field) for field in fields)
    def __call__(self, row):
        """Returns a record from `row`"""
        if ignore_empty:
            items = ((k, v) for (k, v) in zip(field_names, row) if v is not None)
            return dict(items)
        else:
            return dict(list(zip(field_names, row)))

class RecordToRowConverter(object):
    def __init__(self, fields):
        """Creates a converter from records (dict) to rows (list)."""
        self.field_names = (str(field) for field in fields)

    def __call__(self, record):
        row = (record.get(name) for name in self.field_names)
        return row


########NEW FILE########
__FILENAME__ = operation
# -*- Encoding: utf8 -*-
"""DataObject operations"""

from .errors import *
from .objects import *
from .extensions import collect_subclasses
from .common import get_logger
from collections import OrderedDict, namedtuple
from .dev import is_experimental

import itertools
import inspect

__all__ = (
            "Signature",
            "Operation",
            "operation",
            "common_representations",
            "get_representations"
        )

Operand = namedtuple("Operand", ["rep", "islist", "isany"])

def rep_to_operand(rep):
    """Converts representation to `Operand` definition"""

    if rep.endswith("[]"):
        optype = rep[:-2]
        islist = True
    else:
        optype = rep
        islist = False

    isany = (optype == "*")

    return Operand(optype, islist, isany)


class Signature(object):
    def __init__(self, *signature):
        """Creates an operation signature. Sitnature is a list of object
        representations for operation arguments.

        * An unary operation on a SQL object has signature: `Signature("sql")`
        * Binary operation on two iterators has signature:
          `Signature("rows", "rows")`
        * Operation on a list of iterators, such as operation that appends
          multiople iterators, has signature `Signature("rows[]")`
        * Operation that accepts any kind of object: `Signature("*")`
        * Operation that accepts a list of any objects: `Signature("*[]")`
        """

        self.signature = tuple(signature)
        self.operands = tuple(rep_to_operand(rep) for rep in signature)

    def __getitem__(self, index):
        return self.signature[index]

    def __len__(self):
        return len(self.signature)

    def __eq__(self, obj):
        """Signatures can be compared to lists or tuples of strings"""

        if isinstance(obj, Signature):
            return self.signature == obj.signature
        elif isinstance(obj, (list, tuple)):
            return self.signature == tuple(obj)
        else:
            return False

    def __ne__(self, obj):
        return not self.__eq__(obj)

    def __repr__(self):
        args = ", ".join(self.signature)

        return "Signature(%s)" % args

    def __str__(self):
         return ", ".join(self.signature)

    def matches(self, *operands):
        """Returns `True` if the signature matches signature of `operands`.
        `operands` is a list of strings.

        Rules:
            * ``rep`` matches ``rep`` and ``*``
            * ``rep[]`` matches ``rep[]`` and ``*[]``

        Example matches:

            * `Signature("sql")` matches ``"sql")``
            * `Signature("*")` matches ``"sql"``
            * `Signature("sql[]")` matches ``"sql[]"``
            * `Signature("*[]")` matches ``"sql[]"``

            * `Signature("sql")` does not match ``"rows"``
            * `Signature("sql")` does not match ``"sql[]"``
        """

        if len(operands) != len(self.operands):
            return False

        operands = [rep_to_operand(rep) for rep in operands]

        for mine, their in zip(self.operands, operands):
            if mine.islist != their.islist:
                return False
            if not mine.isany and mine.rep != their.rep:
                return False

        return True

    def __hash__(self):
        return hash(self.signature)

    def description(self):
        return {"args":[str(s) for s in self.signature]}

    def has_any(self):
        """Returns `True` if at least one operand is `any` (``*``)"""
        return any(op.isany for op in self.operands)

    # TODO: DEPRECIATE!
    def as_prototype(self):
        """Returns a `Signature` object that serves as a prototype for similar
        signatures. All representations in the prototype signature are set to
        `any`, number of arguments is preserved."""

        sig = []

        for op in self.operands:
            if op.islist:
                sig.append("*[]")
            else:
                sig.append("*")
        return Signature(*sig)


def common_representations(*objects):
    """Return list of common representations of `objects`"""

    common = list(objects[0].representations())
    for obj in objects[1:]:
        common = [rep for rep in common if rep in obj.representations()]

    return common

def get_representations(*operands):
    """For every operand get list of it's representations. Returns list of
    lists."""

    reps = []
    # Get representations of objects
    for obj in operands:
        if hasattr(obj, "representations"):
            reps.append(obj.representations())
        elif isinstance(obj, (list, tuple)):
            common = common_representations(*obj)
            common = [sig + "[]" for sig in common]
            reps.append(common)
        else:
            raise ArgumentError("Unknown type of operation argument "\
                                "%s (not a data object)" % type(obj).__name__)

    return reps

class Operation(object):
    def __init__(self, name, operands=None, parameters=None):
        """Creates an operation with name `name` and `operands`. If `operands`
        is `none`, then one operand is assumed with name `obj`.

        The `func` should have a signature compatible with `function(context,
        *operands, **parameters)`
        """

        self.name = name

        if operands is None:
            operands = ["obj"]

        if not operands:
            raise ArgumentError("Operand list sohuld not be empty.")

        self.operands = operands
        self.opcount = len(operands)
        self.parameters = parameters

        self.registry = OrderedDict()

        self.experimental = False

    def __eq__(self, other):
        if not isinstance(other, Operation):
            return False
        return other.name == self.name \
                and other.registry == self.registry

    def __call__(self, *args, **kwargs):
        """Dispatch the operation in the default context."""
        default_context.dispatch(self, *args, **kwargs)

    def signatures(self):
        """Return list of registered signatures."""
        return list(self.registry.keys())

    def function(self, signature):
        """Returns a function for `signature`"""
        return self.registry[signature]

    def resolution_order(self, representations):
        """Returns ordered list of signatures for `operands`. The generic
        signatures (those containing at least one ``*``/`any` type are placed
        at the end of the list.

        Note: The order of the generics is undefined."""

        # TODO: make the order well known, for example by having signatures
        # sortable
        generics = []
        signatures = []

        for sig in self.signatures():
            if sig.has_any():
                generics.append(sig)
            else:
                signatures.append(sig)

        matches = []
        gen_matches = []
        for repsig in itertools.product(*representations):
            matches += [sig for sig in signatures if sig.matches(*repsig)]
            gen_matches += [sig for sig in generics if sig.matches(*repsig)]

        matches += gen_matches

        if not matches:
            raise OperationError("No matching signature found for operation '%s' "
                                 " (args: %s)" %
                                    (self.name, representations))
        return matches


    def register(self, *signature, name=None):
        sig = None

        def register_function(func):
            nonlocal sig
            nonlocal name

            # TODO Test for non-keyword arguments for better error reporting
            func_sig = inspect.signature(func)
            if len(func_sig.parameters) < (1 + self.opcount):
                raise ArgumentError("Expected at least %d arguments in %s. "
                                    "Missing context argument?"
                                    % (self.opcount + 1, func.__name__))

            self.registry[sig] = func
            return func

        if signature and callable(signature[0]):
            func, *signature = signature

            # Create default signature if none is provided
            if not signature:
                signature = ["*"] * self.opcount

            sig = Signature(*signature)
            return register_function(func)
        else:
            sig = Signature(*signature)
            return register_function

    def __str__(self):
        return self.name


def operation(*args):
    """Creates an operation prototype. The operation will have the same name
    as the function. Optionaly a number of operands can be specified. If no
    arguments are given then one operand is assumed."""
    opcount = 1

    def decorator(func):
        nonlocal opcount

        # No arguments, this is the decorator
        # Set default values for the arguments
        # Extract just parameter names (sig.parameters is a mapping)
        sig = inspect.signature(func)
        names = tuple(sig.parameters.keys())
        # Set operand names from function parameter names, skip the context
        # parameter and use only as many parameters as operands in the
        # signature
        operands = names[1:1 + opcount]
        # ... rest of the names are considered operation parameters
        parameters = names[1 + opcount:]
        op = Operation(func.__name__, operands, parameters)
        sig = ["*"] * len(operands)
        op.register(func, *sig)
        return op


    if len(args) == 1 and callable(args[0]):
        return decorator(args[0])
    else:
        # This is just returning the decorator
        opcount = args[0]
        return decorator


########NEW FILE########
__FILENAME__ = audit
# -*- coding: utf-8 -*-

from ..metadata import *
from ..operation import operation
from ..prototypes import *
from ..datautil import guess_type

class BasicAuditProbe(object):
    def __init__(self, key=None, distinct_threshold=10):
        self.field = key
        self.value_count = 0
        self.record_count = 0
        self.value_ratio = 0

        self.distinct_values = set()
        self.distinct_overflow = False
        self.storage_types = set()

        self.null_count = 0
        self.null_value_ratio = 0
        self.null_record_ratio = 0
        self.empty_string_count = 0

        self.min_len = 0
        self.max_len = 0

        self.distinct_threshold = distinct_threshold

        self.unique_storage_type = None

        self.probes = []

    def probe(self, value):
        storage_type = value.__class__
        self.storage_types.add(storage_type.__name__)

        self.value_count += 1

        # FIXME: check for existence in field.empty_values
        if value is None:
            self.null_count += 1

        if value == '':
            self.empty_string_count += 1

        try:
            l = len(value)
            self.min_len = min(self.min_len, l)
            self.max_len = max(self.max_len, l)
        except TypeError:
            pass

        self._probe_distinct(value)

        for probe in self.probes:
            probe.probe(value)

    def _probe_distinct(self, value):
        """"""
        if self.distinct_overflow:
            return

        # We are not testing lists, dictionaries and object IDs
        storage_type = value.__class__

        if not self.distinct_threshold or \
                len(self.distinct_values) < self.distinct_threshold:
            try:
                self.distinct_values.add(value)
            except:
                # FIXME: Should somehow handle invalid values that can not be added
                pass
        else:
            self.distinct_overflow = True

    def finalize(self, record_count = None):
        if record_count:
            self.record_count = record_count
        else:
            self.record_count = self.value_count

        if self.record_count:
            self.value_ratio = float(self.value_count) / float(self.record_count)
            self.null_record_ratio = float(self.null_count) / float(self.record_count)

        if self.value_count:
            self.null_value_ratio = float(self.null_count) / float(self.value_count)

        if len(self.storage_types) == 1:
            self.unique_storage_type = list(self.storage_types)[0]

    def to_dict(self):
        """Return dictionary representation of receiver."""
        d = {
            "field": self.field,
            "value_count": self.value_count,
            "record_count": self.record_count,
            "value_ratio": self.value_ratio,
            "storage_types": list(self.storage_types),
            "null_count": self.null_count,
            "null_value_ratio": self.null_value_ratio,
            "null_record_ratio": self.null_record_ratio,
            "empty_string_count": self.empty_string_count,
            "unique_storage_type": self.unique_storage_type,
            "min_len": self.min_len,
            "max_len": self.max_len
        }

        d["distinct_overflow"] = self.distinct_overflow
        d["distinct_count"] = len(self.distinct_values)
        if self.distinct_overflow:
            d["distinct_values"] = []
        else:
            d["distinct_values"] = list(self.distinct_values)

        return d

@basic_audit.register("rows")
def _(ctx, obj, distinct_threshold=100):

    fields = obj.fields
    out_fields= FieldList(
        Field("field", "string"),
        Field("record_count", "integer"),
        Field("value_ratio", "integer"),
        Field("null_count", "integer"),
        Field("null_value_ratio", "number"),
        Field("null_record_ratio", "number"),
        Field("empty_string_count", "integer"),
        Field("min_len", "integer"),
        Field("max_len", "integer"),
        Field("distinct_count", "integer"),
        Field("distinct_overflow", "boolean")
    )

    stats = []
    for field in fields:
        stat = BasicAuditProbe(field.name, distinct_threshold)
        stats.append(stat)

    for row in obj:
        for i, value in enumerate(row):
            stats[i].probe(value)

    result = []
    for stat in stats:
        stat.finalize()
        if stat.distinct_overflow:
            dist_count = None
        else:
            dist_count = len(stat.distinct_values)

        result.append(stat.to_dict())

    return IterableRecordsDataSource(result, out_fields)

@infer_types.register("rows")
def _(ctx, obj, date_format=None):
    rownum = 0
    probes = defaultdict(set)

    for row in obj:
        rownum += 1
        for i, value in enumerate(row):
            probes[i].add(guess_type(value, date_format))

    keys = list(probes.keys())
    keys.sort()

    types = [probes[key] for key in keys]

    if field_names and len(types) != len(field_names):
        raise Exception("Number of detected fields differs from number"
                        " of fields specified in the header row")
    out_fields = FieldList()
    fields = FieldList(
            Field("field", "string"),
            Field("type", "string")
    )

    result = []
    for name, types in zip(field_names, types):
        if "string" in types:
            t = "string"
        elif "integer"in types:
            t = "integer"
        elif "float" in types:
            t = "float"
        elif "date" in types:
            t = "date"
        else:
            t = "string"
        result.append( (name, t) )

    return IterableDataSource(result, out_fields)

########NEW FILE########
__FILENAME__ = generic
from ..operation import operation
from ..dev import experimental
from ..prototypes import *

#############################################################################
# Convenience

@experimental
@rename_fields.register("*")
def rename_fields(ctx, obj, rename):
    return ctx.op.field_filter(obj, rename=rename)

@experimental
@drop_fields.register("*")
def drop_fields(ctx, obj, drop):
    return ctx.op.field_filter(obj, drop=drop)

@experimental
@keep_fields.register("*")
def keep_fields(ctx, obj, keep):
    return ctx.op.field_filter(obj, keep=keep)

#############################################################################
# Debug

@debug_fields.register("*")
def debug_fields(ctx, obj, label=None):
    if label:
        label = " (%s)" % label
    else:
        label = ""

    ctx.logger.info("fields%s: %s" % (label, obj.fields))
    return obj


########NEW FILE########
__FILENAME__ = rows
# -*- coding: utf-8 -*-
"""Iterator composing operations."""
import itertools
import functools
import operator
import sys
import datetime
from collections import OrderedDict, namedtuple
from ..metadata import *
from ..common import get_logger
from ..errors import *
from ..operation import operation
from ..objects import *
from ..dev import experimental
from ..prototypes import *
from ..datautil import to_bool

from datetime import datetime
from time import strptime
from base64 import b64decode
import json

# FIXME: add cheaper version for already sorted data
# FIXME: BasicAuditProbe was removed

__all__ = ()

def unary_iterator(func):
    """Wraps a function that provides an operation returning an iterator.
    Assumes return fields are the same fields as first argument object"""
    @functools.wraps(func)
    def decorator(ctx, obj, *args, **kwargs):
        result = func(ctx, obj, *args, **kwargs)
        return IterableDataSource(result, obj.fields.clone())

    return decorator

#############################################################################
# Metadata Operations

_default_type_converters = {
    "integer": int,
    "number": float,
    "boolean": to_bool,
    "date": lambda val: datetime.strptime(val, '%Y-%m-%d').date(),
    "time": lambda val: strptime(val, '%H:%M'),
    "datetime": lambda val: datetime.strptime(val, '%Y-%m-%dT%H:%M:%S%Z'),
    "binary": b64decode,
    "object": json.loads,
    "geojson": json.loads,
    "array": ValueError
}

@retype.register("rows")
def _(ctx, obj, typemap):
    def converter(converters, iterator):
        for row in iterator:
            result = []
            for conv, value in zip(converters, row):
                if conv:
                    result.append(conv(value))
                else:
                    result.append(value)
            yield result

    fields = FieldList()
    converters = []
    for field in obj.fields:
        new_type = typemap.get(field.name)
        if new_type and new_type != field.storage_type:
            conv = _default_type_converters[new_type]
        else:
            conv = None
        converters.append(conv)

        field = field.clone(storage_type=new_type or field.storage_type)
        fields.append(field)

    return IterableDataSource(converter(converters, obj), fields)

@field_filter.register("rows")
def _(ctx, iterator, keep=None, drop=None, rename=None, filter=None):
    """Filters fields in `iterator` according to the `field_filter`.
    `iterator` should be a rows iterator and `fields` is list of iterator's
    fields."""
    if filter:
        if keep or drop or rename:
            raise OperationError("Either filter or keep, drop, rename should "
                                 "be used")
        else:
            field_filter = filter
    else:
        field_filter = FieldFilter(keep=keep, drop=drop, rename=rename)

    row_filter = field_filter.row_filter(iterator.fields)

    new_iterator = map(row_filter, iterator)
    new_fields = field_filter.filter(iterator.fields)

    return IterableDataSource(new_iterator, new_fields)


#############################################################################
# Row Operations


@filter_by_value.register("rows")
@unary_iterator
def _(ctx, iterator, key, value, discard=False):
    """Select rows where value of `field` belongs to the set of `values`. If
    `discard` is ``True`` then the matching rows are discarded instead
    (operation is inverted)."""

    fields = iterator.fields
    index = fields.index(str(key))

    if discard:
        predicate = lambda row: row[index] != value
    else:
        predicate = lambda row: row[index] == value

    return filter(predicate, iterator)

@filter_by_set.register("rows")
@unary_iterator
def _(ctx, iterator, field, values, discard=False):
    """Select rows where value of `field` belongs to the set of `values`. If
    `discard` is ``True`` then the matching rows are discarded instead
    (operation is inverted)."""
    fields = iterator.fields
    index = fields.index(field)

    # Convert the values to more efficient set
    values = set(values)

    if discard:
        predicate = lambda row: row[index] not in values
    else:
        predicate = lambda row: row[index] in values

    return filter(predicate, iterator)

@filter_by_range.register("rows")
@unary_iterator
def _(ctx, iterator, field, low, high, discard=False):
    """Select rows where value `low` <= `field` <= `high`. If
    `discard` is ``True`` then the matching rows are discarded instead
    (operation is inverted). To check only agains one boundary set the other
    to ``None``. Equivalend of SQL ``BETWEEN``"""

    fields = iterator.fields
    index = fields.index(field)

    if discard:
        if high is None and low is not None:
            predicate = lambda row: not (low <= row[index])
        elif low is None and high is not None:
            predicate = lambda row: not (row[index] <= high)
        else:
            predicate = lambda row: not (low <= row[index] <= high)
    else:
        if high is None and low is not None:
            predicate = lambda row: low <= row[index]
        elif low is None and high is not None:
            predicate = lambda row: row[index] <= high
        else:
            predicate = lambda row: low <= row[index] <= high

    return filter(predicate, iterator)

@filter_not_empty.register("rows")
@unary_iterator
def _(ctx, iterator, field):
    """Select rows where value of `field` is not None"""

    fields = iterator.fields
    index = fields.index(field)

    predicate = lambda row: row[index] is not None

    return filter(predicate, iterator)

@filter_empty.register("rows")
@unary_iterator
def _(ctx, iterator, field):
    """Select rows where value of `field` is None or empty string"""

    fields = iterator.fields
    index = fields.index(field)

    predicate = lambda row: row[index] is None or row[index] == ""

    return filter(predicate, iterator)

@filter_by_predicate.register("rows")
@unary_iterator
def _(ctx, obj, predicate, fields, discard=False,
                        **kwargs):
    """Returns an interator selecting fields where `predicate` is true.
    `predicate` should be a python callable. `arg_fields` are names of fields
    to be passed to the function (in that order). `kwargs` are additional key
    arguments to the predicate function."""

    def iterator(indexes):
        for row in iter(obj):
            values = [row[index] for index in indexes]
            flag = predicate(*values, **kwargs)
            if (flag and not discard) or (not flag and discard):
                yield row

    key = prepare_key(fields)
    indexes = obj.fields.indexes(key)

    return iterator(indexes)


@filter_by_predicate.register("records")
def _(ctx, iterator, predicate, fields, discard=False,
                        **kwargs):
    """Returns an interator selecting fields where `predicate` is true.
    `predicate` should be a python callable. `arg_fields` are names of fields
    to be passed to the function (in that order). `kwargs` are additional key
    arguments to the predicate function."""

    for record in iterator:
        args = [record[str(f)] for f in fields]
        flag = predicate(*args, **kwargs)
        if (flag and not discard) or (not flag and discard):
            yield record


@distinct.register("rows")
def _(ctx, obj, key=None, is_sorted=False):
    """Return distinct `keys` from `iterator`. `iterator` does
    not have to be sorted. If iterator is sorted by the keys and
    `is_sorted` is ``True`` then more efficient version is used."""
    # TODO: remove is_sorted hint, objects should store metadata about
    # that

    def iterator(row_filter):
        if is_sorted:
            last_key = object()

            # FIXME: use itertools equivalent
            for row in obj:
                key_tuple = (row_filter(row))
                if key_tuple != last_key:
                    yield key_tuple

        else:
            distinct_values = set()
            for row in obj:
                # Construct key tuple from distinct fields
                key_tuple = tuple(row_filter(row))
                if key_tuple not in distinct_values:
                    distinct_values.add(key_tuple)
                    yield key_tuple

    fields = obj.fields
    if key:
        key = prepare_key(key)
        row_filter = FieldFilter(keep=key).row_filter(fields)
    else:
        row_filter = FieldFilter().row_filter(fields)

    # Retain original order of fields
    fields = FieldList(*row_filter(obj.fields))

    return IterableDataSource(iterator(row_filter), fields)


@distinct_rows.register("rows")
@unary_iterator
def _(ctx, obj, key=None, is_sorted=False):
    """Return distinct rows based on `key` from `iterator`. `iterator`
    does not have to be sorted. If iterator is sorted by the keys and
    `is_sorted` is ``True`` then more efficient version is used."""
    # TODO: remove is_sorted hint, objects should store metadata about
    # that

    fields = obj.fields
    if key:
        key = prepare_key(key)
        row_filter = FieldFilter(keep=key).row_filter(fields)
    else:
        row_filter = FieldFilter().row_filter(fields)

    if is_sorted:
        last_key = object()

        # FIXME: use itertools equivalent
        for value in obj:
            key_tuple = (row_filter(row))
            if key_tuple != last_key:
                yield row

    else:
        distinct_values = set()
        for row in obj:
            # Construct key tuple from distinct fields
            key_tuple = tuple(row_filter(row))
            if key_tuple not in distinct_values:
                distinct_values.add(key_tuple)
                yield row


@first_unique.register("rows")
@unary_iterator
def _(ctx, iterator, keys=None, discard=False):
    """Return rows that are unique by `keys`. If `discard` is `True` then the
    action is reversed and duplicate rows are returned."""

    # FIXME: add is_sorted version
    # FIXME: use prepare key

    row_filter = FieldFilter(keep=keys).row_filter(iterator.fields)

    distinct_values = set()

    for row in iterator:
        # Construct key tuple from distinct fields
        key_tuple = tuple(row_filter(row))

        if key_tuple not in distinct_values:
            distinct_values.add(key_tuple)
            if not discard:
                yield row
        else:
            if discard:
                # We already have one found record, which was discarded
                # (because discard is true), now we pass duplicates
                yield row


@sample.register("rows")
@unary_iterator
def _(ctx, iterator, value, discard=False, mode="first"):
    """Returns sample from the iterator. If `mode` is ``first`` (default),
    then `value` is number of first records to be returned. If `mode` is
    ``nth`` then one in `value` records is returned."""

    if mode == "first":
        if discard:
            return itertools.islice(iterator, value, None)
        else:
            return itertools.islice(iterator, value)
    elif mode == "nth":
        if discard:
            return discard_nth(iterator, value)
        else:
            return itertools.islice(iterator, None, None, value)
    elif mode == "random":
        raise NotImplementedError("random sampling is not yet implemented")
    else:
        raise Exception("Unknown sample mode '%s'" % mode)


@unary_iterator
def discard_nth_base(ctx, iterator, step):
    """Discards every step-th item from `iterator`"""
    for i, value in enumerate(iterator):
        if i % step != 0:
            yield value

discard_nth.register("rows")(discard_nth_base)
discard_nth.register("records")(discard_nth_base)

@sort.register("rows")
@unary_iterator
def _(ctx, obj, orderby):
    iterator = obj.rows()

    orderby = prepare_order_list(orderby)

    for field, order in reversed(orderby):
        index = obj.fields.index(field)

        if order.startswith("asc"):
            reverse = False
        elif order.startswith("desc"):
            reverse = True
        else:
            raise ValueError("Unknown order %s for column %s") % (order, column)

        iterator = sorted(iterator, key=operator.itemgetter(index),
                                    reverse=reverse)

    return iterator


###
# Simple and naive aggregation in Python

def agg_sum(a, value):
    return a+value

def agg_average(a, value):
    return (a[0]+1, a[1]+value)

def agg_average_finalize(a):
    return a[1]/a[0]

AggregationFunction = namedtuple("AggregationFunction",
                            ["func", "start", "finalize"])
aggregation_functions = {
            "sum": AggregationFunction(agg_sum, 0, None),
            "min": AggregationFunction(min, 0, None),
            "max": AggregationFunction(max, 0, None),
            "average": AggregationFunction(agg_average, (0,0), agg_average_finalize)
        }

@aggregate.register("rows")
def _(ctx, obj, key, measures=None, include_count=True,
      count_field="record_count"):
    """Aggregates measure fields in `iterator` by `keys`. `fields` is a field
    list of the iterator, `keys` is a list of fields that will be used as
    keys. `aggregations` is a list of measures to be aggregated.

    `measures` should be a list of tuples in form (`measure`, `aggregate`).
    See `distill_measure_aggregates()` for how to convert from arbitrary list
    of measures into this form.

    Output of this iterator is an iterator that yields rows with fields that
    contain: key fields, measures (as specified in the measures list) and
    optional record count if `include_count` is ``True`` (default).

    Result is not ordered even the input was ordered.

    .. note:

        This is naïve, pure Python implementation of aggregation. Might not
        fit your expectations in regards of speed and memory consumption for
        large datasets.
    """

    def aggregation_result(keys, aggregates, measure_aggregates):
        # Pass results to output
        for key in keys:
            row = list(key[:])

            key_aggregate = aggregates[key]
            for i, (measure, index, function) in enumerate(measure_aggregates):
                aggregate = key_aggregate[i]
                finalize = aggregation_functions[function].finalize
                if finalize:
                    row.append(finalize(aggregate))
                else:
                    row.append(aggregate)

            if include_count:
                row.append(key_aggregate[-1])

            yield row

    # TODO: create sorted version
    # TODO: include SQL style COUNT(field) to count non-NULL values

    # Coalesce to a list if just one is specified
    keys = prepare_key(key)

    measures = prepare_aggregation_list(measures)

    # Prepare output fields
    out_fields = FieldList()
    out_fields += obj.fields.fields(keys)

    measure_fields = set()
    measure_aggregates = []
    for measure in measures:
        name = measure[0]
        index = obj.fields.index(name)
        aggregate = measure[1]

        measure_aggregates.append( (name, index, aggregate) )
        measure_fields.add(name)

        field = obj.fields.field(name)
        field = field.clone(name="%s_%s" % (name, aggregate),
                            analytical_type="measure")
        out_fields.append(field)

    if include_count:
        out_fields.append(Field(count_field,
                            storage_type="integer",
                            analytical_type="measure"))

    if keys:
        key_selectors = obj.fields.indexes(keys)
    else:
        key_selectors = []

    keys = set()

    # key -> list of aggregates
    aggregates = {}

    for row in obj.rows():
        # Create aggregation key
        key = tuple(row[s] for s in key_selectors)

        # Create new aggregate record for key if it does not exist
        #
        try:
            key_aggregate = aggregates[key]
        except KeyError:
            keys.add(key)
            key_aggregate = []
            for measure, index, function in measure_aggregates:
                start = aggregation_functions[function].start
                key_aggregate.append(start)
            if include_count:
                key_aggregate.append(0)

            aggregates[key] = key_aggregate

        for i, (measure, index, function) in enumerate(measure_aggregates):
            func = aggregation_functions[function].func
            key_aggregate[i] = func(key_aggregate[i], row[index])

        if include_count:
            key_aggregate[-1] += 1

    iterator = aggregation_result(keys, aggregates, measure_aggregates)

    return IterableDataSource(iterator, out_fields)


#############################################################################
# Transpose

@transpose_by.register("rows")
def _(ctx, obj, key, column_field, value_field):

    def iterator():
        nonlocal keep_filter, transpose_filter, transpose_names

        for row in obj:
            keep = keep_filter(row)
            transposed = transpose_filter(row)

            for name, value in zip(transpose_names, transposed):
                yield list(keep) + [name, value]

    key = prepare_key(key)
    out_fields = FieldList()
    out_fields += obj.fields.fields(key)

    # TODO: set type of the field as type of the first transposed field
    out_fields.append(Field(column_field, "string"))
    out_fields.append(Field(value_field))

    keep_filter = FieldFilter(keep=key).row_filter(obj.fields)
    transpose_filter = FieldFilter(drop=key).row_filter(obj.fields)
    transpose_names = transpose_filter(obj.fields.names())

    return IterableDataSource(iterator(), out_fields)

#############################################################################
# Field Operations


@append_constant_fields.register("rows")
def _(ctx, obj, fields, value):
    def iterator(constants):
        for row in obj.rows():
            yield list(row) + constants

    if not isinstance(value, (list, tuple)):
        constants = (value, )
    else:
        constants = value

    output_fields = obj.fields + fields

    return IterableDataSource(iterator(constants), output_fields)


@dates_to_dimension.register("rows")
def _(ctx, obj, fields=None, unknown_date=0):
    def iterator(indexes):
        for row in obj.rows():
            row = list(row)

            for index in indexes:
                row[index] = row[index].strftime("%Y%m%d")

            yield row

    if fields:
        date_fields = obj.fields(fields)
    else:
        date_fields = obj.fields.fields(storage_type="date")

    indexes = obj.fields.indexes(date_fields)

    fields = []
    for field in obj.fields:
        if field in date_fields:
            fields.append(field.clone(storage_type="integer",
                                      concrete_storage_type=None))
        else:
            fields.append(field.clone())

    fields = FieldList(*fields)

    return IterableDataSource(iterator(indexes), fields)


@string_to_date.register("rows")
def _(ctx, obj, fields, fmt="%Y-%m-%dT%H:%M:%S.Z"):
    def iterator(indexes):
        for row in obj.rows():
            row = list(row)
            for index in indexes:
                date_str = row[index]
                value = None
                if date_str:
                    try:
                        value = datetime.datetime.strptime(row[index], fmt)
                    except ValueError:
                        pass

                row[index] = value
            yield row

    date_fields = prepare_key(fields)
    indexes = obj.fields.indexes(date_fields)

    # Prepare output fields
    fields = FieldList()
    for field in obj.fields:
        if str(field) in date_fields:
            fields.append(field.clone(storage_type="date",
                                      concrete_storage_type=None))
        else:
            fields.append(field.clone())

    return IterableDataSource(iterator(indexes), fields)

@split_date.register("rows")
def _(ctx, obj, fields, parts=["year", "month", "day"]):
    """Extract `parts` from date objects"""

    def iterator(indexes):
        for row in obj.rows():
            new_row = []
            for i, value in enumerate(row):
                if i in indexes:
                    for part in parts:
                        new_row.append(getattr(value, part))
                else:
                    new_row.append(value)

            yield new_row

    date_fields = prepare_key(fields)

    indexes = obj.fields.indexes(date_fields)

    # Prepare output fields
    fields = FieldList()
    proto = Field(name="p", storage_type="integer", analytical_type="ordinal")

    for field in obj.fields:
        if str(field) in date_fields:
            for part in parts:
                name = "%s_%s" % (str(field), part)
                fields.append(proto.clone(name=name))
        else:
            fields.append(field.clone())

    return IterableDataSource(iterator(indexes), fields)

@text_substitute.register("rows")
@unary_iterator
def _(ctx, iterator, field, substitutions):
    """Substitute field using text substitutions"""
    # Compile patterns
    fields = iterator.fields
    substitutions = [(re.compile(patt), r) for (patt, r) in subsitutions]
    index = fields.index(field)
    for row in iterator:
        row = list(row)

        value = row[index]
        for (pattern, repl) in substitutions:
            value = re.sub(pattern, repl, value)
        row[index] = value

        yield row

@empty_to_missing.register("rows")
@unary_iterator
@experimental
def _(ctx, iterator, fields=None, strict=False):
    """Converts empty strings into `None` values."""
    if fields:
        if strict:
            fields = iterator.fields.fields(fields)
        else:
            array = []
            for field in fields:
                if field in iterator.fields:
                    array.append(field)
            fields = array
    else:
        fields = iterator.fields.fields(storage_type="string")

    indexes = iterator.fields.indexes(fields)

    for row in iterator:
        row = list(row)
        for index in indexes:
            row[index] = row[index] if row[index] else None
        yield row

@string_strip.register("rows")
@unary_iterator
def _(ctx, iterator, strip_fields=None, chars=None):
    """Strip characters from `strip_fields` in the iterator. If no
    `strip_fields` is provided, then it strips all `string` or `text` storage
    type objects."""

    fields = iterator.fields
    if not strip_fields:
        strip_fields = []
        for field in fields:
            if field.storage_type =="string" or field.storage_type == "text":
                strip_fields.append(field)

    indexes = fields.indexes(strip_fields)

    for row in iterator:
        row = list(row)
        for index in indexes:
            value = row[index]
            if value:
                row[index] = value.strip(chars)
        yield row

@string_split_fixed.register("rows")
def _(ctx, iterato, split_fields=None, new_fields=None, widths=None):

    raise NotImplemented
#############################################################################
# Compositions


@append.register("rows[]")
def _(ctx, objects):
    """Appends iterators"""
    # TODO: check for field equality
    iterators = [iter(obj) for obj in objects]
    iterator = itertools.chain(*iterators)

    return IterableDataSource(iterator, objects[0].fields)



@join_details.register("rows", "rows")
def _(self, master, detail, master_key, detail_key):
    """"Simple master-detail join"""

    def _join_detail_iterator(master, detail, master_key, detail_key):
        """Simple iterator implementation of the left inner join"""

        djoins = []
        detail_map = {}

        # TODO: support compound keys
        detail_index = detail.fields.index(detail_key[0])
        detail_dict = {}

        for row in detail:
            row = list(row)
            key = row.pop(detail_index)
            detail_dict[key] = row

        for master_row in master:
            row = list(master_row)

            master_index = master.fields.index(master_key[0])
            key = master_row[master_index]
            try:
                detail_row = detail_dict[key]
            except KeyError:
                continue
            else:
                yield row + detail_row

    master_key = prepare_key(master_key)
    detail_key = prepare_key(detail_key)

    if len(master_key) > 1 or len(detail_key) > 1:
        raise ArgumentError("Compound keys are not supported yet")

    result = _join_detail_iterator(master, detail, master_key, detail_key)

    # Prepare output fields and columns selection - the selection skips detail
    # columns that are used as key, because they are already present in the
    # master table.

    out_fields = master.fields.clone()
    for field in detail.fields:
        if str(field) not in detail_key:
            out_fields.append(field)

    return IterableDataSource(result, out_fields)


#############################################################################
# Output


@pretty_print.register("records")
def _(ctx, obj, target=None):
    if not target:
        target = sys.stdout

    names = obj.fields.names()
    widths = [len(field.name) for field in obj.fields]
    # Consume data to be pretty-printed
    text_rows = []
    for row in obj.rows():
        line = [str(value) for value in row]
        widths = [max(w, len(val)) for w,val in zip(widths, line)]
        text_rows.append(line)

    format_str = "|"
    for i, field in enumerate(obj.fields):
        width = widths[i]
        if field.storage_type in ("integer", "float"):
            fmt = ">%d" % width
        else:
            fmt = "<%d" % width

        format_str += "{%d:%s}|" % (i, fmt)

    field_borders = [u"-"*w for w in widths]
    border = u"+" + u"+".join(field_borders) + u"+\n"

    format_str += "\n"

    target.write(border)
    header = format_str.format(*names)
    target.write(header)
    target.write(border)

    for row in text_rows:
        target.write(format_str.format(*row))
    target.write(border)

    target.flush()

# def threshold(value, low, high, bins=None):
#     """Returns threshold value for `value`. `bins` should be names of bins. By
#     default it is ``['low', 'medium', 'high']``
#     """
#
#     if not bins:
#         bins = ['low', 'medium', 'high']
#     elif len(bins) != 3:
#         raise Exception("bins should be a list of three elements")
#
#     if low is None and high is None:
#         raise Exception("low and hight threshold values should not be "
#                         "both none at the same time.")
#


#############################################################################
# Conversions


@as_records.register("rows")
@unary_iterator
def _(ctx, obj):
    """Returns iterator of dictionaries where keys are defined in
    fields."""
    names = [str(field) for field in obj.fields]
    for row in obj:
        yield dict(zip(names, row))


@fetch_all.register("rows")
def _(ctx, obj):
    """Loads all data from the iterable object and stores them in a python
    list. Useful for smaller datasets, not recommended for big data."""

    data = list(obj)

    return RowListDataObject(data, fields=obj.fields)


@as_dict.register("rows")
def _(ctx, obj, key=None, value=None):
    """Returns dictionary constructed from the iterator.  `key` is name of a
    field or list of fields that will be used as a simple key or composite
    key. `value` is a field or list of fields that will be used as values.

    If no `key` is provided, then the first field is used as key. If no
    `value` is provided, then whole rows are used as values.

    Keys are supposed to be unique. If they are not, result might be
    unpredictable.

    .. warning::

        This method consumes whole iterator. Might be very costly on large
        datasets.
    """

    fields = obj.fields

    if not key:
        index = 0
        indexes = None
    elif isinstance(key, (str, Field)):
        index = fields.index(key)
        indexes = None
    else:
        indexes = fields.indexes(key)

    if not value:
        if indexes is None:
            d = dict( (row[index], row) for row in obj)
        else:
            for row in obj:
                key_value = (row[index] for index in indexes)
                d[key_value] = row
    elif isinstance(value, (str, Field)):
        value_index = fields.index(value)
        if indexes is None:
            d = dict( (row[index], row[value_index]) for row in obj)
        else:
            for row in obj:
                key_value = (row[index] for index in indexes)
                d[key_value] = row[value_index]
    else:
        raise NotImplementedError("Specific composite value is not implemented")
    return d



########NEW FILE########
__FILENAME__ = prototypes
from .operation import operation

# Operation prototypes – empty operations

@operation
def retype(ctx, iterator, typemap):
    raise NotImplementedError

#############################################################################
# Metadata Operations

@operation
def field_filter(ctx, iterator, keep=None, drop=None, rename=None,
                 filter=None):
    raise NotImplementedError

@operation
def rename_fields(ctx, obj, rename):
    raise NotImplementedError

@operation
def drop_fields(ctx, obj, drop):
    raise NotImplementedError

@operation
def keep_fields(ctx, obj, keep):
    raise NotImplementedError

@operation
def debug_fields(ctx, obj, label=None):
    raise NotImplementedError

#############################################################################
# Row Filters

@operation
def filter_by_value(ctx, iterator, key, value, discard=False):
    raise NotImplementedError

@operation
def filter_by_set(ctx, iterator, field, values, discard=False):
    raise NotImplementedError

@operation
def filter_by_range(ctx, iterator, field, low, high, discard=False):
    raise NotImplementedError

@operation
def filter_not_empty(ctx, iterator, field):
    raise NotImplementedError

@operation
def filter_empty(ctx, iterator, field):
    raise NotImplementedError

@operation
def filter_by_predicate(ctx, obj, predicate, fields, discard=False,
                        **kwargs):
    raise NotImplementedError

@operation
def distinct(ctx, obj, key=None, is_sorted=False):
    raise NotImplementedError

@operation
def distinct_rows(ctx, obj, key=None, is_sorted=False):
    raise NotImplementedError

@operation
def first_unique(ctx, iterator, keys=None, discard=False):
    raise NotImplementedError

@operation
def sample(ctx, iterator, value, discard=False, mode="first"):
    raise NotImplementedError

@operation
def discard_nth(ctx, iterator, step):
    raise NotImplementedError


#############################################################################
# Ordering

@operation
def sort(ctx, obj, orderby):
    raise NotImplementedError


#############################################################################
# Aggregate

@operation
def aggregate(ctx, obj, key, measures=None, include_count=True,
      count_field="record_count"):
    raise NotImplementedError


#############################################################################
# Field Operations

@operation
def append_constant_fields(ctx, obj, fields, value):
    raise NotImplementedError

@operation
def dates_to_dimension(ctx, obj, fields=None, unknown_date=0):
    raise NotImplementedError

@operation
def string_to_date(ctx, obj, fields, fmt="%Y-%m-%dT%H:%M:%S.Z"):
    raise NotImplementedError

@operation
def split_date(ctx, obj, fields, parts=["year", "month", "day"]):
    raise NotImplementedError

@operation
def text_substitute(ctx, iterator, field, substitutions):
    raise NotImplementedError

@operation
def empty_to_missing(ctx, iterator, fields=None, strict=False):
    raise NotImplementedError

@operation
def string_strip(ctx, iterator, strip_fields=None, chars=None):
    """Strip characters from `strip_fields` in the iterator. If no
    `strip_fields` is provided, then it strips all `string` or `text` storage
    type objects."""

    raise NotImplementedError

@operation
def string_split_fixed(ctx, iterator, split_fields=None,
                       new_fields=None, widths=None):
    raise NotImplementedError

@operation
def transpose_by(ctx, iterator, key, new_field):
    raise NotImplementedError


#############################################################################
# Compositions

@operation
def append(ctx, objects):
    raise NotImplementedError

@operation(2)
def join_details(self, master, detail, master_key, detail_key):
    raise NotImplementedError


#############################################################################
# Comparison and Inspection

@operation(2)
def added_keys(self, master, detail, src_key, target_key):
    raise NotImplementedError

@operation(2)
def added_rows(self, master, detail, src_key, target_key):
    raise NotImplementedError

@operation(2)
def changed_rows(self, master, detail, dim_key, source_key, fields,
                    version_field):
    raise NotImplementedError

@operation
def count_duplicates(ctx, obj, keys=None, threshold=1,
                     record_count_label="record_count"):
    raise NotImplementedError

@operation
def duplicate_stats(ctx, obj, fields=None, threshold=1):
    raise NotImplementedError

@operation
def nonempty_count(ctx, obj, fields=None):
    raise NotImplementedError

@operation
def distinct_count(ctx, obj, fields=None):
    raise NotImplementedError

#############################################################################
# Audit

@operation
def basic_audit(ctx, iterable, distinct_threshold):
    raise NotImplementedError

@operation
def infer_types(ctx, iterable, date_format=None):
    raise NotImplementedError


#############################################################################
# Loading

@operation(2)
def insert(ctx, source, target):
    raise NotImplementedError

@operation(2)
def load_versioned_dimension(ctx, dim, source, dim_key, fields,
                             version_fields=None, source_key=None):
    raise NotImplementedError


#############################################################################
# Misc

@operation
def pretty_print(ctx, obj, target=None):
    raise NotImplementedError

#############################################################################
# Conversion

@operation
def as_records(ctx, obj):
    raise NotImplementedError

@operation
def fetch_all(ctx, obj):
    raise NotImplementedError

@operation
def as_dict(ctx, obj, key=None, value=None):
    raise NotImplementedError

#############################################################################
# Assertions

@operation
def assert_unique(ctx, obj, key=None):
    raise NotImplementedError

@operation
def assert_contains(ctx, obj, field, value):
    raise NotImplementedError

@operation
def assert_missing(ctx, obj, field, value):
    raise NotImplementedError

@operation
def _(ctx, obj, field, value):
    raise NotImplementedError

@operation
def _(ctx, obj, field, value):
    raise NotImplementedError



########NEW FILE########
__FILENAME__ = resource
# -*- coding: utf-8 -*-

from contextlib import ContextDecorator

import urllib.request
import urllib.parse
import codecs
import json

__all__ = (
    "Resource",
    "is_local",
    "read_json",
)


class Resource(ContextDecorator):
    def __init__(self, url=None, handle=None, opener=None, encoding=None,
                 binary=False):
        """Creates a data resource for reading. Arguments:

        * `url` – resource URL or a local path
        * `handle` – handle of an opened file. If provided, then the resource
          `url` is just informative (for display or debugging purposes) and
          should correspond to the opened handle.
        * `opener` – a function that opens the URL
        * `encoding` – encoding used on local files or URLs with default URL
          opener
        * `binary` – `True` if the resource is binary, `False` (default) if it
          is a text

        The resource can be used as a context manager: `with Resource(url) as
        f: ...`.

        """

        if url is None and handle is None:
            raise ArgumentError("Either resource url or handle should be "
                                "provided. None was given.")

        self.url = url
        self.binary = binary
        self.encoding = encoding

        self.reader = None

        if not opener:
            if is_local(url):
                self.opener = None
            else:
                self.opener = urllib.request.urlopen
                if self.encoding:
                    self.reader = codecs.getreader(self.encoding)
        else:
            self.opener = opener

        self.handle = handle
        self.should_close = handle is not None

    def open(self):
        if self.handle:
            return self.handle

        if self.opener:
            self.handle = self.opener(self.url)
        else:
            mode = "rb" if self.binary else "r"
            self.handle = open(self.url, mode=mode, encoding=self.encoding)
        if self.reader:
            self.handle = self.reader(self.handle)

        return self.handle

    def close(self):
        if self.should_close:
            self.handle.close()

    def __enter__(self):
        return self.open()

    def __exit__(self, *exc):
        self.close()


def is_local(url):
    """Returns path to a local file from `url`. Returns `None` if the `url`
    does not represent a local file."""
    parts = urllib.parse.urlparse(url)
    return parts.scheme == '' or parts.scheme == 'file'

def open_resource(resource, mode=None, encoding=None, binary=False):
    raise NotImplementedError("open_resource is depreciated, use Resource")

def read_json(url):
    """Reads JSON from `url`. The `url` can be also a local file path."""
    with Resource(url) as f:
        try:
            data = json.load(f)
        except Exception as e:
            raise Exception("Unable to read JSON from %s: %s"
                            % (url, str(e)))
    return data

########NEW FILE########
__FILENAME__ = stores
# -*- Encoding: utf8 -*-

from .errors import *
from .metadata import *
from .extensions import Extensible, extensions
from .objects import data_object
import os.path

__all__ = [
        "DataStore",
        "SimpleMemoryStore",
        "open_store",
        "copy_object"
        ]


def open_store(type_, *args, **kwargs):
    """Opens datastore of `type`."""

    store = extensions.store(type_, *args, **kwargs)
    return store


class DataStore(Extensible):
    __extension_type__ = "store"
    __extension_suffix__ = "Store"
    def __init__(self, **options):
        pass

    def close(self):
        pass

    def clone(self, *args, **kwargs):
        """Returns a clone of the store with different options. Objects coming
        from cloned store might be composable with objects in the original
        store."""
        raise NotImplementedError

    def object_names(self):
        """Returns list of all object names contained in the store"""
        raise NotImplementedError

    def objects(self, names=None, autoload=False):
        """Return list of objects, if available

        * `names`: only objects with given names are returned
        * `autoload`: load object list if necessary, otherwise cached version
          is used if store cachces object metadata.

        Note that loading list of objects might be costly operation in some
        cases.
        """
        raise NotImplementedError

    def get_object(self, name, **args):
        """Subclasses should implement this"""
        raise NotImplementedError

    def __getitem__(self, name):
        return self.get_object(name)

    def create(self, name, fields, replace=False, from_obj=None, temporary=False,
               **options):
        """Args:
            * replace
            * form_obj: object from which the target is created
            * temporary: table is destroyed after store is closed or
              disconnected
        """
        pass

    def exists(self, name):
        """Return `True` if object with `name` exists, otherwise returns
        `False`. Subclasses should implement this method."""
        raise NotImplementedError

    def create_temporary(fields, from_obj=None, **options):
        """Creates a temporary data object"""
        raise NotImplementedError

    def truncate(self, name, *args, **kwargs):
        obj = self.get_object(name, *args, **kwargs)
        obj.truncate()

    def rename(name, new_name, force=False):
        """Renames object from `name` to `new_name`. If `force` is ``True``
        then target is lost"""
        raise NotImplementedError


class SimpleMemoryStore(DataStore):
    def __init__(self):
        """Creates simple in-memory data object store. Useful for temporarily
        store objects. Creates list based objects with `rows` and `records`
        representations."""

        super(SimpleMemoryStore, self).__init__()
        catalogue = {}

    def objects(self):
        return list(catalogue.keys())

    def get_object(self, name):
        try:
            return catalogue[name]
        except KeyError:
            raise NoSuchObjectError(name)

    def create(name, fields, replace=False, from_obj=None, temporary=False,
               **options):
        """Creates and returns a data object that wraps a Python list as a
        data container."""

        if not replace and self.exists(name):
            raise ObjectExistsError(name)

        obj = RowListDataObject(fields)
        catalogue[name] = obj
        return obj

    def exists(name):
        return name in catalogue


class FileSystemStore(DataStore):
    __identifier__ = "file"

    def __init__(self, path):
        """Creates a store for source objects stored on a local file system.
        The type of the object is determined from the file extension.
        Supported extensions and file types:

        * `csv` - CSV source object (read-only)
        * `xls` – MS Excel object
        """

        super().__init__()
        self.path = path

    def get_object(self, name):
        """Returns a CSVSource object with filename constructed from store's
        path and extension"""
        path = os.path.join(self.path, name)
        ext = os.path.splitext(name)[1]

        ext = ext[1:] if ext else ext

        if ext == "csv":
            return data_object("csv_source", path)
        elif ext == "xls":
            return data_object("xls", path)
        else:
            raise ArgumentError("Unknown extension '%s'" % ext)


def copy_object(source_store, source_name, target_store,
                target_name=None, create=False, replace=False):
    """Convenience method that copies object data from source store to target
    store. `source_object` and `target_object` should be object names within
    the respective stores. If `target_name` is not specified, then
    `source_name` is used."""

    target_name = target_name or source_name

    source = source_store.get_object(source_name)
    if create:
        if not replace and target_store.exists(target_name):
            raise Exception("Target object already exists. Use reaplce=True to "
                            "delete the object object and create replacement")
        target = target_store.create(target_name, source.fields, replace=True,
                                     from_obj=source)
    else:
        target = target_store.get_object(target_name)
        target.append_from(source)
        target.flush()

    return target


########NEW FILE########
__FILENAME__ = threadlocal
# -*- Encoding: utf8 -*-
import threading

thread_locals = threading.local()

# Modified version of LocalProxy from werkzeug by mitsuhiko

class LocalProxy(object):
    """Acts as a proxy for a local.  Forwards all operations to a proxied
    object.  The only operations not supported for forwarding are right handed
    operands and any kind of assignment.
    """

    def __init__(self, name, factory=None):
        object.__setattr__(self, '_local_name', name)
        object.__setattr__(self, '_locals', thread_locals)
        object.__setattr__(self, '_proxy_factory', factory)

    def _represented_local_object(self):
        """Return the represented thread local object. If object does not
        exist, try to create it using factory.
        """
        try:
            return getattr(self._locals, self._local_name)
        except AttributeError:
            if self._proxy_factory:
                setattr(self._locals, self._local_name, self._proxy_factory())
                return getattr(self._locals, self._local_name)
            else:
                raise RuntimeError("No thread local '%s' and no factory "
                                    "provided" % self._local_name)

    @property
    def __idict__(self):
        try:
            return self._represented_local_object().__dict__
        except RuntimeError:
            raise AttributeError('__dict__')

    def __repr__(self):
        try:
            obj = self._represented_local_object()
        except RuntimeError:
            return '<%s unbound>' % self.__class__.__name__
        return repr(obj)

    def __bool__(self):
        try:
            return bool(self._represented_local_object())
        except RuntimeError:
            return False

    def __unicode__(self):
        try:
            return unicode(self._represented_local_object())
        except RuntimeError:
            return repr(self)

    def __dir__(self):
        try:
            return dir(self._represented_local_object())
        except RuntimeError:
            return []

    def __getattr__(self, name):
        if name == '__members__':
            return dir(self._represented_local_object())
        return getattr(self._represented_local_object(), name)

    def __setitem__(self, key, value):
        self._represented_local_object()[key] = value

    def __delitem__(self, key):
        del self._represented_local_object()[key]

    __setattr__ = lambda x, n, v: setattr(x._represented_local_object(), n, v)
    __delattr__ = lambda x, n: delattr(x._represented_local_object(), n)
    __str__ = lambda x: str(x._represented_local_object())
    __lt__ = lambda x, o: x._represented_local_object() < o
    __le__ = lambda x, o: x._represented_local_object() <= o
    __eq__ = lambda x, o: x._represented_local_object() == o
    __ne__ = lambda x, o: x._represented_local_object() != o
    __gt__ = lambda x, o: x._represented_local_object() > o
    __ge__ = lambda x, o: x._represented_local_object() >= o
    __cmp__ = lambda x, o: cmp(x._represented_local_object(), o)
    __hash__ = lambda x: hash(x._represented_local_object())
    __call__ = lambda x, *a, **kw: x._represented_local_object()(*a, **kw)
    __len__ = lambda x: len(x._represented_local_object())
    __getitem__ = lambda x, i: x._represented_local_object()[i]
    __iter__ = lambda x: iter(x._represented_local_object())
    __contains__ = lambda x, i: i in x._represented_local_object()
    __add__ = lambda x, o: x._represented_local_object() + o
    __sub__ = lambda x, o: x._represented_local_object() - o
    __mul__ = lambda x, o: x._represented_local_object() * o
    __floordiv__ = lambda x, o: x._represented_local_object() // o
    __mod__ = lambda x, o: x._represented_local_object() % o
    __divmod__ = lambda x, o: x._represented_local_object().__divmod__(o)
    __pow__ = lambda x, o: x._represented_local_object() ** o
    __lshift__ = lambda x, o: x._represented_local_object() << o
    __rshift__ = lambda x, o: x._represented_local_object() >> o
    __and__ = lambda x, o: x._represented_local_object() & o
    __xor__ = lambda x, o: x._represented_local_object() ^ o
    __or__ = lambda x, o: x._represented_local_object() | o
    __div__ = lambda x, o: x._represented_local_object().__div__(o)
    __truediv__ = lambda x, o: x._represented_local_object().__truediv__(o)
    __neg__ = lambda x: -(x._represented_local_object())
    __pos__ = lambda x: +(x._represented_local_object())
    __abs__ = lambda x: abs(x._represented_local_object())
    __invert__ = lambda x: ~(x._represented_local_object())
    __complex__ = lambda x: complex(x._represented_local_object())
    __int__ = lambda x: int(x._represented_local_object())
    __long__ = lambda x: long(x._represented_local_object())
    __float__ = lambda x: float(x._represented_local_object())
    __oct__ = lambda x: oct(x._represented_local_object())
    __hex__ = lambda x: hex(x._represented_local_object())
    __index__ = lambda x: x._represented_local_object().__index__()
    __coerce__ = lambda x, o: x._represented_local_object().__coerce__(x, o)
    __enter__ = lambda x: x._represented_local_object().__enter__()
    __exit__ = lambda x, *a, **kw: x._represented_local_object().__exit__(*a, **kw)
    __radd__ = lambda x, o: o + x._represented_local_object()
    __rsub__ = lambda x, o: o - x._represented_local_object()
    __rmul__ = lambda x, o: o * x._represented_local_object()
    __rdiv__ = lambda x, o: o / x._represented_local_object()
    __rtruediv__ = __rdiv__
    __rfloordiv__ = lambda x, o: o // x._represented_local_object()
    __rmod__ = lambda x, o: o % x._represented_local_object()
    __rdivmod__ = lambda x, o: x._represented_local_object().__rdivmod__(o)


########NEW FILE########
__FILENAME__ = conf
# -*- coding: utf-8 -*-
#
# Bubbles documentation build configuration file, created by
# sphinx-quickstart on Thu Apr 25 01:42:53 2013.
#
# This file is execfile()d with the current directory set to its containing dir.
#
# Note that not all possible configuration values are present in this
# autogenerated file.
#
# All configuration values have a default; values that are commented out
# serve to show the default.

import sys, os

# If extensions (or modules to document with autodoc) are in another directory,
# add these directories to sys.path here. If the directory is relative to the
# documentation root, use os.path.abspath to make it absolute, like shown here.
sys.path.insert(0, os.path.abspath('_ext'))
sys.path.insert(0, os.path.abspath('..'))

# -- General configuration -----------------------------------------------------

# If your documentation needs a minimal Sphinx version, state it here.
#needs_sphinx = '1.0'

# Add any Sphinx extension module names here, as strings. They can be extensions
# coming with Sphinx (named 'sphinx.ext.*') or your custom ones.
extensions = ['sphinx.ext.autodoc']
autoclass_content = 'init'
autodoc_default_flags = ['members']

# Add any paths that contain templates here, relative to this directory.
templates_path = ['_templates']

# The suffix of source filenames.
source_suffix = '.rst'

# The encoding of source files.
#source_encoding = 'utf-8-sig'

# The master toctree document.
master_doc = 'index'

# General information about the project.
project = u'Bubbles'
copyright = u'2013, Stefan Urbanek'

# The version info for the project you're documenting, acts as replacement for
# |version| and |release|, also used in various other places throughout the
# built documents.
#
# The short X.Y version.
version = '2.0.0'
# The full version, including alpha/beta/rc tags.
release = '2.0.0'

# The language for content autogenerated by Sphinx. Refer to documentation
# for a list of supported languages.
#language = None

# There are two options for replacing |today|: either, you set today to some
# non-false value, then it is used:
#today = ''
# Else, today_fmt is used as the format for a strftime call.
#today_fmt = '%B %d, %Y'

# List of patterns, relative to source directory, that match files and
# directories to ignore when looking for source files.
exclude_patterns = ['_build']

# The reST default role (used for this markup: `text`) to use for all documents.
#default_role = None

# If true, '()' will be appended to :func: etc. cross-reference text.
#add_function_parentheses = True

# If true, the current module name will be prepended to all description
# unit titles (such as .. function::).
#add_module_names = True

# If true, sectionauthor and moduleauthor directives will be shown in the
# output. They are ignored by default.
#show_authors = False

# The name of the Pygments (syntax highlighting) style to use.
pygments_style = 'sphinx'

# A list of ignored prefixes for module index sorting.
#modindex_common_prefix = []


# -- Options for HTML output ---------------------------------------------------

# The theme to use for HTML and HTML Help pages.  See the documentation for
# a list of builtin themes.
html_theme = 'default'

# Theme options are theme-specific and customize the look and feel of a theme
# further.  For a list of options available for each theme, see the
# documentation.
#html_theme_options = {}

# Add any paths that contain custom themes here, relative to this directory.
#html_theme_path = []

# The name for this set of Sphinx documents.  If None, it defaults to
# "<project> v<release> documentation".
#html_title = None

# A shorter title for the navigation bar.  Default is the same as html_title.
#html_short_title = None

# The name of an image file (relative to this directory) to place at the top
# of the sidebar.
#html_logo = None

# The name of an image file (within the static path) to use as favicon of the
# docs.  This file should be a Windows icon file (.ico) being 16x16 or 32x32
# pixels large.
#html_favicon = None

# Add any paths that contain custom static files (such as style sheets) here,
# relative to this directory. They are copied after the builtin static files,
# so a file named "default.css" will overwrite the builtin "default.css".
html_static_path = ['_static']

# If not '', a 'Last updated on:' timestamp is inserted at every page bottom,
# using the given strftime format.
#html_last_updated_fmt = '%b %d, %Y'

# If true, SmartyPants will be used to convert quotes and dashes to
# typographically correct entities.
#html_use_smartypants = True

# Custom sidebar templates, maps document names to template names.
#html_sidebars = {}

# Additional templates that should be rendered to pages, maps page names to
# template names.
#html_additional_pages = {}

# If false, no module index is generated.
#html_domain_indices = True

# If false, no index is generated.
#html_use_index = True

# If true, the index is split into individual pages for each letter.
#html_split_index = False

# If true, links to the reST sources are added to the pages.
#html_show_sourcelink = True

# If true, "Created using Sphinx" is shown in the HTML footer. Default is True.
#html_show_sphinx = True

# If true, "(C) Copyright ..." is shown in the HTML footer. Default is True.
#html_show_copyright = True

# If true, an OpenSearch description file will be output, and all pages will
# contain a <link> tag referring to it.  The value of this option must be the
# base URL from which the finished HTML is served.
#html_use_opensearch = ''

# This is the file name suffix for HTML files (e.g. ".xhtml").
#html_file_suffix = None

# Output file base name for HTML help builder.
htmlhelp_basename = 'bubbles'


# -- Options for LaTeX output --------------------------------------------------

latex_elements = {
# The paper size ('letterpaper' or 'a4paper').
#'papersize': 'letterpaper',

# The font size ('10pt', '11pt' or '12pt').
#'pointsize': '10pt',

# Additional stuff for the LaTeX preamble.
#'preamble': '',
}

# Grouping the document tree into LaTeX files. List of tuples
# (source start file, target name, title, author, documentclass [howto/manual]).
latex_documents = [
  ('index', 'Bubbles.tex', u'Bubbles Documentation',
   u'Stefan Urbanek', 'manual'),
]

# The name of an image file (relative to this directory) to place at the top of
# the title page.
#latex_logo = None

# For "manual" documents, if this is true, then toplevel headings are parts,
# not chapters.
#latex_use_parts = False

# If true, show page references after internal links.
#latex_show_pagerefs = False

# If true, show URL addresses after external links.
#latex_show_urls = False

# Documents to append as an appendix to all manuals.
#latex_appendices = []

# If false, no module index is generated.
#latex_domain_indices = True


# -- Options for manual page output --------------------------------------------

# One entry per manual page. List of tuples
# (source start file, name, description, authors, manual section).
man_pages = [
    ('index', 'bubbles', u'Bubbles Documentation',
     [u'Stefan Urbanek'], 1)
]

# If true, show URL addresses after external links.
#man_show_urls = False


# -- Options for Texinfo output ------------------------------------------------

# Grouping the document tree into Texinfo files. List of tuples
# (source start file, target name, title, author,
#  dir menu entry, description, category)
texinfo_documents = [
  ('index', 'Bubbles', u'Bubbles Documentation',
   u'Stefan Urbanek', 'Bubbles', 'One line description of project.',
   'Miscellaneous'),
]

# Documents to append as an appendix to all manuals.
#texinfo_appendices = []

# If false, no module index is generated.
#texinfo_domain_indices = True

# How to display URL addresses: 'footnote', 'no', or 'inline'.
#texinfo_show_urls = 'footnote'


# -- Options for Epub output ---------------------------------------------------

# Bibliographic Dublin Core info.
epub_title = u'Bubbles'
epub_author = u'Stefan Urbanek'
epub_publisher = u'Stefan Urbanek'
epub_copyright = u'2013, Stefan Urbanek'

# The language of the text. It defaults to the language option
# or en if the language is not set.
#epub_language = ''

# The scheme of the identifier. Typical schemes are ISBN or URL.
#epub_scheme = ''

# The unique identifier of the text. This can be a ISBN number
# or the project homepage.
#epub_identifier = ''

# A unique identification for the text.
#epub_uid = ''

# A tuple containing the cover image and cover page html template filenames.
#epub_cover = ()

# HTML files that should be inserted before the pages created by sphinx.
# The format is a list of tuples containing the path and title.
#epub_pre_files = []

# HTML files shat should be inserted after the pages created by sphinx.
# The format is a list of tuples containing the path and title.
#epub_post_files = []

# A list of files that should not be packed into the epub file.
#epub_exclude_files = []

# The depth of the table of contents in toc.ncx.
#epub_tocdepth = 3

# Allow duplicate toc entries.
#epub_tocdup = True

########NEW FILE########
__FILENAME__ = aggregate_over_window
from bubbles import Pipeline, FieldList, data_object, open_store

# Sample order data with fields:
fields = FieldList(
            ["id", "integer"],
            ["customer_id", "integer"],
            ["year", "integer"],
            ["amount", "integer"]
        )

data = [
    [1, 1, 2009, 10],
    [2, 1, 2010, 20],
    [3, 1, 2011, 20],
    [4, 1, 2012, 50],
    [5, 2, 2010, 50],
    [6, 2, 2012, 40],
    [7, 3, 2011, 100],
    [8, 3, 2012, 150],
    [9, 3, 2013, 120]
]

# Stores for SQL alternative, if enabled (see below)
stores = { "default": open_store("sql","sqlite:///") }

#
# Create the pipeline
#

p = Pipeline(stores=stores)
p.source_object("iterable", iterable=data, fields=fields)

# Uncomment this to get SQL operations instead of python iterator
p.create("default", "data")

# Find last purchase date
last_purchase = p.fork()
last_purchase.aggregate(["customer_id"],
                        [["year", "max"]],
                        include_count=False)
last_purchase.rename_fields({"year_max": "last_purchase_year"})
p.join_details(last_purchase, "customer_id", "customer_id")

p.pretty_print()

p.run()


########NEW FILE########
__FILENAME__ = hello
import bubbles

URL = "https://raw.github.com/Stiivi/cubes/master/examples/hello_world/data.csv"

p = bubbles.Pipeline()
p.source_object("csv_source", resource=URL, encoding="utf8")
p.retype({"Amount (US$, Millions)": "integer"})
p.aggregate("Category", "Amount (US$, Millions)")
p.pretty_print()

p.run()

########NEW FILE########
__FILENAME__ = hello_sql
import bubbles

# Follow the comments – there is a line to be uncommented

URL = "https://raw.github.com/Stiivi/cubes/master/examples/hello_world/data.csv"

# Prepare list of stores, we just need one temporary SQL store

stores = {
    "target": bubbles.open_store("sql", "sqlite:///")
}


p = bubbles.Pipeline(stores=stores)
p.source_object("csv_source", resource=URL, encoding="utf8")
p.retype({"Amount (US$, Millions)": "integer"})

# We create a table
# Uncomment this line and see the difference in debug messages
p.create("target", "data")

p.aggregate("Category", "Amount (US$, Millions)")
p.pretty_print()
p.run()


########NEW FILE########
__FILENAME__ = test_text
import unittest
from ..common import data_path

from bubbles.errors import *
from bubbles.backends.text.objects import CSVSource, CSVTarget
from bubbles.metadata import FieldList

class TextBackendTestCase(unittest.TestCase):
    def test_load(self):
        obj = CSVSource(data_path("fruits-sk.csv"))
        self.assertEqual(["id", "fruit", "type"], obj.fields.names())

        rows = list(obj.rows())
        self.assertEqual(16, len(rows))

        self.assertEqual(["1", "jablko", "malvice"], rows[0])
        obj.release()

    @unittest.skip("Moved to an oepration")
    def test_infer_types(self):
        obj = CSVSource(data_path("fruits-sk.csv"), infer_fields=True)
        self.assertEqual("integer", obj.fields[0].storage_type)

        rows = list(obj.rows())
        self.assertEqual([1, "jablko", "malvice"], rows[0])
        obj.release()

    def test_no_header(self):
        with self.assertRaises(ArgumentError):
            obj = CSVSource(data_path("fruits-sk.csv"), read_header=False)

        fields = FieldList("id", "fruit", "type")
        obj = CSVSource(data_path("fruits-sk.csv"), read_header=False,
                                fields=fields)
        self.assertEqual(["id", "fruit", "type"], obj.fields.names())

        rows = list(obj.rows())
        self.assertEqual(17, len(rows))

        self.assertEqual(["id", "fruit", "type"], rows[0])
        obj.release()

    def test_encoding(self):
        obj_l2 = CSVSource(data_path("fruits-sk-latin2.csv"), encoding="latin2")
        rows_l2 = list(obj_l2.rows())
        obj_l2.release()

        obj_utf = CSVSource(data_path("fruits-sk.csv"))
        rows_utf = list(obj_utf.rows())
        obj_utf.release()
        self.assertEqual(rows_l2, rows_utf)

if __name__ == "__main__":
    unittest.main()

########NEW FILE########
__FILENAME__ = test_xls
import unittest
from ..common import data_path

from bubbles import *
from bubbles.backends.xls import XLSStore, XLSObject

class XLSBackendTestCase(unittest.TestCase):
    def test_load(self):
        obj = XLSObject(data_path("data.xls"), encoding="latin1")
        self.assertEqual(["id", "name", "amount"], obj.fields.names())

        rows = list(obj.rows())
        self.assertEqual(4, len(rows))
        self.assertEqual(4, len(obj))

        self.assertSequenceEqual([1, "Adam", 10], rows[0])

    def test_skip(self):
        obj = XLSObject(data_path("data.xls"), FieldList("number", "name"),
                        skip_rows=2)
        self.assertEqual(["number", "name"], obj.fields.names())

        rows = list(obj.rows())
        self.assertEqual(2, len(obj))
        self.assertEqual(2, len(rows))

        self.assertSequenceEqual([3.0, "Cecil"], rows[0])

    def test_skip_too_much(self):
        with self.assertRaises(ArgumentError):
            obj = XLSObject(data_path("data.xls"), FieldList("numbers"), skip_rows=20)

    def test_store(self):
        store = XLSStore(data_path("data.xls"))
        self.assertSequenceEqual(["amounts", "numbers"], store.object_names())

        obj = store.get_object("numbers", skip_rows=2)
        self.assertEqual(10, len(obj))

        store = open_store("xls", data_path("data.xls"))

if __name__ == "__main__":
    unittest.main()

########NEW FILE########
__FILENAME__ = test_xlsx
import unittest
from ..common import data_path

from bubbles.backends.xlsx import XLSXStore, XLSXObject
from bubbles.metadata import FieldList
from bubbles.errors import ArgumentError
from bubbles.stores import open_store


class XLSXBackendTestCase(unittest.TestCase):
    def test_load(self):
        obj = XLSXObject(data_path("data.xlsx"), encoding="latin1")
        self.assertEqual(["id", "name", "amount"], obj.fields.names())

        rows = list(obj.rows())
        self.assertEqual(4, len(rows))
        self.assertEqual(4, len(obj))

        self.assertSequenceEqual([1, "Adam", 10], rows[0])

    def test_skip(self):
        obj = XLSXObject(data_path("data.xlsx"), FieldList("number", "name"),
                        skip_rows=2)
        self.assertEqual(["number", "name"], obj.fields.names())

        rows = list(obj.rows())
        self.assertEqual(2, len(obj))
        self.assertEqual(2, len(rows))

        self.assertSequenceEqual([3.0, "Cecil"], rows[0])

    def test_store(self):
        store = XLSXStore(data_path("data.xlsx"))
        self.assertSequenceEqual(["amounts", "numbers"], store.object_names())

        obj = store.get_object("numbers", skip_rows=2)
        self.assertEqual(10, len(obj))

        store = open_store("xlsx", data_path("data.xlsx"))

if __name__ == "__main__":
    unittest.main()

########NEW FILE########
__FILENAME__ = common
import os.path

def data_path(filename):
    """Returns a full path to data file with name `filename`"""
    root_path = os.path.join(os.path.dirname(__file__), "data")
    return os.path.join(root_path, filename)


########NEW FILE########
__FILENAME__ = test_core
import unittest
from bubbles import *
# import bubbles.iterator

# FIXME: clean this up
import inspect

def default(ctx, left):
    pass

def unary(ctx, left):
    pass

def binary(ctx, left, right):
    pass

class DummyDataObject(DataObject):
    def __init__(self, reps=None, data=None):
        """Creates a dummy data object with bogus representations `reps` and
        arbitrary data `data`"""
        self.reps = reps or []
        self.data = data

    def representations(self):
        return self.reps

class TextObject(DataObject):
    def __init__(self, string):
        self.string = string

    def representations(self):
        return ["rows", "text"]

    def rows(self):
        return iter(self.string)

    def text(self):
        return self.string

class OperationTestCase(unittest.TestCase):
    def test_match(self):
        self.assertTrue(Signature("sql").matches("sql"))
        self.assertTrue(Signature("*").matches("sql"))
        self.assertTrue(Signature("sql[]").matches("sql[]"))
        self.assertTrue(Signature("*[]").matches("sql[]"))

        self.assertFalse(Signature("sql").matches("rows"))
        self.assertFalse(Signature("sql").matches("sql[]"))

    def test_common_reps(self):
        objs = [
                DummyDataObject(["a", "b", "c"]),
                DummyDataObject(["a", "b", "d"]),
                DummyDataObject(["b", "d", "e"])
            ]
        self.assertEqual(["b"], list(common_representations(*objs)))

        objs = [
                DummyDataObject(["a", "b", "c"]),
                DummyDataObject(["a", "b", "d"]),
                DummyDataObject(["d", "d", "e"])
            ]
        self.assertEqual([], list(common_representations(*objs)))

    def test_prototype(self):
        proto = Signature("sql", "sql")
        match = Signature("*", "*")
        self.assertEqual(match, proto.as_prototype())

        proto = Signature("sql[]", "sql")
        match = Signature("*[]", "*")
        self.assertEqual(match, proto.as_prototype())

        proto = Signature("*[]", "*")
        match = Signature("*[]", "*")
        self.assertEqual(match, proto.as_prototype())

    def test_create_op(self):
        def one(ctx, obj):
            pass
        def fun(ctx, left, right):
            pass

        with self.assertRaises(ArgumentError):
            Operation("fun", [])

        op = Operation("fun")
        self.assertEqual(1, op.opcount)
        self.assertEqual(["obj"], op.operands)

        op = Operation("fun", ["left", "right"])
        self.assertEqual(2, op.opcount)

    def test_register(self):
        def func(ctx, obj):
            pass

        def func_sql(ctx, obj):
            pass

        def func_text(ctx, obj):
            pass

        op = Operation("select", ["obj"])

        op.register(func)
        self.assertEqual(1, len(op.registry))
        self.assertEqual([Signature("*")], list(op.registry.keys()))

        op.register(func_sql, "sql")
        op.register(func_text, "text")
        self.assertEqual(3, len(op.registry))

        sigs = []
        for s in op.signatures():
            sigs.append([op.rep for op in s.operands])

        self.assertSequenceEqual(sigs, [["*"], ["sql"], ["text"]])

        f = op.function(Signature("*"))
        self.assertEqual(f, func)
        f = op.function(Signature("sql"))
        self.assertEqual(f, func_sql)
        f = op.function(Signature("text"))
        self.assertEqual(f, func_text)

    def test_register_invalid(self):
        def func_invalid(obj):
            pass

        def func_invalid2(ctx, obj):
            pass

        op = Operation("select")

        with self.assertRaisesRegexp(ArgumentError, "Expected at least"):
            op.register(func_invalid)

    def test_decorator(self):

        op = Operation("select", ["obj"])

        @op.register
        def func(ctx, obj):
            pass

        sig = Signature("*")
        self.assertEqual(1, len(op.registry))
        self.assertEqual([sig], list(op.registry.keys()))
        self.assertEqual(func, op.registry[sig])

        @op.register("sql")
        def func_sql(ctx, obj):
            pass

        sig = Signature("sql")
        self.assertEqual(2, len(op.registry))
        self.assertEqual(func_sql, op.registry[sig])

        # Test decorator implicitly creating Operation
        @operation
        def implicit(ctx, obj):
            pass

        self.assertIsInstance(implicit, Operation)
        self.assertEqual("implicit", implicit.name)

    def test_context(self):
        c = OperationContext()

        with self.assertRaises(KeyError):
            c.operations["select"]

        with self.assertRaises(OperationError):
            c.operation("select")

        c.add_operation(Operation("select"))
        op = c.operations["select"]

        @operation
        def touch(ctx, obj):
            pass

        c.add_operation(touch)
        op = c.operations["touch"]

    def test_resolution_order(self):
        @operation
        def upper(ctx, obj):
            pass

        obj = DummyDataObject(["rows"])

        order = upper.resolution_order(get_representations(obj))
        self.assertEqual([Signature("*")], order)

        @upper.register("rows")
        def _(ctx, obj):
            pass

        order = upper.resolution_order(get_representations(obj))
        self.assertEqual([Signature("rows"),
                          Signature("*")], order)

    def test_call(self):
        @operation
        def upper(ctx, obj):
            return obj.text().upper()

        c = OperationContext()
        c.add_operation(upper)

        obj = TextObject("hi there")
        result = c.call("upper", obj)
        self.assertEqual("HI THERE", result)

    def test_call_context_op(self):
        op = Operation("upper")
        @op.register("rows")
        def _(ctx, obj):
            rows = obj.rows()
            text = "".join(rows)
            return list(text.upper())

        @op.register("text")
        def _(ctx, obj):
            text = obj.text()
            return list(text.upper())


        c = OperationContext()
        c.add_operation(op)

        obj = TextObject("windchimes")

        result = c.op.upper(obj)
        self.assertEqual(list("WINDCHIMES"), result)

    def test_get_representations(self):
        obj = DummyDataObject(["rows", "sql"])
        self.assertEqual( [["rows", "sql"]], get_representations(obj))

        obj = DummyDataObject(["rows", "sql"])
        extr = get_representations([obj])
        self.assertEqual( [["rows[]", "sql[]"]], extr)

    def test_comparison(self):
        sig1 = Signature("a", "b", "c")
        sig2 = Signature("a", "b", "c")
        sig3 = Signature("a", "b")

        self.assertTrue(sig1 == sig1)
        self.assertTrue(sig1 == sig2)
        self.assertFalse(sig1 == sig3)

        self.assertTrue(sig1 == ["a", "b", "c"])
        self.assertFalse(sig1 == ["a", "b"])

    def test_retry(self):
        op = Operation("join", ["left", "right"])

        @op.register("sql", "sql")
        def _(ctx, l, r):
            if l.data == r.data:
                return "SQL"
            else:
                raise RetryOperation(["sql", "rows"])

        @op.register("sql", "rows")
        def _(ctx, l, r):
            return "ITERATOR"

        local = DummyDataObject(["sql", "rows"], "local")
        remote = DummyDataObject(["sql", "rows"], "remote")

        c = OperationContext()
        c.add_operation(op)

        result = c.op.join(local, local)
        self.assertEqual(result, "SQL")

        result = c.op.join(local, remote)
        self.assertEqual(result, "ITERATOR")

        fail = Operation("fail", ["left", "right"])
        @fail.register("sql", "rows")
        def _(ctx, l, r):
            raise RetryOperation(["sql", "sql"])
        c.add_operation(fail)

        with self.assertRaises(OperationError):
            c.op.fail(local, local)

        # Already visited
        repeat = Operation("repeat", ["left", "right"])
        @repeat.register("sql", "sql")
        def _(ctx, l, r):
            raise RetryOperation(["sql", "sql"])
        c.add_operation(repeat)

        with self.assertRaises(RetryError):
            c.op.repeat(local, local)

    def test_allow_deny_retry(self):
        swim = Operation("swim", ["obj"])
        @swim.register("sql")
        def _(ctx, obj):
            raise RetryOperation(["rows"])

        @swim.register("rows")
        def _(ctx, obj):
            obj.data = "good"
            return obj

        obj = DummyDataObject(["sql", "rows"], "")

        c = OperationContext()
        c.add_operation(swim)

        result = c.op.swim(obj)
        self.assertEqual("good", result.data)

        c.retry_deny = ["swim"]
        c.retry_allow = []
        with self.assertRaises(RetryError):
            c.op.swim(obj)

        c.retry_deny = []
        c.retry_allow = ["swim"]
        result = c.op.swim(obj)
        self.assertEqual("good", result.data)

        c.retry_deny = ["swim"]
        c.retry_allow = ["swim"]
        with self.assertRaises(RetryError):
            c.op.swim(obj)


    def test_retry_nested(self):
        """Test whether failed nested operation fails correctly (Because of
        Issue #4)."""

        aggregate = Operation("aggregate", ["obj"])
        @aggregate.register("sql")
        def _(ctx, obj, fail):
            if fail:
                raise RetryOperation(["rows"])
            else:
                obj.data += "-SQL-"
            return obj

        @aggregate.register("rows")
        def _(ctx, obj, fail):
            obj.data += "-ROWS-"
            return obj

        window_aggregate = Operation("window_aggregate", ["obj"])
        @window_aggregate.register("sql")
        def _(ctx, obj, fail):
            obj.data += "START"
            ctx.op.aggregate(obj, fail)
            obj.data += "END"

        c = OperationContext()
        c.add_operation(aggregate)
        c.add_operation(window_aggregate)

        # Expected order:
        # 1. window_aggregate is called
        # 2. sql aggregate is called, but fails
        # 3. row aggregate is called
        # 4. window aggregate continues

        obj = DummyDataObject(["sql"], "")

        c.op.window_aggregate(obj, fail=True)
        self.assertEqual("START-ROWS-END", obj.data)

        obj.data = ""
        c.op.window_aggregate(obj, fail=False)
        self.assertEqual("START-SQL-END", obj.data)

    def test_priority(self):
        objsql = DummyDataObject(["sql", "rows"])
        objrows = DummyDataObject(["rows", "sql"])

        meditate = Operation("meditate", ["obj"])

        @meditate.register("sql")
        def fsql(ctx, obj):
            return "sql"

        @meditate.register("rows")
        def frows(ctx, obj):
            return "rows"

        c = OperationContext()
        c.add_operation(meditate)

        self.assertEqual("sql", c.op.meditate(objsql))
        self.assertEqual("rows", c.op.meditate(objrows))

if __name__ == "__main__":
    unittest.main()

########NEW FILE########
__FILENAME__ = test_graph
import unittest
from bubbles import *

class GraphTestCase(unittest.TestCase):
    def test_basic(self):
        g = Graph()
        g.add(Node("src"), "n1")
        g.add(Node("distinct"),"n2")
        g.add(Node("pretty_print"), "n3")

        self.assertEqual(3, len(g.nodes))

        g.connect("n1", "n2")

        sources = g.sources("n2")
        self.assertEqual(1, len(sources))
        self.assertTrue(isinstance(sources["default"], Node))
        self.assertEqual("src", sources["default"].opname)

    def test_ports(self):
        g = Graph()
        g.add(Node("dim"), "dim")
        g.add(Node("src"), "src")
        g.add(Node("join_detail"), "j")

        g.connect("dim", "j", "master")

        with self.assertRaises(GraphError):
            g.connect("src", "j", "master")

        g.connect("src", "j", "detail")

        sources = g.sources("j")
        self.assertEqual(2, len(sources))
        self.assertEqual(["detail", "master"], sorted(sources.keys()))

if __name__ == "__main__":
    unittest.main()

########NEW FILE########
__FILENAME__ = test_locals
import unittest
from bubbles.threadlocal import LocalProxy, thread_locals

class TestLocalProxy(unittest.TestCase):
    def setUp(self):
        try:
            del thread_locals.test
        except AttributeError:
            pass

    def test_no_local(self):

        proxy = LocalProxy("test")
        with self.assertRaises(RuntimeError):
            proxy["foo"] = "bar"

    def test_with_local(self):
        thread_locals.test = {}
        proxy = LocalProxy("test")
        proxy["foo"] = "bar"

        self.assertEqual("bar", thread_locals.test["foo"])

    def test_factory(self):
        proxy = LocalProxy("test", factory=dict)
        self.assertTrue(isinstance(proxy._represented_local_object(), dict))
        self.assertEqual(0, len(proxy))

        proxy["foo"] = "bar"
        self.assertEqual("bar", thread_locals.test["foo"])

########NEW FILE########
__FILENAME__ = test_metadata
import unittest
from copy import copy
from bubbles import FieldList, Field, FieldFilter, to_field
from bubbles.errors import *

class FieldListTestCase(unittest.TestCase):
    def test_list_creation(self):
        fields = FieldList("foo", "bar")

        for field in fields:
            self.assertEqual(type(field), Field)
            self.assertIsInstance(field.name, str)

        self.assertEqual("foo", fields[0].name, 'message')
        self.assertEqual(2, len(fields))

    def test_list_add(self):
        fields = FieldList("foo", "bar")
        fields.append("baz")
        self.assertEqual(3, len(fields))

    def test_indexes(self):
        fields = FieldList("a", "b", "c", "d")
        indexes = fields.indexes(["a", "c", "d"])
        self.assertEqual((0,2,3), indexes)

        indexes = fields.indexes( fields.fields() )
        self.assertEqual((0,1,2,3), indexes)

    def test_deletion(self):
        fields = FieldList("a", "b", "c", "d")
        del fields[0]

        self.assertEqual(["b", "c", "d"], fields.names())

        del fields[2]
        self.assertEqual(["b", "c"], fields.names())

        self.assertRaises(NoSuchFieldError, fields.field, "d")
        self.assertEqual(2, len(fields))

    def test_contains(self):
        fields = FieldList("a", "b", "c", "d")
        field = Field("a")
        self.assertEqual(True, "a" in fields)
        self.assertEqual(True, field in fields._fields)

    #def test_retype(self):
    #    fields = FieldList(["a", "b", "c", "d"])
    #    self.assertEqual("unknown", fields.field("a").storage_type)
    #    retype_dict = {"a": {"storage_type":"integer"}}
    #    fields.retype(retype_dict)
    #    self.assertEqual("integer", fields.field("a").storage_type)

    #    retype_dict = {"a": {"name":"foo"}}
    #    self.assertRaises(Exception, fields.retype, retype_dict)

    def test_mask(self):
        fields = FieldList("a", "b", "c", "d")
        mask = fields.mask(["b", "d"])
        self.assertEqual([False, True, False, True], mask)

class MetadataTestCase(unittest.TestCase):
    def test_names(self):
        field = Field("bar")
        self.assertEqual("bar", field.name)
        self.assertEqual("bar", str(field))

    def test_to_field(self):
        field = to_field("foo")
        self.assertIsInstance(field, Field)
        self.assertEqual("foo", field.name)
        self.assertIsInstance(field.name, str)
        # self.assertEqual("unknown", field.storage_type)
        # self.assertEqual("typeless", field.analytical_type)

        field = to_field(["bar", "string", "flag"])
        self.assertEqual("bar", field.name)
        self.assertEqual("string", field.storage_type)
        self.assertEqual("flag", field.analytical_type)

        desc = {
                "name":"baz",
                "storage_type":"integer",
                "analytical_type": "flag"
            }
        field = to_field(desc)
        self.assertEqual("baz", field.name)
        self.assertEqual("integer", field.storage_type)
        self.assertEqual("flag", field.analytical_type)

    def test_field_to_dict(self):
        desc = {
                "name":"baz",
                "storage_type":"integer",
                "analytical_type": "flag"
            }
        field = to_field(desc)
        field2 = to_field(field.to_dict())
        self.assertEqual(field, field2)

    @unittest.skip("skipping")
    def test_coalesce_value(self):
        self.assertEqual(1, coalesce_value("1", "integer"))
        self.assertEqual("1", coalesce_value(1, "string"))
        self.assertEqual(1.5, coalesce_value("1.5", "float"))
        self.assertEqual(1000, coalesce_value("1 000", "integer", strip=True))
        self.assertEqual(['1','2','3'], coalesce_value("1,2,3", "list", strip=True))


    def setUp(self):
        self.fields = FieldList("a", "b", "c", "d")

    def test_init(self):
        self.assertRaises(MetadataError, FieldFilter, drop=["foo"], keep=["bar"])

    def test_map(self):

        m = FieldFilter(drop=["a","c"])
        self.assertListEqual(["b", "d"], m.filter(self.fields).names())

        m = FieldFilter(keep=["a","c"])
        self.assertListEqual(["a", "c"], m.filter(self.fields).names())

        m = FieldFilter(rename={"a":"x","c":"y"})
        self.assertListEqual(["x", "b", "y", "d"], m.filter(self.fields).names())

    def test_selectors(self):
        m = FieldFilter(keep=["a","c"])
        self.assertListEqual([True, False, True, False],
                                m.field_mask(self.fields))

        m = FieldFilter(drop=["b","d"])
        self.assertListEqual([True, False, True, False],
                                m.field_mask(self.fields))

        m = FieldFilter()
        self.assertListEqual([True, True, True, True],
                                m.field_mask(self.fields))
def test_suite():
   suite = unittest.TestSuite()

   suite.addTest(unittest.makeSuite(FieldListTestCase))
   suite.addTest(unittest.makeSuite(MetadataTestCase))

   return suite


########NEW FILE########
