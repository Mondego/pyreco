__FILENAME__ = couchdb
#!/usr/bin/env python

"""
Example showing how to import data from a CouchDB instance.

Uses Couch's _changes feed to propogate updates and deletes into PANDA.
"""

import json

import requests

PANDA_API = 'http://localhost:8000/api/1.0'
PANDA_AUTH_PARAMS = {
    'email': 'panda@pandaproject.net',
    'api_key': 'edfe6c5ffd1be4d3bf22f69188ac6bc0fc04c84b'
}
PANDA_DATASET_SLUG = 'couchdb-example'

PANDA_DATASET_URL = '%s/dataset/%s/' % (PANDA_API, PANDA_DATASET_SLUG)
PANDA_DATA_URL = '%s/dataset/%s/data/' % (PANDA_API, PANDA_DATASET_SLUG)
PANDA_BULK_UPDATE_SIZE = 1000

COUCHDB_ROOT_URL = 'http://datacouch.com/db/dc07acde3002cb1f62a08de546916097cd'
COUCHDB_ROWS_URL = 'http://datacouch.com/db/dc07acde3002cb1f62a08de546916097cd/rows'
COUCHDB_CHANGES_URL = 'http://datacouch.com/db/dc07acde3002cb1f62a08de546916097cd/_changes'

COLUMNS = ['First Name', 'Last Name', 'Employer']

LAST_SEQ_FILENAME = 'last_seq'

# Utility functions
def panda_get(url, params={}):
    params.update(PANDA_AUTH_PARAMS)
    return requests.get(url, params=params)

def panda_put(url, data, params={}):
    params.update(PANDA_AUTH_PARAMS)
    return requests.put(url, data, params=params, headers={ 'Content-Type': 'application/json' })

def panda_delete(url):
    return requests.delete(url, params=PANDA_AUTH_PARAMS, headers={ 'Content-Type': 'application/json' })

def write_last_seq(last_seq):
    with open(LAST_SEQ_FILENAME, 'w') as f:
        f.write(str(last_seq))

def read_last_seq():
    with open(LAST_SEQ_FILENAME) as f:
        return f.read().strip()

def couchdb_row_to_panda_data(row):
    return {
        'data': [row['first_name'], row['last_name'], row['employer']],
        'external_id': row['_id'] 
    }

# Check if dataset exists
response = panda_get(PANDA_DATASET_URL)

# Create dataset if necessary
if response.status_code == 404:
    dataset = {
        'name': 'CouchDB: PANDA Contributors',
        'description': 'A list of contributors to PANDA imported from a dataset on DataCouch: <a href="http://datacouch.com/edit/#/dc07acde3002cb1f62a08de546916097cd">http://datacouch.com/edit/#/dc07acde3002cb1f62a08de546916097cd</a>.'
    }

    response = panda_put(PANDA_DATASET_URL, json.dumps(dataset), params={ 'columns': ','.join(COLUMNS) })

    # Get changes that have come before so we can skip them in the future
    response = requests.get(COUCHDB_CHANGES_URL)
    data = json.loads(response.content)

    write_last_seq(data['last_seq'])

    # Do a complete import of all data from CouchDB 
    response = requests.get(COUCHDB_ROWS_URL)
    data = json.loads(response.content)

    put_data = {
        'objects': []
    }

    for i, row in enumerate(data['rows']):
        put_data['objects'].append(couchdb_row_to_panda_data(row['value']))

        if i and i % PANDA_BULK_UPDATE_SIZE == 0:
            print 'Updating %i rows...' % PANDA_BULK_UPDATE_SIZE

            panda_put(PANDA_DATA_URL, json.dumps(put_data))
            put_data['objects'] = []
            
    if put_data['objects']:
        print 'Updating %i rows' % len(put_data['objects'])
        panda_put(PANDA_DATA_URL, json.dumps(put_data))

# Update existing dataset
else:
    # Where did we leave off?
    last_seq = read_last_seq()

    response = requests.get(COUCHDB_CHANGES_URL, params={ 'since': last_seq })
    data = json.loads(response.content)
    
    delete_ids = []

    put_data = {
        'objects': []
    }

    for i, row in enumerate(data['results']):
        # Is this a deletion?
        if row.get('deleted', False):
            delete_ids.append(row['id'])
            continue

        doc_id = row['id']

        detail_response = requests.get('%s/%s' % (COUCHDB_ROOT_URL, doc_id))
        detail_data = json.loads(detail_response.content)

        put_data['objects'].append(couchdb_row_to_panda_data(detail_data))

        if i and i % PANDA_BULK_UPDATE_SIZE == 0:
            print 'Updating %i rows...' % PANDA_BULK_UPDATE_SIZE

            panda_put(PANDA_DATA_URL, json.dumps(put_data))
            put_data['objects'] = []
            
    if put_data['objects']:
        print 'Updating %i rows' % len(put_data['objects'])
        panda_put(PANDA_DATA_URL, json.dumps(put_data))

    # Process deletes
    if delete_ids:
        print 'Deleting %i rows' % len(delete_ids)

        for deleted in delete_ids:
            response = panda_delete('%s%s/' % (PANDA_DATA_URL, deleted))

    # Update location for next run
    write_last_seq(data['last_seq'])

print 'Done'


########NEW FILE########
__FILENAME__ = google_docs
#!/usr/bin/env python

"""
Example showing how to import data from the Public Google Spreadsheet.
"""

import json
from StringIO import StringIO

from csvkit import CSVKitReader
import requests

PANDA_API = 'http://localhost:8000/api/1.0'
PANDA_AUTH_PARAMS = {
    'email': 'panda@pandaproject.net',
    'api_key': 'edfe6c5ffd1be4d3bf22f69188ac6bc0fc04c84b'
}
PANDA_DATASET_SLUG = 'news-developer-jobs'

PANDA_DATASET_URL = '%s/dataset/%s/' % (PANDA_API, PANDA_DATASET_SLUG)
PANDA_DATA_URL = '%s/dataset/%s/data/' % (PANDA_API, PANDA_DATASET_SLUG)
PANDA_BULK_UPDATE_SIZE = 1000

SPREADSHEET_ID = '0AmqohgGX3YQadE1VSktrWG1nNFF6RUFNT1RKa0k0a2c'
COLUMNS = ['Employer', 'Date Entered', 'More Info', 'Job Title', 'City / State', 'Contact person', 'Contact email / phone', 'Country', 'Latitude', 'Longitude']

# Utility functions
def panda_get(url, params={}):
    params.update(PANDA_AUTH_PARAMS)
    return requests.get(url, params=params)

def panda_put(url, data, params={}):
    params.update(PANDA_AUTH_PARAMS)
    return requests.put(url, data, params=params, headers={ 'Content-Type': 'application/json' })

def panda_delete(url, params={}):
    params.update(PANDA_AUTH_PARAMS)
    return requests.delete(url, params=params)

# Check if dataset exists
response = panda_get(PANDA_DATASET_URL)

# Create dataset if necessary
if response.status_code == 404:
    dataset = {
        'name': 'Google Docs: News Developer Jobs',
        'description': 'The crowdsourced jobs list that powers http://www.newsnerdjobs.com/.'
    }

    response = panda_put(PANDA_DATASET_URL, json.dumps(dataset), params={ 'columns': ','.join(COLUMNS) })

# Open connection to Google
response = requests.get('https://docs.google.com/spreadsheet/pub?key=%s&single=true&gid=4&output=csv' % SPREADSHEET_ID)
csv = StringIO(response.content)

reader = CSVKitReader(csv)
reader.next()

put_data = {
    'objects': []
}

# Delete existing data in panda
response = panda_delete(PANDA_DATA_URL)

for i, row in enumerate(reader):
    put_data['objects'].append({
        'data': row
    })

    if i and i % PANDA_BULK_UPDATE_SIZE == 0:
        print 'Updating %i rows...' % PANDA_BULK_UPDATE_SIZE

        panda_put(PANDA_DATA_URL, json.dumps(put_data))
        put_data['objects'] = []
        
if put_data['objects']:
    print 'Updating %i rows' % len(put_data['objects'])
    panda_put(PANDA_DATA_URL, json.dumps(put_data))

print 'Done'


########NEW FILE########
__FILENAME__ = scraperwiki
#!/usr/bin/env python

"""
Example showing how to import data from the Scraperwiki API.
"""

import json
import re

import requests

PANDA_API = 'http://localhost:8000/api/1.0'
PANDA_AUTH_PARAMS = {
    'email': 'panda@pandaproject.net',
    'api_key': 'edfe6c5ffd1be4d3bf22f69188ac6bc0fc04c84b'
}
PANDA_DATASET_SLUG = 'smith-county-criminal-cases'

PANDA_DATASET_URL = '%s/dataset/%s/' % (PANDA_API, PANDA_DATASET_SLUG)
PANDA_DATA_URL = '%s/dataset/%s/data/' % (PANDA_API, PANDA_DATASET_SLUG)
PANDA_BULK_UPDATE_SIZE = 1000

SCRAPERWIKI_URL = 'https://api.scraperwiki.com/api/1.0/datastore/sqlite?format=jsondict&name=tyler_criminal_records&query=select%20*%20from%20%60swdata%60'
COLUMNS = ['cause_number', 'date_filed', 'defendant_name', 'defendant_birthdate', 'offense', 'crime_date', 'degree', 'disposed', 'court', 'warrant_status', 'attorney', 'view_url']
COLUMN_TYPES = ['', 'date', '', '', '', '', '', '', '', '', '', '']

# Utility functions
def panda_get(url, params={}):
    params.update(PANDA_AUTH_PARAMS)
    return requests.get(url, params=params)

def panda_put(url, data, params={}):
    params.update(PANDA_AUTH_PARAMS)
    return requests.put(url, data, params=params, headers={ 'Content-Type': 'application/json' })

def slugify(value):
    """
    Graciously borrowed from Django core.
    """
    import unicodedata
    value = unicodedata.normalize('NFKD', value).encode('ascii', 'ignore')
    value = unicode(re.sub('[^\w\s-]', '', value).strip().lower())
    return re.sub('[-\s]+', '-', value)

# Check if dataset exists
response = panda_get(PANDA_DATASET_URL)

# Create dataset if necessary
if response.status_code == 404:
    dataset = {
        'name': 'Scraperwiki: Smith County Criminal Case Records',
        'description': 'Results of the scraper at <a href="https://scraperwiki.com/scrapers/tyler_criminal_records/">https://scraperwiki.com/scrapers/tyler_criminal_records/</a>.'
    }

    response = panda_put(PANDA_DATASET_URL, json.dumps(dataset), params={
        'columns': ','.join(COLUMNS),
        'typed_columns': ','.join(['true' if t else '' for t in COLUMN_TYPES]),
        'column_types': ','.join(COLUMN_TYPES) 
    })

# Fetch latest data from Scraperwiki
print 'Fetching latest data'
response = requests.get(SCRAPERWIKI_URL)

data = json.loads(response.content)

put_data = {
    'objects': []
}

for i, row in enumerate(data):
    put_data['objects'].append({
        'data': [row[c] for c in COLUMNS],
        'external_id': slugify(row['cause_number']) # Slugify because a few have errants commas and such
    })

    if i and i % PANDA_BULK_UPDATE_SIZE == 0:
        print 'Updating %i rows...' % PANDA_BULK_UPDATE_SIZE

        panda_put(PANDA_DATA_URL, json.dumps(put_data))
        put_data['objects'] = []
        
if put_data['objects']:
    print 'Updating %i rows' % len(put_data['objects'])
    panda_put(PANDA_DATA_URL, json.dumps(put_data))

print 'Done'


########NEW FILE########
__FILENAME__ = scraperwiki_twitter
#!/usr/bin/env python

"""
Example showing how to import Twitter data from the Scraperwiki API.
"""

import json

import requests

PANDA_API = 'http://localhost:8000/api/1.0'
PANDA_AUTH_PARAMS = {
    'email': 'panda@pandaproject.net',
    'api_key': 'edfe6c5ffd1be4d3bf22f69188ac6bc0fc04c84b'
}
PANDA_DATASET_SLUG = 'twitter-pandaproject'

PANDA_DATASET_URL = '%s/dataset/%s/' % (PANDA_API, PANDA_DATASET_SLUG)
PANDA_DATA_URL = '%s/dataset/%s/data/' % (PANDA_API, PANDA_DATASET_SLUG)
PANDA_BULK_UPDATE_SIZE = 1000

SCRAPERWIKI_URL = 'https://api.scraperwiki.com/api/1.0/datastore/sqlite?format=jsonlist&name=basic_twitter_scraper_437&query=select%20*%20from%20%60swdata%60'
COLUMNS = ['text', 'id', 'from_user']

# Utility functions
def panda_get(url, params={}):
    params.update(PANDA_AUTH_PARAMS)
    return requests.get(url, params=params)

def panda_put(url, data, params={}):
    params.update(PANDA_AUTH_PARAMS)
    return requests.put(url, data, params=params, headers={ 'Content-Type': 'application/json' })

# Check if dataset exists
response = panda_get(PANDA_DATASET_URL)

# Create dataset if necessary
if response.status_code == 404:
    dataset = {
        'name': 'PANDA Project Twitter Search',
        'description': 'Results of the scraper at <a href="https://scraperwiki.com/scrapers/basic_twitter_scraper_437/">https://scraperwiki.com/scrapers/basic_twitter_scraper_437/</a>.'
    }

    response = panda_put(PANDA_DATASET_URL, json.dumps(dataset), params={
        'columns': ','.join(COLUMNS),
    })

# Fetch latest data from Scraperwiki
print 'Fetching latest data'
response = requests.get(SCRAPERWIKI_URL)

data = json.loads(response.content)

put_data = {
    'objects': []
}

for i, row in enumerate(data['data']):
    put_data['objects'].append({
        'data': row,
        'external_id': unicode(row[1])
    })

    if i and i % PANDA_BULK_UPDATE_SIZE == 0:
        print 'Updating %i rows...' % PANDA_BULK_UPDATE_SIZE

        panda_put(PANDA_DATA_URL, json.dumps(put_data))
        put_data['objects'] = []
        
if put_data['objects']:
    print 'Updating %i rows' % len(put_data['objects'])
    response = panda_put(PANDA_DATA_URL, json.dumps(put_data))

print 'Done'


########NEW FILE########
__FILENAME__ = socrata
#!/usr/bin/env python

"""
Example showing how to import data from the Socrata API.
"""

import json

import requests

PANDA_API = 'http://localhost:8000/api/1.0'
PANDA_AUTH_PARAMS = {
    'email': 'panda@pandaproject.net',
    'api_key': 'edfe6c5ffd1be4d3bf22f69188ac6bc0fc04c84b'
}
PANDA_DATASET_SLUG = 'foia-request-log-311'

PANDA_DATASET_URL = '%s/dataset/%s/' % (PANDA_API, PANDA_DATASET_SLUG)
PANDA_DATA_URL = '%s/dataset/%s/data/' % (PANDA_API, PANDA_DATASET_SLUG)
PANDA_BULK_UPDATE_SIZE = 1000

SOCRATA_URL = 'http://data.cityofchicago.org/api/views/j2p9-gdf5/rows.json?unwrapped=true'
COLUMNS = ['Requestor Name', 'Organization', 'Description of Request', 'Date Received', 'Due Date']

# Utility functions
def panda_get(url, params={}):
    params.update(PANDA_AUTH_PARAMS)
    return requests.get(url, params=params)

def panda_put(url, data, params={}):
    params.update(PANDA_AUTH_PARAMS)
    return requests.put(url, data, params=params, headers={ 'Content-Type': 'application/json' })

# Check if dataset exists
response = panda_get(PANDA_DATASET_URL)

# Create dataset if necessary
if response.status_code == 404:
    dataset = {
        'name': 'Socrata: FOIA Request Log - Office of the Mayor',
        'description': 'FOIA requests made to the mayor\'s office imported from the Socrata dataset at <a href="http://data.cityofchicago.org/FOIA/FOIA-Request-Log-Office-of-the-Mayor/srzw-dcvg">http://data.cityofchicago.org/FOIA/FOIA-Request-Log-Office-of-the-Mayor/srzw-dcvg</a>.'
    }

    response = panda_put(PANDA_DATASET_URL, json.dumps(dataset), params={ 'columns': ','.join(COLUMNS) })

# Fetch latest data from Socrata
print 'Fetching latest data'
response = requests.get(SOCRATA_URL)

data = json.loads(response.content)

put_data = {
    'objects': []
}

for i, row in enumerate(data):
    # First 8 columns are metadata
    put_data['objects'].append({
        'data': row[-5:],
        'external_id': unicode(row[0])   # per-dataset id
    })

    if i and i % PANDA_BULK_UPDATE_SIZE == 0:
        print 'Updating %i rows...' % PANDA_BULK_UPDATE_SIZE

        panda_put(PANDA_DATA_URL, json.dumps(put_data))
        put_data['objects'] = []
        
if put_data['objects']:
    print 'Updating %i rows' % len(put_data['objects'])
    panda_put(PANDA_DATA_URL, json.dumps(put_data))

print 'Done'


########NEW FILE########
__FILENAME__ = application
#!/usr/bin/env python

import os

import django.core.handlers.wsgi

# When serving under WSGI (rather than runserver) use deployed config
os.environ["DJANGO_SETTINGS_MODULE"] = "config.deployed.settings"

application = django.core.handlers.wsgi.WSGIHandler()


########NEW FILE########
__FILENAME__ = application_jumpstart
#!/usr/bin/env python

import os

import django.core.handlers.wsgi

os.environ["DJANGO_SETTINGS_MODULE"] = "config.jumpstart.settings"

application = django.core.handlers.wsgi.WSGIHandler()


########NEW FILE########
__FILENAME__ = urls
#!/usr/bin/env python

from django.conf.urls.defaults import patterns, url

from client import views

urlpatterns = patterns('',
    url(r'^templates.js$', views.jst, name='jst'),
    url(r'^i18n.js$', 'django.views.i18n.javascript_catalog'),
    url(r'^dashboard/$', views.dashboard, name='dashboard'),
    url(r'^$', views.index, name='index')
)


########NEW FILE########
__FILENAME__ = utils
#!/usr/bin/env python

import os

def get_total_disk_space(p):
    """
    Calculate the total disk space of the device on which a given file path resides.
    """
    s = os.statvfs(p)
    return s.f_frsize * s.f_blocks   

def get_free_disk_space(p):
    """
    Returns the number of free bytes on the drive that ``p`` is on
    """
    s = os.statvfs(p)
    return s.f_frsize * s.f_bavail


########NEW FILE########
__FILENAME__ = views
#!/usr/bin/env python

import datetime
import os
import re
from urllib import unquote

from django.conf import settings
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import render_to_response
from django.utils.timezone import now

from livesettings import config_value
from tastypie.serializers import Serializer

from client import utils
from panda.api.category import CategoryResource
from panda.models import ActivityLog, Category, Dataset, SearchLog, UserProxy

def index(request):
    """
    Page shell for the client-side application.

    Bootstraps read-once data onto the page.
    """
    serializer = Serializer()
    cr = CategoryResource()

    categories = list(Category.objects.annotate(dataset_count=Count('datasets')))

    bundles = [cr.build_bundle(obj=c) for c in categories]
    categories_bootstrap = [cr.full_dehydrate(b) for b in bundles]

    uncategorized = Category(
        id=settings.PANDA_UNCATEGORIZED_ID,
        slug=settings.PANDA_UNCATEGORIZED_SLUG,
        name=settings.PANDA_UNCATEGORIZED_NAME)
    uncategorized.__dict__['dataset_count'] = Dataset.objects.filter(categories=None).count() 
    uncategorized_bundle = cr.full_dehydrate(cr.build_bundle(obj=uncategorized))

    categories_bootstrap.append(uncategorized_bundle)

    return render_to_response('index.html', {
        'settings': settings,
        'warn_upload_size': int(config_value('MISC', 'WARN_UPLOAD_SIZE')),
        'max_upload_size': int(config_value('MISC', 'MAX_UPLOAD_SIZE')),
        'email_enabled': int(config_value('EMAIL', 'EMAIL_ENABLED')),
        'demo_mode_enabled': int(config_value('MISC', 'DEMO_MODE_ENABLED')),
        'bootstrap_data': serializer.to_json({
            'categories': categories_bootstrap
        }),
        'moment_lang_code': settings.MOMENT_LANGUAGE_MAPPING.get(settings.LANGUAGE_CODE, None),
    })

def dashboard(request):
    """
    Render HTML for dashboard/metrics view.
    """
    # Datasets
    dataset_count = Dataset.objects.all().count()

    datasets_without_descriptions = [(unquote(dataset['name']), dataset['slug']) for dataset in Dataset.objects.filter(description='').values('name', 'slug')]
    datasets_without_categories = [(unquote(dataset['name']), dataset['slug']) for dataset in Dataset.objects.filter(categories=None).values('name', 'slug')]

    # Users
    user_count = UserProxy.objects.all().count()
    activated_user_count = UserProxy.objects.filter(is_active=True).count()

    today = now().date()
    thirty_days_ago = today - datetime.timedelta(days=30)

    active_users = list(UserProxy.objects.raw('SELECT auth_user.*, count(panda_activitylog.id) AS activity_logs__count FROM auth_user LEFT JOIN panda_activitylog ON panda_activitylog.user_id = auth_user.id WHERE auth_user.is_active = True AND panda_activitylog.when > %s GROUP BY auth_user.id ORDER BY activity_logs__count DESC, auth_user.id ASC', [thirty_days_ago]))

    most_active_users = active_users[:10]

    if len(active_users) > 10:
        least_active_users = active_users[-10:]
        least_active_users.reverse()
    else:
        least_active_users = []

    inactive_users = UserProxy.objects.all() \
        .annotate(Count('activity_logs')) \
        .filter(activity_logs__count=0)

    _active_users_by_day = \
        list(ActivityLog.objects.filter(when__gt=thirty_days_ago) \
        .values('when') \
        .annotate(Count('id')) \
        .order_by('when'))

    dates = [thirty_days_ago + datetime.timedelta(days=x) for x in range(0, 31)]

    active_users_by_day = []

    for d in dates:
        if _active_users_by_day and _active_users_by_day[0]['when'] == d:
            _d = _active_users_by_day.pop(0)
            active_users_by_day.append(_d)
        else:
            active_users_by_day.append({ 'when': d, 'id__count': 0 })

    # Searches

    total_searches = SearchLog.objects.count()

    most_searched_datasets = [(unquote(dataset['name']), dataset['slug'], dataset['searches__count']) for dataset in \
        Dataset.objects.all() \
        .annotate(Count('searches')) \
        .filter(searches__count__gt=0) \
        .order_by('-searches__count') \
        .values('name', 'slug', 'searches__count')[:10]]

    _searches_by_day = \
        list(SearchLog.objects.filter(when__gt=thirty_days_ago) \
        .extra(select={ 'day': '"when"::date' }) \
        .values('day') \
        .annotate(Count('when')) \
        .order_by('day'))

    dates = [thirty_days_ago + datetime.timedelta(days=x) for x in range(0, 31)]

    searches_by_day = []

    for d in dates:
        if _searches_by_day and _searches_by_day[0]['day'] == d:
            _d = _searches_by_day.pop(0)
            searches_by_day.append(_d)
        else:
            searches_by_day.append({ 'day': d, 'when__count': 0 })

    # Disk space
    root_disk = os.stat('/').st_dev
    upload_disk = os.stat(settings.MEDIA_ROOT).st_dev
    indices_disk = os.stat(settings.SOLR_DIRECTORY).st_dev

    root_disk_total = utils.get_total_disk_space('/')
    root_disk_free = utils.get_free_disk_space('/')
    root_disk_percent_used = 100 - (float(root_disk_free) / root_disk_total * 100)

    if upload_disk != root_disk:    
        upload_disk_total = utils.get_total_disk_space(settings.MEDIA_ROOT)
        upload_disk_free = utils.get_free_disk_space(settings.MEDIA_ROOT)
        upload_disk_percent_used = 100 - (float(upload_disk_free) / upload_disk_total * 100)
    else:
        upload_disk_total = None
        upload_disk_free = None
        upload_disk_percent_used = None

    if indices_disk != root_disk:
        indices_disk_total = utils.get_total_disk_space(settings.SOLR_DIRECTORY)
        indices_disk_free = utils.get_free_disk_space(settings.SOLR_DIRECTORY)
        indices_disk_percent_used = 100 - (float(indices_disk_free) / indices_disk_total * 100)
    else:
        indices_disk_total = None
        indices_disk_free = None
        indices_disk_percent_used = None

    return render_to_response('dashboard.html', {
        'settings': settings,
        'dataset_count': dataset_count,
        'datasets_without_descriptions': datasets_without_descriptions,
        'datasets_without_categories': datasets_without_categories,
        'user_count': user_count,
        'activated_user_count': activated_user_count,
        'most_active_users': most_active_users,
        'least_active_users': least_active_users,
        'inactive_users': inactive_users,
        'active_users_by_day': active_users_by_day,
        'total_searches': total_searches,
        'most_searched_datasets': most_searched_datasets,
        'searches_by_day': searches_by_day,
        'root_disk_total': root_disk_total,
        'root_disk_free': root_disk_free,
        'root_disk_percent_used': root_disk_percent_used,
        'upload_disk_total': upload_disk_total,
        'upload_disk_free': upload_disk_free,
        'upload_disk_percent_used': upload_disk_percent_used,
        'indices_disk_total': indices_disk_total,
        'indices_disk_free': indices_disk_free,
        'indices_disk_percent_used': indices_disk_percent_used,
        'storage_documentation_url': 'http://panda.readthedocs.org/en/%s/storage.html' % settings.PANDA_VERSION
    })

def jst(request):
    """
    Compile JST templates into a javascript module.
    """
    templates_path = os.path.join(settings.SITE_ROOT, 'client/static/templates')

    compiled = ''

    for dirpath, dirnames, filenames in os.walk(templates_path):
        for filename in filenames:
            name, extension = os.path.splitext(filename)

            if extension != '.jst':
                continue

            fullpath = os.path.join(dirpath, filename)

            with open(fullpath, 'r') as f:
                contents = f.read()

            # Borrowed from django-pipeline
            contents = re.sub(r"\r?\n", "", contents)
            contents = re.sub(r"'", "\\'", contents)

            compiled += "PANDA.templates['%s'] = _.template('%s');\n" % (
                name,
                contents
            )

    return HttpResponse(compiled, mimetype='text/javascript')


########NEW FILE########
__FILENAME__ = settings
#!/usr/bin/env python

from config.settings import *

# Running in deployed mode
SETTINGS = 'deployed'

# Debug
DEBUG = False 
TEMPLATE_DEBUG = DEBUG

# Static media
STATIC_ROOT = '/var/lib/panda/media'

# Uploads 
MEDIA_ROOT = '/var/lib/panda/uploads' 
EXPORT_ROOT = '/var/lib/panda/exports'

# Solr
SOLR_DIRECTORY = '/opt/solr/panda/solr'

# Django-compressor
COMPRESS_ENABLED = True 

# Celery
CELERYBEAT_SCHEDULE_FILENAME = '/var/celery/celerybeat-schedule'

try:
    from local_settings import *
except ImportError:
    pass


########NEW FILE########
__FILENAME__ = settings
#!/usr/bin/env python

from config.settings import *
from config.deployed.settings import *

# Running in jumpstart mode
# This means the app will have root access!
SETTINGS = 'jumpstart'

DAEMON_PID_PATH = '/tmp/jumpstart-restart.pid'
DAEMON_LOG_PATH = '/var/log/jumpstart-restart.log'

try:
    from local_settings import *
except ImportError:
    pass


########NEW FILE########
__FILENAME__ = settings
#!/usr/bin/env python

import datetime
import os

import django
from django.utils.translation import ugettext_lazy as _

# Which settings are we using?
# Useful for debugging.
SETTINGS = 'base'

# Base paths
DJANGO_ROOT = os.path.dirname(os.path.realpath(django.__file__))
SITE_ROOT = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))

# Debugging
DEBUG = True
TEMPLATE_DEBUG = DEBUG

ADMINS = (
    # ('Your Name', 'your_email@domain.com'),
)

MANAGERS = ADMINS

LOGIN_URL = '/admin/login/'
LOGOUT_URL = '/admin/logout/'
LOGIN_REDIRECT_URL = '/admin/'

SITE_ID = 1

# Default connection to socket
DATABASES = {
    'default': {
        'ENGINE': 'django.db.backends.postgresql_psycopg2',
        'HOST': 'localhost',
        'PORT': '5432',
        'NAME': 'panda',
        'USER': 'panda',
        'PASSWORD': 'panda'
    }
}

TIME_ZONE = 'Etc/UTC' 
USE_TZ = True 

LANGUAGE_CODE = 'en-us'
USE_I18N = True
USE_L10N = False 

LOCALE_PATHS = (os.path.join(SITE_ROOT, 'locale'),)

# Media
STATIC_ROOT = os.path.join(SITE_ROOT, 'media')
STATIC_URL = '/site_media/'
ADMIN_MEDIA_PREFIX = '/site_media/admin/'

STATICFILES_FINDERS = (
    'django.contrib.staticfiles.finders.FileSystemFinder',
    'django.contrib.staticfiles.finders.AppDirectoriesFinder',
    'compressor.finders.CompressorFinder',
)

# Uploads
MEDIA_ROOT = '/tmp/panda'
EXPORT_ROOT = '/tmp/panda_exports'

# Make this unique, and don't share it with anybody.
SECRET_KEY = '-lyd+@8@=9oni01+gjvb(txz3%hh_7a9m5*n0q^ce5+&c1fkm('

# List of callables that know how to import templates from various sources.
TEMPLATE_LOADERS = (
    'django.template.loaders.filesystem.Loader',
    'django.template.loaders.app_directories.Loader',
    'django.template.loaders.eggs.Loader',
)

TEMPLATE_CONTEXT_PROCESSORS = (
    'django.core.context_processors.media',
    'django.contrib.auth.context_processors.auth',
    'django.contrib.messages.context_processors.messages',
    'django.core.context_processors.csrf',
    'django.core.context_processors.i18n'
)

MIDDLEWARE_CLASSES = (
    'django.contrib.sessions.middleware.SessionMiddleware',
    'django.middleware.common.CommonMiddleware',
    'django.middleware.csrf.CsrfViewMiddleware',
    'django.contrib.auth.middleware.AuthenticationMiddleware',
    'django.contrib.messages.middleware.MessageMiddleware',
    'panda.middleware.CsrfCookieUsedMiddleware'
)

ROOT_URLCONF = 'config.urls'

TEMPLATE_DIRS = (
    os.path.join(SITE_ROOT, 'templates')
)

INSTALLED_APPS = (
    'longerusername',
    'django.contrib.auth',
    'django.contrib.contenttypes',
    'django.contrib.sessions',
    'django.contrib.messages',
    'django.contrib.admin',
    'django.contrib.humanize',
    'django.contrib.sites',
    'django.contrib.staticfiles',

    'south',
    'tastypie',
    'djcelery',
    'compressor',
    'livesettings',

    'jumpstart',
    'panda',
    'client'
)

SESSION_COOKIE_AGE = 2592000    # 30 days

AUTH_PROFILE_MODULE = 'panda.UserProfile'

# Django-compressor
COMPRESS_ENABLED = False 

# Celery
import djcelery
djcelery.setup_loader()

BROKER_TRANSPORT = 'sqlalchemy'
BROKER_URL = 'postgresql://%(USER)s:%(PASSWORD)s@%(HOST)s/%(NAME)s' % DATABASES['default']
CELERY_RESULT_DBURI = 'postgresql://%(USER)s:%(PASSWORD)s@%(HOST)s/%(NAME)s' % DATABASES['default'] 
CELERYD_HIJACK_ROOT_LOGGER = False
CELERYD_CONCURRENCY = 1
CELERY_IGNORE_RESULT = True
CELERY_STORE_ERRORS_EVEN_IF_IGNORED = True
CELERYBEAT_SCHEDULE_FILENAME = 'celerybeat-schedule'

from celery.schedules import crontab

CELERYBEAT_SCHEDULE = {
    'purge_orphaned_uploads': {
        'task': 'panda.tasks.cron.purge_orphaned_uploads',
        'schedule': crontab(minute=0, hour=2),
        'kwargs': { 'fake': False }
    },
    'run_subscriptions': {
        'task': 'panda.tasks.cron.run_subscriptions',
        'schedule': crontab(minute=30, hour=2)
    },
    'run_admin_alerts': {
        'task': 'panda.tasks.cron.run_admin_alerts',
        'schedule': crontab(minute=0, hour=4)
    }
}

# South
SOUTH_TESTS_MIGRATE = False

# Hack, see: http://stackoverflow.com/questions/3898239/souths-syncdb-migrate-creates-pages-of-output
import south.logger

# Logging
LOGGING = {
    'version': 1,
    'disable_existing_loggers': True,
    'formatters': {
        'standard': {
            'format': '%(asctime)s [%(levelname)s] %(name)s: %(message)s'
        },
    },
    'handlers': {  
        'console': {
            'level':'DEBUG',
            'class':'logging.StreamHandler',
            'formatter': 'standard'
        },
        'default': {
            'level':'INFO',
            'class':'loghandlers.GroupWriteRotatingFileHandler',
            'filename': '/var/log/panda/panda.log',
            'maxBytes': 1024*1024*5, # 5 MB
            'backupCount': 5,
            'formatter':'standard',
        },
        'request_handler': {
                'level':'INFO',
                'class':'loghandlers.GroupWriteRotatingFileHandler',
                'filename': '/var/log/panda/requests.log',
                'maxBytes': 1024*1024*5, # 5 MB
                'backupCount': 5,
                'formatter':'standard',
        },  
        'backend_handler': {
                'level':'DEBUG',
                'class':'django.utils.log.NullHandler',
        },
    },
    'loggers': {
        '': {
            'handlers': ['default', 'console'],
            'level': 'DEBUG',
            'propagate': True
        },
        'django.request': {
            'handlers': ['request_handler', 'console'],
            'level': 'DEBUG',
            'propagate': False
        },
        'django.db': { 
            'handlers': ['backend_handler'],
            'level': 'DEBUG',
            'propagate': False
        },
        'south': {
            'handlers': ['console'],
            'level': 'INFO',
            'propogate': False
        },
        'keyedcache': {
            'handlers': ['console'],
            'level': 'ERROR',
            'propogate': False
        },
        'requests.packages.urllib3.connectionpool': {
            'handlers': ['console'],
            'level': 'ERROR',
            'propogate': False
        }
    }
}

# Solr
SOLR_ENDPOINT = 'http://localhost:8983/solr'
SOLR_DATA_CORE = 'data'
SOLR_DATASETS_CORE = 'datasets'
SOLR_DIRECTORY = '/var/solr'

# Miscellaneous configuration
PANDA_VERSION = '1.1.2'
PANDA_DEFAULT_SEARCH_GROUPS = 10
PANDA_DEFAULT_SEARCH_ROWS_PER_GROUP = 5
PANDA_DEFAULT_SEARCH_ROWS = 50
PANDA_SNIFFER_MAX_SAMPLE_SIZE = 1024 * 100  # 100 KB
PANDA_SAMPLE_DATA_ROWS = 5
PANDA_SCHEMA_SAMPLE_ROWS = 100
PANDA_ACTIVATION_PERIOD = datetime.timedelta(days=30)
PANDA_AVAILABLE_SPACE_WARN = 1024 * 1024 * 1024 * 2 # 2GB
PANDA_AVAILABLE_SPACE_CRITICAL = 1024 * 1024 * 1024 * 1 # 1GB
PANDA_NOTIFICATIONS_TO_SHOW = 50

PANDA_UNCATEGORIZED_ID = 0
PANDA_UNCATEGORIZED_SLUG = 'uncategorized'
# running this through gettext causes file uploads not to work, so disabled until solved!
PANDA_UNCATEGORIZED_NAME = _('Uncategorized')

MOMENT_LANGUAGE_MAPPING = {
    'en': None,
    'es': 'es',
    'de': 'de'
}

# Allow for local (per-user) override
try:
    from local_settings import *
except ImportError:
    pass


########NEW FILE########
__FILENAME__ = urls
#!/usr/bin/env python

from django.conf import settings
from django.conf.urls.defaults import include, patterns
from django.contrib import admin

from longerusername.forms import AuthenticationForm

# Jumpstart mode
if settings.SETTINGS == 'jumpstart':
    urlpatterns = patterns('',
        (r'', include('jumpstart.urls')),
        (r'^site_media/(?P<path>.*)$', 'django.views.static.serve', {
            'document_root': settings.STATIC_ROOT,
            'show_indexes': True
        }),
    )
# Normal mode
else:
    admin.autodiscover()
    admin.site.index_template = 'admin/panda_index.html'
    admin.site.login_form = AuthenticationForm

    urlpatterns = patterns('',
        (r'', include('panda.urls')),
        (r'', include('client.urls')),
        (r'^admin/settings/', include('livesettings.urls')),
        (r'^admin/', include(admin.site.urls)),

        # Should never be used in production, as nginx will serve these paths
        (r'^site_media/(?P<path>.*)$', 'django.views.static.serve', {
            'document_root': settings.STATIC_ROOT,
            'show_indexes': True
        }),
    )

########NEW FILE########
__FILENAME__ = conf
# -*- coding: utf-8 -*-
#
# PANDA documentation build configuration file, created by
# sphinx-quickstart on Fri Apr 15 21:52:09 2011.
#
# This file is execfile()d with the current directory set to its containing dir.
#
# Note that not all possible configuration values are present in this
# autogenerated file.
#
# All configuration values have a default; values that are commented out
# serve to show the default.

import os

# If extensions (or modules to document with autodoc) are in another directory,
# add these directories to sys.path here. If the directory is relative to the
# documentation root, use os.path.abspath to make it absolute, like shown here.
#sys.path.insert(0, os.path.abspath('..'))

# -- General configuration -----------------------------------------------------

# If your documentation needs a minimal Sphinx version, state it here.
#needs_sphinx = '1.0'

# Add any Sphinx extension module names here, as strings. They can be extensions
# coming with Sphinx (named 'sphinx.ext.*') or your custom ones.
extensions = ['sphinx.ext.autodoc']

# Add any paths that contain templates here, relative to this directory.
templates_path = ['_templates']

# The suffix of source filenames.
source_suffix = '.rst'

# The encoding of source files.
#source_encoding = 'utf-8-sig'

# The master toctree document.
master_doc = 'index'

# General information about the project.
project = u'panda'
copyright = u'2012, PANDA Project'

# The version info for the project you're documenting, acts as replacement for
# |version| and |release|, also used in various other places throughout the
# built documents.
#
# The short X.Y version.
version = '1.1.2'
# The full version, including alpha/beta/rc tags.
release = '1.1.2'

# The language for content autogenerated by Sphinx. Refer to documentation
# for a list of supported languages.
#language = None

# There are two options for replacing |today|: either, you set today to some
# non-false value, then it is used:
#today = ''
# Else, today_fmt is used as the format for a strftime call.
#today_fmt = '%B %d, %Y'

# List of patterns, relative to source directory, that match files and
# directories to ignore when looking for source files.
exclude_patterns = ['_build']

# The reST default role (used for this markup: `text`) to use for all documents.
#default_role = None

# If true, '()' will be appended to :func: etc. cross-reference text.
#add_function_parentheses = True

# If true, the current module name will be prepended to all description
# unit titles (such as .. function::).
#add_module_names = True

# If true, sectionauthor and moduleauthor directives will be shown in the
# output. They are ignored by default.
#show_authors = False

# The name of the Pygments (syntax highlighting) style to use.
pygments_style = 'sphinx'

# A list of ignored prefixes for module index sorting.
#modindex_common_prefix = []


# -- Options for HTML output ---------------------------------------------------

# The theme to use for HTML and HTML Help pages.  See the documentation for
# a list of builtin themes.
html_theme = 'default'

# Theme options are theme-specific and customize the look and feel of a theme
# further.  For a list of options available for each theme, see the
# documentation.
#html_theme_options = {}

# Add any paths that contain custom themes here, relative to this directory.
#html_theme_path = []

# The name for this set of Sphinx documents.  If None, it defaults to
# "<project> v<release> documentation".
#html_title = None

# A shorter title for the navigation bar.  Default is the same as html_title.
#html_short_title = None

# The name of an image file (relative to this directory) to place at the top
# of the sidebar.
#html_logo = None

# The name of an image file (within the static path) to use as favicon of the
# docs.  This file should be a Windows icon file (.ico) being 16x16 or 32x32
# pixels large.
#html_favicon = None

# Add any paths that contain custom static files (such as style sheets) here,
# relative to this directory. They are copied after the builtin static files,
# so a file named "default.css" will overwrite the builtin "default.css".
html_static_path = ['_static']

# If not '', a 'Last updated on:' timestamp is inserted at every page bottom,
# using the given strftime format.
#html_last_updated_fmt = '%b %d, %Y'

# If true, SmartyPants will be used to convert quotes and dashes to
# typographically correct entities.
#html_use_smartypants = True

# Custom sidebar templates, maps document names to template names.
html_sidebars = {
   '*': ['localtoc.html', 'searchbox.html', 'relations.html']
}

# Additional templates that should be rendered to pages, maps page names to
# template names.
#html_additional_pages = {}

# If false, no module index is generated.
#html_domain_indices = True

# If false, no index is generated.
#html_use_index = True

# If true, the index is split into individual pages for each letter.
#html_split_index = False

# If true, links to the reST sources are added to the pages.
#html_show_sourcelink = True

# If true, "Created using Sphinx" is shown in the HTML footer. Default is True.
#html_show_sphinx = True

# If true, "(C) Copyright ..." is shown in the HTML footer. Default is True.
#html_show_copyright = True

# If true, an OpenSearch description file will be output, and all pages will
# contain a <link> tag referring to it.  The value of this option must be the
# base URL from which the finished HTML is served.
#html_use_opensearch = ''

# This is the file name suffix for HTML files (e.g. ".xhtml").
#html_file_suffix = None

# Output file base name for HTML help builder.
htmlhelp_basename = 'pandadoc'

html_domain_indices = False
html_use_index = False

# -- Options for LaTeX output --------------------------------------------------

# The paper size ('letter' or 'a4').
#latex_paper_size = 'letter'

# The font size ('10pt', '11pt' or '12pt').
#latex_font_size = '10pt'

# Grouping the document tree into LaTeX files. List of tuples
# (source start file, target name, title, author, documentclass [howto/manual]).
latex_documents = [
  ('index', 'panda.tex', u'PANDA Project Documentation',
   u'PANDA Project', 'manual'),
]

# The name of an image file (relative to this directory) to place at the top of
# the title page.
#latex_logo = None

# For "manual" documents, if this is true, then toplevel headings are parts,
# not chapters.
#latex_use_parts = False

# If true, show page references after internal links.
#latex_show_pagerefs = False

# If true, show URL addresses after external links.
#latex_show_urls = False

# Additional stuff for the LaTeX preamble.
#latex_preamble = ''

# Documents to append as an appendix to all manuals.
#latex_appendices = []

# If false, no module index is generated.
#latex_domain_indices = True


########NEW FILE########
__FILENAME__ = fabfile
#!/usr/bin/env python

from fabric.api import *

"""
Base configuration
"""
env.user = 'ubuntu'
env.project_name = 'panda'
env.database_password = 'panda'
env.path = '/opt/%(project_name)s' % env
env.solr_path = '/opt/solr/panda/solr'
env.repository_url = 'git://github.com/pandaproject/panda.git'
env.hosts = ['panda.beta.tribapps.com']
env.vars = 'DEPLOYMENT_TARGET="deployed"'

env.local_solr = 'apache-solr-3.4.0/example'
env.local_solr_home = '/var/solr'

env.local_test_email = 'panda@pandaproject.net'
env.local_test_api_key = 'edfe6c5ffd1be4d3bf22f69188ac6bc0fc04c84b'
env.local_test_xhr_path = 'client/static/js/spec/mock_xhr_responses.js'
    
"""
Branches
"""
def stable():
    """
    Work on stable branch.
    """
    env.branch = 'stable'

def master():
    """
    Work on development branch.
    """
    env.branch = 'master'

def branch(branch_name):
    """
    Work on any specified branch.
    """
    env.branch = branch_name
    
"""
Commands - setup
"""
def setup():
    """
    Setup a fresh virtualenv, install everything we need, and fire up the database.
    
    Does NOT perform the functions of deploy().
    """
    require('branch', provided_by=[stable, master, branch])
    
    setup_directories()
    clone_repo()
    checkout_latest()
    install_requirements()
    destroy_database()
    create_database()
    syncdb()

def setup_directories():
    """
    Create directories necessary for deployment.
    """
    sudo('mkdir -p %(path)s' % env)

def clone_repo():
    """
    Do initial clone of the git repository.
    """
    sudo('git clone %(repository_url)s %(path)s' % env)

def checkout_latest():
    """
    Pull the latest code on the specified branch.
    """
    sudo('cd %(path)s; git checkout %(branch)s; git pull origin %(branch)s' % env)

def install_requirements():
    """
    Install the required packages using pip.
    """
    sudo('pip install -r %(path)s/requirements.txt' % env)

"""
Commands - deployment
"""
def deploy():
    """
    Deploy the latest version of the site to the server and restart Apache2.
    """
    require('branch', provided_by=[stable, master, branch])
    
    checkout_latest()
    collect_static_files()
    reload_app()

def collect_static_files():
    """
    Collect static files on the server.
    """
    with cd('%(path)s' % env):
        sudo('%(vars)s python manage.py collectstatic --noinput' % env, user="panda")
       
def reload_app(): 
    """
    Restart the uwsgi server.
    """
    sudo('service uwsgi restart')
    sudo('service celeryd restart')

def update_requirements():
    """
    Update the installed dependencies the server.
    """
    sudo('pip install -U -r %(path)s/requirements.txt' % env)
    
"""
Commands - data
"""
def reset_database():
    """
    Drop and recreate the project database.
    """
    with settings(warn_only=True):
        sudo('service celeryd stop')

    sudo('service postgresql restart') # disconnect any active users

    destroy_database()
    create_database()
    syncdb()

    sudo('service celeryd start') 

def create_database():
    """
    Creates the user and database for this project.
    """
    sudo('echo "CREATE USER %(project_name)s WITH PASSWORD \'%(database_password)s\';" | psql postgres' % env, user='postgres')
    sudo('createdb -O %(project_name)s %(project_name)s' % env, user='postgres')
    
def destroy_database():
    """
    Destroys the user and database for this project.
    
    Will not cause the fab to fail if they do not exist.
    """
    with settings(warn_only=True):
        sudo('dropdb %(project_name)s' % env, user='postgres')
        sudo('dropuser %(project_name)s' % env, user='postgres')
        
def syncdb():
    """
    Sync the Django models to the database.
    """
    with cd('%(path)s' % env):
        sudo('%(vars)s python manage.py syncdb --noinput' % env, user='panda')
        sudo('%(vars)s python manage.py migrate --noinput' % env, user='panda')
        sudo('%(vars)s python manage.py loaddata panda/fixtures/init_panda.json' % env)

def reset_solr():
    """
    Update configuration, blow away current data, and restart Solr.
    """
    with settings(warn_only=True):
        sudo('service solr stop')

    sudo('sudo mkdir -p %(solr_path)s' % env)

    sudo('cp %(path)s/setup_panda/solr.xml %(solr_path)s/solr.xml' % env)

    # data
    sudo('mkdir -p %(solr_path)s/pandadata/conf' % env)
    sudo('mkdir -p %(solr_path)s/pandadata/lib' % env)
    sudo('rm -rf %(solr_path)s/pandadata/data' % env)

    sudo('cp %(path)s/setup_panda/data_schema.xml %(solr_path)s/pandadata/conf/schema.xml' % env)
    sudo('cp %(path)s/setup_panda/english_names.txt %(solr_path)s/pandadata/conf/english_names.txt' % env)
    sudo('cp %(path)s/setup_panda/solrconfig.xml %(solr_path)s/pandadata/conf/solrconfig.xml' % env)
    sudo('cp %(path)s/setup_panda/panda.jar %(solr_path)s/pandadata/lib/panda.jar' % env)
    sudo('rm -rf %(solr_path)s/pandadata/data' % env)

    # data_test
    sudo('mkdir -p %(solr_path)s/pandadata_test/conf' % env)
    sudo('mkdir -p %(solr_path)s/pandadata_test/lib' % env)
    sudo('rm -rf %(solr_path)s/pandadata_test/data' % env)

    sudo('cp %(path)s/setup_panda/data_schema.xml %(solr_path)s/pandadata_test/conf/schema.xml' % env)
    sudo('cp %(path)s/setup_panda/english_names.txt %(solr_path)s/pandadata_test/conf/english_names.txt' % env)
    sudo('cp %(path)s/setup_panda/solrconfig.xml %(solr_path)s/pandadata_test/conf/solrconfig.xml' % env)
    sudo('cp %(path)s/setup_panda/panda.jar %(solr_path)s/pandadata_test/lib/panda.jar' % env)
    sudo('rm -rf %(solr_path)s/pandadata_test/data' % env)

    # datasets
    sudo('mkdir -p %(solr_path)s/pandadatasets/conf' % env)
    sudo('rm -rf %(solr_path)s/pandadatasets/data' % env)

    sudo('cp %(path)s/setup_panda/solrconfig.xml %(solr_path)s/pandadatasets/conf/solrconfig.xml' % env)
    sudo('cp %(path)s/setup_panda/datasets_schema.xml %(solr_path)s/pandadatasets/conf/schema.xml' % env)

    # datasets_test
    sudo('mkdir -p %(solr_path)s/pandadatasets_test/conf' % env)
    sudo('rm -rf %(solr_path)s/pandadatasets_test/data' % env)

    sudo('cp %(path)s/setup_panda/solrconfig.xml %(solr_path)s/pandadatasets_test/conf/solrconfig.xml' % env)
    sudo('cp %(path)s/setup_panda/datasets_schema.xml %(solr_path)s/pandadatasets_test/conf/schema.xml' % env)

    sudo('chown -R solr:solr %(solr_path)s' % env)
    sudo('service solr start')

def reset_jumpstart():
    """
    Reset the configuration to run the jumpstart server.
    """
    sudo('service uwsgi stop')
    sudo('sudo cp %(path)s/setup_panda/uwsgi_jumpstart.conf /etc/init/uwsgi.conf' % env)
    sudo('service uwsgi start')

"""
Commands - Local development
"""
def local_reset():
    """
    Reset the local database and Solr instance.
    """
    local_reset_database()
    local_reset_solr()

def local_reset_database():
    """
    Reset the local database.
    """
    local('dropdb %(project_name)s' % env)
    local('createdb -O %(project_name)s %(project_name)s' % env)
    local('python manage.py syncdb --noinput' % env)
    local('python manage.py migrate --noinput' % env)
    local('python manage.py loaddata panda/fixtures/init_panda.json' % env)
    local('python manage.py loaddata panda/fixtures/test_users.json' % env)

def local_reset_solr():
    """
    Reset the local solr configuration.
    """
    local('cp setup_panda/solr.xml %(local_solr_home)s/solr.xml' % env)

    # data
    local('mkdir -p %(local_solr_home)s/pandadata/conf' % env)
    local('mkdir -p %(local_solr_home)s/pandadata/lib' % env)
    local('rm -rf %(local_solr_home)s/pandadata/data' % env)

    local('cp setup_panda/panda.jar %(local_solr_home)s/pandadata/lib/panda.jar' % env)
    local('cp setup_panda/solrconfig.xml %(local_solr_home)s/pandadata/conf/solrconfig.xml' % env)
    local('cp setup_panda/data_schema.xml %(local_solr_home)s/pandadata/conf/schema.xml' % env)
    local('cp setup_panda/english_names.txt %(local_solr_home)s/pandadata/conf/english_names.txt' % env)

    # data_test
    local('mkdir -p %(local_solr_home)s/pandadata_test/conf' % env)
    local('mkdir -p %(local_solr_home)s/pandadata_test/lib' % env)
    local('rm -rf %(local_solr_home)s/pandadata_test/data' % env)

    local('cp setup_panda/panda.jar %(local_solr_home)s/pandadata_test/lib/panda.jar' % env)
    local('cp setup_panda/solrconfig.xml %(local_solr_home)s/pandadata_test/conf/solrconfig.xml' % env)
    local('cp setup_panda/data_schema.xml %(local_solr_home)s/pandadata_test/conf/schema.xml' % env)
    local('cp setup_panda/english_names.txt %(local_solr_home)s/pandadata_test/conf/english_names.txt' % env)

    # datasets
    local('mkdir -p %(local_solr_home)s/pandadatasets/conf' % env)
    local('rm -rf %(local_solr_home)s/pandadatasets/data' % env)

    local('cp setup_panda/solrconfig.xml %(local_solr_home)s/pandadatasets/conf/solrconfig.xml' % env)
    local('cp setup_panda/datasets_schema.xml %(local_solr_home)s/pandadatasets/conf/schema.xml' % env)

    # datasets_test
    local('mkdir -p %(local_solr_home)s/pandadatasets_test/conf' % env)
    local('rm -rf %(local_solr_home)s/pandadatasets_test/data' % env)

    local('cp setup_panda/solrconfig.xml %(local_solr_home)s/pandadatasets_test/conf/solrconfig.xml' % env)
    local('cp setup_panda/datasets_schema.xml %(local_solr_home)s/pandadatasets_test/conf/schema.xml' % env)

def local_solr():
    """
    Start the local Solr instance.
    """
    local('cd %(local_solr)s && java -Xms256M -Xmx512G -Dsolr.solr.home=%(local_solr_home)s -jar start.jar' % env)

def local_email():
    local('python -m smtpd -n -c DebuggingServer localhost:1025')

def make_fixtures():
    """
    Creates a consistent set of local test data and generates fixtures.

    Notes:
    * Will reset the database.
    * Local server (runserver, celeryd and solr) must be running.
    """
    local('python manage.py flush --noinput')
    local('python manage.py loaddata panda/fixtures/init_panda.json' % env)
    local('curl --data-binary "{ \\"delete\\": { \\"query\\": \\"*:*\\" } }" -H "Content-type:application/xml" "http://localhost:8983/solr/data/update?commit=true"')
    local('curl --data-binary "{ \\"delete\\": { \\"query\\": \\"*:*\\" } }" -H "Content-type:application/xml" "http://localhost:8983/solr/datasets/update?commit=true"')

    local('curl -H "PANDA_EMAIL: %(local_test_email)s" -H "PANDA_API_KEY: %(local_test_api_key)s" -H "Content-Type: application/json" --data-binary "{ \\"name\\": \\"Test\\" }" "http://localhost:8000/api/1.0/dataset/"' % env)
    local('curl -H "PANDA_EMAIL: %(local_test_email)s" -H "PANDA_API_KEY: %(local_test_api_key)s" -F file=@test_data/contributors.csv -F dataset_slug=test "http://localhost:8000/data_upload/"' % env)
    local('curl -H "PANDA_EMAIL: %(local_test_email)s" -H "PANDA_API_KEY: %(local_test_api_key)s" "http://localhost:8000/api/1.0/dataset/test/import/1/"' % env)

    mock_xhr_responses = ['window.MOCK_XHR_RESPONSES = {};']

    response = local('curl "http://localhost:8000/api/1.0/task/1/?format=json&email=%(local_test_email)s&api_key=%(local_test_api_key)s"' % env, capture=True)
    mock_xhr_responses.append('MOCK_XHR_RESPONSES.task = \'' + response.replace('\\', '\\\\') + '\';')

    response = local('curl "http://localhost:8000/api/1.0/task/?format=json&email=%(local_test_email)s&api_key=%(local_test_api_key)s"' % env, capture=True)
    mock_xhr_responses.append('MOCK_XHR_RESPONSES.tasks = \'' + response.replace('\\', '\\\\') + '\';')

    response = local('curl "http://localhost:8000/api/1.0/dataset/test/?format=json&email=%(local_test_email)s&api_key=%(local_test_api_key)s"' % env, capture=True)
    mock_xhr_responses.append('MOCK_XHR_RESPONSES.dataset = \'' + response.replace('\\', '\\\\') + '\';')

    response = local('curl "http://localhost:8000/api/1.0/dataset/?format=json&email=%(local_test_email)s&api_key=%(local_test_api_key)s"' % env, capture=True)
    mock_xhr_responses.append('MOCK_XHR_RESPONSES.datasets = \'' + response.replace('\\', '\\\\') + '\';')

    response = local('curl "http://localhost:8000/api/1.0/data/?q=Tribune&format=json&email=%(local_test_email)s&api_key=%(local_test_api_key)s"' % env, capture=True)
    mock_xhr_responses.append('MOCK_XHR_RESPONSES.search = \'' + response.replace('\\', '\\\\') + '\';')

    response = local('curl "http://localhost:8000/api/1.0/dataset/test/data/?q=Tribune&format=json&email=%(local_test_email)s&api_key=%(local_test_api_key)s"' % env, capture=True)
    mock_xhr_responses.append('MOCK_XHR_RESPONSES.dataset_search = \'' + response.replace('\\', '\\\\') + '\';')

    # Task
    with open('%(local_test_xhr_path)s' % env, 'w') as f:
        f.write('\n'.join(mock_xhr_responses))

def coverage():
    local('coverage erase')
    local('coverage run --source panda manage.py test panda')
    local('coverage html --omit "panda/migrations/*,panda/tests/*" -d coverage_html')
    local('open coverage_html/index.html')

def makemessages():
    local('python manage.py makemessages -a -i _site -i media -i docs -i coverage_html')
    local('python manage.py makemessages -a -d djangojs -e js -i _site -i media -i docs -i coverage_html')

def compilemessages():
    local('python manage.py compilemessages')


########NEW FILE########
__FILENAME__ = daemon
#!/usr/bin/env python

"""
From: http://www.jejik.com/articles/2007/02/a_simple_unix_linux_daemon_in_python/
"""

import sys, os, time, atexit
from signal import SIGTERM 

class Daemon:
	"""
	A generic daemon class.
	
	Usage: subclass the Daemon class and override the run() method
	"""
	def __init__(self, pidfile, stdin='/dev/null', stdout='/dev/null', stderr='/dev/null'):
		self.stdin = stdin
		self.stdout = stdout
		self.stderr = stderr
		self.pidfile = pidfile
	
	def daemonize(self):
		"""
		do the UNIX double-fork magic, see Stevens' "Advanced 
		Programming in the UNIX Environment" for details (ISBN 0201563177)
		http://www.erlenstar.demon.co.uk/unix/faq_2.html#SEC16
		"""
		try: 
			pid = os.fork() 
			if pid > 0:
				# exit first parent
				sys.exit(0) 
		except OSError, e: 
			sys.stderr.write("fork #1 failed: %d (%s)\n" % (e.errno, e.strerror))
			sys.exit(1)
	
		# decouple from parent environment
		os.chdir("/") 
		os.setsid() 
		os.umask(0) 
	
		# do second fork
		try: 
			pid = os.fork() 
			if pid > 0:
				# exit from second parent
				sys.exit(0) 
		except OSError, e: 
			sys.stderr.write("fork #2 failed: %d (%s)\n" % (e.errno, e.strerror))
			sys.exit(1) 
	
		# redirect standard file descriptors
		sys.stdout.flush()
		sys.stderr.flush()
		si = file(self.stdin, 'r')
		so = file(self.stdout, 'a+')
		se = file(self.stderr, 'a+', 0)
		os.dup2(si.fileno(), sys.stdin.fileno())
		os.dup2(so.fileno(), sys.stdout.fileno())
		os.dup2(se.fileno(), sys.stderr.fileno())
	
		# write pidfile
		atexit.register(self.delpid)
		pid = str(os.getpid())
		file(self.pidfile,'w+').write("%s\n" % pid)
	
	def delpid(self):
		os.remove(self.pidfile)

	def start(self):
		"""
		Start the daemon
		"""
		# Check for a pidfile to see if the daemon already runs
		try:
			pf = file(self.pidfile,'r')
			pid = int(pf.read().strip())
			pf.close()
		except IOError:
			pid = None
	
		if pid:
			message = "pidfile %s already exist. Daemon already running?\n"
			sys.stderr.write(message % self.pidfile)
			sys.exit(1)
		
		# Start the daemon
		self.daemonize()
		self.run()

	def stop(self):
		"""
		Stop the daemon
		"""
		# Get the pid from the pidfile
		try:
			pf = file(self.pidfile,'r')
			pid = int(pf.read().strip())
			pf.close()
		except IOError:
			pid = None
	
		if not pid:
			message = "pidfile %s does not exist. Daemon not running?\n"
			sys.stderr.write(message % self.pidfile)
			return # not an error in a restart

		# Try killing the daemon process	
		try:
			while 1:
				os.kill(pid, SIGTERM)
				time.sleep(0.1)
		except OSError, err:
			err = str(err)
			if err.find("No such process") > 0:
				if os.path.exists(self.pidfile):
					os.remove(self.pidfile)
			else:
				print str(err)
				sys.exit(1)

	def restart(self):
		"""
		Restart the daemon
		"""
		self.stop()
		self.start()

	def run(self):
		"""
		You should override this method when you subclass Daemon. It will be called after the process has been
		daemonized by start() or restart().
		"""
		pass


########NEW FILE########
__FILENAME__ = urls
#!/usr/bin/env python

from django.conf.urls.defaults import patterns, url

from jumpstart import views

urlpatterns = patterns('',
    url(r'^wait$', views.wait, name='wait'),
    url(r'^$', views.jumpstart, name='jumpstart')
)


########NEW FILE########
__FILENAME__ = views
#!/usr/bin/env python

import os
import random
import subprocess
import time

from django.conf import settings
from django.shortcuts import render_to_response
from django.template import RequestContext
from django.utils.translation import ugettext as _
from pytz import common_timezones
from tastypie.models import ApiKey

from daemon import Daemon
from panda.models import UserProxy

LOCAL_SETTINGS_PATH = '%s/local_settings.py' % settings.SITE_ROOT
RESTART_SCRIPT_PATH = '%s/jumpstart/restart-uwsgi.sh' % settings.SITE_ROOT 

class RestartDaemon(Daemon):
    """
    Simple daemon so that a uwsgi process can reboot itself
    """
    def run(self):
        # Sleep for a moment to give uwsgi a chance to return a response
        time.sleep(5)

        subprocess.call(['sudo', RESTART_SCRIPT_PATH])
        
        if os.path.exists(self.pidfile):
            os.remove(self.pidfile)

def jumpstart(request):
    context = RequestContext(request, {
        'settings': settings,
        'languages': (
            ('en', _('English')),
            ('de', _('German')),
            ('es', _('Spanish')),
            ('it', _('Italian'))
        ),
        'timezones': common_timezones
    })

    return render_to_response('jumpstart/index.html', context)

def wait(request):
    language = request.POST.get('language', 'en')
    timezone = request.POST['timezone']
    email = request.POST['email']
    password = request.POST['password']

    secret_key = ''.join([random.choice('abcdefghijklmnopqrstuvwxyz0123456789!@#$%^&*(-_=+)') for i in range(50)])

    # Test if running under runserver
    # (Via: http://stackoverflow.com/questions/10962703/django-distinguish-development-server-manage-py-runserver-from-the-regular-o)
    wsgi_wrapper = request.META.get('wsgi.file_wrapper', None)
    wsgi_wrapper_path = wsgi_wrapper.__module__ if wsgi_wrapper else None

    with open(LOCAL_SETTINGS_PATH, 'w') as f:
        f.write("LANGUAGE_CODE = '%s'\n" % language)
        f.write("TIME_ZONE = '%s'\n" % timezone)

        f.write("SECRET_KEY = '%s'\n" % secret_key)

        if wsgi_wrapper_path:
            f.write("DEBUG = 'True'\n")

    admin = UserProxy.objects.create_user(email, email, password)
    admin.is_staff = True
    admin.is_superuser = True
    admin.save()

    ApiKey.objects.get_or_create(user=admin)

    if not wsgi_wrapper_path:
        daemon = RestartDaemon(settings.DAEMON_PID_PATH, stdout=settings.DAEMON_LOG_PATH)
        daemon.start()

    return render_to_response('jumpstart/wait.html', { 'settings': settings })


########NEW FILE########
__FILENAME__ = loghandlers
#!/usr/bin/env python

import logging
import os

class GroupWriteRotatingFileHandler(logging.handlers.RotatingFileHandler):
    """
    Rotating logger which also adds group+writable permissions.

    Tip from:
    http://stackoverflow.com/questions/1407474/does-python-logging-handlers-rotatingfilehandler-allow-creation-of-a-group-writab/6779307#6779307
    """
    def _open(self):
        old_umask = os.umask(0o002)
        f = logging.handlers.RotatingFileHandler._open(self)
        os.umask(old_umask)

        return f

########NEW FILE########
__FILENAME__ = manage
#!/usr/bin/env python
import os
from django.core.management import execute_manager

if not os.environ.has_key("DJANGO_SETTINGS_MODULE"):
    if not os.environ.has_key("DEPLOYMENT_TARGET"):
        os.environ["DJANGO_SETTINGS_MODULE"] = "config.settings"
    else:
        os.environ["DJANGO_SETTINGS_MODULE"] = "config.%s.settings" % os.environ["DEPLOYMENT_TARGET"]

settings_module = os.environ["DJANGO_SETTINGS_MODULE"]

try:
    settings = __import__(settings_module, globals(), locals(), ['settings'], -1)
except ImportError:
    import sys
    sys.stderr.write("Error: Can't find the file '%s.py'.\n" % settings_module)
    sys.exit(1)

if __name__ == "__main__":
    execute_manager(settings)

########NEW FILE########
__FILENAME__ = admin
#!/usr/bin/env python

from StringIO import StringIO

from django import forms
from django.conf import settings
from django.conf.urls import patterns, url
from django.contrib import admin
from django.contrib.admin import helpers
from django.contrib.auth.admin import UserAdmin
from django.contrib.auth.forms import UserChangeForm
from django.contrib.auth.models import Group, User
from django.contrib.sites.models import Site
from django.core.exceptions import PermissionDenied
from django.core.urlresolvers import reverse
from django.db import models, transaction
from django.http import HttpResponseRedirect
from django.shortcuts import get_object_or_404, render_to_response
from django.template import RequestContext
from django.utils.encoding import force_unicode
from django.utils.translation import ugettext_lazy as _
from djcelery.models import CrontabSchedule, IntervalSchedule, PeriodicTask, TaskState, WorkerState
from livesettings import config_value
from tastypie.admin import ApiKeyInline
from tastypie.models import ApiKey

from csvkit import CSVKitReader
from csvkit.sniffer import sniff_dialect as csvkit_sniff
from panda import solr
from panda.models import Category, TaskStatus, UserProfile, UserProxy

# Hide celery monitors
admin.site.unregister(CrontabSchedule)
admin.site.unregister(IntervalSchedule)
admin.site.unregister(PeriodicTask)
admin.site.unregister(TaskState)
admin.site.unregister(WorkerState)

class PandaUserCreationForm(forms.ModelForm):
    """
    Custom User creation form that eliminates duplication between username
    and email.
    """
    class Meta:
        model = UserProxy
        fields = ("email",)

    email = forms.EmailField(label=_("E-mail"), max_length=75)

    def clean_email(self):
        email = self.cleaned_data["email"]
        
        try:
            UserProxy.objects.get(email=email)
        except UserProxy.DoesNotExist:
            return email

        raise forms.ValidationError(_("A user with that email address already exists."))

    def save(self, commit=True):
        user = super(PandaUserCreationForm, self).save(commit=False)
        user.email = user.email.lower()
        user.username = user.email
        user.is_active = False
        user.set_unusable_password()

        if commit:
            user.save()

        return user

class PandaUserChangeForm(UserChangeForm):
    """
    Customized User change form that allows password to be blank.
    (for editing unactivated accounts)
    """
    class Media:
        js = ('panda_user_change_form.js',)

    def __init__(self, *args, **kwargs):
        super(PandaUserChangeForm, self).__init__(*args, **kwargs)

        # We edit the email field and copy it to the username field
        del self.fields['username']
        
        self.fields['password'].required = False

    def save(self, commit=True):
        user = super(PandaUserChangeForm, self).save(commit=False)
        user.email = user.email.lower()
        user.username = user.email

        if commit:
            user.save()

        return user

class PandaApiKeyInline(ApiKeyInline):
    """
    Customized ApiKeyInline that doesn't allow the creation date to be modified.
    """
    readonly_fields = ('created',)

class UserProfileInline(admin.StackedInline):
    """
    Inline for UserProfile which does not allow the activation key to be modified. 
    """
    model = UserProfile
    
    readonly_fields = ('activation_key', 'activation_key_expiration')

class UserModelAdmin(UserAdmin):
    """
    Heavily modified admin page for editing Users. Eliminates duplication between
    username and email fields. Hides unnecessary cruft. Makes timestamp fields
    readonly. Etc.
    """
    inlines = [UserProfileInline, PandaApiKeyInline]
    add_form = PandaUserCreationForm
    form = PandaUserChangeForm

    add_form_template = 'admin/panda/userproxy/add_form.html'

    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('email', 'first_name', 'last_name')}
        ),
    )

    fieldsets = (
        (None, {'fields': ('email', 'password')}),
        (_('Personal info'), {'fields': ('first_name', 'last_name')}),
        (_('Permissions'), {'fields': ('is_active', 'is_staff', 'is_superuser')}),
        (_('Important dates'), {'fields': ('last_login', 'date_joined')}),
    )

    list_display = ('email', 'first_name', 'last_name', 'is_staff', 'is_active')
    search_fields = ('first_name', 'last_name', 'email')
    ordering = ('email',)

    readonly_fields = ('last_login', 'date_joined')

    actions = ['resend_activation']

    def get_urls(self):
        urls = super(UserModelAdmin, self).get_urls()
        custom_urls = patterns('',
            url(r'^(.+)/resend_activation$',
                self.admin_site.admin_view(self.resend_activation_single),
                name='%s_%s_resend_activation' % (self.model._meta.app_label, self.model._meta.module_name)
            ),
            url(r'^add_many/$',
                self.admin_site.admin_view(self.add_many),
                name='%s_%s_add_many' % (self.model._meta.app_label, self.model._meta.module_name)
            ),
        )

        return custom_urls + urls

    def resend_activation_single(self, request, pk):
        if not config_value('EMAIL', 'EMAIL_ENABLED'):
            self.message_user(request, _('Email is not configured for your PANDA.'))

            return HttpResponseRedirect(
                reverse('admin:panda_userproxy_change', args=[pk])
            )

        user = get_object_or_404(UserProxy, pk=pk)
        user_profile = user.get_profile()

        user_profile.generate_activation_key()
        user_profile.save()

        user_profile.send_activation_email()
        self.message_user(request, _('Activation email sent.'))

        return HttpResponseRedirect(
            reverse('admin:panda_userproxy_change', args=[pk])
        )

    def resend_activation(self, request, queryset):
        if not config_value('EMAIL', 'EMAIL_ENABLED'):
            self.message_user(request, _('Email is not configured for your PANDA.'))
            return HttpResponseRedirect(
                reverse('admin:panda_userproxy_changelist')
            )

        users = list(queryset)

        for user in users:
            user_profile = user.get_profile()

            user_profile.generate_activation_key()
            user_profile.save()

            user_profile.send_activation_email()

        self.message_user(request, _('Sent %i activation emails.') % len(users))

    resend_activation.short_description = _('Resend activation email(s)')

    @transaction.commit_on_success
    def add_many(self, request, extra_context=None):
        model = self.model
        opts = model._meta

        context = RequestContext(request, {
            'opts': opts,
            'title': _('Add %s') % force_unicode(opts.verbose_name_plural),
            'media': self.media,
            'error': [],
            'app_label': opts.app_label,
            'email_enabled': config_value('EMAIL', 'EMAIL_ENABLED')
        })
        
        context.update(extra_context or {})

        if request.method == 'POST':
            try:
                user_data = request.POST.get('user-data', '') 

                if not user_data:
                    raise Exception(_('No user data provided.'))

                context['user_data'] = user_data

                try:
                    csv_dialect = csvkit_sniff(user_data)
                except UnicodeDecodeError:
                    raise Exception(_('Only UTF-8 data is supported.'))

                if not csv_dialect:
                    raise Exception(_('Unable to determine the format of the data you entered. Please ensure it is valid CSV data.'))

                reader = CSVKitReader(StringIO(user_data), dialect=csv_dialect)

                emails = 0

                for i, row in enumerate(reader):
                    if len(row) < 4:
                        raise Exception(_('Row %i has less than 4 columns.') % i)
                    if len(row) > 4:
                        raise Exception(_('Row %i has more than 4 columns.') % i)

                    if UserProxy.objects.filter(email=row[0]).count():
                        raise Exception(_('User "%s" already exists')  % row[0])

                    user = UserProxy.objects.create_user(row[0], row[0], row[1] or None)
                    user.is_active = bool(row[1]) # active if a password is provided
                    user.first_name = row[2]
                    user.last_name = row[3]
                    user.save()

                    ApiKey.objects.get_or_create(user=user)

                    if not row[1] and config_value('EMAIL', 'EMAIL_ENABLED'):
                        emails += 1

                self.message_user(request, _('Successfully created %i user(s)') % (i + 1))

                if emails:
                    self.message_user(request, _('Sent %i activation email(s)') % emails)
            except Exception, e:
                context['error'] = e.message

        return render_to_response('admin/panda/userproxy/add_many_form.html', context)

    @transaction.commit_on_success
    def add_view(self, request, form_url='', extra_context=None):
        """
        This method is overriden in its entirety so that the ApiKey inline won't be
        displayed/parsed on the add_form page.
        """
        model = self.model
        opts = model._meta

        if not self.has_add_permission(request):
            raise PermissionDenied

        ModelForm = self.get_form(request)
        formsets = []
        inline_instances = self.get_inline_instances(request)
        if request.method == 'POST':
            form = ModelForm(request.POST, request.FILES)
            if form.is_valid():
                new_object = self.save_form(request, form, change=False)
                form_validated = True
            else:
                form_validated = False
                new_object = self.model()
            
            PANDA_SKIP_INLINES="""prefixes = {}
            for FormSet, inline in zip(self.get_formsets(request), inline_instances):
                prefix = FormSet.get_default_prefix()
                prefixes[prefix] = prefixes.get(prefix, 0) + 1
                if prefixes[prefix] != 1 or not prefix:
                    prefix = "%s-%s" % (prefix, prefixes[prefix])
                formset = FormSet(data=request.POST, files=request.FILES,
                                  instance=new_object,
                                  save_as_new="_saveasnew" in request.POST,
                                  prefix=prefix, queryset=inline.queryset(request))
                formsets.append(formset)
            if all_valid(formsets) and form_validated:"""

            if form_validated:
                self.save_model(request, new_object, form, False)
                self.save_related(request, form, formsets, False)
                self.log_addition(request, new_object)
                return self.response_add(request, new_object)
        else:
            # Prepare the dict of initial data from the request.
            # We have to special-case M2Ms as a list of comma-separated PKs.
            initial = dict(request.GET.items())
            for k in initial:
                try:
                    f = opts.get_field(k)
                except models.FieldDoesNotExist:
                    continue
                if isinstance(f, models.ManyToManyField):
                    initial[k] = initial[k].split(",")
            form = ModelForm(initial=initial)

            PANDA_SKIP_INLINES = """prefixes = {}
            
            for FormSet, inline in zip(self.get_formsets(request), inline_instances):
                prefix = FormSet.get_default_prefix()
                prefixes[prefix] = prefixes.get(prefix, 0) + 1
                if prefixes[prefix] != 1 or not prefix:
                    prefix = "%s-%s" % (prefix, prefixes[prefix])
                formset = FormSet(instance=self.model(), prefix=prefix,
                                  queryset=inline.queryset(request))
                formsets.append(formset)"""

        adminForm = helpers.AdminForm(form, list(self.get_fieldsets(request)),
            self.get_prepopulated_fields(request),
            self.get_readonly_fields(request),
            model_admin=self)
        media = self.media + adminForm.media

        inline_admin_formsets = []
        for inline, formset in zip(inline_instances, formsets):
            fieldsets = list(inline.get_fieldsets(request))
            readonly = list(inline.get_readonly_fields(request))
            prepopulated = dict(inline.get_prepopulated_fields(request))
            inline_admin_formset = helpers.InlineAdminFormSet(inline, formset,
                fieldsets, prepopulated, readonly, model_admin=self)
            inline_admin_formsets.append(inline_admin_formset)
            media = media + inline_admin_formset.media

        context = {
            'title': _('Add %s') % force_unicode(opts.verbose_name),
            'adminform': adminForm,
            'is_popup': "_popup" in request.REQUEST,
            'show_delete': False,
            'media': media,
            'inline_admin_formsets': inline_admin_formsets,
            'errors': helpers.AdminErrorList(form, formsets),
            'app_label': opts.app_label,
            'email_enabled': config_value('EMAIL', 'EMAIL_ENABLED')
        }
        context.update(extra_context or {})
        return self.render_change_form(request, context, form_url=form_url, add=True)

admin.site.unregister(Group)
admin.site.unregister(User)
admin.site.register(UserProxy, UserModelAdmin)

# Hide sites framework
admin.site.unregister(Site)

class CategoryAdmin(admin.ModelAdmin):
    fields = ('name', 'slug')
    prepopulated_fields = { 'slug': ('name', ) }

    def save_model(self, request, obj, form, change):
        """
        On save, update full text metadata of related datasets. 
        """
        if change:
            datasets = list(obj.datasets.all())
            obj.save()

            for dataset in datasets:
                dataset.update_full_text(commit=False)

            solr.commit(settings.SOLR_DATASETS_CORE)
        else:
            obj.save()

    def delete_model(self, request, obj):
        """
        On delete, update full text metadata of related datasets. 
        """
        datasets = list(obj.datasets.all())
        obj.delete()

        for dataset in datasets:
            dataset.update_full_text()

        solr.commit(settings.SOLR_DATASETS_CORE)

admin.site.register(Category, CategoryAdmin)

class TaskStatusAdmin(admin.ModelAdmin):
    fields = ('task_name', 'task_description', 'status', 'message', 'start', 'end', 'traceback', 'creator')
    readonly_fields = ('task_name', 'task_description', 'status', 'message', 'start', 'end', 'traceback', 'creator')
    
    list_display = ('task_name', 'task_description',  'status', 'start', 'end', 'creator')
    list_display_links = ('task_name', 'task_description')
    list_filter = ('status', )

    actions = ['abort_task']

    def get_urls(self):
        urls = super(TaskStatusAdmin, self).get_urls()
        custom_urls = patterns('',
            url(r'^(.+)/abort$',
                self.admin_site.admin_view(self.abort_single),
                name='%s_%s_abort' % (self.model._meta.app_label, self.model._meta.module_name)
            ),
        )

        return custom_urls + urls

    def abort_single(self, request, pk):
        task = get_object_or_404(TaskStatus, pk=pk)

        if task.end:
            self.message_user(request, _('You can not abort a task that has already ended.'))
        else:
            task.request_abort()
            self.message_user(request, _('Attempting to abort task.'))

        return HttpResponseRedirect(
            reverse('admin:panda_taskstatus_changelist')
        )

    def abort_task(self, request, queryset):
        tasks = list(queryset)

        for task in tasks:
            if task.end:
                self.message_user(request, _('You can not abort tasks that have already ended.'))
                return

        for task in tasks:
            task.request_abort()
        
        self.message_user(request, _('Attempting to abort %i task(s).') % len(tasks))

    abort_task.short_description = _('Abort task(s)')

admin.site.register(TaskStatus, TaskStatusAdmin)

admin.site.disable_action('delete_selected')


########NEW FILE########
__FILENAME__ = activity_log
#!/usr/bin/env python

from tastypie import fields
from tastypie.authorization import DjangoAuthorization
from tastypie.exceptions import ImmediateHttpResponse
from tastypie.http import HttpConflict
from django.utils.translation import ugettext_lazy as _

from panda.api.utils import PandaAuthentication, PandaModelResource, PandaSerializer
from django.db import IntegrityError
from panda.models import ActivityLog, UserProxy

class ActivityLogResource(PandaModelResource):
    """
    API resource for DataUploads.
    """
    from panda.api.users import UserResource

    creator = fields.ForeignKey(UserResource, 'user', full=True)

    class Meta:
        queryset = ActivityLog.objects.all()
        resource_name = 'activity_log'
        allowed_methods = ['get', 'post']

        authentication = PandaAuthentication()
        authorization = DjangoAuthorization()
        serializer = PandaSerializer()

    def obj_create(self, bundle, request=None, **kwargs):
        """
        Create an activity log for the accessing user.
        """
        # Because users may have authenticated via headers the request.user may
        # not be a full User instance. To be sure, we fetch one.
        user = UserProxy.objects.get(id=request.user.id)

        try:
            bundle = super(ActivityLogResource, self).obj_create(bundle, request=request, user=user, **kwargs)
        except IntegrityError:
            raise ImmediateHttpResponse(response=HttpConflict(_('Activity has already been recorded.')))

        return bundle


########NEW FILE########
__FILENAME__ = category
#!/usr/bin/env python

from django.conf import settings
from django.db.models import Count
from tastypie.authorization import DjangoAuthorization

from panda.api.utils import PandaAuthentication, SluggedModelResource, PandaSerializer
from panda.models import Category, Dataset

class CategoryResource(SluggedModelResource):
    """
    Simple API for Category objects. 
    """
    class Meta:
        queryset = Category.objects.annotate(dataset_count=Count('datasets'))
        resource_name = 'category'
        allowed_methods = ['get']

        authentication = PandaAuthentication()
        authorization = DjangoAuthorization()
        serializer = PandaSerializer()

    def dehydrate(self, bundle):
        """
        If using an annotated queryset, return the dataset counts as well.
        (This happens when accessing the Category API directly, but not when
        a Category is embedded in another object, such as a Dataset.)
        """
        if hasattr(bundle.obj, 'dataset_count') and bundle.obj.dataset_count is not None:
            bundle.data['dataset_count'] = bundle.obj.dataset_count

        return bundle

    def get_list(self, request, **kwargs):
        """
        Overriden from underlying implementation in order to insert a fake category
        for "Uncategorized" datasets.
        """
        # TODO: Uncached for now. Invalidation that works for everyone may be
        #       impossible.
        objects = self.obj_get_list(request=request, **self.remove_api_resource_names(kwargs))
        sorted_objects = self.apply_sorting(objects, options=request.GET)

        paginator = self._meta.paginator_class(request.GET, sorted_objects, resource_uri=self.get_resource_list_uri(), limit=self._meta.limit)
        to_be_serialized = paginator.page()

        # Dehydrate the bundles in preparation for serialization.
        bundles = [self.build_bundle(obj=obj, request=request) for obj in to_be_serialized['objects']]
        to_be_serialized['objects'] = [self.full_dehydrate(bundle) for bundle in bundles]

        # Insert fake category
        uncategorized = Category(
            id=settings.PANDA_UNCATEGORIZED_ID,
            slug=settings.PANDA_UNCATEGORIZED_SLUG,
            name=settings.PANDA_UNCATEGORIZED_NAME)
        uncategorized.__dict__['dataset_count'] = Dataset.objects.filter(categories=None).count() 
        uncategorized_bundle = self.full_dehydrate(self.build_bundle(obj=uncategorized))

        to_be_serialized['objects'].append(uncategorized_bundle)

        to_be_serialized = self.alter_list_data_to_serialize(request, to_be_serialized)

        return self.create_response(request, to_be_serialized)


########NEW FILE########
__FILENAME__ = data
#!/usr/bin/env python

import re

from django.conf import settings
from django.core.urlresolvers import get_script_prefix, resolve, reverse
from django.utils import simplejson as json
from django.utils.translation import ugettext_lazy as _
from tastypie import fields, http
from tastypie.authorization import DjangoAuthorization
from tastypie.bundle import Bundle
from tastypie.exceptions import BadRequest, NotFound, ImmediateHttpResponse
from tastypie.utils import dict_strip_unicode_keys 
from tastypie.utils.mime import build_content_type
from tastypie.validation import Validation

from panda import solr
from panda.api.datasets import DatasetResource
from panda.exceptions import DatasetLockedError
from panda.api.utils import PandaAuthentication, PandaPaginator, PandaResource, PandaSerializer
from panda.models import Category, Dataset, SearchLog, TaskStatus, UserProxy
from panda.tasks import ExportSearchTask, PurgeDataTask

class SolrObject(object):
    """
    A lightweight wrapper around a Solr response object for use when
    querying Solr via Tastypie.
    """
    def __init__(self, initial=None, **kwargs):
        self.__dict__['_data'] = {}

        if hasattr(initial, 'items'):
            self.__dict__['_data'] = initial

        self.__dict__['_data'].update(kwargs)

    def __getattr__(self, name):
        return self._data.get(name, None)

    def __setattr__(self, name, value):
        self.__dict__['_data'][name] = value

    def __str__(self):
        return str(self.__dict__['_data'])

    def __unicode__(self):
        return unicode(self.__dict__['_data'])

    def to_dict(self):
        return self._data

class DataValidation(Validation):
    """
    Tastypie Validation for Data objects.
    """
    def is_valid(self, bundle, request=None):
        errors = {}

        if 'data' not in bundle.data or not bundle.data['data']:
            errors['data'] = [_('The data field is required.')]

        if 'external_id' in bundle.data:
            if not isinstance(bundle.data['external_id'], basestring):
                errors['external_id'] = [_('external_id must be a string.')]
            elif not re.match('^[\w\d_-]+$', bundle.data['external_id']):
                errors['external_id'] = [_('external_id can only contain letters, numbers, underscores and dashes.')]

        return errors

class DataResource(PandaResource):
    """
    API resource for data.
    """
    dataset_slug = fields.CharField(attribute='dataset_slug',
        help_text=_('Slug of the dataset this row of data belongs to.'))
    external_id = fields.CharField(attribute='external_id', null=True, blank=True,
        help_text=_('Per-dataset unique identifier for this row of data.'))
    data = fields.CharField(attribute='data',
        help_text=_('An ordered list of values corresponding to the columns in the parent dataset.'))

    class Meta:
        resource_name = 'data'
        allowed_methods = ['get', 'post', 'put', 'delete']
        always_return_data = True

        authentication = PandaAuthentication()
        authorization = DjangoAuthorization()
        serializer = PandaSerializer()
        validation = DataValidation()

        object_class = SolrObject

    def dehydrate_data(self, bundle):
        """
        Convert csv data into a proper array for JSON serialization
        """
        return json.loads(bundle.data['data'])

    def dehydrate(self, bundle):
        """
        Trim the dataset_slug field and add a proper relationship.
        """
        dataset = Dataset.objects.get(slug=bundle.data['dataset_slug'])

        del bundle.data['dataset_slug']
        bundle.data['dataset'] = DatasetResource().get_resource_uri(dataset)

        return bundle

    def get_resource_uri(self, bundle_or_obj):
        """
        Build a canonical uri for a datum.

        If the resource doesn't have an external_id it is
        considered "unaddressable" and this will return None.
        """
        dr = DatasetResource()

        kwargs = {
            'api_name': self._meta.api_name,
            'dataset_resource_name': dr._meta.resource_name,
            'resource_name': self._meta.resource_name,
        }

        if isinstance(bundle_or_obj, Bundle):
            kwargs['dataset_slug'] = bundle_or_obj.obj.dataset_slug
            kwargs['external_id'] = bundle_or_obj.obj.external_id
        else:
            kwargs['dataset_slug'] = bundle_or_obj.dataset_slug
            kwargs['external_id'] = bundle_or_obj.external_id
 
        if not kwargs['external_id']:
            return None

        return dr._build_reverse_url('api_dataset_data_detail', kwargs=kwargs)

    def get_dataset_from_kwargs(self, bundle, **kwargs):
        """
        Extract a dataset from one of the variety of places it might be hiding.
        """
        kwargs_slug = kwargs['dataset_slug']
        
        bundle_uri = None
        bundle_slug = None

        if bundle:
            bundle_uri = bundle.data.pop('dataset', None)

        if bundle_uri:
            prefix = get_script_prefix()

            if prefix and bundle_uri.startswith(prefix):
                bundle_uri = bundle_uri[len(prefix)-1:]

            view, args, kwargs = resolve(bundle_uri)

            bundle_slug = kwargs['slug']

        if bundle_slug and bundle_slug != kwargs_slug:
            raise BadRequest(_('Dataset specified in request body does not agree with dataset API endpoint used.'))

        return Dataset.objects.get(slug=kwargs_slug) 

    def validate_bundle_data(self, bundle, request, dataset):
        """
        Perform additional validation that isn't possible with the Validation object.
        """
        errors = {}

        field_count = len(bundle.data['data'])

        if dataset.initial_upload and not dataset.row_count:
            errors['dataset'] = [_('Can not create or modify data for a dataset which has initial_upload, but has not completed the import process.')]

        if dataset.column_schema is None:
            errors['dataset'] = [_('Can not create or modify data for a dataset without columns.')]
        else:
            expected_field_count = len(dataset.column_schema)

            if field_count != expected_field_count:
                errors['data'] = [_('Got %(field_count)i data fields. Expected %(expected_field_count)i.') \
                    % {'field_count': field_count, 'expected_field_count': expected_field_count}]

        # Cribbed from is_valid()
        if errors:
            if request:
                desired_format = self.determine_format(request)
            else:
                desired_format = self._meta.default_format

            serialized = self.serialize(request, errors, desired_format)
            response = http.HttpBadRequest(content=serialized, content_type=build_content_type(desired_format))
            raise ImmediateHttpResponse(response=response)

    # Data access methods

    def get_object_list():
        """
        Bypassed, should never be invoked. 

        Since Solr queries are not lazy, fetching a complete list
        of objects never makes sense.
        """
        raise NotImplementedError() 

    def obj_get_list(self, request=None, **kwargs):
        """
        Bypassed, should never be invoked. 
        
        See ``get_list``.
        """
        raise NotImplementedError() 

    def obj_get(self, request=None, **kwargs):
        """
        Query Solr for a single item by primary key.
        """
        dataset = Dataset.objects.get(slug=kwargs['dataset_slug'])

        row = dataset.get_row(kwargs['external_id'])

        if not row:
            raise NotFound()

        return SolrObject(row)

    def obj_create(self, bundle, request=None, **kwargs):
        """
        Add one Data to a Dataset.
        """
        dataset = self.get_dataset_from_kwargs(bundle, **kwargs)

        self.validate_bundle_data(bundle, request, dataset)

        if 'external_id' in bundle.data:
            external_id = bundle.data['external_id']
        elif 'external_id' in kwargs:
            external_id = kwargs['external_id']
        else:
            external_id = None

        # Because users may have authenticated via headers the request.user may
        # not be a full User instance. To be sure, we fetch one.
        user = UserProxy.objects.get(id=request.user.id)

        try:
            row = dataset.add_row(user, bundle.data['data'], external_id=external_id)
        except DatasetLockedError:
            raise ImmediateHttpResponse(response=http.HttpForbidden(_('Dataset is currently locked by another process.')))

        bundle.obj = SolrObject(row)

        return bundle

    def obj_update(self, bundle, request=None, **kwargs):
        """
        Overwrite an existing Data.
        """
        return self.obj_create(bundle, request, **kwargs)

    def obj_delete_list(self, request=None, **kwargs):
        """
        See ``put_list``. 
        """
        raise NotImplementedError()

    def obj_delete(self, request=None, **kwargs):
        """
        Delete a ``Data``.
        """
        dataset = Dataset.objects.get(slug=kwargs['dataset_slug'])

        # Because users may have authenticated via headers the request.user may
        # not be a full User instance. To be sure, we fetch one.
        user = UserProxy.objects.get(id=request.user.id)

        try:
            dataset.delete_row(user, kwargs['external_id'])
        except DatasetLockedError:
            raise ImmediateHttpResponse(response=http.HttpForbidden(_('Dataset is currently locked by another process.')))

    def rollback(self, bundles):
        """
        See ``put_list``.
        """
        raise NotImplementedError()

    # Views

    def get_list(self, request, **kwargs):
        """
        Retrieve a list of ``Data`` objects, optionally applying full-text search.

        Bypasses ``obj_get_list``, making it unnecessary.
        """
        results = self.search_dataset_data(request, **kwargs)

        return self.create_response(request, results)

    def get_detail(self, request, **kwargs):
        """
        Handled by the underlying implementation.

        See ``obj_get``.
        """
        return super(DataResource, self).get_detail(request, **kwargs)

    def put_list(self, request, **kwargs):
        """
        A custom bulk create/update handler. Notes: 

        * ``obj_delete_list`` is never called, objects are overwritten instead.
        * All objects are validated before any objects are created, so ``rollback`` is unnecessary.
        * A single dataset save and Solr commit are made at the end (optimization!).
        """
        deserialized = self.deserialize(request, request.raw_post_data, format=request.META.get('CONTENT_TYPE', 'application/json'))
        deserialized = self.alter_deserialized_list_data(request, deserialized)

        if not 'objects' in deserialized:
            raise BadRequest(_("Invalid data sent."))

        bundles = []
        data = []

        dataset = self.get_dataset_from_kwargs(None, **kwargs)

        for object_data in deserialized['objects']:
            bundle = self.build_bundle(data=dict_strip_unicode_keys(object_data), request=request)

            self.is_valid(bundle, request)

            if bundle.errors:
                self.error_response(bundle.errors, request)

            bundles.append(bundle)
            data.append((
                bundle.data['data'],
                bundle.data.get('external_id', None) 
            ))
        
            self.validate_bundle_data(bundle, request, dataset)

        # Because users may have authenticated via headers the request.user may
        # not be a full User instance. To be sure, we fetch one.
        user = UserProxy.objects.get(id=request.user.id)
        
        try:
            solr_rows = dataset.add_many_rows(user, data)
        except DatasetLockedError:
            raise ImmediateHttpResponse(response=http.HttpForbidden(_('Dataset is currently locked by another process.')))

        for bundle, solr_row in zip(bundles, solr_rows):
            bundle.obj = SolrObject(solr_row)

        if not self._meta.always_return_data:
            return http.HttpNoContent()
        else:
            to_be_serialized = {}
            to_be_serialized['objects'] = [self.full_dehydrate(bundle) for bundle in bundles]
            to_be_serialized = self.alter_list_data_to_serialize(request, to_be_serialized)

            return self.create_response(request, to_be_serialized, response_class=http.HttpAccepted)

    def put_detail(self, request, **kwargs):
        """
        Handled by the underlying implementation.

        See ``obj_update``.
        """
        return super(DataResource, self).put_detail(request, **kwargs)

    def post_list(self, request, **kwargs):
        """
        Handled by the underlying implementation.

        See ``obj_create``.
        """
        return super(DataResource, self).post_list(request, **kwargs)

    def post_detail(self, request, **kwargs):
        """
        Handled by the underlying implementation, which means this is
        not supported.
        """
        return super(DataResource, self).post_detail(request, **kwargs)

    def delete_list(self, request, **kwargs):
        """
        Delete all ``Data`` in a ``Dataset``. Must be called from a data
        url nested under a Dataset. Deleting *all* ``Data`` objects is
        not supported.
        """
        dataset = Dataset.objects.get(slug=kwargs['dataset_slug'])

        # Because users may have authenticated via headers the request.user may
        # not be a full User instance. To be sure, we fetch one.
        user = UserProxy.objects.get(id=request.user.id)
        
        try:
            dataset.delete_all_rows(user) 
        except DatasetLockedError:
            raise ImmediateHttpResponse(response=http.HttpForbidden(_('Dataset is currently locked by another process.')))

        return http.HttpNoContent()

    def delete_detail(self, request, **kwargs):
        return super(DataResource, self).delete_detail(request, **kwargs)

    # Search

    def search_all_data(self, request, **kwargs):
        """
        List endpoint using Solr. Provides full-text search via the "q" parameter."
        """
        self.method_check(request, allowed=['get'])
        self.is_authenticated(request)
        self.throttle_check(request)

        try:
            query = '(%s)' % request.GET['q']
        except KeyError:
            query = ''

        category = request.GET.get('category', '')
        since = request.GET.get('since', None)
        limit = int(request.GET.get('limit', settings.PANDA_DEFAULT_SEARCH_GROUPS))
        offset = int(request.GET.get('offset', 0))
        group_limit = int(request.GET.get('group_limit', settings.PANDA_DEFAULT_SEARCH_ROWS_PER_GROUP))
        group_offset = int(request.GET.get('group_offset', 0))
        export = bool(request.GET.get('export', False))

        solr_query_bits = [query]

        if category:
            if category != 'uncategorized':
                category = Category.objects.get(slug=category)
                dataset_slugs = category.datasets.values_list('slug', flat=True)
            else:
                dataset_slugs = Dataset.objects.filter(categories=None).values_list('slug', flat=True) 

            solr_query_bits.append('dataset_slug:(%s)' % ' '.join(dataset_slugs))

        if since:
            solr_query_bits.append('last_modified:[' + since + 'Z TO *]')

        # Because users may have authenticated via headers the request.user may
        # not be a full User instance. To be sure, we fetch one.
        user = UserProxy.objects.get(id=request.user.id)

        if export:
            task_type = ExportSearchTask

            task = TaskStatus.objects.create(
                task_name=task_type.name,
                task_description=_('Export search results for "%s".') % query,
                creator=user
            )

            task_type.apply_async(
                args=[query, task.id],
                kwargs={},
                task_id=task.id
            )
        else:
            response = solr.query_grouped(
                settings.SOLR_DATA_CORE,
                ' AND '.join(solr_query_bits),
                'dataset_slug',
                offset=offset,
                limit=limit,
                group_limit=group_limit,
                group_offset=group_offset
            )
            groups = response['grouped']['dataset_slug']['groups']

            page = PandaPaginator(
                request.GET,
                groups,
                resource_uri=request.path_info,
                count=response['grouped']['dataset_slug']['ngroups']
            ).page()

            datasets = []

            for group in groups:
                dataset_slug = group['groupValue']
                results = group['doclist']
                
                try:
                    dataset = Dataset.objects.get(slug=dataset_slug)
                # In the event that stale data exists in Solr, skip this dataset,
                # request the invalid data be purged and return the other results.
                # Pagination may be wrong, but this is the most functional solution. (#793)
                except Dataset.DoesNotExist:
                    PurgeDataTask.apply_async(args=[dataset_slug])
                    solr.delete(settings.SOLR_DATASETS_CORE, 'slug:%s' % dataset_slug)

                    page['meta']['total_count'] -= 1

                    continue
                
                dataset_resource = DatasetResource()
                dataset_bundle = dataset_resource.build_bundle(obj=dataset, request=request)
                dataset_bundle = dataset_resource.full_dehydrate(dataset_bundle)
                dataset_bundle = dataset_resource.simplify_bundle(dataset_bundle)

                objects = [SolrObject(obj) for obj in results['docs']]
                
                dataset_search_url = reverse('api_dataset_data_list', kwargs={ 'api_name': self._meta.api_name, 'dataset_resource_name': 'dataset', 'resource_name': 'data', 'dataset_slug': dataset.slug })

                data_page = PandaPaginator(
                    { 'limit': str(group_limit), 'offset': str(group_offset), 'q': query },
                    objects,
                    resource_uri=dataset_search_url,
                    count=results['numFound']
                ).page()

                dataset_bundle.data.update(data_page)
                dataset_bundle.data['objects'] = []

                for obj in objects:
                    data_bundle = self.build_bundle(obj=obj, request=request)
                    data_bundle = self.full_dehydrate(data_bundle)
                    dataset_bundle.data['objects'].append(data_bundle)

                datasets.append(dataset_bundle.data)

            page['objects'] = datasets
            
            # Log query
            SearchLog.objects.create(user=user, dataset=None, query=query)

        self.log_throttled_access(request)

        if export:
            return self.create_response(request, _('Export queued.'))
        else:
            return self.create_response(request, page)

    def search_dataset_data(self, request, **kwargs):
        """
        Perform a full-text search on only one dataset.

        See ``get_list``.
        """
        dataset = Dataset.objects.get(slug=kwargs['dataset_slug'])

        try:
            query = '(%s)' % request.GET['q']
        except KeyError:
            query = ''

        since = request.GET.get('since', None)
        limit = int(request.GET.get('limit', settings.PANDA_DEFAULT_SEARCH_ROWS))
        offset = int(request.GET.get('offset', 0))
        sort = request.GET.get('sort', '_docid_ asc')

        solr_query_bits = [query]
        solr_query_bits.append('dataset_slug:%s' % dataset.slug)

        if since:
            solr_query_bits.append('last_modified:[' + since + 'Z TO *]')

        response = solr.query(
            settings.SOLR_DATA_CORE,
            ' AND '.join(solr_query_bits),
            offset=offset,
            sort=sort,
            limit=limit
        )

        dataset_resource = DatasetResource()
        dataset_bundle = dataset_resource.build_bundle(obj=dataset, request=request)
        dataset_bundle = dataset_resource.full_dehydrate(dataset_bundle)
        dataset_bundle = dataset_resource.simplify_bundle(dataset_bundle)
       
        results = [SolrObject(d) for d in response['response']['docs']]

        page = PandaPaginator(
            request.GET,
            results,
            resource_uri=request.path_info,
            count=response['response']['numFound']
        ).page() 
        
        dataset_bundle.data.update(page)
        dataset_bundle.data['objects'] = []

        for obj in results:
            bundle = self.build_bundle(obj=obj, request=request)
            bundle = self.full_dehydrate(bundle)
            dataset_bundle.data['objects'].append(bundle.data)

        # Because users may have authenticated via headers the request.user may
        # not be a full User instance. To be sure, we fetch one.
        user = UserProxy.objects.get(id=request.user.id)
        
        SearchLog.objects.create(user=user, dataset=dataset, query=query)

        return dataset_bundle


########NEW FILE########
__FILENAME__ = datasets
#!/usr/bin/env python

from django.conf import settings
from django.conf.urls.defaults import url
from django.utils.translation import ugettext_lazy as _
from tastypie import fields
from tastypie import http
from tastypie.authorization import DjangoAuthorization
from tastypie.exceptions import BadRequest, ImmediateHttpResponse
from tastypie.utils.urls import trailing_slash
from tastypie.validation import Validation

from panda import solr
from panda.api.utils import PandaAuthentication, PandaPaginator, JSONApiField, SluggedModelResource, PandaSerializer
from panda.exceptions import DataImportError, DatasetLockedError
from panda.models import Category, Dataset, DataUpload, UserProxy
from panda.utils.column_schema import make_column_schema

class DatasetValidation(Validation):
    def is_valid(self, bundle, request=None):
        errors = {}

        if 'name' not in bundle.data or not bundle.data['name']:
            errors['name'] = [_('This field is required.')]

        return errors

class DatasetResource(SluggedModelResource):
    """
    API resource for Datasets.
    """
    from panda.api.category import CategoryResource
    from panda.api.tasks import TaskResource
    from panda.api.data_uploads import DataUploadResource
    from panda.api.users import UserResource

    categories = fields.ToManyField(CategoryResource, 'categories', full=True, null=True)
    creator = fields.ForeignKey(UserResource, 'creator', full=True, readonly=True)
    current_task = fields.ToOneField(TaskResource, 'current_task', full=True, null=True, readonly=True)
    related_uploads = fields.ToManyField('panda.api.related_uploads.RelatedUploadResource', 'related_uploads', full=True, null=True)
    data_uploads = fields.ToManyField('panda.api.data_uploads.DataUploadResource', 'data_uploads', full=True, null=True)
    last_modified_by = fields.ForeignKey(UserResource, 'last_modified_by', full=True, readonly=True, null=True)
    initial_upload = fields.ForeignKey(DataUploadResource, 'initial_upload', readonly=True, null=True)

    slug = fields.CharField(attribute='slug')
    column_schema = JSONApiField(attribute='column_schema', readonly=True, null=True)
    sample_data = JSONApiField(attribute='sample_data', readonly=True, null=True)
    row_count = fields.IntegerField(attribute='row_count', readonly=True, null=True)
    creation_date = fields.DateTimeField(attribute='creation_date', readonly=True, null=True)
    last_modified = fields.DateTimeField(attribute='last_modified', readonly=True, null=True)
    last_modification = fields.CharField(attribute='last_modification', readonly=True, null=True)
    last_modified_by = fields.ForeignKey(UserResource, 'last_modified_by', full=True, null=True, readonly=True)
    locked = fields.BooleanField(attribute='locked', readonly=True, null=True)
    locked_at = fields.DateTimeField(attribute='locked_at', readonly=True, null=True)

    class Meta:
        queryset = Dataset.objects.all()
        resource_name = 'dataset'
        allowed_methods = ['get', 'post', 'put', 'delete']
        always_return_data = True

        authentication = PandaAuthentication()
        authorization = DjangoAuthorization()
        validation = DatasetValidation()
        serializer = PandaSerializer()

    def simplify_bundle(self, bundle):
        """
        Takes a dehydrated bundle and removes attributes to create a "simple"
        view that is faster over the wire.
        """
        del bundle.data['data_uploads']
        del bundle.data['related_uploads']
        del bundle.data['sample_data']
        del bundle.data['current_task']

        return bundle

    def override_urls(self):
        """
        Add urls for search endpoint.
        """
        from panda.api.data import DataResource
        
        data_resource = DataResource(api_name=self._meta.api_name)

        return [
            url(r"^(?P<resource_name>%s)/schema%s$" % (self._meta.resource_name, trailing_slash()), self.wrap_view('get_schema'), name="api_get_schema"),
            url(r"^(?P<resource_name>%s)/(?P<slug>[\w\d_-]+)%s$" % (self._meta.resource_name, trailing_slash()), self.wrap_view('dispatch_detail'), name="api_dispatch_detail"),
            url(r'^(?P<resource_name>%s)/(?P<slug>[\w\d_-]+)/import/(?P<upload_id>\d+)%s$' % (self._meta.resource_name, trailing_slash()), self.wrap_view('import_data'), name='api_import_data'),
            url(r'^(?P<resource_name>%s)/(?P<slug>[\w\d_-]+)/export%s$' % (self._meta.resource_name, trailing_slash()), self.wrap_view('export_data'), name='api_export_data'),
            url(r'^(?P<resource_name>%s)/(?P<slug>[\w\d_-]+)/reindex%s$' % (self._meta.resource_name, trailing_slash()), self.wrap_view('reindex_data'), name='api_reindex_data'),
            
            # Nested urls for accessing data
            url(r'^(?P<dataset_resource_name>%s)/(?P<dataset_slug>[\w\d_-]+)/(?P<resource_name>%s)%s$' % (self._meta.resource_name, data_resource._meta.resource_name, trailing_slash()), data_resource.wrap_view('dispatch_list'), name='api_dataset_data_list'),
            url(r'^(?P<dataset_resource_name>%s)/(?P<dataset_slug>[\w\d_-]+)/(?P<resource_name>%s)/(?P<external_id>[\w\d_-]+)%s$' % (self._meta.resource_name, data_resource._meta.resource_name, trailing_slash()), data_resource.wrap_view('dispatch_detail'), name='api_dataset_data_detail'),
            url(r'^data%s' % trailing_slash(), data_resource.wrap_view('search_all_data'), name='api_data_search')
        ]

    def get_list(self, request, **kwargs):
        """
        List endpoint using Solr. Provides full-text search via the "q" parameter."
        """
        limit = int(request.GET.get('limit', settings.PANDA_DEFAULT_SEARCH_ROWS))
        offset = int(request.GET.get('offset', 0))
        category_slug = request.GET.get('category', None)
        creator_email = request.GET.get('creator_email', None)
        query = request.GET.get('q', '')
        simple = True if request.GET.get('simple', 'false').lower() == 'true' else False

        if category_slug == settings.PANDA_UNCATEGORIZED_SLUG:
            category_id = settings.PANDA_UNCATEGORIZED_ID
        elif category_slug:
            category_id = Category.objects.get(slug=category_slug).id
        else:
            category_id = None

        if category_id is not None and query:
            q = 'categories:%s %s' % (category_id, query)
        elif category_id is not None:
            q = 'categories:%s' % category_id
        else:
            q = query

        if creator_email:
            datasets = Dataset.objects.filter(creator__email=creator_email)
            count = datasets.count()
            datasets = datasets[offset:offset + limit]
        else:
            response = solr.query(settings.SOLR_DATASETS_CORE, q, offset=offset, limit=limit, sort='creation_date desc')
            count = response['response']['numFound']
            
            dataset_slugs = [d['slug'] for d in response['response']['docs']]
            datasets = Dataset.objects.filter(slug__in=dataset_slugs)

        paginator = PandaPaginator(request.GET, datasets, resource_uri=request.path_info, count=count)
        page = paginator.page()

        objects = []

        for obj in datasets:
            bundle = self.build_bundle(obj=obj, request=request)
            bundle = self.full_dehydrate(bundle)

            # Prune attributes we don't care about
            if simple:
                bundle = self.simplify_bundle(bundle)

            objects.append(bundle)

        page['objects'] = objects

        return self.create_response(request, page)

    def put_detail(self, request, **kwargs):
        """
        Allow emulating a ``PATCH`` request by passing ``?patch=true``.
        (As a workaround for IE's broken XMLHttpRequest.)
        """
        if request.GET.get('patch', 'false').lower() == 'true':
            return super(DatasetResource, self).patch_detail(request, **kwargs)
        else:
            return super(DatasetResource, self).put_detail(request, **kwargs)

    def obj_create(self, bundle, request=None, **kwargs):
        """
        Set creator and update full text.
        """
        # Because users may have authenticated via headers the request.user may
        # not be a full User instance. To be sure, we fetch one.
        user = UserProxy.objects.get(id=request.user.id)

        bundle = super(DatasetResource, self).obj_create(bundle, request=request, creator=user, **kwargs)

        if 'columns' in request.GET:
            columns = request.GET['columns'].split(',')
        else:
            columns = None

        if 'typed_columns' in request.GET:
            if not columns:
                raise BadRequest(_('The "columns" argument must also be specified when specifying "typed_columns".'))

            typed_columns = [True if c.lower() == 'true' else False for c in request.GET['typed_columns'].split(',')]

            if len(typed_columns) != len(columns):
                raise BadRequest(_('The "typed_columns" argument must be a comma-separated list of True/False values with the same number of values as the "columns" argument.'))
        else:
            typed_columns = None

        if 'column_types' in request.GET:
            if not columns:
                raise BadRequest(_('The "columns" argument must also be specified when specifying "column_types".'))

            column_types = [None if c.lower() == '' else c.lower() for c in request.GET['column_types'].split(',')]

            if len(column_types) != len(columns):
                raise BadRequest(_('The "column_types" argument must be a comma-separated list of types with the same number of values as the "columns" argument.'))
        else:
            column_types = None

        if columns:
            bundle.obj.column_schema = make_column_schema(columns, typed_columns, column_types)
            bundle.obj.save()

        # After ALL changes have been made to the object and its relations, update its full text in Solr.
        bundle.obj.update_full_text()

        return bundle

    def obj_update(self, bundle, request=None, **kwargs):
        """
        Update full text.
        """
        bundle = super(DatasetResource, self).obj_update(bundle, request=request, **kwargs)

        # After ALL changes have been made to the object and its relations, update its full text in Solr.
        bundle.obj.update_full_text()

        return bundle

    def import_data(self, request, **kwargs):
        """
        Dummy endpoint for kicking off data import tasks.
        """
        self.method_check(request, allowed=['get'])
        self.is_authenticated(request)
        self.throttle_check(request)

        if 'slug' in kwargs:
            slug = kwargs['slug']
        else:
            slug = request.GET.get('slug')

        dataset = Dataset.objects.get(slug=slug)
        upload = DataUpload.objects.get(id=kwargs['upload_id'])

        # Because users may have authenticated via headers the request.user may
        # not be a full User instance. To be sure, we fetch one.
        user = UserProxy.objects.get(id=request.user.id)

        try:
            dataset.import_data(user, upload)
        except DatasetLockedError:
            raise ImmediateHttpResponse(response=http.HttpForbidden(_('Dataset is currently locked by another process.')))
        except DataImportError, e:
            raise ImmediateHttpResponse(response=http.HttpForbidden(e.message))

        dataset.update_full_text()

        bundle = self.build_bundle(obj=dataset, request=request)
        bundle = self.full_dehydrate(bundle)

        self.log_throttled_access(request)

        return self.create_response(request, bundle)

    def reindex_data(self, request, **kwargs):
        """
        Dummy endpoint for kicking off data reindexing tasks.
        """
        self.method_check(request, allowed=['get'])
        self.is_authenticated(request)
        self.throttle_check(request)

        if 'slug' in kwargs:
            slug = kwargs['slug']
        else:
            slug = request.GET.get('slug')

        dataset = Dataset.objects.get(slug=slug)

        if not dataset.column_schema:
            raise BadRequest(_('This dataset has no data to reindex.'))

        if 'typed_columns' in request.GET:
            typed_columns = [True if c.lower() == 'true' else False for c in request.GET['typed_columns'].split(',')]

            if len(typed_columns) != len(dataset.column_schema):
                raise BadRequest(_('typed_columns must be a comma-separated list of True/False values with the same number of values as the dataset has columns.'))
        else:
            typed_columns = None

        if 'column_types' in request.GET:
            column_types = [None if c.lower() == '' else c.lower() for c in request.GET['column_types'].split(',')]

            if len(column_types) != len(dataset.column_schema):
                raise BadRequest(_('column_types must be a comma-separated list of types with the same number of values as the dataset has columns.'))
        else:
            column_types = None

        # Because users may have authenticated via headers the request.user may
        # not be a full User instance. To be sure, we fetch one.
        user = UserProxy.objects.get(id=request.user.id)

        try:
            dataset.reindex_data(user, typed_columns=typed_columns, column_types=column_types)
        except DatasetLockedError:
            raise ImmediateHttpResponse(response=http.HttpForbidden(_('Dataset is currently locked by another process.')))

        dataset.update_full_text()

        bundle = self.build_bundle(obj=dataset, request=request)
        bundle = self.full_dehydrate(bundle)

        self.log_throttled_access(request)

        return self.create_response(request, bundle)

    def export_data(self, request, **kwargs):
        """
        Dummy endpoint for kicking off data export tasks.

        NB: This endpoint is used for both exporting complete datasets
        (without a query arg) and exporting dataset search results
        (with a query arg).
        """
        self.method_check(request, allowed=['get'])
        self.is_authenticated(request)
        self.throttle_check(request)

        if 'slug' in kwargs:
            slug = kwargs['slug']
        else:
            slug = request.GET.get('slug')

        dataset = Dataset.objects.get(slug=slug)

        query = request.GET.get('q', '')
        since = request.GET.get('since', None)

        if since:
            query = 'last_modified:[' + since + 'Z TO *] AND (%s)' % query

        # Because users may have authenticated via headers the request.user may
        # not be a full User instance. To be sure, we fetch one.
        user = UserProxy.objects.get(id=request.user.id)

        try:
            dataset.export_data(user, query=query)
        except DatasetLockedError:
            raise ImmediateHttpResponse(response=http.HttpForbidden(_('Dataset is currently locked by another process.')))

        bundle = self.build_bundle(obj=dataset, request=request)
        bundle = self.full_dehydrate(bundle)

        self.log_throttled_access(request)

        return self.create_response(request, bundle)


########NEW FILE########
__FILENAME__ = data_uploads
#!/usr/bin/env python

from mimetypes import guess_type

from django.conf.urls.defaults import url
from django.core.exceptions import ObjectDoesNotExist
from django.core.servers.basehttp import FileWrapper
from django.http import HttpResponse
from django.utils.translation import ugettext_lazy as _
from tastypie import fields
from tastypie import http
from tastypie.authorization import DjangoAuthorization
from tastypie.exceptions import ImmediateHttpResponse, NotFound
from tastypie.utils.urls import trailing_slash

from panda.api.utils import JSONApiField, PandaAuthentication, PandaModelResource, PandaSerializer
from panda.exceptions import DataUploadNotDeletable
from panda.models import DataUpload

class DataUploadResource(PandaModelResource):
    """
    API resource for DataUploads.
    """
    from panda.api.users import UserResource

    filename = fields.CharField('filename', readonly=True)
    original_filename = fields.CharField('original_filename', readonly=True)
    size = fields.IntegerField('size', readonly=True)
    creator = fields.ForeignKey(UserResource, 'creator', full=True, readonly=True)
    creation_date = fields.DateTimeField('creation_date', readonly=True)
    title = fields.CharField('title', null=True)
    dataset = fields.ForeignKey('panda.api.datasets.DatasetResource', 'dataset', null=True, readonly=True)
    data_type = fields.CharField('data_type', null=True, readonly=True)
    encoding = fields.CharField('encoding', readonly=True)
    dialect = fields.CharField('dialect', null=True, readonly=True)
    columns = JSONApiField('columns', null=True, readonly=True)
    sample_data = JSONApiField('sample_data', null=True, readonly=True)
    guessed_types = JSONApiField('guessed_types', null=True, readonly=True)
    imported = fields.BooleanField('imported', readonly=True)

    class Meta:
        queryset = DataUpload.objects.all()
        resource_name = 'data_upload'
        allowed_methods = ['get', 'put', 'delete']
        always_return_data = True

        authentication = PandaAuthentication()
        authorization = DjangoAuthorization()
        serializer = PandaSerializer()

    def override_urls(self):
        """
        Add urls for search endpoint.
        """
        return [
            url(r'^(?P<resource_name>%s)/(?P<pk>\w[\w/-]*)/download%s$' % (self._meta.resource_name, trailing_slash()), self.wrap_view('download'), name='api_download_data_upload'),
        ]

    def obj_delete(self, request=None, **kwargs):
        """
        Override delete to also update related Dataset's metadata.
        """
        obj = kwargs.pop('_obj', None)

        if not hasattr(obj, 'delete'):
            try:
                obj = self.obj_get(request, **kwargs)
            except ObjectDoesNotExist:
                raise NotFound(_("A model instance matching the provided arguments could not be found."))

        try:
            obj.delete()
        except DataUploadNotDeletable, e:
            raise ImmediateHttpResponse(response=http.HttpForbidden(e.message))

        if obj.dataset:
            obj.dataset.update_full_text()

    def download(self, request, **kwargs):
        """
        Download the original file that was uploaded.
        """
        # Allow POST so csrf token can come through
        self.method_check(request, allowed=['get', 'post'])
        self.is_authenticated(request)
        self.throttle_check(request)

        if 'pk' in kwargs:
            get_id = kwargs['pk']
        else:
            get_id = request.GET.get('id', '')

        upload = DataUpload.objects.get(id=get_id)
        path = upload.get_path()

        self.log_throttled_access(request)

        response = HttpResponse(FileWrapper(open(path, 'r')), content_type=guess_type(upload.original_filename)[0])
        response['Content-Disposition'] = 'attachment; filename=%s' % upload.original_filename
        response['Content-Length'] = upload.size

        return response


########NEW FILE########
__FILENAME__ = exports
#!/usr/bin/env python

from mimetypes import guess_type

from django.conf.urls.defaults import url
from django.core.servers.basehttp import FileWrapper
from django.http import HttpResponse
from tastypie import fields
from tastypie.authorization import DjangoAuthorization
from tastypie.utils.urls import trailing_slash

from panda.api.utils import PandaAuthentication, PandaModelResource, PandaSerializer
from panda.models import Export

class ExportResource(PandaModelResource):
    """
    API resource for Exports.
    """
    from panda.api.users import UserResource

    creator = fields.ForeignKey(UserResource, 'creator')
    dataset = fields.ForeignKey('panda.api.datasets.DatasetResource', 'dataset', null=True)

    class Meta:
        queryset = Export.objects.all()
        resource_name = 'export'
        allowed_methods = ['get']

        authentication = PandaAuthentication()
        authorization = DjangoAuthorization()
        serializer = PandaSerializer()

    def override_urls(self):
        """
        Add urls for search endpoint.
        """
        return [
            url(r'^(?P<resource_name>%s)/(?P<pk>\w[\w/-]*)/download%s$' % (self._meta.resource_name, trailing_slash()), self.wrap_view('download'), name='api_download_export'),
        ]

    def download(self, request, **kwargs):
        """
        Download the original file that was uploaded.
        """
        # Allow POST so csrf token can come through
        self.method_check(request, allowed=['get', 'post'])
        self.is_authenticated(request)
        self.throttle_check(request)

        if 'pk' in kwargs:
            get_id = kwargs['pk']
        else:
            get_id = request.GET.get('id', '')

        export = Export.objects.get(id=get_id)
        path = export.get_path()

        self.log_throttled_access(request)

        response = HttpResponse(FileWrapper(open(path, 'r')), content_type=guess_type(export.filename)[0])
        response['Content-Disposition'] = 'attachment; filename=%s' % export.filename
        response['Content-Length'] = export.size

        return response


########NEW FILE########
__FILENAME__ = notifications
#!/usr/bin/env python

from tastypie.authorization import DjangoAuthorization

from panda.api.utils import PandaAuthentication, PandaSerializer, PandaModelResource
from panda.models import Notification, UserProxy 

class NotificationResource(PandaModelResource):
    """
    Access to user notifications.
    """
    class Meta:
        queryset = Notification.objects.all()
        resource_name = 'notification'
        
        authentication = PandaAuthentication()
        authorization = DjangoAuthorization()
        serializer = PandaSerializer()

        filtering = {
            "read_at": ('isnull')
        }

    def obj_create(self, bundle, request=None, **kwargs):
        # Because users may have authenticated via headers the request.user may
        # not be a full User instance. To be sure, we fetch one.
        user = UserProxy.objects.get(id=request.user.id)

        return super(NotificationResource, self).obj_create(bundle, request, recipient=user)

    def apply_authorization_limits(self, request, object_list):
        # Because users may have authenticated via headers the request.user may
        # not be a full User instance. To be sure, we fetch one.
        user = UserProxy.objects.get(id=request.user.id)

        return object_list.filter(recipient=user)

    def save_related(self, bundle):
        """
        Overriding this is a bit of a dirty-hack, but we really don't want
        Dataset being saved whenever a notification is updated (because it
        kicks off Solr indexing).
        """
        pass

    def put_list(self, request, **kwargs):
        """
        Allow emulating a ``PATCH`` request by passing ``?patch=true``.
        (As a workaround for IE's broken XMLHttpRequest.)
        """
        if request.GET.get('patch', 'false').lower() == 'true':
            return super(NotificationResource, self).patch_list(request, **kwargs)
        else:
            return super(NotificationResource, self).put_list(request, **kwargs)


########NEW FILE########
__FILENAME__ = related_uploads
#!/usr/bin/env python

from mimetypes import guess_type

from django.conf.urls.defaults import url
from django.core.exceptions import ObjectDoesNotExist
from django.core.servers.basehttp import FileWrapper
from django.http import HttpResponse
from django.utils.translation import ugettext_lazy as _
from tastypie import fields
from tastypie.authorization import DjangoAuthorization
from tastypie.exceptions import NotFound
from tastypie.utils.urls import trailing_slash

from panda.api.utils import PandaAuthentication, PandaModelResource, PandaSerializer
from panda.models import RelatedUpload

class RelatedUploadResource(PandaModelResource):
    """
    API resource for DataUploads.
    """
    from panda.api.users import UserResource

    filename = fields.CharField('filename', readonly=True)
    original_filename = fields.CharField('original_filename', readonly=True)
    size = fields.IntegerField('size', readonly=True)
    creator = fields.ForeignKey(UserResource, 'creator', full=True, readonly=True)
    creation_date = fields.DateTimeField('creation_date', readonly=True)
    dataset = fields.ForeignKey('panda.api.datasets.DatasetResource', 'dataset', null=True, readonly=True)
    title = fields.CharField('title', null=True)

    class Meta:
        queryset = RelatedUpload.objects.all()
        resource_name = 'related_upload'
        allowed_methods = ['get', 'put', 'delete']

        authentication = PandaAuthentication()
        authorization = DjangoAuthorization()
        serializer = PandaSerializer()

    def override_urls(self):
        """
        Add urls for search endpoint.
        """
        return [
            url(r'^(?P<resource_name>%s)/(?P<pk>\w[\w/-]*)/download%s$' % (self._meta.resource_name, trailing_slash()), self.wrap_view('download'), name='api_download_related_upload'),
        ]

    def obj_delete(self, request=None, **kwargs):
        """
        Override delete to also update related Dataset's metadata.
        """
        obj = kwargs.pop('_obj', None)

        if not hasattr(obj, 'delete'):
            try:
                obj = self.obj_get(request, **kwargs)
            except ObjectDoesNotExist:
                raise NotFound(_("A model instance matching the provided arguments could not be found."))

        obj.delete()

        if obj.dataset:
            obj.dataset.update_full_text()

    def download(self, request, **kwargs):
        """
        Download the original file that was uploaded.
        """
        # Allow POST so csrf token can come through
        self.method_check(request, allowed=['get', 'post'])
        self.is_authenticated(request)
        self.throttle_check(request)

        if 'pk' in kwargs:
            get_id = kwargs['pk']
        else:
            get_id = request.GET.get('id', '')

        upload = RelatedUpload.objects.get(id=get_id)
        path = upload.get_path()

        self.log_throttled_access(request)

        response = HttpResponse(FileWrapper(open(path, 'r')), content_type=guess_type(upload.original_filename)[0])
        response['Content-Disposition'] = 'attachment; filename=%s' % upload.original_filename
        response['Content-Length'] = upload.size

        return response


########NEW FILE########
__FILENAME__ = search_subscriptions
#!/usr/bin/env python

from tastypie import fields
from tastypie.authorization import DjangoAuthorization

from panda.api.utils import PandaAuthentication, PandaSerializer, PandaModelResource
from panda.models import SearchSubscription, UserProxy 

class SearchSubscriptionResource(PandaModelResource):
    """
    Access to user subscriptions.
    """
    from panda.api.category import CategoryResource
    from panda.api.datasets import DatasetResource

    dataset = fields.ForeignKey(DatasetResource, 'dataset', null=True, full=True)
    category = fields.ForeignKey(CategoryResource, 'category', null=True, full=True)

    class Meta:
        queryset = SearchSubscription.objects.all()
        resource_name = 'search_subscription'
        allowed_methods = ['get', 'post', 'delete']
        
        authentication = PandaAuthentication()
        authorization = DjangoAuthorization()
        serializer = PandaSerializer()

    def obj_create(self, bundle, request=None, **kwargs):
        # Because users may have authenticated via headers the request.user may
        # not be a full User instance. To be sure, we fetch one.
        user = UserProxy.objects.get(id=request.user.id)

        return super(SearchSubscriptionResource, self).obj_create(bundle, request, user=user)

    def apply_authorization_limits(self, request, object_list):
        # Because users may have authenticated via headers the request.user may
        # not be a full User instance. To be sure, we fetch one.
        user = UserProxy.objects.get(id=request.user.id)

        return object_list.filter(user=user)

    def save_related(self, bundle):
        """
        Overriding this is a bit of a dirty-hack, but we really don't want
        Dataset being saved whenever a subscription is updated (because it
        kicks off Solr indexing).
        """
        pass


########NEW FILE########
__FILENAME__ = tasks
#!/usr/bin/env python

from tastypie import fields
from tastypie.authorization import DjangoAuthorization

from panda.api.utils import PandaAuthentication, PandaSerializer, PandaModelResource
from panda.models import TaskStatus

class TaskResource(PandaModelResource):
    """
    Simple wrapper around django-celery's task API.
    """
    from panda.api.users import UserResource

    creator = fields.ForeignKey(UserResource, 'creator')

    class Meta:
        queryset = TaskStatus.objects.all()
        resource_name = 'task'
        allowed_methods = ['get']
        
        filtering = {
            'status': ('exact', 'in', ),
            'end': ('year', 'month', 'day')
        }

        authentication = PandaAuthentication()
        authorization = DjangoAuthorization()
        serializer = PandaSerializer()


########NEW FILE########
__FILENAME__ = users
#!/usr/bin/env python

from django.conf import settings
from django.conf.urls.defaults import url
from django.core.exceptions import ObjectDoesNotExist
from django.core.validators import email_re
from django.utils.translation import ugettext_lazy as _
from tastypie import fields
from tastypie import http
from tastypie.authorization import Authorization
from tastypie.exceptions import BadRequest, NotFound, ImmediateHttpResponse
from tastypie.resources import NOT_AVAILABLE
from tastypie.utils.urls import trailing_slash
from tastypie.validation import Validation

from panda.api.utils import PandaAuthentication, PandaSerializer, PandaModelResource
from panda.models import Export, UserProxy

class UserValidation(Validation):
    def is_valid(self, bundle, request=None):
        errors = {}

        if 'email' not in bundle.data or not bundle.data['email']:
            errors['email'] = [_('This field is required.')]
        elif not email_re.match(bundle.data['email']):
            errors['email'] = [_('Email address is not valid.')]

        return errors

class UserAuthorization(Authorization):
    def is_authorized(self, request, obj=None):
        """
        Superusers can change any other user. Regular users can only GET and
        PUT to their own user.
        """
        if request.method in ['GET', 'PUT']:
            return True
        else:
            return request.user.is_superuser

class UserResource(PandaModelResource):
    """
    API resource for Uploads.
    """
    # Write-only! See dehydrate().
    password = fields.CharField(attribute='password')

    class Meta:
        queryset = UserProxy.objects.all()
        resource_name = 'user'
        allowed_methods = ['get', 'post', 'put', 'delete']
        excludes = ['username', 'is_staff', 'is_superuser']
        always_return_data = True

        authentication = PandaAuthentication()
        authorization = UserAuthorization()
        validation = UserValidation()
        serializer = PandaSerializer()

    def override_urls(self):
        """
        Add urls for search endpoint.
        """
        return [
            url(r'^(?P<resource_name>%s)/(?P<pk>\w[\w/-]*)/login_help%s$' % (self._meta.resource_name, trailing_slash()), self.wrap_view('login_help'), name='api_user_login_help'),
        ]

    def hydrate_email(self, bundle):
        """
        Copy the email to the username field.
        """
        if 'email' in bundle.data:
            bundle.data['email'] = bundle.data['email'].lower()

        return bundle

    def dehydrate(self, bundle):
        """
        Always remove the password form the serialized bundle.
        """
        del bundle.data['password']

        user = bundle.obj

        if 'notifications' in bundle.request.GET and bundle.request.GET['notifications'].lower() == 'true':
            from panda.api.notifications import NotificationResource
            
            resource = NotificationResource()

            notifications = user.notifications.all()[:settings.PANDA_NOTIFICATIONS_TO_SHOW]

            bundles = [resource.build_bundle(obj=n) for n in notifications]
            notifications = [resource.full_dehydrate(b) for b in bundles]

            bundle.data['notifications'] = notifications

        if 'exports' in bundle.request.GET and bundle.request.GET['exports'].lower() == 'true':
            from panda.api.exports import ExportResource

            resource = ExportResource()

            exports = Export.objects.filter(creator=user)

            bundles = [resource.build_bundle(obj=e) for e in exports]
            exports = [resource.full_dehydrate(b) for b in bundles]

            bundle.data['exports'] = exports

        if 'datasets' in bundle.request.GET and bundle.request.GET['datasets'].lower() == 'true':
            from panda.api.datasets import DatasetResource

            resource = DatasetResource()

            datasets = user.datasets.all()

            bundles = [resource.build_bundle(obj=d) for d in datasets]
            datasets = [resource.simplify_bundle(resource.full_dehydrate(b)) for b in bundles]

            bundle.data['datasets'] = datasets

        if 'search_subscriptions' in bundle.request.GET and bundle.request.GET['search_subscriptions'].lower() == 'true':
            from panda.api.search_subscriptions import SearchSubscriptionResource

            resource = SearchSubscriptionResource()

            subscriptions = user.search_subscriptions.all()

            bundles = [resource.build_bundle(obj=s) for s in subscriptions]
            datasets = [resource.full_dehydrate(b) for b in bundles]

            bundle.data['subscriptions'] = datasets

        return bundle

    def obj_create(self, bundle, request=None, **kwargs):
        """
        Create user using email as username and optionally using a supplied password.
        """
        if not request.user.is_superuser:
            raise ImmediateHttpResponse(response=http.HttpUnauthorized())

        bundle.obj = self._meta.object_class()

        for key, value in kwargs.items():
            setattr(bundle.obj, key, value)

        bundle = self.full_hydrate(bundle)
        self.is_valid(bundle,request)

        bundle.obj.username = bundle.obj.email

        # Set password before saving so the post-save signal will correctly send email
        if bundle.data.get('password'):
            bundle.obj.set_password(bundle.data.get('password'))
        else:
            bundle.obj.set_unusable_password()

        if bundle.errors:
            self.error_response(bundle.errors, request)

        # Save FKs just in case.
        self.save_related(bundle)

        # Save parent
        bundle.obj.save()

        # Now pick up the M2M bits.
        m2m_bundle = self.hydrate_m2m(bundle)
        self.save_m2m(m2m_bundle)

        return bundle

    def put_detail(self, request, **kwargs):
        """
        Allow emulating a ``PATCH`` request by passing ``?patch=true``.
        (As a workaround for IE's broken XMLHttpRequest.)
        """
        if request.GET.get('patch', 'false').lower() == 'true':
            return super(UserResource, self).patch_detail(request, **kwargs)
        else:
            return super(UserResource, self).put_detail(request, **kwargs)

    def obj_update(self, bundle, request=None, skip_errors=False, **kwargs):
        """
        A ORM-specific implementation of ``obj_update``.
        """
        if not bundle.obj or not bundle.obj.pk:
            # Attempt to hydrate data from kwargs before doing a lookup for the object.
            # This step is needed so certain values (like datetime) will pass model validation.
            try:
                bundle.obj = self.get_object_list(bundle.request).model()
                bundle.data.update(kwargs)
                bundle = self.full_hydrate(bundle)
                lookup_kwargs = kwargs.copy()

                for key in kwargs.keys():
                    if key == 'pk':
                        continue
                    elif getattr(bundle.obj, key, NOT_AVAILABLE) is not NOT_AVAILABLE:
                        lookup_kwargs[key] = getattr(bundle.obj, key)
                    else:
                        del lookup_kwargs[key]
            except:
                # if there is trouble hydrating the data, fall back to just
                # using kwargs by itself (usually it only contains a "pk" key
                # and this will work fine.
                lookup_kwargs = kwargs

            try:
                bundle.obj = self.obj_get(bundle.request, **lookup_kwargs)
            except ObjectDoesNotExist:
                raise NotFound(_("A model instance matching the provided arguments could not be found."))

        # CHECK AUTHORIZATION 
        if request and not request.user.is_superuser and bundle.obj.id != request.user.id:
            raise ImmediateHttpResponse(response=http.HttpUnauthorized())

        bundle = self.full_hydrate(bundle)
        self.is_valid(bundle,request)

        if bundle.errors and not skip_errors:
            self.error_response(bundle.errors, request)

        # SET USERNAME FROM EMAIL
        bundle.obj.username = bundle.obj.email

        # SET PASSWORD
        if 'password' in bundle.data:
            bundle.obj.set_password(bundle.data.get('password'))

        # Save FKs just in case.
        self.save_related(bundle)

        # Save the main object.
        bundle.obj.save()

        # Now pick up the M2M bits.
        m2m_bundle = self.hydrate_m2m(bundle)
        self.save_m2m(m2m_bundle)

        return bundle

    def login_help(self, request, **kwargs):
        """
        Set the status of the "show_login_help" flag.
        """
        self.method_check(request, allowed=['post'])
        self.is_authenticated(request)
        self.throttle_check(request)

        if 'pk' in kwargs:
            get_id = int(kwargs['pk'])
        else:
            get_id = int(request.GET.get('id', ''))

        # CHECK AUTHORIZATION 
        if request and not request.user.is_superuser and get_id != request.user.id:
            raise ImmediateHttpResponse(response=http.HttpUnauthorized())

        deserialized = self.deserialize(request, request.raw_post_data, format=request.META.get('CONTENT_TYPE', 'application/json'))
        deserialized = self.alter_deserialized_list_data(request, deserialized)

        if not 'show_login_help' in deserialized:
            raise BadRequest(_("Invalid data sent."))

        user = UserProxy.objects.get(id=get_id)
        profile = user.get_profile()

        profile.show_login_help = deserialized['show_login_help']
        profile.save()

        return self.create_response(request, {}, response_class=http.HttpAccepted) 


########NEW FILE########
__FILENAME__ = utils
#!/usr/bin/env python

from urllib import unquote

from django.conf import settings
from django.conf.urls.defaults import url
from django.http import HttpResponse
from django.middleware.csrf import _sanitize_token, constant_time_compare
from django.utils.http import same_origin
from tastypie.authentication import ApiKeyAuthentication
from tastypie.bundle import Bundle
from tastypie.fields import ApiField, CharField
from tastypie.paginator import Paginator
from tastypie.resources import ModelResource, Resource
from tastypie.serializers import Serializer
from tastypie.utils.urls import trailing_slash

from panda.fields import JSONField
from panda.models import UserProxy

PANDA_CACHE_CONTROL = 'max-age=0,no-cache,no-store'

class JSONApiField(ApiField):
    """
    Custom ApiField for dealing with data from custom JSONFields.
    """
    dehydrated_type = 'json'
    help_text = 'JSON structured data.'
    
    def dehydrate(self, obj):
        return self.convert(super(JSONApiField, self).dehydrate(obj))
    
    def convert(self, value):
        if value is None:
            return None

        return value

class PandaResource(Resource):
    """
    Resource subclass that overrides cache headers.
    """
    def create_response(self, request, data, response_class=HttpResponse, **response_kwargs):
        """
        Override response generation to add ``Cache-Control: no-cache`` header.
        """
        response = super(PandaResource, self).create_response(request, data, response_class, **response_kwargs)
        response['Cache-Control'] = PANDA_CACHE_CONTROL

        return response

class PandaModelResource(ModelResource):
    """
    ModelResource subclass that supports JSONFields.
    """
    @classmethod
    def api_field_from_django_field(cls, f, default=CharField):
        """
        Overrides default field handling to support custom ListField and JSONField.
        """
        if isinstance(f, JSONField):
            return JSONApiField
    
        return super(PandaModelResource, cls).api_field_from_django_field(f, default)

    def create_response(self, request, data, response_class=HttpResponse, **response_kwargs):
        """
        Override response generation to add ``Cache-Control: no-cache`` header.
        """
        response = super(PandaModelResource, self).create_response(request, data, response_class, **response_kwargs)
        response['Cache-Control'] = PANDA_CACHE_CONTROL

        return response

class SluggedModelResource(PandaModelResource):
    """
    ModelResource that uses slugs for URLs.

    Also supports JSONFields, for simplicity.
    """
    def get_resource_uri(self, bundle_or_obj):
        """
        Handles generating a resource URI for a single resource.
        """
        kwargs = {
            'resource_name': self._meta.resource_name,
        }

        if isinstance(bundle_or_obj, Bundle):
            kwargs['slug'] = bundle_or_obj.obj.slug
        else:
            kwargs['slug'] = bundle_or_obj.slug

        if self._meta.api_name is not None:
            kwargs['api_name'] = self._meta.api_name

        return self._build_reverse_url("api_dispatch_detail", kwargs=kwargs)

    def base_urls(self):
        """
        The standard URLs this ``Resource`` should respond to.
        """
        # Due to the way Django parses URLs, ``get_multiple`` won't work without
        # a trailing slash.
        return [
            url(r"^(?P<resource_name>%s)%s$" % (self._meta.resource_name, trailing_slash()), self.wrap_view('dispatch_list'), name="api_dispatch_list"),
            url(r"^(?P<resource_name>%s)/schema%s$" % (self._meta.resource_name, trailing_slash()), self.wrap_view('get_schema'), name="api_get_schema"),
            url(r"^(?P<resource_name>%s)/set/(?P<slug_list>[\w\d_-]+)/$" % self._meta.resource_name, self.wrap_view('get_multiple'), name="api_get_multiple"),
            url(r"^(?P<resource_name>%s)/(?P<slug>[\w\d_-]+)%s$" % (self._meta.resource_name, trailing_slash()), self.wrap_view('dispatch_detail'), name="api_dispatch_detail"),
        ]

class PandaAuthentication(ApiKeyAuthentication):
    """
    Custom API Auth that authenticates via sessions, headers or querystring parameters. 
    """
    def try_sessions(self, request, **kwargs):
        """
        Attempt to authenticate with sessions.

        Cribbed from a newer version of Tastypie than we're using.
        """
        csrf_token = _sanitize_token(request.COOKIES.get(settings.CSRF_COOKIE_NAME, ''))

        if request.is_secure():
            referer = request.META.get('HTTP_REFERER')

            if referer is None:
                return False

            good_referer = 'https://%s/' % request.get_host()

            if not same_origin(referer, good_referer):
                return False

        # Tastypie docstring says accessing POST here isn't safe, but so far it's not causing any problems...
        # This is necessary for downloads that post the csrf token from an iframe
        request_csrf_token = request.META.get('HTTP_X_CSRFTOKEN', '') or request.POST.get('csrfmiddlewaretoken', '')

        if not constant_time_compare(request_csrf_token, csrf_token):
            return False

        return request.user.is_authenticated()

    def try_api_keys(self, request, **kwargs):
        """
        Attempt to authenticate with API keys in headers or parameters.
        """
        email = request.META.get('HTTP_PANDA_EMAIL') or request.GET.get('email')
        api_key = request.META.get('HTTP_PANDA_API_KEY') or request.GET.get('api_key')

        if email:
            email = unquote(email)

        if not email or not api_key:
            return False

        try:
            user = UserProxy.objects.get(username=email.lower())
        except (UserProxy.DoesNotExist, UserProxy.MultipleObjectsReturned):
            return False 

        if not user.is_active:
            return False
        
        request.user = user

        return self.get_key(user, api_key)

    def is_authenticated(self, request, **kwargs):
        authenticated = self.try_sessions(request, **kwargs)

        if authenticated:
            return True

        authenticated = self.try_api_keys(request, **kwargs)

        if authenticated:
            return True

        return self._unauthorized()

    def get_identifier(self, request):
        """
        Provides a unique string identifier for the requestor.

        This implementation returns the user's username.
        """
        return request.user.username

class PandaSerializer(Serializer):
    """
    A custom serializer that truncates microseconds from iso8601.
    """
    def format_datetime(self, data):
        return data.strftime('%Y-%m-%dT%H:%M:%S')

class PandaPaginator(Paginator):
    """
    A customized paginator that accepts count as a property, rather
    then inferring it from the length of the object array.
    """
    def __init__(self, request_data, objects, resource_uri=None, limit=None, offset=0, count=None):
        self.count = count
        super(PandaPaginator, self).__init__(request_data, objects, resource_uri, limit, offset)

    def get_count(self):
        if self.count is not None:
            return self.count
        
        return super(PandaPaginator, self).get_count()


########NEW FILE########
__FILENAME__ = config
#!/usr/bin/env python

from livesettings import config_register, BooleanValue, ConfigurationGroup, FloatValue, PositiveIntegerValue, StringValue
from django.utils.translation import ugettext_lazy as _

# Site domain settings
DOMAIN_GROUP = ConfigurationGroup(
    'DOMAIN',
    _('Site domain'),
    ordering=0
)

config_register(StringValue(
    DOMAIN_GROUP,
    'SITE_DOMAIN',
    description=_('Site domain to be referenced in outgoing email.'),
    default='localhost:8000'
))

# Email settings
EMAIL_GROUP = ConfigurationGroup(
    'EMAIL',
    _('Email'),
    ordering=1
)

config_register(BooleanValue(
    EMAIL_GROUP,
    'EMAIL_ENABLED',
    description=_('Enable email?'),
    help_text=_('If enabled, notifications and activation messages will be sent via email.'),
    default=False,
    ordering=0
))

config_register(StringValue(
    EMAIL_GROUP,
    'EMAIL_HOST',
    description=_('Hostname or IP of the SMTP server.'),
    default='localhost',
    ordering=1
))

config_register(PositiveIntegerValue(
    EMAIL_GROUP,
    'EMAIL_PORT',
    description=_('Port number of the SMTP server.'),
    default=1025,
    ordering=2
))

config_register(StringValue(
    EMAIL_GROUP,
    'EMAIL_HOST_USER',
    description=_('Username for the SMTP server.'),
    default='',
    ordering=3
))

config_register(StringValue(
    EMAIL_GROUP,
    'EMAIL_HOST_PASSWORD',
    description=_('Password for the SMTP server.'),
    default='',
    ordering=4
))

config_register(BooleanValue(
    EMAIL_GROUP,
    'EMAIL_USE_TLS',
    description=_('Use TLS encryption when connecting to the SMTP server?'),
    default=False,
    ordering=5
))

config_register(StringValue(
    EMAIL_GROUP,
    'DEFAULT_FROM_EMAIL',
    description=_('Email address that PANDA messages should appear to come from.'),
    default='do.not.reply@pandaproject.net',
    ordering=6
))

# Miscellaneous settings
MISC_GROUP = ConfigurationGroup(
    'MISC',
    _('Miscellaneous'),
    ordering=2
)

config_register(BooleanValue(
    MISC_GROUP,
    'DEMO_MODE_ENABLED',
    description=_('Enable demo mode?'),
    help_text=_('In demo mode the login fields will automatically be prepopulated with the default username and password.'),
    default=False,
    ordering=0
))

config_register(PositiveIntegerValue(
    MISC_GROUP,
    'WARN_UPLOAD_SIZE',
    description=_('File size at which a warning about large file uploads is issued, in bytes.'),
    help_text=_('The default value is equivalent to 100MB.'),
    default=104857600,
    ordering=1
))

config_register(PositiveIntegerValue(
    MISC_GROUP,
    'MAX_UPLOAD_SIZE',
    description=_('Maximum size allowed for user-uploaded files, in bytes.'),
    help_text=_('The default value is equivalent to 1GB.'),
    default=1073741824,
    ordering=2
))

# Performance settings
PERF_GROUP = ConfigurationGroup(
    'PERF',
    _('Performance'),
    ordering=3
)

config_register(FloatValue(
    PERF_GROUP,
    'TASK_THROTTLE',
    description=_('Number of seconds to wait between processing batches of data.'),
    help_text=_('A larger number will result in slower imports and exports, but better responsiveness from the PANDA user interface.'),
    default=0.5,
    ordering=1
))


########NEW FILE########
__FILENAME__ = exceptions
#!/usr/bin/env python
from django.utils.translation import ugettext as _

class DatasetLockedError(Exception):
    """
    Exception raised when a lock can not be acquired on a dataset.
    """
    pass

class DataUploadNotDeletable(Exception):
    """
    Exception raised when a DataUpload can not be deleted.
    """
    pass

class DataSamplingError(Exception):
    """
    Exception raised when data can't be sampled from a file,
    such as when unexpected encodings are encountered.
    """
    pass

class DataImportError(Exception):
    """
    Exception raised when a DataImport fails synchronously
    due to an unsupported file type, mismatched columns, etc.
    """
    pass

class NotSniffableError(Exception):
    """
    Exception raised when a file's dialect could not be inferred
    automatically.
    """
    pass

class TypeInferenceError(Exception):
    """
    Exception raised when a column's type can not be inferred.
    """
    pass

class TypeCoercionError(Exception):
    """
    Exception raised when a value can not be coerced to a given type.
    """
    def __init__(self, value, normal_type):
        self.value = value
        self.normal_type = normal_type
        msg = _('Unable to convert "%(value)s" to type %(normal_type)s') \
            % {'value': value, 'normal_type': normal_type}
        super(TypeCoercionError, self).__init__(value, normal_type)


########NEW FILE########
__FILENAME__ = fields
#!/usr/bin/env python

from django.core.serializers.json import DjangoJSONEncoder
from django.db import models
from django.utils import simplejson as json
from south.modelsinspector import add_introspection_rules

class JSONField(models.TextField):
    """
    Store arbitrary JSON in a Model field.
    """
    # Used so to_python() is called
    __metaclass__ = models.SubfieldBase

    def to_python(self, value):
        """
        Convert string value to JSON after its loaded from the database.
        """
        if value == "":
            return None

        try:
            if isinstance(value, basestring):
                return json.loads(value)
        except ValueError:
            pass

        return value

    def get_prep_value(self, value):
        """
        Convert our JSON object to a string before being saved.
        """
        if value == "":
            return None

        if isinstance(value, dict) or isinstance(value, list):
            value = json.dumps(value, cls=DjangoJSONEncoder)

        return super(JSONField, self).get_prep_value(value)

    def value_to_string(self, obj):
        """
        Called by the serializer.
        """
        value = self._get_val_from_obj(obj)

        return self.get_db_prep_value(value)

add_introspection_rules([], ["^panda\.fields\.JSONField"])


########NEW FILE########
__FILENAME__ = manual_import
#!/usr/bin/env python

import os

from django.conf import settings
from django.core.management.base import BaseCommand
from django.utils.translation import ugettext as _
from livesettings import config_value

from panda.models import Dataset, DataUpload, UserProxy

class Command(BaseCommand):
    args = '<dataset_filename user_email>'
    help = _('Manually import data for when the web UI fails. See http://panda.readthedocs.org/en/latest/manual_imports.html')

    def handle(self, *args, **options):
        if len(args) < 2:
            self.stderr.write(_('You must specify a filename and user.\n'))
            return

        filename = args[0]
        email = args[1]

        path = os.path.join(settings.MEDIA_ROOT, filename)

        if not os.path.exists(path):
            self.stderr.write(_('File does not exist!\n'))
            return

        size = os.path.getsize(path)

        try:
            creator = UserProxy.objects.get(email=email)
        except UserProxy.DoesNotExist:
            self.stderr.write(_('User does not exist!\n'))
            return

        upload = DataUpload.objects.create(
            filename=filename,
            original_filename=filename,
            size=size,
            creator=creator,
            dataset=None,
            encoding='utf-8')
     
        dataset = Dataset.objects.create(
            name=filename,
            creator=creator,
            initial_upload=upload)

        self.stdout.write('%s http://%s/#dataset/%s\n' % (_('Dataset created:'), config_value('DOMAIN', 'SITE_DOMAIN'), dataset.slug))

        dataset.import_data(creator, upload)
        
        dataset.update_full_text()

        self.stdout.write(_('Import started. Check dataset page for progress.\n'))

########NEW FILE########
__FILENAME__ = purge_orphaned_uploads
#!/usr/bin/env python

from itertools import chain
import os

from django.conf import settings
from django.core.management.base import BaseCommand
from django.utils.translation import ugettext as _

from optparse import make_option
from panda.models import DataUpload, RelatedUpload

class Command(BaseCommand):
    help = _('Audit uploads and local files, deleting any not associated with a dataset.')
    option_list = BaseCommand.option_list + (
        make_option('--fake',
            action='store_true',
            dest='fake',
            default=False,
            help=_('Only describe what files would be deleted, don\'t actually delete them.')),
        )

    def handle(self, *args, **options):
        if options['fake']:
            self.stdout.write(_('Running in fake mode! No files will actually be deleted!'))

        local_files = os.listdir(settings.MEDIA_ROOT)
        data_uploads = DataUpload.objects.all()
        related_uploads = RelatedUpload.objects.all()

        for upload in chain(data_uploads, related_uploads):
            # This file is accounted for
            try:
                local_files.remove(upload.filename)
            except ValueError:
                pass

            if not upload.dataset:
                if options['fake']:
                    self.stdout.write(_('Would delete upload: %s\n') % upload)
                else:
                    self.stdout.write(_('Deleted upload: %s\n') % upload)
                    upload.delete()

        for f in local_files:
            path = os.path.join(settings.MEDIA_ROOT, f)

            if options['fake']:
                self.stdout.write(_('Would delete file: %s\n') % path)
            else:
                self.stdout.write(_('Deleted file: %s\n') % path)
                os.remove(path)


########NEW FILE########
__FILENAME__ = reindex_datasets
#!/usr/bin/env python

from django.core.management.base import NoArgsCommand
from django.utils.translation import ugettext as _

from panda.models import Dataset

class Command(NoArgsCommand):
    help = _('Reindex all datasets')

    def handle_noargs(self, **options):
        for dataset in Dataset.objects.all():
            dataset.update_full_text()
            self.stdout.write(_('Updated: %s\n') % dataset.name)
        
        self.stdout.write(_('Done!\n'))


########NEW FILE########
__FILENAME__ = middleware
#!/usr/bin/env python

from django.middleware.csrf import get_token 

class CsrfCookieUsedMiddleware(object): 
    """ 
    Simple middleware that ensures Django's CSRF middleware will 
    always include the CSRF cookie on outgoing responses. 

    See: https://groups.google.com/d/msg/django-developers/Zi9_AyfBd_0/t_TlsL8-CHMJ
    """ 
    def process_request(self, request): 
        get_token(request)


########NEW FILE########
__FILENAME__ = 0001_initial
# encoding: utf-8
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models

class Migration(SchemaMigration):

    def forwards(self, orm):
        
        # Adding model 'Category'
        db.create_table('panda_category', (
            ('id', self.gf('django.db.models.fields.AutoField')(primary_key=True)),
            ('slug', self.gf('django.db.models.fields.SlugField')(max_length=256, db_index=True)),
            ('name', self.gf('django.db.models.fields.CharField')(max_length=64)),
        ))
        db.send_create_signal('panda', ['Category'])

        # Adding model 'TaskStatus'
        db.create_table('panda_taskstatus', (
            ('id', self.gf('django.db.models.fields.AutoField')(primary_key=True)),
            ('task_name', self.gf('django.db.models.fields.CharField')(max_length=255)),
            ('status', self.gf('django.db.models.fields.CharField')(default='PENDING', max_length=50)),
            ('message', self.gf('django.db.models.fields.CharField')(max_length=255, blank=True)),
            ('start', self.gf('django.db.models.fields.DateTimeField')(null=True)),
            ('end', self.gf('django.db.models.fields.DateTimeField')(null=True)),
            ('traceback', self.gf('django.db.models.fields.TextField')(default=None, null=True, blank=True)),
            ('creator', self.gf('django.db.models.fields.related.ForeignKey')(related_name='tasks', null=True, to=orm['auth.User'])),
        ))
        db.send_create_signal('panda', ['TaskStatus'])

        # Adding model 'Dataset'
        db.create_table('panda_dataset', (
            ('id', self.gf('django.db.models.fields.AutoField')(primary_key=True)),
            ('slug', self.gf('django.db.models.fields.SlugField')(max_length=256, db_index=True)),
            ('name', self.gf('django.db.models.fields.CharField')(max_length=256)),
            ('description', self.gf('django.db.models.fields.TextField')(blank=True)),
            ('initial_upload', self.gf('django.db.models.fields.related.ForeignKey')(blank=True, related_name='initial_upload_for', null=True, to=orm['panda.DataUpload'])),
            ('columns', self.gf('panda.fields.JSONField')(default=None, null=True)),
            ('sample_data', self.gf('panda.fields.JSONField')(default=None, null=True)),
            ('row_count', self.gf('django.db.models.fields.IntegerField')(null=True, blank=True)),
            ('current_task', self.gf('django.db.models.fields.related.ForeignKey')(to=orm['panda.TaskStatus'], null=True, blank=True)),
            ('creation_date', self.gf('django.db.models.fields.DateTimeField')(null=True)),
            ('creator', self.gf('django.db.models.fields.related.ForeignKey')(related_name='datasets', to=orm['auth.User'])),
            ('last_modified', self.gf('django.db.models.fields.DateTimeField')(default=None, null=True, blank=True)),
            ('last_modification', self.gf('django.db.models.fields.TextField')(default=None, null=True, blank=True)),
            ('last_modified_by', self.gf('django.db.models.fields.related.ForeignKey')(to=orm['auth.User'], null=True, blank=True)),
        ))
        db.send_create_signal('panda', ['Dataset'])

        # Adding M2M table for field categories on 'Dataset'
        db.create_table('panda_dataset_categories', (
            ('id', models.AutoField(verbose_name='ID', primary_key=True, auto_created=True)),
            ('dataset', models.ForeignKey(orm['panda.dataset'], null=False)),
            ('category', models.ForeignKey(orm['panda.category'], null=False))
        ))
        db.create_unique('panda_dataset_categories', ['dataset_id', 'category_id'])

        # Adding model 'DataUpload'
        db.create_table('panda_dataupload', (
            ('id', self.gf('django.db.models.fields.AutoField')(primary_key=True)),
            ('filename', self.gf('django.db.models.fields.CharField')(max_length=256)),
            ('original_filename', self.gf('django.db.models.fields.CharField')(max_length=256)),
            ('size', self.gf('django.db.models.fields.IntegerField')()),
            ('creator', self.gf('django.db.models.fields.related.ForeignKey')(to=orm['auth.User'])),
            ('creation_date', self.gf('django.db.models.fields.DateTimeField')()),
            ('dataset', self.gf('django.db.models.fields.related.ForeignKey')(related_name='data_uploads', null=True, to=orm['panda.Dataset'])),
            ('data_type', self.gf('django.db.models.fields.CharField')(max_length=4, null=True, blank=True)),
            ('encoding', self.gf('django.db.models.fields.CharField')(default='utf-8', max_length=32)),
            ('dialect', self.gf('panda.fields.JSONField')(null=True)),
            ('columns', self.gf('panda.fields.JSONField')(null=True)),
            ('sample_data', self.gf('panda.fields.JSONField')(null=True)),
            ('imported', self.gf('django.db.models.fields.BooleanField')(default=False)),
        ))
        db.send_create_signal('panda', ['DataUpload'])

        # Adding model 'Export'
        db.create_table('panda_export', (
            ('id', self.gf('django.db.models.fields.AutoField')(primary_key=True)),
            ('filename', self.gf('django.db.models.fields.CharField')(max_length=256)),
            ('original_filename', self.gf('django.db.models.fields.CharField')(max_length=256)),
            ('size', self.gf('django.db.models.fields.IntegerField')()),
            ('creator', self.gf('django.db.models.fields.related.ForeignKey')(to=orm['auth.User'])),
            ('creation_date', self.gf('django.db.models.fields.DateTimeField')()),
            ('dataset', self.gf('django.db.models.fields.related.ForeignKey')(related_name='exports', to=orm['panda.Dataset'])),
        ))
        db.send_create_signal('panda', ['Export'])

        # Adding model 'Notification'
        db.create_table('panda_notification', (
            ('id', self.gf('django.db.models.fields.AutoField')(primary_key=True)),
            ('recipient', self.gf('django.db.models.fields.related.ForeignKey')(related_name='notifications', to=orm['auth.User'])),
            ('message', self.gf('django.db.models.fields.TextField')()),
            ('type', self.gf('django.db.models.fields.CharField')(default='Info', max_length=16)),
            ('sent_at', self.gf('django.db.models.fields.DateTimeField')(auto_now=True, blank=True)),
            ('read_at', self.gf('django.db.models.fields.DateTimeField')(default=None, null=True, blank=True)),
            ('related_task', self.gf('django.db.models.fields.related.ForeignKey')(default=None, to=orm['panda.TaskStatus'], null=True)),
            ('related_dataset', self.gf('django.db.models.fields.related.ForeignKey')(default=None, to=orm['panda.Dataset'], null=True)),
        ))
        db.send_create_signal('panda', ['Notification'])

        # Adding model 'RelatedUpload'
        db.create_table('panda_relatedupload', (
            ('id', self.gf('django.db.models.fields.AutoField')(primary_key=True)),
            ('filename', self.gf('django.db.models.fields.CharField')(max_length=256)),
            ('original_filename', self.gf('django.db.models.fields.CharField')(max_length=256)),
            ('size', self.gf('django.db.models.fields.IntegerField')()),
            ('creator', self.gf('django.db.models.fields.related.ForeignKey')(to=orm['auth.User'])),
            ('creation_date', self.gf('django.db.models.fields.DateTimeField')()),
            ('dataset', self.gf('django.db.models.fields.related.ForeignKey')(related_name='related_uploads', to=orm['panda.Dataset'])),
        ))
        db.send_create_signal('panda', ['RelatedUpload'])

        # Adding model 'UserProfile'
        db.create_table('panda_userprofile', (
            ('id', self.gf('django.db.models.fields.AutoField')(primary_key=True)),
            ('user', self.gf('django.db.models.fields.related.OneToOneField')(to=orm['auth.User'], unique=True)),
            ('activation_key', self.gf('django.db.models.fields.CharField')(max_length=40)),
        ))
        db.send_create_signal('panda', ['UserProfile'])


    def backwards(self, orm):
        
        # Deleting model 'Category'
        db.delete_table('panda_category')

        # Deleting model 'TaskStatus'
        db.delete_table('panda_taskstatus')

        # Deleting model 'Dataset'
        db.delete_table('panda_dataset')

        # Removing M2M table for field categories on 'Dataset'
        db.delete_table('panda_dataset_categories')

        # Deleting model 'DataUpload'
        db.delete_table('panda_dataupload')

        # Deleting model 'Export'
        db.delete_table('panda_export')

        # Deleting model 'Notification'
        db.delete_table('panda_notification')

        # Deleting model 'RelatedUpload'
        db.delete_table('panda_relatedupload')

        # Deleting model 'UserProfile'
        db.delete_table('panda_userprofile')


    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256', 'db_index': 'True'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'columns': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256', 'db_index': 'True'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']

########NEW FILE########
__FILENAME__ = 0002_auto__add_field_dataset_locked__add_field_dataset_locked_at
# encoding: utf-8
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models

class Migration(SchemaMigration):

    def forwards(self, orm):
        
        # Adding field 'Dataset.locked'
        db.add_column('panda_dataset', 'locked', self.gf('django.db.models.fields.BooleanField')(default=False), keep_default=False)

        # Adding field 'Dataset.locked_at'
        db.add_column('panda_dataset', 'locked_at', self.gf('django.db.models.fields.DateTimeField')(default=None, null=True), keep_default=False)


    def backwards(self, orm):
        
        # Deleting field 'Dataset.locked'
        db.delete_column('panda_dataset', 'locked')

        # Deleting field 'Dataset.locked_at'
        db.delete_column('panda_dataset', 'locked_at')


    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256', 'db_index': 'True'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'columns': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256', 'db_index': 'True'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']

########NEW FILE########
__FILENAME__ = 0003_auto__add_field_dataupload_guessed_types
# encoding: utf-8
import datetime
import os

from south.db import db
from south.v2 import SchemaMigration
from django.conf import settings

from panda.models import DataUpload
from panda import utils

class Migration(SchemaMigration):

    def get_path(self, data_upload):
        """
        From BaseUpload abstract model.
        """
        return os.path.join(settings.MEDIA_ROOT, data_upload.filename)

    def dialect_as_parameters(self, data_upload):
        """
        From DataUpload model.
        """
        dialect_params = {}

        # This code is absolutely terrifying
        # (Also, it works.)
        for k, v in data_upload.dialect.items():
            if isinstance(v, basestring):
                dialect_params[k] = v.decode('string_escape')
            else:
                dialect_params[k] = v

        return dialect_params

    def forwards(self, orm):
        
        # Adding field 'DataUpload.guessed_types'
        db.add_column('panda_dataupload', 'guessed_types', self.gf('panda.fields.JSONField')(null=True), keep_default=False)

        db.commit_transaction()     # Commit the first transaction
        db.start_transaction()      # Start the second, committed on completion

        if not db.dry_run:
            for data_upload in orm.DataUpload.objects.all():
                path = self.get_path(data_upload) 
                try:
                    data_upload.guessed_types = utils.guess_column_types(data_upload.data_type, path, self.dialect_as_parameters(data_upload), encoding=data_upload.encoding)
                    data_upload.save()
                except IOError:
                    # File does not exist on disk
                    continue

    def backwards(self, orm):
        
        # Deleting field 'DataUpload.guessed_types'
        db.delete_column('panda_dataupload', 'guessed_types')


    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256', 'db_index': 'True'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'columns': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256', 'db_index': 'True'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']

########NEW FILE########
__FILENAME__ = 0004_auto__add_field_dataset_column_types__add_field_dataset_typed_column_n
# encoding: utf-8
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models

class Migration(SchemaMigration):

    def forwards(self, orm):
        
        # Adding field 'Dataset.column_types'
        db.add_column('panda_dataset', 'column_types', self.gf('panda.fields.JSONField')(default=None, null=True), keep_default=False)

        # Adding field 'Dataset.typed_column_names'
        db.add_column('panda_dataset', 'typed_column_names', self.gf('panda.fields.JSONField')(default=None, null=True), keep_default=False)

        db.commit_transaction()     # Commit the first transaction
        db.start_transaction()      # Start the second, committed on completion

        if not db.dry_run:
            for dataset in orm.Dataset.objects.all():
                if dataset.initial_upload:
                    dataset.column_types = dataset.initial_upload.guessed_types

                    # Account for bug where columns sometimes were not copied across
                    if not dataset.columns:
                        dataset.columns = dataset.initial_upload.columns
                else:
                    dataset.column_types = ['unicode' for c in dataset.columns]

                dataset.typed_column_names = [None for c in dataset.columns]

                dataset.save()

    def backwards(self, orm):
        
        # Deleting field 'Dataset.column_types'
        db.delete_column('panda_dataset', 'column_types')

        # Deleting field 'Dataset.typed_column_names'
        db.delete_column('panda_dataset', 'typed_column_names')


    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256', 'db_index': 'True'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_types': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'columns': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256', 'db_index': 'True'}),
            'typed_column_names': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']

########NEW FILE########
__FILENAME__ = 0005_auto__add_field_dataset_typed_columns
# encoding: utf-8
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models

class Migration(SchemaMigration):

    def forwards(self, orm):
        
        # Adding field 'Dataset.typed_columns'
        db.add_column('panda_dataset', 'typed_columns', self.gf('panda.fields.JSONField')(default=None, null=True), keep_default=False)

        db.commit_transaction()     # Commit the first transaction
        db.start_transaction()      # Start the second, committed on completion

        if not db.dry_run:
            for dataset in orm.Dataset.objects.all():
                dataset.typed_columns = [False for c in dataset.columns]

                dataset.save()

    def backwards(self, orm):
        
        # Deleting field 'Dataset.typed_columns'
        db.delete_column('panda_dataset', 'typed_columns')


    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256', 'db_index': 'True'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_types': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'columns': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256', 'db_index': 'True'}),
            'typed_column_names': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'typed_columns': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']

########NEW FILE########
__FILENAME__ = 0006_auto__add_field_dataset_column_schema
# encoding: utf-8
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models

class Migration(SchemaMigration):

    def forwards(self, orm):
        
        # Adding field 'Dataset.column_schema'
        db.add_column('panda_dataset', 'column_schema', self.gf('panda.fields.JSONField')(default=None, null=True), keep_default=False)

        db.commit_transaction()     # Commit the first transaction
        db.start_transaction()      # Start the second, committed on completion

        if not db.dry_run:
            for dataset in orm.Dataset.objects.all():
                schema = []

                for i, name in enumerate(dataset.columns):
                    c = { 'name': name }

                    if dataset.typed_columns:
                        c['indexed'] = dataset.typed_columns[i]
                    else:
                        c['indexed'] = False

                    if dataset.column_types:
                        c['type'] = dataset.column_types[i]
                    else:
                        c['type'] = None

                    if dataset.typed_column_names:
                        c['indexed_name'] = dataset.typed_column_names[i]
                    else:
                        c['indexed_name'] = None

                    c['min'] = None
                    c['max'] = None

                    schema.append(c)

                dataset.column_schema = schema
                dataset.save()

    def backwards(self, orm):
        
        # Deleting field 'Dataset.column_schema'
        db.delete_column('panda_dataset', 'column_schema')


    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256', 'db_index': 'True'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'column_types': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'columns': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256', 'db_index': 'True'}),
            'typed_column_names': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'typed_columns': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']

########NEW FILE########
__FILENAME__ = 0007_auto__del_field_dataset_typed_columns__del_field_dataset_column_types_
# encoding: utf-8
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models

class Migration(SchemaMigration):

    def forwards(self, orm):
        
        # Deleting field 'Dataset.typed_columns'
        db.delete_column('panda_dataset', 'typed_columns')

        # Deleting field 'Dataset.column_types'
        db.delete_column('panda_dataset', 'column_types')

        # Deleting field 'Dataset.typed_column_names'
        db.delete_column('panda_dataset', 'typed_column_names')

        # Deleting field 'Dataset.columns'
        db.delete_column('panda_dataset', 'columns')


    def backwards(self, orm):
        
        # Adding field 'Dataset.typed_columns'
        db.add_column('panda_dataset', 'typed_columns', self.gf('panda.fields.JSONField')(default=None, null=True), keep_default=False)

        # Adding field 'Dataset.column_types'
        db.add_column('panda_dataset', 'column_types', self.gf('panda.fields.JSONField')(default=None, null=True), keep_default=False)

        # Adding field 'Dataset.typed_column_names'
        db.add_column('panda_dataset', 'typed_column_names', self.gf('panda.fields.JSONField')(default=None, null=True), keep_default=False)

        # Adding field 'Dataset.columns'
        db.add_column('panda_dataset', 'columns', self.gf('panda.fields.JSONField')(default=None, null=True), keep_default=False)

        db.commit_transaction()     # Commit the first transaction
        db.start_transaction()      # Start the second, committed on completion

        if not db.dry_run:
            for dataset in orm.Dataset.objects.all():
                columns = []
                typed_columns = []
                column_types = []
                typed_column_names = []

                for schema in dataset.column_schema:
                    columns.append(schema['name'])
                    typed_columns.append(schema['indexed'])
                    column_types.append(schema['type'])
                    typed_column_names.append(schema['indexed_name'])

                dataset.columns = columns
                dataset.typed_columns = typed_columns
                dataset.column_types = column_types
                dataset.typed_column_names = typed_column_names
                dataset.save()


    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256', 'db_index': 'True'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256', 'db_index': 'True'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']

########NEW FILE########
__FILENAME__ = 0008_unlock_datasets
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import DataMigration
from django.db import models

class Migration(DataMigration):

    def forwards(self, orm):
        if not db.dry_run:
            # Unlock all datasets that may have been incorrectly locked. Issue #603.
            for dataset in orm.Dataset.objects.all():
                if dataset.locked:
                    dataset.locked = False
                    dataset.save()

    def backwards(self, orm):
        pass


    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
    symmetrical = True

########NEW FILE########
__FILENAME__ = 0009_auto__add_activitylog__add_unique_activitylog_user_when
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding model 'ActivityLog'
        db.create_table('panda_activitylog', (
            ('id', self.gf('django.db.models.fields.AutoField')(primary_key=True)),
            ('user', self.gf('django.db.models.fields.related.ForeignKey')(related_name='activity_logs', to=orm['auth.User'])),
            ('when', self.gf('django.db.models.fields.DateField')()),
        ))
        db.send_create_signal('panda', ['ActivityLog'])

        # Adding unique constraint on 'ActivityLog', fields ['user', 'when']
        db.create_unique('panda_activitylog', ['user_id', 'when'])

    def backwards(self, orm):
        # Removing unique constraint on 'ActivityLog', fields ['user', 'when']
        db.delete_unique('panda_activitylog', ['user_id', 'when'])

        # Deleting model 'ActivityLog'
        db.delete_table('panda_activitylog')

    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
########NEW FILE########
__FILENAME__ = 0010_auto__add_searchlog
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding model 'SearchLog'
        db.create_table('panda_searchlog', (
            ('id', self.gf('django.db.models.fields.AutoField')(primary_key=True)),
            ('user', self.gf('django.db.models.fields.related.ForeignKey')(related_name='search_logs', to=orm['auth.User'])),
            ('dataset', self.gf('django.db.models.fields.related.ForeignKey')(default=None, related_name='searches', null=True, to=orm['panda.Dataset'])),
            ('query', self.gf('django.db.models.fields.CharField')(max_length=256)),
            ('when', self.gf('django.db.models.fields.DateTimeField')()),
        ))
        db.send_create_signal('panda', ['SearchLog'])

    def backwards(self, orm):
        # Deleting model 'SearchLog'
        db.delete_table('panda_searchlog')

    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']

########NEW FILE########
__FILENAME__ = 0011_add_activitylog_permissions
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import DataMigration
from django.db import models

class Migration(DataMigration):

    def forwards(self, orm):
        """
        This migration will fail if run against a clean database (fresh setup)
        This is fine because the permission will be installed from the fixture.
        """
        try:
            perm = orm['auth.permission'].objects.get(codename='add_activitylog')
            group = orm['auth.group'].objects.get(name='panda_user')

            group.permissions.add(perm)
        except:
            pass

    def backwards(self, orm):
        perm = orm['auth.permission'].objects.get(codename='add_activitylog')
        group = orm['auth.group'].objects.get(name='panda_user')

        group.permissions.remove(perm)

    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
    symmetrical = True

########NEW FILE########
__FILENAME__ = 0012_auto__chg_field_export_dataset
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):

        # Changing field 'Export.dataset'
        db.alter_column('panda_export', 'dataset_id', self.gf('django.db.models.fields.related.ForeignKey')(null=True, to=orm['panda.Dataset']))
    def backwards(self, orm):

        # Changing field 'Export.dataset'
        db.alter_column('panda_export', 'dataset_id', self.gf('django.db.models.fields.related.ForeignKey')(default=None, to=orm['panda.Dataset']))
    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
########NEW FILE########
__FILENAME__ = 0013_auto__chg_field_userprofile_activation_key
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):

        # Changing field 'UserProfile.activation_key'
        db.alter_column('panda_userprofile', 'activation_key', self.gf('django.db.models.fields.CharField')(max_length=40, null=True))
    def backwards(self, orm):

        # Changing field 'UserProfile.activation_key'
        db.alter_column('panda_userprofile', 'activation_key', self.gf('django.db.models.fields.CharField')(default=None, max_length=40))
    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
########NEW FILE########
__FILENAME__ = 0014_delete_used_activation_keys
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import DataMigration
from django.db import models

class Migration(DataMigration):

    def forwards(self, orm):
        for user_profile in orm.UserProfile.objects.all():
            if user_profile.user.is_active:
                user_profile.activation_key = None
                user_profile.save()

    def backwards(self, orm):
        pass


    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
    symmetrical = True

########NEW FILE########
__FILENAME__ = 0015_lowercase_emails
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import DataMigration
from django.db import models

class Migration(DataMigration):

    def forwards(self, orm):
        for user in orm['auth.user'].objects.all():
            user.username = user.username.lower()
            user.email = user.username
            user.save()


    def backwards(self, orm):
        pass


    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
    symmetrical = True

########NEW FILE########
__FILENAME__ = 0016_auto__add_field_notification_related_export
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding field 'Notification.related_export'
        db.add_column('panda_notification', 'related_export',
                      self.gf('django.db.models.fields.related.ForeignKey')(default=None, to=orm['panda.Export'], null=True),
                      keep_default=False)

    def backwards(self, orm):
        # Deleting field 'Notification.related_export'
        db.delete_column('panda_notification', 'related_export_id')

    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_export': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Export']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
########NEW FILE########
__FILENAME__ = 0017_auto__add_field_userprofile_activation_key_expiration
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding field 'UserProfile.activation_key_expiration'
        db.add_column('panda_userprofile', 'activation_key_expiration',
                      self.gf('django.db.models.fields.DateTimeField')(default=datetime.datetime(2012, 5, 9, 0, 0)),
                      keep_default=False)

    def backwards(self, orm):
        # Deleting field 'UserProfile.activation_key_expiration'
        db.delete_column('panda_userprofile', 'activation_key_expiration')

    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_export': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Export']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'activation_key_expiration': ('django.db.models.fields.DateTimeField', [], {}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
########NEW FILE########
__FILENAME__ = 0018_auto__add_field_taskstatus_task_description
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding field 'TaskStatus.task_description'
        db.add_column('panda_taskstatus', 'task_description',
                      self.gf('django.db.models.fields.TextField')(default=''),
                      keep_default=False)

    def backwards(self, orm):
        # Deleting field 'TaskStatus.task_description'
        db.delete_column('panda_taskstatus', 'task_description')

    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_export': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Export']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_description': ('django.db.models.fields.TextField', [], {}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'activation_key_expiration': ('django.db.models.fields.DateTimeField', [], {}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
########NEW FILE########
__FILENAME__ = 0019_auto__add_searchsubscription
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding model 'SearchSubscription'
        db.create_table('panda_searchsubscription', (
            ('id', self.gf('django.db.models.fields.AutoField')(primary_key=True)),
            ('user', self.gf('django.db.models.fields.related.ForeignKey')(related_name='subscribed_searches', to=orm['auth.User'])),
            ('dataset', self.gf('django.db.models.fields.related.ForeignKey')(default=None, related_name='subscribed_searches', null=True, to=orm['panda.Dataset'])),
            ('query', self.gf('django.db.models.fields.CharField')(max_length=256)),
            ('last_run', self.gf('django.db.models.fields.DateTimeField')(null=True)),
        ))
        db.send_create_signal('panda', ['SearchSubscription'])

    def backwards(self, orm):
        # Deleting model 'SearchSubscription'
        db.delete_table('panda_searchsubscription')

    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_export': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Export']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.searchsubscription': {
            'Meta': {'object_name': 'SearchSubscription'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'subscribed_searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'last_run': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'subscribed_searches'", 'to': "orm['auth.User']"})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_description': ('django.db.models.fields.TextField', [], {}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'activation_key_expiration': ('django.db.models.fields.DateTimeField', [], {}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
########NEW FILE########
__FILENAME__ = 0020_auto__add_field_searchsubscription_query_url__chg_field_searchsubscrip
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding field 'SearchSubscription.query_url'
        db.add_column('panda_searchsubscription', 'query_url',
                      self.gf('django.db.models.fields.CharField')(default='*', max_length=256),
                      keep_default=False)


        # Changing field 'SearchSubscription.last_run'
        db.alter_column('panda_searchsubscription', 'last_run', self.gf('django.db.models.fields.DateTimeField')(auto_now=True, default=datetime.datetime(2012, 6, 4, 0, 0)))
    def backwards(self, orm):
        # Deleting field 'SearchSubscription.query_url'
        db.delete_column('panda_searchsubscription', 'query_url')


        # Changing field 'SearchSubscription.last_run'
        db.alter_column('panda_searchsubscription', 'last_run', self.gf('django.db.models.fields.DateTimeField')(null=True))
    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_export': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Export']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.searchsubscription': {
            'Meta': {'object_name': 'SearchSubscription'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'subscribed_searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'last_run': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'query_url': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'subscribed_searches'", 'to': "orm['auth.User']"})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_description': ('django.db.models.fields.TextField', [], {}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'activation_key_expiration': ('django.db.models.fields.DateTimeField', [], {}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
########NEW FILE########
__FILENAME__ = 0021_auto__add_field_notification_url
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding field 'Notification.url'
        db.add_column('panda_notification', 'url',
                      self.gf('django.db.models.fields.URLField')(default=None, max_length=200, null=True),
                      keep_default=False)

    def backwards(self, orm):
        # Deleting field 'Notification.url'
        db.delete_column('panda_notification', 'url')

    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_export': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Export']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'}),
            'url': ('django.db.models.fields.URLField', [], {'default': 'None', 'max_length': '200', 'null': 'True'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.searchsubscription': {
            'Meta': {'object_name': 'SearchSubscription'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'subscribed_searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'last_run': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'query_url': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'subscribed_searches'", 'to': "orm['auth.User']"})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_description': ('django.db.models.fields.TextField', [], {}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'activation_key_expiration': ('django.db.models.fields.DateTimeField', [], {}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
########NEW FILE########
__FILENAME__ = 0022_create_notification_urls
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import DataMigration
from django.db import models

class Migration(DataMigration):

    def forwards(self, orm):
        "Write your forwards methods here."
        if not db.dry_run:
            for notification in orm.Notification.objects.all():
                if notification.related_export:
                    notification.url = '#export/%i' % notification.related_export.id
                elif notification.related_dataset:
                    notification.url = '#dataset/%s' % notification.related_dataset.slug

                notification.save()

    def backwards(self, orm):
        "Write your backwards methods here."
        pass

    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'related_dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Dataset']", 'null': 'True'}),
            'related_export': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.Export']", 'null': 'True'}),
            'related_task': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'to': "orm['panda.TaskStatus']", 'null': 'True'}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'}),
            'url': ('django.db.models.fields.URLField', [], {'default': 'None', 'max_length': '200', 'null': 'True'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.searchsubscription': {
            'Meta': {'object_name': 'SearchSubscription'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'subscribed_searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'last_run': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'query_url': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'subscribed_searches'", 'to': "orm['auth.User']"})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_description': ('django.db.models.fields.TextField', [], {}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'activation_key_expiration': ('django.db.models.fields.DateTimeField', [], {}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
    symmetrical = True

########NEW FILE########
__FILENAME__ = 0023_auto__del_field_notification_related_export__del_field_notification_re
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Deleting field 'Notification.related_export'
        db.delete_column('panda_notification', 'related_export_id')

        # Deleting field 'Notification.related_dataset'
        db.delete_column('panda_notification', 'related_dataset_id')

        # Deleting field 'Notification.related_task'
        db.delete_column('panda_notification', 'related_task_id')

    def backwards(self, orm):
        # Adding field 'Notification.related_export'
        db.add_column('panda_notification', 'related_export',
                      self.gf('django.db.models.fields.related.ForeignKey')(default=None, to=orm['panda.Export'], null=True),
                      keep_default=False)

        # Adding field 'Notification.related_dataset'
        db.add_column('panda_notification', 'related_dataset',
                      self.gf('django.db.models.fields.related.ForeignKey')(default=None, to=orm['panda.Dataset'], null=True),
                      keep_default=False)

        # Adding field 'Notification.related_task'
        db.add_column('panda_notification', 'related_task',
                      self.gf('django.db.models.fields.related.ForeignKey')(default=None, to=orm['panda.TaskStatus'], null=True),
                      keep_default=False)

    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'}),
            'url': ('django.db.models.fields.URLField', [], {'default': 'None', 'max_length': '200', 'null': 'True'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.searchsubscription': {
            'Meta': {'object_name': 'SearchSubscription'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'subscribed_searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'last_run': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'query_url': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'subscribed_searches'", 'to': "orm['auth.User']"})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_description': ('django.db.models.fields.TextField', [], {}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'activation_key_expiration': ('django.db.models.fields.DateTimeField', [], {}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
########NEW FILE########
__FILENAME__ = 0024_auto__add_field_searchsubscription_query_human
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding field 'SearchSubscription.query_human'
        db.add_column('panda_searchsubscription', 'query_human',
                      self.gf('django.db.models.fields.TextField')(default=''),
                      keep_default=False)

    def backwards(self, orm):
        # Deleting field 'SearchSubscription.query_human'
        db.delete_column('panda_searchsubscription', 'query_human')

    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'}),
            'url': ('django.db.models.fields.URLField', [], {'default': 'None', 'max_length': '200', 'null': 'True'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.searchsubscription': {
            'Meta': {'object_name': 'SearchSubscription'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'subscribed_searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'last_run': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'query_human': ('django.db.models.fields.TextField', [], {}),
            'query_url': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'subscribed_searches'", 'to': "orm['auth.User']"})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_description': ('django.db.models.fields.TextField', [], {}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'activation_key_expiration': ('django.db.models.fields.DateTimeField', [], {}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
########NEW FILE########
__FILENAME__ = 0025_add_subscription_permissions
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import DataMigration
from django.db import models

class Migration(DataMigration):

    def forwards(self, orm):
        """
        This migration will fail if run against a clean database (fresh setup)
        This is fine because the permission will be installed from the fixture.
        """

        try:
            group = orm['auth.group'].objects.get(name='panda_user')

            perm = orm['auth.permission'].objects.get(codename='add_searchsubscription')
            group.permissions.add(perm)

            perm = orm['auth.permission'].objects.get(codename='delete_searchsubscription')
            group.permissions.add(perm)
        except:
            pass

    def backwards(self, orm):
        group = orm['auth.group'].objects.get(name='panda_user')

        perm = orm['auth.permission'].objects.get(codename='add_searchsubscription')
        group.permissions.remove(perm)

        perm = orm['auth.permission'].objects.get(codename='delete_searchsubscription')
        group.permissions.remove(perm)

    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'}),
            'url': ('django.db.models.fields.URLField', [], {'default': 'None', 'max_length': '200', 'null': 'True'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.searchsubscription': {
            'Meta': {'object_name': 'SearchSubscription'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'subscribed_searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'last_run': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'query_human': ('django.db.models.fields.TextField', [], {}),
            'query_url': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'subscribed_searches'", 'to': "orm['auth.User']"})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_description': ('django.db.models.fields.TextField', [], {}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'activation_key_expiration': ('django.db.models.fields.DateTimeField', [], {}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
    symmetrical = True

########NEW FILE########
__FILENAME__ = 0026_auto__add_field_userprofile_show_login_help
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding field 'UserProfile.show_login_help'
        db.add_column('panda_userprofile', 'show_login_help',
                      self.gf('django.db.models.fields.BooleanField')(default=True),
                      keep_default=False)

    def backwards(self, orm):
        # Deleting field 'UserProfile.show_login_help'
        db.delete_column('panda_userprofile', 'show_login_help')

    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'}),
            'url': ('django.db.models.fields.URLField', [], {'default': 'None', 'max_length': '200', 'null': 'True'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.searchsubscription': {
            'Meta': {'object_name': 'SearchSubscription'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'subscribed_searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'last_run': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'query_human': ('django.db.models.fields.TextField', [], {}),
            'query_url': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'subscribed_searches'", 'to': "orm['auth.User']"})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_description': ('django.db.models.fields.TextField', [], {}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'activation_key_expiration': ('django.db.models.fields.DateTimeField', [], {}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'show_login_help': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
########NEW FILE########
__FILENAME__ = 0027_auto__add_field_searchsubscription_category
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding field 'SearchSubscription.category'
        db.add_column('panda_searchsubscription', 'category',
                      self.gf('django.db.models.fields.related.ForeignKey')(default=None, related_name='subscribes_searches', null=True, to=orm['panda.Category']),
                      keep_default=False)

    def backwards(self, orm):
        # Deleting field 'SearchSubscription.category'
        db.delete_column('panda_searchsubscription', 'category_id')

    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'}),
            'url': ('django.db.models.fields.URLField', [], {'default': 'None', 'max_length': '200', 'null': 'True'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.searchsubscription': {
            'Meta': {'object_name': 'SearchSubscription'},
            'category': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'subscribes_searches'", 'null': 'True', 'to': "orm['panda.Category']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'subscribed_searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'last_run': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'query_human': ('django.db.models.fields.TextField', [], {}),
            'query_url': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'subscribed_searches'", 'to': "orm['auth.User']"})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_description': ('django.db.models.fields.TextField', [], {}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'activation_key_expiration': ('django.db.models.fields.DateTimeField', [], {}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'show_login_help': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
########NEW FILE########
__FILENAME__ = 0028_auto__add_field_relatedupload_title__add_field_export_title__add_field
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding field 'RelatedUpload.title'
        db.add_column('panda_relatedupload', 'title',
                      self.gf('django.db.models.fields.TextField')(default='', max_length=256),
                      keep_default=False)

        # Adding field 'Export.title'
        db.add_column('panda_export', 'title',
                      self.gf('django.db.models.fields.TextField')(default='', max_length=256),
                      keep_default=False)

        # Adding field 'DataUpload.title'
        db.add_column('panda_dataupload', 'title',
                      self.gf('django.db.models.fields.TextField')(default='', max_length=256),
                      keep_default=False)

    def backwards(self, orm):
        # Deleting field 'RelatedUpload.title'
        db.delete_column('panda_relatedupload', 'title')

        # Deleting field 'Export.title'
        db.delete_column('panda_export', 'title')

        # Deleting field 'DataUpload.title'
        db.delete_column('panda_dataupload', 'title')

    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {}),
            'title': ('django.db.models.fields.TextField', [], {'max_length': '256'})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {}),
            'title': ('django.db.models.fields.TextField', [], {'max_length': '256'})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'}),
            'url': ('django.db.models.fields.URLField', [], {'default': 'None', 'max_length': '200', 'null': 'True'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {}),
            'title': ('django.db.models.fields.TextField', [], {'max_length': '256'})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.searchsubscription': {
            'Meta': {'object_name': 'SearchSubscription'},
            'category': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'search_subscriptions'", 'null': 'True', 'to': "orm['panda.Category']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'search_subscriptions'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'last_run': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'query_human': ('django.db.models.fields.TextField', [], {}),
            'query_url': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_subscriptions'", 'to': "orm['auth.User']"})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_description': ('django.db.models.fields.TextField', [], {}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'activation_key_expiration': ('django.db.models.fields.DateTimeField', [], {}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'show_login_help': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
########NEW FILE########
__FILENAME__ = 0029_populate_upload_titles
# -*- coding: utf-8 -*-

from itertools import chain

from south.v2 import DataMigration

class Migration(DataMigration):

    def forwards(self, orm):
        for upload in chain(
                orm['panda.Export'].objects.all(),
                orm['panda.RelatedUpload'].objects.all(),
                orm['panda.DataUpload'].objects.all()):
            if not upload.title:
                upload.title = upload.original_filename
                upload.save()

    def backwards(self, orm):
        "Write your backwards methods here."


    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {}),
            'title': ('django.db.models.fields.TextField', [], {'max_length': '256'})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {}),
            'title': ('django.db.models.fields.TextField', [], {'max_length': '256'})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'}),
            'url': ('django.db.models.fields.URLField', [], {'default': 'None', 'max_length': '200', 'null': 'True'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {}),
            'title': ('django.db.models.fields.TextField', [], {'max_length': '256'})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.searchsubscription': {
            'Meta': {'object_name': 'SearchSubscription'},
            'category': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'search_subscriptions'", 'null': 'True', 'to': "orm['panda.Category']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'search_subscriptions'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'last_run': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'query_human': ('django.db.models.fields.TextField', [], {}),
            'query_url': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_subscriptions'", 'to': "orm['auth.User']"})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_description': ('django.db.models.fields.TextField', [], {}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'activation_key_expiration': ('django.db.models.fields.DateTimeField', [], {}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'show_login_help': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
    symmetrical = True

########NEW FILE########
__FILENAME__ = 0030_auto__add_field_dataset_related_stories
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding field 'Dataset.related_stories'
        db.add_column('panda_dataset', 'related_stories',
                      self.gf('panda.fields.JSONField')(default=[]),
                      keep_default=False)

    def backwards(self, orm):
        # Deleting field 'Dataset.related_stories'
        db.delete_column('panda_dataset', 'related_stories')

    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '30'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'related_stories': ('panda.fields.JSONField', [], {'default': '[]'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {}),
            'title': ('django.db.models.fields.TextField', [], {'max_length': '256'})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {}),
            'title': ('django.db.models.fields.TextField', [], {'max_length': '256'})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'}),
            'url': ('django.db.models.fields.URLField', [], {'default': 'None', 'max_length': '200', 'null': 'True'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {}),
            'title': ('django.db.models.fields.TextField', [], {'max_length': '256'})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.searchsubscription': {
            'Meta': {'object_name': 'SearchSubscription'},
            'category': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'search_subscriptions'", 'null': 'True', 'to': "orm['panda.Category']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'search_subscriptions'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'last_run': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'query_human': ('django.db.models.fields.TextField', [], {}),
            'query_url': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_subscriptions'", 'to': "orm['auth.User']"})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_description': ('django.db.models.fields.TextField', [], {}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'activation_key_expiration': ('django.db.models.fields.DateTimeField', [], {}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'show_login_help': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
########NEW FILE########
__FILENAME__ = 0031_rename_dataset_related_stories
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        db.rename_column('panda_dataset', 'related_stories', 'related_links')

    def backwards(self, orm):
        db.rename_column('panda_dataset', 'related_links', 'related_stories')

    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '255'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'related_links': ('panda.fields.JSONField', [], {'default': '[]'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {}),
            'title': ('django.db.models.fields.TextField', [], {'max_length': '256'})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {}),
            'title': ('django.db.models.fields.TextField', [], {'max_length': '256'})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'}),
            'url': ('django.db.models.fields.URLField', [], {'default': 'None', 'max_length': '200', 'null': 'True'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {}),
            'title': ('django.db.models.fields.TextField', [], {'max_length': '256'})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.searchsubscription': {
            'Meta': {'object_name': 'SearchSubscription'},
            'category': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'search_subscriptions'", 'null': 'True', 'to': "orm['panda.Category']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'search_subscriptions'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'last_run': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'query_human': ('django.db.models.fields.TextField', [], {}),
            'query_url': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_subscriptions'", 'to': "orm['auth.User']"})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_description': ('django.db.models.fields.TextField', [], {}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'activation_key_expiration': ('django.db.models.fields.DateTimeField', [], {}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'show_login_help': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']

########NEW FILE########
__FILENAME__ = 0032_auto__add_field_dataupload_deletable
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):
        # Adding field 'DataUpload.deletable'
        db.add_column('panda_dataupload', 'deletable',
                      self.gf('django.db.models.fields.BooleanField')(default=False),
                      keep_default=False)

    def backwards(self, orm):
        # Deleting field 'DataUpload.deletable'
        db.delete_column('panda_dataupload', 'deletable')

    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '255'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'related_links': ('panda.fields.JSONField', [], {'default': '[]'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'deletable': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {}),
            'title': ('django.db.models.fields.TextField', [], {'max_length': '256'})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {}),
            'title': ('django.db.models.fields.TextField', [], {'max_length': '256'})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'}),
            'url': ('django.db.models.fields.URLField', [], {'default': 'None', 'max_length': '200', 'null': 'True'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {}),
            'title': ('django.db.models.fields.TextField', [], {'max_length': '256'})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.searchsubscription': {
            'Meta': {'object_name': 'SearchSubscription'},
            'category': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'search_subscriptions'", 'null': 'True', 'to': "orm['panda.Category']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'search_subscriptions'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'last_run': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'query_human': ('django.db.models.fields.TextField', [], {}),
            'query_url': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_subscriptions'", 'to': "orm['auth.User']"})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_description': ('django.db.models.fields.TextField', [], {}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'activation_key_expiration': ('django.db.models.fields.DateTimeField', [], {}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'show_login_help': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']

########NEW FILE########
__FILENAME__ = 0033_auto__chg_field_searchlog_query
# -*- coding: utf-8 -*-
import datetime
from south.db import db
from south.v2 import SchemaMigration
from django.db import models


class Migration(SchemaMigration):

    def forwards(self, orm):

        # Changing field 'SearchLog.query'
        db.alter_column('panda_searchlog', 'query', self.gf('django.db.models.fields.CharField')(max_length=4096))
    def backwards(self, orm):

        # Changing field 'SearchLog.query'
        db.alter_column('panda_searchlog', 'query', self.gf('django.db.models.fields.CharField')(max_length=256))
    models = {
        'auth.group': {
            'Meta': {'object_name': 'Group'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '80'}),
            'permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'})
        },
        'auth.permission': {
            'Meta': {'ordering': "('content_type__app_label', 'content_type__model', 'codename')", 'unique_together': "(('content_type', 'codename'),)", 'object_name': 'Permission'},
            'codename': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'content_type': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['contenttypes.ContentType']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '50'})
        },
        'auth.user': {
            'Meta': {'object_name': 'User'},
            'date_joined': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'email': ('django.db.models.fields.EmailField', [], {'max_length': '75', 'blank': 'True'}),
            'first_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'groups': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Group']", 'symmetrical': 'False', 'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'is_active': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'is_staff': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'is_superuser': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'last_login': ('django.db.models.fields.DateTimeField', [], {'default': 'datetime.datetime.now'}),
            'last_name': ('django.db.models.fields.CharField', [], {'max_length': '30', 'blank': 'True'}),
            'password': ('django.db.models.fields.CharField', [], {'max_length': '128'}),
            'user_permissions': ('django.db.models.fields.related.ManyToManyField', [], {'to': "orm['auth.Permission']", 'symmetrical': 'False', 'blank': 'True'}),
            'username': ('django.db.models.fields.CharField', [], {'unique': 'True', 'max_length': '255'})
        },
        'contenttypes.contenttype': {
            'Meta': {'ordering': "('name',)", 'unique_together': "(('app_label', 'model'),)", 'object_name': 'ContentType', 'db_table': "'django_content_type'"},
            'app_label': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'model': ('django.db.models.fields.CharField', [], {'max_length': '100'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '100'})
        },
        'panda.activitylog': {
            'Meta': {'unique_together': "(('user', 'when'),)", 'object_name': 'ActivityLog'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'activity_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.category': {
            'Meta': {'object_name': 'Category'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '64'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataset': {
            'Meta': {'ordering': "['-creation_date']", 'object_name': 'Dataset'},
            'categories': ('django.db.models.fields.related.ManyToManyField', [], {'blank': 'True', 'related_name': "'datasets'", 'null': 'True', 'symmetrical': 'False', 'to': "orm['panda.Category']"}),
            'column_schema': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'datasets'", 'to': "orm['auth.User']"}),
            'current_task': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['panda.TaskStatus']", 'null': 'True', 'blank': 'True'}),
            'description': ('django.db.models.fields.TextField', [], {'blank': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'initial_upload': ('django.db.models.fields.related.ForeignKey', [], {'blank': 'True', 'related_name': "'initial_upload_for'", 'null': 'True', 'to': "orm['panda.DataUpload']"}),
            'last_modification': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'last_modified_by': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']", 'null': 'True', 'blank': 'True'}),
            'locked': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'locked_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True'}),
            'name': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'related_links': ('panda.fields.JSONField', [], {'default': '[]'}),
            'row_count': ('django.db.models.fields.IntegerField', [], {'null': 'True', 'blank': 'True'}),
            'sample_data': ('panda.fields.JSONField', [], {'default': 'None', 'null': 'True'}),
            'slug': ('django.db.models.fields.SlugField', [], {'max_length': '256'})
        },
        'panda.dataupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'DataUpload'},
            'columns': ('panda.fields.JSONField', [], {'null': 'True'}),
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'data_type': ('django.db.models.fields.CharField', [], {'max_length': '4', 'null': 'True', 'blank': 'True'}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'data_uploads'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'deletable': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'dialect': ('panda.fields.JSONField', [], {'null': 'True'}),
            'encoding': ('django.db.models.fields.CharField', [], {'default': "'utf-8'", 'max_length': '32'}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'guessed_types': ('panda.fields.JSONField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'imported': ('django.db.models.fields.BooleanField', [], {'default': 'False'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'sample_data': ('panda.fields.JSONField', [], {'null': 'True'}),
            'size': ('django.db.models.fields.IntegerField', [], {}),
            'title': ('django.db.models.fields.TextField', [], {'max_length': '256'})
        },
        'panda.export': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'Export'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'exports'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {}),
            'title': ('django.db.models.fields.TextField', [], {'max_length': '256'})
        },
        'panda.notification': {
            'Meta': {'ordering': "['-sent_at']", 'object_name': 'Notification'},
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.TextField', [], {}),
            'read_at': ('django.db.models.fields.DateTimeField', [], {'default': 'None', 'null': 'True', 'blank': 'True'}),
            'recipient': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'notifications'", 'to': "orm['auth.User']"}),
            'sent_at': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'type': ('django.db.models.fields.CharField', [], {'default': "'Info'", 'max_length': '16'}),
            'url': ('django.db.models.fields.URLField', [], {'default': 'None', 'max_length': '200', 'null': 'True'})
        },
        'panda.relatedupload': {
            'Meta': {'ordering': "['creation_date']", 'object_name': 'RelatedUpload'},
            'creation_date': ('django.db.models.fields.DateTimeField', [], {}),
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'to': "orm['auth.User']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'related_uploads'", 'to': "orm['panda.Dataset']"}),
            'filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'original_filename': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'size': ('django.db.models.fields.IntegerField', [], {}),
            'title': ('django.db.models.fields.TextField', [], {'max_length': '256'})
        },
        'panda.searchlog': {
            'Meta': {'object_name': 'SearchLog'},
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'searches'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '4096'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_logs'", 'to': "orm['auth.User']"}),
            'when': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'})
        },
        'panda.searchsubscription': {
            'Meta': {'object_name': 'SearchSubscription'},
            'category': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'search_subscriptions'", 'null': 'True', 'to': "orm['panda.Category']"}),
            'dataset': ('django.db.models.fields.related.ForeignKey', [], {'default': 'None', 'related_name': "'search_subscriptions'", 'null': 'True', 'to': "orm['panda.Dataset']"}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'last_run': ('django.db.models.fields.DateTimeField', [], {'auto_now': 'True', 'blank': 'True'}),
            'query': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'query_human': ('django.db.models.fields.TextField', [], {}),
            'query_url': ('django.db.models.fields.CharField', [], {'max_length': '256'}),
            'user': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'search_subscriptions'", 'to': "orm['auth.User']"})
        },
        'panda.taskstatus': {
            'Meta': {'object_name': 'TaskStatus'},
            'creator': ('django.db.models.fields.related.ForeignKey', [], {'related_name': "'tasks'", 'null': 'True', 'to': "orm['auth.User']"}),
            'end': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'message': ('django.db.models.fields.CharField', [], {'max_length': '255', 'blank': 'True'}),
            'start': ('django.db.models.fields.DateTimeField', [], {'null': 'True'}),
            'status': ('django.db.models.fields.CharField', [], {'default': "'PENDING'", 'max_length': '50'}),
            'task_description': ('django.db.models.fields.TextField', [], {}),
            'task_name': ('django.db.models.fields.CharField', [], {'max_length': '255'}),
            'traceback': ('django.db.models.fields.TextField', [], {'default': 'None', 'null': 'True', 'blank': 'True'})
        },
        'panda.userprofile': {
            'Meta': {'object_name': 'UserProfile'},
            'activation_key': ('django.db.models.fields.CharField', [], {'max_length': '40', 'null': 'True', 'blank': 'True'}),
            'activation_key_expiration': ('django.db.models.fields.DateTimeField', [], {}),
            'id': ('django.db.models.fields.AutoField', [], {'primary_key': 'True'}),
            'show_login_help': ('django.db.models.fields.BooleanField', [], {'default': 'True'}),
            'user': ('django.db.models.fields.related.OneToOneField', [], {'to': "orm['auth.User']", 'unique': 'True'})
        }
    }

    complete_apps = ['panda']
########NEW FILE########
__FILENAME__ = activity_log
#!/usr/bin/env python

from django.db import models
from django.utils.translation import ugettext_lazy as _

from panda.models.user_proxy import UserProxy

class ActivityLog(models.Model):
    """
    A daily log of activity by a users.
    """
    user = models.ForeignKey(UserProxy, related_name='activity_logs',
        help_text=_('The user who was active.'),
        verbose_name=_('user'))
    when = models.DateField(_('when'), auto_now=True,
        help_text=_('The date this activity was recorded.'))

    class Meta:
        app_label = 'panda'
        verbose_name = _('ActivityLog')
        verbose_name_plural = _('ActivityLogs')
        unique_together = ('user', 'when')

    def __unicode__(self):
        return _('%(user)s at %(when)s') % { 'user': self.user, 'when': self.when }


########NEW FILE########
__FILENAME__ = base_upload
#!/usr/bin/env python

import os.path

from django.db import models
from django.utils.timezone import now 
from django.utils.translation import ugettext_lazy as _

from panda.models.user_proxy import UserProxy

class BaseUpload(models.Model):
    """
    Base class for any file uploaded to PANDA.
    """
    filename = models.CharField(_('filename'), 
        max_length=256,
        help_text=_('Filename as stored in PANDA.'))
    original_filename = models.CharField(_('original_filename'), 
        max_length=256,
        help_text=_('Filename as originally uploaded.'))
    size = models.IntegerField(_('size'),
        help_text=_('Size of the file in bytes.'))
    creator = models.ForeignKey(UserProxy,
        help_text=_('The user who uploaded this file.'),
        verbose_name=_('creator'))
    creation_date = models.DateTimeField(_('creation_date'),
        help_text=_('The date this file was uploaded.'))
    title = models.TextField(_('title'),
        max_length=256,
        help_text=_('A user-friendly name for this file.'))

    class Meta:
        app_label = 'panda'
        abstract = True

    def __unicode__(self):
        return self.filename

    def save(self, *args, **kwargs):
        if not self.creation_date:
            self.creation_date = now()

        if not self.title:
            self.title = self.original_filename

        super(BaseUpload, self).save(*args, **kwargs)

    def delete(self, *args, **kwargs):
        """
        When deleting an upload, it will attempt to clean
        up its own associated files.
        """
        try:
            os.remove(self.get_path())
        except:
            pass

        super(BaseUpload, self).delete(*args, **kwargs)

    def get_path(self):
        """
        Get the absolute path to this upload on disk.
        """
        return os.path.join(self.file_root, self.filename)


########NEW FILE########
__FILENAME__ = category
#!/usr/bin/env python

from django.db import models
from django.utils.translation import ugettext_lazy as _

from panda.models.slugged_model import SluggedModel

class Category(SluggedModel):
    """
    A category that contains Datasets.
    """
    name = models.CharField(_('name'),
        max_length=64,
        help_text=_('Category name.'))

    class Meta:
        app_label = 'panda'
        verbose_name = _('Category')
        verbose_name_plural = _('Categories')

    def __unicode__(self):
        return self.name


########NEW FILE########
__FILENAME__ = dataset
#!/usr/bin/env python

from urllib import unquote

from django.conf import settings
from django.db import models
from django.utils.timezone import now 
from django.utils.translation import ugettext_lazy as _

from panda import solr, utils
from panda.exceptions import DataImportError, DatasetLockedError
from panda.fields import JSONField
from panda.models.category import Category
from panda.models.slugged_model import SluggedModel
from panda.models.task_status import TaskStatus
from panda.models.user_proxy import UserProxy
from panda.tasks import get_import_task_type_for_upload, ExportCSVTask, PurgeDataTask, ReindexTask 
from panda.utils.column_schema import make_column_schema, update_indexed_names
from panda.utils.typecoercion import DataTyper

class Dataset(SluggedModel):
    """
    A PANDA dataset (one table & associated metadata).
    """
    name = models.CharField(_('name'), max_length=256,
        help_text=_('User-supplied dataset name.'))
    description = models.TextField(_('description'), blank=True,
        help_text=_('User-supplied dataset description.'))
    # related_uploads =  models.ToMany(RelatedUpload, null=True)
    # data_uploads =  models.ToMany(DataUpload, null=True)
    initial_upload = models.ForeignKey('DataUpload', null=True, blank=True, related_name='initial_upload_for',
        help_text=_('The data upload used to create this dataset, if any was used.'),
        verbose_name=_('initial_upload'))
    column_schema = JSONField(_('column_schema'), null=True, default=None,
        help_text=_('Metadata about columns.'))
    sample_data = JSONField(_('sample_data'), null=True, default=None,
        help_text=_('Example data rows from the dataset.'))
    row_count = models.IntegerField(_('row_count'), null=True, blank=True,
        help_text=_('The number of rows in this dataset. Null if no data has been added/imported.'))
    current_task = models.ForeignKey(TaskStatus, blank=True, null=True,
        help_text=_('The currently executed or last finished task related to this dataset.'),
        verbose_name=_('current_task')) 
    creation_date = models.DateTimeField(_('creation_date'), null=True,
        help_text=_('The date this dataset was initially created.'))
    creator = models.ForeignKey(UserProxy, related_name='datasets',
        help_text=_('The user who created this dataset.'),
        verbose_name=_('creator'))
    categories = models.ManyToManyField(Category, related_name='datasets', blank=True, null=True,
        help_text=_('Categories containing this Dataset.'),
        verbose_name=_('categories'))
    last_modified = models.DateTimeField(_('last_modified'), null=True, blank=True, default=None,
        help_text=_('When, if ever, was this dataset last modified via the API?'))
    last_modification = models.TextField(_('last_modification'), null=True, blank=True, default=None,
        help_text=_('Description of the last modification made to this Dataset.'))
    last_modified_by = models.ForeignKey(UserProxy, null=True, blank=True,
        help_text=_('The user, if any, who last modified this dataset.'),
        verbose_name=_('last_modified_by'))
    locked = models.BooleanField(_('locked'), default=False,
        help_text=_('Is this table locked for writing?'))
    locked_at = models.DateTimeField(_('locked_at'), null=True, default=None,
        help_text=_('Time this dataset was last locked.'))
    related_links = JSONField(default=[])

    class Meta:
        app_label = 'panda'
        ordering = ['-creation_date']
        verbose_name = _('Dataset')
        verbose_name_plural = _('Datasets')

    def __unicode__(self):
        return self.name

    def save(self, *args, **kwargs):
        """
        Save the date of creation.
        """
        if not self.creation_date:
            self.creation_date = now()

        super(Dataset, self).save(*args, **kwargs)

    def lock(self):
        """
        Obtain an editing lock on this dataset.
        """
        # Ensure latest state has come over from the database
        before_lock = self.__class__.objects.get(pk=self.pk)
        self.locked = before_lock.locked
        self.locked_at = before_lock.locked_at

        if self.locked:
            # Already locked
            raise DatasetLockedError(_('This dataset is currently locked by another process.'))

        new_locked_at = now()

        self.locked = True
        self.locked_at = new_locked_at

        self.save()

        # Refresh from database
        after_lock = Dataset.objects.get(id=self.id)
        self.locked = after_lock.locked
        self.locked_at = after_lock.locked_at

        if self.locked_at != new_locked_at:
            # Somebody else got the lock
            raise DatasetLockedError(_('This dataset is currently locked by another process.'))

    def unlock(self):
        """
        Unlock this dataset so it can be edited.
        """
        self.locked = False
        self.lock_id = None

        self.save()

    def update_full_text(self, commit=True):
        """
        Update the full-text search metadata for this dataset stored in Solr.
        """
        category_ids = []

        full_text_data = [
            unquote(self.name),
            unquote(self.description),
            '%s %s' % (self.creator.first_name, self.creator.last_name),
            self.creator.email
        ]

        for category in self.categories.all():
            category_ids.append(category.id)
            full_text_data.append(category.name)

        if not category_ids:
            category_ids.append(settings.PANDA_UNCATEGORIZED_ID)
            full_text_data.append(settings.PANDA_UNCATEGORIZED_NAME)

        for data_upload in self.data_uploads.all():
            full_text_data.append(data_upload.original_filename)

        for related_upload in self.related_uploads.all():
            full_text_data.append(related_upload.original_filename)

        if self.column_schema is not None:
            full_text_data.extend([c['name'] for c in self.column_schema])

        full_text = u'\n'.join(map(unicode, full_text_data)) # convert any i18n proxies into strings

        solr.add(settings.SOLR_DATASETS_CORE, [{
            'slug': self.slug,
            'creation_date': self.creation_date.isoformat() + 'Z',
            'categories': category_ids,
            'full_text': full_text
        }], commit=commit)

    def delete(self, *args, **kwargs):
        """
        Cancel any in progress task.
        """
        # Cancel import if necessary 
        if self.current_task:
            self.current_task.request_abort()

        # Manually delete related uploads so their delete method is called
        for upload in self.data_uploads.all():
            upload.delete(skip_purge=True, force=True)

        for upload in self.related_uploads.all():
            upload.delete()

        # Cleanup data in Solr
        PurgeDataTask.apply_async(args=[self.slug])
        solr.delete(settings.SOLR_DATASETS_CORE, 'slug:%s' % self.slug)

        super(Dataset, self).delete(*args, **kwargs)

    def import_data(self, user, upload, external_id_field_index=None):
        """
        Import data into this ``Dataset`` from a given ``DataUpload``.
        """
        self.lock()

        try:
            if upload.imported:
                raise DataImportError(_('This file has already been imported.'))

            task_type = get_import_task_type_for_upload(upload)

            if not task_type:
                # This is normally caught on the client.
                raise DataImportError(_('This file type is not supported for data import.'))
            
            if self.column_schema:
                # This is normally caught on the client.
                if upload.columns != [c['name'] for c in self.column_schema]:
                    raise DataImportError(_('The columns in this file do not match those in the dataset.'))
            else:
                self.column_schema = make_column_schema(upload.columns, types=upload.guessed_types)
                
            if self.sample_data is None:
                self.sample_data = upload.sample_data

            # If this is the first import and the API hasn't been used, save that information
            if self.initial_upload is None and self.row_count is None:
                self.initial_upload = upload

            self.current_task = TaskStatus.objects.create(
                task_name=task_type.name,
                task_description=_('Import data from %(filename)s into %(slug)s.') \
                    % {'filename': upload.filename, 'slug': self.slug},
                creator=user
            )
            self.save()

            task_type.apply_async(
                args=[self.slug, upload.id],
                kwargs={ 'external_id_field_index': external_id_field_index },
                task_id=self.current_task.id
            )
        except:
            self.unlock()
            raise

    def reindex_data(self, user, typed_columns=None, column_types=None):
        """
        Reindex the data currently stored for this ``Dataset``.
        """
        self.lock()
        
        task_type = ReindexTask

        try:
            typed_column_count = 0

            if typed_columns:
                for i, t in enumerate(typed_columns):
                    self.column_schema[i]['indexed'] = t
                    
                    if t:
                        typed_column_count += 1

            if column_types:
                for i, t in enumerate(column_types):
                    self.column_schema[i]['type'] = t

            self.column_schema = update_indexed_names(self.column_schema)

            self.current_task = TaskStatus.objects.create(
                task_name=task_type.name,
                task_description=_('Reindex %(slug)s with %(typed_column_count)i column filters.') \
                     % {'slug': self.slug, 'typed_column_count': typed_column_count}, 
                creator=user
            )

            self.save()

            task_type.apply_async(
                args=[self.slug],
                kwargs={},
                task_id=self.current_task.id
            )
        except:
            self.unlock()
            raise

    def export_data(self, user, query=None, filename=None):
        """
        Execute the data export task for this ``Dataset``.
        """
        task_type = ExportCSVTask

        if query:
            description = _('Export search results for "%(query)s" in %(slug)s.') \
                % {'query': query, 'slug': self.slug}
        else:
            description = _('Exporting data in %s.') % self.slug

        self.current_task = TaskStatus.objects.create(
            task_name=task_type.name,
            task_description=description,
            creator=user
        )

        self.save()

        task_type.apply_async(
            args=[self.slug],
            kwargs={ 'query': query, 'filename': filename },
            task_id=self.current_task.id
        )

    def get_row(self, external_id):
        """
        Fetch a row from this dataset.
        """
        response = solr.query(settings.SOLR_DATA_CORE, 'dataset_slug:%s AND external_id:%s' % (self.slug, external_id), limit=1)

        if len(response['response']['docs']) < 1:
            return None

        return response['response']['docs'][0]

    def add_row(self, user, data, external_id=None):
        """
        Add (or overwrite) a row to this dataset.
        """
        self.lock()

        try:
            data_typer = DataTyper(self.column_schema)

            solr_row = utils.solr.make_data_row(self, data, external_id=external_id)
            solr_row = data_typer(solr_row, data)

            solr.add(settings.SOLR_DATA_CORE, [solr_row], commit=True)

            self.schema = data_typer.schema

            if not self.sample_data:
                self.sample_data = []
            
            if len(self.sample_data) < 5:
                self.sample_data.append(data)

            old_row_count = self.row_count
            self.row_count = self._count_rows()
            added = self.row_count - (old_row_count or 0)
            self.last_modified = now()
            self.last_modified_by = user
            self.last_modification = _('1 row %s') % ('added' if added else 'updated')
            self.save()

            return solr_row
        finally:
            self.unlock()

    def add_many_rows(self, user, data):
        """
        Shortcut for adding rows in bulk. 

        ``data`` must be an array of tuples in the format (data_array, external_id)
        """
        self.lock()

        try:
            data_typer = DataTyper(self.column_schema)

            solr_rows = [utils.solr.make_data_row(self, d[0], external_id=d[1]) for d in data]
            solr_rows = [data_typer(s, d[0]) for s, d in zip(solr_rows, data)]

            solr.add(settings.SOLR_DATA_CORE, solr_rows, commit=True)
            
            self.schema = data_typer.schema

            if not self.sample_data:
                self.sample_data = []
            
            if len(self.sample_data) < 5:
                needed = 5 - len(self.sample_data)
                self.sample_data.extend([d[0] for d in data[:needed]])

            old_row_count = self.row_count
            self.row_count = self._count_rows()
            added = self.row_count - (old_row_count or 0)
            updated = len(data) - added
            self.last_modified = now()
            self.last_modified_by = user

            if added and updated: 
                self.last_modification = _('%(added)i rows added and %(updated)i updated') \
                    % {'added': added, 'updated': updated}
            elif added:
                self.last_modification = _('%i rows added') % added
            else:
                self.last_modification = _('%i rows updated') % updated

            self.save()

            return solr_rows
        finally:
            self.unlock()
        
    def delete_row(self, user, external_id):
        """
        Delete a row in this dataset.
        """
        self.lock()

        try:
            solr.delete(settings.SOLR_DATA_CORE, 'dataset_slug:%s AND external_id:%s' % (self.slug, external_id), commit=True)
        
            self.row_count = self._count_rows()
            self.last_modified = now()
            self.last_modified_by = user
            self.last_modification = _('1 row deleted')
            self.save()
        finally:
            self.unlock()

    def delete_all_rows(self, user,):
        """
        Delete all rows in this dataset.
        """
        self.lock()

        try:
            solr.delete(settings.SOLR_DATA_CORE, 'dataset_slug:%s' % self.slug, commit=True)

            old_row_count = self.row_count
            self.row_count = 0
            self.last_modified = now()
            self.last_modification = _('All %i rows deleted') % old_row_count or 0
            self.save()
        finally:
            self.unlock()

    def _count_rows(self):
        """
        Count the number of rows currently stored in Solr for this Dataset.
        Useful for sanity checks.
        """
        return solr.query(settings.SOLR_DATA_CORE, 'dataset_slug:%s' % self.slug)['response']['numFound']


########NEW FILE########
__FILENAME__ = data_upload
#!/usr/bin/env python

import os.path

from django.conf import settings
from django.db import models
from django.utils.translation import ugettext_lazy as _

from panda import utils
from panda.exceptions import DataUploadNotDeletable
from panda.fields import JSONField
from panda.models.base_upload import BaseUpload
from panda.tasks import PurgeDataTask

class DataUpload(BaseUpload):
    """
    A data file uploaded to PANDA (either a table or metadata file).
    """
    from panda.models.dataset import Dataset

    dataset = models.ForeignKey(Dataset, related_name='data_uploads', null=True,
        help_text=_('The dataset this upload is associated with.'),
        verbose_name=_('dataset'))

    data_type = models.CharField(_('data_type'), max_length=4, null=True, blank=True,
        help_text=_('The type of this file, if known.'))
    encoding = models.CharField(_('encoding'), max_length=32, default='utf-8',
        help_text=_('The character encoding of this file. Defaults to utf-8'))
    dialect = JSONField(_('dialect'), null=True,
        help_text=_('Description of the formatting of this file.'))
    columns = JSONField(_('columns'), null=True,
        help_text=_('A list of names for the columns in this upload.'))
    sample_data = JSONField(_('sample_data'), null=True,
        help_text=_('Example data from this file.'))
    guessed_types = JSONField(_('guessed_types'), null=True,
        help_text=_('Column types guessed based on a sample of data.'))
    imported = models.BooleanField(_('imported'), default=False,
        help_text=_('Has this upload ever been imported into its parent dataset.'))
    deletable = models.BooleanField(_('deletable'), default=True,
        help_text=_('Can this data upload be deleted? False for uploads prior to 1.0.'))
    
    file_root = settings.MEDIA_ROOT

    class Meta:
        app_label = 'panda'
        ordering = ['creation_date']
        verbose_name = _('DataUpload')
        verbose_name_plural = _('DataUploads')

    def __unicode__(self):
        return self.filename

    def save(self, *args, **kwargs):
        if self.data_type is None:
            self.data_type = self._infer_data_type()

        if self.data_type:
            path = self.get_path()

            if self.dialect is None:
                self.dialect = utils.sniff_dialect(self.data_type, path, encoding=self.encoding)

            if self.columns is None:
                self.columns = utils.extract_column_names(self.data_type, path, self.dialect_as_parameters(), encoding=self.encoding)

            if self.sample_data is None:
                self.sample_data = utils.sample_data(self.data_type, path, self.dialect_as_parameters(), encoding=self.encoding)

            if self.guessed_types is None:
                self.guessed_types = utils.guess_column_types(self.data_type, path, self.dialect_as_parameters(), encoding=self.encoding)

        super(DataUpload, self).save(*args, **kwargs)

    def _infer_data_type(self):
        """
        Get the data type of this file. Returns an empty string
        if the file is not a recognized type.
        """
        extension = os.path.splitext(self.filename)[1]

        if extension == '.csv':
            return 'csv' 
        elif extension == '.xls':
            return 'xls'
        elif extension == '.xlsx':
            return 'xlsx'

        return ''

    def dialect_as_parameters(self):
        """
        Dialect parameters are stored as a JSON document, which causes
        certain characters to be escaped. This method reverses this so
        they can be used as arguments.
        """
        dialect_params = {}

        # This code is absolutely terrifying
        # (Also, it works.)
        for k, v in self.dialect.items():
            if isinstance(v, basestring):
                dialect_params[k] = v.decode('string_escape')
            else:
                dialect_params[k] = v

        return dialect_params

    def delete(self, *args, **kwargs):
        """
        Cancel any in progress task.
        """
        skip_purge = kwargs.pop('skip_purge', False)
        force = kwargs.pop('force', False)

        # Don't allow deletion of dated uploads unless forced
        if not self.deletable and not force:
            raise DataUploadNotDeletable(_('This data upload was created before deleting individual data uploads was supported. In order to delete it you must delete the entire dataset.'))

        # Update related datasets so deletes will not cascade
        if self.initial_upload_for.count():
            for dataset in self.initial_upload_for.all():
                dataset.initial_upload = None
                dataset.save()

        # Cleanup data in Solr
        if self.dataset and self.imported and not skip_purge:
            PurgeDataTask.apply_async(args=[self.dataset.slug, self.id])

        super(DataUpload, self).delete(*args, **kwargs)


########NEW FILE########
__FILENAME__ = export
#!/usr/bin/env python

from django.conf import settings
from django.db import models
from django.utils.translation import ugettext_lazy as _

from panda.models.base_upload import BaseUpload

class Export(BaseUpload):
    """
    A dataset exported to a file.
    """
    from panda.models.dataset import Dataset

    dataset = models.ForeignKey(Dataset, related_name='exports', null=True,
        help_text=_('The dataset this export is from.'),
        verbose_name=_('dataset'))

    file_root = settings.EXPORT_ROOT

    class Meta:
        app_label = 'panda'
        ordering = ['creation_date']
        verbose_name = _('Export')
        verbose_name_plural = _('Exports')


########NEW FILE########
__FILENAME__ = notification
#!/user/bin/env python

from django.db import models
from django.utils.translation import ugettext_lazy as _

from panda.models.user_proxy import UserProxy

NOTIFICATION_TYPE_CHOICES = (
    ('Info', _('Info')),
    ('Warning', _('Warning')),
    ('Error', _('Error'))
)

class Notification(models.Model):
    """
    A user notification related to a task.
    """
    recipient = models.ForeignKey(UserProxy, related_name='notifications',
        help_text=_('The user who should receive this notification.'),
        verbose_name=_('recipient'))
    message = models.TextField(_('message'),
        help_text=_('The message to deliver.'))
    type = models.CharField(_('type'), max_length=16, choices=NOTIFICATION_TYPE_CHOICES, default='Info',
        help_text=_('The type of message: info, warning or error'))
    sent_at = models.DateTimeField(_('sent_at'), auto_now=True,
        help_text=_('When this notification was created'))
    read_at = models.DateTimeField(_('read_at'), null=True, blank=True, default=None,
        help_text=_('When this notification was read by the user.'))
    url = models.URLField(_('url'), null=True, default=None,
        help_text=_('A url to link to when displaying this notification.')) 

    class Meta:
        app_label = 'panda'
        ordering = ['-sent_at'] 
        verbose_name = _('Notification')
        verbose_name_plural = _('Notifications')

########NEW FILE########
__FILENAME__ = related_upload
#!/usr/bin/env python

from django.conf import settings
from django.db import models
from django.utils.translation import ugettext_lazy as _

from panda.models.base_upload import BaseUpload

class RelatedUpload(BaseUpload):
    """
    A file related to a dataset file uploaded to PANDA.
    """
    from panda.models.dataset import Dataset

    dataset = models.ForeignKey(Dataset, related_name='related_uploads',
        help_text=_('The dataset this upload is associated with.'),
        verbose_name=_('dataset'))

    file_root = settings.MEDIA_ROOT

    class Meta:
        app_label = 'panda'
        ordering = ['creation_date']
        verbose_name = _('RelatedUpload')
        verbose_name_plural = _('RelatedUploads')

########NEW FILE########
__FILENAME__ = search_log
#!/usr/bin/env python

from django.db import models
from django.utils.translation import ugettext_lazy as _

from panda.models.dataset import Dataset
from panda.models.user_proxy import UserProxy

class SearchLog(models.Model):
    """
    A log of a user search.
    """
    user = models.ForeignKey(UserProxy, related_name='search_logs',
        help_text=_('The user who executed the search.'),
        verbose_name=_('user'))
    dataset = models.ForeignKey(Dataset, related_name='searches', null=True, default=None,
        help_text=_('The data set searched, or null if all were searched.'),
        verbose_name=_('dataset'))
    query = models.CharField(_('query'), max_length=4096, 
        help_text=_('The search query that was executed'))
    when = models.DateTimeField(_('when'), auto_now=True,
        help_text=_('The date and time this search was logged.'))

    class Meta:
        app_label = 'panda'
        verbose_name = _('SearchLog')
        verbose_name_plural = _('SearchLogs')

    def __unicode__(self):
        if self.dataset:
            return _('%(user)s searched %(dataset)s for %(query)s') \
                % {'user': self.user, 'dataset': self.dataset, 'query': self.query}
        else:
            return _('%(user)s searched for %(query)s') \
                % {'user': self.user, 'query': self.query}


########NEW FILE########
__FILENAME__ = search_subscription
#!/usr/bin/env python

from django.db import models
from django.utils.translation import ugettext_lazy as _

from panda.models.category import Category
from panda.models.dataset import Dataset
from panda.models.user_proxy import UserProxy

class SearchSubscription(models.Model):
    """
    A log of a user search.
    """
    user = models.ForeignKey(UserProxy, related_name='search_subscriptions',
        help_text=_('The user who subscribed to the search.'),
        verbose_name=_('user'))
    dataset = models.ForeignKey(Dataset, related_name='search_subscriptions', null=True, default=None,
        help_text=_('The dataset to be searched or null if all are to be searched.'),
        verbose_name=_('dataset'))
    category = models.ForeignKey(Category, related_name='search_subscriptions', null=True, default=None,
        help_text=_('A category to be searched or null if all are to be searched.'),
        verbose_name=_('category'))
    query = models.CharField(_('query'), max_length=256, 
        help_text=_('The search query to executed.'))
    query_url = models.CharField(_('query_url'), max_length=256,
        help_text=_('Query encoded for URL.'))
    query_human = models.TextField(_('query_human'),
        help_text=_('Human-readable description of the query being run.'))
    last_run = models.DateTimeField(_('last_run'), auto_now=True,
        help_text=_('The last time this search this was run.'))

    class Meta:
        app_label = 'panda'
        verbose_name = _('SearchSubscription')
        verbose_name_plural = _('SearchSubscriptions')

    def __unicode__(self):
        if self.dataset:
            return _('%(user)s is searching for %(query)s in %(dataset)s') \
                % {'user': self.user, 'query': self.query, 'dataset': self.dataset}
        else:
            return _('%(user)s is searching for %(query)s in all datasets') \
                % {'user': self.user, 'query': self.query}



########NEW FILE########
__FILENAME__ = slugged_model
#!/usr/bin/env python

import re

from django.db import models
from django.template.defaultfilters import slugify
from django.utils.translation import ugettext_lazy as _

class SluggedModel(models.Model):
    """
    Extend this class to get a slug field and slug generated from a model
    field. We call the 'get_slug_text', '__unicode__' or '__str__'
    methods (in that order) on save() to get text to slugify. The slug may
    have numbers appended to make sure the slug is unique.
    """
    slug = models.SlugField(_('slug'), max_length=256)
    
    class Meta:
        abstract = True
    
    def save(self, *args, **kwargs):
        if not self.slug:
            self.slug = self.generate_unique_slug()  
        
        super(SluggedModel, self).save(*args, **kwargs)

    def generate_unique_slug(self):
        """
        Customized unique_slug function
        """
        if hasattr(self, 'get_slug_text') and callable(self.get_slug_text):
            slug_txt = self.get_slug_text()
        elif hasattr(self, '__unicode__'):
            slug_txt = unicode(self)
        elif hasattr(self, '__str__'):
            slug_txt = str(self)
        else:
            return

        slug = slugify(slug_txt)
        all_slugs = set(sl.values()[0] for sl in self.__class__.objects.values("slug"))

        if slug in all_slugs:
            counterFinder = re.compile(r'-\d+$')
            counter = 2
            slug = '%s-%i' % (slug, counter)

            while slug in all_slugs:
                slug = re.sub(counterFinder, '-%i' % counter, slug)
                counter += 1

        return slug 


########NEW FILE########
__FILENAME__ = task_status
#!/usr/bin/env python

from celery import states
from celery.contrib.abortable import AbortableAsyncResult
from django.db import models
from django.utils.timezone import now 
from django.utils.translation import ugettext_lazy as _
from djcelery.models import TASK_STATE_CHOICES

from panda.models.user_proxy import UserProxy

TASK_STATUS_CHOICES = TASK_STATE_CHOICES
TASK_STATUS_CHOICES.extend([
    ('ABORTED', 'ABORTED'),
    ('ABORT REQUESTED', 'ABORT REQUESTED')
])

class TaskStatus(models.Model):
    """
    An object to track the status of a Celery task, as the
    data available in AsyncResult is not sufficient.
    """
    task_name = models.CharField(_('task_name'), max_length=255,
        help_text=_('Identifying name for this task.'))
    task_description = models.TextField(_('task_description'),
        help_text=_('Description of the task.'))
    status = models.CharField(_('status'), max_length=50, default=states.PENDING, choices=TASK_STATUS_CHOICES,
        help_text=_('Current state of this task.'))
    message = models.CharField(_('message'), max_length=255, blank=True,
        help_text=_('A human-readable message indicating the progress of this task.'))
    start = models.DateTimeField(_('start'), null=True,
        help_text=_('Date and time that this task began processing.'))
    end = models.DateTimeField(_('end'), null=True,
        help_text=_('Date and time that this task ceased processing (either complete or failed).'))
    traceback = models.TextField(_('traceback'), blank=True, null=True, default=None,
        help_text=_('Traceback that exited this task, if it failed.'))
    creator = models.ForeignKey(UserProxy, null=True, related_name='tasks',
        help_text=_('The user who initiated this task.'),
        verbose_name=_('creator'))

    class Meta:
        app_label = 'panda'
        verbose_name = _('Task')
        verbose_name_plural = _('Tasks')

    def __unicode__(self):
        return self.task_description or self.task_name

    def request_abort(self):
        """
        Set flag to abort this task if it is still running.
        """
        if not self.end:
            async_result = AbortableAsyncResult(self.id)
            async_result.abort()

            self.status = 'ABORT REQUESTED'
            self.save()

    def begin(self, message):
        """
        Mark that task has begun.
        """
        self.status = 'STARTED'
        self.start = now()
        self.message = message 
        self.save()

    def update(self, message):
        """
        Update task status message.
        """
        self.message = message 
        self.save()

    def abort(self, message):
        """
        Mark that task has aborted.
        """
        self.status = 'ABORTED'
        self.end = now()
        self.message = message
        self.save()

    def complete(self, message):
        """
        Mark that task has completed.
        """
        self.status = 'SUCCESS'
        self.end = now()
        self.message = message
        self.save()

    def exception(self, message, formatted_traceback):
        """
        Mark that task raised an exception
        """
        self.status = 'FAILURE'
        self.end = now()
        self.message = message 
        self.traceback = formatted_traceback
        self.save()


########NEW FILE########
__FILENAME__ = user_profile
#!/user/bin/env python

import random
import sha

from django.conf import settings
from django.db import models
from django.utils.timezone import now
from django.utils.translation import ugettext as _
from livesettings import config_value

from panda.models.user_proxy import UserProxy
from panda.utils.mail import send_mail

class UserProfile(models.Model):
    """
    User metadata such as their activation key.
    """
    user = models.OneToOneField(UserProxy,
        verbose_name=_('user'))

    activation_key = models.CharField(_('activation_key'), max_length=40, null=True, blank=True)
    activation_key_expiration = models.DateTimeField(_('activation_key_expiration'))

    # NB: This field is no longer used.
    show_login_help = models.BooleanField(_('show_login_help'), default=True, help_text='This field is no longer used.')

    class Meta:
        app_label = 'panda'
        verbose_name = _('UserProfile')
        verbose_name_plural = _('UserProfiles')

    def generate_activation_key(self):
        salt = sha.new(str(random.random())).hexdigest()[:5]
        self.activation_key = sha.new(salt + self.user.username).hexdigest()
        self.activation_key_expiration=now() + settings.PANDA_ACTIVATION_PERIOD

    def send_activation_email(self):
        email_subject = _('Welcome to PANDA, please activate your account!')
        email_body = _('Hello there, the administrator of your organization\'s PANDA has signed you up for an account.\n\nTo activate your account, click this link:\n\nhttp://%(site_domain)s/#activate/%(activation_key)s') \
             % {'site_domain': config_value('DOMAIN', 'SITE_DOMAIN'), 'activation_key': self.activation_key}

        send_mail(email_subject,
                  email_body,
                  [self.user.email])


########NEW FILE########
__FILENAME__ = user_proxy
#!/usr/bin/env python

from django.contrib.auth.models import User
from django.conf import settings
from django.utils.translation import ugettext_lazy as _

from panda import solr

class UserProxy(User):
    """
    User Django's ProxyModel concept to track changes to the User
    model without overriding it. This way related datasets can be
    updated when user details change. Inspired by:

    http://stackoverflow.com/questions/1817244/override-default-user-model-method
    http://stackoverflow.com/questions/1355150/django-when-saving-how-can-you-check-if-a-field-has-changed
    """
    __original_first_name = None
    __original_last_name = None
    __original_email = None

    class Meta:
        proxy = True
        app_label = 'panda'
        verbose_name = _('User')
        verbose_name_plural = _('Users')

    def __init__(self, *args, **kwargs):
        super(User, self).__init__(*args, **kwargs)
        self.__original_first_name = self.first_name
        self.__original_last_name = self.last_name
        self.__original_email = self.email

    def save(self, *args, **kwargs):
        super(User, self).save(*args, **kwargs)
        
        if self.first_name != self.__original_first_name or \
            self.last_name != self.__original_last_name or \
            self.email != self.__original_email:

            if self.datasets.count():
                for dataset in self.datasets.all():
                    dataset.update_full_text(commit=False)
                
                solr.commit(settings.SOLR_DATASETS_CORE)

        self.__original_first_name = self.first_name
        self.__original_last_name = self.last_name
        self.__original_email = self.email


########NEW FILE########
__FILENAME__ = solr
#!/usr/bin/env python

"""
Ultra-lightweight wrapper around Solr's JSON API.

Replaces sunburnt in PANDA. Not a generic solution.
"""
import datetime

from django.conf import settings
from django.core.serializers.json import DjangoJSONEncoder
from django.utils import datetime_safe
from django.utils import simplejson

import requests

class SolrJSONEncoder(DjangoJSONEncoder):
    """
    Custom JSONEncoder based on DjangoJSONEncoder that formats datetimes the way Solr likes them. 
    """
    def default(self, o):
        if isinstance(o, datetime.datetime):
            d = datetime_safe.new_datetime(o)
            return d.strftime('%Y-%m-%dT%H:%M:%SZ')
        else:
            return super(SolrJSONEncoder, self).default(o)

def dumps(data):
    return simplejson.dumps(data, cls=SolrJSONEncoder)

def loads(data):
    return simplejson.loads(data)

class SolrError(Exception):
    """
    Exceptionr raised when a Solr requests fails for any reason.
    """
    def __init__(self, response, *args, **kwargs):
        self.status_code = response.status_code
        self.response_body = response.content

        super(SolrError, self).__init__(*args, **kwargs)

    def __unicode__(self):
        return self.response_body

def add(core, documents, commit=False):
    """
    Add a document or list of documents to Solr.

    Does not commit changes by default.
    """
    url = ''.join([settings.SOLR_ENDPOINT, '/', core, '/update'])
    params = { 'commit': 'true' } if commit else {}
    response = requests.post(url, dumps(documents), params=params, headers={ 'Content-Type': 'application/json' })

    if response.status_code != 200:
        raise SolrError(response)
    
    return loads(response.content)

def commit(core):
    """
    Commit all staged changes to the Solr index.
    """
    url = ''.join([settings.SOLR_ENDPOINT, '/', core, '/update'])
    response = requests.post(url, '[]', params={ 'commit': 'true' }, headers={ 'Content-Type': 'application/json' })

    if response.status_code != 200:
        raise SolrError(response)
    
    return loads(response.content)

def delete(core, q, commit=True):
    """
    Delete documents by query from the Solr index.

    Commits changes by default.
    """
    url = ''.join([settings.SOLR_ENDPOINT, '/', core, '/update'])
    params = { 'commit': 'true' } if commit else {}
    response = requests.post(url, dumps({ 'delete': { 'query': q } }), params=params, headers={ 'Content-Type': 'application/json' })
    
    if response.status_code != 200:
        raise SolrError(response)
    
    return loads(response.content)

def query(core, q, limit=10, offset=0, sort='_docid_ asc'):
    """
    Execute a simple, raw query against the Solr index.
    """
    url = ''.join([settings.SOLR_ENDPOINT, '/', core, '/select'])
    response = requests.get(url, params={ 'q': q, 'mm': '1', 'start': offset, 'rows': limit, 'sort': sort }, headers={ 'Content-Type': 'application/json' })

    if response.status_code != 200:
        raise SolrError(response)
    
    return loads(response.content)

def query_grouped(core, q, group_field, limit=10, offset=0, sort='_docid_ asc', group_limit=settings.PANDA_DEFAULT_SEARCH_ROWS_PER_GROUP, group_offset=0):
    """
    Execute a query and return results in a grouped format
    appropriate for the PANDA API.
    """
    url = ''.join([settings.SOLR_ENDPOINT, '/', core, '/select'])
    response = requests.get(url, params={ 'q': q, 'mm': '1', 'start': offset, 'rows': limit, 'sort': sort, 'group': 'true', 'group.field': group_field, 'group.limit': group_limit, 'group.offset': group_offset, 'group.ngroups': 'true' }, headers={ 'Content-Type': 'application/json' })

    if response.status_code != 200:
        raise SolrError(response)

    return loads(response.content)


########NEW FILE########
__FILENAME__ = storage
#!/usr/bin/env python

from io import BufferedWriter, FileIO
import os

from ajaxuploader.backends.base import AbstractUploadBackend
from django.conf import settings

from panda.api import DataUploadResource, RelatedUploadResource, UserResource
from panda.models import Dataset, DataUpload, RelatedUpload, UserProxy

class PANDAAbstractUploadBackend(AbstractUploadBackend):
    """
    Customized backend to handle AJAX uploads.
    """
    def update_filename(self, request, filename):
        """
        Verify that the filename is unique, if it isn't append and iterate
        a counter until it is.
        """
        self._original_filename = filename

        filename = self._original_filename
        root, ext = os.path.splitext(self._original_filename)
        path = os.path.join(settings.MEDIA_ROOT, filename)

        i = 1

        while os.path.exists(path):
            filename = '%s%i%s' % (root, i, ext)
            path = os.path.join(settings.MEDIA_ROOT, filename)
            i += 1

        return filename 

    def setup(self, filename):
        """
        Open the destination file for writing.
        """
        self._path = os.path.join(settings.MEDIA_ROOT, filename)

        try:
            os.makedirs(os.path.realpath(os.path.dirname(self._path)))
        except:
            pass

        self._dest = BufferedWriter(FileIO(self._path, "w"))

    def upload_chunk(self, chunk):
        """
        Write a chunk of data to the destination.
        """
        self._dest.write(chunk)

    def upload_complete(self, request, filename):
        """
        Close the destination file.
        """
        self._dest.close()

class PANDADataUploadBackend(PANDAAbstractUploadBackend):
    """
    Backend specifically for DataUploads.
    """
    def upload_complete(self, request, filename):
        """
        Create a DataUpload object.
        """
        try:
            super(PANDADataUploadBackend, self).upload_complete(request, filename)

            root, ext = os.path.splitext(filename)
            path = os.path.join(settings.MEDIA_ROOT, filename)
            size = os.path.getsize(path)

            if 'dataset_slug' in request.REQUEST:
                dataset = Dataset.objects.get(slug=request.REQUEST['dataset_slug'])
            else:
                dataset = None

            encoding = request.REQUEST.get('encoding', 'utf-8')

            if not encoding:
                encoding = 'utf-8'

            # Because users may have authenticated via headers the request.user may
            # not be a full User instance. To be sure, we fetch one.
            creator = UserProxy.objects.get(id=request.user.id)

            upload = DataUpload.objects.create(
                filename=filename,
                original_filename=self._original_filename,
                size=size,
                creator=creator,
                dataset=dataset,
                encoding=encoding)

            if dataset:
                dataset.update_full_text()

            resource = DataUploadResource()
            bundle = resource.build_bundle(obj=upload, request=request)
            data = resource.full_dehydrate(bundle).data

            # django-ajax-upoader does not use the Tastypie serializer
            # so we must 'manually' serialize the embedded resource bundle
            resource = UserResource()
            bundle = data['creator']
            user_data = resource.full_dehydrate(bundle).data

            data['creator'] = user_data
        except Exception, e:
            # This global error handler is a kludge to ensure IE8 can properly handle the responses
            return { 'error_message': e.message, 'success': False }

        return data 

class PANDARelatedUploadBackend(PANDAAbstractUploadBackend):
    """
    Backend specifically for RelatedUploads.
    """
    def upload_complete(self, request, filename):
        """
        Create a RelatedUpload object.
        """
        try:
            super(PANDARelatedUploadBackend, self).upload_complete(request, filename)

            root, ext = os.path.splitext(filename)
            path = os.path.join(settings.MEDIA_ROOT, filename)
            size = os.path.getsize(path)

            dataset = Dataset.objects.get(slug=request.REQUEST['dataset_slug'])

            # Because users may have authenticated via headers the request.user may
            # not be a full User instance. To be sure, we fetch one.
            creator = UserProxy.objects.get(id=request.user.id)

            upload = RelatedUpload.objects.create(
                filename=filename,
                original_filename=self._original_filename,
                size=size,
                creator=creator,
                dataset=dataset)

            dataset.update_full_text()

            resource = RelatedUploadResource()
            bundle = resource.build_bundle(obj=upload, request=request)
            data = resource.full_dehydrate(bundle).data

            # django-ajax-upoader does not use the Tastypie serializer
            # so we must 'manually' serialize the embedded resource bundle
            resource = UserResource()
            bundle = data['creator']
            user_data = resource.full_dehydrate(bundle).data

            data['creator'] = user_data
        except Exception, e:
            # This global error handler is a kludge to ensure IE8 can properly handle the responses
            return { 'error_message': e.message, 'success': False }

        return data


########NEW FILE########
__FILENAME__ = base
from celery.task import Task as CeleryTask
from celery.contrib.abortable import AbortableTask as CeleryAbortableTask
from django.utils import translation
from django.conf import settings
import logging
log = logging.getLogger('panda.tasks.base')

class Task(CeleryTask):
    def __init__(self,*args,**kwargs):
        if settings.USE_I18N:
            try:
                translation.activate(settings.LANGUAGE_CODE)
            except Exception, e:
                log.warn("Error activating translation library", e)
        super(CeleryTask,self).__init__(*args,**kwargs)

class AbortableTask(CeleryAbortableTask):
    def __init__(self,*args,**kwargs):
        if settings.USE_I18N:
            try:
                translation.activate(settings.LANGUAGE_CODE)
            except Exception, e:
                log.warn("Error activating translation library", e)
        super(CeleryAbortableTask,self).__init__(*args,**kwargs)


########NEW FILE########
__FILENAME__ = export_csv
#!/usr/bin/env python

import datetime
import logging
from math import floor
import os.path
import time

from csvkit import CSVKitWriter
from django.conf import settings
from django.utils import simplejson as json
from django.utils.translation import ugettext
from livesettings import config_value

from panda import solr
from panda.tasks.export_file import ExportFileTask 

SOLR_PAGE_SIZE = 500

class ExportCSVTask(ExportFileTask):
    """
    Task to export all data for a dataset to a CSV.
    """
    name = 'panda.tasks.export.csv'

    def run(self, dataset_slug, query=None, filename=None, *args, **kwargs):
        """
        Execute export.
        """
        from panda.models import Dataset

        log = logging.getLogger(self.name)
        log.info('Beginning export, dataset_slug:%s %s' % (dataset_slug, query))

        try:
            dataset = Dataset.objects.get(slug=dataset_slug)
        except Dataset.DoesNotExist:
            log.warning('Export failed due to Dataset being deleted, dataset_slug: %s' % dataset_slug)

            return

        task_status = dataset.current_task
        task_status.begin(ugettext('Preparing to export'))

        if not filename:
            filename = '%s_%s.csv' % (dataset_slug, datetime.datetime.utcnow().isoformat())
        else:
            filename = '%s.csv' % filename

        path = os.path.join(settings.EXPORT_ROOT, filename)

        try:
            os.makedirs(os.path.realpath(os.path.dirname(path)))
        except:
            pass

        f = open(path, 'w')
        writer = CSVKitWriter(f)

        # Header
        writer.writerow([c['name'] for c in dataset.column_schema])

        solr_query_bits = []

        if query:
            solr_query_bits.append('(%s)' % query) 

        solr_query_bits.append('dataset_slug:%s' % dataset_slug)

        solr_query = ' AND '.join(solr_query_bits)

        response = solr.query(
            settings.SOLR_DATA_CORE,
            solr_query,
            offset=0,
            limit=0
        )

        total_count = response['response']['numFound']
        n = 0
        throttle = config_value('PERF', 'TASK_THROTTLE')

        while n < total_count:
            response = solr.query(
                settings.SOLR_DATA_CORE,
                solr_query,
                offset=n,
                limit=SOLR_PAGE_SIZE
            )

            results = response['response']['docs']

            for row in results:
                data = json.loads(row['data'])

                writer.writerow(data)

            task_status.update(ugettext('%.0f%% complete') % floor(float(n) / float(total_count) * 100))

            if self.is_aborted():
                task_status.abort(ugettext('Aborted after exporting %.0f%%') % floor(float(n) / float(total_count) * 100))

                log.warning('Export aborted, dataset_slug: %s' % dataset_slug)

                return

            n += SOLR_PAGE_SIZE
            
            time.sleep(throttle)

        f.close()

        task_status.update(ugettext('100% complete'))

        log.info('Finished export, dataset_slug:%s %s' % (dataset_slug, query))

        return filename


########NEW FILE########
__FILENAME__ = export_file
#!/usr/bin/env python

import logging
import os.path
import traceback

from panda.tasks.base import AbortableTask
from django.conf import settings
from django.utils.translation import ugettext

from panda.utils.notifications import notify

class ExportFileTask(AbortableTask):
    """
    Base type for file export tasks. 
    """
    abstract = True

    # All subclasses should be within this namespace
    name = 'panda.tasks.export'

    def run(self, dataset_slug, *args, **kwargs):
        """
        Execute export.
        """
        raise NotImplementedError() 

    def after_return(self, status, retval, task_id, args, kwargs, einfo):
        """
        Save final status, results, etc.
        """
        from panda.models import Dataset

        log = logging.getLogger(self.name)

        try:
            dataset = Dataset.objects.get(slug=args[0])
        except Dataset.DoesNotExist:
            log.warning('Can not send export notifications due to Dataset being deleted, dataset_slug: %s' % args[0])

            return
        
        query = kwargs.get('query', None)

        self.send_notifications(dataset, query, retval, einfo) 

    def send_notifications(self, dataset, query, retval, einfo):
        """
        Send user notifications this task has finished.
        """
        from panda.models import Export

        task_status = dataset.current_task 

        export = None
        extra_context = {
            'query': query,
            'related_dataset': dataset
        }
        url = None

        if einfo:
            if hasattr(einfo, 'traceback'):
                tb = einfo.traceback
            else:
                tb = ''.join(traceback.format_tb(einfo[2]))

            task_status.exception(
                ugettext('Export failed'),
                u'%s\n\nTraceback:\n%s' % (unicode(retval), tb)
            )

            template_prefix = 'export_failed'
            extra_context['error'] = unicode(retval)
            extra_context['traceback'] = tb
            notification_type = 'Error'
        elif self.is_aborted():
            template_prefix = 'export_aborted'
            notification_type = 'Info'
        else:
            task_status.complete(ugettext('Export complete'))

            export = Export.objects.create(
                filename=retval,
                original_filename=retval,
                size=os.path.getsize(os.path.join(settings.EXPORT_ROOT, retval)),
                creator=task_status.creator,
                creation_date=task_status.start,
                dataset=dataset)

            extra_context['related_export'] = export

            url = '#export/%i' % export.id

            template_prefix = 'export_complete'
            notification_type = 'Info'
            
        if task_status.creator:
            notify(
                task_status.creator,
                template_prefix,
                notification_type,
                url,
                extra_context=extra_context
            )


########NEW FILE########
__FILENAME__ = export_search
#!/usr/bin/env python

import logging
from math import floor
import os.path
import time
from traceback import format_tb
from zipfile import ZipFile

from panda.tasks.base import AbortableTask
from csvkit import CSVKitWriter
from django.conf import settings
from django.utils import simplejson as json
from django.utils.timezone import now 
from django.utils.translation import ugettext
from livesettings import config_value

from panda import solr
from panda.utils.notifications import notify

SOLR_PAGE_SIZE = 500

class ExportSearchTask(AbortableTask):
    """
    Task to export all search results to a batch of CSV files.
    """
    name = 'panda.tasks.export.search'

    def run(self, query, task_status_id, filename=None, *args, **kwargs):
        """
        Execute export.
        """
        from panda.models import Dataset, TaskStatus

        log = logging.getLogger(self.name)
        log.info('Beginning export, query: %s' % query)

        task_status = TaskStatus.objects.get(id=task_status_id)
        task_status.begin('Preparing to import')

        if not filename:
            filename = 'search_export_%s' % (now().isoformat())

        zip_name = '%s.zip' % filename

        path = os.path.join(settings.EXPORT_ROOT, filename)
        zip_path = os.path.join(settings.EXPORT_ROOT, zip_name)

        try:
            os.makedirs(os.path.realpath(path))
        except:
            pass
        
        zipfile = ZipFile(zip_path, 'w')

        response = solr.query_grouped(
            settings.SOLR_DATA_CORE,
            query,
            'dataset_slug',
            offset=0,
            limit=1000,
            group_limit=0,
            group_offset=0
        )
        groups = response['grouped']['dataset_slug']['groups']

        datasets = {}

        for group in groups:
            dataset_slug = group['groupValue']
            count = group['doclist']['numFound']

            datasets[dataset_slug] = count

        total_n = 0
        throttle = config_value('PERF', 'TASK_THROTTLE')

        for dataset_slug in datasets:
            try:
                dataset = Dataset.objects.get(slug=dataset_slug)
            except Dataset.DoesNotExist:
                log.warning('Skipping part of export due to Dataset being deleted, dataset_slug: %s' % dataset_slug)

                continue

            filename = '%s.csv' % dataset_slug
            file_path = os.path.join(path, filename)

            f = open(file_path, 'w')
            writer = CSVKitWriter(f)
            
            # Header
            writer.writerow([c['name'] for c in dataset.column_schema])
                
            response = solr.query(
                settings.SOLR_DATA_CORE,
                query,
                offset=0,
                limit=0
            )

            # Update dataset and total counts for progress tracking
            datasets[dataset_slug] = response['response']['numFound']
            total_count = sum(datasets.values())

            n = 0

            while n < datasets[dataset_slug]:

                response = solr.query(
                    settings.SOLR_DATA_CORE,
                    'dataset_slug:%s AND (%s)' % (dataset_slug, query),
                    offset=n,
                    limit=SOLR_PAGE_SIZE
                )

                results = response['response']['docs']

                for row in results:
                    data = json.loads(row['data'])

                    writer.writerow(data)

                task_status.update(ugettext('%.0f%% complete') % floor(float(total_n) / float(total_count) * 100))

                if self.is_aborted():
                    task_status.abort(ugettext('Aborted after exporting %.0f%%') % floor(float(total_n) / float(total_count) * 100))

                    log.warning('Export aborted, query: %s' % query)

                    return

                n += SOLR_PAGE_SIZE
                total_n += response['response']['numFound'] 
                
                time.sleep(throttle)

            f.close()

            # Add to zip and nuke temp file
            zipfile.write(file_path, filename)
            os.remove(file_path)

        # Finish zip file and nuke temp directory
        zipfile.close()
        os.rmdir(path)

        task_status.update(ugettext('100% complete'))

        log.info('Finished export, query: %s' % query)

        return zip_name

    def after_return(self, status, retval, task_id, args, kwargs, einfo):
        """
        Save final status, results, etc.
        """
        from panda.models import TaskStatus

        query = args[0]
        task_status = TaskStatus.objects.get(id=args[1])

        self.send_notifications(query, task_status, retval, einfo) 

    def send_notifications(self, query, task_status, retval, einfo):
        """
        Send user notifications this task has finished.
        """
        from panda.models import Export

        export = None
        extra_context = { 'query': query }
        url = None

        if einfo:
            if isinstance(einfo, tuple):
                tb = '\n'.join(format_tb(einfo[2]))
            else:
                tb = einfo.traceback

            task_status.exception(
                ugettext('Export failed'),
                u'%s\n\nTraceback:\n%s' % (unicode(retval), tb)
            )
            
            template_prefix = 'export_search_failed'
            extra_context['error'] = unicode(retval)
            extra_context['traceback'] = tb
            notification_type = 'Error'
        elif self.is_aborted():
            template_prefix = 'export_search_aborted'
            notification_type = 'Info'
        else:
            task_status.complete(ugettext('Export complete'))

            export = Export.objects.create(
                filename=retval,
                original_filename=retval,
                size=os.path.getsize(os.path.join(settings.EXPORT_ROOT, retval)),
                creator=task_status.creator,
                creation_date=task_status.start,
                dataset=None)

            extra_context['related_export'] = export

            url = '#export/%i' % export.id

            template_prefix = 'export_search_complete'
            notification_type = 'Info'

        if task_status.creator:
            notify(
                task_status.creator,
                template_prefix,
                notification_type,
                url,
                extra_context=extra_context
            )


########NEW FILE########
__FILENAME__ = import_csv
#!/usr/bin/env python

import logging
from math import floor
import time

from csvkit import CSVKitReader
from django.conf import settings
from django.utils.translation import ugettext
from livesettings import config_value

from panda import solr, utils
from panda.exceptions import DataImportError
from panda.tasks.import_file import ImportFileTask 
from panda.utils.typecoercion import DataTyper

SOLR_ADD_BUFFER_SIZE = 500

class ImportCSVTask(ImportFileTask):
    """
    Task to import all data for a dataset from a CSV.
    """
    name = 'panda.tasks.import.csv'

    def _count_lines(self, filename):
        """
        Efficiently count the number of lines in a file.
        """
        with open(filename) as f:
            for i, l in enumerate(f):
                pass
        return i + 1

    def run(self, dataset_slug, upload_id, external_id_field_index=None, *args, **kwargs):
        """
        Execute import.
        """
        from panda.models import Dataset, DataUpload
        
        log = logging.getLogger(self.name)
        log.info('Beginning import, dataset_slug: %s' % dataset_slug)

        try:
            dataset = Dataset.objects.get(slug=dataset_slug)
        except Dataset.DoesNotExist:
            log.warning('Import failed due to Dataset being deleted, dataset_slug: %s' % dataset_slug)

            return

        upload = DataUpload.objects.get(id=upload_id)

        task_status = dataset.current_task
        task_status.begin(ugettext('Preparing to import'))

        line_count = self._count_lines(upload.get_path())

        if self.is_aborted():
            task_status.abort('Aborted during preperation')

            log.warning('Import aborted, dataset_slug: %s' % dataset_slug)

            return

        f = open(upload.get_path(), 'r')

        reader = CSVKitReader(f, encoding=upload.encoding, **upload.dialect_as_parameters())
        reader.next()

        add_buffer = []
        data_typer = DataTyper(dataset.column_schema)
        throttle = config_value('PERF', 'TASK_THROTTLE')

        i = 0

        while True:
            # The row number which is about to be read, for error handling and indexing
            i += 1

            try:
                row = reader.next()
            except StopIteration:
                i -= 1
                break
            except UnicodeDecodeError:
                raise DataImportError(ugettext('This CSV file contains characters that are not %(encoding)s encoded in or after row %(row)i. You need to re-upload this file and input the correct encoding in order to import data from this file.') % { 'encoding': upload.encoding, 'row': i })

            external_id = None

            if external_id_field_index is not None:
                external_id = row[external_id_field_index]

            data = utils.solr.make_data_row(dataset, row, data_upload=upload, external_id=external_id)
            data = data_typer(data, row)

            add_buffer.append(data)

            if i % SOLR_ADD_BUFFER_SIZE == 0:
                solr.add(settings.SOLR_DATA_CORE, add_buffer)

                add_buffer = []

                task_status.update(ugettext('%.0f%% complete (estimated)') % floor(float(i) / float(line_count) * 100))

                if self.is_aborted():
                    task_status.abort(ugettext('Aborted after importing %.0f%% (estimated)') % floor(float(i) / float(line_count) * 100))

                    log.warning('Import aborted, dataset_slug: %s' % dataset_slug)

                    return

                time.sleep(throttle)

        if add_buffer:
            solr.add(settings.SOLR_DATA_CORE, add_buffer)
            add_buffer = []

        solr.commit(settings.SOLR_DATA_CORE)

        f.close()

        task_status.update('100% complete')

        # Refresh dataset from database so there is no chance of crushing changes made since the task started
        try:
            dataset = Dataset.objects.get(slug=dataset_slug)
        except Dataset.DoesNotExist:
            log.warning('Import could not be completed due to Dataset being deleted, dataset_slug: %s' % dataset_slug)

            return

        if not dataset.row_count:
            dataset.row_count = i
        else:
            dataset.row_count += i

        dataset.column_schema = data_typer.schema

        dataset.save()

        # Refres
        upload = DataUpload.objects.get(id=upload_id)

        upload.imported = True
        upload.save()

        log.info('Finished import, dataset_slug: %s' % dataset_slug)

        return data_typer


########NEW FILE########
__FILENAME__ = import_file
#!/usr/bin/env python

import logging
import traceback

from panda.tasks.base import AbortableTask
from django.conf import settings
from django.utils.translation import ugettext

from panda import solr
from panda.utils.notifications import notify

SOLR_ADD_BUFFER_SIZE = 500

class ImportFileTask(AbortableTask):
    """
    Base type for file import tasks. 
    """
    abstract = True

    # All subclasses should be within this namespace
    name = 'panda.tasks.import'

    def run(self, dataset_slug, upload_id, *args, **kwargs):
        """
        Execute import.
        """
        raise NotImplementedError() 

    def after_return(self, status, retval, task_id, args, kwargs, einfo):
        """
        Save final status, results, etc.
        """
        from panda.models import Dataset

        log = logging.getLogger(self.name)

        try:
            dataset = Dataset.objects.get(slug=args[0])
        except Dataset.DoesNotExist:
            log.warning('Can not send import notifications due to Dataset being deleted, dataset_slug: %s' % args[0])

            return

        try:
            try:
                self.send_notifications(dataset, retval, einfo)
            finally:
                # If import failed, clear any data that might be staged
                if dataset.current_task.status == 'FAILURE':
                    solr.delete(settings.SOLR_DATA_CORE, 'dataset_slug:%s' % args[0], commit=True)
        finally:
            dataset.unlock()

    def send_notifications(self, dataset, retval, einfo):
        """
        Send user notifications this task has finished.
        """
        task_status = dataset.current_task 

        extra_context = {
            'related_dataset': dataset
        }

        if einfo:
            if hasattr(einfo, 'traceback'):
                tb = einfo.traceback
            else:
                tb = ''.join(traceback.format_tb(einfo[2]))

            task_status.exception(
                ugettext('Import failed'),
                u'%s\n\nTraceback:\n%s' % (unicode(retval), tb)
            )
            
            template_prefix = 'import_failed'
            extra_context['error'] = unicode(retval)
            extra_context['traceback'] = tb
            notification_type = 'Error'
        elif self.is_aborted():
            template_prefix = 'import_aborted'
            notification_type = 'Info'
        else:
            task_status.complete(ugettext('Import complete'))

            template_prefix = 'import_complete'
            extra_context['type_summary'] = retval.summarize()
            notification_type = 'Info'
        
        if task_status.creator:
            notify(
                task_status.creator,
                template_prefix,
                notification_type,
                url='#dataset/%s' % dataset.slug,
                extra_context=extra_context
            )


########NEW FILE########
__FILENAME__ = import_xls
#!/usr/bin/env python

import logging
from math import floor
import time

from django.conf import settings
from django.utils.translation import ugettext
import xlrd
from livesettings import config_value

from panda import solr, utils
from panda.tasks.import_file import ImportFileTask
from panda.utils.typecoercion import DataTyper

SOLR_ADD_BUFFER_SIZE = 500

class ImportXLSTask(ImportFileTask):
    """
    Task to import all data for a dataset from an Excel XLS file.
    """
    name = 'panda.tasks.import.xls'

    def run(self, dataset_slug, upload_id, external_id_field_index=None, *args, **kwargs):
        """
        Execute import.
        """
        from panda.models import Dataset, DataUpload
        
        log = logging.getLogger(self.name)
        log.info('Beginning import, dataset_slug: %s' % dataset_slug)
    
        try:
            dataset = Dataset.objects.get(slug=dataset_slug)
        except Dataset.DoesNotExist:
            log.warning('Import failed due to Dataset being deleted, dataset_slug: %s' % dataset_slug)

            return

        upload = DataUpload.objects.get(id=upload_id)

        task_status = dataset.current_task
        task_status.begin(ugettext('Preparing to import'))

        book = xlrd.open_workbook(upload.get_path(), on_demand=True)
        sheet = book.sheet_by_index(0)
        row_count = sheet.nrows
        
        add_buffer = []
        data_typer = DataTyper(dataset.column_schema)
        throttle = config_value('PERF', 'TASK_THROTTLE')

        for i in range(1, row_count):
            values = sheet.row_values(i)
            types = sheet.row_types(i)

            normal_values = []

            for v, t in zip(values, types):
                if t == xlrd.biffh.XL_CELL_DATE:
                    v = utils.xls.normalize_date(v, book.datemode)
                elif t == xlrd.biffh.XL_CELL_NUMBER:
                    if v % 1 == 0:
                        v = int(v)

                normal_values.append(unicode(v))

            external_id = None

            if external_id_field_index is not None:
                external_id = values[external_id_field_index]

            data = utils.solr.make_data_row(dataset, normal_values, data_upload=upload, external_id=external_id)
            data = data_typer(data, normal_values)

            add_buffer.append(data)

            if i % SOLR_ADD_BUFFER_SIZE == 0:
                solr.add(settings.SOLR_DATA_CORE, add_buffer)
                add_buffer = []

                task_status.update(ugettext('%.0f%% complete') % floor(float(i) / float(row_count) * 100))

                if self.is_aborted():
                    task_status.abort(ugettext('Aborted after importing %.0f%%') % floor(float(i) / float(row_count) * 100))

                    log.warning('Import aborted, dataset_slug: %s' % dataset_slug)

                    return
            
                time.sleep(throttle)

        if add_buffer:
            solr.add(settings.SOLR_DATA_CORE, add_buffer)
            add_buffer = []

        solr.commit(settings.SOLR_DATA_CORE)

        task_status.update(ugettext('100% complete'))

        # Refresh dataset from database so there is no chance of crushing changes made since the task started
        try:
            dataset = Dataset.objects.get(slug=dataset_slug)
        except Dataset.DoesNotExist:
            log.warning('Import could not be completed due to Dataset being deleted, dataset_slug: %s' % dataset_slug)

            return

        if not dataset.row_count:
            dataset.row_count = i
        else:
            dataset.row_count += i

        dataset.column_schema = data_typer.schema

        dataset.save()

        # Refres
        upload = DataUpload.objects.get(id=upload_id)

        upload.imported = True
        upload.save()

        log.info('Finished import, dataset_slug: %s' % dataset_slug)
        
        return data_typer


########NEW FILE########
__FILENAME__ = import_xlsx
#!/usr/bin/env python

import datetime
import logging
from math import floor
import time

from django.conf import settings
from django.utils.translation import ugettext
from livesettings import config_value
from openpyxl.reader.excel import load_workbook

from panda import solr, utils
from panda.tasks.import_file import ImportFileTask
from panda.utils.typecoercion import DataTyper

SOLR_ADD_BUFFER_SIZE = 500

class ImportXLSXTask(ImportFileTask):
    """
    Task to import all data for a dataset from an Excel/OpenOffice XLSX file.
    """
    name = 'panda.tasks.import.xlsx'

    def run(self, dataset_slug, upload_id, external_id_field_index=None, *args, **kwargs):
        """
        Execute import.
        """
        from panda.models import Dataset, DataUpload
        
        log = logging.getLogger(self.name)
        log.info('Beginning import, dataset_slug: %s' % dataset_slug)

        try:
            dataset = Dataset.objects.get(slug=dataset_slug)
        except Dataset.DoesNotExist:
            log.warning('Import failed due to Dataset being deleted, dataset_slug: %s' % dataset_slug)

            return

        upload = DataUpload.objects.get(id=upload_id)

        task_status = dataset.current_task
        task_status.begin(ugettext('Preparing to import'))

        book = load_workbook(upload.get_path(), use_iterators=True)
        sheet = book.get_active_sheet()
        row_count = sheet.get_highest_row()
        
        add_buffer = []
        data_typer = DataTyper(dataset.column_schema)
        throttle = config_value('PERF', 'TASK_THROTTLE')

        for i, row in enumerate(sheet.iter_rows()):
            # Skip header
            if i == 0:
                continue

            values = []

            for c in row:
                value = c.internal_value

                if value.__class__ is datetime.datetime:
                    value = utils.xlsx.normalize_date(value)
                elif value.__class__ is float:
                    if value % 1 == 0:
                        value = int(value)

                if value.__class__ in (datetime.datetime, datetime.date, datetime.time):
                    value = value.isoformat()

                values.append(value)

            external_id = None

            if external_id_field_index is not None:
                external_id = values[external_id_field_index]

            data = utils.solr.make_data_row(dataset, values, data_upload=upload, external_id=external_id)
            data = data_typer(data, values)

            add_buffer.append(data)

            if i % SOLR_ADD_BUFFER_SIZE == 0:
                solr.add(settings.SOLR_DATA_CORE, add_buffer)
                add_buffer = []

                task_status.update(ugettext('%.0f%% complete') % floor(float(i) / float(row_count) * 100))

                if self.is_aborted():
                    task_status.abort(ugettext('Aborted after importing %.0f%%') % floor(float(i) / float(row_count) * 100))

                    log.warning('Import aborted, dataset_slug: %s' % dataset_slug)

                    return
                
                time.sleep(throttle)

        if add_buffer:
            solr.add(settings.SOLR_DATA_CORE, add_buffer)
            add_buffer = []

        solr.commit(settings.SOLR_DATA_CORE)

        task_status.update(ugettext('100% complete'))

        # Refresh dataset from database so there is no chance of crushing changes made since the task started
        try:
            dataset = Dataset.objects.get(slug=dataset_slug)
        except Dataset.DoesNotExist:
            log.warning('Import could not be completed due to Dataset being deleted, dataset_slug: %s' % dataset_slug)

            return

        if not dataset.row_count:
            dataset.row_count = i
        else:
            dataset.row_count += i
        
        dataset.column_schema = data_typer.schema

        dataset.save()

        # Refres
        upload = DataUpload.objects.get(id=upload_id)

        upload.imported = True
        upload.save()

        log.info('Finished import, dataset_slug: %s' % dataset_slug)

        return data_typer


########NEW FILE########
__FILENAME__ = purge_data
#!/usr/bin/env python

import logging

from django.conf import settings
from panda.tasks.base import Task

from panda import solr

class PurgeDataTask(Task):
    """
    Purge a dataset from Solr.
    """
    name = 'panda.tasks.purge.data'

    def run(self, dataset_slug, data_upload_id=None):
        from panda.models import Dataset

        log = logging.getLogger(self.name)
        log.info('Beginning purge, dataset_slug: %s' % dataset_slug)

        if data_upload_id:
            q = 'data_upload_id:%i' % data_upload_id
        else:
            q = 'dataset_slug:%s' % dataset_slug

        solr.delete(settings.SOLR_DATA_CORE, q)

        try:
            # If the dataset hasn't been deleted, update its row count
            dataset = Dataset.objects.get(slug=dataset_slug)
            dataset.row_count = dataset._count_rows()
            dataset.save()
        except Dataset.DoesNotExist:
            pass

        log.info('Finished purge, dataset_slug: %s' % dataset_slug)


########NEW FILE########
__FILENAME__ = purge_orphaned_uploads
#!/usr/bin/env python

from itertools import chain
import logging
import os

from panda.tasks.base import Task
from django.conf import settings

SOLR_ADD_BUFFER_SIZE = 500

class PurgeOrphanedUploadsTask(Task):
    """
    Task to import all data for a dataset from a CSV.
    """
    name = 'panda.tasks.cron.purge_orphaned_uploads'

    def run(self, fake=False, *args, **kwargs):
        """
        Execute import.
        """
        from panda.models import DataUpload, RelatedUpload

        log = logging.getLogger(self.name)
        log.info('Purging orphaned uploads')

        local_files = os.listdir(settings.MEDIA_ROOT)
        data_uploads = DataUpload.objects.all()
        related_uploads = RelatedUpload.objects.all()

        for upload in chain(data_uploads, related_uploads):
            # This file is accounted for
            try:
                local_files.remove(upload.filename)
            except ValueError:
                pass

            if not upload.dataset:
                if fake:
                    log.info('Would delete upload: %s\n' % upload)
                else:
                    log.info('Deleted upload: %s\n' % upload)
                    upload.delete()

        for f in local_files:
            path = os.path.join(settings.MEDIA_ROOT, f)

            if fake:
                log.info('Would delete file: %s\n' % path)
            else:
                log.info('Deleted file: %s\n' % path)
                os.remove(path)

        log.info('Purge complete')


########NEW FILE########
__FILENAME__ = reindex
#!/usr/bin/env python

import logging
from math import floor
import time
import traceback

from panda.tasks.base import AbortableTask
from django.conf import settings
from django.utils import simplejson as json
from django.utils.translation import ugettext
from livesettings import config_value

from panda import solr, utils
from panda.utils.notifications import notify
from panda.utils.typecoercion import DataTyper 

SOLR_READ_BUFFER_SIZE = 500
SOLR_ADD_BUFFER_SIZE = 500

class ReindexTask(AbortableTask):
    """
    Task to import all data for a dataset from a CSV.
    """
    name = 'panda.tasks.reindex'

    def run(self, dataset_slug, *args, **kwargs):
        """
        Execute reindex.
        """
        from panda.models import Dataset
        
        log = logging.getLogger(self.name)
        log.info('Beginning reindex, dataset_slug: %s' % dataset_slug)

        try:
            dataset = Dataset.objects.get(slug=dataset_slug)
        except Dataset.DoesNotExist:
            log.warning('Reindexing failed due to Dataset being deleted, dataset_slug: %s' % dataset_slug)

            return

        task_status = dataset.current_task
        task_status.begin(ugettext('Preparing to reindex'))

        if self.is_aborted():
            task_status.abort(ugettext('Aborted during preparation'))

            log.warning('Reindex aborted, dataset_slug: %s' % dataset_slug)

            return

        read_buffer = []
        add_buffer = []
        data_typer = DataTyper(dataset.column_schema)
        throttle = config_value('PERF', 'TASK_THROTTLE')

        i = 0

        while i < dataset.row_count:
            if not read_buffer:
                query = 'dataset_slug: %s' % (dataset.slug)
                response = solr.query(settings.SOLR_DATA_CORE, query, limit=SOLR_READ_BUFFER_SIZE, offset=i)
                read_buffer = response['response']['docs']

            data = read_buffer.pop(0)
            row = json.loads(data['data'])

            new_data = utils.solr.make_data_row(dataset, row)
            new_data['id'] = data['id'] 
            new_data['data_upload_id'] = data['data_upload_id']
            new_data = data_typer(new_data, row)

            add_buffer.append(new_data)

            if i % SOLR_ADD_BUFFER_SIZE == 0:
                solr.add(settings.SOLR_DATA_CORE, add_buffer)

                add_buffer = []

                task_status.update(ugettext('%.0f%% complete') % floor(float(i) / float(dataset.row_count) * 100))

                if self.is_aborted():
                    task_status.abort(ugettext('Aborted after reindexing %.0f%%') % floor(float(i) / float(dataset.row_count) * 100))

                    log.warning('Reindex aborted, dataset_slug: %s' % dataset_slug)

                    return
            
                time.sleep(throttle)

            i += 1

        if add_buffer:
            solr.add(settings.SOLR_DATA_CORE, add_buffer)
            add_buffer = []

        solr.commit(settings.SOLR_DATA_CORE)

        task_status.update(ugettext('100% complete'))

        # Refresh dataset
        try:
            dataset = Dataset.objects.get(slug=dataset_slug)
        except Dataset.DoesNotExist:
            log.warning('Reindexing could not be completed due to Dataset being deleted, dataset_slug: %s' % dataset_slug)

            return

        dataset.column_schema = data_typer.schema 
        dataset.save()

        log.info('Finished reindex, dataset_slug: %s' % dataset_slug)

        return data_typer

    def after_return(self, status, retval, task_id, args, kwargs, einfo):
        """
        Save final status, results, etc.
        """
        from panda.models import Dataset

        log = logging.getLogger(self.name)

        try:
            dataset = Dataset.objects.get(slug=args[0])
        except Dataset.DoesNotExist:
            log.warning('Can not send reindexing notifications due to Dataset being deleted, dataset_slug: %s' % args[0])

            return

        try:
            try:
                self.send_notifications(dataset, retval, einfo)
            finally:
                # If reindex failed, clear any data that might be staged
                if dataset.current_task.status == 'FAILURE':
                    solr.delete(settings.SOLR_DATA_CORE, 'dataset_slug:%s' % args[0], commit=True)
        finally:
            dataset.unlock()

    def send_notifications(self, dataset, retval, einfo):
        """
        Send user notifications this task has finished.
        """
        task_status = dataset.current_task 

        extra_context = {
            'related_dataset': dataset
        }

        if einfo:
            if hasattr(einfo, 'traceback'):
                tb = einfo.traceback
            else:
                tb = ''.join(traceback.format_tb(einfo[2]))

            task_status.exception(
                ugettext('Reindex failed'),
                u'%s\n\nTraceback:\n%s' % (unicode(retval), tb)
            )
            
            template_prefix = 'reindex_failed'
            notification_type = 'Error'
        elif self.is_aborted():
            template_prefix = 'reindex_aborted'
            notification_type = 'Info'
        else:
            task_status.complete(ugettext('Reindex complete'))

            template_prefix = 'reindex_complete'
            extra_context['type_summary'] = retval.summarize()
            notification_type = 'Info'
        
        if task_status.creator:
            notify(
                task_status.creator,
                template_prefix,
                notification_type,
                url='#dataset/%s' % dataset.slug,
                extra_context=extra_context
            )


########NEW FILE########
__FILENAME__ = run_admin_alerts
#!/usr/bin/env python

import logging
import os

from panda.tasks.base import Task
from django.conf import settings
from django.template import Context
from livesettings import config_value

from client.utils import get_total_disk_space, get_free_disk_space
from panda.utils.mail import send_mail
from panda.utils.notifications import get_email_subject_template, get_email_body_template

class RunAdminAlertsTask(Task):
    """
    Notify administrators of anything which requires their attention (disk space, etc).
    """
    name = 'panda.tasks.cron.run_admin_alerts'

    def run(self, *args, **kwargs):
        from panda.models import UserProxy

        log = logging.getLogger(self.name)
        log.info('Running admin alerts')

        # Disk space
        root_disk = os.stat('/').st_dev
        upload_disk = os.stat(settings.MEDIA_ROOT).st_dev
        indices_disk = os.stat(settings.SOLR_DIRECTORY).st_dev

        root_disk_total = get_total_disk_space('/')
        root_disk_free = get_free_disk_space('/')
        root_disk_percent_used = 100 - (float(root_disk_free) / root_disk_total * 100)

        if upload_disk != root_disk:    
            upload_disk_total = get_total_disk_space(settings.MEDIA_ROOT)
            upload_disk_free = get_free_disk_space(settings.MEDIA_ROOT)
            upload_disk_percent_used = 100 - (float(upload_disk_free) / upload_disk_total * 100)
        else:
            upload_disk_total = None
            upload_disk_free = None
            upload_disk_percent_used = None

        if indices_disk != root_disk:
            indices_disk_total = get_total_disk_space(settings.SOLR_DIRECTORY)
            indices_disk_free = get_free_disk_space(settings.SOLR_DIRECTORY)
            indices_disk_percent_used = 100 - (float(indices_disk_free) / indices_disk_total * 100)
        else:
            indices_disk_total = None
            indices_disk_free = None
            indices_disk_percent_used = None

        notify = False

        for free in (root_disk_free, upload_disk_free, indices_disk_free):
            if free is None:
                continue
            
            if free < settings.PANDA_AVAILABLE_SPACE_WARN:
                notify = True

        if notify:
            context = Context({
                'root_disk': root_disk,
                'upload_disk': upload_disk,
                'indices_disk': indices_disk,
                'root_disk_total': root_disk_total,
                'root_disk_free': root_disk_free,
                'root_disk_percent_used': root_disk_percent_used,
                'upload_disk_total': upload_disk_total,
                'upload_disk_free': upload_disk_free,
                'upload_disk_percent_used': upload_disk_percent_used,
                'indices_disk_total': indices_disk_total,
                'indices_disk_free': indices_disk_free,
                'indices_disk_percent_used': indices_disk_percent_used,
                'settings': settings,
                'site_domain': config_value('DOMAIN', 'SITE_DOMAIN')
            })

            # Don't HTML escape plain-text emails
            context.autoescape = False

            email_subject = get_email_subject_template('disk_space_alert').render(context)
            email_message = get_email_body_template('disk_space_alert').render(context)

            recipients = UserProxy.objects.filter(is_superuser=True, is_active=True)

            send_mail(email_subject.strip(), email_message, [r.email for r in recipients])

        log.info('Finished running admin alerts')


########NEW FILE########
__FILENAME__ = run_subscriptions
#!/usr/bin/env python

import logging

from panda.tasks.base import Task
from django.conf import settings
from django.utils.timezone import now 

from panda import solr
from panda.utils.notifications import notify

class RunSubscriptionsTask(Task):
    """
    Execute all user-subscribed searches. 
    """
    name = 'panda.tasks.cron.run_subscriptions'

    def run(self, *args, **kwargs):
        from panda.models import SearchSubscription

        log = logging.getLogger(self.name)
        log.info('Running subscribed searches')

        subscriptions = SearchSubscription.objects.all()

        for sub in subscriptions:
            log.info('Running subscription: %s' % sub)

            since = sub.last_run.replace(microsecond=0, tzinfo=None)
            since = since.isoformat('T')

            sub.last_run = now()
            sub.save()
   
            solr_query = 'last_modified:[%s TO *] AND (%s)' % (since + 'Z', sub.query)

            if sub.dataset:
                solr_query += ' dataset_slug:%s' % (sub.dataset.slug)
            elif sub.category:
                dataset_slugs = sub.category.datasets.values_list('slug', flat=True)
                solr_query += ' dataset_slug:(%s)' % ' '.join(dataset_slugs)

            response = solr.query(
                settings.SOLR_DATA_CORE,
                solr_query,
                offset=0,
                limit=0
            )

            count = response['response']['numFound'] 

            log.info('Found %i new results' % count)

            if count:
                if sub.dataset:
                    url = '#dataset/%s/search/%s/%s' % (sub.dataset.slug, sub.query_url, since)
                elif sub.category:
                    url = '#search/%s/%s/%s' % (sub.category.slug, sub.query, since)
                else:
                    url = '#search/all/%s/%s' % (sub.query, since)
                    
                notify(
                    sub.user,
                    'subscription_results',
                    'info',
                    url=url,
                    extra_context={
                        'query': sub.query,
                        'query_url': sub.query_url,
                        'category': sub.category,
                        'related_dataset': sub.dataset,
                        'count': count,
                        'since': since
                    }
                )

        log.info('Finished running subscribed searches')


########NEW FILE########
__FILENAME__ = test_admin
#!/usr/bin/env python

from django.core.urlresolvers import reverse
from django.test import TransactionTestCase
from django.test.client import Client

from panda.models import UserProxy
from panda.tests import utils

class TestUserAdmin(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        self.user = utils.get_panda_user()
        
        self.client = Client()
        self.client.login(username='panda@pandaproject.net', password='panda')

    def tearDown(self):
        self.client.logout()

    def test_add_user(self):
        # Test fetching the form
        response = self.client.get(reverse('admin:panda_userproxy_add'))
        
        self.assertEqual(response.status_code, 200)

        new_user = {
            'email': 'foo@bar.com',
            'last_name': 'Barman'
        }

        # Test submitting the form
        response = self.client.post(reverse('admin:panda_userproxy_add'), new_user)

        self.assertEqual(response.status_code, 302)

        created_user = UserProxy.objects.get(username='foo@bar.com')
        self.assertEqual(created_user.last_name, 'Barman')


########NEW FILE########
__FILENAME__ = test_api_category
#!/usr/bin/env python

from django.conf import settings
from django.test import TransactionTestCase
from django.test.client import Client
from django.utils import simplejson as json

from panda.models import Category 
from panda.tests import utils

class TestAPICategories(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        settings.CELERY_ALWAYS_EAGER = True
        
        self.user = utils.get_panda_user()
        self.dataset = utils.get_test_dataset(self.user)
        self.upload = utils.get_test_data_upload(self.user, self.dataset)

        self.auth_headers = utils.get_auth_headers()

        self.client = Client()

    def test_get(self):
        category = Category.objects.get(slug='crime')

        # No datasets in category
        response = self.client.get('/api/1.0/category/%s/' % category.slug, **self.auth_headers) 

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(body['name'], 'Crime')
        self.assertEqual(body['slug'], 'crime')
        self.assertEqual(body['dataset_count'], 0)

        # One dataset in category
        category.datasets.add(self.dataset)

        response = self.client.get('/api/1.0/category/%s/' % category.slug, **self.auth_headers) 

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(body['dataset_count'], 1)


    def test_list(self):
        categories = Category.objects.all()

        # Dataset not in category
        response = self.client.get('/api/1.0/category/', data={ 'limit': 5 }, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(len(body['objects']), len(categories) + 1)
        self.assertEqual(body['meta']['total_count'], len(categories))
        self.assertEqual(body['meta']['limit'], 5)
        self.assertEqual(body['meta']['offset'], 0)
        self.assertEqual(body['meta']['next'], None)
        self.assertEqual(body['meta']['previous'], None)
        
        uncategorized = next(c for c in body['objects'] if c['slug'] == 'uncategorized')

        self.assertEqual(uncategorized['dataset_count'], 1)

        # Dataset in category
        categories[0].datasets.add(self.dataset)

        response = self.client.get('/api/1.0/category/', data={ 'limit': 5 }, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        uncategorized = next(c for c in body['objects'] if c['slug'] == 'uncategorized')

        self.assertEqual(uncategorized['dataset_count'], 0)
        

########NEW FILE########
__FILENAME__ = test_api_data
#!/usr/bin/env python

from time import sleep

from django.conf import settings
from django.test import TransactionTestCase
from django.test.client import Client
from django.utils import simplejson as json
from django.utils.timezone import now
from tastypie.bundle import Bundle
from tastypie.exceptions import BadRequest

from panda import solr
from panda.api.data import DataResource, DataValidation
from panda.models import Category, Dataset
from panda.tests import utils

class TestDataValidation(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        self.validator = DataValidation()

        self.user = utils.get_panda_user()
        self.dataset = utils.get_test_dataset(self.user)
        self.upload = utils.get_test_data_upload(self.user, self.dataset)

    def test_required_fields(self):
        bundle = Bundle(data={})

        errors = self.validator.is_valid(bundle, None)

        self.assertIn('data', errors)
        self.assertIn('required', errors['data'][0])

    def test_external_id_valid(self):
        bundle = Bundle(data={ 'external_id': 'a1_-' })
        errors = self.validator.is_valid(bundle, None)

        self.assertNotIn('external_id', errors)

    def test_external_id_invalid(self):
        bundle = Bundle(data={ 'external_id': 'no spaces' })
        errors = self.validator.is_valid(bundle, None)

        self.assertIn('external_id', errors)

class TestAPIData(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        settings.CELERY_ALWAYS_EAGER = True

        utils.setup_test_solr() 

        self.user = utils.get_panda_user()
        self.dataset = utils.get_test_dataset(self.user)
        self.upload = utils.get_test_data_upload(self.user, self.dataset)

        self.auth_headers = utils.get_auth_headers()

        self.client = Client()
    
    def test_get(self):
        self.dataset.import_data(self.user, self.upload, 0)

        response = self.client.get('/api/1.0/dataset/%s/data/' % self.dataset.slug, **self.auth_headers)
        self.assertEqual(response.status_code, 200)
        body = json.loads(response.content)

        # Returned as a list of datasets
        self.assertEqual(body['meta']['total_count'], 4)
        self.assertEqual(len(body['objects']), 4)

        datum = body['objects'][0]

        response = self.client.get('/api/1.0/dataset/%s/data/%s/' % (self.dataset.slug, datum['external_id']), **self.auth_headers)
        self.assertEqual(response.status_code, 200)
        get_result = json.loads(response.content)

        self.assertEqual(datum, get_result)

    def test_get_404(self):
        self.dataset.import_data(self.user, self.upload, 0)

        response = self.client.get('/api/1.0/dataset/%s/data/not-a-valid-id/' % self.dataset.id, **self.auth_headers)
        self.assertEqual(response.status_code, 404)

    def test_list(self):
        self.dataset.import_data(self.user, self.upload, 0)

        response = self.client.get('/api/1.0/data/', **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        # Returned as a list of datasets
        self.assertEqual(body['meta']['total_count'], 1)
        self.assertEqual(len(body['objects']), 1)
        self.assertEqual(int(body['objects'][0]['id']), self.dataset.id)
        self.assertEqual(body['objects'][0]['meta']['total_count'], 4)
        self.assertEqual(len(body['objects'][0]['objects']), 4)

        self.assertIn('data', body['objects'][0]['objects'][0])
        self.assertIn('resource_uri', body['objects'][0]['objects'][0])
        self.assertIn('external_id', body['objects'][0]['objects'][0])

    def test_get_dataset_from_kwargs(self):
        data_resource = DataResource()

        bundle = Bundle(data={})
        
        dataset = data_resource.get_dataset_from_kwargs(bundle, dataset_slug=self.dataset.slug)

        self.assertEqual(dataset.id, self.dataset.id)

    def test_get_dataset_from_kwargs_agree(self):
        data_resource = DataResource()

        bundle = Bundle(data={ 'dataset': '/api/1.0/dataset/%s/' % self.dataset.slug })
        
        dataset = data_resource.get_dataset_from_kwargs(bundle, dataset_slug=self.dataset.slug)

        self.assertEqual(dataset.id, self.dataset.id)

    def test_get_dataset_from_kwargs_conflict(self):
        data_resource = DataResource()

        second_dataset = Dataset.objects.create(
            name='Second dataset',
            creator=self.dataset.creator)

        bundle = Bundle(data={ 'dataset': '/api/1.0/dataset/%s/' % second_dataset.slug })
        
        with self.assertRaises(BadRequest):
            data_resource.get_dataset_from_kwargs(bundle, dataset_slug=self.dataset.slug)

    def test_create(self):
        self.dataset.import_data(self.user, self.upload, 0)

        new_data = {
            'data': ['5', 'A', 'B', 'C']
        }

        response = self.client.post('/api/1.0/dataset/%s/data/' % self.dataset.slug, content_type='application/json', data=json.dumps(new_data), **self.auth_headers)

        self.assertEqual(response.status_code, 201)
        body = json.loads(response.content)
        self.assertEqual(body['data'], new_data['data'])
        self.assertIn('dataset', body)
        self.assertIn('resource_uri', body)
        self.assertIn('external_id', body)

        # Refresh
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        self.assertEqual(self.dataset.row_count, 5)

    def test_create_bulk(self):
        self.dataset.import_data(self.user, self.upload, 0)

        new_data = { 'objects': [
            {
                'data': ['5', 'A', 'B', 'C']
            },
            {
                'data': ['6', 'D', 'E', 'F']
            }
        ]}

        response = self.client.put('/api/1.0/dataset/%s/data/' % self.dataset.slug, content_type='application/json', data=json.dumps(new_data), **self.auth_headers)

        self.assertEqual(response.status_code, 202)
        body = json.loads(response.content)
        self.assertEqual(len(body['objects']), 2)

        # Refresh
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        self.assertEqual(self.dataset.row_count, 6)

    def test_create_no_columns(self):
        new_data = {
            'data': ['5', 'A', 'B', 'C']
        }

        response = self.client.post('/api/1.0/dataset/%s/data/' % self.dataset.slug, content_type='application/json', data=json.dumps(new_data), **self.auth_headers)

        self.assertEqual(response.status_code, 400)
        body = json.loads(response.content)
        self.assertIn('dataset', body)

    def test_create_makes_sample(self):
        self.dataset.import_data(self.user, self.upload, 0)

        new_data = {
            'data': ['5', 'A', 'B', 'C']
        }

        response = self.client.post('/api/1.0/dataset/%s/data/' % self.dataset.slug, content_type='application/json', data=json.dumps(new_data), **self.auth_headers)

        self.assertEqual(response.status_code, 201)
        
        # Refresh
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        self.assertEqual(len(self.dataset.sample_data), 5)

    def test_created_search(self):
        self.dataset.import_data(self.user, self.upload, 0)

        new_data = {
            'data': ['5', 'Flibbity!', 'B', 'C']
        }

        response = self.client.post('/api/1.0/dataset/%s/data/' % self.dataset.slug, content_type='application/json', data=json.dumps(new_data), **self.auth_headers)

        self.assertEqual(response.status_code, 201)
        
        response = self.client.get('/api/1.0/data/?q=flibbity', **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        # Verify that the group count is correct
        self.assertEqual(body['meta']['total_count'], 1)
        self.assertEqual(len(body['objects']), 1)

    def test_create_too_few_fields(self):
        self.dataset.import_data(self.user, self.upload, 0)

        new_data = {
            'data': ['5', 'Mr.', 'PANDA']
        }

        response = self.client.post('/api/1.0/dataset/%s/data/' % self.dataset.slug, content_type='application/json', data=json.dumps(new_data), **self.auth_headers)

        self.assertEqual(response.status_code, 400)
        body = json.loads(response.content)
        self.assertIn('data', body)

    def test_create_too_many_fields(self):
        self.dataset.import_data(self.user, self.upload, 0)

        new_data = {
            'data': ['5', 'Mr.', 'PANDA', 'PANDA Project', 'PANDAs everywhere']
        }

        response = self.client.post('/api/1.0/dataset/%s/data/' % self.dataset.slug, content_type='application/json', data=json.dumps(new_data), **self.auth_headers)

        self.assertEqual(response.status_code, 400)
        body = json.loads(response.content)
        self.assertIn('data', body)

    def test_update(self):
        self.dataset.import_data(self.user, self.upload, 0)

        update_data = {
            'dataset': '/api/1.0/dataset/%s/' % self.dataset.slug,
            'data': ['5', 'A', 'B', 'C']
        }

        response = self.client.get('/api/1.0/dataset/%s/data/' % self.dataset.slug, **self.auth_headers)

        self.assertEqual(response.status_code, 200)
        body = json.loads(response.content)

        data = body['objects'][0]

        response = self.client.put('/api/1.0/dataset/%s/data/%s/' % (self.dataset.slug, data['external_id']), content_type='application/json', data=json.dumps(update_data), **self.auth_headers)

        self.assertEqual(response.status_code, 202)
        body = json.loads(response.content)
        self.assertEqual(body['data'], update_data['data'])
        self.assertEqual(body['dataset'], data['dataset'])
        self.assertEqual(body['resource_uri'], data['resource_uri'])
        self.assertEqual(body['external_id'], data['external_id'])

    def test_update_bulk(self):
        self.dataset.import_data(self.user, self.upload, 0)

        update_data = {
            'dataset': '/api/1.0/dataset/%s/' % self.dataset.slug,
            'data': ['5', 'Flibbity!', 'B', 'C']
        }

        response = self.client.get('/api/1.0/data/', **self.auth_headers)

        self.assertEqual(response.status_code, 200)
        body = json.loads(response.content)

        # Dataset objects were returned
        data = body['objects'][0]['objects'][0]

        response = self.client.put('/api/1.0/dataset/%s/data/%s/' % (self.dataset.slug, data['external_id']), content_type='application/json', data=json.dumps(update_data), **self.auth_headers)

        self.assertEqual(response.status_code, 202)

        response = self.client.get('/api/1.0/data/?q=flibbity', **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        # Verify that the group count is correct
        self.assertEqual(body['meta']['total_count'], 1)
        self.assertEqual(len(body['objects']), 1)

    def test_delete(self):
        self.dataset.import_data(self.user, self.upload, 0)

        response = self.client.get('/api/1.0/dataset/%s/data/' % self.dataset.slug, **self.auth_headers)

        self.assertEqual(response.status_code, 200)
        body = json.loads(response.content)

        data = body['objects'][0]

        response = self.client.delete('/api/1.0/dataset/%s/data/%s/' % (self.dataset.slug, data['external_id']), content_type='application/json', **self.auth_headers)

        self.assertEqual(response.status_code, 204)

        # Refresh
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        self.assertEqual(self.dataset.row_count, 3)

    def test_delete_list(self):
        self.dataset.import_data(self.user, self.upload, 0)

        response = self.client.delete('/api/1.0/dataset/%s/data/' % self.dataset.slug, **self.auth_headers)

        self.assertEqual(response.status_code, 204)

        # Refresh
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        self.assertEqual(self.dataset.row_count, 0)

    def test_deleted_search(self):
        self.dataset.import_data(self.user, self.upload, 0)

        response = self.client.get('/api/1.0/dataset/%s/data/' % self.dataset.slug, **self.auth_headers)

        self.assertEqual(response.status_code, 200)
        body = json.loads(response.content)

        # Dataset objects were returned
        data = body['objects'][0]

        response = self.client.delete('/api/1.0/dataset/%s/data/%s/' % (self.dataset.slug, data['external_id']), content_type='application/json', **self.auth_headers)

        self.assertEqual(response.status_code, 204)

        response = self.client.get('/api/1.0/data/?q=%s' % data['data'][0], **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        # Verify that the group count is correct
        self.assertEqual(body['meta']['total_count'], 0)
        self.assertEqual(len(body['objects']), 0)

    def test_post_detail(self):
        new_data = {
            'dataset': '/api/1.0/dataset/%s/' % self.dataset.slug,
            'data': ['5', 'A', 'B', 'C']
        }

        response = self.client.post('/api/1.0/dataset/%s/data/im-a-fake-uuid/' % self.dataset.slug, content_type='application/json', data=json.dumps(new_data), **self.auth_headers)

        self.assertEqual(response.status_code, 501)

    def test_search(self):
        self.dataset.import_data(self.user, self.upload, 0)

        # Import second dataset so we can make sure both match 
        second_dataset = Dataset.objects.create(
            name='Second dataset',
            creator=self.dataset.creator)

        second_dataset.import_data(self.user, self.upload, 0)

        response = self.client.get('/api/1.0/data/?q=Christopher', **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        # Verify that the group count is correct
        self.assertEqual(body['meta']['total_count'], 2)
        self.assertEqual(len(body['objects']), 2)

        # Verify that each matched dataset includes one result
        for result_dataset in body['objects']:
            self.assertEqual(result_dataset['meta']['total_count'], 1)
            self.assertEqual(len(result_dataset['objects']), 1)

            db_dataset = Dataset.objects.get(id=result_dataset['id'])
            
            self.assertEqual(result_dataset['name'], db_dataset.name)
            self.assertEqual(result_dataset['row_count'], db_dataset.row_count)
            self.assertEqual(result_dataset['column_schema'], db_dataset.column_schema)

            self.assertEqual(result_dataset['objects'][0]['data'][1], 'Christopher')
            self.assertIn('resource_uri', result_dataset['objects'][0])
            self.assertIn('external_id', result_dataset['objects'][0])

    def test_search_category(self):
        category = Category.objects.get(slug='politics')

        self.dataset.import_data(self.user, self.upload, 0)
        self.dataset = Dataset.objects.get(id=self.dataset.id)
        category.datasets.add(self.dataset)

        # Import second dataset so we can make sure both match 
        second_dataset = Dataset.objects.create(
            name='Second dataset',
            creator=self.dataset.creator)

        second_dataset.import_data(self.user, self.upload, 0)

        response = self.client.get('/api/1.0/data/?q=Christopher&category=politics', **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        # Verify that the group count is correct
        self.assertEqual(body['meta']['total_count'], 1)
        self.assertEqual(len(body['objects']), 1)

        # Verify that each matched dataset includes one result
        result_dataset = body['objects'][0]
        self.assertEqual(result_dataset['meta']['total_count'], 1)
        self.assertEqual(len(result_dataset['objects']), 1)

        self.assertEqual(result_dataset['name'], self.dataset.name)
        self.assertEqual(result_dataset['row_count'], self.dataset.row_count)
        self.assertEqual(result_dataset['column_schema'], self.dataset.column_schema)

        self.assertEqual(result_dataset['objects'][0]['data'][1], 'Christopher')
        self.assertIn('resource_uri', result_dataset['objects'][0])
        self.assertIn('external_id', result_dataset['objects'][0])

    def test_search_since(self):
        self.dataset.import_data(self.user, self.upload, 0)

        # Import second dataset so we can make sure only one matches 
        second_dataset = Dataset.objects.create(
            name='Second dataset',
            creator=self.dataset.creator)

        second_dataset.import_data(self.user, self.upload, 0)

        sleep(1)
        between_time = now().replace(microsecond=0, tzinfo=None)
        between_time = between_time.isoformat('T') 
        
        # Import second dataset twice, to verify only one is matched
        second_dataset.import_data(self.user, self.upload, 0)

        response = self.client.get('/api/1.0/data/?q=Christopher&since=%s' % between_time, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        # Verify that the group count is correct
        self.assertEqual(body['meta']['total_count'], 1)
        self.assertEqual(len(body['objects']), 1)

        # Verify that each matched dataset includes one result
        result_dataset = body['objects'][0]
        self.assertEqual(result_dataset['meta']['total_count'], 1)
        self.assertEqual(len(result_dataset['objects']), 1)

    def test_search_stale_dataset(self):
        self.dataset.import_data(self.user, self.upload, 0)
        self.dataset.update_full_text()

        # Import second dataset so we can make sure both match 
        second_dataset = Dataset.objects.create(
            name='Second dataset',
            creator=self.dataset.creator)

        second_dataset.import_data(self.user, self.upload, 0)
        second_dataset.update_full_text()

        # Manually delete second dataset to simulate an integrity issue
        from django.db import connection, transaction
        cursor = connection.cursor()

        cursor.execute("DELETE FROM panda_dataset WHERE slug='%s'" % second_dataset.slug)
        transaction.commit_unless_managed()

        # Verify Solr data still exists
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'dataset_slug:%s' % self.dataset.slug)['response']['numFound'], 4)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'dataset_slug:%s' % second_dataset.slug)['response']['numFound'], 4)
        self.assertEqual(solr.query(settings.SOLR_DATASETS_CORE, 'slug:%s' % self.dataset.slug)['response']['numFound'], 1)
        self.assertEqual(solr.query(settings.SOLR_DATASETS_CORE, 'slug:%s' % second_dataset.slug)['response']['numFound'], 1)

        # Execute search, which should invoke purge as a side-effect
        response = self.client.get('/api/1.0/data/?q=Christopher', **self.auth_headers)
        self.assertEqual(response.status_code, 200)

        # Verify Solr data has been purged
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'dataset_slug:%s' % self.dataset.slug)['response']['numFound'], 4)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'dataset_slug:%s' % second_dataset.slug)['response']['numFound'], 0)
        self.assertEqual(solr.query(settings.SOLR_DATASETS_CORE, 'slug:%s' % self.dataset.slug)['response']['numFound'], 1)
        self.assertEqual(solr.query(settings.SOLR_DATASETS_CORE, 'slug:%s' % second_dataset.slug)['response']['numFound'], 0)

        body = json.loads(response.content)

        # Verify that the group count is correct
        self.assertEqual(body['meta']['total_count'], 1)
        self.assertEqual(len(body['objects']), 1)

        # Verify that each matched dataset includes one result
        result_dataset = body['objects'][0]
        self.assertEqual(result_dataset['slug'], self.dataset.slug)
        self.assertEqual(result_dataset['meta']['total_count'], 1)
        self.assertEqual(len(result_dataset['objects']), 1)

    def test_search_meta(self):
        self.dataset.import_data(self.user, self.upload, 0)

        # Import second dataset so we can make sure both match 
        second_dataset = Dataset.objects.create(
            name='Second dataset',
            creator=self.dataset.creator)

        second_dataset.import_data(self.user, self.upload, 0)

        response = self.client.get('/api/1.0/data/?q=Ryan&limit=1', **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        # Verify that the group count is correct
        self.assertEqual(body['meta']['limit'], 1)
        self.assertEqual(body['meta']['offset'], 0)
        self.assertEqual(body['meta']['total_count'], 2)
        self.assertIs(body['meta']['previous'], None)
        self.assertIsNot(body['meta']['next'], None)
        self.assertEqual(len(body['objects']), 1)

    def test_search_boolean_query(self):
        self.dataset.import_data(self.user, self.upload, 0)

        response = self.client.get('/api/1.0/data/?q=Brian+and+Tribune', **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(body['meta']['total_count'], 1)
        self.assertEqual(len(body['objects']), 1)

    def test_search_unauthorized(self):
        response = self.client.get('/api/1.0/data/?q=Christopher')

        self.assertEqual(response.status_code, 401)   

    def test_export_data(self):
        self.dataset.import_data(self.user, self.upload, 0)

        response = self.client.get('/api/1.0/data/?q=joseph&export=true', **self.auth_headers)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content, '"Export queued."')


########NEW FILE########
__FILENAME__ = test_api_dataset
#!/usr/bin/env python

from time import sleep

from django.conf import settings
from django.test import TestCase, TransactionTestCase
from django.test.client import Client
from django.utils import simplejson as json
from django.utils.timezone import now
from tastypie.bundle import Bundle

from panda import solr
from panda.api.datasets import DatasetValidation
from panda.models import Category, Dataset
from panda.tests import utils

class TestDatasetValidation(TestCase):
    def setUp(self):
        self.validator = DatasetValidation()

    def test_required_fields(self):
        bundle = Bundle(data={})

        errors = self.validator.is_valid(bundle)

        self.assertIn('name', errors)

class TestAPIDataset(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        settings.CELERY_ALWAYS_EAGER = True

        utils.setup_test_solr() 

        self.user = utils.get_panda_user()
        self.dataset = utils.get_test_dataset(self.user)
        self.upload = utils.get_test_data_upload(self.user, self.dataset)

        self.auth_headers = utils.get_auth_headers() 

        self.client = Client()

    def test_get(self):
        # Import so that there will be a task object
        self.dataset.import_data(self.user, self.upload, 0)

        # Refetch dataset so that attributes will be updated
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        response = self.client.get('/api/1.0/dataset/%s/' % self.dataset.slug, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(body['name'], self.dataset.name)
        self.assertEqual(body['description'], self.dataset.description)
        self.assertEqual(body['row_count'], self.dataset.row_count)
        self.assertEqual(body['sample_data'], self.dataset.sample_data)
        self.assertEqual(body['column_schema'], self.dataset.column_schema)
        self.assertEqual(body['creator']['email'], self.dataset.creator.email)

        task_response = self.client.get('/api/1.0/task/%i/' % self.dataset.current_task.id, **self.auth_headers)

        self.assertEqual(task_response.status_code, 200)

        self.assertEqual(body['current_task'], json.loads(task_response.content))

        self.assertEqual(len(body['related_uploads']), 0)
        self.assertEqual(len(body['data_uploads']), 1)
        self.assertEqual(body['initial_upload'], '/api/1.0/data_upload/%i/' % self.dataset.initial_upload.id)

    def test_get_unauthorized(self):
        response = self.client.get('/api/1.0/dataset/%s/' % self.dataset.slug)

        self.assertEqual(response.status_code, 401)

    def test_get_inactive(self):
        self.user.is_active = False
        self.user.save()

        response = self.client.get('/api/1.0/dataset/%s/' % self.dataset.slug, **self.auth_headers)

        self.assertEqual(response.status_code, 401)

        self.user.is_active = True
        self.user.save()

    def test_list(self):
        response = self.client.get('/api/1.0/dataset/', data={ 'limit': 5 }, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(len(body['objects']), 1)
        self.assertEqual(body['meta']['total_count'], 1)
        self.assertEqual(body['meta']['limit'], 5)
        self.assertEqual(body['meta']['offset'], 0)
        self.assertEqual(body['meta']['next'], None)
        self.assertEqual(body['meta']['previous'], None)

    def test_list_filtered_by_category_miss(self):
        response = self.client.get('/api/1.0/dataset/', data={ 'category': 'crime' }, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(len(body['objects']), 0)
        self.assertEqual(body['meta']['total_count'], 0)

    def test_list_filtered_by_category_hit(self):
        category = Category.objects.get(slug='crime')
        self.dataset.categories.add(category)
        self.dataset.save()
        self.dataset.update_full_text()

        response = self.client.get('/api/1.0/dataset/', data={ 'category': 'crime' }, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(len(body['objects']), 1)
        self.assertEqual(body['meta']['total_count'], 1)
        self.assertEqual(int(body['objects'][0]['id']), self.dataset.id)

    def test_create_post(self):
        new_dataset = {
            'name': 'New dataset!',
            'description': 'Its got yummy data!'
        }

        response = self.client.post('/api/1.0/dataset/', content_type='application/json', data=json.dumps(new_dataset), **self.auth_headers)

        self.assertEqual(response.status_code, 201)

        body = json.loads(response.content)

        self.assertEqual(body['name'], 'New dataset!')
        self.assertEqual(body['slug'], 'new-dataset')
        self.assertEqual(body['description'], 'Its got yummy data!')
        self.assertEqual(body['row_count'], None)
        self.assertEqual(body['column_schema'], None)
        self.assertEqual(body['sample_data'], None)
        self.assertEqual(body['current_task'], None)
        self.assertEqual(body['initial_upload'], None)
        self.assertEqual(body['related_uploads'], [])
        self.assertEqual(body['data_uploads'], [])

        new_dataset = Dataset.objects.get(id=body['id'])

        self.assertEqual(new_dataset.name, 'New dataset!')
        self.assertEqual(new_dataset.description, 'Its got yummy data!')
        self.assertEqual(new_dataset.row_count, None)
        self.assertEqual(new_dataset.column_schema, None)
        self.assertEqual(new_dataset.sample_data, None)
        self.assertEqual(new_dataset.current_task, None)
        self.assertEqual(new_dataset.initial_upload, None)
        self.assertEqual(new_dataset.related_uploads.count(), 0)
        self.assertEqual(new_dataset.data_uploads.count(), 0)

    def test_create_post_slug(self):
        # Verify that new slugs are NOT created via POST.
        new_dataset = {
            'slug': 'new-slug',
            'name': 'New dataset!',
            'description': 'Its got yummy data!'
        }

        response = self.client.post('/api/1.0/dataset/', content_type='application/json', data=json.dumps(new_dataset), **self.auth_headers)

        self.assertEqual(response.status_code, 201)

        body = json.loads(response.content)

        self.assertEqual(body['slug'], 'new-slug')

        new_dataset = Dataset.objects.get(id=body['id'])

        self.assertEqual(new_dataset.slug, 'new-slug')

    def test_create_put(self):
        new_dataset = {
            'name': 'New dataset!',
            'description': 'Its got yummy data!'
        }

        response = self.client.put('/api/1.0/dataset/new-id/', content_type='application/json', data=json.dumps(new_dataset), **self.auth_headers)

        self.assertEqual(response.status_code, 201)

        body = json.loads(response.content)

        self.assertEqual(body['name'], 'New dataset!')
        self.assertEqual(body['slug'], 'new-id')
        self.assertEqual(body['description'], 'Its got yummy data!')
        self.assertEqual(body['row_count'], None)
        self.assertEqual(body['column_schema'], None)
        self.assertEqual(body['sample_data'], None)
        self.assertEqual(body['current_task'], None)
        self.assertEqual(body['initial_upload'], None)
        self.assertEqual(body['data_uploads'], [])

        new_dataset = Dataset.objects.get(id=body['id'])

        self.assertEqual(new_dataset.name, 'New dataset!')
        self.assertEqual(new_dataset.slug, 'new-id')
        self.assertEqual(new_dataset.description, 'Its got yummy data!')
        self.assertEqual(new_dataset.row_count, None)
        self.assertEqual(new_dataset.column_schema, None)
        self.assertEqual(new_dataset.sample_data, None)
        self.assertEqual(new_dataset.current_task, None)
        self.assertEqual(new_dataset.initial_upload, None)
        self.assertEqual(new_dataset.data_uploads.count(), 0)

    def test_create_put_twice(self):
        new_dataset = {
            'name': 'New dataset!',
            'description': 'Its got yummy data!'
        }

        response = self.client.put('/api/1.0/dataset/new-slug/', content_type='application/json', data=json.dumps(new_dataset), **self.auth_headers)

        self.assertEqual(response.status_code, 201)

        update_dataset = {
            'name': 'Updated dataset!'
        }
        
        body = json.loads(response.content)

        self.assertEqual(body['name'], 'New dataset!')
        self.assertEqual(body['slug'], 'new-slug')
        dataset_id = body['id']

        response = self.client.put('/api/1.0/dataset/new-slug/', content_type='application/json', data=json.dumps(update_dataset), **self.auth_headers)

        self.assertEqual(response.status_code, 202)

        body = json.loads(response.content)

        self.assertEqual(body['name'], 'Updated dataset!')
        self.assertEqual(body['slug'], 'new-slug')
        self.assertEqual(body['id'], dataset_id)

        # One dataset is created by setup
        self.assertEqual(Dataset.objects.all().count(), 2)

    def test_put_different_slug(self):
        new_dataset = {
            'name': 'New dataset!',
            'description': 'Its got yummy data!'
        }

        response = self.client.put('/api/1.0/dataset/new-slug/', content_type='application/json', data=json.dumps(new_dataset), **self.auth_headers)

        self.assertEqual(response.status_code, 201)

        update_dataset = {
            'slug': 'changed-slug',
            'name': 'Updated dataset!'
        }

        response = self.client.put('/api/1.0/dataset/new-slug/', content_type='application/json', data=json.dumps(update_dataset), **self.auth_headers)

        self.assertEqual(response.status_code, 202)

        body = json.loads(response.content)

        self.assertEqual(body['slug'], 'new-slug')
        
        new_dataset = Dataset.objects.get(id=body['id'])

        self.assertEqual(new_dataset.slug, 'new-slug')

    def test_create_as_new_user(self):
        new_user = {
            'email': 'tester@tester.com',
            'password': 'test',
            'first_name': 'Testy',
            'last_name': 'McTester'
        }

        response = self.client.post('/api/1.0/user/', content_type='application/json', data=json.dumps(new_user), **utils.get_auth_headers('panda@pandaproject.net'))

        self.assertEqual(response.status_code, 201)
        
        new_dataset = {
            'name': 'New dataset!',
            'description': 'Its got yummy data!'
        }

        response = self.client.post('/api/1.0/dataset/', content_type='application/json', data=json.dumps(new_dataset), **utils.get_auth_headers('tester@tester.com'))

        self.assertEqual(response.status_code, 201)        

    def test_update_readonly(self):
        response = self.client.get('/api/1.0/dataset/%s/' % self.dataset.slug, content_type='application/json', **utils.get_auth_headers('panda@pandaproject.net'))

        data = json.loads(response.content)

        row_count = data['row_count']
        data['row_count'] = 2717

        # Fixes issue with deserialization of users embedded in data_uploads -- is this a bug?
        data['data_uploads'] = [du['resource_uri'] for du in data['data_uploads']]

        response = self.client.put('/api/1.0/dataset/%s/' % self.dataset.slug, content_type='application/json', data=json.dumps(data), **utils.get_auth_headers('panda@pandaproject.net'))

        new_data = json.loads(response.content)

        self.assertEqual(new_data['row_count'], row_count)

        # Refresh
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        self.assertEqual(self.dataset.row_count, row_count)

    def test_create_with_schema(self):
        new_dataset = {
            'name': 'New dataset!'
        }

        response = self.client.post('/api/1.0/dataset/?columns=foo,bar,baz&typed_columns=True,,False&column_types=int,unicode,date', content_type='application/json', data=json.dumps(new_dataset), **self.auth_headers)

        self.assertEqual(response.status_code, 201)

        body = json.loads(response.content)

        self.assertEqual([c['name'] for c in body['column_schema']], ['foo', 'bar', 'baz'])
        self.assertEqual([c['indexed'] for c in body['column_schema']], [True, False, False])
        self.assertEqual([c['type'] for c in body['column_schema']], ['int', 'unicode', 'date'])

        new_dataset = Dataset.objects.get(id=body['id'])

        self.assertEqual([c['name'] for c in new_dataset.column_schema], ['foo', 'bar', 'baz'])
        self.assertEqual([c['indexed'] for c in new_dataset.column_schema], [True, False, False])
        self.assertEqual([c['type'] for c in new_dataset.column_schema], ['int', 'unicode', 'date'])

    def test_import_data(self):
        response = self.client.get('/api/1.0/dataset/%s/import/%i/' % (self.dataset.slug, self.upload.id), **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertNotEqual(body['current_task'], None)
        self.assertEqual(body['current_task']['task_name'], 'panda.tasks.import.csv')
        
        # Refetch dataset so that attributes will be updated
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        self.assertEqual(self.dataset.row_count, 4)
        self.assertEqual([c['name'] for c in self.dataset.column_schema], self.upload.columns)
        self.assertEqual(self.dataset.initial_upload, self.upload)
        self.assertEqual(self.dataset.sample_data, self.upload.sample_data)

        task = self.dataset.current_task

        self.assertNotEqual(task, None)
        self.assertEqual(task.status, 'SUCCESS')
        self.assertEqual(task.task_name, 'panda.tasks.import.csv')
        self.assertNotEqual(task.start, None)
        self.assertNotEqual(task.end, None)
        self.assertEqual(task.traceback, None)

        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'Christopher')['response']['numFound'], 1)

    def test_import_data_locked(self):
        # Note - testing a race condition here, should find a better way
        response = self.client.get('/api/1.0/dataset/%s/import/%i/' % (self.dataset.slug, self.upload.id), **self.auth_headers)
        response = self.client.get('/api/1.0/dataset/%s/import/%i/' % (self.dataset.slug, self.upload.id), **self.auth_headers)

        self.assertEqual(response.status_code, 403)

    def test_import_data_unauthorized(self):
        response = self.client.get('/api/1.0/dataset/%s/import/%i/' % (self.dataset.slug, self.upload.id))

        self.assertEqual(response.status_code, 401)

    def test_reindex_data(self):
        response = self.client.get('/api/1.0/dataset/%s/import/%i/' % (self.dataset.slug, self.upload.id), **self.auth_headers)

        response = self.client.get('/api/1.0/dataset/%s/reindex/?typed_columns=True,False,False,False' % (self.dataset.slug), **self.auth_headers)

        self.assertEqual(response.status_code, 200)
        
        # Refetch dataset so that attributes will be updated
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        self.assertEqual(self.dataset.row_count, 4)
        self.assertEqual([c['name'] for c in self.dataset.column_schema], self.upload.columns)
        self.assertEqual(self.dataset.initial_upload, self.upload)
        self.assertEqual(self.dataset.sample_data, self.upload.sample_data)

        task = self.dataset.current_task

        self.assertNotEqual(task, None)
        self.assertEqual(task.status, 'SUCCESS')
        self.assertEqual(task.task_name, 'panda.tasks.reindex')
        self.assertNotEqual(task.start, None)
        self.assertNotEqual(task.end, None)
        self.assertEqual(task.traceback, None)

        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_int_id:3')['response']['numFound'], 1)

    def test_reindex_data_no_data(self):
        response = self.client.get('/api/1.0/dataset/%s/reindex/' % (self.dataset.slug), **self.auth_headers)

        self.assertEqual(response.status_code, 400)

    def test_reindex_data_invalid_columns(self):
        response = self.client.get('/api/1.0/dataset/%s/import/%i/' % (self.dataset.slug, self.upload.id), **self.auth_headers)

        response = self.client.get('/api/1.0/dataset/%s/reindex/?typed_columns=True,False,False' % (self.dataset.slug), **self.auth_headers)

        self.assertEqual(response.status_code, 400)

    def test_export_data(self):
        self.dataset.import_data(self.user, self.upload, 0)

        response = self.client.get('/api/1.0/dataset/%s/export/' % self.dataset.slug, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertNotEqual(body['current_task'], None)
        self.assertEqual(body['current_task']['task_name'], 'panda.tasks.export.csv')
        
        # Refetch dataset so that attributes will be updated
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        task = self.dataset.current_task

        self.assertNotEqual(task, None)
        self.assertEqual(task.status, 'SUCCESS')
        self.assertEqual(task.task_name, 'panda.tasks.export.csv')
        self.assertNotEqual(task.start, None)
        self.assertNotEqual(task.end, None)
        self.assertEqual(task.traceback, None)

    def test_get_datum(self):
        self.dataset.import_data(self.user, self.upload, 0)

        # Refetch dataset so that attributes will be updated
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        # Get id of a datum in Solr
        datum = solr.query(settings.SOLR_DATA_CORE, 'dataset_slug:%s AND Brian' % self.dataset.slug)['response']['docs'][0]

        response = self.client.get('/api/1.0/dataset/%s/data/%s/' % (self.dataset.slug, datum['external_id']), **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)
        
        # Verify that correct attributes of the dataset are attached
        self.assertEqual(body['external_id'], datum['external_id'])
        self.assertEqual(body['dataset'], '/api/1.0/dataset/%s/' % self.dataset.slug)

    def test_get_data(self):
        self.dataset.import_data(self.user, self.upload, 0)

        # Refetch dataset so that attributes will be updated
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        # Import second dataset so we can make sure only one is matched
        second_dataset = Dataset.objects.create(
            name='Second dataset',
            creator=self.dataset.creator)

        # Bending a rules a bit since this upload is associated with the other dataset
        second_dataset.import_data(self.user, self.upload, 0)

        response = self.client.get('/api/1.0/dataset/%s/data/' % self.dataset.slug, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)
        
        # Verify that correct attributes of the dataset are attached
        self.assertEqual(int(body['id']), self.dataset.id)
        self.assertEqual(body['name'], self.dataset.name)
        self.assertEqual(body['row_count'], self.dataset.row_count)
        self.assertEqual(body['column_schema'], self.dataset.column_schema)

        # Test that only one dataset was matched
        self.assertEqual(body['meta']['total_count'], 4)
        self.assertEqual(len(body['objects']), 4)
        self.assertEqual(body['objects'][0]['data'][1], 'Brian')

    def test_search_data(self):
        self.dataset.import_data(self.user, self.upload, 0)

        # Refetch dataset so that attributes will be updated
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        # Import second dataset so we can make sure only one is matched
        second_dataset = Dataset.objects.create(
            name='Second dataset',
            creator=self.dataset.creator)

        # Bending the rules again...
        second_dataset.import_data(self.user, self.upload, 0)

        response = self.client.get('/api/1.0/dataset/%s/data/?q=Christopher' % self.dataset.slug, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)
        
        # Verify that correct attributes of the dataset are attached
        self.assertEqual(int(body['id']), self.dataset.id)
        self.assertEqual(body['name'], self.dataset.name)
        self.assertEqual(body['row_count'], self.dataset.row_count)
        self.assertEqual(body['column_schema'], self.dataset.column_schema)

        # Test that only one dataset was matched
        self.assertEqual(body['meta']['total_count'], 1)
        self.assertEqual(len(body['objects']), 1)
        self.assertEqual(body['objects'][0]['data'][1], 'Christopher')

    def test_search_data_limit(self):
        self.dataset.import_data(self.user, self.upload, 0)

        response = self.client.get('/api/1.0/dataset/%s/data/?q=Tribune&limit=1' % self.dataset.slug, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(body['meta']['limit'], 1)
        self.assertEqual(body['meta']['offset'], 0)
        self.assertEqual(body['meta']['total_count'], 2)
        self.assertIs(body['meta']['previous'], None)
        self.assertIsNot(body['meta']['next'], None)
        self.assertEqual(len(body['objects']), 1)

    def test_search_data_unauthorized(self):
        response = self.client.get('/api/1.0/dataset/%s/data/?q=Christopher' % self.dataset.slug)

        self.assertEqual(response.status_code, 401)

    def test_search_data_since(self):
        self.dataset.import_data(self.user, self.upload, 0)

        # Refetch dataset so that attributes will be updated
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        # Import second dataset so we can make sure only one is matched
        second_dataset = Dataset.objects.create(
            name='Second dataset',
            creator=self.dataset.creator)

        second_dataset.import_data(self.user, self.upload, 0)

        sleep(1)
        between_time = now().replace(microsecond=0, tzinfo=None)
        between_time = between_time.isoformat('T')

        # Import 2nd dataset again, to verify only one is matched
        second_dataset.import_data(self.user, self.upload, 0)

        response = self.client.get('/api/1.0/dataset/%s/data/?q=Christopher&since=%s' % (second_dataset.slug, between_time), **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)
        
        # Verify that correct attributes of the dataset are attached
        self.assertEqual(int(body['id']), second_dataset.id)

        # Test that only one dataset and one import was matched
        self.assertEqual(body['meta']['total_count'], 1)
        self.assertEqual(len(body['objects']), 1)

    def test_search_datasets(self):
        second_dataset = Dataset.objects.create(
            name='Second dataset',
            description='contributors',
            creator=self.dataset.creator)
        second_dataset.update_full_text()

        # Should match both
        response = self.client.get('/api/1.0/dataset/?q=contributors', **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(body['meta']['total_count'], 2)
        self.assertEqual(len(body['objects']), 2)

        # Should match only the second dataset
        response = self.client.get('/api/1.0/dataset/?q=second', **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(body['meta']['total_count'], 1)
        self.assertEqual(len(body['objects']), 1)
        self.assertEqual(int(body['objects'][0]['id']), second_dataset.id)

    def test_search_datasets_simple(self):
        response = self.client.get('/api/1.0/dataset/?q=contributors&simple=true', **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(body['meta']['total_count'], 1)
        self.assertEqual(len(body['objects']), 1)
        self.assertEqual(int(body['objects'][0]['id']), self.dataset.id)
        self.assertNotIn('related_uploads', body['objects'][0])
        self.assertNotIn('data_uploads', body['objects'][0])
        self.assertNotIn('sample_data', body['objects'][0])
        self.assertNotIn('current_task', body['objects'][0])

    def test_delete(self):
        dataset_id = self.dataset.id
        response = self.client.delete('/api/1.0/dataset/%s/' % (self.dataset.slug), **self.auth_headers)

        self.assertEqual(response.status_code, 204)

        response = self.client.get('/api/1.0/dataset/%s/' % (self.dataset.slug), **self.auth_headers)

        self.assertEqual(response.status_code, 404)

        with self.assertRaises(Dataset.DoesNotExist):
            Dataset.objects.get(id=dataset_id)

    def test_creator_email_filter(self):
        response = self.client.get('/api/1.0/dataset/', data={ 'creator_email': self.user.email }, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(len(body['objects']), 1)
        self.assertEqual(body['meta']['total_count'], 1)

        response = self.client.get('/api/1.0/dataset/', data={ 'creator_email': utils.get_admin_user().email }, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(len(body['objects']), 0)
        self.assertEqual(body['meta']['total_count'], 0)


########NEW FILE########
__FILENAME__ = test_api_data_upload
#!/usr/bin/env python

import os.path

from django.conf import settings
from django.test import TransactionTestCase
from django.test.client import Client
from django.utils import simplejson as json

from panda.models import DataUpload
from panda.tests import utils

class TestAPIDataUpload(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        self.user = utils.get_panda_user()
        self.dataset = utils.get_test_dataset(self.user)
        self.upload = utils.get_test_data_upload(self.user, self.dataset)

        self.auth_headers = utils.get_auth_headers()

        self.client = Client()

    def test_get(self):
        response = self.client.get('/api/1.0/data_upload/%i/' % self.upload.id, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(body['filename'], self.upload.filename)
        self.assertEqual(body['original_filename'], self.upload.original_filename)
        self.assertEqual(body['size'], self.upload.size)
        self.assertEqual(body['creator']['resource_uri'], '/api/1.0/user/%i/' % self.user.id)
        self.assertNotEqual(body['creation_date'], None)
        self.assertEqual(body['dataset'], '/api/1.0/dataset/%s/' % self.dataset.slug)
        self.assertEqual(body['data_type'], 'csv')
        self.assertEqual(body['columns'], ['id', 'first_name', 'last_name', 'employer'])
        self.assertEqual(len(body['sample_data']), 4)
        self.assertEqual(body['sample_data'][0], ['1', 'Brian', 'Boyer', 'Chicago Tribune'])

    def test_get_unauthorized(self):
        response = self.client.get('/api/1.0/data_upload/%i/' % self.upload.id)

        self.assertEqual(response.status_code, 401)

    def test_list(self):
        response = self.client.get('/api/1.0/data_upload/', data={ 'limit': 5 }, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(len(body['objects']), 1)
        self.assertEqual(body['meta']['total_count'], 1)
        self.assertEqual(body['meta']['limit'], 5)
        self.assertEqual(body['meta']['offset'], 0)
        self.assertEqual(body['meta']['next'], None)
        self.assertEqual(body['meta']['previous'], None)

    def test_create_denied(self):
        new_upload = {
            'filename': 'test.csv',
            'original_filename': 'test.csv',
            'size': 20
        }

        response = self.client.post('/api/1.0/data_upload/', content_type='application/json', data=json.dumps(new_upload), **self.auth_headers)

        self.assertEqual(response.status_code, 405)

    def test_download(self):
        response = self.client.get('/api/1.0/data_upload/%i/download/' % self.upload.id, **self.auth_headers)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Disposition'], 'attachment; filename=%s' % self.upload.original_filename)
        self.assertEqual(int(response['Content-Length']), self.upload.size)

        with open(os.path.join(settings.MEDIA_ROOT, utils.TEST_DATA_FILENAME)) as f:
            self.assertEqual(f.read(), response.content)

    def test_download_unauthorized(self):
        response = self.client.get('/api/1.0/data_upload/%i/download/' % self.upload.id)

        self.assertEqual(response.status_code, 401)

    def test_upload_file(self):
        with open(os.path.join(settings.MEDIA_ROOT, utils.TEST_DATA_FILENAME)) as f:
            response = self.client.post('/data_upload/', data={ 'file': f, 'dataset_slug': self.dataset.slug }, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)
        
        self.assertEqual(body['success'], True)

        upload = DataUpload.objects.get(id=body['id'])

        self.assertEqual(body['original_filename'], upload.original_filename)
        self.assertEqual(body['size'], os.path.getsize(os.path.join(settings.MEDIA_ROOT, utils.TEST_DATA_FILENAME)))
        self.assertEqual(body['size'], upload.size)
        self.assertEqual(body['creator']['resource_uri'], '/api/1.0/user/%i/' % self.user.id)

    def test_upload_unauthorized(self):
        with open(os.path.join(settings.MEDIA_ROOT, utils.TEST_DATA_FILENAME)) as f:
            response = self.client.post('/data_upload/', data={ 'file': f })

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(body['success'], False)
        self.assertEqual(body['forbidden'], True)

    def test_delete(self):
        path = self.upload.get_path()
        self.assertEqual(os.path.isfile(path), True)

        response = self.client.delete('/api/1.0/data_upload/%i/' % self.upload.id, **self.auth_headers)

        self.assertEqual(response.status_code, 204)
        self.assertEqual(os.path.exists(path), False)


########NEW FILE########
__FILENAME__ = test_api_export
#!/usr/bin/env python

import os

from django.conf import settings
from django.test import TransactionTestCase
from django.test.client import Client

from panda.models import Export
from panda.tests import utils

class TestAPIExport(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        settings.CELERY_ALWAYS_EAGER = True
        
        utils.setup_test_solr()
        
        self.user = utils.get_panda_user()
        self.dataset = utils.get_test_dataset(self.user)
        self.upload = utils.get_test_data_upload(self.user, self.dataset)

        self.dataset.import_data(self.user, self.upload, 0)

        self.auth_headers = utils.get_auth_headers()

        self.client = Client()

    def test_download(self):
        self.dataset.export_data(self.user)

        export = Export.objects.get(dataset=self.dataset)

        response = self.client.get('/api/1.0/export/%i/download/' % export.id, **self.auth_headers) 

        self.assertEqual(response.status_code, 200)

        with open(os.path.join(utils.TEST_DATA_PATH, utils.TEST_DATA_FILENAME)) as f:
            self.assertEqual(response.content, f.read())


########NEW FILE########
__FILENAME__ = test_api_notification
#!/usr/bin/env python

from datetime import datetime

from django.conf import settings
from django.test import TransactionTestCase
from django.test.client import Client
from django.utils import simplejson as json
import pytz

from panda.models import Notification, UserProxy 
from panda.tests import utils

class TestAPINotifications(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        settings.CELERY_ALWAYS_EAGER = True
        
        utils.setup_test_solr()
        
        self.user = utils.get_panda_user()
        self.dataset = utils.get_test_dataset(self.user)
        self.upload = utils.get_test_data_upload(self.user, self.dataset)

        self.dataset.import_data(self.user, self.upload, 0)

        self.auth_headers = utils.get_auth_headers()

        self.client = Client()

    def test_get(self):
        notification = Notification.objects.all()[0]

        response = self.client.get('/api/1.0/notification/%i/' % notification.id, **self.auth_headers) 

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        sent_at = datetime.strptime(body['sent_at'], "%Y-%m-%dT%H:%M:%S").replace(tzinfo=pytz.timezone('Etc/UTC'))
        self.assertEqual(sent_at, notification.sent_at.replace(microsecond=0))
        self.assertEqual(body['read_at'], None)
        self.assertEqual(body['message'], notification.message)

    def test_get_not_recipient(self):
        response = self.client.get('/api/1.0/notification/%i/' % self.dataset.current_task.id) 

        self.assertEqual(response.status_code, 401)

    def test_get_unauthorized(self):
        UserProxy.objects.create_user('nobody@nobody.com', 'nobody@nobody.com', 'password')

        notification = Notification.objects.all()[0]

        response = self.client.get('/api/1.0/notification/%i/' % notification.id, **utils.get_auth_headers('nobody@nobody.com')) 

        self.assertEqual(response.status_code, 404)

    def test_list(self):
        response = self.client.get('/api/1.0/notification/', data={ 'limit': 5 }, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(len(body['objects']), 1)
        self.assertEqual(body['meta']['total_count'], 1)
        self.assertEqual(body['meta']['limit'], 5)
        self.assertEqual(body['meta']['offset'], 0)
        self.assertEqual(body['meta']['next'], None)
        self.assertEqual(body['meta']['previous'], None)

    def test_list_unauthorized(self):
        UserProxy.objects.create_user('nobody@nobody.com', 'nobody@nobody.com', 'password')

        response = self.client.get('/api/1.0/notification/?', data={ 'limit': 5 }, **utils.get_auth_headers('nobody@nobody.com')) 

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(len(body['objects']), 0)
        self.assertEqual(body['meta']['total_count'], 0)
        self.assertEqual(body['meta']['limit'], 5)
        self.assertEqual(body['meta']['offset'], 0)
        self.assertEqual(body['meta']['next'], None)
        self.assertEqual(body['meta']['previous'], None)

    def test_update(self):
        notification = Notification.objects.all()[0]

        data = json.dumps({ 'read_at': datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S') })

        response = self.client.put('/api/1.0/notification/%i/' % notification.id, data=data, content_type='application/json', **self.auth_headers) 

        self.assertEqual(response.status_code, 204)

        # Refresh
        notification = Notification.objects.all()[0]

        self.assertNotEqual(notification.read_at, None)

    def test_update_unauthorized(self):
        UserProxy.objects.create_user('nobody@nobody.com', 'nobody@nobody.com', 'password')

        notification = Notification.objects.all()[0]

        data = json.dumps({ 'read_at': datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S') })

        response = self.client.put('/api/1.0/notification/%i/' % notification.id, data=data, content_type='application/json', **utils.get_auth_headers('nobody@nobody.com')) 
        # This returns 201 (rather than 401), because the PUT fails to match an
        # existing notification that the user has access to and thus falls
        # back to creating a new one.
        # This is probably not ideal, but works.
        self.assertEqual(response.status_code, 201)


########NEW FILE########
__FILENAME__ = test_api_related_upload
#!/usr/bin/env python

import os.path

from django.conf import settings
from django.test import TransactionTestCase
from django.test.client import Client
from django.utils import simplejson as json

from panda.models import RelatedUpload
from panda.tests import utils

class TestAPIRelatedUpload(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        self.user = utils.get_panda_user()
        self.dataset = utils.get_test_dataset(self.user)
        self.upload = utils.get_test_related_upload(self.user, self.dataset)

        self.auth_headers = utils.get_auth_headers()

        self.client = Client()

    def test_get(self):
        response = self.client.get('/api/1.0/related_upload/%i/' % self.upload.id, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(body['filename'], self.upload.filename)
        self.assertEqual(body['original_filename'], self.upload.original_filename)
        self.assertEqual(body['size'], self.upload.size)
        self.assertEqual(body['creator']['resource_uri'], '/api/1.0/user/%i/' % self.user.id)
        self.assertNotEqual(body['creation_date'], None)

    def test_get_unauthorized(self):
        response = self.client.get('/api/1.0/related_upload/%i/' % self.upload.id)

        self.assertEqual(response.status_code, 401)

    def test_list(self):
        response = self.client.get('/api/1.0/related_upload/', data={ 'limit': 5 }, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(len(body['objects']), 1)
        self.assertEqual(body['meta']['total_count'], 1)
        self.assertEqual(body['meta']['limit'], 5)
        self.assertEqual(body['meta']['offset'], 0)
        self.assertEqual(body['meta']['next'], None)
        self.assertEqual(body['meta']['previous'], None)

    def test_create_denied(self):
        new_upload = {
            'filename': 'test.csv',
            'original_filename': 'test.csv',
            'size': 20
        }

        response = self.client.post('/api/1.0/related_upload/', content_type='application/json', data=json.dumps(new_upload), **self.auth_headers)

        self.assertEqual(response.status_code, 405)

    def test_download(self):
        response = self.client.get('/api/1.0/related_upload/%i/download/' % self.upload.id, **self.auth_headers)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Disposition'], 'attachment; filename=%s' % self.upload.original_filename)
        self.assertEqual(int(response['Content-Length']), self.upload.size)

        with open(os.path.join(settings.MEDIA_ROOT, utils.TEST_DATA_FILENAME)) as f:
            self.assertEqual(f.read(), response.content)

    def test_download_unauthorized(self):
        response = self.client.get('/api/1.0/related_upload/%i/download/' % self.upload.id)

        self.assertEqual(response.status_code, 401)

    def test_upload_file(self):
        with open(os.path.join(settings.MEDIA_ROOT, utils.TEST_DATA_FILENAME)) as f:
            response = self.client.post('/related_upload/', data={ 'file': f, 'dataset_slug': self.dataset.slug }, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)
        
        self.assertEqual(body['success'], True)

        upload = RelatedUpload.objects.get(id=body['id'])

        self.assertEqual(body['original_filename'], upload.original_filename)
        self.assertEqual(body['size'], os.path.getsize(os.path.join(settings.MEDIA_ROOT, utils.TEST_DATA_FILENAME)))
        self.assertEqual(body['size'], upload.size)
        self.assertEqual(body['creator']['resource_uri'], '/api/1.0/user/%i/' % self.user.id)

    def test_upload_unauthorized(self):
        with open(os.path.join(settings.MEDIA_ROOT, utils.TEST_DATA_FILENAME)) as f:
            response = self.client.post('/related_upload/', data={ 'file': f })

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)
        
        self.assertEqual(body['success'], False)
        self.assertEqual(body['forbidden'], True)

    def test_delete(self):
        path = self.upload.get_path()
        self.assertEqual(os.path.isfile(path), True)
        
        response = self.client.delete('/api/1.0/related_upload/%i/' % self.upload.id, **self.auth_headers)

        self.assertEqual(response.status_code, 204)
        self.assertEqual(os.path.exists(path), False)


########NEW FILE########
__FILENAME__ = test_api_search_subscriptions
#!/usr/bin/env python

from django.conf import settings
from django.test import TransactionTestCase
from django.test.client import Client
from django.utils import simplejson as json

from panda.models import SearchSubscription, UserProxy 
from panda.tests import utils

class TestAPISearchSubscriptions(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        settings.CELERY_ALWAYS_EAGER = True
        
        utils.setup_test_solr()
        
        self.user = utils.get_panda_user()
        self.dataset = utils.get_test_dataset(self.user)
        self.upload = utils.get_test_data_upload(self.user, self.dataset)

        self.dataset.import_data(self.user, self.upload, 0)

        self.auth_headers = utils.get_auth_headers()

        self.client = Client()

    def test_get(self):
        sub = SearchSubscription.objects.create(
            user=self.user,
            dataset=self.dataset,
            query='*'
        )

        response = self.client.get('/api/1.0/search_subscription/%i/' % sub.id, **self.auth_headers) 

        self.assertEqual(response.status_code, 200)

    def test_get_not_user(self):
        sub = SearchSubscription.objects.create(
            user=self.user,
            dataset=self.dataset,
            query='*'
        )

        response = self.client.get('/api/1.0/search_subscription/%i/' % sub.id) 

        self.assertEqual(response.status_code, 401)

    def test_get_unauthorized(self):
        UserProxy.objects.create_user('nobody@nobody.com', 'nobody@nobody.com', 'password')

        sub = SearchSubscription.objects.create(
            user=self.user,
            dataset=self.dataset,
            query='*'
        )

        response = self.client.get('/api/1.0/search_subscription/%i/' % sub.id, **utils.get_auth_headers('nobody@nobody.com')) 

        self.assertEqual(response.status_code, 404)

    def test_list(self):
        SearchSubscription.objects.create(
            user=self.user,
            dataset=self.dataset,
            query='*'
        )

        response = self.client.get('/api/1.0/search_subscription/', data={ 'limit': 5 }, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(len(body['objects']), 1)
        self.assertEqual(body['meta']['total_count'], 1)
        self.assertEqual(body['meta']['limit'], 5)
        self.assertEqual(body['meta']['offset'], 0)
        self.assertEqual(body['meta']['next'], None)
        self.assertEqual(body['meta']['previous'], None)

    def test_list_unauthorized(self):
        UserProxy.objects.create_user('nobody@nobody.com', 'nobody@nobody.com', 'password')

        response = self.client.get('/api/1.0/search_subscription/', data={ 'limit': 5 }, **utils.get_auth_headers('nobody@nobody.com')) 

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(len(body['objects']), 0)
        self.assertEqual(body['meta']['total_count'], 0)
        self.assertEqual(body['meta']['limit'], 5)
        self.assertEqual(body['meta']['offset'], 0)
        self.assertEqual(body['meta']['next'], None)
        self.assertEqual(body['meta']['previous'], None)

    def test_update(self):
        sub = SearchSubscription.objects.create(
            user=self.user,
            dataset=self.dataset,
            query='*'
        )

        response = self.client.put('/api/1.0/search_subscription/%i/' % sub.id, data=json.dumps({}), content_type='application/json', **self.auth_headers) 

        self.assertEqual(response.status_code, 405)

    def test_update_unauthorized(self):
        UserProxy.objects.create_user('nobody@nobody.com', 'nobody@nobody.com', 'password')

        sub = SearchSubscription.objects.create(
            user=self.user,
            dataset=self.dataset,
            query='*'
        )

        response = self.client.put('/api/1.0/search_subscription/%i/' % sub.id, data=json.dumps({}), content_type='application/json', **utils.get_auth_headers('nobody@nobody.com')) 
        # This returns 201 (rather than 401), because the PUT fails to match an
        # existing subscription that the user has access to and thus falls
        # back to creating a new one.
        # This is probably not ideal, but works.
        self.assertEqual(response.status_code, 405)

    def test_delete(self):
        sub = SearchSubscription.objects.create(
            user=self.user,
            dataset=self.dataset,
            query='*'
        )

        response = self.client.delete('/api/1.0/search_subscription/%i/' % sub.id, **self.auth_headers)

        self.assertEqual(response.status_code, 204)

        response = self.client.get('/api/1.0/search_subscription/%i/' % sub.id, **self.auth_headers)

        self.assertEqual(response.status_code, 404)

        with self.assertRaises(SearchSubscription.DoesNotExist):
            SearchSubscription.objects.get(id=sub.id)

    def test_delete_unauthorized(self):
        UserProxy.objects.create_user('nobody@nobody.com', 'nobody@nobody.com', 'password')

        sub = SearchSubscription.objects.create(
            user=self.user,
            dataset=self.dataset,
            query='*'
        )

        response = self.client.delete('/api/1.0/search_subscription/%i/' % sub.id, **utils.get_auth_headers('nobody@nobody.com'))

        self.assertEqual(response.status_code, 404)

        response = self.client.get('/api/1.0/search_subscription/%i/' % sub.id, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        # Verify no exception is raised
        SearchSubscription.objects.get(id=sub.id)


########NEW FILE########
__FILENAME__ = test_api_task_status
#!/usr/bin/env python

from datetime import datetime

from django.conf import settings
from django.test import TransactionTestCase
from django.test.client import Client
from django.utils import simplejson as json
import pytz

from panda.models import TaskStatus
from panda.tests import utils

class TestAPITaskStatus(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        settings.CELERY_ALWAYS_EAGER = True
        
        utils.setup_test_solr()
        
        self.user = utils.get_panda_user()
        self.dataset = utils.get_test_dataset(self.user)
        self.upload = utils.get_test_data_upload(self.user, self.dataset)

        self.dataset.import_data(self.user, self.upload, 0)

        self.auth_headers = utils.get_auth_headers()

        self.client = Client()

    def test_get(self):
        task = TaskStatus.objects.get(id=self.dataset.current_task.id)

        response = self.client.get('/api/1.0/task/%i/' % task.id, **self.auth_headers) 

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(body['status'], task.status)
        self.assertEqual(body['task_name'], task.task_name)
        start = datetime.strptime(body['start'], "%Y-%m-%dT%H:%M:%S").replace(tzinfo=pytz.timezone('Etc/UTC'))
        self.assertEqual(start, task.start.replace(microsecond=0))
        end = datetime.strptime(body['end'], "%Y-%m-%dT%H:%M:%S").replace(tzinfo=pytz.timezone('Etc/UTC'))
        self.assertEqual(end, task.end.replace(microsecond=0))
        self.assertEqual(body['message'], task.message)
        self.assertEqual(body['traceback'], None)
        self.assertNotEqual(body['creator'], None)

    def test_get_unauthorized(self):
        response = self.client.get('/api/1.0/task/%i/' % self.dataset.current_task.id) 

        self.assertEqual(response.status_code, 401)

    def test_list(self):
        response = self.client.get('/api/1.0/task/', data={ 'limit': 5 }, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(len(body['objects']), 1)
        self.assertEqual(body['meta']['total_count'], 1)
        self.assertEqual(body['meta']['limit'], 5)
        self.assertEqual(body['meta']['offset'], 0)
        self.assertEqual(body['meta']['next'], None)
        self.assertEqual(body['meta']['previous'], None)

    def test_create_denied(self):
        new_task = {
            'task_name': 'panda.tasks.ImportDatasetTask'
        }

        response = self.client.post('/api/1.0/task/', content_type='application/json', data=json.dumps(new_task), **self.auth_headers)

        self.assertEqual(response.status_code, 405)


########NEW FILE########
__FILENAME__ = test_api_user
#!/usr/bin/env python

from django.contrib.auth.models import User
from django.contrib.auth import authenticate
from django.contrib.auth.models import Group
from django.conf import settings
from django.test import TestCase, TransactionTestCase
from django.test.client import Client
from django.utils import simplejson as json
from tastypie.bundle import Bundle

from panda.api.users import UserValidation
from panda.models import UserProxy
from panda.tests import utils

class TestUserValidation(TestCase):
    def setUp(self):
        self.validator = UserValidation()

    def test_required_fields(self):
        bundle = Bundle(data={})

        errors = self.validator.is_valid(bundle)

        self.assertIn("email", errors)

    def test_invalid_emails(self):
        for email in ['nobody.com', 'nobody@', 'nobody@nobody', 'nobody@.com', '']:
            bundle = Bundle(data={ 'email': email })
        
            errors = self.validator.is_valid(bundle)

            self.assertIn("email", errors)

    def test_valid_emails(self):
        for email in ['nobody@nobody.com', 'nobody.nobody@somewhere.com', 'no_body@no-body.re']:
            bundle = Bundle(data={ 'email': email })
        
            errors = self.validator.is_valid(bundle)

            self.assertNotIn("email", errors)

class TestAPIUser(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        settings.CELERY_ALWAYS_EAGER = True

        self.user = utils.get_panda_user() 
        self.panda_user_group = Group.objects.get(name='panda_user')
        
        self.auth_headers = utils.get_auth_headers()

        self.client = Client()

    def test_get(self):
        response = self.client.get('/api/1.0/user/%i/' % self.user.id, **self.auth_headers) 

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertNotIn('username', body)
        self.assertNotIn('password', body)
        self.assertNotIn('is_superuser', body)
        self.assertNotIn('is_staff', body)

    def test_get_unauthorized(self):
        response = self.client.get('/api/1.0/user/%i/' % self.user.id) 

        self.assertEqual(response.status_code, 401)

    def test_list(self):
        response = self.client.get('/api/1.0/user/', data={ 'limit': 5 }, **self.auth_headers)

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(len(body['objects']), 2)
        self.assertEqual(body['meta']['total_count'], 2)
        self.assertEqual(body['meta']['limit'], 5)
        self.assertEqual(body['meta']['offset'], 0)
        self.assertEqual(body['meta']['next'], None)
        self.assertEqual(body['meta']['previous'], None)

    def test_create_as_admin(self):
        new_user = {
            'email': 'tester@tester.com',
            'password': 'test',
            'first_name': 'Testy',
            'last_name': 'McTester'
        }

        response = self.client.post('/api/1.0/user/', content_type='application/json', data=json.dumps(new_user), **utils.get_auth_headers('panda@pandaproject.net'))

        self.assertEqual(response.status_code, 201)
        
        body = json.loads(response.content)

        self.assertEqual(body['email'], 'tester@tester.com')
        self.assertEqual(body['first_name'], 'Testy')
        self.assertEqual(body['last_name'], 'McTester')
        
        new_user = User.objects.get(username='tester@tester.com')

        self.assertEqual(new_user.username, 'tester@tester.com')
        self.assertEqual(new_user.email, 'tester@tester.com')
        self.assertEqual(new_user.first_name, 'Testy')
        self.assertEqual(new_user.last_name, 'McTester')
        self.assertNotEqual(new_user.api_key, None)

        self.assertEqual(list(new_user.groups.all()), [self.panda_user_group])

        self.assertEqual(authenticate(username='tester@tester.com', password='test'), new_user)

    def test_create_as_user(self):
        new_user = {
            'email': 'tester@tester.com',
            'password': 'test',
            'first_name': 'Testy',
            'last_name': 'McTester'
        }

        response = self.client.post('/api/1.0/user/', content_type='application/json', data=json.dumps(new_user), **self.auth_headers)

        self.assertEqual(response.status_code, 401)

    def test_update_as_user(self):
        update_user = {
            'email': 'tester@tester.com',
            'first_name': 'Testy',
            'last_name': 'McTester'
        }

        before_user = self.user

        response = self.client.put('/api/1.0/user/%i/' % self.user.id, content_type='application/json', data=json.dumps(update_user), **self.auth_headers)

        self.assertEqual(response.status_code, 202)

        after_user = UserProxy.objects.get(id=self.user.id)

        self.assertEqual(after_user.email, 'tester@tester.com')
        self.assertEqual(after_user.username, 'tester@tester.com')
        self.assertEqual(after_user.first_name, 'Testy')
        self.assertEqual(after_user.last_name, 'McTester')
        self.assertEqual(before_user.date_joined, after_user.date_joined)
        self.assertEqual(before_user.is_active, after_user.is_active)
        self.assertEqual(before_user.last_login, after_user.last_login)
        self.assertEqual(before_user.password, after_user.password)

    def test_update_as_different_user(self):
        new_user = {
            'email': 'tester@tester.com',
            'password': 'test',
            'first_name': 'Testy',
            'last_name': 'McTester'
        }

        response = self.client.post('/api/1.0/user/', content_type='application/json', data=json.dumps(new_user), **utils.get_auth_headers('panda@pandaproject.net'))

        self.assertEqual(response.status_code, 201)

        update_user = {
            'email': 'foo@bar.com',
            'first_name': 'Testy',
            'last_name': 'McTester'
        }

        response = self.client.put('/api/1.0/user/%i/' % self.user.id, content_type='application/json', data=json.dumps(update_user), **utils.get_auth_headers('tester@tester.com'))

        self.assertEqual(response.status_code, 401)

    def test_update_as_admin(self):
        update_user = {
            'email': 'tester@tester.com',
            'first_name': 'Testy',
            'last_name': 'McTester'
        }

        before_user = self.user

        response = self.client.put('/api/1.0/user/%i/' % self.user.id, content_type='application/json', data=json.dumps(update_user), **utils.get_auth_headers('panda@pandaproject.net'))

        self.assertEqual(response.status_code, 202)

        after_user = UserProxy.objects.get(id=self.user.id)

        self.assertEqual(after_user.email, 'tester@tester.com')
        self.assertEqual(after_user.username, 'tester@tester.com')
        self.assertEqual(after_user.first_name, 'Testy')
        self.assertEqual(after_user.last_name, 'McTester')
        self.assertEqual(before_user.date_joined, after_user.date_joined)
        self.assertEqual(before_user.is_active, after_user.is_active)
        self.assertEqual(before_user.last_login, after_user.last_login)
        self.assertEqual(before_user.password, after_user.password)

    def test_change_password(self):
        update_user = {
            'email': 'tester@tester.com',
            'first_name': 'Testy',
            'last_name': 'McTester',
            'password': 'foobarbaz'
        }

        before_user = self.user

        response = self.client.put('/api/1.0/user/%i/' % self.user.id, content_type='application/json', data=json.dumps(update_user), **self.auth_headers)

        self.assertEqual(response.status_code, 202)

        after_user = UserProxy.objects.get(id=self.user.id)

        self.assertEqual(after_user.email, 'tester@tester.com')
        self.assertEqual(after_user.username, 'tester@tester.com')
        self.assertEqual(after_user.first_name, 'Testy')
        self.assertEqual(after_user.last_name, 'McTester')
        self.assertEqual(before_user.date_joined, after_user.date_joined)
        self.assertEqual(before_user.is_active, after_user.is_active)
        self.assertEqual(before_user.last_login, after_user.last_login)
        self.assertNotEqual(before_user.password, after_user.password)
        self.assertNotEqual(after_user.password, 'foobarbaz')


########NEW FILE########
__FILENAME__ = test_dataset
#!/usr/bin/env python

import os.path

from django.conf import settings
from django.test import TransactionTestCase
from django.utils import simplejson as json

from panda import solr
from panda.exceptions import DatasetLockedError, DataImportError, DataSamplingError
from panda.models import Dataset, DataUpload, RelatedUpload, TaskStatus
from panda.tests import utils
from panda.utils.column_schema import update_indexed_names

class TestDataset(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        settings.CELERY_ALWAYS_EAGER = True

        utils.setup_test_solr() 

        self.user = utils.get_panda_user()
        self.dataset = utils.get_test_dataset(self.user)
        self.upload = utils.get_test_data_upload(self.user, self.dataset)

    def test_lock(self):
        self.dataset.lock()
        self.assertEqual(self.dataset.locked, True)

    def test_lock_fail(self):
        self.dataset.lock()
        self.assertRaises(DatasetLockedError, self.dataset.lock)

    def test_unlock(self):
        self.dataset.lock()
        self.dataset.unlock()
        self.dataset.lock()
        self.dataset.unlock()

        self.assertEqual(self.dataset.locked, False)

    def test_metadata_searchable(self):
        response = solr.query(settings.SOLR_DATASETS_CORE, 'contributors', sort='slug asc')

        self.assertEqual(response['response']['numFound'], 1)

    def test_sample_encoding_success(self):
        utils.get_test_data_upload(self.user, self.dataset, utils.TEST_LATIN1_FILENAME, encoding='latin1')

    def test_sample_encoding_fails(self):
        with self.assertRaises(DataSamplingError):
            utils.get_test_data_upload(self.user, self.dataset, utils.TEST_LATIN1_FILENAME)

    def test_import_csv(self):
        self.dataset.import_data(self.user, self.upload)

        task = self.dataset.current_task

        self.assertNotEqual(task, None)
        self.assertNotEqual(task.id, None)
        self.assertEqual(task.task_name, 'panda.tasks.import.csv')

        # Refresh from database
        dataset = Dataset.objects.get(id=self.dataset.id)
        upload = DataUpload.objects.get(id=self.upload.id)
        task = TaskStatus.objects.get(id=task.id)

        self.assertEqual([c['name'] for c in dataset.column_schema], ['id', 'first_name', 'last_name', 'employer'])
        self.assertEqual([c['type'] for c in dataset.column_schema], ['int', 'unicode', 'unicode', 'unicode'])
        self.assertEqual([c['indexed_name'] for c in dataset.column_schema], [None, None, None, None])
        self.assertEqual(dataset.row_count, 4)
        self.assertEqual(upload.imported, True)
        self.assertEqual(task.status, 'SUCCESS')
        self.assertNotEqual(task.start, None)
        self.assertNotEqual(task.end, None)
        self.assertEqual(task.traceback, None)
        self.assertEqual(dataset.locked, False)

        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'Christopher')['response']['numFound'], 1)

    def test_import_xls(self):
        xls_upload = utils.get_test_data_upload(self.user, self.dataset, utils.TEST_XLS_FILENAME)

        self.dataset.import_data(self.user, xls_upload)

        task = self.dataset.current_task

        self.assertNotEqual(task, None)
        self.assertNotEqual(task.id, None)
        self.assertEqual(task.task_name, 'panda.tasks.import.xls')

        # Refresh from database
        dataset = Dataset.objects.get(id=self.dataset.id)
        xls_upload = DataUpload.objects.get(id=xls_upload.id)
        task = TaskStatus.objects.get(id=task.id)

        self.assertEqual([c['name'] for c in dataset.column_schema], ['id', 'first_name', 'last_name', 'employer'])
        self.assertEqual([c['type'] for c in dataset.column_schema], ['int', 'unicode', 'unicode', 'unicode'])
        self.assertEqual([c['indexed_name'] for c in dataset.column_schema], [None, None, None, None])
        self.assertEqual(dataset.row_count, 4)
        self.assertEqual(xls_upload.imported, True)
        self.assertEqual(task.status, 'SUCCESS')
        self.assertNotEqual(task.start, None)
        self.assertNotEqual(task.end, None)
        self.assertEqual(task.traceback, None)
        self.assertEqual(dataset.locked, False)

        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'Christopher')['response']['numFound'], 1)

    def test_import_excel_xlsx(self):
        xlsx_upload = utils.get_test_data_upload(self.user, self.dataset, utils.TEST_EXCEL_XLSX_FILENAME)

        self.dataset.import_data(self.user, xlsx_upload)

        task = self.dataset.current_task

        self.assertNotEqual(task, None)
        self.assertNotEqual(task.id, None)
        self.assertEqual(task.task_name, 'panda.tasks.import.xlsx')

        # Refresh from database
        dataset = Dataset.objects.get(id=self.dataset.id)
        xlsx_upload = DataUpload.objects.get(id=xlsx_upload.id)
        task = TaskStatus.objects.get(id=task.id)

        self.assertEqual([c['name'] for c in dataset.column_schema], ['id', 'first_name', 'last_name', 'employer'])
        self.assertEqual(dataset.row_count, 4)
        self.assertEqual(xlsx_upload.imported, True)
        self.assertEqual(task.status, 'SUCCESS')
        self.assertNotEqual(task.start, None)
        self.assertNotEqual(task.end, None)
        self.assertEqual(task.traceback, None)
        self.assertEqual(dataset.locked, False)

        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'Christopher')['response']['numFound'], 1)

    def test_import_oo_xlsx(self):
        xlsx_upload = utils.get_test_data_upload(self.user, self.dataset, utils.TEST_OO_XLSX_FILENAME)

        self.dataset.import_data(self.user, xlsx_upload)

        task = self.dataset.current_task

        self.assertNotEqual(task, None)
        self.assertNotEqual(task.id, None)
        self.assertEqual(task.task_name, 'panda.tasks.import.xlsx')

        # Refresh from database
        dataset = Dataset.objects.get(id=self.dataset.id)
        xlsx_upload = DataUpload.objects.get(id=xlsx_upload.id)
        task = TaskStatus.objects.get(id=task.id)

        self.assertEqual([c['name'] for c in dataset.column_schema], ['id', 'first_name', 'last_name', 'employer'])
        self.assertEqual([c['type'] for c in dataset.column_schema], ['int', 'unicode', 'unicode', 'unicode'])
        self.assertEqual([c['indexed_name'] for c in dataset.column_schema], [None, None, None, None])
        self.assertEqual(dataset.row_count, 4)
        self.assertEqual(xlsx_upload.imported, True)
        self.assertEqual(task.status, 'SUCCESS')
        self.assertNotEqual(task.start, None)
        self.assertNotEqual(task.end, None)
        self.assertEqual(task.traceback, None)
        self.assertEqual(dataset.locked, False)

        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'Christopher')['response']['numFound'], 1)

    def test_import_additional_data_same_columns(self):
        self.dataset.import_data(self.user, self.upload)

        xls_upload = utils.get_test_data_upload(self.user, self.dataset, utils.TEST_XLS_FILENAME)
        
        # Refresh from database
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        self.dataset.import_data(self.user, xls_upload)

        # Refresh from database
        dataset = Dataset.objects.get(id=self.dataset.id)
        upload = DataUpload.objects.get(id=self.upload.id)
        xls_upload = DataUpload.objects.get(id=xls_upload.id)
        
        self.assertEqual([c['name'] for c in dataset.column_schema], ['id', 'first_name', 'last_name', 'employer'])
        self.assertEqual([c['type'] for c in dataset.column_schema], ['int', 'unicode', 'unicode', 'unicode'])
        self.assertEqual([c['indexed_name'] for c in dataset.column_schema], [None, None, None, None])
        self.assertEqual(dataset.row_count, 8)
        self.assertEqual(upload.imported, True)
        self.assertEqual(xls_upload.imported, True)
        self.assertEqual(dataset.locked, False)

        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'Christopher')['response']['numFound'], 2)
    
    def test_import_additional_data_different_columns(self):
        self.dataset.import_data(self.user, self.upload)

        xls_upload = utils.get_test_data_upload(self.user, self.dataset, utils.TEST_XLS_FILENAME)
        xls_upload.columns = ['id', 'first_name', 'last_name', 'employer', 'MORE COLUMNS!']
        xls_upload.save()
        
        # Refresh from database
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        self.assertRaises(DataImportError, self.dataset.import_data, self.user, xls_upload)

        # Refresh from database
        dataset = Dataset.objects.get(id=self.dataset.id)
        upload = DataUpload.objects.get(id=self.upload.id)
        xls_upload = DataUpload.objects.get(id=xls_upload.id)
        
        self.assertEqual([c['name'] for c in dataset.column_schema], ['id', 'first_name', 'last_name', 'employer'])
        self.assertEqual(dataset.row_count, 4)
        self.assertEqual(upload.imported, True)
        self.assertEqual(xls_upload.imported, False)
        self.assertEqual(dataset.locked, False)

        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'Christopher')['response']['numFound'], 1)

    def test_import_additional_csv_typed_columns(self):
        self.dataset.import_data(self.user, self.upload)

        # Refresh from database
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        self.dataset.reindex_data(self.user, typed_columns=[True, False, True, True])

        second_upload = utils.get_test_data_upload(self.user, self.dataset, utils.TEST_DATA_FILENAME)
        
        # Refresh from database
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        self.dataset.import_data(self.user, second_upload)

        # Refresh from database
        dataset = Dataset.objects.get(id=self.dataset.id)
        
        self.assertEqual([c['name'] for c in dataset.column_schema], ['id', 'first_name', 'last_name', 'employer'])
        self.assertEqual([c['type'] for c in dataset.column_schema], ['int', 'unicode', 'unicode', 'unicode'])
        self.assertEqual([c['indexed'] for c in dataset.column_schema], [True, False, True, True])
        self.assertEqual([c['indexed_name'] for c in dataset.column_schema], ['column_int_id', None, 'column_unicode_last_name', 'column_unicode_employer'])
        self.assertEqual(dataset.row_count, 8)
        self.assertEqual(dataset.locked, False)

        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'Christopher')['response']['numFound'], 2)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_int_id:2')['response']['numFound'], 2)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_unicode_last_name:Germuska')['response']['numFound'], 2)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_unicode_first_name:Joseph')['response']['numFound'], 0)

    def test_import_additional_xls_typed_columns(self):
        self.dataset.import_data(self.user, self.upload)

        # Refresh from database
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        self.dataset.reindex_data(self.user, typed_columns=[True, False, True, True])

        second_upload = utils.get_test_data_upload(self.user, self.dataset, utils.TEST_XLS_FILENAME)
        
        # Refresh from database
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        self.dataset.import_data(self.user, second_upload)

        # Refresh from database
        dataset = Dataset.objects.get(id=self.dataset.id)
        
        self.assertEqual([c['name'] for c in dataset.column_schema], ['id', 'first_name', 'last_name', 'employer'])
        self.assertEqual([c['type'] for c in dataset.column_schema], ['int', 'unicode', 'unicode', 'unicode'])
        self.assertEqual([c['indexed'] for c in dataset.column_schema], [True, False, True, True])
        self.assertEqual([c['indexed_name'] for c in dataset.column_schema], ['column_int_id', None, 'column_unicode_last_name', 'column_unicode_employer'])
        self.assertEqual(dataset.row_count, 8)
        self.assertEqual(dataset.locked, False)

        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'Christopher')['response']['numFound'], 2)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_int_id:2')['response']['numFound'], 2)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_unicode_last_name:Germuska')['response']['numFound'], 2)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_unicode_first_name:Joseph')['response']['numFound'], 0)

    def test_import_additional_xlsx_typed_columns(self):
        self.dataset.import_data(self.user, self.upload)

        # Refresh from database
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        self.dataset.reindex_data(self.user, typed_columns=[True, False, True, True])

        second_upload = utils.get_test_data_upload(self.user, self.dataset, utils.TEST_EXCEL_XLSX_FILENAME)
        
        # Refresh from database
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        self.dataset.import_data(self.user, second_upload)

        # Refresh from database
        dataset = Dataset.objects.get(id=self.dataset.id)
        
        self.assertEqual([c['name'] for c in dataset.column_schema], ['id', 'first_name', 'last_name', 'employer'])
        self.assertEqual([c['type'] for c in dataset.column_schema], ['int', 'unicode', 'unicode', 'unicode'])
        self.assertEqual([c['indexed'] for c in dataset.column_schema], [True, False, True, True])
        self.assertEqual([c['indexed_name'] for c in dataset.column_schema], ['column_int_id', None, 'column_unicode_last_name', 'column_unicode_employer'])
        self.assertEqual(dataset.row_count, 8)
        self.assertEqual(dataset.locked, False)

        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'Christopher')['response']['numFound'], 2)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_int_id:2')['response']['numFound'], 2)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_unicode_last_name:Germuska')['response']['numFound'], 2)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_unicode_first_name:Joseph')['response']['numFound'], 0)

    def test_delete(self):
        self.dataset.import_data(self.user, self.upload)

        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'Christopher')['response']['numFound'], 1)

        dataset_id = self.dataset.id
        self.dataset.delete()

        with self.assertRaises(Dataset.DoesNotExist):
            Dataset.objects.get(id=dataset_id)

        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'Christopher')['response']['numFound'], 0)

        response = solr.query(settings.SOLR_DATASETS_CORE, 'contributors', sort='slug asc')

        self.assertEqual(response['response']['numFound'], 0)

    def test_data_uploads_deleted(self):
        path = self.upload.get_path()

        self.assertEquals(os.path.exists(path), True)

        self.dataset.delete()

        with self.assertRaises(DataUpload.DoesNotExist):
            DataUpload.objects.get(id=self.upload.id)

        self.assertEquals(os.path.exists(path), False)

    def test_related_uploads_deleted(self):
        related_upload = utils.get_test_related_upload(self.user, self.dataset)

        path = related_upload.get_path()

        self.assertEquals(os.path.exists(path), True)

        self.dataset.delete()

        with self.assertRaises(RelatedUpload.DoesNotExist):
            RelatedUpload.objects.get(id=related_upload.id)

        self.assertEquals(os.path.exists(path), False)

    def test_related_uploads_deleted(self):
        path = self.upload.get_path()

        self.assertEquals(os.path.exists(path), True)

        self.dataset.delete()

        with self.assertRaises(DataUpload.DoesNotExist):
            DataUpload.objects.get(id=self.upload.id)

        self.assertEquals(os.path.exists(path), False)

    def test_get_row(self):
        self.dataset.import_data(self.user, self.upload, 0)

        row = self.dataset.get_row('1')

        self.assertEqual(row['external_id'], '1')
        self.assertEqual(json.loads(row['data']), ['1', 'Brian', 'Boyer', 'Chicago Tribune'])

    def test_add_row(self):
        self.dataset.import_data(self.user, self.upload, 0)

        # Refresh dataset so row_count is available
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        new_row =['5', 'Somebody', 'Else', 'Somewhere']

        self.dataset.add_row(self.user, new_row, external_id='5')
        row = self.dataset.get_row('5')

        self.assertEqual(row['external_id'], '5')
        self.assertEqual(json.loads(row['data']), new_row)
        self.assertEqual(self.dataset.row_count, 5)
        self.assertNotEqual(self.dataset.last_modified, None)
        self.assertEqual(self.dataset._count_rows(), 5)

    def test_add_many_rows(self):
        self.dataset.import_data(self.user, self.upload, 0)

        # Refresh dataset so row_count is available
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        new_rows = [
            (['5', 'Somebody', 'Else', 'Somewhere'], 5),
            (['6', 'Another', 'Person', 'Somewhere'], 6)
        ]

        self.dataset.add_many_rows(self.user, new_rows)
        row = self.dataset.get_row('6')

        self.assertEqual(row['external_id'], '6')
        self.assertEqual(json.loads(row['data']), new_rows[1][0])
        self.assertEqual(self.dataset.row_count, 6)
        self.assertNotEqual(self.dataset.last_modified, None)
        self.assertEqual(self.dataset._count_rows(), 6)

    def test_add_row_typed(self):
        self.dataset.import_data(self.user, self.upload, 0)

        self.dataset.reindex_data(self.user, typed_columns=[True, False, True, True])

        # Refresh from database
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        new_row =['5', 'Somebody', 'Else', 'Somewhere']

        self.dataset.add_row(self.user, new_row, external_id='5')
        row = self.dataset.get_row('5')

        self.assertEqual(row['external_id'], '5')
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_int_id:5')['response']['numFound'], 1)

    def test_add_many_rows_typed(self):
        self.dataset.import_data(self.user, self.upload, 0)

        self.dataset.reindex_data(self.user, typed_columns=[True, False, True, True])

        # Refresh dataset so row_count is available
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        new_rows = [
            (['5', 'Somebody', 'Else', 'Somewhere'], 5),
            (['6', 'Another', 'Person', 'Somewhere'], 6)
        ]

        self.dataset.add_many_rows(self.user, new_rows)
        row = self.dataset.get_row('6')

        self.assertEqual(row['external_id'], '6')
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_int_id:[5 TO 6]')['response']['numFound'], 2)

    def test_delete_row(self):
        self.dataset.import_data(self.user, self.upload, 0)

        # Refresh dataset so row_count is available
        self.dataset = Dataset.objects.get(id=self.dataset.id)

        self.dataset.delete_row(self.user, '1')
        row = self.dataset.get_row('1')

        self.assertEqual(row, None)
        self.assertEqual(self.dataset.row_count, 3)
        self.assertNotEqual(self.dataset.last_modified, None)
        self.assertEqual(self.dataset._count_rows(), 3)

    def test_export_csv(self):
        self.dataset.import_data(self.user, self.upload)
        
        dataset = Dataset.objects.get(id=self.dataset.id)

        dataset.export_data(self.user, filename='test_export')
        self.assertEqual(dataset.locked, False)

        # Refresh from database
        dataset = Dataset.objects.get(id=self.dataset.id)
        task = TaskStatus.objects.get(id=dataset.current_task.id)

        self.assertEqual(task.task_name, 'panda.tasks.export.csv')
        self.assertEqual(task.status, 'SUCCESS')
        self.assertNotEqual(task.start, None)
        self.assertNotEqual(task.end, None)
        self.assertEqual(task.traceback, None)
        self.assertEqual(dataset.locked, False)

        with open(os.path.join(utils.TEST_DATA_PATH, utils.TEST_DATA_FILENAME), 'r') as f:
            imported_csv = f.read()

        with open(os.path.join(settings.EXPORT_ROOT, 'test_export.csv')) as f:
            exported_csv = f.read()

        self.assertEqual(imported_csv, exported_csv)

    def test_export_query_csv(self):
        self.dataset.import_data(self.user, self.upload)
        
        dataset = Dataset.objects.get(id=self.dataset.id)

        dataset.export_data(self.user, query='tribune', filename='test_export')
        self.assertEqual(dataset.locked, False)

        # Refresh from database
        dataset = Dataset.objects.get(id=self.dataset.id)
        task = TaskStatus.objects.get(id=dataset.current_task.id)

        self.assertEqual(task.task_name, 'panda.tasks.export.csv')
        self.assertEqual(task.status, 'SUCCESS')
        self.assertNotEqual(task.start, None)
        self.assertNotEqual(task.end, None)
        self.assertEqual(task.traceback, None)
        self.assertEqual(dataset.locked, False)

        with open(os.path.join(settings.EXPORT_ROOT, 'test_export.csv')) as f:
            self.assertEqual('id,first_name,last_name,employer\n', f.next())
            self.assertEqual('1,Brian,Boyer,Chicago Tribune\n', f.next())
            self.assertEqual('2,Joseph,Germuska,Chicago Tribune\n', f.next())

            with self.assertRaises(StopIteration):
                f.next()

    def test_reindex(self):
        self.dataset.import_data(self.user, self.upload)

        # Refresh from database
        dataset = Dataset.objects.get(id=self.dataset.id)

        dataset.reindex_data(self.user, typed_columns=[True, False, True, True])

        # Refresh from database
        dataset = Dataset.objects.get(id=self.dataset.id)
        task = dataset.current_task

        self.assertEqual(task.status, 'SUCCESS')
        self.assertNotEqual(task.start, None)
        self.assertNotEqual(task.end, None)
        self.assertEqual(task.traceback, None)

        self.assertEqual([c['name'] for c in dataset.column_schema], ['id', 'first_name', 'last_name', 'employer'])
        self.assertEqual([c['type'] for c in dataset.column_schema], ['int', 'unicode', 'unicode', 'unicode'])
        self.assertEqual([c['indexed'] for c in dataset.column_schema], [True, False, True, True])
        self.assertEqual([c['indexed_name'] for c in dataset.column_schema], ['column_int_id', None, 'column_unicode_last_name', 'column_unicode_employer'])
        self.assertEqual([c['min'] for c in dataset.column_schema], [1, None, None, None])
        self.assertEqual([c['max'] for c in dataset.column_schema], [4, None, None, None])
        self.assertEqual(dataset.row_count, 4)
        self.assertEqual(dataset.locked, False)

        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_int_id:2')['response']['numFound'], 1)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_unicode_last_name:Germuska')['response']['numFound'], 1)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_unicode_first_name:Joseph')['response']['numFound'], 0)

    def test_reindex_complex(self):
        upload = utils.get_test_data_upload(self.user, self.dataset, filename=utils.TEST_CSV_TYPES_FILENAME)
        self.dataset.import_data(self.user, upload)

        # Refresh from database
        dataset = Dataset.objects.get(id=self.dataset.id)

        dataset.reindex_data(self.user, typed_columns=[True for c in upload.columns])

        # Refresh from database
        dataset = Dataset.objects.get(id=self.dataset.id)
        task = dataset.current_task

        self.assertEqual(task.status, 'SUCCESS')
        self.assertNotEqual(task.start, None)
        self.assertNotEqual(task.end, None)
        self.assertEqual(task.traceback, None)

        self.assertEqual([c['name'] for c in dataset.column_schema], ['text', 'date', 'integer', 'boolean', 'float', 'time', 'datetime', 'empty_column', ''])
        self.assertEqual([c['type'] for c in dataset.column_schema], ['unicode', 'date', 'int', 'bool', 'float', 'time', 'datetime', None, 'unicode'])
        self.assertEqual([c['indexed'] for c in dataset.column_schema], [True for c in upload.columns])
        self.assertEqual([c['indexed_name'] for c in dataset.column_schema], ['column_unicode_text', 'column_date_date', 'column_int_integer', 'column_bool_boolean', 'column_float_float', 'column_time_time', 'column_datetime_datetime', None, 'column_unicode_'])
        self.assertEqual([c['min'] for c in dataset.column_schema], [None, u'1920-01-01T00:00:00', 40, None, 1.0, u'9999-12-31T00:00:00', u'1971-01-01T04:14:00', None, None])
        self.assertEqual([c['max'] for c in dataset.column_schema], [None, u'1971-01-01T00:00:00', 164, None, 41800000.01, u'9999-12-31T14:57:13', u'2048-01-01T14:57:00', None, None])
        self.assertEqual(dataset.row_count, 5)
        self.assertEqual(dataset.locked, False)

        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_bool_boolean:true')['response']['numFound'], 2)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_unicode_text:"Chicago Tribune"')['response']['numFound'], 1)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_datetime_datetime:[1971-01-01T01:01:01Z TO NOW]')['response']['numFound'], 1)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_time_time:[9999-12-31T04:13:01Z TO *]')['response']['numFound'], 2)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_date_date:[1971-01-01T00:00:00Z TO NOW]')['response']['numFound'], 1)

    def test_generate_typed_column_names_none(self):
        self.dataset.import_data(self.user, self.upload)

        self.assertEqual([c['indexed_name'] for c in self.dataset.column_schema], [None, None, None, None])

    def test_generate_typed_column_names_some(self):
        self.dataset.import_data(self.user, self.upload)

        typed_columns = [True, False, True, True]

        for i, c in enumerate(self.dataset.column_schema):
            self.dataset.column_schema[i]['indexed'] = typed_columns.pop(0)

        self.dataset.column_schema = update_indexed_names(self.dataset.column_schema)

        self.assertEqual([c['indexed_name'] for c in self.dataset.column_schema], ['column_int_id', None, 'column_unicode_last_name', 'column_unicode_employer'])

    def test_generate_typed_column_names_conflict(self):
        self.dataset.import_data(self.user, self.upload)

        typed_columns = [True, False, True, True]

        for i, c in enumerate(self.dataset.column_schema):
            self.dataset.column_schema[i]['name'] = 'test'
            self.dataset.column_schema[i]['indexed'] = typed_columns.pop(0)

        self.dataset.column_schema = update_indexed_names(self.dataset.column_schema)

        self.assertEqual([c['indexed_name'] for c in self.dataset.column_schema], ['column_int_test', None, 'column_unicode_test', 'column_unicode_test2'])

    def test_reindex_with_currency(self):
        upload = utils.get_test_data_upload(self.user, self.dataset, filename=utils.TEST_MONEY)
        self.dataset.import_data(self.user, upload)

        # Refresh from database
        dataset = Dataset.objects.get(id=self.dataset.id)

        dataset.reindex_data(self.user, typed_columns=[False, True], column_types=['unicode', 'float'])

        # Refresh from database
        dataset = Dataset.objects.get(id=self.dataset.id)

        self.assertEqual([c['name'] for c in dataset.column_schema], ['product', 'price'])
        self.assertEqual([c['type'] for c in dataset.column_schema], ['unicode', 'float'])
        self.assertEqual([c['indexed'] for c in dataset.column_schema], [False, True])
        self.assertEqual([c['indexed_name'] for c in dataset.column_schema], [None, 'column_float_price'])
        self.assertEqual([c['min'] for c in dataset.column_schema], [None, 39.99])
        self.assertEqual([c['max'] for c in dataset.column_schema], [None, 2599.00])

        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_float_price:39.99')['response']['numFound'], 2)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_float_price:[1500 TO *]')['response']['numFound'], 2)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'column_float_price:*')['response']['numFound'], 8)

    def test_import_encoded_data(self):
        """
        This tests for a complicated case where a UnicodeDecodeError
        during import could be masked by an AttrbiuteError in the
        return handler.
        """
        old_sniffer_size = settings.PANDA_SNIFFER_MAX_SAMPLE_SIZE
        settings.PANDA_SNIFFER_MAX_SAMPLE_SIZE = 50

        data_upload = utils.get_test_data_upload(self.user, self.dataset, utils.TEST_LATIN1_DATA_FILENAME)

        self.dataset.import_data(self.user, data_upload)

        task = self.dataset.current_task

        self.assertNotEqual(task, None)
        self.assertNotEqual(task.id, None)
        self.assertEqual(task.task_name, 'panda.tasks.import.csv')

        # Refresh from database
        dataset = Dataset.objects.get(id=self.dataset.id)
        data_upload = DataUpload.objects.get(id=data_upload.id)
        task = TaskStatus.objects.get(id=task.id)

        self.assertEqual(len(dataset.column_schema), 8)
        self.assertEqual(dataset.row_count, None)
        self.assertEqual(data_upload.imported, False)
        self.assertEqual(task.status, 'FAILURE')
        self.assertNotEqual(task.start, None)
        self.assertNotEqual(task.end, None)
        self.assertEqual('encoded' in task.traceback, True)
        self.assertEqual(dataset.locked, False)

        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'walking')['response']['numFound'], 0)

        settings.PANDA_SNIFFER_MAX_SAMPLE_SIZE = old_sniffer_size


########NEW FILE########
__FILENAME__ = test_data_upload
#!/usr/bin/env python

import os.path

from django.conf import settings
from django.test import TransactionTestCase

from panda import solr
from panda.exceptions import DataUploadNotDeletable
from panda.models import Dataset, DataUpload
from panda.tests import utils

class TestDataUpload(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        settings.CELERY_ALWAYS_EAGER = True

        self.user = utils.get_panda_user()
        self.dataset = utils.get_test_dataset(self.user)
        self.upload = utils.get_test_data_upload(self.user, self.dataset)

    def test_created(self):
        upload = utils.get_test_data_upload(self.user, self.dataset)

        self.assertEqual(upload.original_filename, utils.TEST_DATA_FILENAME)
        self.assertEqual(upload.creator, self.user)
        self.assertNotEqual(upload.creation_date, None)
        self.assertEqual(upload.dataset, self.dataset)

        self.assertEqual(upload.data_type, 'csv')
        self.assertNotEqual(self.upload.dialect, None)
        self.assertEqual(self.upload.columns, ['id', 'first_name', 'last_name', 'employer']);
        self.assertEqual(len(self.upload.sample_data), 4)
        self.assertEqual(self.upload.sample_data[0], ['1', 'Brian', 'Boyer', 'Chicago Tribune']);
        
        self.assertEqual(len(self.upload.guessed_types), 4)
        self.assertEqual(self.upload.guessed_types, ['int', 'unicode', 'unicode', 'unicode']);
        
        self.assertEqual(upload.deletable, True)

    def test_delete(self):
        upload = utils.get_test_data_upload(self.user, self.dataset)
        upload_id = upload.id
        path = upload.get_path()

        self.assertEqual(os.path.isfile(path), True)

        solr.delete(settings.SOLR_DATA_CORE, '*:*')
        self.dataset.import_data(self.user, upload)
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'Christopher')['response']['numFound'], 1)

        upload = DataUpload.objects.get(id=upload_id)
        
        dataset = Dataset.objects.get(id=self.dataset.id)
        self.assertEqual(dataset.initial_upload, upload)
        self.assertEqual(dataset.row_count, 4)

        upload.delete()

        # Ensure dataset still exists
        dataset = Dataset.objects.get(id=self.dataset.id)
        self.assertEqual(dataset.initial_upload, None)
        self.assertEqual(dataset.row_count, 0)

        self.assertEqual(os.path.exists(path), False)

        with self.assertRaises(DataUpload.DoesNotExist):
            DataUpload.objects.get(id=upload_id)
        
        self.assertEqual(solr.query(settings.SOLR_DATA_CORE, 'Christopher')['response']['numFound'], 0)

    def test_undeletable(self):
        upload = utils.get_test_data_upload(self.user, self.dataset)

        upload.deletable = False
        upload.save()
        
        with self.assertRaises(DataUploadNotDeletable):
            upload.delete()


########NEW FILE########
__FILENAME__ = test_export_search
#!/usr/bin/env python

import os.path
from zipfile import ZipFile

from django.conf import settings
from django.test import TransactionTestCase

from panda.models import TaskStatus
from panda.tasks import ExportSearchTask
from panda.tests import utils

class TestExportSearch(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        settings.CELERY_ALWAYS_EAGER = True

        utils.setup_test_solr() 

        self.user = utils.get_panda_user()
        self.dataset = utils.get_test_dataset(self.user)
        self.dataset2 = utils.get_test_dataset(self.user)
        self.upload = utils.get_test_data_upload(self.user, self.dataset)

    def test_export_query_csv(self):
        self.dataset.import_data(self.user, self.upload)
        self.dataset2.import_data(self.user, self.upload)

        task_type = ExportSearchTask

        task = TaskStatus.objects.create(task_name=task_type.name, creator=self.user)

        task_type.apply_async(
            args=['tribune', task.id],
            kwargs={ 'filename': 'test' },
            task_id=task.id
        )

        # Refresh from database
        task = TaskStatus.objects.get(id=task.id)

        self.assertEqual(task.status, 'SUCCESS')
        self.assertNotEqual(task.start, None)
        self.assertNotEqual(task.end, None)
        self.assertEqual(task.traceback, None)

        self.assertEqual(os.path.exists(os.path.join(settings.EXPORT_ROOT, 'test.zip')), True)
        self.assertEqual(os.path.exists(os.path.join(settings.EXPORT_ROOT, 'test')), False)

        zipfile = ZipFile(os.path.join(settings.EXPORT_ROOT, 'test.zip'))

        expected_filenames = ['contributors.csv', 'contributors-2.csv']

        self.assertEqual(set(zipfile.namelist()), set(expected_filenames))

        for filename in expected_filenames:
            with zipfile.open(filename) as f:
                self.assertEqual('id,first_name,last_name,employer\n', f.next())
                self.assertEqual('1,Brian,Boyer,Chicago Tribune\n', f.next())
                self.assertEqual('2,Joseph,Germuska,Chicago Tribune\n', f.next())

                with self.assertRaises(StopIteration):
                    f.next()

        os.remove(os.path.join(settings.EXPORT_ROOT, 'test.zip'))

    def test_export_query_no_results(self):
        self.dataset.import_data(self.user, self.upload)
        self.dataset2.import_data(self.user, self.upload)

        task_type = ExportSearchTask

        task = TaskStatus.objects.create(task_name=task_type.name, creator=self.user)

        task_type.apply_async(
            args=['foobar', task.id],
            kwargs={ 'filename': 'test' },
            task_id=task.id
        )

        # Refresh from database
        task = TaskStatus.objects.get(id=task.id)

        self.assertEqual(task.status, 'SUCCESS')
        self.assertNotEqual(task.start, None)
        self.assertNotEqual(task.end, None)
        self.assertEqual(task.traceback, None)

        self.assertEqual(os.path.exists(os.path.join(settings.EXPORT_ROOT, 'test.zip')), True)
        self.assertEqual(os.path.exists(os.path.join(settings.EXPORT_ROOT, 'test')), False)

        zipfile = ZipFile(os.path.join(settings.EXPORT_ROOT, 'test.zip'))
        self.assertEqual(set(zipfile.namelist()), set())

        os.remove(os.path.join(settings.EXPORT_ROOT, 'test.zip'))


########NEW FILE########
__FILENAME__ = test_purge_orphaned_uploads
#!/usr/bin/env python

import os

from django.conf import settings
from django.test import TestCase

from panda.models import DataUpload
from panda.tasks import PurgeOrphanedUploadsTask
from panda.tests import utils

class TestPurgeOrphanedUploads(TestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        settings.CELERY_ALWAYS_EAGER = True

        utils.setup_test_solr() 

        self.user = utils.get_panda_user()
        self.dataset = utils.get_test_dataset(self.user)
        self.upload = utils.get_test_data_upload(self.user, self.dataset)
        self.related = utils.get_test_related_upload(self.user, self.dataset)

    def test_delete_orphaned_file(self):
        orphan_filepath = os.path.join(settings.MEDIA_ROOT, 'IM_AN_ORPHANED_FILE.csv')
        open(orphan_filepath, 'w').close()

        PurgeOrphanedUploadsTask.apply_async()

        self.assertEqual(os.path.exists(orphan_filepath), False)

    def test_dont_delete_data_file(self):
        PurgeOrphanedUploadsTask.apply_async()

        self.assertEqual(os.path.exists(self.upload.get_path()), True)

    def test_dont_delete_related_file(self):
        PurgeOrphanedUploadsTask.apply_async()

        self.assertEqual(os.path.exists(self.related.get_path()), True)

    def test_delete_orphaned_data_upload(self):
        self.upload.dataset = None
        self.upload.save()

        PurgeOrphanedUploadsTask.apply_async()

        with self.assertRaises(DataUpload.DoesNotExist):
            DataUpload.objects.get(id=self.upload.id)

        self.assertEqual(os.path.exists(self.upload.get_path()), False)


########NEW FILE########
__FILENAME__ = test_related_upload
#!/usr/bin/env python

import os.path

from django.conf import settings
from django.test import TransactionTestCase

from panda.tests import utils

class TestRelatedUpload(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        settings.CELERY_ALWAYS_EAGER = True

        self.user = utils.get_panda_user()
        self.dataset = utils.get_test_dataset(self.user)
        self.upload = utils.get_test_related_upload(self.user, self.dataset)

    def test_created(self):
        self.assertEqual(self.upload.original_filename, utils.TEST_DATA_FILENAME)
        self.assertEqual(self.upload.creator, self.user)
        self.assertNotEqual(self.upload.creation_date, None)
        self.assertEqual(self.upload.dataset, self.dataset)

    def test_delete(self):
        path = self.upload.get_path()

        self.assertEqual(os.path.isfile(path), True)

        self.upload.delete()

        self.assertEqual(os.path.exists(path), False)


########NEW FILE########
__FILENAME__ = test_search_subscriptions
#!/usr/bin/env python

from django.conf import settings
from django.test import TransactionTestCase

from panda.models import Category, Notification, SearchSubscription
from panda.tasks import RunSubscriptionsTask
from panda.tests import utils

class TestSearchSubscriptions(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        settings.CELERY_ALWAYS_EAGER = True

        utils.setup_test_solr() 

        self.user = utils.get_panda_user()
        self.dataset = utils.get_test_dataset(self.user)
        self.dataset2 = utils.get_test_dataset(self.user)
        self.upload = utils.get_test_data_upload(self.user, self.dataset)

    def test_subscription_dataset(self):
        # TODO: FAILING ITERMITTENTLY
        self.dataset.import_data(self.user, self.upload)

        sub = SearchSubscription.objects.create(
            user=self.user,
            dataset=self.dataset,
            query='*'
        )

        last_run = sub.last_run

        RunSubscriptionsTask.apply_async()

        # Refresh from database
        sub = SearchSubscription.objects.get(pk=sub.pk)

        self.assertNotEqual(last_run, sub.last_run)

        self.assertEqual(Notification.objects.filter(recipient=self.user).count(), 2)

    def test_subscription_category(self):
        self.dataset.import_data(self.user, self.upload)
        category = Category.objects.get(slug="politics")
        category.datasets.add(self.dataset)

        sub = SearchSubscription.objects.create(
            user=self.user,
            category=category,
            query='*'
        )

        last_run = sub.last_run

        RunSubscriptionsTask.apply_async()

        # Refresh from database
        sub = SearchSubscription.objects.get(pk=sub.pk)

        self.assertNotEqual(last_run, sub.last_run)

        self.assertEqual(Notification.objects.filter(recipient=self.user).count(), 2)

    def test_subscription_global(self):
        self.dataset.import_data(self.user, self.upload)

        sub = SearchSubscription.objects.create(
            user=self.user,
            dataset=None,
            query='*'
        )

        last_run = sub.last_run

        RunSubscriptionsTask.apply_async()

        # Refresh from database
        sub = SearchSubscription.objects.get(pk=sub.pk)

        self.assertNotEqual(last_run, sub.last_run)

        self.assertEqual(Notification.objects.filter(recipient=self.user).count(), 2)


########NEW FILE########
__FILENAME__ = test_solr
#!/usr/bin/env python

import datetime

from django.test import TestCase

from panda import solr as solrjson

class TestSolrJSONEncoder(TestCase):
    def test_datetime(self):
        v = { 'datetime': datetime.datetime(2012, 4, 11, 11, 3, 0) }
        self.assertEqual(solrjson.dumps(v), '{"datetime": "2012-04-11T11:03:00Z"}')

    def test_date(self):
        v = { 'date': datetime.date(2012, 4, 11) }
        self.assertEqual(solrjson.dumps(v), '{"date": "2012-04-11"}')

    def test_time(self):
        v = { 'time': datetime.time(11, 3, 0) }
        self.assertEqual(solrjson.dumps(v), '{"time": "11:03:00"}')

    def test_int(self):
        v = { 'int': 123 }
        self.assertEqual(solrjson.dumps(v), '{"int": 123}')


########NEW FILE########
__FILENAME__ = test_user
#!/usr/bin/env python

from django.conf import settings
from django.test import TransactionTestCase
from django.utils.timezone import now

from panda import solr
from panda.models import UserProxy
from panda.tests import utils
from tastypie.models import ApiKey

class TestUser(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        settings.CELERY_ALWAYS_EAGER = True

        self.user = utils.get_panda_user()

    def test_create_user(self):
        new_user = UserProxy.objects.create_user(
            'foo@bar.com',
            'foo@bar.com'
        )
        new_user.is_active = False
        new_user.save()

        ApiKey.objects.get(user=new_user)
        new_user.groups.get(name="panda_user")
        user_profile = new_user.get_profile()

        self.assertNotEqual(user_profile, None)
        self.assertNotEqual(user_profile.activation_key, None)
        self.assertGreater(user_profile.activation_key_expiration, now())

    def test_long_email(self):
        long_email = ''.join('F' for x in range(60))

        new_user = UserProxy.objects.create_user(long_email, long_email)
        new_user.is_active = False
        new_user.save()

        ApiKey.objects.get(user=new_user)
        new_user.groups.get(name="panda_user")
        user_profile = new_user.get_profile()

        self.assertNotEqual(user_profile, None)
        self.assertNotEqual(user_profile.activation_key, None)
        self.assertGreater(user_profile.activation_key_expiration, now())

    def test_change_user_reindex(self):
        solr.delete(settings.SOLR_DATASETS_CORE, '*:*') 

        self.user.first_name = 'bazbarfoo'
        self.user.save()

        dataset = utils.get_test_dataset(self.user)
        upload = utils.get_test_data_upload(self.user, dataset)
        
        self.assertEqual(solr.query(settings.SOLR_DATASETS_CORE, dataset.creator.first_name)['response']['numFound'], 1)
        old_name = dataset.creator.first_name

        dataset.creator.first_name = 'foobarbaz'
        dataset.creator.save()

        self.assertEqual(solr.query(settings.SOLR_DATASETS_CORE, old_name)['response']['numFound'], 0)
        self.assertEqual(solr.query(settings.SOLR_DATASETS_CORE, dataset.creator.first_name)['response']['numFound'], 1)


########NEW FILE########
__FILENAME__ = test_utils
#!/usr/bin/env python

from datetime import date, time, datetime
import os.path

from django.test import TestCase

from panda import utils
from panda.exceptions import TypeCoercionError
from panda.tests import utils as test_utils

class TestCSV(TestCase):
    def setUp(self):
        self.path = os.path.join(test_utils.TEST_DATA_PATH, test_utils.TEST_DATA_FILENAME)
        self.dialect = {
            'delimiter': ',',
            'doublequote': False,
            'lineterminator': '\r\n',
            'quotechar': '"',
            'quoting': 0,
            'skipinitialspace': False
        }

    def test_csv_sniff_dialect(self):
        dialect = utils.csv.sniff_dialect(self.path)

        self.assertEqual(dialect, self.dialect)

    def test_csv_sniff_dialect_latin1(self):
        path = os.path.join(test_utils.TEST_DATA_PATH, test_utils.TEST_LATIN1_FILENAME)

        dialect = utils.csv.sniff_dialect(path, encoding='Latin-1')

        self.assertEqual(dialect, self.dialect)

    def test_csv_extract_column_names(self):
        columns = utils.csv.extract_column_names(self.path, self.dialect)

        self.assertEqual(columns, ['id', 'first_name', 'last_name', 'employer'])

    def test_csv_extract_column_names_latin1(self):
        path = os.path.join(test_utils.TEST_DATA_PATH, test_utils.TEST_LATIN1_FILENAME)

        columns = utils.csv.extract_column_names(path, self.dialect, encoding='Latin-1')

        self.assertEqual(columns, ['activity', 'unsupplemented mean', 'dir', 'sem', 'n', 'supplemented mean', 'dir', 'sem'])

    def test_csv_sample_data(self):
        samples = utils.csv.sample_data(self.path, self.dialect, 2)

        self.assertEqual(samples, [['1', 'Brian', 'Boyer', 'Chicago Tribune'], ['2', 'Joseph', 'Germuska', 'Chicago Tribune']])

    def test_csv_sample_data_latin1(self):
        path = os.path.join(test_utils.TEST_DATA_PATH, test_utils.TEST_LATIN1_FILENAME)

        samples = utils.csv.sample_data(path, self.dialect, 2, encoding='Latin-1')

        self.assertEqual(samples, [[u'sitting', u'1.21', u'\xb1', u'0.02', u'76', u'1.27', u'\xb1', u'0.02*'], [u'standing', u'1.23', u'\xb1', u'0.03', u'58', u'1.28', u'\xb1', u'0.03']])

    def test_csv_guess_column_types(self):
        guessed_types = utils.csv.guess_column_types(self.path, self.dialect, 5, encoding='Latin-1')

        self.assertEqual(guessed_types, ['int', 'unicode', 'unicode', 'unicode'])

class TestXLS(TestCase):
    def setUp(self):
        self.path = os.path.join(test_utils.TEST_DATA_PATH, test_utils.TEST_XLS_FILENAME)
        self.dialect = {}

    def test_xls_sniff_dialect(self):
        dialect = utils.xls.sniff_dialect(self.path)

        self.assertEqual(dialect, self.dialect)

    def test_xls_extract_column_names(self):
        columns = utils.xls.extract_column_names(self.path, self.dialect)

        self.assertEqual(columns, ['id', 'first_name', 'last_name', 'employer'])

    def test_xls_sample_data(self):
        samples = utils.xls.sample_data(self.path, self.dialect, 2)

        self.assertEqual(samples, [['1', 'Brian', 'Boyer', 'Chicago Tribune'], ['2', 'Joseph', 'Germuska', 'Chicago Tribune']])

    def test_xls_guess_column_types(self):
        self.path = os.path.join(test_utils.TEST_DATA_PATH, test_utils.TEST_XLS_TYPES_FILENAME)

        guessed_types = utils.xls.guess_column_types(self.path, self.dialect, 5, encoding='Latin-1')

        self.assertEqual(guessed_types, ['unicode', 'date', 'int', 'bool', 'float', 'time', 'datetime', None, 'unicode'])

class TestXLSX(TestCase):
    def setUp(self):
        self.path = os.path.join(test_utils.TEST_DATA_PATH, test_utils.TEST_EXCEL_XLSX_FILENAME)
        self.dialect = {}

    def test_xlsx_sniff_dialect(self):
        dialect = utils.xlsx.sniff_dialect(self.path)

        self.assertEqual(dialect, self.dialect)

    def test_xlsx_extract_column_names(self):
        columns = utils.xlsx.extract_column_names(self.path, self.dialect)

        self.assertEqual(columns, ['id', 'first_name', 'last_name', 'employer'])

    def test_xlsx_sample_data(self):
        samples = utils.xlsx.sample_data(self.path, self.dialect, 2)

        self.assertEqual(samples, [['1', 'Brian', 'Boyer', 'Chicago Tribune'], ['2', 'Joseph', 'Germuska', 'Chicago Tribune']])

    def test_xlsx_guess_column_types(self):
        self.path = os.path.join(test_utils.TEST_DATA_PATH, test_utils.TEST_XLSX_TYPES_FILENAME)

        guessed_types = utils.xlsx.guess_column_types(self.path, self.dialect, 5, encoding='Latin-1')

        self.assertEqual(guessed_types, ['unicode', 'date', 'int', 'bool', 'float', 'time', 'datetime', None, 'unicode'])

class TestTypeCoercion(TestCase):
    def setUp(self):
        self.data_typer = utils.typecoercion.DataTyper([])
        self.coerce_type = self.data_typer.coerce_type

    def test_coerce_nulls(self):
        self.assertEqual(self.coerce_type(None, bool), None)
        self.assertEqual(self.coerce_type('N/A', int), None)
        self.assertEqual(self.coerce_type('n/a', datetime), None)

    def test_coerce_int_from_str(self):
        self.assertEqual(self.coerce_type('171', int), 171)

    def test_coerce_int_from_str_fails(self):
        with self.assertRaises(TypeCoercionError):
            self.assertEqual(self.coerce_type('#171', int), 171)

    def test_coerce_int_from_unicode(self):
        self.assertEqual(self.coerce_type(u'171', int), 171)

    def test_coerce_int_from_currency_str(self):
        self.assertEqual(self.coerce_type('$171,000', int), 171000)

    def test_coerce_int_from_currency_float(self):
        self.assertEqual(self.coerce_type(u'$171,000', int), 171000)

    def test_coerce_float_from_str(self):
        self.assertEqual(self.coerce_type('171.59', float), 171.59)

    def test_coerce_float_from_unicode(self):
        self.assertEqual(self.coerce_type(u'171.59', float), 171.59)

    def test_coerce_float_from_currency_str(self):
        self.assertEqual(self.coerce_type('$171,000.59', float), 171000.59)

    def test_coerce_float_from_currency_float(self):
        self.assertEqual(self.coerce_type(u'$171,000.59', float), 171000.59)

    def test_coerce_bool_from_str(self):
        self.assertEqual(self.coerce_type('True', bool), True)
        self.assertEqual(self.coerce_type('true', bool), True)
        self.assertEqual(self.coerce_type('T', bool), True)
        self.assertEqual(self.coerce_type('yes', bool), True)

    def test_coerce_bool_from_unicode(self):
        self.assertEqual(self.coerce_type(u'True', bool), True)
        self.assertEqual(self.coerce_type(u'true', bool), True)
        self.assertEqual(self.coerce_type(u'T', bool), True)
        self.assertEqual(self.coerce_type(u'yes', bool), True)

    def test_coerce_datetime_from_str(self):
        self.assertEqual(self.coerce_type('2011-4-13 8:28 AM', datetime), datetime(2011, 4, 13, 8, 28, 0))

    def test_coerce_date_from_str(self):
        self.assertEqual(self.coerce_type('2011-4-13', date), datetime(2011, 4, 13, 0, 0, 0))
        
    def test_coerce_time_from_str(self):
        self.assertEqual(self.coerce_type('8:28 AM', time), datetime(9999, 12, 31, 8, 28, 0))


########NEW FILE########
__FILENAME__ = test_views
#!/usr/bin/env python

from django.contrib.auth import authenticate
from django.test import TransactionTestCase
from django.test.client import Client
from django.utils import simplejson as json
from django.utils.timezone import now

from panda.models import UserProfile, UserProxy
from panda.tests import utils

class TestLogin(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        self.user = utils.get_panda_user()
        
        self.client = Client()

    def test_login_success(self):
        response = self.client.post('/login/', { 'email': 'user@pandaproject.net', 'password': 'user' }) 

        self.assertEqual(response.status_code, 200)

        body = json.loads(response.content)

        self.assertEqual(body['email'], 'user@pandaproject.net')
        self.assertEqual(body['notifications'], [])

        # Verify old code is dead
        self.assertNotIn('api_key', body)

    def test_login_disabled(self):
        self.user.is_active = False
        self.user.save()

        response = self.client.post('/login/', { 'email': 'user@pandaproject.net', 'password': 'user' }) 

        self.assertEqual(response.status_code, 400)

        body = json.loads(response.content)

        self.assertIn('disabled', body['__all__'])

        self.user.is_active = True
        self.user.save()

    def test_login_invalid_email(self):
        response = self.client.post('/login/', { 'email': 'NOTPANDA@pandaproject.net', 'password': 'panda' }) 

        self.assertEqual(response.status_code, 400)

        body = json.loads(response.content)

        self.assertIn('incorrect', body['__all__'])

    def test_login_incorrect_password(self):
        response = self.client.post('/login/', { 'email': 'user@pandaproject.net', 'password': 'NOPANDA' }) 

        self.assertEqual(response.status_code, 400)

        body = json.loads(response.content)

        self.assertIn('incorrect', body['__all__'])

    def test_no_get(self):
        response = self.client.get('/login/', { 'email': 'user@pandaproject.net', 'password': 'NOPANDA' }) 

        self.assertEqual(response.status_code, 400)

        body = json.loads(response.content)

        self.assertEqual(body, None)

class TestActivate(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        self.user = utils.get_panda_user()
        
        self.client = Client()

    def test_check_activation_key_valid(self):
        new_user = UserProxy.objects.create_user(
            'foo@bar.com',
            'foo@bar.com'
        )
        new_user.is_active = False
        new_user.save()

        user_profile = new_user.get_profile()

        response = self.client.get('/check_activation_key/%s/' % user_profile.activation_key)

        self.assertEqual(response.status_code, 200)
        
        body = json.loads(response.content) 

        self.assertEqual(body['activation_key'], user_profile.activation_key)
        self.assertEqual(body['email'], new_user.email)
        self.assertEqual(body['first_name'], '')
        self.assertEqual(body['last_name'], '')

    def test_check_activation_key_invalid(self):
        response = self.client.get('/check_activation_key/NOT_A_VALID_KEY/')

        self.assertEqual(response.status_code, 400)
        
    def test_activate(self):
        new_user = UserProxy.objects.create_user(
            'foo@bar.com',
            'foo@bar.com'
        )
        new_user.is_active = False
        new_user.save()

        user_profile = new_user.get_profile()
        self.assertNotEqual(user_profile.activation_key, None)
        self.assertGreater(user_profile.activation_key_expiration, now())

        activation_data = {
            'activation_key': user_profile.activation_key,
            'email': 'foo@bar.com',
            'password': 'foobarbaz',
            'reenter_password': 'foobarbaz',
            'first_name': 'Foo',
            'last_name': ''
        }

        response = self.client.post('/activate/', activation_data) 

        self.assertEqual(response.status_code, 200)

        self.assertEqual(authenticate(username='foo@bar.com', password='foobarbaz').pk, new_user.pk)
        
        # Refresh
        user_profile = UserProfile.objects.get(id=user_profile.id)
       
        self.assertNotEqual(user_profile.activation_key, None)
        self.assertLess(user_profile.activation_key_expiration, now())

class  TestForgotPassword(TransactionTestCase):
    fixtures = ['init_panda.json', 'test_users.json']

    def setUp(self):
        self.user = utils.get_panda_user()
        
        self.client = Client()

    def test_forgot_password(self):
        new_user = UserProxy.objects.create_user(
            'foo@bar.com',
            'foo@bar.com',
            'foobarbaz'
        )

        self.assertEqual(authenticate(username='foo@bar.com', password='foobarbaz').pk, new_user.pk)
        
        # Force expiration date into the past
        user_profile = new_user.get_profile() 
        user_profile.activation_key_expiration = now()
        user_profile.save()

        response = self.client.post('/forgot_password/', { 'email': 'foo@bar.com' }) 

        self.assertEqual(response.status_code, 200)

        # Refresh
        user_profile = UserProfile.objects.get(id=user_profile.id)
       
        # Expiration date should be pushed back into the future
        self.assertNotEqual(user_profile.activation_key, None)
        self.assertGreater(user_profile.activation_key_expiration, now())


########NEW FILE########
__FILENAME__ = utils
#!/usr/bin/env python

import os.path
from shutil import copyfile

from django.conf import settings
from livesettings import config_get

from panda import solr
from panda.models import Dataset, DataUpload, RelatedUpload, UserProxy

TEST_DATA_PATH = os.path.join(settings.SITE_ROOT, 'test_data')
TEST_DATA_FILENAME = 'contributors.csv'
TEST_XLS_FILENAME = 'contributors.xls'
TEST_CSV_TYPES_FILENAME = 'test_types.csv'
TEST_XLS_TYPES_FILENAME = 'test_types.xls'
TEST_XLSX_TYPES_FILENAME = 'test_types.xlsx'
TEST_EXCEL_XLSX_FILENAME = 'contributors.excel.xlsx'
TEST_OO_XLSX_FILENAME = 'contributors.oo.xlsx'
TEST_LATIN1_FILENAME = 'test_not_unicode_sample.csv'
TEST_LATIN1_DATA_FILENAME = 'test_not_unicode_data.csv'
TEST_MONEY = 'test_money.csv'

def setup_test_solr():
    settings.SOLR_DATA_CORE = 'data_test'
    settings.SOLR_DATASETS_CORE = 'datasets_test'
    config_get('PERF', 'TASK_THROTTLE').update(0.0) 
    solr.delete(settings.SOLR_DATA_CORE, '*:*')
    solr.delete(settings.SOLR_DATASETS_CORE, '*:*')

def get_auth_headers(email='user@pandaproject.net'):
    user = UserProxy.objects.get(email=email)

    return {
        'HTTP_PANDA_EMAIL': email,
        'HTTP_PANDA_API_KEY': user.api_key.key
    }

def get_admin_user():
    return UserProxy.objects.get(email='panda@pandaproject.net')

def get_panda_user():
    return UserProxy.objects.get(email='user@pandaproject.net')

def get_test_dataset(creator):
    dataset = Dataset.objects.create(
        name='Contributors',
        description='Biographic information about contributors to the PANDA project.',
        creator=creator)
    
    dataset.update_full_text()

    return dataset

def get_test_data_upload(creator, dataset, filename=TEST_DATA_FILENAME, encoding='utf8'):
    # Ensure panda subdir has been created
    try:
        os.mkdir(settings.MEDIA_ROOT)
    except OSError:
        pass

    src = os.path.join(TEST_DATA_PATH, filename)
    dst = os.path.join(settings.MEDIA_ROOT, filename)
    copyfile(src, dst)

    return DataUpload.objects.create(
        filename=filename,
        original_filename=filename,
        size=os.path.getsize(dst),
        creator=creator,
        dataset=dataset,
        encoding=encoding)

def get_test_related_upload(creator, dataset, filename=TEST_DATA_FILENAME):
    # Ensure panda subdir has been created
    try:
        os.mkdir(settings.MEDIA_ROOT)
    except OSError:
        pass

    src = os.path.join(TEST_DATA_PATH, filename)
    dst = os.path.join(settings.MEDIA_ROOT, filename)
    copyfile(src, dst)

    return RelatedUpload.objects.create(
        filename=filename,
        original_filename=filename,
        size=os.path.getsize(dst),
        creator=creator,
        dataset=dataset)


########NEW FILE########
__FILENAME__ = urls
#!/usr/bin/env python

from django.conf.urls.defaults import include, patterns, url
from tastypie.api import Api
from tastypie.utils.urls import trailing_slash

from panda.api import ActivityLogResource, CategoryResource, DatasetResource, DataUploadResource, ExportResource, NotificationResource, RelatedUploadResource, SearchSubscriptionResource, TaskResource, UserResource
from panda import views

api_1_0 = Api(api_name='1.0')
api_1_0.register(ActivityLogResource())
api_1_0.register(CategoryResource())
api_1_0.register(DatasetResource())
api_1_0.register(DataUploadResource())
api_1_0.register(ExportResource())
api_1_0.register(NotificationResource())
api_1_0.register(RelatedUploadResource())
api_1_0.register(SearchSubscriptionResource())
api_1_0.register(TaskResource())
api_1_0.register(UserResource())

urlpatterns = patterns('',
    url(r'^login%s$' % trailing_slash(), views.panda_login, name="login"),
    url(r'^logout%s$' % trailing_slash(), views.panda_logout, name="logout"),
    url(r'^check_activation_key/(?P<activation_key>[\w\d]+)%s$' % trailing_slash(), views.check_activation_key, name="check_activation_key"),
    url(r'^activate%s$' % trailing_slash(), views.activate, name="activate"),
    url(r'^forgot_password%s$' % trailing_slash(), views.forgot_password, name="forgot_password"),
    url(r'^check_available_space%s$' % trailing_slash(), views.check_available_space, name="check_available_space"),
    url(r'^data_upload%s$' % trailing_slash(), views.data_upload, name="data_upload"),
    url(r'^related_upload%s$' % trailing_slash(), views.related_upload, name="related_upload"),

    (r'^api/', include(api_1_0.urls)),
)


########NEW FILE########
__FILENAME__ = column_schema
#!/usr/bin/env python

import re
import unicodedata

def slugify(name):
    """
    Slugify a column header for use as a name.

    Adapted from Django.
    """
    slug = name
    slug = unicodedata.normalize('NFKD', unicode(slug)).encode('ascii', 'ignore')
    slug = unicode(re.sub('[^\w\s-]', '', slug).strip().lower())
    slug = re.sub('[-\s]+', '_', slug)

    return slug

def update_indexed_names(column_schema):
    """
    Update a column schema with appropriate indexed column names.
    """
    indexed_names = []

    for i, c in enumerate(column_schema):
        if c['indexed'] and c['type']:
            slug = slugify(c['name'])

            indexed_name = 'column_%s_%s' % (c['type'], slug)

            # Deduplicate within dataset
            if indexed_name in indexed_names:
                n = 2
                test_name = '%s%i' % (indexed_name, n)

                while test_name in indexed_names:
                    n += 1
                    test_name = '%s%i' % (indexed_name, n)

                indexed_name = test_name

            column_schema[i]['indexed_name'] = indexed_name
            indexed_names.append(indexed_name)

    return column_schema 

def make_column_schema(columns, indexed=None, types=None):
    """
    Generate a column schema from parallel arrays of columns, index booleans, and index types.
    """
    column_schema = []

    for i, name in enumerate(columns):
        c = {
            'name': name,
            'indexed': indexed[i] if indexed else False,
            'type': types[i] if types else None,
            'indexed_name': None,
            'min': None,
            'max': None
        }

        column_schema.append(c)

    column_schema = update_indexed_names(column_schema)

    return column_schema


########NEW FILE########
__FILENAME__ = csvdata
#!/usr/bin/env python

import codecs
from itertools import islice
from types import NoneType

from csvkit import CSVKitReader
from csvkit.sniffer import sniff_dialect as csvkit_sniff
from csvkit.typeinference import normalize_table
from django.conf import settings
from django.utils.translation import ugettext as _

from panda.exceptions import DataSamplingError, NotSniffableError

def sniff_dialect(path, encoding='utf-8'):
    with codecs.open(path, 'r', encoding=encoding) as f:
        try:
            csv_dialect = csvkit_sniff(f.read(settings.PANDA_SNIFFER_MAX_SAMPLE_SIZE))
        except UnicodeDecodeError:
            raise DataSamplingError(_('This CSV file contains characters that are not %s encoded. You need to input the correct encoding in order to import data from this file. Read our documentation about <a href="http://pandaproject.net/docs/determining-the-encoding-of-a-csv-file.html">determining the encoding of a CSV file</a>.') % (encoding))

        if not csv_dialect:
            raise NotSniffableError(_('CSV dialect could not be automatically inferred.')) 

        return {
            'lineterminator': csv_dialect.lineterminator,
            'skipinitialspace': csv_dialect.skipinitialspace,
            'quoting': csv_dialect.quoting,
            'delimiter': csv_dialect.delimiter,
            'quotechar': csv_dialect.quotechar,
            'doublequote': csv_dialect.doublequote
        }

def extract_column_names(path, dialect_parameters, encoding='utf-8'):
    with open(path, 'r') as f:
        reader = CSVKitReader(f, encoding=encoding, **dialect_parameters)

        try:
            headers = reader.next()
        except UnicodeDecodeError:
            raise DataSamplingError(_('This CSV file contains characters that are not %s encoded. You need to input the correct encoding in order to import data from this file.') % encoding)

        return headers

def sample_data(path, dialect_parameters, sample_size, encoding='utf-8'):
    with open(path, 'r') as f:
        reader = CSVKitReader(f, encoding=encoding, **dialect_parameters)

        try:
            reader.next() # skip headers
            samples = []

            for row in islice(reader, sample_size):
                samples.append(row)
        except UnicodeDecodeError:
            raise DataSamplingError(_('This CSV file contains characters that are not %s encoded. You need to input the correct encoding in order to import data from this file.') % (encoding))

        return samples

def guess_column_types(path, dialect, sample_size, encoding='utf-8'):
    """
    Guess column types based on a sample of data.
    """
    with open(path, 'r') as f:
        reader = CSVKitReader(f, encoding=encoding, **dialect)
        headers = reader.next()

        sample = islice(reader, sample_size)
        normal_types, normal_values = normalize_table(sample)

        type_names = []

        for t in normal_types:
            if t is NoneType:
                type_names.append(None)
            else:
                type_names.append(t.__name__)

        # If a final column had no values csvkit will have dropped it
        while len(type_names) < len(headers):
            type_names.append(None)

        return type_names 


########NEW FILE########
__FILENAME__ = mail
#!/usr/bin/env python

import logging
import socket

from django.core import mail
from livesettings import config_value

def get_connection():
    return mail.get_connection(
        host=config_value('EMAIL', 'EMAIL_HOST'),
        port=config_value('EMAIL', 'EMAIL_PORT'),
        # See http://bugs.python.org/issue8489
        username=str(config_value('EMAIL', 'EMAIL_HOST_USER')),
        password=str(config_value('EMAIL', 'EMAIL_HOST_PASSWORD')),
        use_tls=config_value('EMAIL', 'EMAIL_USE_TLS')) 

def send_mail(subject, message, recipients):
    log = logging.getLogger('panda.utils.mail')

    if not config_value('EMAIL', 'EMAIL_ENABLED'):
        log.info('Email is disabled, not sending message to %s' % recipients)
        return

    try:
        mail.send_mail('[PANDA] %s' % subject, message, str(config_value('EMAIL', 'DEFAULT_FROM_EMAIL')), recipients, connection=get_connection())
    except socket.error:
        log.error('Failed connecting to email server. (Sending to %s.)' % recipients)


########NEW FILE########
__FILENAME__ = notifications
#!/usr/bin/env python

from django.template.loader import get_template
from django.template import Context, TemplateDoesNotExist
from livesettings import config_value

from panda.utils.mail import send_mail

def get_message_template(prefix):
    return get_template('/'.join(['notifications', prefix, 'message.html']))

def get_email_subject_template(prefix):
    return get_template('/'.join(['notifications', prefix, 'email_subject.txt']))

def get_email_body_template(prefix):
    return get_template('/'.join(['notifications', prefix, 'email_body.txt']))

def notify(recipient, template_prefix, note_type, url=None, extra_context={}):
    """
    Notify a user of an event using the Notification system and
    email.
    """
    from panda.models import Notification

    context = Context({
        'recipient': recipient,
        'type': note_type,
        'url': url,
        'site_domain': config_value('DOMAIN', 'SITE_DOMAIN')
    })
    context.update(extra_context)

    message = get_message_template(template_prefix).render(context) 
    
    Notification.objects.create(
        recipient=recipient,
        message=message,
        type=note_type,
        url=url
    )

    # Don't HTML escape plain-text emails
    context.autoescape = False

    try:
        email_subject = get_email_subject_template(template_prefix).render(context)
    except TemplateDoesNotExist:
        email_subject = message

    try:
        email_message = get_email_body_template(template_prefix).render(context)
    except TemplateDoesNotExist:
        email_message = message
    
    send_mail(email_subject.strip(), email_message, [recipient.username])


########NEW FILE########
__FILENAME__ = solr
#!/usr/bin/env python

from uuid import uuid4

from django.utils import simplejson as json
from django.utils.timezone import now

def make_data_row(dataset, data, data_upload=None, external_id=None):
    last_modified = now().replace(microsecond=0, tzinfo=None)
    last_modified = last_modified.isoformat('T') + 'Z' 

    solr_row = {
        'dataset_slug': dataset.slug,
        'data_upload_id': data_upload.id if data_upload else None,
        'full_text': '\n'.join([unicode(d) for d in data]),
        'data': json.dumps(data),
        'last_modified': last_modified 
    }

    if external_id:
        solr_row['id'] = '%s-%s' % (dataset.slug, external_id)
        solr_row['external_id'] = external_id
    else:
        solr_row['id'] = unicode(uuid4())

    return solr_row


########NEW FILE########
__FILENAME__ = typecoercion
#!/usr/bin/env python
# -*- coding: utf-8 -*-

from datetime import date, time, datetime

from csvkit.typeinference import NULL_VALUES, TRUE_VALUES, FALSE_VALUES, DEFAULT_DATETIME
from dateutil.parser import parse
from django.utils.translation import ugettext as _

from panda.exceptions import TypeCoercionError

TYPE_NAMES_MAPPING = {
    'unicode': unicode,
    'int': int,
    'bool': bool,
    'float': float,
    'datetime': datetime,
    'date': date,
    'time': time
}

CURRENCY_SYMBOLS_ASCII = '$,'

# Via http://en.wikipedia.org/wiki/Currency_sign
CURRENCY_SYMBOLS_UNICODE_TRANSLATE_TABLE = dict([(ord(c), None) for c in '$,€£₱؋฿₵₡₫ƒ₣₲₴₭ლ₥₦£៛₹₪৳₮₩¥'])

class DataTyper(object):
    """
    A callable object that adds typed columns to a Solr object based on a Dataset schema.
    Along the way it also updates the schema based on the new data.
    """
    def __init__(self, schema):
        self.schema = schema

        # Min/max values for dates/times/datetimes get stored as strings and need to be coerced back
        for n, c in enumerate(self.schema):
            if c['indexed'] and c['type']:
                t = TYPE_NAMES_MAPPING[c['type']]
            
                if t in (date, time, datetime):
                    self.schema[n]['min'] = self.coerce_type(c['min'], datetime)
                    self.schema[n]['max'] = self.coerce_type(c['max'], datetime)

        self.errors = [[] for c in self.schema]

    def __call__(self, data, row):
        """
        Given a Solr data object and a row of data, will ad typed columns to the data
        object and then return it.
        """
        for n, c in enumerate(self.schema):
            if c['indexed'] and c['type']:
                try:
                    t = TYPE_NAMES_MAPPING[c['type']]
                    value = self.coerce_type(row[n], t)
                    data[c['indexed_name']] = value

                    if t in [int, float, date, time, datetime] and value is not None:
                        if c['min'] is None or value < c['min']:
                            self.schema[n]['min'] = value

                        if c['max'] is None or value > c['max']:
                            self.schema[n]['max'] = value
                except TypeCoercionError, e:
                    self.errors[n].append(e)

        return data
    
    def summarize(self):
        """
        Generate a plain-text summary of typing, suitable for an email notification.
        """
        if any([c['indexed'] for c in self.schema]):
            summary = 'Summary of column filters:\n\n'

            for n, c in enumerate(self.schema):
                if c['indexed'] and c['type']:
                    error_count = len(self.errors[n])

                    if not error_count:
                        summary += _('%(name)s: all values succesfully converted to type "%(type)s"\n') \
                            % {'name': c['name'], 'type': c['type']}
                    else:
                        summary += _('%(name)s: failed to convert %(error_count)i values to type "%(type)s"\n') \
                            % {'name': c['name'], 'error_count': error_count, 'type': c['type']}

            return summary
        else:
            return None

    def coerce_type(self, value, normal_type):
        """
        Coerce a single value into a type supported by PANDA.
        
        All one function for performance.
        """
        if isinstance(value, basestring) and value.lower() in NULL_VALUES:
            value = None

        # All types support nulls
        if value is None:
            return None

        try:
            # unicode
            if normal_type is unicode:
                return unicode(value)
            # int
            elif normal_type is int:
                # Filter currency symbols
                if isinstance(value, str):
                    value = value.translate(None, CURRENCY_SYMBOLS_ASCII)
                elif isinstance(value, unicode):
                    value = value.translate(CURRENCY_SYMBOLS_UNICODE_TRANSLATE_TABLE)

                return int(value) 
            # bool
            elif normal_type is bool:
                if isinstance(value, basestring):
                    lcase = value.lower()

                    if lcase in TRUE_VALUES:
                        value = True
                    elif lcase in FALSE_VALUES:
                        value = False
                    else:
                        raise ValueError()

                return bool(value)
            # float
            elif normal_type is float:
                # Filter currency symbols
                if isinstance(value, str):
                    value = value.translate(None, CURRENCY_SYMBOLS_ASCII)
                elif isinstance(value, unicode):
                    value = value.translate(CURRENCY_SYMBOLS_UNICODE_TRANSLATE_TABLE)

                return float(value)
            # date, time, datetime
            elif normal_type in [date, time, datetime]:
                # Don't parse empty strings!
                if not value:
                    raise ValueError()

                try:
                    d = parse(value, default=DEFAULT_DATETIME)
                except OverflowError:
                    raise ValueError()
                except TypeError:
                    raise ValueError()

                return d
        except ValueError:
            raise TypeCoercionError(value, normal_type)


########NEW FILE########
__FILENAME__ = xls
#!/usr/bin/env python

import datetime

from csvkit.convert.xls import determine_column_type
import xlrd
from django.utils.translation import ugettext as _

from panda.exceptions import TypeInferenceError

def sniff_dialect(path, **kwargs):
    return {} 

def extract_column_names(path, dialect, **kwargs):
    book = xlrd.open_workbook(path, on_demand=True)
    sheet = book.sheet_by_index(0)
    headers = sheet.row_values(0)

    return headers

def normalize_date(v, datemode):
    """
    Convert an xldate to a date, time, or datetime
    depending on its value.
    """
    v_tuple = xlrd.xldate_as_tuple(v, datemode)

    if v_tuple == (0, 0, 0, 0, 0, 0):
        # Midnight 
        dt = datetime.time(*v_tuple[3:])
    elif v_tuple[3:] == (0, 0, 0):
        # Date only
        dt = datetime.date(*v_tuple[:3])
    elif v_tuple[:3] == (0, 0, 0):
        # Time only
        dt = datetime.time(*v_tuple[3:])
    else:
        # Date and time
        dt = datetime.datetime(*v_tuple)

    return dt.isoformat()

def sample_data(path, dialect, sample_size, **kwargs):
    book = xlrd.open_workbook(path, on_demand=True)
    sheet = book.sheet_by_index(0)

    samples = []

    for i in range(1, min(sheet.nrows, sample_size + 1)):
        values = sheet.row_values(i)
        types = sheet.row_types(i)

        normal_values = []

        for v, t in zip(values, types):
            if t == xlrd.biffh.XL_CELL_DATE:
                v = normalize_date(v, book.datemode)
            elif t == xlrd.biffh.XL_CELL_NUMBER:
                if v % 1 == 0:
                    v = int(v)

            normal_values.append(unicode(v))

        samples.append(normal_values)

    return samples

def determine_number_type(values):
    """
    Determine if a column of numbers in an XLS file are integral.
    """
    # Test if all values are whole numbers, if so coerce floats it ints
    integral = True

    for v in values:
        if v and v % 1 != 0:
            integral = False
            break

    if integral:
        return int
    else:
        return float

def determine_date_type(values, datemode=0):
    """
    Determine if an Excel date column really contains dates... 
    """
    normal_types_set = set()

    for v in values:
        # Skip blanks 
        if v == '':
            continue

        v_tuple = xlrd.xldate_as_tuple(v, datemode)

        if v_tuple == (0, 0, 0, 0, 0, 0):
            # Midnight 
            normal_types_set.add(datetime.time)
        elif v_tuple[3:] == (0, 0, 0):
            # Date only
            normal_types_set.add(datetime.date)
        elif v_tuple[:3] == (0, 0, 0):
            # Time only
            normal_types_set.add(datetime.time)
        else:
            # Date and time
            normal_types_set.add(datetime.datetime)

    if len(normal_types_set) == 1:
        # No special handling if column contains only one type
        return normal_types_set.pop()
    elif normal_types_set == set([datetime.datetime, datetime.date]):
        # If a mix of dates and datetimes, up-convert dates to datetimes
        return datetime.datetime
    elif normal_types_set == set([datetime.datetime, datetime.time]):
        # Datetimes and times don't mix
        return unicode
    elif normal_types_set == set([datetime.date, datetime.time]):
        # Dates and times don't mix
        return unicode

def guess_column_types(path, dialect, sample_size, encoding='utf-8'):
    """
    Guess column types based on a sample of data.
    """
    book = xlrd.open_workbook(path, on_demand=True)
    sheet = book.sheet_by_index(0)

    column_types = []

    for i in range(sheet.ncols):
        values = sheet.col_values(i)[1:sample_size + 1]
        types = sheet.col_types(i)[1:sample_size + 1]
        nominal_type = determine_column_type(types)

        if nominal_type == xlrd.biffh.XL_CELL_EMPTY:
            column_types.append(None)
        elif nominal_type == xlrd.biffh.XL_CELL_TEXT:
            column_types.append(unicode)
        elif nominal_type == xlrd.biffh.XL_CELL_NUMBER:
            column_types.append(determine_number_type(values))
        elif nominal_type == xlrd.biffh.XL_CELL_DATE:
            column_types.append(determine_date_type(values, datemode=book.datemode))
        elif nominal_type == xlrd.biffh.XL_CELL_BOOLEAN:
            column_types.append(bool)
        elif nominal_type == xlrd.biffh.XL_CELL_ERROR:
            column_types.append(unicode)
        else:
            raise TypeInferenceError(_('Unknown column type found in xls file: %s') % nominal_type) 

    return [t.__name__ if t else None for t in column_types]


########NEW FILE########
__FILENAME__ = xlsx
#!/usr/bin/env python

import datetime
from itertools import islice
from types import NoneType

from csvkit.typeinference import NULL_TIME
from openpyxl.reader.excel import load_workbook

def sniff_dialect(path, **kwargs):
    return {}

def extract_column_names(path, dialect, **kwargs):
    book = load_workbook(path, use_iterators=True)
    sheet = book.get_active_sheet()
    headers = sheet.iter_rows().next()

    return [unicode(h.internal_value) if h.internal_value is not None else '' for h in headers]

def normalize_date(dt):
    if dt.time() == NULL_TIME:
        return dt.date().isoformat()

    if dt.microsecond == 0:
        return dt.isoformat()

    ms = dt.microsecond

    if ms < 1000:
        return dt.replace(microsecond=0).isoformat()
    elif ms > 999000:
        return dt.replace(second=dt.second + 1, microsecond=0).isoformat()

    return dt.isoformat()

def sample_data(path, dialect, sample_size, **kwargs):
    book = load_workbook(path, use_iterators=True)
    sheet = book.get_active_sheet()

    samples = []

    for i, row in enumerate(sheet.iter_rows()):
        if i == 0:
            continue

        if i == sample_size + 1:
            break

        values = []

        for c in row:
            value = c.internal_value

            if value.__class__ is datetime.datetime:
                value = normalize_date(value)
            elif value.__class__ is float:
                if value % 1 == 0:
                    value = int(value)

            if value.__class__ in (datetime.datetime, datetime.date, datetime.time):
                value = value.isoformat()

            values.append(unicode(value))

        samples.append(values)

    return samples

def determine_column_type(types):
    """
    Determine the correct type for a column from a list of cell types.
    """
    types_set = set(types)
    types_set.discard(NoneType)

    # Normalize mixed types to text
    if len(types_set) > 1:
        return unicode

    try:
        return types_set.pop()
    except KeyError:
        return None 

def determine_number_type(values):
    """
    Determine if a column of numbers in an XLS file are integral.
    """
    # Test if all values are whole numbers, if so coerce floats it ints
    integral = True

    for v in values:
        if v and v % 1 != 0:
            integral = False
            break

    if integral:
        return int
    else:
        return float

def determine_date_type(values):
    """
    Determine if a column of numbers in an XLS file are only dates.
    """
    if any([dt and dt.time() != NULL_TIME for dt in values]):
        return datetime.datetime
    else:
        return datetime.date

def guess_column_types(path, dialect, sample_size, encoding='utf-8'):
    """
    Guess column types based on a sample of data.
    """
    book = load_workbook(path, use_iterators=True)
    sheet = book.get_active_sheet()

    rows = islice(sheet.iter_rows(), 0, sample_size + 1)
    rows.next()

    columns = zip(*rows)
    column_types = []

    for column in columns:
        values = [c.internal_value for c in column]

        t = determine_column_type([v.__class__ for v in values])

        if t is float:
            t = determine_number_type(values) 
        elif t is datetime.datetime:
            t = determine_date_type(values)

        column_types.append(t)

    return [c.__name__ if c else None for c in column_types]


########NEW FILE########
__FILENAME__ = views
#!/usr/bin/env python

import os

from ajaxuploader.views import AjaxFileUploader
from csvkit.exceptions import FieldSizeLimitError
from django.conf import settings
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponse
from django.utils.timezone import now
from django.utils.translation import ugettext as _
from livesettings import config_value
from tastypie.bundle import Bundle
from tastypie.serializers import Serializer

from client.utils import get_free_disk_space
from panda.api.notifications import NotificationResource
from panda.api.users import UserValidation
from panda.api.utils import PandaAuthentication
from panda.models import UserProfile, UserProxy
from panda.storage import PANDADataUploadBackend, PANDARelatedUploadBackend
from panda.utils.mail import send_mail

class JSONResponse(HttpResponse):
    """
    A shortcut for an HTTPResponse containing data serialized as json.

    Note: Uses Tastypie's serializer to transparently support serializing bundles.
    """
    def __init__(self, contents, **kwargs):
        serializer = Serializer()

        super(JSONResponse, self).__init__(serializer.to_json(contents), content_type='application/json', **kwargs)
                
class SecureAjaxFileUploader(AjaxFileUploader):
    """
    A custom version of AjaxFileUploader that checks for authorization.
    """
    def __call__(self, request):
        auth = PandaAuthentication()

        if auth.is_authenticated(request) != True:
            # Valum's FileUploader only parses the response if the status code is 200.
            return JSONResponse({ 'success': False, 'forbidden': True }, status=200)

        try:
            return self._ajax_upload(request)
        except FieldSizeLimitError:
            return JSONResponse({ 'error_message': _('CSV contains fields longer than maximum length of 131072 characters.') })
        except Exception, e:
            return JSONResponse({ 'error_message': unicode(e) })

data_upload = SecureAjaxFileUploader(backend=PANDADataUploadBackend)
related_upload = SecureAjaxFileUploader(backend=PANDARelatedUploadBackend)

def make_user_login_response(user):
    """
    Generate a response to a login request.
    """
    nr = NotificationResource()

    notifications = user.notifications.all()[:settings.PANDA_NOTIFICATIONS_TO_SHOW]

    bundles = [nr.build_bundle(obj=n) for n in notifications]
    notifications = [nr.full_dehydrate(b) for b in bundles]

    return {
        'id': user.id,
        'email': user.email,
        'is_staff': user.is_staff,
        'show_login_help': user.get_profile().show_login_help,
        'notifications': notifications
    }

def panda_login(request):
    """
    PANDA login: takes a username and password and returns an API key
    for querying the API.
    """
    if request.method == 'POST':
        email = request.POST['email']
        password = request.POST['password']

        user = authenticate(username=email.lower(), password=password)

        if user is not None:
            # Convert authenticated user to a proxy model
            _user_proxy = UserProxy()
            _user_proxy.__dict__ = user.__dict__
            user = _user_proxy

            if user.is_active:
                login(request, user)

                # Success
                return JSONResponse(make_user_login_response(user))
            else:
                # Disabled account
                return JSONResponse({ '__all__': _('This account is disabled') }, status=400)
        else:
            # Invalid login
            return JSONResponse({ '__all__': _('Email or password is incorrect') }, status=400)
    else:
        # Invalid request
        return JSONResponse(None, status=400)

def panda_logout(request):
    """
    Logout any active session.
    """
    logout(request)

    return JSONResponse({ '__all__': _('Successfully logged out') }, status=200)

def check_activation_key(request, activation_key):
    """
    Test if an activation key is valid and if so fetch information
    about the user to populate the form.
    """
    try:
        user_profile = UserProfile.objects.get(activation_key=activation_key)
    except UserProfile.DoesNotExist:
        return JSONResponse({ '__all__': _('Invalid activation key') }, status=400)

    user = user_profile.user 

    if user_profile.activation_key_expiration <= now():
        return JSONResponse({ '__all__': _('Expired activation key. Contact your administrator') }, status=400)

    return JSONResponse({
        'activation_key': user_profile.activation_key,
        'email': user.email,
        'first_name': user.first_name,
        'last_name': user.last_name
    })

def activate(request):
    """
    PANDA user activation.
    """
    if request.method == 'POST':
        validator = UserValidation()

        data = dict([(k, v) for k, v in request.POST.items()])

        try:
            user_profile = UserProfile.objects.get(activation_key=data['activation_key'])
        except UserProfile.DoesNotExist:
            return JSONResponse({ '__all__': _('Invalid activation key!') }, status=400)

        user = user_profile.user

        if user_profile.activation_key_expiration <= now():
            return JSONResponse({ '__all__': _('Expired activation key. Contact your administrator.') }, status=400)

        if 'password' not in data:
            return JSONResponse({ 'password': _('This field is required.') }, status=400)

        if 'reenter_password' in data:
            del data['reenter_password']

        bundle = Bundle(data=data)

        errors = validator.is_valid(bundle)

        if errors:
            return JSONResponse(errors, status=400) 

        user.username = bundle.data['email']
        user.email = bundle.data['email']
        user.first_name = bundle.data.get('first_name', '')
        user.last_name = bundle.data.get('last_name', '')
        user.set_password(bundle.data['password'])
        user.is_active = True

        user.save()

        user_profile.activation_key_expiration = now()
        user_profile.save()

        # Success
        return JSONResponse(make_user_login_response(user))
    else:
        # Invalid request
        return JSONResponse(None, status=400)

def forgot_password(request):
    """
    PANDA user password reset and notification.
    """
    if request.method == 'POST':
        try:
            user = UserProxy.objects.get(email=request.POST.get('email'))
        except UserProfile.DoesNotExist:
            return JSONResponse({ '__all__': _('Unknown or inactive email address.') }, status=400)

        if not user.is_active:
            return JSONResponse({ '__all__': _('Unknown or inactive email address.') }, status=400)

        user_profile = user.get_profile()
        user_profile.generate_activation_key()
        user_profile.save()

        email_subject = _('Forgotten password')
        email_body = _('PANDA received a request to change your password.\n\nTo set your new password follow this link:\n\nhttp://%(site_domain)s/#reset-password/%(activation_key)s\n\nIf you did not request this email you should notify your administrator.') \
            % {'site_domain': config_value('DOMAIN', 'SITE_DOMAIN'), 'activation_key': user_profile.activation_key}

        send_mail(email_subject,
                  email_body,
                  [user.email])

        # Success
        return JSONResponse(make_user_login_response(user))
    else:
        # Invalid request
        return JSONResponse(None, status=400)

def check_available_space(request):
    """
    Check the amount of space left on each disk.
    """
    return JSONResponse({
        'root': {
            'device': os.stat('/').st_dev,
            'free_space': get_free_disk_space('/')
        },
        'uploads': {
            'device':  os.stat(settings.MEDIA_ROOT).st_dev,
            'free_space': get_free_disk_space(settings.MEDIA_ROOT)
        },
        'indices': {
            'device': os.stat(settings.SOLR_DIRECTORY).st_dev,
            'free_space': get_free_disk_space(settings.SOLR_DIRECTORY)
        }
    })


########NEW FILE########
__FILENAME__ = audit_uploads
#!/usr/bin/env python

"""
Review the database and filesystem to identify possible orphaned files and datasets.
"""

import psycopg2
import os, os.path
from sys import argv

UPLOAD_DIRECTORY = '/var/lib/panda/uploads'
DATABASE='panda'
DB_USER='panda'
DB_PASSWORD='panda'

files = os.listdir(UPLOAD_DIRECTORY)
connection = psycopg2.connect(database=DATABASE,user=DB_USER,password=DB_PASSWORD)
cursor = connection.cursor()

cursor.execute("""select du.id, du.filename, ds.name from panda_dataupload du 
                  left outer join panda_dataset ds on (du.dataset_id = ds.id)""")

files_to_delete = []
uploads_to_delete = []
for id, filename, dataset_name in cursor:
    if dataset_name:
        files.remove(filename)
    if not dataset_name:
        print "Orphaned upload: %i" % id
        uploads_to_delete.append(id)
        try:
            files.remove(filename)
            files_to_delete.append(filename)
        except ValueError: pass
for f in files:
    print "not even a dataupload for %s" % f
    files_to_delete.append(f)

if uploads_to_delete:
    print "Data upload IDs with no corresponding dataset:"
    print "\n".join(map(str,uploads_to_delete))
    print
if files_to_delete:
    print "Files with no corresponding dataset:"
    print "\n".join(files_to_delete)
    print

if files_to_delete and uploads_to_delete:
    question = "Enter 'yes' to delete ALL of these files and data upload records: "
elif files_to_delete:
    question = "Enter 'yes' to delete ALL of these files: "
elif uploads_to_delete:
    question = "Enter 'yes' to delete ALL of these data upload records: "
else:
    question = None

if question:
    do_it = raw_input()
    if do_it == 'yes':
        for f in files_to_delete:
            os.remove(os.path.join(UPLOAD_DIRECTORY,f))
        for id in uploads_to_delete:
            cursor.execute("delete from panda_dataupload where id = %s",(id,))
        connection.commit()
        cursor.close()
        print "Cleanup complete."
    else:
        "No changes were made to the database or filesystem."
else:
    print "Everything looks pretty clean."

connection.close()

    


########NEW FILE########
__FILENAME__ = backup_volumes
#!/usr/bin/env python

"""
Comprehensive script to handle backing up PANDA's EBS volumes.
"""

from datetime import datetime
from getpass import getpass
import os
import subprocess
import sys
import time

from boto.ec2.connection import EC2Connection
from boto.utils import get_instance_metadata

# Utilities
def safe_dismount(mount_point):
    dismounted = False

    while not dismounted:
        try:
            subprocess.check_output(['umount', mount_point], stderr=subprocess.STDOUT)
            dismounted = True
        except:
            time.sleep(1)

def mount_point_from_device_name(device):
    df = subprocess.check_output(['df', device])
    return df.split()[-1]

# Sanity checks
if not os.geteuid() == 0:
    sys.exit('You must run this script with sudo!')

metadata = get_instance_metadata()
instance_id = metadata['instance-id'] 

# Prompt for parameters
aws_key = getpass('Enter your AWS Access Key: ')
secret_key = getpass('Enter your AWS Secret Key: ')

print 'Beginning PANDA backup'

sys.stdout.write('Connecting to EC2... ')
conn = EC2Connection(aws_key, secret_key)
print 'connected'

sys.stdout.write('Identifying attached volumes...')
volumes = [v for v in conn.get_all_volumes() if v.attach_data.instance_id == instance_id]
print volumes

sys.stdout.write('Stopping services... ')
subprocess.check_output(['service', 'uwsgi', 'stop'], stderr=subprocess.STDOUT)
subprocess.check_output(['service', 'celeryd', 'stop'], stderr=subprocess.STDOUT)
subprocess.check_output(['service', 'nginx', 'stop'], stderr=subprocess.STDOUT)
subprocess.check_output(['service', 'postgresql', 'stop'], stderr=subprocess.STDOUT)
subprocess.check_output(['service', 'solr', 'stop'], stderr=subprocess.STDOUT)
print 'stopped'

for v in volumes:
    device = v.attach_data.device.replace('/dev/sd', '/dev/xvd')
    mount_point = mount_point_from_device_name(device)

    description = 'PANDA backup of %s, attached at %s and mounted at %s (created at %s)' % (v.id, v.attach_data.device, mount_point, datetime.today().isoformat(' '))
    
    sys.stdout.write('Creating snapshot of %s (%s)... ' % (v.id, mount_point))
    v.create_snapshot(description)

    snapshot = conn.get_all_snapshots(filters={ 'description': description})[0]

    while snapshot.status == 'pending':
        time.sleep(2)
        snapshot.update()

    print 'created'

sys.stdout.write('Restarting services... ')
subprocess.check_output(['service', 'solr', 'start'], stderr=subprocess.STDOUT)
subprocess.check_output(['service', 'postgresql', 'start'], stderr=subprocess.STDOUT)
subprocess.check_output(['service', 'nginx', 'start'], stderr=subprocess.STDOUT)
subprocess.check_output(['service', 'celeryd', 'start'], stderr=subprocess.STDOUT)
subprocess.check_output(['service', 'uwsgi', 'start'], stderr=subprocess.STDOUT)
print 'restarted'

print 'Done!'


########NEW FILE########
__FILENAME__ = check_po
#!/usr/bin/env python
import os, os.path
import re

import sys


placeholder_pat = re.compile("%\((.+?)\)(\w)")
def extract_placeholders(s):
    return set(placeholder_pat.findall(s))

def check_file(fn):
    msgid = ''
    msgstr = ''
    workingon = 'msgid'
    mismatches = []
    for line in open(fn):
        if line.startswith('#'): continue
        text = ''
        line = line.rstrip()
        if line.startswith('msg'):
            workingon, text = line.split(' ',1)
            if workingon == 'msgid':
                if msgid and msgstr and len(msgstr.strip()) > 0:
                    id_placeholders = extract_placeholders(msgid)
                    str_placeholders = extract_placeholders(msgstr)
                    if len(id_placeholders) != len(str_placeholders) or (len(id_placeholders.difference(str_placeholders)) != 0):
                        mismatches.append((msgid,msgstr))
                msgid = msgstr = ''
        else:
            text = line
        text = text.strip('"')
        if text:
            if workingon == 'msgid':
                msgid += text
            else:
                msgstr += text

    if mismatches:
        print "WARNING: %i mismatches in %s" % (len(mismatches),fn)
        for msgid, msgstr in mismatches:
            print 'msgid:' + msgid
            print 'msgstr:' + msgstr
            print


if __name__ == '__main__':
    try:
        start_dir = sys.argv[1]
    except:
        start_dir = '../locale'

    for path, dirs, files in os.walk(start_dir):
        for f in files:
            if f.endswith('.po'):
                check_file(os.path.join(path,f))

########NEW FILE########
__FILENAME__ = migrate_files_volume
#!/usr/bin/env python

"""
Comprehensive script to handle migrating PANDA's files (uploads and exports)
to a larger EBS volume.
Handles all stages of device creation, attachment, file movement, etc.
It will work whether the files are currently on another EBS or on local storage.

The only thing this script does not do is detach and destroy any old volume.
"""

from getpass import getpass
import os
import shutil
import string
import subprocess
import sys
import time
from datetime import datetime

from boto.ec2.connection import EC2Connection
from boto.utils import get_instance_metadata

TEMP_MOUNT_POINT = '/mnt/filesmigration'
PANDA_DIR = '/var/lib/panda'
FSTAB_BACKUP = '/etc/fstab.filesmigration.bak'

# Utilities
def safe_dismount(mount_point):
    dismounted = False

    while not dismounted:
        try:
            subprocess.check_output(['umount', mount_point], stderr=subprocess.STDOUT)
            dismounted = True
        except:
            time.sleep(1)

# Sanity checks
if not os.geteuid() == 0:
    sys.exit('You must run this script with sudo!')

metadata = get_instance_metadata()

backed_up = raw_input('Migrating your PANDA files is a complicated and potentially destructive operation. Have you backed up your data? (y/N): ')

if backed_up.lower() != 'y':
    sys.exit('Back up your data before running this script! Aborting.')

# Prompt for parameters
aws_key = getpass('Enter your AWS Access Key: ')
secret_key = getpass('Enter your AWS Secret Key: ')
size_gb = raw_input('How many GB would you like your new PANDA files volume to be? ')

print 'Beginning PANDA files migration'

sys.stdout.write('Connecting to EC2... ')
conn = EC2Connection(aws_key, secret_key)
print 'connected'

sys.stdout.write('Identifying running instance... ')
instance_id = metadata['instance-id'] 

reservations = conn.get_all_instances()

instance = None

for r in reservations:
    for i in r.instances:
        if i.id == instance_id:
            instance = i
            break

    if instance:
        break

if not instance:
    sys.exit('Unable to determine running instance! Aborting.')

print instance_id

sys.stdout.write('Creating new volume... ')
vol = conn.create_volume(size_gb, instance.placement)
conn.create_tags([vol.id], {'Name': 'PANDA Uploads volume %s' % datetime.now().strftime('%Y-%m-%d')})
print vol.id

sys.stdout.write('Waiting for volume to become available... ')

while not vol.status == 'available':
    time.sleep(5)
    vol.update()

print 'available'

sys.stdout.write('Backing up fstab... ')
shutil.copy2('/etc/fstab', FSTAB_BACKUP)
print FSTAB_BACKUP

sys.stdout.write('Finding an available device path... ')
ec2_device_name = None
device_path = None

for letter in string.lowercase[6:]:
    ec2_device_name = '/dev/sd%s' % letter
    device_path = '/dev/xvd%s' % letter

    if not os.path.exists(device_path):
        break

print device_path

sys.stdout.write('Attaching new volume... ')
vol.attach(instance.id, ec2_device_name) 

while not os.path.exists(device_path):
    time.sleep(1)
print 'attached'

sys.stdout.write('Formatting volume... ')
subprocess.check_output(['mkfs.ext3', device_path], stderr=subprocess.STDOUT)
print 'formatted'

sys.stdout.write('Creating temporary mount point... ')
if os.path.exists(TEMP_MOUNT_POINT):
    shutil.rmtree(TEMP_MOUNT_POINT)

os.mkdir(TEMP_MOUNT_POINT)
print TEMP_MOUNT_POINT

sys.stdout.write('Mounting volume... ')
subprocess.check_output(['mount', device_path, TEMP_MOUNT_POINT], stderr=subprocess.STDOUT)
print 'mounted' 

sys.stdout.write('Stopping services... ')
subprocess.check_output(['service', 'uwsgi', 'stop'], stderr=subprocess.STDOUT)
subprocess.check_output(['service', 'celeryd', 'stop'], stderr=subprocess.STDOUT)
print 'stopped'

sys.stdout.write('Copying indexes... ')
names = os.listdir(PANDA_DIR)

for name in names:
    if name == 'lost+found':
        continue

    src_path = os.path.join(PANDA_DIR, name)
    dest_path = os.path.join(TEMP_MOUNT_POINT, name)

    if os.path.isdir(src_path):
        shutil.copytree(src_path, dest_path)
    else:
        shutil.copy2(src_path, dest_path)

print 'copied'

if os.path.ismount(PANDA_DIR):
    sys.stdout.write('Dismounting old storage device... ')
    safe_dismount(PANDA_DIR)
    print 'dismounted'

    sys.stdout.write('Removing device from fstab... ')
    new_fstab = subprocess.check_output(['grep', '-Ev', PANDA_DIR, '/etc/fstab'], stderr=subprocess.STDOUT)
    print 'removed'

    with open('/etc/fstab', 'w') as f:
        f.write(new_fstab)
else:
    sys.stdout.write('Removing old indexes... ')
    shutil.rmtree(PANDA_DIR)
    os.mkdir(PANDA_DIR)

    print 'removed'

sys.stdout.write('Dismounting from temporary mount point...')
safe_dismount(TEMP_MOUNT_POINT)
print 'dismounted'

sys.stdout.write('Remounting at final mount point... ')
subprocess.check_output(['mount', device_path, PANDA_DIR], stderr=subprocess.STDOUT)
print 'mounted'

sys.stdout.write('Reseting permissions... ')
subprocess.check_output(['chown', '-R', 'panda:panda', PANDA_DIR], stderr=subprocess.STDOUT)
print 'reset'

sys.stdout.write('Restarting services... ')
subprocess.check_output(['service', 'celeryd', 'start'], stderr=subprocess.STDOUT)
subprocess.check_output(['service', 'uwsgi', 'start'], stderr=subprocess.STDOUT)
print 'restarted'

sys.stdout.write('Configuring fstab... ')
with open('/etc/fstab', 'a') as f:
    f.write('\n%s\t%s\text3\tdefaults,noatime\t0\t0\n' % (device_path, PANDA_DIR))
print 'configured'

print 'Done!'


########NEW FILE########
__FILENAME__ = migrate_solr_volume
#!/usr/bin/env python

"""
Comprehensive script to handle migrating PANDA's Solr indices to a larger EBS volume.
Handles all stages of device creation, attachment, file movement, etc.
It will work whether the indices are currently on another EBS or on local storage.

The only thing this script does not do is detach and destroy any old volume.
"""

from getpass import getpass
import os
import shutil
import string
import subprocess
import sys
import time
from datetime import datetime

from boto.ec2.connection import EC2Connection
from boto.utils import get_instance_metadata

TEMP_MOUNT_POINT = '/mnt/solrmigration'
SOLR_DIR = '/opt/solr/panda/solr'
FSTAB_BACKUP = '/etc/fstab.solrmigration.bak'

# Utilities
def safe_dismount(mount_point):
    dismounted = False

    while not dismounted:
        try:
            subprocess.check_output(['umount', mount_point], stderr=subprocess.STDOUT)
            dismounted = True
        except:
            time.sleep(1)

# Sanity checks
if not os.geteuid() == 0:
    sys.exit('You must run this script with sudo!')

metadata = get_instance_metadata()

backed_up = raw_input('Migrating your Solr indexes is a complicated and potentially destructive operation. Have you backed up your data? (y/N): ')

if backed_up.lower() != 'y':
    sys.exit('Back up your data before running this script! Aborting.')

# Prompt for parameters
aws_key = getpass('Enter your AWS Access Key: ')
secret_key = getpass('Enter your AWS Secret Key: ')
size_gb = raw_input('How many GB would you like your new Solr volume to be? ')

print 'Beginning Solr migration'

sys.stdout.write('Connecting to EC2... ')
conn = EC2Connection(aws_key, secret_key)
print 'connected'

sys.stdout.write('Identifying running instance... ')
instance_id = metadata['instance-id'] 

reservations = conn.get_all_instances()

instance = None

for r in reservations:
    for i in r.instances:
        if i.id == instance_id:
            instance = i
            break

    if instance:
        break

if not instance:
    sys.exit('Unable to determine running instance! Aborting.')

print instance_id

sys.stdout.write('Creating new volume... ')
vol = conn.create_volume(size_gb, instance.placement)
conn.create_tags([vol.id], {'Name': 'PANDA Solr volume %s' % datetime.now().strftime('%Y-%m-%d')})
print vol.id

sys.stdout.write('Waiting for volume to become available... ')

while not vol.status == 'available':
    time.sleep(5)
    vol.update()

print 'available'

sys.stdout.write('Backing up fstab... ')
shutil.copy2('/etc/fstab', FSTAB_BACKUP)
print FSTAB_BACKUP

sys.stdout.write('Finding an available device path... ')
ec2_device_name = None
device_path = None

for letter in string.lowercase[6:]:
    ec2_device_name = '/dev/sd%s' % letter
    device_path = '/dev/xvd%s' % letter

    if not os.path.exists(device_path):
        break

print device_path

sys.stdout.write('Attaching new volume... ')
vol.attach(instance.id, ec2_device_name) 

while not os.path.exists(device_path):
    time.sleep(1)
print 'attached'

sys.stdout.write('Formatting volume... ')
subprocess.check_output(['mkfs.ext3', device_path], stderr=subprocess.STDOUT)
print 'formatted'

sys.stdout.write('Creating temporary mount point... ')
if os.path.exists(TEMP_MOUNT_POINT):
    shutil.rmtree(TEMP_MOUNT_POINT)

os.mkdir(TEMP_MOUNT_POINT)
print TEMP_MOUNT_POINT

sys.stdout.write('Mounting volume... ')
subprocess.check_output(['mount', device_path, TEMP_MOUNT_POINT], stderr=subprocess.STDOUT)
print 'mounted' 

sys.stdout.write('Stopping services... ')
subprocess.check_output(['service', 'uwsgi', 'stop'], stderr=subprocess.STDOUT)
subprocess.check_output(['service', 'celeryd', 'stop'], stderr=subprocess.STDOUT)
subprocess.check_output(['service', 'solr', 'stop'], stderr=subprocess.STDOUT)
print 'stopped'

sys.stdout.write('Copying indexes... ')
names = os.listdir(SOLR_DIR)

for name in names:
    if name == 'lost+found':
        continue

    src_path = os.path.join(SOLR_DIR, name)
    dest_path = os.path.join(TEMP_MOUNT_POINT, name)

    if os.path.isdir(src_path):
        shutil.copytree(src_path, dest_path)
    else:
        shutil.copy2(src_path, dest_path)

print 'copied'

if os.path.ismount(SOLR_DIR):
    sys.stdout.write('Dismounting old storage device... ')
    safe_dismount(SOLR_DIR)
    print 'dismounted'

    sys.stdout.write('Removing device from fstab... ')
    new_fstab = subprocess.check_output(['grep', '-Ev', SOLR_DIR, '/etc/fstab'], stderr=subprocess.STDOUT)
    print 'removed'

    with open('/etc/fstab', 'w') as f:
        f.write(new_fstab)
else:
    sys.stdout.write('Removing old indexes... ')
    shutil.rmtree(SOLR_DIR)
    os.mkdir(SOLR_DIR)

    print 'removed'

sys.stdout.write('Dismounting from temporary mount point...')
safe_dismount(TEMP_MOUNT_POINT)
print 'dismounted'

sys.stdout.write('Remounting at final mount point... ')
subprocess.check_output(['mount', device_path, SOLR_DIR], stderr=subprocess.STDOUT)
print 'mounted'

sys.stdout.write('Reseting permissions... ')
subprocess.check_output(['chown', '-R', 'solr:solr', SOLR_DIR], stderr=subprocess.STDOUT)
print 'reset'

sys.stdout.write('Restarting services... ')
subprocess.check_output(['service', 'solr', 'start'], stderr=subprocess.STDOUT)
subprocess.check_output(['service', 'celeryd', 'start'], stderr=subprocess.STDOUT)
subprocess.check_output(['service', 'uwsgi', 'start'], stderr=subprocess.STDOUT)
print 'restarted'

sys.stdout.write('Configuring fstab... ')
with open('/etc/fstab', 'a') as f:
    f.write('\n%s\t%s\text3\tdefaults,noatime\t0\t0\n' % (device_path, SOLR_DIR))
print 'configured'

print 'Done!'


########NEW FILE########
