__FILENAME__ = counting_sort
#!/usr/bin/env python

import collections
import sys


def usage():
    sys.stderr.write("USAGE: counting-sort [FILE]\n")
    sys.exit(1)


def counting_sort(input_stream, output_stream):
    buckets = collections.defaultdict(int)
    for key in input_stream:
        buckets[key] += 1
    for key in sorted(buckets.keys()):
        for _ in xrange(0, buckets[key]):
            output_stream.write(key)

if __name__ == '__main__':
    if len(sys.argv) == 1:
        counting_sort(sys.stdin, sys.stdout)
    elif len(sys.argv) == 2:
        if sys.argv[1] == '--help':
            usage()
        else:
            with open(sys.argv[1]) as input_stream:
                counting_sort(input_stream, sys.stdout)
    else:
        usage()

########NEW FILE########
__FILENAME__ = csv_to_json
#!/usr/bin/env python

import argparse
import codecs
import csv
import json
import sys

ENCODING = 'utf-8'

sys.stdin = codecs.getreader(ENCODING)(sys.stdin)
sys.stdout = codecs.getwriter(ENCODING)(sys.stdout)
sys.stderr = codecs.getwriter(ENCODING)(sys.stderr)


def utf_8_encoder(unicode_csv_data):
    for line in unicode_csv_data:
        yield line.encode('utf-8')


def unicode_csv_reader(unicode_csv_data, dialect=csv.excel, **kwargs):
    csv_reader = csv.reader(utf_8_encoder(unicode_csv_data),
                            dialect=dialect, **kwargs)
    for row in csv_reader:
        yield [unicode(cell, 'utf-8') for cell in row]


def csv_to_json(input_stream, output_stream, header_str, delimiter, quotechar):
    reader = unicode_csv_reader(input_stream,
                                delimiter=delimiter,
                                quotechar=quotechar)

    if header_str:
        header = header_str.split(',')
    else:
        header = reader.next()

    for row in reader:
        output_stream.write(json.dumps(dict(zip(header, row))))
        output_stream.write('\n')


if __name__ == '__main__':
    parser = argparse.ArgumentParser()

    parser.add_argument('input', nargs='?')
    parser.add_argument('--delimiter', '-d',
                        dest='delimiter',
                        default=',')
    parser.add_argument('--header',
                        dest='header',
                        metavar='NAME[,NAME..]')
    parser.add_argument('--quotechar', '-q',
                        dest='quotechar',
                        default='"')

    args = parser.parse_args()

    if args.input:
        f = codecs.open(args.input, encoding=ENCODING)
    else:
        f = sys.stdin

    csv_to_json(f, sys.stdout, args.header, args.delimiter, args.quotechar)

########NEW FILE########
__FILENAME__ = csv_to_tsv
#!/usr/bin/env python

# Convert CSV and similar formats to TSV (tab and newline delimited).

import argparse
import codecs
import csv
import re
import sys

# Unicode Newline Guidelines:
#
# http://www.unicode.org/standard/reports/tr13/tr13-5.html

STRIPPER_CHARS = u"\f\n\r\t\v\x85\u2028\u2029"
PROHIBITED_REGEX = re.compile(u'([\f\n\r\t\v\x85\u2028\u2029])')
SPACES_REGEX = re.compile(' +')
ENCODING = 'utf-8'

sys.stdin = codecs.getreader(ENCODING)(sys.stdin)
sys.stdout = codecs.getwriter(ENCODING)(sys.stdout)
sys.stderr = codecs.getwriter(ENCODING)(sys.stderr)


def utf_8_encoder(unicode_csv_data):
    for line in unicode_csv_data:
        yield line.encode('utf-8')


def unicode_csv_reader(unicode_csv_data, dialect=csv.excel, **kwargs):
    csv_reader = csv.reader(utf_8_encoder(unicode_csv_data),
                            dialect=dialect, **kwargs)
    for row in csv_reader:
        yield [unicode(cell, 'utf-8') for cell in row]


def stripper(row):
    return [field.translate(None, STRIPPER_CHARS) for field in row]


def escape(field):

    str_builder = []

    for ch in field:
        if ch == '\\':
            str_builder.append('\\')
            str_builder.append('\\')
        elif ch == '\f':
            str_builder.append('\\')
            str_builder.append('f')
        elif ch == '\n':
            str_builder.append('\\')
            str_builder.append('n')
        elif ch == '\r':
            str_builder.append('\\')
            str_builder.append('r')
        elif ch == '\t':
            str_builder.append('\\')
            str_builder.append('t')
        elif ch == '\v':
            str_builder.append('\\')
            str_builder.append('v')
        elif ch == '\x85':
            str_builder.append('\\')
            str_builder.append('x85')
        elif ch == u'\u2028':
            str_builder.append('\\')
            str_builder.append('u2028')
        elif ch == u'\u2029':
            str_builder.append('\\')
            str_builder.append('u2029')
        else:
            str_builder.append(ch)

    return ''.join(str_builder)


def escaper(row):
    return [escape(field) for field in row]


def replacer(row):

    return [PROHIBITED_REGEX.sub(' ', field)
            for field
            in row]


def squeezer(row):

    return [SPACES_REGEX.sub(' ', PROHIBITED_REGEX.sub(' ', field))
            for field
            in row]


def detecter(row):
    for field in row:
        if PROHIBITED_REGEX.search(field):
            raise ValueError('prohibited character in field: {}'.format(field))

    return row


def csv_to_tsv(input_stream,
               output_stream,
               header,
               delimiter=',',
               quotechar='"',
               sanitizer=stripper):

    rows = unicode_csv_reader(input_stream,
                              delimiter=delimiter,
                              quotechar=quotechar)

    if header:
        output_stream.write('\t'.join(header.split(',')) + '\n')
    for row in rows:
        output_stream.write('\t'.join(sanitizer(row)) + '\n')


if __name__ == '__main__':

    parser = argparse.ArgumentParser()

    parser.add_argument('--delimiter', '-d',
                        dest='delimiter',
                        default=',')
    parser.add_argument('--quotechar', '-q',
                        dest='quotechar',
                        default='"')
    parser.add_argument('--escape', '-e',
                        dest='escape',
                        action='store_const',
                        const=True)
    parser.add_argument('--header',
                        dest='header',
                        metavar='NAME[,NAME..]')
    parser.add_argument('--strip', '-x',
                        dest='strip',
                        action='store_const',
                        const=True)
    parser.add_argument('--replace', '-r',
                        dest='replace',
                        action='store_const',
                        const=True)
    parser.add_argument('--squeeze', '-z',
                        dest='squeeze',
                        action='store_const',
                        const=True)

    args = parser.parse_args()

    exclusive_cnt = 0
    if args.escape:
        exclusive_cnt += 1
    if args.strip:
        exclusive_cnt += 1
    if args.replace:
        exclusive_cnt += 1
    if args.squeeze:
        exclusive_cnt += 1
    if exclusive_cnt > 1:
        raise Exception('The flags --escape, --strip, --replace, and --squeeze'
                        ' cannot be used together.')

    if args.escape:
        sanitizer = escaper
    elif args.strip:
        sanitizer = stripper
    elif args.replace:
        sanitizer = replacer
    elif args.squeeze:
        sanitizer = squeezer
    else:
        sanitizer = detecter

    csv_to_tsv(sys.stdin,
               sys.stdout,
               args.header,
               delimiter=args.delimiter,
               quotechar=args.quotechar,
               sanitizer=sanitizer)

########NEW FILE########
__FILENAME__ = csv_to_xlsx
#!/usr/bin/env python

import argparse
import codecs
import csv
import openpyxl
import re
import sys

REGEX_CSV_SUFFIX = re.compile(r'.csv$', re.I)
REGEX_XLSX_SUFFIX = re.compile(r'.xlsx$', re.I)
REGEX_INVALID_SHEETNAME_CHARS = re.compile(r'[][*?/\.]')
REGEX_SPACES = re.compile(' +')
MAX_SHEETNAME_LENGTH = 31
ENCODING = 'utf-8'


def path_to_sheetname(path):
    sheetname = REGEX_CSV_SUFFIX.sub('', path)
    sheetname = REGEX_INVALID_SHEETNAME_CHARS.sub(' ', sheetname)
    sheetname = REGEX_SPACES.sub(' ', sheetname)

    return sheetname[0:MAX_SHEETNAME_LENGTH].strip()


def utf_8_encoder(unicode_csv_data):
    for line in unicode_csv_data:
        yield line.encode('utf-8')


def unicode_csv_reader(unicode_csv_data, dialect=csv.excel, **kwargs):
    csv_reader = csv.reader(utf_8_encoder(unicode_csv_data),
                            dialect=dialect, **kwargs)
    for row in csv_reader:
        yield [unicode(cell, 'utf-8') for cell in row]


def csv_to_xlsx(input_files, output_file):
    wb = openpyxl.Workbook()
    sheetnames = {}

    for filenum, input_file in enumerate(input_files):
        with codecs.open(input_file, encoding=ENCODING) as f:
            rows = unicode_csv_reader(f)
            if filenum == 0:
                ws = wb.get_active_sheet()
            else:
                ws = wb.create_sheet()
            sheetname = path_to_sheetname(input_file)
            if sheetname in sheetnames:
                raise ValueError('files {} and {} result in the same sheet '
                                 'name: "{}"'.format(input_file,
                                                     sheetnames[sheetname],
                                                     sheetname))
            sheetnames[sheetname] = input_file
            ws.title = sheetname
            for rownum, row in enumerate(rows):
                for colnum, value in enumerate(row):
                    # TODO: WHAT ABOUT DATES
                    ws.cell(row=rownum, column=colnum).value = value

    wb.save(output_file)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('input_files',
                        nargs='+',
                        metavar='CSV_FILE')
    parser.add_argument('--output-file', '-o',
                        dest='output_file',
                        required=True)

    args = parser.parse_args()

    if not REGEX_XLSX_SUFFIX.search(args.output_file):
        sys.stderr.write('ERROR: output file must have .xlsx '
                         'suffix: {}\n'.format(args.output_file))
        sys.exit(1)

    csv_to_xlsx(args.input_files, args.output_file)

########NEW FILE########
__FILENAME__ = date_seq
#!/usr/bin/env python

import argparse
import datetime
import re
import sys

import pprint
PP = pprint.PrettyPrinter()

REGEX_INPUT_DATE = re.compile(r'^\d{4,14}$')
WEEKDAY_TO_NUMBER = {
    'mon': 1,
    'tue': 2,
    'wed': 3,
    'thu': 4,
    'fri': 5,
    'sat': 6,
    'sun': 7
}


def check(dt, fmt, regex_date_filter):
    if not regex_date_filter:
        return True
    if regex_date_filter.match(dt.strftime(fmt)):
        return True
    return False


def make_year_iterator(start,
                       end,
                       regex_date_filter,
                       fmt):
    def date_iter(start_i, end_i):
        i = start_i
        while True:
            if i > end_i:
                return
            dt = datetime.datetime.strptime(str(i), fmt)
            if check(dt, fmt, regex_date_filter):
                yield dt
            i += 1

    return date_iter(int(start), int(end))


def make_month_iterator(start,
                        end,
                        regex_date_filter,
                        fmt):

    start_yyyy = int(start[0:4])
    start_mm = int(start[4:6])
    end_yyyy = int(end[0:4])
    end_mm = int(end[4:6])

    def date_iter(start_yyyy, start_mm, end_yyyy, end_mm):
        yyyy = start_yyyy
        mm = start_mm
        while True:
            if yyyy > end_yyyy or (yyyy == end_yyyy and mm > end_mm):
                return
            dt = datetime.datetime.strptime('%04d%02d' % (yyyy, mm), fmt)
            if check(dt, fmt, regex_date_filter):
                yield dt
            mm += 1
            if mm == 13:
                mm = 1
                yyyy += 1

    return date_iter(start_yyyy, start_mm, end_yyyy, end_mm)


def make_date_iterator(start,
                       end,
                       weekday_numbers,
                       regex_date_filter,
                       delta,
                       fmt):

    start_dt = datetime.datetime.strptime(start, fmt)
    end_dt = datetime.datetime.strptime(end, fmt)

    def date_iter(start_dt, end_dt):
        dt = start_dt
        while True:
            if dt > end_dt:
                return
            if not weekday_numbers:
                if check(dt, fmt, regex_date_filter):
                    yield dt
            else:
                weekday_number = int(dt.strftime('%u'))
                if weekday_number in weekday_numbers:
                    if check(dt, fmt, regex_date_filter):
                        yield dt
            dt += delta

    return date_iter(start_dt, end_dt)


def date_seq(start,
             end,
             weekdays,
             date_filter,
             output_fmt,
             output_stream):

    if len(start) != len(end):
        raise Exception('Start and end date must be same length')

    if not REGEX_INPUT_DATE.search(start):
        raise Exception(
            'Start date must be in YYYY[MM[DD[HH[MI[SS]]]]] format.')

    if not REGEX_INPUT_DATE.search(end):
        raise Exception('End date must be in YYYY[MM[DD[HH[MI[SS]]]]] format.')

    if weekdays:
        weekday_numbers = [WEEKDAY_TO_NUMBER[wkday.lower()[0:3]]
                           for wkday
                           in weekdays.split(',')]
    else:
        weekday_numbers = []

    regex_date_filter = re.compile(date_filter) if date_filter else None
    date_iter = None

    if len(start) == 4:
        fmt = '%Y'
        date_iter = make_year_iterator(start, end, regex_date_filter, fmt)
    elif len(start) == 6:
        fmt = '%Y%m'
        date_iter = make_month_iterator(start, end, regex_date_filter, fmt)
    elif len(start) == 8:
        delta = datetime.timedelta(days=1)
        fmt = '%Y%m%d'
    elif len(start) == 10:
        delta = datetime.timedelta(hours=1)
        fmt = '%Y%m%d%H'
    elif len(start) == 12:
        delta = datetime.timedelta(minutes=1)
        fmt = '%Y%m%d%H%S'
    elif len(start) == 14:
        delta = datetime.timedelta(seconds=1)
        fmt = '%Y%m%d%H%M%S'
    else:
        raise Exception('unexpected argument length: {}'.format(len(start)))

    if not date_iter:
        date_iter = make_date_iterator(start,
                                       end,
                                       weekday_numbers,
                                       regex_date_filter,
                                       delta,
                                       fmt)

    if output_fmt is None:
        output_fmt = fmt

    for dt in date_iter:
        output_stream.write(dt.strftime(output_fmt) + '\n')


parser = argparse.ArgumentParser()

parser.add_argument('start', help='YYYY[MM[DD[HH[MI[SS]]]]]')
parser.add_argument('end', help='YYYY[MM[DD[HH[MI[SS]]]]]')

parser.add_argument('--format', '-f',
                    dest='format', help='strftime style format for output',
                    default=None)

parser.add_argument('--regex', '-r',
                    dest='date_filter', help='date filter regex.',
                    default=None)

parser.add_argument('--weekdays', '-w',
                    dest='weekdays', help='comma separated: Sun,Mon,...',
                    default=None)

args = parser.parse_args()


date_seq(args.start,
         args.end,
         args.weekdays,
         args.date_filter,
         args.format,
         sys.stdout)

########NEW FILE########
__FILENAME__ = highlight
#!/usr/bin/env python

import argparse
import re
import sys

NORMAL = '\033[m'
BLACK_FOREGROUND = '\033[01;30m'
RED_FOREGROUND = '\033[01;31m'
GREEN_FOREGROUND = '\033[01;32m'
YELLOW_FOREGROUND = '\033[01;33m'
BLUE_FOREGROUND = '\033[01;34m'
MAGENTA_FOREGROUND = '\033[01;35m'
CYAN_FOREGROUND = '\033[01;36m'
WHITE_FOREGROUND = '\033[01;37m'


def highlight(input_stream, output_stream, esc_seq_to_pattern):
    for line in input_stream:
        output_line = line
        for esc_seq, pattern in esc_seq_to_pattern.iteritems():
            rx = re.compile("({})".format(pattern))
            output_line = rx.sub('{}\\1{}'.format(esc_seq, NORMAL),
                                 output_line)
        output_stream.write(output_line)


if __name__ == '__main__':

    parser = argparse.ArgumentParser()
    parser.add_argument('positional', nargs='*')
    parser.add_argument('--black',
                        dest='black',
                        metavar='PATTERN')
    parser.add_argument('--red', '-r',
                        dest='red',
                        metavar='PATTERN')
    parser.add_argument('--green', '-g',
                        dest='green',
                        metavar='PATTERN')
    parser.add_argument('--yellow', '-y',
                        dest='yellow',
                        metavar='PATTERN')
    parser.add_argument('--blue', '-b',
                        dest='blue',
                        metavar='PATTERN')
    parser.add_argument('--magenta', '-m',
                        dest='magenta',
                        metavar='PATTERN')
    parser.add_argument('--cyan', '-c',
                        dest='cyan',
                        metavar='PATTERN')
    parser.add_argument('--white', '-w',
                        dest='white',
                        metavar='PATTERN')

    args = parser.parse_args()
    pattern = None
    input_path = None

    if len(args.positional) == 1:
        if args.red or args.black or args.green or args.yellow or args.blue \
           or args.magenta or args.cyan or args.white:
            input_path = args.positional[0]
        else:
            pattern = args.positional[0]
    elif len(args.positional) == 2:
        pattern, input_path = args.positional
    elif len(args.positional) > 2:
        sys.stderr.write('USAGE: hightlight [OPTIONS] [PATTERN] [FILE]\n')
        sys.exit(1)

    esc_seq_to_pattern = {}
    if pattern and args.red:
        raise Exception('--red|-r cannot be used with default pattern')
    if pattern:
        esc_seq_to_pattern[RED_FOREGROUND] = pattern
    if args.red:
        esc_seq_to_pattern[RED_FOREGROUND] = args.red
    if args.black:
        esc_seq_to_pattern[BLACK_FOREGROUND] = args.black
    if args.green:
        esc_seq_to_pattern[GREEN_FOREGROUND] = args.green
    if args.yellow:
        esc_seq_to_pattern[YELLOW_FOREGROUND] = args.yellow
    if args.blue:
        esc_seq_to_pattern[BLUE_FOREGROUND] = args.blue
    if args.magenta:
        esc_seq_to_pattern[MAGENTA_FOREGROUND] = args.magenta
    if args.cyan:
        esc_seq_to_pattern[CYAN_FOREGROUND] = args.cyan
    if args.white:
        esc_seq_to_pattern[WHITE_FOREGROUND] = args.white

    if not esc_seq_to_pattern:
        sys.stderr.write("No PATTERN specified.\n")
        parser.print_help()
        sys.exit(1)

    if input_path:
        with open(input_path) as f:
            highlight(f, sys.stdout, esc_seq_to_pattern)
    else:
        highlight(sys.stdin, sys.stdout, esc_seq_to_pattern)

########NEW FILE########
__FILENAME__ = join_tsv
#!/usr/bin/env python

import argparse
import codecs
import collections
import os
import sys

ENCODING = 'utf-8'
BIG_FIRST = 1
BIG_LAST = 2
JOIN_INNER = 1
JOIN_LEFT = 2
JOIN_RIGHT = 3
JOIN_FULL = 4
DEFAULT_OUTER_NULL = ''
outer_null = None

sys.stdout = codecs.getwriter(ENCODING)(sys.stdout)
sys.stderr = codecs.getwriter(ENCODING)(sys.stderr)


def header_and_column_to_rows(path, column):
    with codecs.open(path, encoding=ENCODING) as f:
        column_to_rows = collections.defaultdict(list)

        header = f.readline().rstrip('\r\n').split('\t')
        row_len = len(header)
        column_index = None
        try:
            column_index = header.index(column)
        except ValueError:
            raise Exception('{} does not have a {} column'.format(
                path, column))
        del(header[column_index])

        for lineno, line in enumerate(f, start=2):
            fields = line.rstrip('\r\n').split('\t')
            if len(fields) != row_len:
                raise Exception('row {} does not have {} fields: {}'.format(
                    lineno,
                    row_len,
                    line))
            column_value = fields[column_index]
            del(fields[column_index])
            column_to_rows[column_value].append(fields)

    return header, column_to_rows


def print_row(join_value, fields1, fields2, f):
    f.write(join_value)
    f.write('\t')
    f.write('\t'.join(fields1))
    f.write('\t')
    f.write('\t'.join(fields2))
    f.write('\n')


def join_tsv(left_join_column,
             right_join_column,
             null,
             join_type,
             path1,
             path2,
             output_stream):

    if os.path.getsize(path1) > os.path.getsize(path2):
        big, small, file_order = path1, path2, BIG_FIRST
        big_join_column = left_join_column
        small_join_column = right_join_column
    else:
        big, small, file_order = path2, path1, BIG_LAST
        big_join_column = right_join_column
        small_join_column = left_join_column

    outer_join_big, outer_join_small = False, False

    small_header, column_to_rows = header_and_column_to_rows(small,
                                                             small_join_column)

    EMPTY_SMALL_HEADER = [outer_null] * len(small_header)

    if join_type == JOIN_FULL:
        outer_join_big, outer_join_small = True, True
    elif join_type == JOIN_LEFT:
        if file_order == BIG_FIRST:
            outer_join_big = True
        else:
            outer_join_small = True
    elif join_type == JOIN_RIGHT:
        if file_order == BIG_FIRST:
            outer_join_small = True
        else:
            outer_join_big = True

    with codecs.open(big, encoding=ENCODING) as f:
        big_header = f.readline().rstrip('\r\n').split('\t')
        row_len = len(big_header)
        column_index = None
        try:
            column_index = big_header.index(big_join_column)
        except ValueError:
            raise Exception('{} does not have a {} column'.format(
                big, big_join_column))
        del(big_header[column_index])
        EMPTY_BIG_HEADER = [outer_null] * len(big_header)

        print_row(left_join_column,
                  big_header if file_order == BIG_FIRST else small_header,
                  small_header if file_order == BIG_FIRST else big_header,
                  output_stream)

        # used if output_join_small is True
        join_values = set()

        for lineno, line in enumerate(f, start=2):
            big_fields = line.rstrip('\r\n').split('\t')
            if len(big_fields) != row_len:
                raise Exception('row {} does not have {} fields: {}'.format(
                    lineno,
                    row_len,
                    line))
            join_value = big_fields[column_index]
            del(big_fields[column_index])

            if join_value != null:
                small_rows = column_to_rows.get(join_value,
                                                [EMPTY_SMALL_HEADER]
                                                if outer_join_big
                                                else [])
                if outer_join_small:
                    join_values.add(join_value)
                for small_fields in small_rows:
                    print_row(
                        join_value,
                        big_fields if file_order == BIG_FIRST
                        else small_fields,
                        small_fields if file_order == BIG_FIRST
                        else big_fields,
                        output_stream)

        if outer_join_small:
            big_fields = EMPTY_BIG_HEADER
            for join_value, small_rows in column_to_rows.iteritems():
                if join_value not in join_values:
                    for small_fields in small_rows:
                        print_row(
                            join_value,
                            big_fields if file_order == BIG_FIRST
                            else small_fields,
                            small_fields if file_order == BIG_FIRST
                            else big_fields,
                            output_stream)


if __name__ == '__main__':

    parser = argparse.ArgumentParser()
    parser.add_argument('files',
                        nargs='+',
                        metavar='TSV_FILE')
    parser.add_argument('--column', '-c', '-C',
                        dest='column')
    parser.add_argument('--left', '-l',
                        dest='left',
                        action='store_true',
                        default=False)
    parser.add_argument('--left-column', '-L',
                        dest='left_column',
                        default=None)
    parser.add_argument('--right', '-r',
                        dest='right',
                        action='store_true',
                        default=False)
    parser.add_argument('--right-column', '-R',
                        dest='right_column',
                        default=None)
    parser.add_argument('--full', '-f',
                        dest='full',
                        action='store_true',
                        default=False)
    parser.add_argument('--null', '-n',
                        dest='null',
                        default='')
    parser.add_argument('--outer-null', '-o',
                        dest='outer_null',
                        default=DEFAULT_OUTER_NULL)
    parser.add_argument('--no-null', '-N',
                        dest='no_null',
                        action='store_true',
                        default=False)
    args = parser.parse_args()

    if len(args.files) != 2:
        sys.stderr.write('must be two files, not {}\n'.format(args.files))
        parser.print_help()
        sys.exit(1)

    left_join_column = None
    right_join_column = None

    if args.column:
        if args.left_column or args.right_column:
            sys.stderr.write('--column flag is incompatible with --left-column'
                             ' and --right-column flags\n')
            parser.print_help()
            sys.exit(1)
        left_join_column, right_join_column = args.column, args.column
    if args.left_column:
        left_join_column = args.left_column
    if args.right_column:
        right_join_column = args.right_column

    if not left_join_column or not right_join_column:
        sys.stderr.write('must specify join column(s)\n')
        parser.print_help()
        sys.exit(1)

    join_type = JOIN_INNER
    flag_cnt = 0
    if args.left:
        join_type = JOIN_LEFT
        flag_cnt += 1
    if args.right:
        join_type = JOIN_RIGHT
        flag_cnt += 1
    if args.full:
        join_type = JOIN_FULL
        flag_cnt += 1

    if flag_cnt > 1:
        sys.stderr.write('left, right or full join flags are exclusive\n')
        parser.print_help()
        sys.exit(1)

    outer_null = args.outer_null

    join_tsv(left_join_column,
             right_join_column,
             None if args.no_null else args.null,
             join_type,
             args.files[0],
             args.files[1],
             sys.stdout)

########NEW FILE########
__FILENAME__ = normalize_utf8
#!/usr/bin/env python

import argparse
import codecs
import sys
import unicodedata

ENCODING = 'utf-8'
NFC = 'NFC'
NFD = 'NFD'
NFKC = 'NFKC'
NFKD = 'NFKD'

sys.stdin = codecs.getreader(ENCODING)(sys.stdin)
sys.stdout = codecs.getwriter(ENCODING)(sys.stdout)
sys.stderr = codecs.getwriter(ENCODING)(sys.stderr)


def normalize_utf8(input_stream, output_stream, normalization_form):
    """
    Form must be 'NFC', 'NFD', 'NFKC', or 'NFKC'.

    Normalization forms are explained at

        http://unicode.org/reports/tr15/

    """
    for line in input_stream:
        output_stream.write(unicodedata.normalize(normalization_form, line))


if __name__ == '__main__':

    parser = argparse.ArgumentParser()
    parser.add_argument('positional_args',
                        nargs='?')
    parser.add_argument('--nfc',
                        action='store_true',
                        dest='nfc')
    parser.add_argument('--nfd',
                        action='store_true',
                        dest='nfd')
    parser.add_argument('--nfkc',
                        action='store_true',
                        dest='nfkc')
    parser.add_argument('--nfkd',
                        action='store_true',
                        dest='nfkd')
    args = parser.parse_args()

    flag_count = 0
    if args.nfc:
        flag_count += 1
        normalization_form = NFC
    if args.nfd:
        flag_count += 1
        normalization_form = NFD
    if args.nfkc:
        flag_count += 1
        normalization_form = NFKC
    if args.nfkd:
        flag_count += 1
        normalization_form = NFKD

    if flag_count == 0:
        args.nfc = True
        normalization_form = NFC
    if flag_count > 1:
        sys.stderr.write('At most one normalization flag can be used.\n')
        parser.print_usage(sys.stderr)
        sys.exit(1)

    if args.positional_args:
        fin = codecs.open(args.positional_args, encoding='utf-8')
    else:
        fin = sys.stdin

    normalize_utf8(fin, sys.stdout, normalization_form)

########NEW FILE########
__FILENAME__ = reservoir_sample
#!/usr/bin/env python

import argparse
import random
import sys


def reservoir_sample(count, input_stream, output_stream):

    n = None
    output = []

    try:
        n = int(count)
        if n < 1:
            raise ValueError
    except ValueError:
        raise Exception('argument not a positive integer')

    for i, line in enumerate(input_stream):
        if i < n:
            output.append(line)
        else:
            choice = random.randint(0, i)
            if choice < n:
                output[choice] = line

    for line in output:
        output_stream.write(line)


if __name__ == '__main__':

    parser = argparse.ArgumentParser()
    parser.add_argument('positional',
                        nargs='?',
                        metavar='FILE')
    parser.add_argument('--size', '-s',
                        dest='size',
                        type=int,
                        metavar='NUM',
                        required=True)
    parser.add_argument('--random-seed', '-r',
                        dest='random_seed',
                        default=None)
    args = parser.parse_args()

    if args.random_seed:
        random.seed(args.random_seed)

    if args.positional:
        with open(args.positional) as f:
            reservoir_sample(args.size, f, sys.stdout)
    else:
        reservoir_sample(args.size, sys.stdin, sys.stdout)

########NEW FILE########
__FILENAME__ = trim_tsv
#!/usr/bin/env python

import sys

DELIMITER = '\t'


def trim_tsv(input_stream, output_stream):
    for line in input_stream:
        row = line.rstrip('\r\n').split(DELIMITER)
        data = [field.strip() for field in row]
        output_stream.write(DELIMITER.join(data))
        output_stream.write('\n')


if __name__ == '__main__':

    if len(sys.argv) == 1:
        trim_tsv(sys.stdin, sys.stdout)
    elif len(sys.argv) == 2:
        with open(sys.argv[1]) as f:
            trim_tsv(f, sys.stdout)
    else:
        sys.stderr.write("USAGE: trim-tsv [FILE]\n")
        sys.exit(1)

########NEW FILE########
__FILENAME__ = tsv_to_csv
#!/usr/bin/env python

# Convert TSV (tab and newline delimited) to CSV.

import argparse
import codecs
import csv
import sys

ENCODING = 'utf-8'
NEWLINE_CHARS = u'\f\n\r\v\x85\u2028\u2029'


def unescaper(field):

    str_builder = []

    backslashed = False

    i = 0
    while i < len(field):
        ch = field[i]
        if ch == '\\':
            backslashed = True
        else:
            if backslashed:
                if ch == 'f':
                    str_builder.append('\f')
                if ch == 'n':
                    str_builder.append('\n')
                elif ch == 'r':
                    str_builder.append('\r')
                elif ch == 't':
                    str_builder.append('\t')
                elif field[i:i + 5] == 'u2028':
                    i += 4
                    str_builder.append(u'\u2028')
                elif field[i:i + 5] == 'u2029':
                    i += 4
                    str_builder.append(u'\u2029')
                elif ch == 'v':
                    str_builder.append('\v')
                elif field[i:i + 3] == 'x85':
                    i += 4
                    str_builder.append('\x85')
                elif ch == '\\':
                    str_builder.append('\\')
                else:
                    # if correctly escaped, this case
                    # won't happen.
                    str_builder.append(ch)
            else:
                str_builder.append(ch)
            backslashed = False
        i += 1

    return ''.join(str_builder)


def tsv_to_csv(input_stream,
               output_stream,
               delimiter=',',
               quotechar='"',
               unescape=False):

    csv_writer = csv.writer(output_stream,
                            delimiter=delimiter,
                            quotechar=quotechar)

    for line in input_stream:
        row = line.rstrip(NEWLINE_CHARS).split('\t')
        if unescape:
            row = [unescaper(field) for field in row]
        csv_writer.writerow(row)


if __name__ == '__main__':

    sys.stdin = codecs.getreader(ENCODING)(sys.stdin)
    sys.stdout = codecs.getwriter(ENCODING)(sys.stdout)
    sys.stderr = codecs.getwriter(ENCODING)(sys.stderr)

    parser = argparse.ArgumentParser()

    parser.add_argument('input', nargs='?')
    parser.add_argument('--delimiter', '-d',
                        dest='delimiter',
                        default=',')
    parser.add_argument('--quotechar', '-q',
                        dest='quotechar',
                        default='"')
    parser.add_argument('--unescape', '-u',
                        action='store_const',
                        const=True)

    args = parser.parse_args()

    if args.input:
        f = codecs.open(args.input, encoding=ENCODING)
    else:
        f = sys.stdin
    tsv_to_csv(f,
               sys.stdout,
               delimiter=args.delimiter,
               quotechar=args.quotechar,
               unescape=args.unescape)

########NEW FILE########
__FILENAME__ = tsv_to_json
#!/usr/bin/env python

import codecs
import json
import sys

ENCODING = 'utf-8'
NEWLINE_CHARS = u'\f\n\r\v\x85\u2028\u2029'

sys.stdin = codecs.getreader(ENCODING)(sys.stdin)
sys.stdout = codecs.getwriter(ENCODING)(sys.stdout)
sys.stderr = codecs.getwriter(ENCODING)(sys.stderr)

if len(sys.argv) == 1:
    f = sys.stdin
elif len(sys.argv) == 2:
    if sys.argv[1] == '--help':
        sys.stderr.write('USAGE: tsv-to-json [TSV_FILE]\n')
        sys.exit(1)
    f = codecs.open(sys.argv[1], encoding=ENCODING)
else:
    sys.stderr.write("USAGE: tsv_to_json.py [FILE]")
    sys.exit(1)

header = f.readline().rstrip(NEWLINE_CHARS).split('\t')

for lineno, line in enumerate(f):
    fields = line.rstrip(NEWLINE_CHARS).split('\t')
    if len(fields) != len(header):
        raise Exception('incorrect number of fields at line {}: {}'.format(
            lineno,
            line))
    print(json.dumps(dict(zip(header, fields))))

########NEW FILE########
__FILENAME__ = xlsx_to_csv
#!/usr/bin/env python

import argparse
import codecs
import datetime
import StringIO
import csv
import os
import pprint
import sys
import xlrd

DATE_FMT = '%Y-%m-%dT%H:%M:%S'
ENCODING = 'utf-8'
CSV_SUFFIX = '.csv'
PP = pprint.PrettyPrinter()


class UnicodeWriter(object):
    """
    This class is lifted from http://docs.python.org/2/library/csv.html
    """

    def __init__(self, f, dialect=csv.excel, encoding=ENCODING, **kwds):
        self.queue = StringIO.StringIO()
        self.writer = csv.writer(self.queue, dialect=dialect, **kwds)
        self.stream = f
        self.encoder = codecs.getincrementalencoder(encoding)()

    def writerow(self, row):
        self.writer.writerow([s.encode(ENCODING) for s in row])
        data = self.queue.getvalue()
        data = data.decode(ENCODING)
        data = self.encoder.encode(data)
        self.stream.write(data)
        self.queue.truncate(0)

    def writerows(self, rows):
        for row in rows:
            self.writerow(row)


def list_xlsx_sheets(xlsx_path, output_stream):
    book = xlrd.open_workbook(xlsx_path)
    for sheet in sorted(book.sheet_names()):
        output_stream.write(sheet)
        output_stream.write('\n')


def sheet_name_to_filename(sheet_name):
    return sheet_name + CSV_SUFFIX


def cell_to_str(cell, date_fmt, datemode):
    if cell.ctype == xlrd.XL_CELL_DATE:
        dt = datetime.datetime(*xlrd.xldate_as_tuple(cell.value, datemode))
        return dt.strftime(date_fmt)
    elif cell.ctype == xlrd.XL_CELL_NUMBER:
        if cell.value == int(cell.value):
            return unicode(int(cell.value))
        else:
            return unicode(cell.value)
    else:
        return unicode(cell.value)


def xlsx_book_to_csv(book, sheet_path, sheet_name, date_fmt):
    sheet = book.sheet_by_name(sheet_name)
    if sheet_path == '-':
        f = sys.stdout
    else:
        f = open(sheet_path, 'wb')
    csvw = UnicodeWriter(f)
    for rownum in range(0, sheet.nrows):
        row = [cell_to_str(cell, date_fmt, book.datemode)
               for cell
               in sheet.row(rownum)]
        csvw.writerow(row)
    if sheet_path != '-':
        f.close()


def xlsx_path_to_csv(xlsx_path, sheet_path, sheet_name, date_fmt):
    book = xlrd.open_workbook(xlsx_path)
    xlsx_book_to_csv(book, sheet_path, sheet_name, date_fmt)


def xlsx_path_to_csvs(xlsx_path, dir_path, date_fmt):
    book = xlrd.open_workbook(xlsx_path)
    for sheet_name in book.sheet_names():
        sheet_path = os.path.join(dir_path,
                                  sheet_name_to_filename(sheet_name))
        xlsx_book_to_csv(book, sheet_path, sheet_name, date_fmt)


if __name__ == '__main__':

    parser = argparse.ArgumentParser()
    parser.add_argument('paths',
                        nargs='*',
                        metavar='PATH')
    parser.add_argument('--date-format', '-d',
                        dest='date_fmt',
                        default=DATE_FMT)
    parser.add_argument('--sheet', '-s',
                        dest='sheet')
    parser.add_argument('--list', '-l',
                        dest='list',
                        action='store_true')
    args = parser.parse_args()

    if args.list:
        if len(args.paths) != 1:
            raise Exception("USAGE: xlsx-to-csv --list XLSX_FILE")
        list_xlsx_sheets(args.paths[0], sys.stdout)
    elif args.sheet:
        if len(args.paths) == 1:
            xlsx_path = args.paths[0]
            output_path = sheet_name_to_filename(args.sheet)
        elif len(args.paths) == 2:
            xlsx_path = args.paths[0]
            output_path = args.paths[1]
        else:
            raise Exception("USAGE xlsx-to-csv --sheet=NAME XLSX_FILE "
                            "[OUTPUT_FILE]")
        xlsx_path_to_csv(xlsx_path, output_path, args.sheet, args.date_fmt)
    else:
        if len(args.paths) != 2:
            raise Exception("USAGE: xlsx-to-csv XLSX_FILE OUTPUT_DIR")
        if os.path.exists(args.paths[1]):
            sys.stderr.write(
                'Something is already at the output path: {}\n'.format(
                    args.paths[1]))
            sys.exit(1)
        os.makedirs(args.paths[1])
        xlsx_path_to_csvs(args.paths[0], args.paths[1], args.date_fmt)

########NEW FILE########
